import logging
import time
import traceback
from fastapi import FastAPI, HTTPException, UploadFile, File, Depends, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi import Request
from pydantic import BaseModel, Field
from typing import Optional, List, Any, Dict
from datetime import date, datetime, timedelta
import uuid, os, json, pathlib, calendar, threading, hashlib, hmac
from collections import defaultdict
from database import db, savepoint, init_db, ensure_stok_yolda_columns, ensure_dusum_modu, ensure_operasyon_event_durum_latent, ensure_rapor_kapanis, ensure_kart_kategori_columns, ensure_kart_ekstre_donem, ensure_kart_satici_kural, ensure_kart_devir_islem_turu, ensure_isletmeci, ensure_abonelik, ensure_gider_kanonik, ensure_odeme_plani_odeme_yontemi
from operasyon_stok_motor import eksik_kullanim_kontrol, tum_subeler_skor_guncelle
from tr_saat import bugun_tr, dt_now_tr_naive
from kasa_service import (
    audit,
    insert_kasa_hareketi,
    iptal_kasa_hareketi,
    kart_plan_guncelle_tx,
    onay_ekle,
    vadeli_alim_kapat,
    vadeli_kasadan_odenen_toplam,
)
from sube_panel import router as sube_panel_router
from ciro_taslak_api import router as ciro_taslak_router
from sube_operasyon import router as sube_operasyon_router
from sube_kapanis_dual import router as sube_kapanis_dual_router
from operasyon_merkez_api import router as operasyon_merkez_router
from sube_personel_api import router as sube_personel_router
from banka_yatirim_api import router as banka_yatirim_router
from kasa_teslim_api import router as kasa_teslim_router
from sube_ici_borc_api import router as sube_ici_borc_router
from isletmeci_api import router as isletmeci_router
from abonelik_api import router as abonelik_router
from tedarikci_api import router as tedarikci_router
from odeme_plani_motor_api import router as odeme_plani_motor_router
from odeme_plani_api import router as odeme_plani_read_router
from evo_sync import router as evo_sync_router
from kart_analiz import router as kart_analiz_router
from gorev_api import router as gorev_router
from is_basvuru_api import router as is_basvuru_router
from ev_tasarim_api import router as ev_tasarim_router
from fatura_api import router as fatura_router  # İZOLE modül (öneri-only, kill switch)
from stok_sayim_api import router as stok_sayim_router  # İZOLE modül (kalibrasyon/kontrol, öneri-only)
from supplier_payment import router as supplier_payment_router  # İZOLE: tedarikçi ödeme olay katmanı (duyu toplar, beyin uyumaz)
import vardiya_v2 as _vv2
from vardiya_v2 import _ad_soyad_split as _vardiya_personel_ad_split


def ay_ekle(d: date, ay: int) -> date:
    """dateutil.relativedelta gerektirmeden tarihe ay ekler. Ay sonu taşmalarını düzeltir."""
    yil = d.year + (d.month - 1 + ay) // 12
    ay_no = (d.month - 1 + ay) % 12 + 1
    gun = min(d.day, calendar.monthrange(yil, ay_no)[1])
    return date(yil, ay_no, gun)
from motors import (
    karar_motoru,
    odeme_strateji_motoru,
    nakit_akis_simulasyon,
    guncel_kasa,
    kasa_detay,
    kart_analiz_hesapla,
    aylik_odeme_plani_uret, kart_kesim_plan_tetikle,
    uyari_motoru,
    finans_ozet_motoru,
    uyari_cache_clear,
)
from finans_core import (
    kart_borc, tum_kart_borclari, kasa_bakiyesi, kasa_bakiyesi_tarihte,
    kart_bakiye_ozeti,          # 🧭 kanonik bakiye modeli (yol haritası ADIM 4/12)
    kart_ekstre, kart_ekstre_donem_override, kart_bu_ay_odenen, kart_faiz_tahmini,
    kart_asgari_orani,
    faiz_hesapla_ve_yaz, tum_kartlar_faiz_hesapla,
    taksit_detay, gelecek_taksit_yuku, tum_kartlar_taksit_yuku,
    kart_ekstre_forecast, tum_kartlar_ekstre_forecast, kart_aktif_donem,
    aktif_kesim_gunu, nakit_akis_sim, nakit_akis_tahmin_dogruluk,
    kesim_tarihi_hesapla, _safe_date,
)

# ⚠️ 2026-09-01 denetimi: `logger` bu dosyada 362. satırda tanımlıydı ama
# 104/110. satırlarda — yani MODÜL YÜKLENİRKEN — kullanılıyordu. İzole router
# import'ları ("modül patlasa bile uygulama ayakta kalır" sözü) hata verdiğinde
# except gövdesi `logger.warning` çağırıp NameError atıyor ve UYGULAMA HİÇ
# AÇILMIYORDU. Tanım router kayıtlarından ÖNCEye alındı; 362'deki basicConfig
# yapılandırması yerinde duruyor (aynı ada ikinci kez atanması zararsız).
logger = logging.getLogger("evvel-erp")

app = FastAPI(title="EVVEL ERP", version="2.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
app.include_router(sube_panel_router)
app.include_router(ciro_taslak_router)
app.include_router(sube_operasyon_router)
app.include_router(sube_kapanis_dual_router)
app.include_router(operasyon_merkez_router)
app.include_router(sube_personel_router)
app.include_router(banka_yatirim_router)
app.include_router(kasa_teslim_router)
app.include_router(sube_ici_borc_router)
app.include_router(isletmeci_router)
app.include_router(abonelik_router)
app.include_router(tedarikci_router)
app.include_router(odeme_plani_motor_router)
app.include_router(odeme_plani_read_router)
# 📂 KAYIT DOSYASI (İz & Belge doktrini, 2026-08-15) — izole, salt-okur.
# try/except: modül patlasa bile uygulama ayakta kalır (duyu modülleriyle aynı disiplin).
try:
    from kayit_dosyasi_api import router as kayit_dosyasi_router
    app.include_router(kayit_dosyasi_router)
except Exception as _e:  # noqa: BLE001
    logger.warning("kayit_dosyasi_api yüklenemedi: %s", _e)
# 🕳️ BAĞSIZ STOK GİRİŞİ DUYUSU (2026-09-02) — izole, salt-okur.
# "Depoya mal girdi ama hiçbir tedarikçi zincirine bağlanamıyor" duyusu.
# ATALAY vakası: 01.08'de ZAFER'e Espresso +10 girmiş, siparişi iptal —
# belge talebi doğmamış, fatura kovalanmamış, cari oluşmamış ve mevcut
# duyuların HİÇBİRİ görmüyordu (acik-teslimat belge_talep'e, telafi-adaylari
# 'teslim_alindi'ye bakıyor; bu kayıt ikisine de düşmüyor).
try:
    from duyu_bagsiz_giris import router as bagsiz_giris_router
    app.include_router(bagsiz_giris_router)
except Exception as _e:  # noqa: BLE001
    logger.warning("duyu_bagsiz_giris yüklenemedi: %s", _e)
# 🔗 TEDARİKÇİ ZİNCİRİ (kimlik kararı + zaman çizgisi, 2026-08-15) — izole.
try:
    from tedarikci_zinciri_api import router as tedarikci_zinciri_router
    app.include_router(tedarikci_zinciri_router)
except Exception as _e:  # noqa: BLE001
    logger.warning("tedarikci_zinciri_api yüklenemedi: %s", _e)
app.include_router(evo_sync_router)
# ⛔ KART-ANALİZ ROUTER'I DEVRE DIŞI (2026-08-17, kart alanı denetimi — sahip
# talimatı "gereksiz kurulmuş yerleri devreden çıkar"). 4 ucu (parse-pdf,
# kartlar-listesi, aktar, kaydet-son-dort-hane) FE'den HİÇ çağrılmıyordu
# (grep: src/ içinde 0) ama /kart-analiz/aktar kart_hareketleri'ne KORUMASIZ
# yazıyordu: islem_turu daima 'HARCAMA' (ödeme harcama olurdu), taksit_sayisi
# yok (taksitli alım tek çekim), kaynak_tablo NULL (çift-yazma freni onları
# 'elle kayıt' sanıp MEŞRU ekstre satırlarını yutardı), audit yok.
# Ayrıca /kart-analiz/kartlar-listesi borcu 5. bir formülle (ödemeleri düşmeden)
# hesaplıyordu. Kapı kapatıldı; DOSYA DURUYOR — main.py:4360 kart_analiz.parse_pdf
# ekstre yüklemenin ANA motoru (import satırı 40'ta kalır).
# app.include_router(kart_analiz_router)
app.include_router(gorev_router)
app.include_router(is_basvuru_router)
app.include_router(ev_tasarim_router)

# Finansal Duyu — İZOLE (Akıllı Denetim duyu altyapısı). Modül patlasa bile ana
# uygulama ayakta kalsın diye try/except ile takılır (kullanıcı izolasyon direktifi).
try:
    from finansal_duyu_api import router as finansal_duyu_router
    app.include_router(finansal_duyu_router)
except Exception as _fin_duyu_err:
    logging.getLogger(__name__).warning(
        f"finansal_duyu modulu yuklenemedi (izole, ana akis etkilenmez): {_fin_duyu_err}"
    )
app.include_router(fatura_router)
# Belge Talep Motoru — İZOLE (teslim alınınca tedarikçiden fatura PDF kovala).
# Modül patlasa bile ana uygulama ayakta kalsın (kullanıcı izolasyon direktifi).
try:
    from belge_talep_api import router as belge_talep_router
    app.include_router(belge_talep_router)
except Exception as _belge_talep_err:
    logging.getLogger(__name__).warning(
        f"belge_talep modulu yuklenemedi (izole, ana akis etkilenmez): {_belge_talep_err}"
    )
# Tedarik Mutabakatı — İZOLE, SALT OKUR (fatura ↔ teslim adet çapası).
try:
    from tedarik_mutabakat_api import router as tedarik_mutabakat_router
    app.include_router(tedarik_mutabakat_router)
except Exception as _tedarik_mut_err:
    logging.getLogger(__name__).warning(
        f"tedarik_mutabakat modulu yuklenemedi (izole, ana akis etkilenmez): {_tedarik_mut_err}"
    )
# PERSONEL KİMLİĞİ — İZOLE: "aynı kişi mi?" + "gerçekten ne zaman başladı?"
# ÖNERİ-ONLY: hiçbir kaydı kendiliğinden birleştirmez/düzeltmez.
try:
    from personel_kimlik_api import router as personel_kimlik_router
    app.include_router(personel_kimlik_router)
except Exception as _pk_err:
    logging.getLogger(__name__).warning(
        f"personel_kimlik modulu yuklenemedi (izole, ana akis etkilenmez): {_pk_err}"
    )
# ARAMA — İZOLE, SALT OKUR: "şu belirli şeyi arıyorum" tek kapısı.
# Fatura aramasını YENİDEN YAZMAZ, mevcut /api/fatura/ara'yı çağırır.
try:
    from arama_api import router as arama_router
    app.include_router(arama_router)
except Exception as _arama_err:
    logging.getLogger(__name__).warning(
        f"arama modulu yuklenemedi (izole, ana akis etkilenmez): {_arama_err}"
    )
# BAKIŞ Ölçüm — İZOLE toplayıcı: "ekran iş üretiyor mu?" (M1-M5).
# Kendi iki tablosuna yazar, başka hiçbir tabloya dokunmaz; tüm uçları hata yutar.
try:
    from bakis_olcum_api import router as bakis_olcum_router
    app.include_router(bakis_olcum_router)
except Exception as _bakis_olcum_err:
    logging.getLogger(__name__).warning(
        f"bakis_olcum modulu yuklenemedi (izole, ana akis etkilenmez): {_bakis_olcum_err}"
    )
# OPS ÖLÇÜM — "iş kuyruğu işe yarıyor mu?" (2026-08-27). AYRI modül, AYRI iki
# tablo. bakis_olcum'a yazılmadı: onun ozet sorguları `gorunum`a göre süzmüyor,
# OPS oturumları oraya girseydi BAKIŞ'ın M1-M5 medyanları iki farklı ekranın
# davranışını tek kovada birleştirirdi (kapsam karışması).
try:
    from ops_olcum_api import router as ops_olcum_router
    app.include_router(ops_olcum_router)
except Exception as _ops_olcum_err:
    logging.getLogger(__name__).warning(
        f"ops_olcum modulu yuklenemedi (izole, ana akis etkilenmez): {_ops_olcum_err}"
    )
# Fatura İstek Motoru (BM-4+4A) — İZOLE (ödenmiş ama faturasız ≥eşik ödemeler).
try:
    from fatura_istek_api import router as fatura_istek_router
    app.include_router(fatura_istek_router)
except Exception as _fatura_istek_err:
    logging.getLogger(__name__).warning(
        f"fatura_istek modulu yuklenemedi (izole, ana akis etkilenmez): {_fatura_istek_err}"
    )
# Teslim Bildirim Akışı — İZOLE (sahip 2026-07-18: 'personel teslim alınca
# haberim olsun; Tamam deyince bir daha çıkmasın'). Salt-okur türetme.
try:
    from teslim_bildirim_api import router as teslim_bildirim_router
    app.include_router(teslim_bildirim_router)
except Exception as _teslim_bild_err:
    logging.getLogger(__name__).warning(
        f"teslim_bildirim modulu yuklenemedi (izole, ana akis etkilenmez): {_teslim_bild_err}"
    )
# Personel Puan Defteri — İZOLE (öneri-only; puan maaşa otomatik bağlanmaz).
try:
    from personel_puan_api import router as personel_puan_router
    app.include_router(personel_puan_router)
except Exception as _puan_err:
    logging.getLogger(__name__).warning(
        f"personel_puan modulu yuklenemedi (izole, ana akis etkilenmez): {_puan_err}"
    )
app.include_router(stok_sayim_router)
app.include_router(supplier_payment_router)
# K1 kart-ödeme tanısı — İZOLE mutabakat toplayıcı (salt-okur tarama + önizleme-varsayılan onarım)
try:
    from k1_kart_odeme_tani import router as k1_tani_router
    app.include_router(k1_tani_router)
except Exception as _k1_err:
    logging.getLogger(__name__).warning(
        f"k1_kart_odeme_tani modulu yuklenemedi (izole, ana akis etkilenmez): {_k1_err}"
    )
# DUYU OMURGASI (FAZ 0) — Katman-2 olay omurgası + ground-truth etiket defteri + cursor.
# Sinir sisteminin çekirdeği: duyular yazar, sinapslar cursor'la okur. Salt-okur uçlar.
try:
    from duyu_omurga import router as duyu_omurga_router
    app.include_router(duyu_omurga_router)
except Exception as _duyu_err:
    logging.getLogger(__name__).warning(
        f"duyu_omurga modulu yuklenemedi (izole, ana akis etkilenmez): {_duyu_err}"
    )
# TÜKETİM DÖRTGENİ (FAZ 1a) — giren↔satış↔kullanım↔sayım yan yana, salt-okur, kesit rozetli
try:
    from dortgen_duyu import router as dortgen_router
    app.include_router(dortgen_router)
except Exception as _dg_err:
    logging.getLogger(__name__).warning(
        f"dortgen_duyu modulu yuklenemedi (izole, ana akis etkilenmez): {_dg_err}"
    )
# DUYU SAF GÖRÜNÜMLERİ (FAZ 1c+1e) — vergi→nakit takvimi + kapanış-fark şube profili
try:
    from duyu_gorunumler import router as duyu_gorunum_router
    app.include_router(duyu_gorunum_router)
except Exception as _dgo_err:
    logging.getLogger(__name__).warning(
        f"duyu_gorunumler modulu yuklenemedi (izole, ana akis etkilenmez): {_dgo_err}"
    )
# FAZ 2 DUYULARI — kayıt disiplini üçlüsü (açıklama yoğunluğu + kapanış-sonrası/backdate +
# ödeme karması); hepsi Sv0 kaydet-gösterme, izole.
try:
    from duyu_faz2 import router as duyu_faz2_router
    app.include_router(duyu_faz2_router)
except Exception as _df2_err:
    logging.getLogger(__name__).warning(
        f"duyu_faz2 modulu yuklenemedi (izole, ana akis etkilenmez): {_df2_err}"
    )
# FAZ 3 SİNAPSLARI — duyular birbirine bağlanır (kase + zincir + kompozit); aday dili, izole.
try:
    from duyu_sinaps import router as duyu_sinaps_router
    app.include_router(duyu_sinaps_router)
except Exception as _ds3_err:
    logging.getLogger(__name__).warning(
        f"duyu_sinaps modulu yuklenemedi (izole, ana akis etkilenmez): {_ds3_err}"
    )
# 5-YZ TAMAMLAMA — örüntü duyuları (sayım-çevresi + fatura örüntüsü + ters zincir +
# ürün sessiz sıfırlanması); Sv0/aday, izole.
try:
    from duyu_oruntu import router as duyu_oruntu_router
    app.include_router(duyu_oruntu_router)
except Exception as _dor_err:
    logging.getLogger(__name__).warning(
        f"duyu_oruntu modulu yuklenemedi (izole, ana akis etkilenmez): {_dor_err}"
    )
# YAVRU ÖRME MOTORU (Y1) — bildirimsel kural kütüphanesi + T1/T2 bağ motoru; izole.
try:
    from duyu_yavru import router as duyu_yavru_router
    app.include_router(duyu_yavru_router)
except Exception as _dyv_err:
    logging.getLogger(__name__).warning(
        f"duyu_yavru modulu yuklenemedi (izole, ana akis etkilenmez): {_dyv_err}"
    )
# ÖZGÜN KURGU DUYULARI — vardiya plan-gerçek + menü fiyat izi + bildirim iletim; izole.
try:
    from duyu_ozgun import router as duyu_ozgun_router
    app.include_router(duyu_ozgun_router)
except Exception as _doz_err:
    logging.getLogger(__name__).warning(
        f"duyu_ozgun modulu yuklenemedi (izole, ana akis etkilenmez): {_doz_err}"
    )
# FAZ 4 ÖN-KURULUM — motor uyanış kapısı (UYUR): kanıt paketi önizleme + hazırlık ölçer +
# etiket köprüsü + operasyon ritmi. Motor bu modülü ÇAĞIRMAZ; insan önizler.
try:
    from duyu_uyanis import router as duyu_uyanis_router
    app.include_router(duyu_uyanis_router)
except Exception as _du4_err:
    logging.getLogger(__name__).warning(
        f"duyu_uyanis modulu yuklenemedi (izole, ana akis etkilenmez): {_du4_err}"
    )
# EVVEL BEYNİ v0.1 (L3 dil/sentez) — salt-okur gözlem katmanı: karar vermez, alarm kapatmaz,
# kişi/niyet atfetmez, operasyon başlatmaz (Codex sınır cümlesi).
try:
    from beyin_api import router as beyin_router
    app.include_router(beyin_router)
except Exception as _by_err:
    logging.getLogger(__name__).warning(
        f"beyin_api modulu yuklenemedi (izole, ana akis etkilenmez): {_by_err}"
    )
# Avans Servisi — İZOLE mini bordro-finans köprüsü (talep→onay→teslim→mahsup).
# Kasa izini (PERSONEL_AVANS) SADECE bu servis yazar; maaş motoru sadece OKUR.
try:
    from avans_service import router as avans_router
    app.include_router(avans_router)
except Exception as _avans_err:
    logging.getLogger(__name__).warning(
        f"avans_service modulu yuklenemedi (izole, ana akis etkilenmez): {_avans_err}"
    )
# Tam Maliyet — İZOLE (genel merkez gideri + tahakkuk + şube dağıtımı). P&L'ye dokunmaz.
try:
    from tam_maliyet_api import router as tam_maliyet_router
    app.include_router(tam_maliyet_router)
except Exception as _tam_maliyet_err:
    logging.getLogger(__name__).warning(
        f"tam_maliyet modulu yuklenemedi (izole, ana akis etkilenmez): {_tam_maliyet_err}"
    )
# Borç Navigasyon Motoru — İZOLE + SALT-OKUR (ABEK çekirdekli karar destek).
# Hiçbir şey yazmaz; modül patlasa bile ana uygulama ayakta kalır.
try:
    from borc_navigasyon_api import router as borc_nav_router
    app.include_router(borc_nav_router)
except Exception as _borc_nav_err:
    logging.getLogger(__name__).warning(
        f"borc_navigasyon modulu yuklenemedi (izole, ana akis etkilenmez): {_borc_nav_err}"
    )
# TV Menü — İZOLE (TULİPİ dijital menü panosu + canlı fiyat). /tv-menu HTML linki
# SPA catch-all'dan ÖNCE kayıtlı olsun diye burada (üstte) include edilir.
try:
    from tv_menu_api import router as tv_menu_router
    app.include_router(tv_menu_router)
except Exception as _tv_menu_err:
    logging.getLogger(__name__).warning(
        f"tv_menu modulu yuklenemedi (izole, ana akis etkilenmez): {_tv_menu_err}"
    )
# Reçete Kontrol — İZOLE (2026-07-08): ürün-aç düşürmeye devam eder; reçete
# yalnız KONTROL eder (Evo satış × reçete = beklenen ↔ stok hareketi = gerçek).
try:
    from recete_api import router as recete_router
    app.include_router(recete_router)
except Exception as _recete_err:
    logging.getLogger(__name__).warning(
        f"recete modulu yuklenemedi (izole, ana akis etkilenmez): {_recete_err}"
    )

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger("evvel-erp")


# ── CFO Panel (admin ana ekranı) şifre kapısı ───────────────────
# Şube paneli/personel QR akışlarından ayrı, sadece admin arayüzünün
# (App.jsx) açılış ekranına basit bir şifre kapısı. Demo aşaması için
# tek bir paylaşılan şifre yeterli — ENV ile değiştirilebilir.
ADMIN_SIFRE = os.environ.get("ADMIN_SIFRE", "evvel2026")


class _AdminGirisBody(BaseModel):
    sifre: str


# ── OTURUM JETONU ────────────────────────────────────────────────────────────
# Sorun (sahip, 2026-08-10): "her sayfa yenilemede şifre soruyor". Kapı yalnız
# istemcide bir bayraktı (localStorage bilinçli kullanılmıyordu) → F5 = yeniden giriş.
#
# Çözüm: sunucunun İMZALADIĞI süreli jeton. İstemci saklar, açılışta doğrulatır.
#   jeton = "<bitiş_zamanı>.<HMAC(ADMIN_SIFRE, bitiş_zamanı)>"
# Gizli anahtar ADMIN_SIFRE'nin kendisi olduğu için ŞİFRE DEĞİŞİNCE tüm eski
# oturumlar kendiliğinden geçersizleşir — ayrıca bir iptal listesi tutmak gerekmez.
# Sunucu yeniden başlarsa jeton yaşar (anahtar ENV'den gelir, bellekte üretilmez).
#
# ⚠️ DÜRÜST SINIR: bu kapı KAZARA açılmaya karşı bir perdedir, API'lerin kendisi
# hâlâ açıktır (bkz. güvenlik backlog'u). Jeton o durumu ne iyileştirir ne kötüleştirir.
ADMIN_OTURUM_GUN = 30


# ⚠️ TEK MERKEZ (2026-09-02): jeton mantigi `admin_oturum.py`e tasindi.
# Sebep: bir router'a kapi koymak isteyen modul (is_basvuru_api) `main`i
# import EDEMEZ — main zaten o router'i import ediyor, dongu olur. Mantigi
# kopyalamak da olmaz: kopya gun gelir ayrisir ve o gun kapi SESSIZCE acilir.
# Buradaki iki ad geriye-uyum icin duruyor, govde ortak modulden gelir.
from admin_oturum import jeton_uret as _admin_jeton_uret, aktor_bilgisi
from evvel_merkez_guard import merkez_mutasyon_korumasi
from admin_oturum import jeton_gecerli as _admin_jeton_gecerli


@app.post("/api/admin-giris")
def admin_giris(body: _AdminGirisBody):
    if (body.sifre or "").strip() != ADMIN_SIFRE:
        raise HTTPException(401, "Şifre yanlış")
    return {"ok": True, "jeton": _admin_jeton_uret(), "gecerlilik_gun": ADMIN_OTURUM_GUN}


@app.get("/api/admin-oturum")
def admin_oturum(jeton: str = ""):
    """Saklanan jeton hâlâ geçerli mi? Şifre değiştiyse burada düşer."""
    gecerli, kalan = _admin_jeton_gecerli(jeton)
    return {"gecerli": gecerli, "kalan_gun": round(kalan / 86400, 1) if gecerli else 0}

@app.middleware("http")
async def log_requests(request: Request, call_next):
    start = time.time()
    response = await call_next(request)
    ms = round((time.time() - start) * 1000)
    logger.info(f"{request.method} {request.url.path} → {response.status_code} ({ms}ms)")
    return response


@app.exception_handler(Exception)
async def hata_yakala(request: Request, exc: Exception):
    tb = traceback.format_exc()
    logger.error(f"HATA: {request.url.path}\n{tb}")
    # Hata ayıklama için tarayıcıya da özet ver (üretimde güvenli: sadece mesaj + ilk trace satırı)
    return JSONResponse(
        status_code=500,
        content={
            "detail": f"{type(exc).__name__}: {str(exc)}",
            "path": str(request.url.path),
            "trace": tb.splitlines()[-6:] if tb else [],
        }
    )


# ── GECE YARISI SCHEDULER ──────────────────────────────────────
def _akilli_denetim_telafi(gun_sayisi: int = 3, gecikme_sn: int = 90):
    """AÇILIŞTA TELAFİ KOŞUSU — atlanan gecenin denetimini kalkışta tamamlar.

    🔴 NEDEN VAR (2026-08-27, canlı bulgu):
    Gece denetimi `_gece_yarisi_scheduler` içinde SÜREÇ-İÇİ bir uyku
    döngüsüyle çalışıyor: `gece yarısına kadar uyu → uyan → sleep(30dk) →
    motoru koştur → BİR SONRAKİ geceye kadar uyu`. Konteyner bu pencerede
    yeniden başlarsa (deploy · Railway restart · uyku) o gecenin koşusu
    TAMAMEN atlanır — döngü baştan başlayınca "bir sonraki gece yarısı"
    yarını işaret eder ve BUGÜNÜ TELAFİ ETMEZ.
    Canlı kanıt: 27 Ağustos'ta 3 aktif şubenin son koşusu 25 Ağustos'tu;
    iki gece atlanmıştı ve ekran hepsine "temiz" diyordu.

    ⚠️ ÖNERİ-ONLY İHLALİ DEĞİL: motor `read_only` modda çalışır ve yalnız
    BULGU ÜRETİR — hiçbir kararı uygulamaz, hiçbir kaydı değiştirmez.
    Doktrin sistemin KARAR VERMESİNİ yasaklar, BAKMASINI değil. Bakmayan
    denetim, denetim değildir.

    ⚠️ İDEMPOTENT: `truth_motor_kararlar` içinde (sube_id, tarih) çifti
    varsa o gün ATLANIR — gecenin normal koşusu zaten aynı freni kullanıyor
    (bkz. scheduler içindeki aynı SELECT). İki kez koşmak mükerrer bulgu
    üretmez.

    ⚠️ SINIRLI GERİYE DÖNÜŞ: en fazla `gun_sayisi` gün geriye bakar. Aylarca
    geriye backfill YAPMAZ — motoru kapalı şube (`sube_aktif_mi` False)
    zaten hiç işlenmez, o şubelerin boşluğu "motor kapalı" olarak kalır ve
    ekranda öyle görünür (sahte doldurma yok).

    ⚠️ KALKIŞI BEKLETMEZ: ayrı thread + `gecikme_sn` gecikme. Uygulama önce
    istek karşılamaya başlar, DB havuzu açılışta dövülmez.
    """
    import time as _time
    try:
        _time.sleep(max(gecikme_sn, 5))
        from tr_saat import bugun_tr as _bugun_tr
        import truth_motor as _tm

        if not _tm._global_aktif():
            logger.info("🩺 Telafi koşusu atlandı — motor global olarak kapalı")
            return

        bugun = _bugun_tr()
        # Motor "dün"ü işler: bugünden geriye gun_sayisi gün.
        hedef_gunler = [str(bugun - timedelta(days=i)) for i in range(1, gun_sayisi + 1)]
        kosan = 0
        atlanan = 0
        hata = 0
        with db() as (conn, cur):
            cur.execute("SELECT id::text AS id, ad FROM subeler WHERE aktif=TRUE")
            subeler = cur.fetchall() or []
            for sb in subeler:
                sid = sb['id']
                try:
                    if not _tm.sube_aktif_mi(cur, sid):
                        continue
                except Exception:
                    continue
                for gun in hedef_gunler:
                    try:
                        cur.execute(
                            "SELECT 1 FROM truth_motor_kararlar WHERE sube_id=%s AND tarih=%s::date LIMIT 1",
                            (sid, gun),
                        )
                        if cur.fetchone():
                            atlanan += 1
                            continue
                        veriler = _tm.veri_topla(cur, sid, gun)
                        sonuc = _tm.motor_calistir(cur, sid, gun, veriler)
                        if sonuc.get("calisti"):
                            kosan += 1
                            logger.info(f"🩺 Telafi koşusu: {sb.get('ad')} · {gun}")
                    except Exception as e:
                        hata += 1
                        logger.warning(f"🩺 Telafi koşusu {sb.get('ad')} {gun}: {str(e)[:120]}")
            conn.commit()

        if kosan or hata:
            logger.info(f"🩺 Telafi koşusu bitti — koşan {kosan} · atlanan {atlanan} · hata {hata}")
        # ⚠️ KOŞTUĞUNU HER HÂLDE YAZ (çıktı üretmese bile): nabız, "motor
        # çalışıyor mu?" sorusunun tek dürüst cevabıdır.
        try:
            from duyu_omurga import duyu_nabiz_yaz as _nabiz
            _nabiz("denetim_telafi", durum=("hata" if hata else "basari"),
                   taranan=(kosan + atlanan), uretilen=kosan, yutulan_hata=hata,
                   not_metin=f"acilis telafisi · son {gun_sayisi} gun")
        except Exception:
            pass
    except Exception as e:  # noqa: BLE001
        logger.warning(f"🩺 Telafi koşusu yutuldu: {str(e)[:160]}")


def _gece_yarisi_scheduler():
    """
    Her gece yarısı çalışır. Restart bağımlılığını kaldırır.
    - Ay başı: aylık ödeme planı üret
    - Ay sonu: faiz hesapla
    - Her gece: kasa anomali kontrolü
    - Her gece 00:15: WhatsApp günlük özet (şubeler kapanışı tamamlar)
    """
    import time as _time

    logger.info("🕐 Scheduler thread aktif")

    while True:
        try:
            # Bir sonraki İstanbul gece yarısına kadar bekle
            bugu = bugun_tr()
            yarin = datetime.combine(bugu + timedelta(days=1), datetime.min.time())
            bekle = (yarin - dt_now_tr_naive()).total_seconds()
            _time.sleep(max(bekle, 60))  # en az 60 saniye

            bugun = bugun_tr()
            ay_son_gun = calendar.monthrange(bugun.year, bugun.month)[1]

            # Ay başı — sabit gider, maaş, taksit vs. (kart asgarisi HARİÇ;
            # kart asgarisi her kartın kendi kesim gününde tetiklenir)
            if bugun.day == 1:
                try:
                    sonuc = aylik_odeme_plani_uret(bugun.year, bugun.month)
                    logger.info(f"⏰ Scheduler: Aylık plan üretildi — {sonuc.get('toplam', 0)} kayıt")
                except Exception as e:
                    logger.error(f"⏰ Scheduler plan hatası: {e}")

            # KART BAZLI KESİM TETİKLEYİCİ — her gece tüm kartları yokla,
            # bugün kesim olan varsa o kart için ekstre planı üret/güncelle.
            # Hafta sonu/tatil kayması motorun içindedir.
            try:
                ks = kart_kesim_plan_tetikle()
                if ks.get("tetiklenen"):
                    logger.info(f"⏰ Scheduler: Kart kesim tetiklendi — {len(ks['tetiklenen'])} kart")
            except Exception as e:
                logger.error(f"⏰ Scheduler kart kesim hatası: {e}")

            # FAİZ — her gece tüm kartlar yoklanır.
            # faiz_hesapla_ve_yaz her kart için kendi kesim/son_odeme döngüsünü
            # değerlendirir. Son ödeme tarihi geçen ve faizi henüz yazılmamış
            # kartlara faiz işler. Diğerleri 'henuz_kapanmis_kesim_yok' /
            # 'zaten_yazilmis' / 'tam_odendi' döner ve atlanır.
            try:
                with db() as (conn, cur):
                    sonuclar = tum_kartlar_faiz_hesapla(cur)
                yazilan = sum(1 for k in sonuclar if k.get('durum') == 'yazildi')
                if yazilan > 0:
                    logger.info(f"⏰ Scheduler: Faiz üretildi — {yazilan} kart")
            except Exception as e:
                logger.error(f"⏰ Scheduler faiz hatası: {e}")

            # 📅 HAFTALIK TEDARİK ÖLÇÜMÜ — her PAZARTESİ (2026-08-24)
            # 11 açık sipariş temizlendi ve şubelere hatırlatma gönderildi. Ama
            # liste temizlemek ALIŞKANLIĞI değiştirmez — ZAFER aynı hatayı dört
            # kez üst üste yapmıştı. "Hatırlatma işe yaradı mı?" sorusunun tek
            # dürüst cevabı RAKAMDIR ve tek ölçüm değil EĞİLİM gerektirir.
            # İZOLE: kendi append-only tablosuna yazar, hata YUTULUR — gece
            # akışının geri kalanını asla bozmaz.
            if bugun.weekday() == 0:
                try:
                    from tedarik_mutabakat_api import haftalik_olcum_al
                    _to = haftalik_olcum_al(yaz=True)
                    logger.info("⏰ Scheduler: tedarik ölçümü — açık %s, patlama %s",
                                _to.get("acik_siparis"), _to.get("patlama"))
                    # Sahibin görmediği ölçüm ölçüm değildir: yalnız GÖRÜLECEK
                    # bir şey varsa yaz (uyarı bütçesi — "0 açık" mesajı atmaz).
                    if (_to.get("acik_siparis") or 0) > 0 or (_to.get("patlama") or 0) > 0:
                        try:
                            from whatsapp_bildirim import whatsapp_gonder
                            _sat = " · ".join(f"{k} {v}" for k, v in
                                              (_to.get("sube_kirilim") or {}).items())
                            whatsapp_gonder(
                                "📦 HAFTALIK TEDARİK ÖLÇÜMÜ\n\n"
                                f"Teslim alınmamış sipariş: {_to.get('acik_siparis')}"
                                + (f" (en eskisi {_to.get('en_eski_gun')} gün)"
                                   if _to.get("en_eski_gun") else "")
                                + (f"\nŞube: {_sat}" if _sat else "")
                                + f"\nSipariş patlaması (son 30 gün): {_to.get('patlama')}"
                                + f"\nMal geldi/sipariş kapanmadı: {_to.get('mal_geldi_kapanmadi')}"
                                + "\n\nTeslim al yapılmazsa stok artmaz ve "
                                  "tedarikçiden fatura istenmez.")
                        except Exception as _ew:  # noqa: BLE001
                            logger.warning("⏰ tedarik ölçümü WhatsApp atlandı: %s", str(_ew)[:120])
                except Exception as _et:  # noqa: BLE001
                    logger.warning("⏰ Scheduler tedarik ölçümü hatası (yutuldu): %s", str(_et)[:150])

            # Her gece — kasa anomali kontrolü
            try:
                with db() as (conn, cur):
                    cur.execute("SELECT COUNT(*) as sorunlu FROM v_kasa_anomali WHERE durum != 'OK'")
                    sorunlu = cur.fetchone()['sorunlu']
                    if sorunlu > 0:
                        logger.warning(f"⏰ Scheduler: {sorunlu} kasa anomali tespit edildi")
            except Exception as e:
                logger.warning(f"⏰ Scheduler anomali kontrol hatası: {e}")

            try:
                with db() as (conn, cur):
                    eksik_kullanim_kontrol(cur)
                    sk = tum_subeler_skor_guncelle(cur)
                    conn.commit()
                if sk:
                    logger.info(f"⏰ Scheduler: eksik kullanım + şube skor güncellendi ({len(sk)} şube)")
            except Exception as e:
                logger.warning(f"⏰ Scheduler stok davranış / skor hatası: {e}")

            # ── BORDRO GECE SENKRONU (2026-08-07, denetim bulgusu) ──────────────
            # ÖNCE: bordro yalnız ELLE tetikleniyordu ("Vardiya verisini maaşa aktar").
            # Kimse basmazsa personel_aylik taslağı doğmuyor, ödeme planına maaş
            # kalemi düşmüyor → 215 K ₺ maaş görünmeden vadesi geçiyordu (canlı vaka:
            # Temmuz dönemi 10 kalem, vade 1 Ağu, 7 Ağu'da hâlâ plan dışıydı).
            # ŞİMDİ: her gece kanonik yol (maas_service.aylik_vardiya_senkronize)
            # koşar. ONAYLI kaydı EZMEZ (servis kendi frenini uygular: durum='onaylandi'
            # ise yalnız planı tazeler). Dönem dışı personele kayıt açmaz.
            # İki dönem koşulur: içinde bulunulan ay + ödeme günü geçmemiş önceki ay
            # (ay başında önceki dönemin bordrosu hâlâ ödenmeyi bekler).
            try:
                _bugun_b = bugun_tr()
                _donemler = [(_bugun_b.year, _bugun_b.month)]
                if _bugun_b.day <= 15:
                    _onc = _bugun_b.replace(day=1) - timedelta(days=1)
                    _donemler.append((_onc.year, _onc.month))
                _bordro_log = []
                for _yb, _ab in _donemler:
                    with db() as (conn, cur):
                        cur.execute(
                            """SELECT * FROM personel
                               WHERE aktif=TRUE
                                  OR (cikis_tarihi IS NOT NULL AND cikis_tarihi >= MAKE_DATE(%s,%s,1))
                               ORDER BY ad_soyad""",
                            (_yb, _ab),
                        )
                        _kisiler = cur.fetchall()
                        _yazilan = 0
                        for _p in _kisiler:
                            try:
                                _r = _personel_aylik_vardiya_senkronize(cur, dict(_p), _yb, _ab)
                                if not (_r or {}).get("atlandi"):
                                    _yazilan += 1
                            except Exception as _e_p:
                                logger.warning(f"⏰ Bordro senkron ({_p.get('ad_soyad')}): {_e_p}")
                        conn.commit()
                    _bordro_log.append(f"{_yb}-{_ab:02d}: {_yazilan}/{len(_kisiler)}")
                logger.info(f"⏰ Scheduler: bordro senkronu — {' · '.join(_bordro_log)}")
            except Exception as e:
                logger.warning(f"⏰ Scheduler bordro senkron hatası: {e}")

            # Her gece — tedarikçi ödeme olay katmanını besle (DUYU duysun, alarm YOK).
            # İdempotent; yeni kart/nakit ödemeleri supplier_payment_event'e akar.
            try:
                from supplier_payment import supplier_payment_sync_v2 as supplier_payment_sync
                with db() as (conn, cur):
                    _sp = supplier_payment_sync(cur)
                    conn.commit()
                logger.info(f"⏰ Scheduler: supplier_payment_event beslendi ({_sp.get('eklenen')} yeni olay)")
            except Exception as e:
                logger.warning(f"⏰ Scheduler supplier_payment hatası: {e}")

            # FAZ 1b (2026-07-06) — her gece kart-kasa mutabakat taraması → duyu omurgası.
            # Fark vakası varsa Katman-2 olay yazar (alarmsız, Sv0); hata-yutar.
            try:
                from k1_kart_odeme_tani import gece_mutabakat_olay_yaz
                gece_mutabakat_olay_yaz()
                logger.info("⏰ Scheduler: K1 kart-kasa mutabakatı omurgaya tarandı")
            except Exception as e:
                logger.warning(f"⏰ Scheduler K1 mutabakat hatası: {e}")

            # FİNANSAL DUYU taraması (2026-08-14) — gecikmiş ödeme +
            # FIN_KART_ODEME_GIRILMEMIS ("ödedim ama deftere girmedim", OPET vakası).
            # Uç yalnız POST-tetikliydi; kimse çağırmazsa duyu hiç duymuyordu.
            # Salt-okur + hata-yutar; gözlem yazımı zaten idempotent (referans_id).
            try:
                from finansal_duyu_api import tarama as _fin_tarama
                _ft = _fin_tarama()
                logger.info(f"⏰ Scheduler: finansal duyu tarandı ({_ft.get('yazilan')})")
            except Exception as e:
                logger.warning(f"⏰ Scheduler finansal duyu hatası: {e}")

            # F5 (2026-08-15, sahip: "bu motor her gece kendi koşacak mı?" → EVET):
            # TESLİMAT ↔ FATURA eşleşme taraması. Açık teslimatlar × bağlanmamış
            # faturalar; yalnız YÜKSEK GÜVENLİ (karşılıklı-en-iyi + tarih yönü temiz
            # + çakışmasız + kalem/tutar kanıtı) çiftler onay kuyruğuna ÖNERİ düşer.
            # Yükleme-anı taramasıyla AYNI fonksiyon — iki kopya eşik yok.
            # İDEMPOTENT: mükerrer engeli (bekleyen aynı-tip öneri + aynı çift daha
            # önce sorulmuş) sayesinde her gece aynı öneri ÇOĞALMAZ.
            try:
                from belge_talep_api import teslimat_fatura_oneri_tara
                _tf = teslimat_fatura_oneri_tara(kaynak="gece")
                logger.info("⏰ Scheduler: teslimat↔fatura eşleşme önerisi "
                            f"({_tf.get('yazilan')} yeni · {_tf.get('mukerrer_atlandi')} mükerrer "
                            f"· {_tf.get('zayif_atlandi')} zayıf)")
            except Exception as e:
                logger.warning(f"⏰ Scheduler teslimat-fatura öneri hatası: {e}")

            # FAZ 1c+ (2026-07-06) — ayın 1'i: geçen ayın KDV pozisyonu dönem olayı (idempotent,
            # TAHMİNİ rozetli — beyanname değil). source_ref=YYYY-MM → ayda tek olay.
            if bugun.day == 1:
                try:
                    from calendar import monthrange as _mr2
                    from operasyon_merkez_api import ops_maliyet_kdv_pozisyon
                    _gy, _gm = (bugun.year, bugun.month - 1) if bugun.month > 1 else (bugun.year - 1, 12)
                    _poz = ops_maliyet_kdv_pozisyon(gun=_mr2(_gy, _gm)[1], sube_id=None)
                    from duyu_omurga import duyu_olay_yaz
                    duyu_olay_yaz(
                        "kdv_pozisyon", "finans.vergi.donem_pozisyonu", f"{_gy}-{_gm:02d}",
                        entity_scope="genel", occurred_at=f"{_gy}-{_gm:02d}-01",
                        signal_name="KDV dönem pozisyonu (tahmini)",
                        payload={"hesaplanan": _poz.get("toplam_hesaplanan_tl"),
                                 "indirilecek": _poz.get("toplam_indirilecek_tl"),
                                 "odenecek": _poz.get("toplam_odenecek_tl")},
                    )
                    logger.info("⏰ Scheduler: KDV dönem pozisyonu olayı yazıldı")
                    try:
                        from duyu_omurga import duyu_nabiz_yaz
                        duyu_nabiz_yaz("kdv_pozisyon", taranan=1, uretilen=1)
                    except Exception:
                        pass
                except Exception as e:
                    logger.warning(f"⏰ Scheduler KDV dönem olayı hatası: {e}")
                    try:
                        from duyu_omurga import duyu_nabiz_yaz
                        duyu_nabiz_yaz("kdv_pozisyon", durum="hata", yutulan_hata=1,
                                       not_metin=str(e)[:200])
                    except Exception:
                        pass

            # FAZ 1f (2026-07-06) — GECE SAĞLIK DEĞERLENDİRMESİ (proprioception): ritmini
            # belirgin aşan duyular için DURUM-GEÇİŞİ meta olayı (spam yok); hata-yutar.
            try:
                from duyu_omurga import gece_saglik_degerlendir
                gece_saglik_degerlendir()
                logger.info("⏰ Scheduler: duyu sağlık değerlendirmesi tamamlandı")
            except Exception as e:
                logger.warning(f"⏰ Scheduler duyu sağlık hatası: {e}")

            # (Beyin sentezi zincirin SONUNA taşındı — 2026-07-07 Codex zarf kararı:
            #  V3 anlatıcı, gecenin tüm olayları doğduktan sonra ve EN SON konuşur.)

            # FAZ V (2026-07-06) — SAHİP-DAHİL müdahale izi: dünün geriye-dönük işlemleri
            # omurgaya günlük özet olayı (Grok kör noktası); hata-yutar.
            try:
                from duyu_gorunumler import gece_mudahale_olay_yaz
                gece_mudahale_olay_yaz()
                logger.info("⏰ Scheduler: müdahale izi taraması tamamlandı")
            except Exception as e:
                logger.warning(f"⏰ Scheduler müdahale izi hatası: {e}")

            # FAZ 2 duyuları (2026-07-06): açıklama yoğunluğu + kapanış-sonrası/backdate +
            # ödeme karması — dünün ham kesitleri omurgaya; her biri kendi hatasını yutar.
            try:
                from duyu_faz2 import gece_faz2_calistir
                gece_faz2_calistir()
                logger.info("⏰ Scheduler: FAZ 2 duyu taraması tamamlandı")
            except Exception as e:
                logger.warning(f"⏰ Scheduler FAZ 2 duyu hatası: {e}")

            # FAZ 3 sinapsları (2026-07-06): kase + zincir + kompozit — duyuların ürettiği
            # olaylardan SONRA koşar ki kompozit o gecenin olaylarını da görsün.
            try:
                from duyu_sinaps import gece_sinaps_calistir
                gece_sinaps_calistir()
                logger.info("⏰ Scheduler: FAZ 3 sinaps taraması tamamlandı")
            except Exception as e:
                logger.warning(f"⏰ Scheduler FAZ 3 sinaps hatası: {e}")

            # FAZ 4 ön-kurulum (2026-07-06): etiket köprüsü (insan kararları → öğretmen
            # defteri) + operasyon ritmi (şube-dilim kesiti). Motor hâlâ UYUYOR.
            try:
                from duyu_uyanis import gece_uyanis_calistir
                gece_uyanis_calistir()
                logger.info("⏰ Scheduler: FAZ 4 uyanış hazırlığı tamamlandı")
            except Exception as e:
                logger.warning(f"⏰ Scheduler FAZ 4 uyanış hatası: {e}")

            # 5-YZ tamamlama (2026-07-06): örüntü duyuları — sayım-çevresi + fatura
            # örüntüsü + ters zincir + ürün sessiz sıfırlanması.
            try:
                from duyu_oruntu import gece_oruntu_calistir
                gece_oruntu_calistir()
                logger.info("⏰ Scheduler: örüntü duyuları taraması tamamlandı")
            except Exception as e:
                logger.warning(f"⏰ Scheduler örüntü duyuları hatası: {e}")

            # Özgün kurgu duyuları (2026-07-06): vardiya plan-gerçek kesiti (fiyat izi ve
            # bildirim iletim olay-güdümlü, kancalardan beslenir).
            try:
                from duyu_ozgun import gece_ozgun_calistir
                gece_ozgun_calistir()
                logger.info("⏰ Scheduler: özgün duyular taraması tamamlandı")
            except Exception as e:
                logger.warning(f"⏰ Scheduler özgün duyular hatası: {e}")

            # Konuşma izleri (2026-07-07, Codex çaprazlı): motor bulgu kesiti (kaba+D+2
            # gecikmeli) + söz→aksiyon adayı. Rapor izi gönderim kancasından beslenir.
            try:
                from duyu_konusma import gece_konusma_calistir
                gece_konusma_calistir()
                logger.info("⏰ Scheduler: konuşma izleri taraması tamamlandı")
            except Exception as e:
                logger.warning(f"⏰ Scheduler konuşma izleri hatası: {e}")

            # ÖZ-SORGU (2026-07-07, 'koltuğuma geç' talimatı): sistem kendine 3 patron
            # sorusu sorar; cevaplayamadıkları kendiliğinden veri dileği olur.
            # MİMARİ DENETİM (2026-07-15): tek try ~10 halkayı sarıyordu — ilk
            # halka çökünce ağır ön-hesap/bağ/öz-sorgu hiç koşmuyor, ertesi gün
            # her uç canlı hesaba düşüyordu. Her halka KENDİ try'ında (ev kuralı:
            # bir halka çökerse diğerleri yaşar — artık yapısal).
            def _halka(ad, fn):
                try:
                    fn()
                    logger.info(f"⏰ Scheduler halka tamam: {ad}")
                except Exception as _he:  # noqa: BLE001
                    logger.warning(f"⏰ Scheduler halka hatası ({ad}): {_he}")

            # OCR KURTARMA zincirin BAŞINDA (mimari denetim: kimlik/istek taraması
            # kurtarılan faturaları aynı gece görsün) + fatura_api içinde artık
            # eşzamanlılık freni var (pool zehirlenmez).
            def _ocr_kurtarma():
                from fatura_api import ocr_yeniden_dene
                _r = ocr_yeniden_dene(limit=50)
                logger.info(f"⏰ OCR gece kurtarma: {_r.get('kuyruga_alinan')} foto kuyruga alindi")
            _halka("ocr_kurtarma", _ocr_kurtarma)
            _halka("recete_kontrol", lambda: __import__("recete_api").gece_recete_kontrol_ozeti())
            _halka("degirmen", lambda: __import__("recete_api").gece_degirmen_izleme())
            _halka("kart_dongu", lambda: __import__("duyu_gorunumler").gece_kart_dongu_izleme())
            _halka("belge_kimlik", lambda: __import__("fatura_api").gece_belge_kimlik())
            # OCR RETRY (2026-08-08, sahip: "ileride de yaşamamak için önlem al"):
            # kota/ağ hatasıyla takılan faturaları yeniden dener. Belge kimliğinden
            # SONRA, kuyruk taramasından ÖNCE koşar — okunan fatura aynı gece
            # borç kuyruğuna girebilsin.
            _halka("ocr_retry", lambda: __import__("fatura_api").gece_ocr_takilanlari())
            # AP KART İZİ (2026-08-08): açık borçlar × kart ekstresi eşleştirmesi.
            # Gece YALNIZ RAPOR üretir (uygula=0) — otomatik bağlama ekstre importu
            # anında ya da sahip tetiklemesiyle olur, uykuda sessizce değil.
            _halka("kart_izi", lambda: __import__("fatura_api").gece_kart_izi_tara())
            _halka("fatura_istek", lambda: __import__("fatura_istek_api").gece_fatura_istek_tara())
            # FAZ A (2026-07-18): okunmuş ama kuyruğa bağlanmamış faturalar
            _halka("fatura_kuyruk", lambda: __import__("fatura_api").gece_fatura_kuyruk_tara())
            # SELF-HEAL (2026-07-19, APS dersi): hayalet sözler kasa iziyle kapanır —
            # mutabakattan ÖNCE koşar ki rapor temiz tabloyu görsün
            _halka("ap_selfheal", lambda: __import__("fatura_api").gece_ap_selfheal())
            # FAZ D (2026-07-18): AP mutabakat sağlığı (cari ↔ kuyruk çift-koşu)
            _halka("ap_mutabakat", lambda: __import__("fatura_api").gece_ap_mutabakat())
            # 📦 GRNI GÖRÜNÜRLÜĞÜ (2026-09-01 zincir denetimi, D-10): belge
            # talebi hiç açılmamış teslimatlar hiçbir borç toplamında
            # görünmüyor. Telafi yolu vardı ama insan çağırmalıydı; artık
            # gece SAYAR ve duyu olayı üretir (uygulama yine insan onayıyla).
            _halka("belge_telafi_gozlem",
                   lambda: __import__("belge_talep_api").gece_belge_telafi_gozlem())
            # 🕳️ Bağsız stok girişi (2026-09-02): tedarikçiye bağlanamayan
            # depo girişleri. Diğer duyuların KÖR NOKTASI — onlar belge_talep
            # ya da 'teslim_alindi' damgasına bakıyor; bu kayıtlarda ikisi de yok.
            _halka("bagsiz_giris",
                   lambda: __import__("duyu_bagsiz_giris").gece_bagsiz_giris_tara())
            _halka("fiyat_bandi", lambda: __import__("fatura_api").gece_fiyat_bandi_izleme())
            _halka("personel_puan", lambda: __import__("personel_puan_api").gece_personel_puan_tara())
            _halka("agir_onhesap", lambda: __import__("duyu_gorunumler").gece_agir_onhesap())
            _halka("bag_defteri", lambda: __import__("duyu_gorunumler").bag_defteri_hesapla())
            _halka("ozsorgu", lambda: __import__("beyin_api").gece_ozsorgu())
            logger.info("⏰ Scheduler: öz-sorgu zinciri tamamlandı")

            # EVVEL BEYNİ gece öz-anlatısı (L3) — ZİNCİRİN SONU (Codex zarf kararı:
            # anlatıcı, gecenin tüm olayları doğduktan sonra ve EN SON konuşur).
            # Arşive yazar + sabah WhatsApp mesajı (çerçeveli); hata-yutar.
            try:
                from beyin_api import gece_sentez
                gece_sentez()
                logger.info("⏰ Scheduler: Evvel Beyni gece sentezi tamamlandı (zincir sonu)")
            except Exception as e:
                logger.warning(f"⏰ Scheduler beyin sentez hatası: {e}")

            # Pazartesi — geçen haftanın kalem bazlı fire raporu
            if bugun.weekday() == 0:  # 0 = Pazartesi
                try:
                    from operasyon_stok_motor import haftalik_fire_hesapla
                    gecen_hf_basi = bugun - timedelta(days=7)
                    with db() as (conn, cur):
                        cur.execute("SELECT id FROM subeler WHERE aktif=TRUE")
                        sube_ids = [r["id"] for r in cur.fetchall()]
                    yazilan = 0
                    for sid in sube_ids:
                        try:
                            with db() as (conn, cur):
                                r = haftalik_fire_hesapla(cur, sid, gecen_hf_basi)
                                conn.commit()
                            if r.get("yazildi"):
                                yazilan += 1
                        except Exception as _e:
                            logger.warning(f"⏰ Fire haftalık {sid}: {_e}")
                    logger.info(f"⏰ Scheduler: Haftalık fire hesaplandı — {yazilan}/{len(sube_ids)} şube")
                except Exception as e:
                    logger.warning(f"⏰ Scheduler haftalık fire hatası: {e}")

            # ── RAPOR CACHE — gecelik özet batch (defensive) ──
            try:
                from rapor_cache import (
                    gunluk_ozet_topla_tum_subeler, aylik_food_cost_hesapla,
                    batch_log_basla, batch_log_bitir,
                )
                import time as _tm
                _t0 = _tm.time()
                dun = bugun - timedelta(days=1)
                bid = None
                with db() as (conn, cur):
                    bid = batch_log_basla(cur, 'gunluk_ozet',
                                          {'tarih': str(dun), 'sebep': 'gece_scheduler'})
                    conn.commit()
                with db() as (conn, cur):
                    sayac_dun = gunluk_ozet_topla_tum_subeler(cur, dun)
                    sayac_bugun = gunluk_ozet_topla_tum_subeler(cur, bugun)
                    conn.commit()
                # Aylık food cost — bu ay + ay başında geçen ay
                ym_bu = bugun.strftime("%Y-%m")
                with db() as (conn, cur):
                    aylik_food_cost_hesapla(cur, ym_bu)
                    if bugun.day <= 3:
                        ym_oncesi = (bugun.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
                        aylik_food_cost_hesapla(cur, ym_oncesi)
                    conn.commit()
                with db() as (conn, cur):
                    batch_log_bitir(cur, bid,
                                    islenen=sayac_dun + sayac_bugun,
                                    sure_ms=int((_tm.time() - _t0) * 1000),
                                    detay={'dun': sayac_dun, 'bugun': sayac_bugun, 'ym': ym_bu})
                    conn.commit()
                logger.info(f"⏰ Rapor cache: dün={sayac_dun} bugün={sayac_bugun} aylık={ym_bu}")
            except Exception as e:
                logger.warning(f"⏰ Scheduler rapor cache hatası: {e}")

            # ── GÜNLÜK FOOD COST (sube_food_cost_gun) — gece hesabı ──
            # 🐞 FIX (2026-08-03, canlı denetim "Food cost —" + "Fire 0" vakası):
            # _food_cost_hesapla_gun YALNIZ manuel butondan çağrılıyordu — gece
            # koşusu yoktu, tablo hep boş kalıyordu → Marj Özeti food cost "—",
            # fire 0 gösteriyordu. Dünün tamamlanmış günü her gece hesaplanır
            # (idempotent: UNIQUE(sube_id,tarih) upsert). Hata-yutar.
            try:
                from operasyon_merkez_api import _food_cost_hesapla_gun
                dun_fc = bugun - timedelta(days=1)
                with db() as (conn, cur):
                    fc_satir = _food_cost_hesapla_gun(cur, str(dun_fc), None)
                    conn.commit()
                logger.info(f"⏰ Günlük food cost: {str(dun_fc)} — {len(fc_satir or [])} şube")
            except Exception as e:
                logger.warning(f"⏰ Scheduler günlük food cost hatası: {e}")

            # ── AKILLI DENETİM MOTORU — gece 00:30, WhatsApp özetinden ÖNCE ──
            # Şubeler kapanışı 00:00–00:30 arası tamamlar, o yüzden 30 dk bekle.
            # Motor, dünün ACILIS/KAPANIS + Evo verisini üçgenleyip
            # truth_motor_kararlar'a yazar — WhatsApp özeti bunu okur.
            try:
                _time.sleep(30 * 60)  # 30 dakika
                from tr_saat import bugun_tr as _bugun_tr
                _dun = _bugun_tr() - timedelta(days=1)

                # ── Otomatik mesai kapatma — gece güvenlik süpürmesi ──────────
                # Mühür polling'i (kapanis-bekleyen) zaten gün içinde otomatik
                # kapatıyor; bu, hiç poll edilmeyen/açık kalan günün son
                # yoklamalarını da kapatır (cikis_tip='otomatik', planlanan saat).
                try:
                    from gorev_api import otomatik_mesai_kapat as _oto_kapat
                    with db() as (conn, cur):
                        cur.execute("SELECT id::text AS id FROM subeler WHERE aktif=TRUE")
                        _kap_subeler = cur.fetchall() or []
                        _kap_toplam = 0
                        for _ks in _kap_subeler:
                            for _kt in (str(_dun), str(_bugun_tr())):
                                try:
                                    _kap_toplam += len(_oto_kapat(cur, _ks['id'], _kt))
                                except Exception:
                                    pass
                        conn.commit()
                    if _kap_toplam:
                        logger.info(f"⏰ Otomatik mesai kapatma: {_kap_toplam} açık yoklama kapatıldı")
                except Exception as _e:
                    logger.warning(f"⏰ Otomatik mesai kapatma hatası: {_e}")

                # ── Fatura fotoğrafı saklama: 6 aydan eski "kanıt" görüntülerini sil
                # (kayıt + OCR kalır; sadece ağır BYTEA düşer → DB şişmez) ──────────
                try:
                    from fatura_api import (
                        fatura_foto_temizle as _ff_temizle,
                        fatura_modul_aktif as _ff_aktif,
                    )
                    if _ff_aktif():
                        with db() as (conn, cur):
                            _ff_n = _ff_temizle(cur)
                            conn.commit()
                        if _ff_n:
                            logger.info(f"🧾 Fatura foto saklama: {_ff_n} eski görüntü (6 ay+) temizlendi")
                except Exception as _e:
                    logger.warning(f"🧾 Fatura foto temizleme hatası: {_e}")

                try:
                    import truth_motor as _tm
                    if _tm._global_aktif():
                        with db() as (conn, cur):
                            cur.execute("SELECT id::text AS id, ad FROM subeler WHERE aktif=TRUE")
                            _subeler = cur.fetchall()
                            _calisan = 0
                            for _s in _subeler:
                                _sid = _s['id']
                                try:
                                    if not _tm.sube_aktif_mi(cur, _sid):
                                        continue
                                    # Aynı gece restart olursa tekrar kaydetmesin
                                    cur.execute(
                                        "SELECT 1 FROM truth_motor_kararlar WHERE sube_id=%s AND tarih=%s::date LIMIT 1",
                                        (_sid, str(_dun)),
                                    )
                                    if cur.fetchone():
                                        continue
                                    _veriler = _tm.veri_topla(cur, _sid, str(_dun))
                                    _sonuc = _tm.motor_calistir(cur, _sid, str(_dun), _veriler)
                                    if _sonuc.get("calisti"):
                                        _calisan += 1
                                except Exception as _e:
                                    logger.warning(f"⏰ Akıllı Denetim {_s.get('ad')}: {_e}")

                                # ── Sprint L: Evo verisi geçmişte eksikse, şimdi
                                # geldiyse son 3 günü yeniden değerlendir ──────────
                                try:
                                    _duzeltilen = _tm.evo_eksik_gunleri_yeniden_degerlendir(cur, _sid, gun_sayisi=3)
                                    if _duzeltilen:
                                        logger.info(f"⏰ Sprint L Evo telafi {_s.get('ad')}: {_duzeltilen}")
                                except Exception as _e:
                                    logger.warning(f"⏰ Sprint L Evo telafi {_s.get('ad')}: {_e}")
                            conn.commit()
                        logger.info(f"⏰ Akıllı Denetim: {_calisan}/{len(_subeler)} şube — tarih={_dun}")
                except Exception as e:
                    logger.warning(f"⏰ Akıllı Denetim motor hatası: {e}")

                # ── P3 Faz 1: sevkiyat çift-kolon tutarlılık duyusu (global, 1 kez) ──
                # Audit Brain'in iç veri-bütünlüğü duyusu: yeni kolon tek gerçek
                # kaynak mı? Bulguları öğrenme defterine yazar (salt okuma).
                try:
                    from siparis_kontrol_kulesi import sevkiyat_kolon_tutarsizlik_tara
                    with db() as (conn, cur):
                        _tut = sevkiyat_kolon_tutarsizlik_tara(cur, gun=120)
                        conn.commit()
                    logger.info(
                        f"⏰ P3 çift-kolon tutarlılık: taranan={_tut.get('taranan')} "
                        f"uyumsuz={_tut.get('uyumsuz')} bos_eski={_tut.get('bos_eski')}"
                    )
                except Exception as _e:
                    logger.warning(f"⏰ P3 çift-kolon tutarlılık tarama hatası: {_e}")

                # ── Kayıt katmanı: sipariş davranış profili (Katman 3, global) ──
                # Tedarik zinciri + sipariş-davranışı denetiminin ilk taşı: şube
                # başına sipariş sıklığını günlük kaydeder (saf veri, hipotez yok).
                try:
                    from siparis_kontrol_kulesi import siparis_davranis_gunluk_gozlemle
                    with db() as (conn, cur):
                        _sdg = siparis_davranis_gunluk_gozlemle(cur, pencere_gun=7)
                        conn.commit()
                    logger.info(f"⏰ Sipariş davranış profili: {_sdg.get('yazilan')} şube kaydedildi")
                except Exception as _e:
                    logger.warning(f"⏰ Sipariş davranış profili hatası: {_e}")
            except Exception as e:
                logger.warning(f"⏰ Akıllı Denetim bekleme hatası: {e}")

            # ── EKSİK CİRO GÜVENCE — gece sweep (Evo öneri üret) ──
            # Personel kapanışı/açılışı unutsa bile: Evo'da satış olan ama Evvel'de
            # ciro olmayan günleri bulup ONAY BEKLEYEN öneri üretir (deftere yazmaz,
            # KAPANIS event'e dokunmaz). WhatsApp özetinden ÖNCE çalışır ki mesaj
            # taze önerileri içersin. Hata-yutar (Evo erişilemezse atlar).
            try:
                from ciro_taslak_api import eksik_gun_ciro_tara, EksikGunTaraBody
                _ek = eksik_gun_ciro_tara(EksikGunTaraBody(gun_sayisi=35, uygula=True))
                _ekyeni = _ek.get("oneri_sayisi", 0)
                if _ekyeni:
                    logger.info(f"⏰ Eksik ciro sweep: {_ekyeni} yeni Evo önerisi onay kuyruğuna düştü")
                if _ek.get("evo_hata"):
                    logger.warning(f"⏰ Eksik ciro sweep — Evo erişim sorunu: {_ek.get('evo_hata')}")
            except Exception as e:
                logger.warning(f"⏰ Eksik ciro sweep hatası: {e}")

            # ── WHATSAPP GÜNLÜK ÖZET — gece 00:30 ──
            try:
                from whatsapp_bildirim import gunluk_ozet_gonder
                sonuc = gunluk_ozet_gonder(_dun)
                if sonuc.get("basarili"):
                    logger.info("⏰ WhatsApp: Günlük özet gönderildi")
                else:
                    logger.warning(f"⏰ WhatsApp: Gönderilemedi — {sonuc.get('hata', '?')}")
            except Exception as e:
                logger.warning(f"⏰ WhatsApp hatası: {e}")

            # ── FIRE/İADE KANIT FOTOĞRAFI TEMİZLİĞİ — depolama maliyeti için ──
            # 7 günden eski kanıt fotoğrafları (BYTEA) silinir. Hash'ler de gittiği
            # için tekrar-kullanım kontrolü bu pencereyle sınırlı kalır (bilinçli karar).
            try:
                from fire_bildirim import foto_eski_temizle
                with db() as (conn, cur):
                    _silinen = foto_eski_temizle(cur, gun=7)
                    conn.commit()
                if _silinen:
                    logger.info(f"⏰ Fire kanıt fotoğrafı temizliği: {_silinen} eski fotoğraf silindi")
            except Exception as e:
                logger.warning(f"⏰ Fire kanıt fotoğrafı temizliği hatası: {e}")

            # ── OPERASYON EVENT — bugünün açılış/kapanış satırları ──
            # Gece yarısı (00:00–02:00) is_gunu_tr() hâlâ ÖNCEKİ iş gününü döndürür
            # (kapanış 02:00'a kadar düne sayılır). Bu yüzden bugünün satırlarını
            # 02:30'da oluştururuz — o saatte is_gunu_tr() = bugün. Dashboard (ensure=False)
            # bu satırları hazır bulur. Idempotent; restart olursa startup telafi eder.
            try:
                hedef = datetime.combine(bugun_tr(), datetime.min.time()) + timedelta(hours=2, minutes=30)
                _bekle2 = (hedef - dt_now_tr_naive()).total_seconds()
                if _bekle2 > 0:
                    _time.sleep(_bekle2)
                from sube_operasyon import ensure_events_tum_subeler
                with db() as (conn, cur):
                    _ne = ensure_events_tum_subeler(cur)
                    conn.commit()
                logger.info(f"⏰ Scheduler: operasyon event satırları oluşturuldu — {_ne} şube")
            except Exception as e:
                logger.warning(f"⏰ Scheduler operasyon event hatası: {e}")

        except Exception as e:
            logger.error(f"⏰ Scheduler genel hata: {e}")
            import time as _t
            _t.sleep(300)  # hata olursa 5 dakika bekle, tekrar dene


def kapanis_hatirlatma_tara():
    """Kapanışa ~1 saat kala, o günün KAPANIŞ SORUMLUSUNA WhatsApp hatırlatma gönderir
    ('kapanış senden bekleniyor'). İZOLE: hata yutar. Gün başına 1 kez (subeler.
    kapanis_hatirlatma_tarih). Şube zaten mühürlendiyse / sorumlu telefonu yoksa atlar."""
    try:
        from tr_saat import is_gunu_tr, dt_now_tr_naive
        from whatsapp_bildirim import whatsapp_gonder_numara
    except Exception:
        return
    try:
        now = dt_now_tr_naive()
    except Exception:
        return
    tarih = str(is_gunu_tr())
    try:
        with db() as (conn, cur):
            cur.execute(
                """SELECT s.id, s.ad, s.kapanis_saati, s.kapanis_hatirlatma_tarih,
                          p.ad_soyad, p.telefon
                   FROM subeler s
                   JOIN personel p ON p.id::text = s.aktif_kapanis_sorumlusu_personel_id
                   WHERE s.aktif = TRUE AND s.aktif_kapanis_sorumlusu_tarih = %s::date
                     AND NULLIF(TRIM(s.kapanis_saati), '') IS NOT NULL""",
                (tarih,),
            )
            rows = [dict(r) for r in (cur.fetchall() or [])]
            for d in rows:
                if str(d.get("kapanis_hatirlatma_tarih") or "") == tarih:
                    continue  # bugün zaten gönderildi
                tel = (d.get("telefon") or "").strip()
                if not tel:
                    continue
                ks = (d.get("kapanis_saati") or "").strip()
                try:
                    hh, mm = int(ks.split(":")[0]), int(ks.split(":")[1])
                except Exception:
                    continue
                kapanis_dt = now.replace(hour=hh, minute=mm, second=0, microsecond=0)
                dk_kala = (kapanis_dt - now).total_seconds() / 60.0
                if not (40 <= dk_kala <= 80):  # ~1 saat kala penceresi
                    continue
                # Zaten mühürlendiyse atla
                cur.execute(
                    "SELECT 1 FROM sube_operasyon_event WHERE sube_id=%s AND tip='KAPANIS' "
                    "AND sira_no=0 AND durum='tamamlandi' AND tarih=%s",
                    (d["id"], tarih),
                )
                if cur.fetchone():
                    continue
                msg = (
                    f"Merhaba {d.get('ad_soyad') or ''} 👋\n"
                    f"*{d.get('ad')}* bugün senin kapanış sorumluluğunda 🔒\n"
                    f"Kapanışa ~1 saat kaldı. Kasa sayımı + ciro girip mührü atmayı unutma. Teşekkürler 🙏"
                )
                try:
                    whatsapp_gonder_numara(tel, msg)
                    cur.execute("UPDATE subeler SET kapanis_hatirlatma_tarih=%s WHERE id=%s", (tarih, d["id"]))
                    conn.commit()
                except Exception:
                    pass
    except Exception as e:
        logger.warning(f"⏰ kapanis hatirlatma tara hatasi (yutuldu): {e}")


_EKSIK_CIRO_EMNIYET = {"gun": None}  # gündüz emniyet koşusu — günde 1 kez


def _kapanis_hatirlatma_scheduler():
    """Gün-içi periyodik (15 dk): kapanışa ~1 saat kala sorumluya hatırlatma."""
    import time as _t
    while True:
        try:
            kapanis_hatirlatma_tara()
        except Exception as e:
            logger.warning(f"⏰ kapanis hatirlatma scheduler hatasi: {e}")
        # 🛟 EKSİK CİRO EMNİYET KOŞUSU (sahip 2026-07-19: Zafer 18.07 girilmemişti,
        # gece 00:30 sweep'i yakalamadı — 'çalışmıyor yine'). Gece koşusu Evo
        # erişim aksaması/yeniden başlatma yüzünden kaçırabilir; TR 11-13 bandında
        # GÜNDE BİR kez kısa sweep (son 3 gün) yeniden dener. Hata-yutar.
        try:
            from tr_saat import dt_now_tr
            _su = dt_now_tr()
            if 11 <= _su.hour < 13 and _EKSIK_CIRO_EMNIYET.get("gun") != _su.date().isoformat():
                from ciro_taslak_api import eksik_gun_ciro_tara, EksikGunTaraBody
                _er = eksik_gun_ciro_tara(EksikGunTaraBody(gun_sayisi=3, uygula=True))
                _EKSIK_CIRO_EMNIYET["gun"] = _su.date().isoformat()
                logger.info(f"🛟 Eksik ciro emniyet koşusu: oneri={_er.get('oneri_sayisi')} "
                            f"islenen={_er.get('toplam_eslesme')} evo_hata={_er.get('evo_hata')}")
        except Exception as e:
            logger.warning(f"🛟 eksik ciro emniyet hatasi (yutuldu): {e}")
        _t.sleep(900)  # 15 dakika


@app.on_event("startup")
def startup():
    init_db()
    try:
        with db() as (conn, cur):
            ensure_stok_yolda_columns(cur)
    except Exception as e:
        logger.warning("stok_yolda kolon migrasyonu (startup): %s", e)
    try:
        with db() as (conn, cur):
            ensure_dusum_modu(cur)
    except Exception as e:
        logger.warning("dusum_modu migrasyonu (startup): %s", e)
    try:
        with db() as (conn, cur):
            ensure_kart_kategori_columns(cur)
    except Exception as e:
        logger.warning("kart kategori (harcama_tipi/sahip) migrasyonu (startup): %s", e)
    try:
        with db() as (conn, cur):
            ensure_kart_ekstre_donem(cur)
    except Exception as e:
        logger.warning("kart_ekstre_donem migrasyonu (startup): %s", e)
    try:
        with db() as (conn, cur):
            ensure_kart_satici_kural(cur)
    except Exception as e:
        logger.warning("kart_satici_kural migrasyonu (startup): %s", e)
    try:
        with db() as (conn, cur):
            ensure_kart_devir_islem_turu(cur)
    except Exception as e:
        logger.warning("kart_devir_islem_turu migrasyonu (startup): %s", e)
    try:
        with db() as (conn, cur):
            ensure_isletmeci(cur)
    except Exception as e:
        logger.warning("isletmeci migrasyonu (startup): %s", e)
    try:
        with db() as (conn, cur):
            ensure_abonelik(cur)
    except Exception as e:
        logger.warning("abonelik migrasyonu (startup): %s", e)
    try:
        with db() as (conn, cur):
            ensure_odeme_plani_odeme_yontemi(cur)
    except Exception as e:
        logger.warning("odeme_plani odeme_yontemi migrasyonu (startup): %s", e)
    try:
        with db() as (conn, cur):
            ensure_gider_kanonik(cur)
    except Exception as e:
        logger.warning("gider_kanonik gorunumu (startup): %s", e)
    try:
        with db() as (conn, cur):
            ensure_operasyon_event_durum_latent(cur)
    except Exception as e:
        logger.warning("operasyon_event durum=latent migrasyonu (startup): %s", e)
    try:
        with db() as (conn, cur):
            ensure_rapor_kapanis(cur)
    except Exception as e:
        logger.warning("rapor_kapanis migrasyonu (startup): %s", e)
    # 🆕 KART GERÇEK MODELİ (yol haritası ADIM 2/12) — SAF EKLEME: tablolar
    # kurulur ama hiçbir uç yazmaz/okumaz; bugünkü davranış birebir sürer.
    # Doldurma ADIM 3, kullanım ADIM 6-7. Hata-yutar: kurulum düşse bile
    # sistem eskisi gibi çalışmaya devam eder.
    try:
        with db() as (conn, cur):
            from database import ensure_kart_gercek_modeli
            ensure_kart_gercek_modeli(cur)
    except Exception as e:
        logger.warning("kart gerçek modeli migrasyonu (startup): %s", e)
    # 💵 KASA TESLİM DEFTERLEŞMESİ + HAYALET MERKEZ KASASI (2026-08-17, sahip
    # onayı "TAŞISIN VE 3 ÖNERİYİ DE KUR"). İki iş: (1) pasif 'sube-merkez'e
    # damgalanmış hareketleri şubesizlik kovasına indir, (2) 144 teslim kaydını
    # çift kayıtla deftere işle. TOPLAM KASAYI DEĞİŞTİRMEZ (net 0); yalnız
    # şube/merkez dağılımı gerçeğe döner. İdempotent — her başlatmada güvenli.
    try:
        with db() as (conn, cur):
            from database import ensure_kasa_teslim_defterlesme
            ensure_kasa_teslim_defterlesme(cur)
    except Exception as e:
        logger.warning("kasa teslim defterleşmesi (startup): %s", e)
    # 💵 FİZİKSEL NAKİT BOYUTU (2026-08-18, sahip teyidi "kiralar şube
    # bankalarından, maaşlarda bankadan"). `tutar` = para pozisyonu (sahibin
    # modeli, DOKUNULMAZ) · `nakit_etki` = çekmecede gerçekten ne oldu.
    # NULL = bilinmiyor (0 DEĞİL) — yanlış çıkarım boş bırakmaktan tehlikelidir.
    try:
        with db() as (conn, cur):
            from database import ensure_nakit_etki
            ensure_nakit_etki(cur)
    except Exception as e:
        logger.warning("nakit etki migrasyonu (startup): %s", e)
    # 🏷️ MALİYET MERKEZİ ≠ ÖDEYEN (2026-08-18, Codex denetimli). Sahip:
    # "krediler ortak kasadan ödeniyormuş gibi düşün; Alsancak kredisi bazen
    # Gazze'den çekilir". Salt EKLEME — kolonlar açılır, davranış değişmez.
    try:
        with db() as (conn, cur):
            from database import ensure_maliyet_merkezi
            ensure_maliyet_merkezi(cur)
    except Exception as e:
        logger.warning("maliyet merkezi migrasyonu (startup): %s", e)
    try:
        with db() as (conn, cur):
            from gorev_api import _seed_sablonlar
            _seed_sablonlar(cur)
            conn.commit()
    except Exception as e:
        logger.warning(f"Görev şablonu seed hatası: {e}")
    # Her başlatmada bu ay için plan üret (yoksa üretir, varsa atlar)
    bugun = bugun_tr()
    try:
        sonuc = aylik_odeme_plani_uret(bugun.year, bugun.month)
        if sonuc['toplam'] > 0:
            logger.info(f"✅ Aylık ödeme planı üretildi: {sonuc['toplam']} kayıt")
        else:
            logger.info(f"ℹ️ Bu ay için ödeme planı zaten mevcut")
    except Exception as e:
        logger.error(f"Ödeme planı üretim hatası: {e}")
    # FAİZ — startup'ta da bir kez yokla. faiz_hesapla_ve_yaz her kart için
    # kendi kesim/son_odeme döngüsünü değerlendirir, son ödeme tarihi geçen
    # ve faizi henüz yazılmamış kartlara faiz işler. Ayrı "ay sonu" veya
    # "kaçırılan telafi" bloklarına gerek yok — guard hem ileriye hem geriye
    # bakar (downtime sonrası restart de telafi olur).
    try:
        with db() as (conn, cur):
            sonuclar = tum_kartlar_faiz_hesapla(cur)
        yazilan = sum(1 for k in sonuclar if k.get('durum') == 'yazildi')
        if yazilan > 0:
            logger.info(f"✅ Faiz yoklaması: {yazilan} kart için faiz yazıldı")
    except Exception as e:
        logger.warning(f"Faiz yoklama hatası: {e}")

    # Kasa tutarlılık kontrolü — hata vermez, sadece uyarı loglar
    try:
        with db() as (conn, cur):
            cur.execute("SELECT COUNT(*) as sorunlu FROM v_kasa_anomali WHERE durum != 'OK'")
            sorunlu = cur.fetchone()['sorunlu']
            if sorunlu > 0:
                logger.warning(f"⚠️ KASA ANOMALİ: {sorunlu} ciro kaydının kasa karşılığı eksik. /api/kasa-kontrol ile kontrol et.")
            else:
                logger.info("✅ Kasa tutarlılık kontrolü: Tüm ciro kayıtları kasa'ya yansımış.")
    except Exception as e:
        logger.warning(f"Kasa kontrol yapılamadı: {e}")

    # Operasyon event satırları (bugünün ACILIS/KAPANIS) — startup'ta garantile.
    # Böylece dashboard salt-okuma (ensure=False) ile çalışsa bile satırlar hazır olur.
    # is_gunu_tr() startup anında doğru iş gününü döndürür.
    try:
        from sube_operasyon import ensure_events_tum_subeler
        with db() as (conn, cur):
            _n = ensure_events_tum_subeler(cur)
            conn.commit()
        logger.info(f"✅ Operasyon event satırları garanti edildi: {_n} şube")
    except Exception as e:
        logger.warning(f"Operasyon event ensure (startup) hatası: {e}")

    # Jenerik 'Kapak' (kapak_adet) merkez katalogdan kaldırıldı — ayrıştırılmış kapaklar
    # (8oz/14oz/plastik) kalır. Mevcut satırı temizle (idempotent; seed artık eklemiyor).
    try:
        with db() as (conn, cur):
            cur.execute("DELETE FROM merkez_stok_kart WHERE kalem_kodu='kapak_adet'")
            cur.execute("DELETE FROM sube_depo_stok WHERE kalem_kodu='kapak_adet'")
            conn.commit()
    except Exception as e:
        logger.warning(f"kapak_adet katalog temizliği (startup): {e}")

    # Scheduler başlat — restart bağımlılığını kaldırır
    _scheduler_thread = threading.Thread(target=_gece_yarisi_scheduler, daemon=True)
    _scheduler_thread.start()
    logger.info("✅ Gece yarısı scheduler başlatıldı")

    # 🩺 AÇILIŞTA TELAFİ KOŞUSU — atlanan gecenin denetimini tamamlar.
    # (Gece scheduler'ı süreç-içi uyku döngüsü olduğu için her restart bir
    #  geceyi düşürebiliyordu; bu thread o boşluğu kapatır.)
    try:
        _telafi_thread = threading.Thread(target=_akilli_denetim_telafi, daemon=True)
        _telafi_thread.start()
        logger.info("✅ Akıllı Denetim telafi koşusu kuyruğa alındı (90 sn sonra)")
    except Exception as _e:
        logger.warning(f"Telafi koşusu başlatılamadı: {_e}")

    # Kapanış hatırlatma scheduler (gün-içi 15 dk) — sorumluya ~1 saat kala WhatsApp
    try:
        _kap_hat_thread = threading.Thread(target=_kapanis_hatirlatma_scheduler, daemon=True)
        _kap_hat_thread.start()
        logger.info("✅ Kapanış hatırlatma scheduler başlatıldı")
    except Exception as _e:
        logger.warning(f"Kapanış hatırlatma scheduler başlatılamadı: {_e}")

def guncelle_borc_envanteri_odeme_plani_sonrasi(cur, plan: dict, ana_para_kismi: float):
    """Kaynak borc_envanteri ise kalan_vade ve toplam_borc güncelle (panel /ode ve onay kuyruğu ortak)."""
    if plan.get('kaynak_tablo') != 'borc_envanteri' or not plan.get('kaynak_id'):
        return
    cur.execute("SELECT * FROM borc_envanteri WHERE id=%s", (plan['kaynak_id'],))
    borc = cur.fetchone()
    if not borc:
        return
    yeni_kalan = (borc['kalan_vade'] - 1) if borc['kalan_vade'] is not None else None
    yeni_toplam = max(0, float(borc['toplam_borc'] or 0) - ana_para_kismi)
    cur.execute("""
        UPDATE borc_envanteri
        SET kalan_vade = %s,
            toplam_borc = %s
        WHERE id = %s
    """, (yeni_kalan, yeni_toplam, plan['kaynak_id']))


# Ödeme planı kaynağı → kasa işlem türü (TEK KAYNAK — tam ve kısmi ödeme aynı tabloyu kullanır)
KAYNAK_KASA_ISLEM_TURU = {
    'sabit_giderler': 'SABIT_GIDER',
    'personel': 'PERSONEL_MAAS',
    'vadeli_alimlar': 'VADELI_ODEME',
    'borc_envanteri': 'BORC_TAKSIT',
    # 💳 CARİ ÖDEME (2026-08-31): tedarikçi cari hesabına ödeme. Haritada YOKTU
    # ve yazma tarafı varsayılan olarak 'KART_ODEME' kullanıyordu — kasa
    # defterine "karta borç ödedim" diye yazılıyordu, oysa tedarikçiye ödeme.
    'cari_odeme': 'FATURA_ODEMESI',
}

# ⚠️ TEK VARSAYILAN (canlı hata, 2026-08-31): yazma tarafı bilinmeyen kaynakta
# 'KART_ODEME', geri alma tarafı 'ODEME' varsayıyordu. İkisi UYUŞMAYINCA
# `iptal_kasa_hareketi` kaydı BULAMIYOR ve iptal sessizce hiçbir şey yapmıyor —
# para kasadan çıkmış görünmeye devam ediyor. Bu tam olarak kodun kendi
# yorumunda uyardığı hata (bkz. odeme_plani_sil FIX O1), yeni bir kaynak
# eklendiğinde tekrarladı. Artık iki taraf da BU sabiti kullanır.
KASA_ISLEM_TURU_VARSAYILAN = 'ODEME'


def kasa_islem_turu(kaynak_tablo: Optional[str]) -> str:
    """kaynak_tablo → kasa işlem türü. Yazma ve İPTAL aynı yerden okur ki
    ayrışmasınlar (ayrıştıkları gün iptal sessizce çalışmaz)."""
    return KAYNAK_KASA_ISLEM_TURU.get(
        (kaynak_tablo or ''), KASA_ISLEM_TURU_VARSAYILAN)


def _kasa_iptal_turleri(cur, kaynak_id, yedek_tur: str) -> list:
    """🔎 İPTAL EDİLECEK KASA TÜRÜNÜ TAHMİN ETME — DEFTERDEN OKU.

    Yazma ve iptal tarafı türü AYRI AYRI türetince kaçınılmaz olarak sapıyor:
    yazıldı KART_ODEME, arandı ODEME → ters kayıt hiç oluşmuyor ve para
    kasadan çıkmış görünmeye devam ediyor (canlı: 120.000 ₺, 2026-08-31).
    Haritayı düzeltmek YETMEZ — geçmişte farklı türle yazılmış satırlar da
    geri alınabilmeli. Kaydın türü zaten defterde YAZILI.

    Birden çok tür varsa hepsi döner (her biri ayrı ayrı iptal edilir).
    Defterde aktif kayıt yoksa `yedek_tur` döner — o zaman hata mesajını
    `iptal_kasa_hareketi` üretir, sessiz geçilmez.
    """
    cur.execute(
        """SELECT DISTINCT islem_turu FROM kasa_hareketleri
            WHERE kaynak_id=%s AND kasa_etkisi=true AND durum='aktif'""",
        (str(kaynak_id),))
    _t = [str(dict(r)["islem_turu"]) for r in (cur.fetchall() or [])]
    return _t or [yedek_tur]


def kasa_ve_faiz_odeme_plani_tam_odeme(
    cur, plan: dict, plan_id: str, odenen: float, tarih: str,
    anapara_aciklama: Optional[str] = None,
    odeme_yontemi: Optional[str] = None,
    odeyen_sube_id: Optional[str] = None,
) -> float:
    """
    Tam ödeme planı nakit: faiz düşümü + doğru kasa türü (SABIT_GIDER, BORC_TAKSIT, …).
    /ode, onay ODEME_PLANI ve /toplu-odeme aynı fonksiyonu kullanır — tutarsız/çift kasa riski azalır.
    anapara_aciklama: kasa satırı açıklaması (None ise plan.aciklama).
    Dönüş: borç envanteri için anapara kısmı.

    💳 odeme_yontemi ('elden' | 'havale') — 2026-08-17, sahip: "şubelerin kasa
    teslimlerini de bankaya akışını takip edebilirsin".
    O AKIŞ ZATEN KURULU (`/api/banka-mutabakat`: teslim alınan − bankaya yatan −
    ELDEN ödenen = elde nakit) AMA BESLENMİYOR: canlıda sınıflama oranı %11,5;
    7.192.110 ₺ / 236 ödeme "belirsiz" kovasında. Yöntem bilinmeyince sistem
    paranın ÇEKMECEDEN mi BANKADAN mı çıktığını bilemiyor ve elde nakit
    −188.303 ₺ gibi imkânsız bir sonuç veriyor.
    Bu parametre o boşluğu KAYNAKTA kapatır: her merkezî ödeme (BORC_TAKSIT,
    PERSONEL_MAAS, SABIT_GIDER, FATURA_ODEMESI, VADELI_ODEME) yöntemini kasa
    hareketine taşır. None geçilirse eski davranış birebir sürer (belirsiz) —
    üç çağıran da kırılmaz.
    """
    odenen = float(odenen)
    # Yöntem yalnız BİLİNEN iki değerden biri olabilir; başka bir şey gelirse
    # None'a düşer. Uydurma etiket, banka mutabakatını sessizce yanlış yönde
    # kaydırırdı (elde nakit hesabı doğrudan bu alandan türüyor).
    _yontem = (odeme_yontemi or '').strip().lower() or None
    if _yontem not in ('elden', 'havale', None):
        _yontem = None
    faiz_kismi = 0.0
    if plan.get('kart_id'):
        cur.execute("""
            SELECT COALESCE(SUM(tutar), 0) as bekleyen_faiz
            FROM kart_hareketleri
            WHERE kart_id=%s AND islem_turu='FAIZ' AND durum='aktif'
        """, (plan['kart_id'],))
        bekleyen_faiz = float(cur.fetchone()['bekleyen_faiz'])
        faiz_kismi = min(bekleyen_faiz, odenen)

    ana_para_kismi = odenen - faiz_kismi

    if faiz_kismi > 0:
        insert_kasa_hareketi(cur, tarih, 'KART_FAIZ', -abs(faiz_kismi),
            f"Kart faiz ödemesi: {plan['aciklama']}", 'odeme_plani', plan_id,
            f"{plan_id}_faiz", 'KART_FAIZ', odeme_yontemi=_yontem)
        kalan_faiz_kapatilacak = faiz_kismi
        # FOR UPDATE: eş zamanlı iki ödeme isteğinde aynı faiz satırlarının
        # çakışmasını önler — satırlar bu transaction bitene kadar kilitlenir.
        cur.execute("""
            SELECT id, tutar FROM kart_hareketleri
            WHERE kart_id=%s AND islem_turu='FAIZ' AND durum='aktif'
            ORDER BY tarih ASC
            FOR UPDATE
        """, (plan['kart_id'],))
        faiz_kayitlari = cur.fetchall()
        for fk in faiz_kayitlari:
            if kalan_faiz_kapatilacak <= 0:
                break
            fk_tutar = float(fk['tutar'])
            if fk_tutar <= kalan_faiz_kapatilacak:
                cur.execute("UPDATE kart_hareketleri SET durum='iptal' WHERE id=%s", (fk['id'],))
                kalan_faiz_kapatilacak -= fk_tutar
            else:
                cur.execute("UPDATE kart_hareketleri SET durum='iptal' WHERE id=%s", (fk['id'],))
                kalan_tutar = fk_tutar - kalan_faiz_kapatilacak
                cur.execute("""INSERT INTO kart_hareketleri
                    (id, kart_id, tarih, islem_turu, tutar, aciklama)
                    VALUES (%s, %s, %s, 'FAIZ', %s, 'Kısmi faiz bakiyesi')
                """, (str(uuid.uuid4()), plan['kart_id'], tarih, kalan_tutar))
                kalan_faiz_kapatilacak = 0

    if ana_para_kismi > 0:
        aciklama_ana = anapara_aciklama if anapara_aciklama is not None else plan['aciklama']
        kaynak = plan.get('kaynak_tablo') or ''
        islem_t = kasa_islem_turu(kaynak)   # iptal tarafiyla AYNI kaynak
        # 🏪 ŞUBE DAMGASI KAYNAKTA (2026-08-09): maaş/sabit gider ödemesi kasaya
        # şubesiz yazılıyordu; şube kârlılığı personel maliyeti OLMADAN
        # hesaplanıyordu (250.487 ₺ "atanmamış" kovasında birikmişti). Geçmiş
        # açıklamadaki addan türetildi ama bu kırılgan bir yol — kaynağı
        # damgalamak kalıcı çözüm. Çözülemezse None kalır, uydurma yapılmaz.
        _plan_sube = None
        try:
            if kaynak == 'personel' and plan.get('kaynak_id'):
                cur.execute("SELECT sube_id FROM personel WHERE id::text=%s",
                            (str(plan['kaynak_id']),))
                _pr = cur.fetchone()
                _plan_sube = (dict(_pr).get('sube_id') if _pr else None) or None
            elif kaynak == 'sabit_giderler' and plan.get('kaynak_id'):
                cur.execute("SELECT sube_id FROM sabit_giderler WHERE id::text=%s",
                            (str(plan['kaynak_id']),))
                _sr = cur.fetchone()
                _plan_sube = (dict(_sr).get('sube_id') if _sr else None) or None
        except Exception as _e:  # noqa: BLE001 — şube çözümü ödemeyi ASLA kilitlemez
            logging.getLogger(__name__).warning(
                "plan sube damgasi cozulemedi: %s", str(_e)[:110])
        if kaynak == 'vadeli_alimlar':
            # vadeli: kasa kaydı kaynağı vadeli_alimlar'a bağlanır (onay guard ile aynı anahtar)
            vk = plan.get('kaynak_id')
            kasa_kt = 'vadeli_alimlar' if vk else 'odeme_plani'
            kasa_kid = vk or plan_id
            insert_kasa_hareketi(
                cur, tarih, islem_t, -abs(ana_para_kismi), aciklama_ana,
                kasa_kt, kasa_kid, plan_id, 'ODEME_PLANI', sube_id=_plan_sube,
                odeme_yontemi=_yontem,
            )
        else:
            insert_kasa_hareketi(cur, tarih, islem_t, -abs(ana_para_kismi),
                aciklama_ana, 'odeme_plani', plan_id, plan_id, 'ODEME_PLANI',
                sube_id=_plan_sube, odeme_yontemi=_yontem)

        # 🔁 ÇAPRAZ ÖDEME → OTOMATİK ŞUBE BORCU (sahip kararı 2026-08-18:
        # "ödemesini yaparken BORÇ OLARAK DA YAZSIN ve bu iyi bir şey bence!")
        # Parayı ÖDEYEN şube, giderin sahibi değilse (ör. Gazze, Alsancak'ın
        # kredisini ödedi) aradaki ilişki görünür bir borca dönüşür.
        # Şişme riski netleştirmeyle çözüldü — bkz. capraz_odeme_borcu_kur.
        # HATA-YUTAR: borç kaydı düşse bile ÖDEME asla kilitlenmez.
        _odeyen = (odeyen_sube_id or "").strip() or None
        if _odeyen and _plan_sube and _odeyen != _plan_sube:
            try:
                from sube_ici_borc_api import capraz_odeme_borcu_kur
                capraz_odeme_borcu_kur(
                    cur, _odeyen, _plan_sube, ana_para_kismi, tarih,
                    aciklama_ana or "ödeme", plan_id)
            except Exception as _e:  # noqa: BLE001
                logging.getLogger(__name__).warning(
                    "capraz odeme borcu kurulamadi: %s", str(_e)[:120])

    # kart_borc() ODEME türündeki kart_hareketleri kaydına bakarak borcu düşürür.
    # Nakit ödeme kasaya gider ama kart borcu bu kayıt olmadan hiç azalmaz.
    # Her kart_id'li plan ödemesinde ODEME kaydı oluşturulmalı.
    if plan.get('kart_id') and odenen > 0:
        cur.execute("""
            INSERT INTO kart_hareketleri
                (id, kart_id, tarih, islem_turu, tutar, aciklama, kaynak_id, kaynak_tablo)
            VALUES (%s, %s, %s, 'ODEME', %s, %s, %s, 'odeme_plani')
            ON CONFLICT DO NOTHING
        """, (
            f"odm_{plan_id}",
            plan['kart_id'],
            tarih,
            abs(odenen),
            f"Ödeme planı: {plan.get('aciklama', '')}",
            plan_id,
        ))

    return ana_para_kismi


# ── PANEL ──────────────────────────────────────────────────────

# ── AY DEVİR (HESAPLANAN — ledger'a yazılmaz) ──────────────────
def devir_hesapla(yil: int = None, ay: int = None):
    """
    Geçen ayın kapanış kasasını hesaplar.
    Ledger'a hiçbir şey yazılmaz — immutable model korunur.
    """
    import calendar
    bugun = bugun_tr()
    yil = yil or bugun.year
    ay  = ay  or bugun.month

    if ay == 1:
        gecen_yil, gecen_ay = yil - 1, 12
    else:
        gecen_yil, gecen_ay = yil, ay - 1

    gecen_ay_son = date(gecen_yil, gecen_ay,
                        calendar.monthrange(gecen_yil, gecen_ay)[1])

    with db() as (conn, cur):
        devir = kasa_bakiyesi_tarihte(cur, gecen_ay_son)

    return {
        "devir_tutar": devir,
        "gecen_ay": f"{gecen_yil}-{gecen_ay:02d}",
        "hesaplandi": True
    }

@app.get("/api/devir")
def devir_goster(yil: int = None, ay: int = None):
    try:
        return devir_hesapla(yil, ay)
    except Exception as e:
        raise HTTPException(500, str(e))


def kart_plan_mutabakat(uygula: bool = False) -> dict:
    """F1 — KART planları self-heal (2026-07-09 derin inceleme; sahip onayı):
    borc_plan_mutabakat yalnız borc_envanteri kapsıyordu, KART planları başıboştu.
      R1 MÜKERRER-AYNI-AY: aynı kart+ay >1 aktif plan → en yeni tarihli kalır,
         diğerleri 'iptal' (çift "ödenecek" kaybolur).
      R2 BAYAT-TAŞIMA: aynı kartta ÖNCEKİ ayda eş-tutarlı (±1) plan varken, kendi
         ayına ait ekstre snapshot'ı OLMAYAN sonraki-ay planı → 'iptal'
         (motorun son-snapshot fallback'i bayat borcu yeni aya kopyalıyordu —
         3018 ve Ziraat 2× vakalarının kökü).
      R3 GEÇMİŞ-AY KAPAMA: plan ayı geçmişte + o ay kart defterinde (kart_hareketleri
         ODEME) toplam ≥ asgari → 'odendi' (tam/asgari notu). İZ YOKSA PLAN KALIR —
         kasa izi = tek gerçek; elle aklama yok.
    uygula=False → yalnız liste döner (kuru çalıştırma, hiçbir şey değişmez).
    İdempotent: ikinci koşuş aynı sonuçları bulmaz (durumlar değişmiş olur)."""
    rapor = {"uygula": uygula, "r1_mukerrer": [], "r2_bayat_tasima": [],
             "r3_kapama": [], "dokunulmayan_bekleyen": [], "hata": None}
    try:
        with db() as (conn, cur):
            cur.execute(
                """SELECT op.id, op.kart_id, k.kart_adi, op.tarih::text AS tarih,
                          TO_CHAR(op.tarih, 'YYYY-MM') AS ay,
                          COALESCE(op.odenecek_tutar,0)::float AS odenecek,
                          COALESCE(op.asgari_tutar,0)::float AS asgari, op.durum
                   FROM odeme_plani op JOIN kartlar k ON k.id = op.kart_id
                   WHERE op.kart_id IS NOT NULL
                     AND op.durum IN ('bekliyor','onay_bekliyor')
                   ORDER BY op.kart_id, TO_CHAR(op.tarih,'YYYY-MM'),
                            op.tarih DESC, op.id"""
            )
            planlar = [dict(r) for r in cur.fetchall() or []]

            # R1b — aynı kart+ay'da ÖDENMİŞ plan varken hâlâ 'bekliyor' satır
            # (2026-07-09 döngü canlı dersi: ekstre-yukle UPDATE yalnız bekleyeni
            # arıyor, satır 'odendi' olunca YENİ bekleyen INSERT ediyordu — HEPSI/
            # WORLD/7015 hem ödendi hem bekliyor görünüyordu)
            iptal_ids = set()
            cur.execute(
                """SELECT op.id, k.kart_adi, TO_CHAR(op.tarih,'YYYY-MM') AS ay,
                          COALESCE(op.odenecek_tutar,0)::float AS tutar
                   FROM odeme_plani op JOIN kartlar k ON k.id = op.kart_id
                   WHERE op.kart_id IS NOT NULL
                     AND op.durum IN ('bekliyor','onay_bekliyor')
                     AND EXISTS (SELECT 1 FROM odeme_plani o2
                                 WHERE o2.kart_id = op.kart_id AND o2.durum = 'odendi'
                                   AND DATE_TRUNC('month', o2.tarih) = DATE_TRUNC('month', op.tarih))"""
            )
            for r in [dict(x) for x in cur.fetchall() or []]:
                iptal_ids.add(r["id"])
                rapor.setdefault("r1b_odendi_yaninda_bekleyen", []).append(
                    {"kart": r["kart_adi"], "ay": r["ay"], "tutar": r["tutar"],
                     "plan_id": r["id"][:8]})

            # R1 — aynı kart+ay mükerrer
            gruplar = {}
            for p in planlar:
                gruplar.setdefault((p["kart_id"], p["ay"]), []).append(p)
            for (_kid, _ay), grp in gruplar.items():
                for fazla in grp[1:]:
                    iptal_ids.add(fazla["id"])
                    rapor["r1_mukerrer"].append(
                        {"kart": fazla["kart_adi"], "ay": fazla["ay"],
                         "tutar": fazla["odenecek"], "plan_id": fazla["id"][:8]})

            kalanlar = [p for p in planlar if p["id"] not in iptal_ids]

            # R2 — bayat-snapshot taşıması (önceki ayda eş-tutar + kendi ayında ekstre yok)
            for p in kalanlar:
                onceki_es = [q for q in planlar
                             if q["kart_id"] == p["kart_id"] and q["ay"] < p["ay"]
                             and abs(q["odenecek"] - p["odenecek"]) <= 1.0
                             and q["id"] not in iptal_ids]
                if not onceki_es:
                    continue
                cur.execute(
                    """SELECT 1 FROM kart_ekstre_donem
                       WHERE kart_id = %s AND TO_CHAR(donem,'YYYY-MM') = %s LIMIT 1""",
                    (p["kart_id"], p["ay"]))
                if cur.fetchone():
                    continue  # kendi ayının gerçek ekstresi var — taşıma değil
                iptal_ids.add(p["id"])
                rapor["r2_bayat_tasima"].append(
                    {"kart": p["kart_adi"], "ay": p["ay"], "tutar": p["odenecek"],
                     "es_tutarli_onceki_ay": onceki_es[0]["ay"], "plan_id": p["id"][:8]})

            if uygula and iptal_ids:
                cur.execute(
                    """UPDATE odeme_plani
                       SET durum='iptal',
                           aciklama = COALESCE(aciklama,'') || ' [kart-mutabakat iptali]'
                       WHERE id = ANY(%s)""", (list(iptal_ids),))

            # R3 — geçmiş-ay kapama (kart defteri ODEME izi ile)
            for p in kalanlar:
                if p["id"] in iptal_ids:
                    continue
                if p["ay"] >= date.today().strftime("%Y-%m"):
                    continue  # bu ay/gelecek — motorun işi, dokunma
                cur.execute(
                    """SELECT COALESCE(SUM(tutar),0)::float AS odenen,
                              MAX(tarih)::text AS son_odeme_gunu
                       FROM kart_hareketleri
                       WHERE kart_id = %s AND durum='aktif' AND islem_turu='ODEME'
                         AND TO_CHAR(tarih,'YYYY-MM') = %s""",
                    (p["kart_id"], p["ay"]))
                oz = dict(cur.fetchone() or {})
                odenen = float(oz.get("odenen") or 0)
                asgari = p["asgari"] or p["odenecek"]
                if odenen >= p["odenecek"] - 0.01:
                    tur = "tam"
                elif asgari > 0 and odenen >= asgari * 0.999:
                    tur = "asgari"
                else:
                    rapor["dokunulmayan_bekleyen"].append(
                        {"kart": p["kart_adi"], "ay": p["ay"], "tutar": p["odenecek"],
                         "o_ay_odenen": odenen,
                         "not": "iz yetersiz — plan KALIR (borç gerçeği)"})
                    continue
                rapor["r3_kapama"].append(
                    {"kart": p["kart_adi"], "ay": p["ay"], "tutar": p["odenecek"],
                     "odenen": odenen, "tur": tur, "plan_id": p["id"][:8]})
                if uygula:
                    ek = (" [kart-mutabakat: %s ödeme iziyle kapatıldı]" % tur
                          if tur == "tam" else
                          " [kart-mutabakat: ASGARİ ödendi (%s) — kalan sonraki döneme]" % odenen)
                    cur.execute(
                        """UPDATE odeme_plani
                           SET durum='odendi', odenen_tutar=%s,
                               odeme_tarihi=COALESCE(odeme_tarihi, %s::date),
                               aciklama = COALESCE(aciklama,'') || %s
                           WHERE id=%s AND durum IN ('bekliyor','onay_bekliyor')""",
                        (odenen, oz.get("son_odeme_gunu") or p["tarih"], ek, p["id"]))
            if uygula:
                conn.commit()
    except Exception as e:  # noqa: BLE001
        rapor["hata"] = str(e)[:200]
        logger.warning("kart_plan_mutabakat: %s", str(e)[:200])
    # R5 SAHTE-ÖDENDİ ONARIMI (2026-07-26, '4 kart ÖM'den kayboldu' vakası):
    # snapshot asgarisi 0/boş gelince plan yazıcı 'asgari ödendi (0)' ile planı
    # ÖDEMESİZ kapatıyordu; 'odendi' damga hem borcu gizler hem yeni planın
    # INSERT'ini bloke eder. KASA İZİ = TEK GERÇEK: 'odendi' ama odenen<=0 VE o ay
    # kart defterinde ODEME izi yoksa damga sahtedir → 'iptal' (R4 yeniden açar).
    try:
        with db() as (conn, cur):
            cur.execute(
                """UPDATE odeme_plani op
                   SET durum='iptal',
                       aciklama = COALESCE(op.aciklama,'') || ' [sahte-odendi onarımı: ödeme izi yok]'
                   WHERE op.kart_id IS NOT NULL AND op.durum='odendi'
                     AND COALESCE(op.odenen_tutar,0) <= 0
                     AND NOT EXISTS (
                         SELECT 1 FROM kart_hareketleri kh
                         WHERE kh.kart_id = op.kart_id AND kh.durum='aktif'
                           AND kh.islem_turu='ODEME'
                           AND DATE_TRUNC('month', kh.tarih) = DATE_TRUNC('month', op.tarih))
                   RETURNING op.id""" if uygula else
                """SELECT op.id FROM odeme_plani op
                   WHERE op.kart_id IS NOT NULL AND op.durum='odendi'
                     AND COALESCE(op.odenen_tutar,0) <= 0
                     AND NOT EXISTS (
                         SELECT 1 FROM kart_hareketleri kh
                         WHERE kh.kart_id = op.kart_id AND kh.durum='aktif'
                           AND kh.islem_turu='ODEME'
                           AND DATE_TRUNC('month', kh.tarih) = DATE_TRUNC('month', op.tarih))""")
            rapor["r5_sahte_odendi_iptal"] = len(cur.fetchall() or [])
            if uygula:
                conn.commit()
    except Exception as e:  # noqa: BLE001
        rapor["hata"] = (rapor.get("hata") or "") + f" r5: {str(e)[:100]}"

    # R4 (2026-07-10 sahip vakası): bu döneme ait EKSTRE SNAPSHOT'ı olan ama aktif
    # planı olmayan kart (eski-odendi blokajı kurbanı) → tek yazıcıyla planı AÇ.
    try:
        from finans_core import kesim_tarihi_hesapla
        from kasa_service import kart_kesim_plani_yaz_tx
        with db() as (conn, cur):
            cur.execute("SELECT * FROM kartlar WHERE aktif=TRUE")
            for k in [dict(r) for r in cur.fetchall() or []]:
                try:
                    bugun = date.today()
                    bu_kesim = kesim_tarihi_hesapla(bugun.year, bugun.month,
                                                    int(k["kesim_gunu"]))
                    if bugun < bu_kesim:
                        continue  # kesim henüz gelmedi
                    cur.execute(
                        """SELECT 1 FROM kart_ekstre_donem
                           WHERE kart_id=%s AND donem=DATE_TRUNC('month',%s::date) LIMIT 1""",
                        (k["id"], str(bu_kesim)))
                    if not cur.fetchone():
                        continue  # bu dönemin ekstresi yok — F3 freni alanı
                    cur.execute(
                        """SELECT 1 FROM odeme_plani
                           WHERE kart_id=%s AND durum IN ('bekliyor','onay_bekliyor')
                             AND tarih >= %s::date LIMIT 1""",
                        (k["id"], str(bu_kesim)))
                    if cur.fetchone():
                        continue  # aktif plan zaten var
                    rapor.setdefault("r4_plan_acilan", []).append(
                        {"kart": k.get("kart_adi"), "kesim": str(bu_kesim)})
                    if uygula:
                        kart_kesim_plani_yaz_tx(cur, k, bugun.year, bugun.month)
                except Exception as e:  # noqa: BLE001
                    logger.warning("R4 %s: %s", k.get("kart_adi"), str(e)[:60])
            if uygula:
                conn.commit()
    except Exception as e:  # noqa: BLE001
        rapor["hata"] = (rapor.get("hata") or "") + f" r4: {str(e)[:100]}"
    rapor["ozet"] = {"r4": len(rapor.get("r4_plan_acilan") or []),
                     "r1b": len(rapor.get("r1b_odendi_yaninda_bekleyen") or []),
                     "r1": len(rapor["r1_mukerrer"]), "r2": len(rapor["r2_bayat_tasima"]),
                     "r3": len(rapor["r3_kapama"]),
                     "r5": rapor.get("r5_sahte_odendi_iptal", 0),
                     "kalan_bekleyen": len(rapor["dokunulmayan_bekleyen"])}
    return rapor


@app.post("/api/kartlar/plan-mutabakat")
def kart_plan_mutabakat_uc(body: dict = None):
    """F1 elle tetik: body.uygula=false → kuru çalıştırma listesi; true → uygular.
    Doğrula-önce-düzelt: önce kuru sonuç incelenir, sonra uygulanır."""
    uygula = bool((body or {}).get("uygula"))
    return kart_plan_mutabakat(uygula=uygula)


def borc_plan_mutabakat(referans_tarih: Optional[date] = None) -> dict:
    """Borç ödeme planı self-healing mutabakatı (idempotent — her panel açılışında güvenli).

    Üç düzeltme yapar:
      1) referans_ay backfill: eski borç planlarına referans_ay yaz (mükerrer index kapsasın).
      2) Ödenmiş kapat: kasa'da o ay BORC_TAKSIT ödemesi olan ama hâlâ bekleyen plan → 'odendi'.
         (Borç plandan ÖNCE ödendiyse /ode kapatamamıştı; üretici sonra bekliyor satır üretmişti.)
      3) Mükerrer iptal: aynı borç+ay için >1 aktif satır → en iyisini tut (ödenmiş > en yeni),
         diğerlerini 'iptal'. Böylece çift "ödenecek" kaybolur ve unique index kurulabilir.
    """
    sonuc = {"backfill": 0, "kapatilan": 0, "mukerrer_iptal": 0, "baslangic_oncesi_iptal": 0, "hata": None}

    # F1 KART BÖLÜMÜ (2026-07-09, kuru çalıştırma doğrulandı → zincire bağlandı):
    # kart planları da self-heal kapsamında — mükerrer/bayat-taşıma/geçmiş-ay kapama.
    try:
        kt = kart_plan_mutabakat(uygula=True)
        sonuc["kart"] = kt.get("ozet")
    except Exception as e:
        sonuc["hata"] = f"kart: {e}"; logger.warning(f"borc_plan_mutabakat kart: {e}")

    # 0) Başlangıç tarihinden ÖNCEKİ bekleyen borç planını iptal et (örn. araba ilk taksit
    #    1 Temmuz ise yanlışlıkla üretilmiş 1 Haziran satırı temizlenir).
    try:
        with db() as (conn, cur):
            cur.execute("""
                UPDATE odeme_plani op
                SET durum='iptal',
                    aciklama = COALESCE(op.aciklama,'') || ' [başlangıçtan-önce-iptal]'
                FROM borc_envanteri b
                WHERE op.kaynak_tablo='borc_envanteri' AND op.kaynak_id = b.id::text
                  AND op.durum IN ('bekliyor','onay_bekliyor')
                  AND b.baslangic_tarihi IS NOT NULL
                  AND op.tarih < b.baslangic_tarihi
            """)
            sonuc["baslangic_oncesi_iptal"] = cur.rowcount or 0
    except Exception as e:
        sonuc["hata"] = f"baslangic: {e}"; logger.warning(f"borc_plan_mutabakat baslangic: {e}")

    # 1) Ödenmiş ama bekleyen borç planını kapat. İKİ eşleşme dalı:
    #    (a) AYNI AY kasa BORC_TAKSIT izi (mevcut kural);
    #    (b) KOÇ FİNANS vakası (2026-07-15, sahip: 'son ödemesi yapıldı ama
    #        bekleyende görünüyor'): ay sonuna yakın vadeli taksit birkaç gün
    #        GEÇ ödenince ay değişiyor, ay-kuralı ıskalıyordu (plan 28.06,
    #        ödeme 04.07). GEÇ ÖDEME PENCERESİ: vade..vade+35g içinde,
    #        kuruşuna aynı tutarlı iz de planı kapatır (kasa izi=tek gerçek).
    try:
        with db() as (conn, cur):
            cur.execute("""
                UPDATE odeme_plani op
                SET durum='odendi',
                    odenen_tutar = COALESCE(op.odenen_tutar, op.odenecek_tutar),
                    odeme_tarihi = COALESCE(op.odeme_tarihi, (
                        SELECT MAX(kh.tarih) FROM kasa_hareketleri kh
                        WHERE kh.kaynak_tablo='borc_envanteri' AND kh.kaynak_id=op.kaynak_id
                          AND kh.islem_turu='BORC_TAKSIT' AND kh.kasa_etkisi=TRUE AND kh.durum='aktif'
                          AND (DATE_TRUNC('month', kh.tarih)=DATE_TRUNC('month', op.tarih)
                               OR (kh.tarih BETWEEN op.tarih AND op.tarih + INTERVAL '35 days'
                                   AND ABS(ABS(kh.tutar) - op.odenecek_tutar) <= 1))))
                WHERE op.kaynak_tablo='borc_envanteri'
                  AND op.durum IN ('bekliyor','onay_bekliyor')
                  AND EXISTS (
                      SELECT 1 FROM kasa_hareketleri kh
                      WHERE kh.kaynak_tablo='borc_envanteri' AND kh.kaynak_id=op.kaynak_id
                        AND kh.islem_turu='BORC_TAKSIT' AND kh.kasa_etkisi=TRUE AND kh.durum='aktif'
                        AND (DATE_TRUNC('month', kh.tarih)=DATE_TRUNC('month', op.tarih)
                             OR (kh.tarih BETWEEN op.tarih AND op.tarih + INTERVAL '35 days'
                                 AND ABS(ABS(kh.tutar) - op.odenecek_tutar) <= 1)))
            """)
            sonuc["kapatilan"] = cur.rowcount or 0
    except Exception as e:
        sonuc["hata"] = f"kapat: {e}"; logger.warning(f"borc_plan_mutabakat kapat: {e}")

    # 2) Mükerrer iptal — aynı borç+ay için 1 aktif satır kalsın (ödenmiş > en yeni)
    try:
        with db() as (conn, cur):
            cur.execute("""
                WITH ranked AS (
                    SELECT id, ROW_NUMBER() OVER (
                        PARTITION BY kaynak_id, DATE_TRUNC('month', tarih)
                        ORDER BY (CASE WHEN durum='odendi' THEN 0 ELSE 1 END), olusturma DESC
                    ) AS rn
                    FROM odeme_plani
                    WHERE kaynak_tablo='borc_envanteri' AND durum <> 'iptal' AND kaynak_id IS NOT NULL
                )
                UPDATE odeme_plani op
                SET durum='iptal',
                    aciklama = COALESCE(op.aciklama,'') || ' [mükerrer-iptal]'
                FROM ranked r
                WHERE op.id = r.id AND r.rn > 1
            """)
            sonuc["mukerrer_iptal"] = cur.rowcount or 0
    except Exception as e:
        sonuc["hata"] = f"mukerrer: {e}"; logger.warning(f"borc_plan_mutabakat mukerrer: {e}")

    # 3) referans_ay backfill (mükerrer index borç satırlarını da korusun) — ÇAKIŞMA KORUMALI:
    #    aynı borç+ay için zaten referans_ay'lı aktif satır varsa atla (unique index patlamasın).
    try:
        with db() as (conn, cur):
            cur.execute("""
                UPDATE odeme_plani op
                SET referans_ay = DATE_TRUNC('month', op.tarih)::date
                WHERE op.kaynak_tablo='borc_envanteri' AND op.referans_ay IS NULL AND op.tarih IS NOT NULL
                  AND NOT EXISTS (
                      SELECT 1 FROM odeme_plani o2
                      WHERE o2.kaynak_tablo='borc_envanteri' AND o2.kaynak_id=op.kaynak_id
                        AND o2.durum <> 'iptal' AND o2.id <> op.id
                        AND o2.referans_ay = DATE_TRUNC('month', op.tarih)::date
                  )
            """)
            sonuc["backfill"] = cur.rowcount or 0
    except Exception as e:
        sonuc["hata"] = f"backfill: {e}"; logger.warning(f"borc_plan_mutabakat backfill: {e}")
    # ÇALIŞMA NABZI (2026-07-06): panel her açılışta koşar — nabız çok sık olmasın diye
    # saatte en fazla 1 yazılır (son nabız <1 saatse atla). Hata-yutar.
    try:
        from duyu_omurga import duyu_nabiz_yaz
        with db() as (_, _c):
            _c.execute(
                "SELECT 1 FROM duyu_nabiz WHERE duyu='borc_plan_selfheal' AND run_ts >= NOW() - INTERVAL '1 hour' LIMIT 1"
            )
            _son_var = _c.fetchone() is not None
        if not _son_var:
            duyu_nabiz_yaz("borc_plan_selfheal", taranan=1,
                           uretilen=sum(v for v in sonuc.values() if isinstance(v, int) and v > 0))
    except Exception:  # noqa: BLE001
        pass
    # DUYU OMURGASI kancası (2026-07-06): self-heal bir şey düzelttiyse günlük idempotent olay
    try:
        _duzeltme = sum(v for v in sonuc.values() if isinstance(v, int) and v > 0)
        if _duzeltme:
            from duyu_omurga import duyu_olay_yaz
            duyu_olay_yaz(
                "borc_plan_selfheal", "finans.plan.kasa_izi_esitleme",
                f"borc-mutabakat:{date.today().isoformat()}",
                entity_scope="genel", signal_name="Borç planı self-heal düzeltmesi",
                evidence_class="mutabakat",
                payload={k: v for k, v in sonuc.items() if isinstance(v, (int, float, str))},
            )
    except Exception:  # noqa: BLE001
        pass
    return sonuc


def odeme_plani_kasa_mutabakat() -> dict:
    """KURAL (kullanıcı, 2026-07-04): 'Kasada ödeme izi varsa ödenmiştir, yoksa ödenmemiştir.'
    Self-healing (idempotent — her panel açılışında güvenli): kasada ödeme izi olan ama hâlâ
    'bekliyor'/'onay_bekliyor' kalmış ödeme planlarını 'odendi' yapar. İz YOKSA DOKUNMAZ —
    plan gerçek borç olarak görünmeye devam eder (ters yön bilerek kapalı: kartla ödenen
    planların kasa izi olmaz, geri açmak yanlış borç diriltir).

    Eşleşme iki türlü:
      (a) PLAN-BAĞLI iz: kh.kaynak_tablo='odeme_plani' AND kh.kaynak_id=op.id — AY FİLTRESİZ.
          Bağ bire bir olduğu için hangi ayda ödendiği fark etmez (POS Temmuz vakasının kökü:
          ay filtresi yüzünden farklı aydaki doğrudan iz görülmüyordu).
      (b) KAYNAK-BAĞLI iz: kh.kaynak = op.kaynak (sabit/personel/vadeli/borç) — SADECE plan
          ayıyla aynı ay (tekrarlayan kalemlerde başka ayın ödemesi bu ayın planını kapatmasın).
    Kart planları (kart_id'li) hariç — kart ödemeleri kart_hareketleri ekseninde yürür.
    """
    sonuc = {"kapatilan": 0, "hata": None}
    _iz_kosulu = """
        (kh.kaynak_tablo = 'odeme_plani' AND kh.kaynak_id = op.id)
        OR (op.kaynak_id IS NOT NULL
            AND kh.kaynak_tablo = op.kaynak_tablo AND kh.kaynak_id = op.kaynak_id
            AND DATE_TRUNC('month', kh.tarih) = DATE_TRUNC('month', op.tarih))
    """
    try:
        with db() as (conn, cur):
            cur.execute(f"""
                UPDATE odeme_plani op
                SET durum='odendi',
                    odenen_tutar = COALESCE(op.odenen_tutar, op.odenecek_tutar),
                    odeme_tarihi = COALESCE(op.odeme_tarihi, (
                        SELECT MAX(kh.tarih) FROM kasa_hareketleri kh
                        WHERE kh.kasa_etkisi=TRUE AND kh.durum='aktif' AND ({_iz_kosulu})))
                WHERE op.kart_id IS NULL
                  AND op.durum IN ('bekliyor','onay_bekliyor')
                  AND EXISTS (
                      SELECT 1 FROM kasa_hareketleri kh
                      WHERE kh.kasa_etkisi=TRUE AND kh.durum='aktif' AND ({_iz_kosulu}))
            """)
            sonuc["kapatilan"] = cur.rowcount or 0
            if sonuc["kapatilan"]:
                # kapananların bekleyen onaylarını da kapat
                cur.execute("""
                    UPDATE onay_kuyrugu ok SET durum='onaylandi', onay_tarihi=NOW()
                    WHERE ok.durum='bekliyor'
                    AND EXISTS (
                        SELECT 1 FROM odeme_plani op
                        WHERE op.durum='odendi'
                          AND (ok.kaynak_id = op.id::text OR
                               (ok.kaynak_tablo = op.kaynak_tablo AND ok.kaynak_id = op.kaynak_id::text))
                    )
                """)
    except Exception as e:
        sonuc["hata"] = str(e)
        logger.warning(f"odeme_plani_kasa_mutabakat: {e}")
    return sonuc


def odeme_plani_kontrol(referans_tarih: Optional[date] = None) -> dict:
    """
    Ay plan üretimi için lazy + idempotent koruma.
    Panel çağrısında tetiklenir; eksik plan varsa üretmeyi dener.
    """
    bugun = referans_tarih or bugun_tr()
    # Self-healing: ödenmiş ama bekleyen / mükerrer borç planlarını düzelt (idempotent)
    borc_mutabakat = borc_plan_mutabakat(bugun)
    # Self-healing (genel): kasa izi olan bekleyen planları kapat — "kasa izi = tek gerçek"
    kasa_mutabakat = odeme_plani_kasa_mutabakat()
    eksik_sabit = eksik_borc = eksik_kart = eksik_personel = 0
    lock_ok = False

    with db() as (conn, cur):
        cur.execute(
            "SELECT pg_try_advisory_xact_lock(hashtext(%s)) AS ok",
            (f"odeme-plan-kontrol:{bugun.year}-{bugun.month}",),
        )
        lock_ok = bool((cur.fetchone() or {}).get("ok"))

        # Sabit gider planı eksik mi?
        cur.execute("""
            SELECT COUNT(*) as eksik FROM sabit_giderler sg
            WHERE sg.aktif = TRUE AND (sg.tip IS NULL OR sg.tip = 'sabit')
            AND NOT EXISTS (
                SELECT 1 FROM odeme_plani op
                WHERE op.kaynak_tablo = 'sabit_giderler'
                AND op.kaynak_id = sg.id
                AND op.durum != 'iptal'
                AND op.referans_ay = DATE_TRUNC('month', %s::date)
            )
        """, (bugun,))
        eksik_sabit = int(cur.fetchone()['eksik'])

        # Borç taksit planı eksik mi?
        cur.execute("""
            SELECT COUNT(*) as eksik FROM borc_envanteri b
            WHERE b.aktif = TRUE AND b.aylik_taksit > 0
            AND (b.kalan_vade IS NULL OR b.kalan_vade > 0)
            AND NOT EXISTS (
                SELECT 1 FROM odeme_plani op
                WHERE op.kaynak_tablo = 'borc_envanteri'
                AND op.kaynak_id = b.id::text
                AND op.durum != 'iptal'
                AND DATE_TRUNC('month', op.tarih) = DATE_TRUNC('month', %s::date)
            )
        """, (bugun,))
        eksik_borc = int(cur.fetchone()['eksik'])

        # Kart asgari ödeme planı eksik mi? (borcu olan aktif kartlar)
        cur.execute("""
            SELECT COUNT(*) as eksik FROM kartlar k
            WHERE k.aktif = TRUE
            AND (
                SELECT COALESCE(SUM(
                    CASE WHEN kh.islem_turu IN ('HARCAMA','FAIZ','DEVIR') THEN kh.tutar
                         WHEN kh.islem_turu='ODEME' THEN -kh.tutar ELSE 0 END
                ), 0) FROM kart_hareketleri kh
                WHERE kh.kart_id = k.id AND kh.durum = 'aktif'
            ) > 0
            AND NOT EXISTS (
                SELECT 1 FROM odeme_plani op
                WHERE op.kart_id = k.id
                AND op.durum != 'iptal'
                AND DATE_TRUNC('month', op.tarih) = DATE_TRUNC('month', %s::date)
            )
        """, (bugun,))
        eksik_kart = int(cur.fetchone()['eksik'])

        # Sürekli personel maaş planı eksik mi?
        cur.execute("""
            SELECT COUNT(*) as eksik FROM personel p
            WHERE p.aktif=TRUE AND p.calisma_turu='surekli'
            AND NOT EXISTS (
                SELECT 1 FROM odeme_plani op
                WHERE op.kaynak_tablo='personel'
                AND op.kaynak_id = p.id::text
                AND op.durum != 'iptal'
                AND DATE_TRUNC('month', op.tarih) = DATE_TRUNC('month', %s::date)
            )
        """, (bugun,))
        eksik_personel = int(cur.fetchone()['eksik'])

    eksik_plan = eksik_sabit + eksik_borc + eksik_kart + eksik_personel
    uretim_denedi = False
    uretilen_adet = 0
    if eksik_plan > 0 and lock_ok:
        uretim_denedi = True
        try:
            sonuc = aylik_odeme_plani_uret(bugun.year, bugun.month)
            uretilen_adet = int(sonuc.get("toplam") or 0)
        except Exception as e:
            logger.warning(f"Lazy odeme_plani_kontrol üretim hatası: {e}")

    return {
        "eksik_toplam": eksik_plan,
        "eksik": {
            "sabit": eksik_sabit,
            "borc": eksik_borc,
            "kart": eksik_kart,
            "personel": eksik_personel,
        },
        "kilit_alindi": lock_ok,
        "uretim_denedi": uretim_denedi,
        "uretilen_adet": uretilen_adet,
        "borc_mutabakat": borc_mutabakat,
        "kasa_mutabakat": kasa_mutabakat,
    }


def bu_ay_plan_var(referans_tarih: Optional[date] = None) -> bool:
    bugun = referans_tarih or bugun_tr()
    with db() as (conn, cur):
        cur.execute(
            """
            SELECT 1
            FROM odeme_plani
            WHERE referans_ay = DATE_TRUNC('month', %s::date)
              AND durum != 'iptal'
            LIMIT 1
            """,
            (bugun,),
        )
        return bool(cur.fetchone())


def aylik_plan_lazy_init(referans_tarih: Optional[date] = None) -> dict:
    """
    Scheduler çalışmasa bile panel ilk açılışında bu ay planlarını üretir.
    İdempotent tasarım: aynı ayda tekrar çağrılar güvenlidir.
    """
    bugun = referans_tarih or bugun_tr()
    if bu_ay_plan_var(bugun):
        return {"uretildi": False, "neden": "plan_mevcut"}

    lock_ok = False
    with db() as (conn, cur):
        cur.execute(
            "SELECT pg_try_advisory_xact_lock(hashtext(%s)) AS ok",
            (f"aylik-plan-lazy-init:{bugun.year}-{bugun.month}",),
        )
        lock_ok = bool((cur.fetchone() or {}).get("ok"))

    if not lock_ok:
        return {"uretildi": False, "neden": "kilit_alinamadi"}

    if bu_ay_plan_var(bugun):
        return {"uretildi": False, "neden": "plan_mevcut"}

    try:
        sonuc = aylik_odeme_plani_uret(bugun.year, bugun.month)
        return {
            "uretildi": True,
            "neden": "uretim",
            "adet": int((sonuc or {}).get("toplam") or 0),
        }
    except Exception as e:
        logger.warning(f"Lazy aylik plan init hatası: {e}")
        return {"uretildi": False, "neden": "hata", "hata": str(e)}


@app.get("/api/panel")
def panel():
    try:
        lazy_plan = aylik_plan_lazy_init()
        plan_kontrol = odeme_plani_kontrol()
        plan_kontrol["lazy_init"] = lazy_plan

        ozet = finans_ozet_motoru()
        ozet['plan_kontrol'] = plan_kontrol
        # Devir: hesaplanır, ledger'a yazılmaz
        devir_bilgi = devir_hesapla()
        ozet['bu_ay_devir'] = devir_bilgi['devir_tutar']
        ozet['gecen_ay_adi'] = devir_bilgi['gecen_ay']

        # ── PERF PANEL REFAKTÖRÜ (2026-07-06) ──────────────────────────────────
        # Eskiden bu gövdede 3 AYRI DB bağlantısı + ~20 tekil SUM sorgusu vardı; her sorgu
        # EXTRACT(YEAR/MONTH)=CURRENT_DATE fonksiyonel filtresi kullanıyordu (index'siz seq scan).
        # Aynı tablo + aynı ay-penceresi sorguları FILTER ile TEK geçişte toplandı (~20 → 7 sorgu,
        # 3 → 1 bağlantı); ay filtresi index-dostu yarı-açık tarih aralığına çevrildi:
        # [ay başı, sonraki ay başı) — EXTRACT ikilisiyle SEMANTİK BİREBİR. Değerler golden'lı.
        with db() as (conn, cur):
            # 1) KASA HAREKETLERİ — bu ayın TÜM metrikleri tek geçişte (eski 9 ayrı sorgu)
            cur.execute("""
                SELECT
                    COALESCE(SUM(tutar)      FILTER (WHERE islem_turu='DIS_KAYNAK'), 0) AS dis_kaynak,
                    COALESCE(SUM(tutar)      FILTER (WHERE islem_turu='CIRO'), 0) AS sadece_ciro,
                    COALESCE(SUM(ABS(tutar)) FILTER (WHERE islem_turu='ANLIK_GIDER'), 0) AS anlik_gider,
                    COALESCE(SUM(ABS(tutar)) FILTER (WHERE islem_turu='SABIT_GIDER'    AND kasa_etkisi), 0) AS sabit_nakit,
                    COALESCE(SUM(ABS(tutar)) FILTER (WHERE islem_turu='FATURA_ODEMESI' AND kasa_etkisi), 0) AS fatura_nakit,
                    COALESCE(SUM(ABS(tutar)) FILTER (WHERE islem_turu='VADELI_ODEME'   AND kasa_etkisi), 0) AS vadeli_nakit,
                    COALESCE(SUM(ABS(tutar)) FILTER (WHERE islem_turu='BORC_TAKSIT'    AND kasa_etkisi), 0) AS borc_odenen,
                    COALESCE(SUM(ABS(tutar)) FILTER (WHERE kasa_etkisi AND tutar < 0
                        AND islem_turu NOT IN ('CIRO_DUZELTME','CIRO_IPTAL','ACILIS_DEVRI')), 0) AS toplam_cikis,
                    COALESCE(SUM(tutar)      FILTER (WHERE kasa_etkisi AND tutar > 0
                        AND islem_turu NOT IN ('CIRO_DUZELTME','CIRO_IPTAL','ACILIS_DEVRI')), 0) AS toplam_giris
                FROM kasa_hareketleri
                WHERE durum='aktif'
                  AND tarih >= date_trunc('month', CURRENT_DATE)
                  AND tarih <  date_trunc('month', CURRENT_DATE) + INTERVAL '1 month'
            """)
            _kh = dict(cur.fetchone())
            ozet['bu_ay_dis_kaynak'] = float(_kh['dis_kaynak'])
            ozet['bu_ay_sadece_ciro'] = float(_kh['sadece_ciro'])
            ozet['bu_ay_anlik_gider'] = float(_kh['anlik_gider'])

            # 2) Bu ay bankaya yatırılan (takip tablosu)
            cur.execute(
                """
                SELECT COALESCE(SUM(tutar), 0) AS toplam,
                       COUNT(*)::int AS adet
                FROM banka_yatirimlari
                WHERE tarih >= date_trunc('month', CURRENT_DATE)
                  AND tarih <  date_trunc('month', CURRENT_DATE) + INTERVAL '1 month'
                """
            )
            _by = cur.fetchone()
            ozet["bu_ay_banka_yatirim"] = float(_by["toplam"] or 0)
            ozet["bu_ay_banka_yatirim_adet"] = int(_by["adet"] or 0)

            # 3) CİRO — nakit/pos/online + finansman kesintileri tek geçişte (eski 2 sorgu).
            #    LEFT JOIN + COALESCE(oran,0): şube eşleşmeyen ciro satır kaybetmez (eski JOIN'de
            #    kesintiye katkısı zaten 0'dı — sonuç birebir aynı).
            cur.execute("""
                SELECT
                    COALESCE(SUM(c.nakit), 0)  AS nakit,
                    COALESCE(SUM(c.pos), 0)    AS pos,
                    COALESCE(SUM(c.online), 0) AS online,
                    COALESCE(SUM(c.pos    * COALESCE(s.pos_oran, 0)    / 100.0), 0) AS pos_kesinti,
                    COALESCE(SUM(c.online * COALESCE(s.online_oran, 0) / 100.0), 0) AS online_kesinti
                FROM ciro c
                LEFT JOIN subeler s ON s.id = c.sube_id
                WHERE c.durum='aktif'
                  AND c.tarih >= date_trunc('month', CURRENT_DATE)
                  AND c.tarih <  date_trunc('month', CURRENT_DATE) + INTERVAL '1 month'
            """)
            breakdown = cur.fetchone()
            ozet['bu_ay_nakit'] = float(breakdown['nakit'])
            ozet['bu_ay_pos'] = float(breakdown['pos'])
            ozet['bu_ay_online'] = float(breakdown['online'])
            ozet['bu_ay_pos_kesinti']    = float(breakdown['pos_kesinti'])
            ozet['bu_ay_online_kesinti'] = float(breakdown['online_kesinti'])

            # 📏 GÜNLÜK ORTALAMA CİRO — "bu tutar bu işletme için büyük mü?"
            # ⚠️ (2026-08-26, Codex denetimi) Bu taban İSTEMCİDE kuruluyordu:
            # `bu_ay_ciro / ayın kaçıncı günü`. Üç ayrı şekilde bozuktu:
            #   1) TAKVİM günü kullanıyordu — kapalı/tatil günler böleni şişirir,
            #      ortalamayı olduğundan düşük gösterir.
            #   2) Ciro GİRİLMEMİŞ günler de bölene giriyordu; veri eksikliği
            #      eşiği küçültüp önemsiz kalemleri öne çıkarıyordu.
            #   3) Ayın 1-3'ünde tek büyük gün ortalamayı aşırı şişirip anlamlı
            #      kalemleri eliyordu.
            # DOĞRU TABAN: son 30 günde CİRO GİRİLMİŞ günlerin ortalaması.
            # Girilmemiş gün böleni büyütmez; ay başı/sonu dalgası 30 güne yayılır.
            # Hiç ciro yoksa None döner — 0 demek "her şey büyüktür" demek olurdu.
            cur.execute("""
                SELECT AVG(gunluk)::float AS ort, COUNT(*) AS gun_adet
                FROM (
                    SELECT tarih,
                           SUM(COALESCE(nakit,0)+COALESCE(pos,0)+COALESCE(online,0)) AS gunluk
                      FROM ciro
                     WHERE COALESCE(durum,'aktif')='aktif'
                       AND tarih >= CURRENT_DATE - INTERVAL '30 days'
                     GROUP BY tarih
                    HAVING SUM(COALESCE(nakit,0)+COALESCE(pos,0)+COALESCE(online,0)) > 0
                ) g
            """)
            _oc = dict(cur.fetchone() or {})
            ozet['gunluk_ort_ciro'] = (round(float(_oc['ort']), 2)
                                       if _oc.get('ort') is not None else None)
            ozet['gunluk_ort_ciro_gun_adet'] = int(_oc.get('gun_adet') or 0)

            # 📅 BU AYIN KAYIT GÜNÜ + GÜNLÜK ORTALAMASI
            # ⚠️ (2026-08-26, Codex) Panel "Günlük ortalama"yı
            # `bu_ay_ciro / d.gunSayisi` ile hesaplıyordu; BÖLEN istemcide
            # `/ciro?limit=600` listesinden sayılıyordu. 600 tavanına
            # ulaşıldığında bölen EKSİK kalıyor ve ortalama OLDUĞUNDAN BÜYÜK
            # çıkıyordu — üstelik uyarı yalnız Ay görünümünde vardı.
            # Kırpılmış listeden bölen saymak, sessiz bir çarpıtmadır.
            # Hem bölen hem sonuç artık sunucudan gelir; istemci BÖLMEZ.
            cur.execute("""
                SELECT COUNT(DISTINCT tarih) AS gun_adet
                  FROM ciro
                 WHERE COALESCE(durum,'aktif')='aktif'
                   AND tarih >= date_trunc('month', CURRENT_DATE)
                   AND tarih <  date_trunc('month', CURRENT_DATE) + INTERVAL '1 month'
            """)
            _ag = int((dict(cur.fetchone() or {})).get('gun_adet') or 0)
            ozet['bu_ay_ciro_gun_adet'] = _ag
            ozet['bu_ay_gunluk_ort'] = (round(ozet.get('bu_ay_ciro', 0.0) / _ag, 2)
                                        if _ag else None)

            # 4) KART HAREKETLERİ — faiz + kart-kırılımları tek geçişte (eski 4 sorgu).
            #    NOT: kart_faizi orijinalinde durum şartı YOKTU — FILTER'larda birebir korunur.
            cur.execute("""
                SELECT
                    COALESCE(SUM(tutar) FILTER (WHERE islem_turu='FAIZ'), 0) AS kart_faizi,
                    COALESCE(SUM(tutar) FILTER (WHERE islem_turu='HARCAMA' AND durum='aktif'
                        AND kaynak_tablo='sabit_giderler'), 0) AS sabit_kart,
                    COALESCE(SUM(tutar) FILTER (WHERE islem_turu='HARCAMA' AND durum='aktif'
                        AND kaynak_tablo='fatura_giderleri'), 0) AS fatura_kart,
                    COALESCE(SUM(tutar) FILTER (WHERE islem_turu='HARCAMA' AND durum='aktif'
                        AND kaynak_tablo='vadeli_alimlar'), 0) AS vadeli_kart
                FROM kart_hareketleri
                WHERE tarih >= date_trunc('month', CURRENT_DATE)
                  AND tarih <  date_trunc('month', CURRENT_DATE) + INTERVAL '1 month'
            """)
            _kt = dict(cur.fetchone())
            ozet['bu_ay_kart_faizi'] = float(_kt['kart_faizi'])
            ozet['bu_ay_finansman_maliyeti'] = ozet['bu_ay_pos_kesinti'] + ozet['bu_ay_online_kesinti'] + ozet['bu_ay_kart_faizi']

            # 5) ÖDEME PLANI — son üretim + bekleyen borç taksitleri tek geçişte (eski 2 sorgu + 1 bağlantı)
            cur.execute("""
                SELECT
                    MAX(olusturma) FILTER (WHERE tarih >= date_trunc('month', CURRENT_DATE)
                        AND tarih < date_trunc('month', CURRENT_DATE) + INTERVAL '1 month') AS son_uretim,
                    COALESCE(SUM(odenecek_tutar) FILTER (WHERE kaynak_tablo='borc_envanteri'
                        AND durum IN ('bekliyor','onay_bekliyor')), 0) AS borc_bekleyen,
                    COUNT(*) FILTER (WHERE kaynak_tablo='borc_envanteri'
                        AND durum IN ('bekliyor','onay_bekliyor')) AS borc_bekleyen_adet
                FROM odeme_plani
            """)
            _op = dict(cur.fetchone())
            ozet['plan_son_uretim'] = str(_op['son_uretim'])[:16] if _op['son_uretim'] else None

            # ── NAKİT / KART KIRILIMLARI (aynı bağlantı — eski 3. with db() bloğu kaldırıldı) ──
            # 6) ANLIK GİDER — kasa=nakit, kart=kart kırılımı (tek tablo, tek sorgu — aynı kaldı)
            cur.execute("""
                SELECT
                    COALESCE(SUM(CASE WHEN ag.odeme_yontemi='nakit' THEN ag.tutar ELSE 0 END), 0) as nakit,
                    COALESCE(SUM(CASE WHEN ag.odeme_yontemi='kart'  THEN ag.tutar ELSE 0 END), 0) as kart
                FROM anlik_giderler ag
                WHERE ag.durum='aktif'
                  AND ag.tarih >= date_trunc('month', CURRENT_DATE)
                  AND ag.tarih <  date_trunc('month', CURRENT_DATE) + INTERVAL '1 month'
            """)
            ag = cur.fetchone()
            ozet['anlik_nakit'] = float(ag['nakit'])
            ozet['anlik_kart']  = float(ag['kart'])

            # Kasa/kart kırılımları — 1) ve 4) numaralı birleşik sorgulardan (eski 7 ayrı sorgu)
            ozet['sabit_nakit']  = float(_kh['sabit_nakit'])
            ozet['sabit_kart']   = float(_kt['sabit_kart'])
            ozet['fatura_nakit'] = float(_kh['fatura_nakit'])
            ozet['fatura_kart']  = float(_kt['fatura_kart'])
            ozet['vadeli_nakit'] = float(_kh['vadeli_nakit'])
            ozet['vadeli_kart']  = float(_kt['vadeli_kart'])

            # 7) PERSONEL — tahmini + gerçekleşen + kayıt bekleyen tek round-trip (eski 3 sorgu)
            # PROD-PANEL-005 FIX: bordro variance POPÜLASYONU HİZALANDI. Eskiden gercek TÜM
            # personel_aylik'i (surekli+part-time) topluyordu ama tahmini SADECE surekli'ydi →
            # part-time gercek'i şişirip "fark"ı yanıltıcı yapıyordu (elma-armut). Üçü de artık
            # SÜREKLİ popülasyon: part-time saatlik/değişken (sabit "tahmini" yok), variance kartına
            # girmez. NOT: ay-içi giriş/çıkış pro-rata + toplam bordro ayrı iş (PROD-PANEL-005b).
            cur.execute("""
                SELECT
                    (SELECT COALESCE(SUM(p.maas + p.yemek_ucreti + p.yol_ucreti), 0)
                     FROM personel p WHERE p.aktif=TRUE AND COALESCE(p.calisma_turu,'surekli')='surekli') AS tahmini,
                    (SELECT COALESCE(SUM(pa.hesaplanan_net), 0)
                     FROM personel_aylik pa
                     JOIN personel p2 ON p2.id = pa.personel_id
                        AND COALESCE(p2.calisma_turu,'surekli')='surekli'
                     WHERE pa.yil = EXTRACT(YEAR FROM CURRENT_DATE)
                       AND pa.ay  = EXTRACT(MONTH FROM CURRENT_DATE)) AS gercek,
                    (SELECT COUNT(*)
                     FROM personel p
                     WHERE p.aktif=TRUE AND COALESCE(p.calisma_turu,'surekli')='surekli'
                       AND NOT EXISTS (
                           SELECT 1 FROM personel_aylik pa
                           WHERE pa.personel_id = p.id
                             AND pa.yil = EXTRACT(YEAR FROM CURRENT_DATE)
                             AND pa.ay  = EXTRACT(MONTH FROM CURRENT_DATE)
                       )) AS bekleyen
            """)
            _pp = dict(cur.fetchone())
            ozet['personel_tahmini'] = float(_pp['tahmini'])
            ozet['personel_gercek'] = float(_pp['gercek'])
            ozet['personel_kayit_bekleyen'] = int(_pp['bekleyen'])

            # BORÇ TAKSİTLERİ — 1) ve 5) numaralı birleşik sorgulardan
            ozet['borc_taksit_odenen'] = float(_kh['borc_odenen'])
            ozet['borc_taksit_bekleyen'] = float(_op['borc_bekleyen'])
            ozet['borc_taksit_bekleyen_adet'] = int(_op['borc_bekleyen_adet'])

            # GENEL TOPLAM (Bu Ay Gider Ödeme Yöntemi özeti — Panel.jsx:1574)
            # PROD-PANEL-003 FIX: özet yalnız anlik+sabit+vadeli'yi topluyordu; panel FATURA ve
            # BORÇ TAKSİT kartlarını da gösterdiğinden "genel" nakit/kart yüzdesi görünen kartlarla
            # TUTMUYORDU. Ödeme-yöntemi ayrıştırılabilir GERÇEK ödemeler dahil edildi (Codex onaylı):
            #  - fatura_nakit/fatura_kart (FATURA_ODEMESI / fatura_giderleri kart)
            #  - borc_taksit_odenen (BORC_TAKSIT = nakit-only kasa çıkışı; kart yolu yok)
            # PERSONEL HARİÇ: personel_gercek = personel_aylik tahakkuku (arrears), nakit/kart kırılımı
            # yok → dahil etmek sahte method-split üretir (ayrı ürün kararı gerektirir).
            ozet['genel_nakit_toplam'] = (ozet['anlik_nakit'] + ozet['sabit_nakit'] + ozet['vadeli_nakit']
                                          + ozet['fatura_nakit'] + ozet['borc_taksit_odenen'])
            ozet['genel_kart_toplam']  = (ozet['anlik_kart'] + ozet['sabit_kart'] + ozet['vadeli_kart']
                                          + ozet['fatura_kart'])

            # BU AY TOPLAM KASA GİRİŞ/ÇIKIŞ — 1) numaralı birleşik sorgudan
            # (CIRO_DUZELTME/CIRO_IPTAL/ACILIS_DEVRI teknik ters kayıtları hariç — filtre aynı)
            ozet['bu_ay_nakit_cikis'] = float(_kh['toplam_cikis'])
            ozet['bu_ay_nakit_giris'] = float(_kh['toplam_giris'])
            ozet['bu_ay_net'] = ozet['bu_ay_nakit_giris'] - ozet['bu_ay_nakit_cikis']

            # 💼 KANAL TOPLAMI — "bu ayın GERÇEK tahsilatı" (devir HARİÇ)
            # ⚠️ (2026-08-26, Codex denetimi) Bu toplam İSTEMCİDE hesaplanıyordu
            # (GenelModulu.jsx tahsilat kanalları bloğu). "Gösterim kendi
            # aritmetiğini kurmaz" doktrininin ihlaliydi: ekran bir rakam
            # TÜRETİYORDU ve o rakamın ikinci bir doğruluk kaynağı oluyordu —
            # kanal tanımı burada değişse ekran eski toplamı göstermeye devam
            # ederdi. Tek kaynak: panel.
            # ⚠️ DEVİR BİLEREK DIŞARIDA: geçmiş aydan devreden bakiye bu ayın
            # tahsilatı değildir; toplama katılırsa "bu ay 3,5 milyon tahsil
            # ettik" yanılsaması doğar.
            ozet['bu_ay_kanal_toplam'] = (ozet.get('bu_ay_nakit', 0.0)
                                          + ozet.get('bu_ay_pos', 0.0)
                                          + ozet.get('bu_ay_online', 0.0)
                                          + ozet.get('bu_ay_dis_kaynak', 0.0))

        return ozet
    except Exception as e:
        import traceback
        logger.error(f"Panel hatası: {e}\n{traceback.format_exc()}")
        raise HTTPException(500, str(e))

@app.get("/api/panel/detay")
def panel_detay():
    """Eski panel endpoint'i — geriye dönük uyumluluk için."""
    try:
        karar = karar_motoru()
        sim = nakit_akis_simulasyon(15)
        with db() as (conn, cur):
            cur.execute("""
                SELECT TO_CHAR(tarih,'YYYY-MM') as ay, SUM(toplam) as ciro
                FROM ciro WHERE tarih >= CURRENT_DATE - INTERVAL '6 months'
                GROUP BY TO_CHAR(tarih,'YYYY-MM') ORDER BY ay DESC LIMIT 6
            """)
            aylik_ciro = [dict(r) for r in cur.fetchall()]
            cur.execute("SELECT COUNT(*) as sayi, COALESCE(SUM(tutar),0) as toplam FROM onay_kuyrugu WHERE durum='bekliyor'")
            bekleyen = dict(cur.fetchone())
            cur.execute("""
                SELECT COALESCE(SUM(CASE WHEN tarih<=CURRENT_DATE+7 THEN odenecek_tutar ELSE 0 END),0) as t7,
                    COALESCE(SUM(CASE WHEN tarih<=CURRENT_DATE+15 THEN odenecek_tutar ELSE 0 END),0) as t15,
                    COALESCE(SUM(CASE WHEN tarih<=CURRENT_DATE+30 THEN odenecek_tutar ELSE 0 END),0) as t30
                FROM odeme_plani WHERE durum='bekliyor' AND tarih BETWEEN CURRENT_DATE AND CURRENT_DATE+30
            """)
            odeme_ozet = dict(cur.fetchone())
            cur.execute("""
SELECT
    COALESCE(SUM(CASE WHEN islem_turu IN ('CIRO','DIS_KAYNAK','KASA_GIRIS','KASA_DUZELTME') AND tutar > 0 THEN tutar ELSE 0 END), 0) as gelir,
    COALESCE(SUM(CASE WHEN islem_turu IN ('ANLIK_GIDER','KART_ODEME','VADELI_ODEME','PERSONEL_MAAS','PERSONEL_AVANS','SABIT_GIDER','BORC_TAKSIT','FATURA_ODEMESI') THEN ABS(tutar) ELSE 0 END), 0) as gider
FROM kasa_hareketleri
WHERE durum='aktif'
""")
            row = cur.fetchone() or {"gelir": 0, "gider": 0}
            # D13 notu (2026-07-10): buradaki 'gider' = KASADAN ÇIKAN NAKİT (nakit
            # akışı; KART_ODEME=banka transferi dahil). P&L gideri DEĞİLDİR —
            # operasyonel P&L kart ödemesini/finansmanı bilinçli dışlar.
            toplam_gelir = float(row.get('gelir', 0) or 0)
            toplam_gider = float(row.get('gider', 0) or 0)

            # Aksiyonlar
            aksiyonlar = []
            kasa_val = karar.get("kasa", 0)
            if kasa_val <= 0:
                aksiyonlar.append({"tip":"kritik","mesaj":"Kasa boş. Önce ciro gir.","aksiyon":"ciro"})
            if odeme_ozet.get("t7", 0) > 0:
                aksiyonlar.append({"tip":"uyari","mesaj":"7 gün içinde ödeme var","aksiyon":"odeme"})

        # Kart analiz — with db() dışında ayrı bağlantıyla
        kart_analiz = kart_analiz_hesapla()

        return {**karar, "simulasyon": sim, "aylik_ciro": aylik_ciro,
                "bekleyen_onay": bekleyen, "odeme_ozet": odeme_ozet,
                "kart_analiz": kart_analiz, "toplam_gelir": toplam_gelir,
                "toplam_gider": toplam_gider, "aksiyonlar": aksiyonlar}
    except Exception as e:
        raise HTTPException(500, str(e))



@app.get("/api/strateji")
def strateji():
    try: return odeme_strateji_motoru()
    except Exception as e: raise HTTPException(500, str(e))

@app.get("/api/simulasyon")
def simulasyon(gun: int = 15):
    try: return nakit_akis_simulasyon(gun)
    except Exception as e: raise HTTPException(500, str(e))


@app.get("/api/nakit-akis-projeksiyon")
def nakit_akis_projeksiyon(gun: int = 30):
    """
    Gelişmiş nakit akış projeksiyonu — /api/simulasyon'un üst kümesi.

    Farklar:
    - Gecikmiş ödemeler (tarih < bugün, hâlâ bekliyor) gün 0'a dahil edilir.
    - Her güne risk_seviye: 'NORMAL' / 'DIKKAT' / 'KRITIK' eklendi.
    - Özet blok: başlangıç kasası, en düşük nokta, risk gün sayısı,
      toplam tahmini gelir/gider, backtest tahmin doğruluğu.
    - /api/simulasyon ile aynı finans_core çekirdeğini kullanır; tüm
      mevcut çağrıcılar değişmez.

    gun: 7–90 gün arası (varsayılan 30).
    """
    gun = max(7, min(90, int(gun)))
    try:
        with db() as (conn, cur):
            baslangic_kasa = kasa_bakiyesi(cur)
            gunler = nakit_akis_sim(cur, gun_sayisi=gun)
            # Backtest doğruluğu aynı cursor ile çekilir — ekstra bağlantı yok.
            dogruluk = nakit_akis_tahmin_dogruluk(cur, gun_sayisi=30)

        # ── Özet istatistikler ──────────────────────────────────
        kasa_degerler = [g['kasa_tahmini'] for g in gunler]
        risk_gunler   = [g for g in gunler if g['risk']]
        dikkat_gunler = [g for g in gunler if g.get('risk_seviye') == 'DIKKAT']
        en_dusuk_kasa = min(kasa_degerler) if kasa_degerler else baslangic_kasa
        en_dusuk_gun  = next(
            (g for g in gunler if g['kasa_tahmini'] == en_dusuk_kasa), None
        )
        gecikmus = gunler[0].get('gecikmus_odeme', 0.0) if gunler else 0.0

        ozet = {
            "gun_sayisi":              gun,
            "baslangic_kasa":          round(baslangic_kasa, 2),
            "tahmini_gun_sonu_kasa":   gunler[-1]['kasa_tahmini'] if gunler else round(baslangic_kasa, 2),
            "en_dusuk_kasa":           round(en_dusuk_kasa, 2),
            "en_dusuk_tarih":          en_dusuk_gun['tarih'] if en_dusuk_gun else None,
            "risk_gun_sayisi":         len(risk_gunler),
            "dikkat_gun_sayisi":       len(dikkat_gunler),
            "ilk_risk_gunu":           risk_gunler[0]['tarih'] if risk_gunler else None,
            "toplam_tahmini_gelir":    round(sum(g['beklenen_gelir'] for g in gunler), 2),
            "toplam_planlanan_gider":  round(sum(g['beklenen_gider'] for g in gunler), 2),
            # Gecikmiş ödeme varsa gün 0'a yüklendiği miktar — kullanıcı bilgilensin.
            "gecikmus_odeme_yuklendi": round(gecikmus, 2),
            # Backtest doğruluğu — yeterli ciro geçmişi yoksa None döner.
            "tahmin_dogruluk_pct":     dogruluk.get('dogruluk_pct'),
            "tahmin_dogruluk_mesaj":   dogruluk.get('mesaj'),
            "tahmin_dogruluk_durum":   dogruluk.get('durum'),
        }

        return {
            "ozet":   ozet,
            "gunler": gunler,
        }
    except Exception as e:
        raise HTTPException(500, str(e))


# ── KASA ───────────────────────────────────────────────────────
@app.get("/api/kasa")
def kasa_durumu():
    with db() as (conn, cur):
        kasa = guncel_kasa()
        cur.execute("""SELECT * FROM kasa_hareketleri WHERE durum='aktif'
            ORDER BY tarih DESC, olusturma DESC LIMIT 100""")
        return {"guncel_bakiye": kasa, "hareketler": [dict(r) for r in cur.fetchall()]}


class MaliyetMerkeziBody(BaseModel):
    # ⚠️ İKİSİ BAĞIMSIZ ALAN — biri olmadan öteki set edilebilir.
    # Sahip (2026-08-18): "katılım evim ve araç kredisi ZAFER'DEN ÖDENİYOR ...
    # aslında anlattığım gibi hangisinin kasasında para varsa ondan ödüyordur!"
    # Bu cümle KİMİN BORCU olduğunu değil, GENELDE KİMİN ÖDEDİĞİni söyler.
    # Mortgage Zafer'den ödeniyor olması onu Zafer'in borcu YAPMAZ.
    # Bu yüzden `maliyet_merkezi_tipi` artık zorunlu değil: sahip yalnız ödeyen
    # bilgisini verdiyse maliyet merkezi BİLİNMİYOR kalır (uydurulmaz).
    maliyet_merkezi_tipi: Optional[str] = None      # 'sube' | 'ortak' | 'sahis'
    maliyet_merkezi_id: Optional[str] = None        # tip='sube' ise sube_id
    varsayilan_odeyen_sube_id: Optional[str] = None # form ön-dolgusu (yalnız ÖNERİ)


@app.post("/api/borclar/{bid}/maliyet-merkezi")
def borc_maliyet_merkezi(bid: str, b: MaliyetMerkeziBody):
    """Kredinin MALİYET MERKEZİni işaretle (kimin borcu) + varsayılan ödeyeni.

    🔴 Neden tip zorunlu (Codex): 10 kredinin 5'i şubeye ait DEĞİL —
    katılım evim (mortgage), KOÇ Finans Araç, QNB (Fethi-Karaman), VAKIF ANNEM,
    YAPI KREDİ ANNEM. Bunlara zorla şube atamak "yalan söylemeye başlamak"tır.
    Tipler: 'sube' (belirli şube) · 'ortak' (işletme geneli) · 'sahis' (şahsi).

    ⚠️ varsayilan_odeyen_sube_id yalnız ÖNERİdir — formu ön-doldurur, kullanıcı
    GÖRÜR ve onaylar. Görmeden otomatik yazılmaz (Codex S4: "istatistiksel
    tahmin yalnız öneri olabilir, gerçek veri olamaz").
    """
    tip = (b.maliyet_merkezi_tipi or "").strip().lower() or None
    if tip is not None and tip not in ("sube", "ortak", "sahis"):
        raise HTTPException(400, "maliyet_merkezi_tipi: sube | ortak | sahis olmalı")
    mid = (b.maliyet_merkezi_id or "").strip() or None
    if tip == "sube" and not mid:
        raise HTTPException(400, "tip 'sube' ise maliyet_merkezi_id (şube) zorunlu")
    if tip and tip != "sube":
        mid = None   # ortak/şahsi merkezde şube kimliği tutulmaz — uydurma yok
    if tip is None and not (b.varsayilan_odeyen_sube_id or "").strip():
        raise HTTPException(400, "En az biri gerekli: maliyet_merkezi_tipi veya "
                                 "varsayilan_odeyen_sube_id")
    with db() as (conn, cur):
        cur.execute("SELECT kurum FROM borc_envanteri WHERE id=%s", (bid,))
        r = cur.fetchone()
        if not r:
            raise HTTPException(404, "Kredi bulunamadı")
        if mid:
            cur.execute("""SELECT 1 FROM subeler WHERE id::text=%s
                             AND UPPER(COALESCE(ad,''))<>'MERKEZ'""", (mid,))
            if not cur.fetchone():
                raise HTTPException(404, f"Şube bulunamadı: {mid}")
        ody = (b.varsayilan_odeyen_sube_id or "").strip() or None
        if ody:
            cur.execute("""SELECT 1 FROM subeler WHERE id::text=%s
                             AND UPPER(COALESCE(ad,''))<>'MERKEZ'""", (ody,))
            if not cur.fetchone():
                raise HTTPException(404, f"Varsayılan ödeyen şube bulunamadı: {ody}")
        # COALESCE ile KISMİ güncelleme: verilmeyen alan mevcut değerini korur —
        # "yalnız ödeyeni söyledim" diyen sahip, maliyet merkezini SIFIRLAMIŞ olmaz.
        cur.execute("""UPDATE borc_envanteri
                          SET maliyet_merkezi_tipi = COALESCE(%s, maliyet_merkezi_tipi),
                              maliyet_merkezi_id   = CASE WHEN %s IS NULL
                                                          THEN maliyet_merkezi_id ELSE %s END,
                              varsayilan_odeyen_sube_id = COALESCE(%s, varsayilan_odeyen_sube_id)
                        WHERE id=%s""", (tip, tip, mid, ody, bid))
        audit(cur, "borc_envanteri", bid, "UPDATE")
    return {"success": True, "kurum": dict(r).get("kurum"),
            "maliyet_merkezi_tipi": tip, "maliyet_merkezi_id": mid,
            "varsayilan_odeyen_sube_id": ody}


class DevirTazeleBody(BaseModel):
    kart_id: Optional[str] = None      # boş = tüm bayat kartlar
    uygula: bool = False               # ⚠️ VARSAYILAN KURU ÇALIŞTIRMA


@app.post("/api/kartlar/devir-tazele")
def kart_devir_tazele(body: DevirTazeleBody):
    """🔄 BAYAT DENKLEŞTİRME YAMASINI TAZELE — sistemin KENDİ mekanizması.

    Codex'in kuralı gereği önce İKİ KARTTA HAM VERİDEN doğrulandı (2026-08-18):
      Garanti Onur (iyi): ham −113.261,15 + yama 597.238,22 = 483.977,07 = BANKA ✓
      WORLD ANNEM (kötü): ham  +41.567,98 + yama 139.842,02 = 181.410  ≠ 480.481 ✗
    Yani mekanizma DOĞRU, yama 5 kartta ESKİMİŞ. Yama yazıldıktan sonra o karta
    ekstre yüklendiyse ham geçmiş büyür ama yama sabit kalır → sapma.

    Bu uç, sistemin zaten kullandığı formülü (`adj = donem_borcu − ham borç`)
    yeniden çalıştırır. Yeni bir kavram İCAT ETMEZ; bayat yamayı günceller.

    ⚠️ NE YAPMAZ: kayıp geçmişi geri getirmez. Yama bir ÖZETtir; tazelemek
    defteri bankaya eşitler ama o 299 K'nın hangi harcamalardan geldiğini
    söylemez. Gerçek geçmiş isteniyorsa eski ekstreler yüklenmeli.
    ⛔ GERİ-ALMA ≠ SİLME: eski yama satırı iptal edilir (iz kalır), yerine
    gerekçeli yenisi yazılır.
    """
    with db() as (conn, cur):
        rapor = kart_devir_denetimi()["kartlar"]
        hedef = [r for r in rapor
                 if r.get("durum") == "YAMA BAYAT"
                 and (not body.kart_id or r["kart_id"] == body.kart_id)]
        plan, uygulanan = [], 0
        for r in hedef:
            eski, yeni = r["stored_devir"], r["expected_devir"]
            plan.append({"kart": r["kart_adi"], "eski_devir": eski, "yeni_devir": yeni,
                         "sapma": r["sapma"], "devir_tarihi": r["devir_tarihi"],
                         "ekstre_borcu": r.get("ekstre_borcu")})
            if not body.uygula:
                continue
            cur.execute("""UPDATE kart_hareketleri
                              SET durum='iptal',
                                  aciklama = COALESCE(aciklama,'') ||
                                    ' [bayat yama — tazelendi 2026-08-18]'
                            WHERE kart_id=%s AND durum='aktif' AND islem_turu='DEVIR'""",
                        (r["kart_id"],))
            hid = str(uuid.uuid4())
            cur.execute("""INSERT INTO kart_hareketleri
                (id, kart_id, tarih, islem_turu, tutar, taksit_sayisi, aciklama)
                VALUES (%s,%s,%s::date,'DEVIR',%s,1,%s)""",
                (hid, r["kart_id"], r["devir_tarihi"], yeni,
                 f"Denkleştirme yaması (tazelendi) — eski {eski:,.2f} → {yeni:,.2f} "
                 f"· sapma {r['sapma']:,.2f} · ekstre {r.get('ekstre_borcu')}"))
            audit(cur, "kart_hareketleri", hid, "DEVIR_TAZELE")
            uygulanan += 1
        if body.uygula:
            kart_plan_guncelle_tx(cur)
    return {
        "kuru_calistirma": not body.uygula,
        "bayat_kart": len(hedef), "tazelenen": uygulanan, "plan": plan,
        "toplam_duzeltme": round(sum(abs(p["sapma"]) for p in plan), 2),
        "uyari": "Bu işlem defteri bankaya EŞİTLER ama kayıp geçmişi geri getirmez. "
                 "Yama bir özettir; hangi harcamalardan geldiğini söylemez.",
        "not": ("Hiçbir şey yazılmadı — uygulamak için uygula=true gönderin."
                if not body.uygula else
                "Eski yamalar İZLİ İPTAL edildi, yerlerine gerekçeli yenisi yazıldı."),
    }


def _tarih_d(v) -> date:
    """'YYYY-MM-DD' / date / None → date. Çözülemezse 1970-01-01 (tarih
    karşılaştırması eşleştirmeyi bozmasın; tutar zaten asıl anahtardır)."""
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return date(1970, 1, 1)


def _satir_vade(tarih: date, taksit_sayisi: int, kesim_gunu: int, i: int) -> date:
    """Bir hareketin i. taksidinin BANKA takvimindeki vade (ekstre kesimi) tarihi.

    Tek çekimde tarih neyse odur. Taksitliyse 1. taksit alımdan SONRAKİ ilk
    kesimde başlar, sonrakiler birer ay arayla (2026-08-24'te canlı öğrenildi —
    bkz. donem-mutabakati'ndaki taksit takvimi notu).
    """
    if (taksit_sayisi or 1) <= 1:
        return tarih
    y, m = tarih.year, tarih.month
    if tarih.day > kesim_gunu:
        m += 1
        if m > 12:
            m, y = 1, y + 1
    m += (i - 1)
    y += (m - 1) // 12
    m = (m - 1) % 12 + 1
    import calendar as _cal
    return date(y, m, min(kesim_gunu, _cal.monthrange(y, m)[1]))


@app.get("/api/kartlar/ekstre-satir-denetimi")
def kart_ekstre_satir_denetimi(kart_id: Optional[str] = None,
                               donem: Optional[str] = None,
                               limit: int = 8):
    """🔬 EKSTRE ↔ DEFTER SATIR DENETİMİ — SALT OKUR. Arşivdeki PDF'i defterle KARŞILAŞTIRIR.

    ── NEDEN ───────────────────────────────────────────────────────────────
    Dönem mutabakatı "bu dönemde X ₺ fark var" diyebiliyor ama HANGİ SATIR
    olduğunu söyleyemiyor. 2026-08-24'te WORLD ANNEM'in ≈3.306 ₺'lik aylık
    açığını bulmak için arşivdeki PDF elle indirilip satır satır defterle
    karşılaştırıldı. Çıkan şey tek satırdı:

        12 Şubat 2026  HEPSIPAY *HEPSIBURADA  3.306,96
        19.841,75 TL'lik işlemin 6 / 6 taksidi
        → defterde: 3.306,96 ₺ · taksit_sayisi = 1   (6 taksitli alım TEK ÇEKİM sanılmış)

    O elle yapılan iş BU UÇTUR. Artık her kart/dönem için kendi yapıyor.

    ── NASIL ───────────────────────────────────────────────────────────────
    1) Dönemin arşivlenmiş PDF'i okunur (kart_ekstre_donem.belge_pdf).
    2) Bankanın O DÖNEME fatura ettiği satırlar çıkarılır. Taksitli satırda
       bankanın yazdığı tutar TAKSİT TUTARIDIR; toplam ve taksit no alt
       satırdan gelir — ikisi de taşınır.
    3) Defterden aynı pencereye (önceki kesim, bu kesim] düşen kalemler
       üretilir. Taksitli hareket, o döneme düşen TAKSİT PAYIYLA temsil edilir
       (banka ne fatura ettiyse onunla kıyaslanabilsin diye).
    4) Tutar+tip üzerinden eşleştirilir; eşleşmeyenler İKİ AYRI listede döner.

    ⚠️ ÖNERİ-ONLY — hiçbir kayıt yazılmaz/silinmez. Eşleşmeme her zaman hata
    değildir: aynı tutarda iki meşru satır, valör kayması, ya da bankanın
    ayrıştırılamayan bir satırı olabilir. Bu yüzden `okuma_guveni` alanı var:
    PDF'ten okunan satırların toplamı bankanın beyan ettiği dönem harcaması ile
    tutmuyorsa OKUMA şüphelidir, defter değil — ve bu AÇIKÇA yazılır.
    """
    limit = max(1, min(24, int(limit or 8)))
    with db() as (conn, cur):
        _ekstre_belge_kolonlari(cur)
        kos, par = ["belge_pdf IS NOT NULL", "donem_borcu IS NOT NULL"], []
        if kart_id:
            kos.append("kart_id=%s"); par.append(kart_id)
        if donem:
            d = donem.strip()[:10]
            if len(d) == 7:
                d += "-01"
            kos.append("donem = DATE_TRUNC('month', %s::date)"); par.append(d)
        cur.execute(f"""
            SELECT e.kart_id::text AS kart_id, k.kart_adi,
                   COALESCE(k.kesim_gunu,1) AS kesim_gunu,
                   e.donem::text AS donem, e.kesim_tarihi::text AS kesim,
                   e.donem_harcama::float AS banka_harcama,
                   e.donem_odeme::float AS banka_odeme,
                   e.donem_borcu::float AS donem_borcu,
                   e.onceki_borc::float AS onceki_borc,
                   e.donem_faizi::float AS donem_faizi,
                   e.belge_pdf
              FROM kart_ekstre_donem e
              JOIN kartlar k ON k.id = e.kart_id
             WHERE {' AND '.join(kos)}
             ORDER BY e.donem DESC LIMIT %s
        """, (*par, limit))
        hedefler = [dict(r) for r in (cur.fetchall() or [])]
        if not hedefler:
            return {"donemler": [], "adet": 0,
                    "not": "Arşivde PDF'li dönem bulunamadı — denetim yapılamaz (HATA ≠ BOŞ)."}

        try:
            import kart_analiz
        except Exception as e:  # noqa: BLE001
            raise HTTPException(500, f"kart_analiz yüklenemedi: {str(e)[:120]}")

        cikti, top_eksik, top_fazla = [], 0.0, 0.0
        for h in hedefler:
            kid, kg = h["kart_id"], int(h["kesim_gunu"] or 1)
            bit = h.get("kesim")
            # Önceki dönemin kesimi = pencerenin başı
            cur.execute("""SELECT kesim_tarihi::text AS k FROM kart_ekstre_donem
                            WHERE kart_id=%s AND donem < DATE_TRUNC('month', %s::date)
                              AND kesim_tarihi IS NOT NULL
                            ORDER BY donem DESC LIMIT 1""", (kid, h["donem"]))
            bas = (cur.fetchone() or {}).get("k")
            satir = {"kart_id": kid, "kart_adi": h["kart_adi"], "donem": h["donem"],
                     "kesim_bas": bas, "kesim_bit": bit}
            if not bas or not bit:
                satir.update({"durum": "PENCERE KURULAMADI — önceki dönemin kesimi yok",
                              "pdf_eksik": [], "defter_fazla": []})
                cikti.append(satir); continue

            # ── 1) PDF satırları ────────────────────────────────────────────
            # 📐 ÖNCE GEOMETRİK OKUYUCU (2026-08-24) — satır denetimi başlangıçta
            # yalnız metin okuyucusunu (kart_analiz) kullanıyordu ve Ziraat'te
            # 126.668,48 ₺'lik dönemin sadece 6.722,50 ₺'sini görüyordu; dönem
            # "OKUMA DOĞRULANAMIYOR" diye ölçüm dışı kalıyordu. Ekstre içe
            # aktarma yolu zaten geometrik okuyucuyu kullanıyor ve o çok daha
            # sağlam (sütun koordinatından okur, yinelenen sayfayı atar).
            # İki ayrı okuyucuyla ölçmek, iki ayrı gerçek üretmek demekti.
            # Geometrik okuyamıyorsa (Axess/CID) metin yoluna DÜŞÜLÜR — orada
            # da EBCDIC çözücü devrede.
            _pdf_ham = bytes(h["belge_pdf"])
            txns, _okuyucu = [], "metin"
            try:
                import ekstre_geometri as _geo2
                _gr = _geo2.geometrik_oku(_pdf_ham)
                if _gr.get("basarili") and _gr.get("satirlar"):
                    _okuyucu = "geometrik"
                    for _r in _gr["satirlar"]:
                        txns.append({
                            "tarih": _r.get("tarih"), "aciklama": _r.get("aciklama") or "",
                            "tutar": _r.get("tutar"), "odeme_mi": _r.get("odeme_mi"),
                            "taksit": (f"{_r.get('taksit_no')}/{_r.get('taksit_sayisi')}"
                                       if _r.get("taksit_sayisi") else None),
                            "taksit_anapara": _r.get("taksit_toplam"),
                        })
            except Exception as _e2:  # noqa: BLE001 — geometrik yol düşerse metin yolu var
                logger.info("satır denetimi: geometrik okuma atlandı (%s)", str(_e2)[:90])
            if not txns:
                try:
                    txns = kart_analiz.parse_pdf(_pdf_ham) or []
                except Exception as e:  # noqa: BLE001
                    satir.update({"durum": f"PDF OKUNAMADI — {str(e)[:90]}",
                                  "pdf_eksik": [], "defter_fazla": []})
                    cikti.append(satir); continue
            satir["okuyucu"] = _okuyucu
            banka = []
            for t in txns:
                tu = abs(float(t.get("tutar") or 0))
                if tu <= 0:
                    continue
                ac = str(t.get("aciklama") or "")
                tip = "ODEME" if t.get("odeme_mi") else "HARCAMA"
                tks = str(t.get("taksit") or "")
                tsay = None
                if "/" in tks:
                    try:
                        tsay = int(tks.split("/")[1])
                    except (ValueError, IndexError):
                        tsay = None
                banka.append({"tarih": t.get("tarih"), "tutar": round(tu, 2), "tip": tip,
                              "aciklama": ac[:60], "taksit": tks or None,
                              "taksit_sayisi": tsay,
                              "taksit_anapara": t.get("taksit_anapara")})
            # OKUMA GÜVENİ — okunan satırlar bankanın beyanını tutuyor mu?
            # 🏦 BANKA GELENEĞİ FARKLI (2026-08-24, Axess açılınca görüldü):
            # Yapı Kredi `dönem harcaması`na FAİZİ DE katıyor; Axess KATMIYOR.
            # Tek yorumla ölçersek bankalardan biri hep "okuma şüpheli" çıkar —
            # Axess'te tam 1.292,38 ₺ (dönem faizinin kendisi) sapıyordu ve
            # okuma aslında KUSURSUZDU. O yüzden iki yorum da denenir, hangisi
            # tutuyorsa o kabul edilir; hangisinin geçerli olduğu da yazılır.
            import re as _re  # modül düzeyinde import yok — yerel al (NameError kalkanı)
            oku_h = round(sum(b["tutar"] for b in banka if b["tip"] != "ODEME"), 2)
            _faiz_re = _re.compile(r"faiz|bsmv|kkdf|[uü]cret", _re.I)
            oku_h_faizsiz = round(sum(b["tutar"] for b in banka
                                      if b["tip"] != "ODEME" and not _faiz_re.search(b["aciklama"] or "")), 2)
            # 🧮 BEYAN YOKSA BORÇ KİMLİĞİNDEN TÜRET (2026-08-24)
            # Ziraat ve Garanti ekstreleri "dönem içi harcamalar" satırını
            # BASMIYOR. Duyu bu yüzden 6 dönemi "OKUMA DOĞRULANAMIYOR" diye
            # ölçüm dışı bırakıyordu — oysa ölçülemeyen şey yoktu, sadece hazır
            # yazılmış bir toplam yoktu. Ekstrenin kendi kimliği zaten söylüyor:
            #     dönem borcu = önceki borç − ödeme + harcama
            # → harcama = dönem borcu − önceki borç + ödeme
            # İçe aktarma yolundaki çapa bunu zaten böyle kuruyor ("borc_kimligi").
            # Duyunun daha az bilgiyle çalışması için bir sebep yok; ölçüt aynı
            # olmalı, yoksa iki uç aynı ekstre için iki farklı hüküm verir.
            bh = h.get("banka_harcama")
            bh_kaynak = "banka beyanı"
            if bh is None and h.get("donem_borcu") is not None and h.get("onceki_borc") is not None:
                # ⚠️ NULL = BİLİNMİYOR, SIFIR DEĞİL (2026-08-24, ilk sürümde düşülen tuzak)
                # Ziraat'te dönem ÖDEMESİ de yazılı değil. İlk denemede `odeme or 0`
                # yazıldı ve türetilen harcama −141.774,87 ₺ çıktı (negatif harcama!);
                # duyu 266.544,30 ₺'lik sahte bir okuma hatası bildirdi. Ödeme
                # bilinmiyorsa sıfır varsaymak, ekstrenin kimliğini bozar.
                # Doğrusu: bankanın yazdığı ödeme yoksa OKUNAN ödemeyi kullan —
                # içe aktarma yolundaki çapa da tam bunu yapıyor. Bu, harcama
                # tarafını iki BASILI çıpayla (dönem borcu, önceki borç) sınar;
                # ödeme tarafının doğruluğu varsayım olarak açıkça etikete yazılır.
                _od = h.get("banka_odeme")
                _od_kaynak = "bankanın yazdığı ödeme"
                if _od is None:
                    _od = round(sum(b["tutar"] for b in banka if b["tip"] == "ODEME"), 2)
                    _od_kaynak = "okunan ödeme (banka ödemeyi de yazmıyor)"
                bh = round(float(h["donem_borcu"]) - float(h["onceki_borc"]) + float(_od), 2)
                bh_kaynak = f"borç kimliğinden türetildi · {_od_kaynak}"
            oku_fark = None
            oku_yorum = None
            if bh is not None:
                f1 = round(float(bh) - oku_h, 2)          # faiz dahil yorumu
                f2 = round(float(bh) - oku_h_faizsiz, 2)  # faiz hariç yorumu
                _adaylar = [(f1, "banka dönem harcamasına FAİZİ KATIYOR (Yapı Kredi geleneği)"),
                            (f2, "banka dönem harcamasına FAİZİ KATMIYOR (Axess geleneği)")]
                # 🍎 ELMA İLE ELMA — GARANTİ FAİZİ TARİHSİZ BASAR (2026-08-24)
                # Garanti/Bonus ekstresinde DÖNEM FAİZİ, KKDF+BSMV, GEÇ ÖDEME ve
                # LİMİT AŞIM faizleri tablonun İÇİNDE değil başlık bloğunda,
                # TARİHSİZ satırlar olarak durur. Geometrik okuyucu tarihsiz satırı
                # işlem saymaz (doğrusu da budur). Ama borç kimliğinden türetilen
                # harcama faizi İÇERİR → okuma haksız yere "şüpheli" çıkıyordu;
                # fark tam faiz kadardı: Garanti Onur 20.979,46 = 15.722,62 DÖNEM
                # FAİZİ + 4.841,41 KKDF/BSMV + 287,93 GEÇ ÖDEME + 127,50 LİMİT AŞIM.
                # Okuma satırları DOĞRUYDU; faiz onun işi değildi. Üçüncü aday:
                # okunan satırlar + ekstrenin kendi beyan ettiği dönem faizi.
                _df = h.get("donem_faizi")
                # ⚠️ KOŞUL KOYMA TUZAĞI: ilk sürüm bu adayı yalnız "okumada HİÇ
                # faiz satırı yoksa" öneriyordu. Garanti'de okuma az miktarda
                # tarihli faiz de içeriyor (46,94 ₺) → koşul tutmadı, aday hiç
                # denenmedi ve 3 dönem boş yere kör kaldı. Oysa formül doğruydu:
                #   Garanti Onur Ağu: 180.606,82 − (159.627,36 + 20.979,46) = 0,00
                # Aday DAİMA denenir; zaten üçü arasından çapaya en yakın olan
                # seçiliyor — yanlış aday ancak daha iyi açıklıyorsa kazanır.
                if _df:
                    _adaylar.append((round(float(bh) - (oku_h + float(_df)), 2),
                                     "okunan satırlar + ekstrenin beyan ettiği dönem faizi "
                                     "(banka faizi TARİHSİZ basıyor — Garanti geleneği)"))
                oku_fark, oku_yorum = min(_adaylar, key=lambda p: abs(p[0]))

            # ── 2) Defter kalemleri (o döneme düşen paylarıyla) ─────────────
            cur.execute("""SELECT id, tarih, islem_turu, tutar::float AS tutar,
                                  COALESCE(taksit_sayisi,1) AS ts, aciklama,
                                  COALESCE(kaynak_tablo,'') AS kaynak
                             FROM kart_hareketleri
                            WHERE kart_id=%s AND durum='aktif' AND islem_turu <> 'DEVIR'
                              AND tarih >= %s::date - 400""", (kid, bit))
            defter = []
            for r in (cur.fetchall() or []):
                r = dict(r)
                ts = max(1, int(r["ts"] or 1))
                for i in range(1, ts + 1):
                    v = _satir_vade(r["tarih"], ts, kg, i)
                    if str(bas) < str(v) <= str(bit):
                        defter.append({
                            "id": r["id"], "tarih": str(r["tarih"])[:10],
                            "tip": "ODEME" if r["islem_turu"] == "ODEME" else "HARCAMA",
                            "tutar": round(float(r["tutar"]) / ts, 2),
                            "toplam": round(float(r["tutar"]), 2), "taksit_sayisi": ts,
                            "taksit_no": i, "aciklama": (r["aciklama"] or "")[:60],
                            "kaynak": r["kaynak"] or "sistem",
                        })

            # ── 3) Eşleştirme — tutar+tip, tarihe en yakın olan ─────────────
            kalan = list(defter)
            pdf_eksik = []
            for b in banka:
                aday = [(abs((_tarih_d(b["tarih"]) - _tarih_d(d["tarih"])).days), j)
                        for j, d in enumerate(kalan)
                        if d["tip"] == b["tip"] and abs(d["tutar"] - b["tutar"]) <= 0.02]
                if aday:
                    aday.sort()
                    kalan.pop(aday[0][1])
                else:
                    pdf_eksik.append(b)
            # ── 3b) YAKIN EŞLEŞME — "eksik" ile "farklı yazılmış"ı ayır ────
            # 🔔 UYARI BÜTÇESİ (2026-08-24): ilk sürüm, aynı ödemenin bankada
            # 50.107/13 Tem, defterde 50.000/13 Tem yazıldığı durumu İKİ AYRI
            # BULGU sayıyordu ("bankada var defterde yok" + "defterde var
            # bankada yok"). Oysa ortada kayıp kayıt yok, KÜÇÜK BİR SAPMA var.
            # Böyle çiftler her ay bağırsaydı duyu gürültüye dönerdi ve gerçek
            # eksik kayıt aralarında kaybolurdu. Artık ayrı bir kovada, farkıyla
            # birlikte raporlanıyor — GİZLENMİYOR ama ALARM da değil.
            yakin = []
            for b in list(pdf_eksik):
                en, ei = None, None
                for j, d in enumerate(kalan):
                    if d["tip"] != b["tip"]:
                        continue
                    tf = abs(d["tutar"] - b["tutar"])
                    gf = abs((_tarih_d(b["tarih"]) - _tarih_d(d["tarih"])).days)
                    # tutarın %1'i veya 250 ₺ (hangisi küçükse gevşek olan) + 7 gün
                    if tf <= max(1.0, min(250.0, b["tutar"] * 0.01)) and gf <= 7:
                        if en is None or (tf, gf) < en:
                            en, ei = (tf, gf), j
                if ei is not None:
                    d = kalan.pop(ei)
                    pdf_eksik.remove(b)
                    yakin.append({"banka": b, "defter": d,
                                  "tutar_farki": round(b["tutar"] - d["tutar"], 2),
                                  "gun_farki": abs((_tarih_d(b["tarih"]) - _tarih_d(d["tarih"])).days)})
            eks_t = round(sum(x["tutar"] for x in pdf_eksik if x["tip"] != "ODEME")
                          - sum(x["tutar"] for x in pdf_eksik if x["tip"] == "ODEME"), 2)
            faz_t = round(sum(x["tutar"] for x in kalan if x["tip"] != "ODEME")
                          - sum(x["tutar"] for x in kalan if x["tip"] == "ODEME"), 2)
            # Toplamlara YALNIZ ölçümü geçerli dönemler girer — okunamayan bir
            # PDF'in "eksiği" para değil, körlüktür; toplayınca sahte açık olur.
            if oku_fark is not None and abs(oku_fark) < 1:
                top_eksik += abs(eks_t); top_fazla += abs(faz_t)
            satir.update({
                # 🎚️ DURUM SIRALAMASI — ÖLÇÜM ALETİ ÖNCE DENETLENİR (2026-08-24)
                # Axess'te PDF'ten okunan harcama 0,00 çıkıyor (CID font) ve okuma
                # farkı dönemin TAMAMI kadar. O hâlde "defterde 27 fazla kayıt var"
                # demek defteri haksız yere suçlamaktır — eksik olan OKUMADIR.
                # Garanti/Ziraat'te ise bankanın dönem harcaması snapshot'ta YOK;
                # okuma doğrulanamıyor. İkisi de "satır farkı" diye bağırsaydı duyu
                # her ay 50+ sahte bulgu üretir, gerçek olanlar içinde kaybolurdu.
                # Sıra: önce aletin sağlamlığı, sonra defterin doğruluğu.
                "durum": ("OKUMA DOĞRULANAMIYOR — dönem harcaması ne yazılı ne türetilebilir"
                          if oku_fark is None else
                          "OKUMA ŞÜPHELİ — önce PDF okuması düzeltilmeli"
                          if abs(oku_fark) >= 1 else
                          "SATIR FARKI VAR" if (pdf_eksik or kalan)
                          else "YALNIZ YAKIN SAPMA" if yakin
                          else "TUTUYOR"),
                "olcum_gecerli": bool(oku_fark is not None and abs(oku_fark) < 1),
                "pdf_satir": len(banka), "defter_kalem": len(defter),
                "pdf_eksik": pdf_eksik,          # bankada var, defterde YOK
                "defter_fazla": kalan,           # defterde var, bankada YOK
                "yakin_eslesme": yakin,          # aynı kalem, küçük sapmayla yazılmış
                "yakin_adet": len(yakin),
                "pdf_eksik_tutar": eks_t, "defter_fazla_tutar": faz_t,
                "okuma_guveni": {
                    "pdf_okunan_harcama": oku_h,
                    "pdf_okunan_harcama_faizsiz": oku_h_faizsiz,
                    "banka_beyan_harcama": bh, "beyan_kaynagi": bh_kaynak,
                    "fark": oku_fark, "gelenek": oku_yorum,
                    "yorum": ("okuma tam — bankanın beyanıyla birebir"
                              if oku_fark is not None and abs(oku_fark) < 1
                              else "⚠ OKUMA ŞÜPHELİ — PDF'ten okunan satırlar bankanın "
                                   "beyan ettiği dönem harcamasını tutmuyor; aşağıdaki "
                                   "farklar DEFTERİ değil OKUMAYI gösteriyor olabilir"),
                },
            })
            cikti.append(satir)
        return {
            "donemler": cikti, "adet": len(cikti),
            "toplam_pdf_eksik": round(top_eksik, 2),
            "toplam_defter_fazla": round(top_fazla, 2),
            "olculebilen_donem": sum(1 for x in cikti if x.get("olcum_gecerli")),
            "okunamayan_donem": sum(1 for x in cikti if not x.get("olcum_gecerli")),
            "not": ("ÖNERİ-ONLY — hiçbir kayıt yazılmadı. 'pdf_eksik' = banka fatura "
                    "etmiş ama defterde yok; 'defter_fazla' = defterde var ama banka "
                    "o dönem fatura etmemiş. Taksitli satırlarda kıyas TAKSİT PAYI "
                    "üzerindendir (banka ne fatura ettiyse o). Bir dönemde "
                    "okuma_guveni.fark sıfırdan sapıyorsa önce OKUMADAN şüphelenin."),
        }


@app.get("/api/kartlar/mukerrer-odeme-adaylari")
def kart_mukerrer_odeme_adaylari(gun: int = 7, tolerans: float = 250.0):
    """🕵️ MÜKERRER KART ÖDEMESİ ADAYLARI — SALT OKUR, ÖNERİ-ONLY. Hiçbir şey silmez.

    ── NEDEN ───────────────────────────────────────────────────────────────
    Bir kart ödemesi sisteme İKİ AYRI KAPIDAN girebiliyor:
      1) Ödeme Merkezi'nden yapılır → "Kart borcu ödemesi" + kasa izi
      2) Ekstre içe aktarılır       → "ÖDEME-İNTERNET BANKACILIĞI" (kaynak_tablo='ekstre_import')
    Aynı ödeme her iki kapıdan geçerse defter borcu OLDUĞUNDAN DÜŞÜK gösterir.

    Canlı bulgu (2026-08-19, WORLD KART ANNEM) — dönem mutabakatı bunu ortaya
    çıkardı: bankanın Ağustos değişimi +55.405,57 ₺ iken defter −118.008,39 ₺.
    Pencerede iki şüpheli çift var:
        13 Tem  −50.000,00 (ekstre)  ↔  13 Tem  −50.107,00 (sistem)
        22 Tem −120.000,00 (ekstre)  ↔  24 Tem −120.107,00 (sistem)
    Toplam 170.000 ₺ — dönemdeki 173.413,96 ₺'lik açığın neredeyse tamamı.
    Aradaki 107 ₺'lik fark, satır-satır mükerrer freninin bu çiftleri
    KAÇIRMASININ sebebi: fren birebir tutar arıyordu.

    ── NE YAPMAZ ───────────────────────────────────────────────────────────
    Hiçbir kaydı iptal ETMEZ. İki kayıt gerçekten iki ayrı ödeme olabilir
    (aynı gün iki taksit yatırmak olağandır). Karar SAHİBİNDİR; bu uç sadece
    kanıtı yan yana koyar. Silmek gerekirse zaten izli iptal yolu var
    (DELETE /api/kart-hareketleri/{id}) — GERİ-ALMA ≠ SİLME.

    Parametreler:
      gun       kaç gün arayla düşen ödemeler çift sayılsın (varsayılan 7)
      tolerans  tutar farkı bu ₺ değerinden küçükse "yakın" sayılır (250 ₺)
    """
    with db() as (conn, cur):
        cur.execute("""
            SELECT a.id AS a_id, b.id AS b_id, k.kart_adi,
                   a.kart_id::text AS kart_id,
                   a.tarih::text AS a_tarih, b.tarih::text AS b_tarih,
                   a.tutar::float AS a_tutar, b.tutar::float AS b_tutar,
                   a.aciklama AS a_aciklama, b.aciklama AS b_aciklama,
                   COALESCE(a.kaynak_tablo,'') AS a_kaynak,
                   COALESCE(b.kaynak_tablo,'') AS b_kaynak
              FROM kart_hareketleri a
              JOIN kart_hareketleri b
                ON b.kart_id = a.kart_id
               AND b.id > a.id
               AND b.durum = 'aktif'
               AND b.islem_turu = 'ODEME'
               AND ABS(b.tarih - a.tarih) <= %s
               AND ABS(b.tutar - a.tutar) <= %s
              JOIN kartlar k ON k.id = a.kart_id
             WHERE a.durum = 'aktif' AND a.islem_turu = 'ODEME'
             ORDER BY k.kart_adi, a.tarih
        """, (gun, tolerans))
        ciftler = []
        for r in (cur.fetchall() or []):
            d = dict(r)
            fark = round(abs(d["a_tutar"] - d["b_tutar"]), 2)
            gun_farki = abs((datetime.strptime(d["b_tarih"], "%Y-%m-%d")
                             - datetime.strptime(d["a_tarih"], "%Y-%m-%d")).days)
            kaynaklar = {d["a_kaynak"], d["b_kaynak"]}
            # 🔴 EN GÜÇLÜ KANIT: biri ekstreden, diğeri sistemden gelmişse aynı
            # ödemenin iki kapıdan girmiş olma ihtimali yüksektir. İkisi de aynı
            # kapıdansa büyük olasılıkla GERÇEKTEN iki ayrı ödemedir.
            iki_kapi = ("ekstre_import" in kaynaklar) and (len(kaynaklar) > 1)
            guc = ("YÜKSEK" if iki_kapi and fark < 1 else
                   "ORTA" if iki_kapi else
                   "DÜŞÜK")
            ciftler.append({
                **d, "tutar_farki": fark, "gun_farki": gun_farki,
                "iki_farkli_kapi": iki_kapi, "supheye_guc": guc,
                "toplam_etki": round(min(d["a_tutar"], d["b_tutar"]), 2),
                "not": ("Biri ekstreden biri sistemden — aynı ödemenin iki kez "
                        "girmiş olma ihtimali yüksek." if iki_kapi else
                        "İkisi de aynı kapıdan — büyük olasılıkla iki ayrı ödeme."),
            })
        yuksek = [c for c in ciftler if c["supheye_guc"] in ("YÜKSEK", "ORTA")]
        return {
            "ciftler": ciftler,
            "adet": len(ciftler),
            "supheli_adet": len(yuksek),
            "supheli_toplam": round(sum(c["toplam_etki"] for c in yuksek), 2),
            "parametreler": {"gun": gun, "tolerans": tolerans},
            "not": ("ÖNERİ-ONLY — hiçbir kayıt silinmedi. İki kayıt gerçekten iki "
                    "ayrı ödeme olabilir; karar sahibindir. Doğrulamak için ilgili "
                    "dönemin /api/kartlar/donem-mutabakati farkına bakın: mükerrer "
                    "ödeme varsa fark POZİTİF (DEFTER EKSİK) çıkar ve büyüklüğü "
                    "mükerrer tutara yakındır."),
        }


@app.get("/api/kartlar/donem-mutabakati")
def kart_donem_mutabakati():
    """🧮 DÖNEM DEĞİŞİMİ MUTABAKATI — SALT OKUR. Kartın TEK dairesel-olmayan ölçümü.

    ── NEDEN BU UÇ VAR ─────────────────────────────────────────────────────
    Bugüne kadarki `mutabakat_farki` şu soruyu soruyordu:
        «defterin TOPLAMI, bankanın dediği TOPLAMA eşit mi?»
    Bu soru DAİRESELDİR. Çünkü defterdeki DEVİR satırı zaten
    `donem_borcu − kart_borc()` diye hesaplanıp yazılmış bir DENKLEŞTİRME
    YAMASIDIR — yani defteri bankaya EŞİTLEMEK için konmuştur. Yamayı içeren
    bir toplamı bankayla kıyaslamak, cevabı soruya yazıp sonra sormaktır.

    Canlı kanıt (2026-08-19, WORLD KART ANNEM):
      mutabakat_farki .......... 3.903,00 ₺  ← "defter tutmuyor" diye okunur
      kesim SONRASI hareketler . 3.903,00 ₺  ← 4 otomatik internet talimatı
    İkisi KURUŞU KURUŞUNA aynı. Çünkü yama, ekstre kesiminden SONRAKİ o dört
    satırı da yutacak şekilde hesaplanmış. Yani o 3.903 ₺ bir uyuşmazlık
    değil, YAMANIN GÖLGESİ. Aynı sebeple diğer kartların 0,00'ı da bir
    başarı değil: yama zaten sıfırlıyor. Ölçüm boş çıkıyordu.

    ── BU UCUN SORDUĞU SORU (dairesel DEĞİL) ───────────────────────────────
        «Bankanın borcu iki ekstre arasında NE KADAR DEĞİŞTİ, defterdeki
         hareketler aynı pencerede NE KADAR DEĞİŞTİRDİ?»
    Yama pencerenin dışında kalır (tek seferlik ve eski tarihli), pencere
    içine düşerse de DEVİR satırları toplama katılmaz + `devir_pencerede`
    bayrağıyla bildirilir. Böylece ölçüm yamadan bağımsızlaşır.

        banka_degisim  = donem_borcu(N) − donem_borcu(N−1)
        defter_degisim = Σ(HARCAMA+FAİZ − ÖDEME),  kesim(N−1) < tarih ≤ kesim(N)
        fark           = banka_degisim − defter_degisim
        fark ≠ 0  →  o dönemde defterde EKSİK/FAZLA satır var (gerçek bulgu)

    Ayrıca ekstrenin KENDİ İÇ ÇAPASI da ölçülür — banka kendi rakamlarıyla
    tutarlı mı: (dönem harcaması − dönem ödemesi + dönem faizi) = değişim?
    Bu ikisi ayrı ayrı bilgidir: `fark` defteri, `capa_farki` okumayı denetler.

    ⚠️ Tek ekstresi olan kartta ölçüm YAPILAMAZ (pencere kurulamaz) — bu
    durum GİZLENMEZ, "tek dönem" olarak raporlanır. HATA ≠ BOŞ.
    """
    with db() as (conn, cur):
        cur.execute("""SELECT id::text AS id, kart_adi FROM kartlar
                        WHERE aktif=TRUE ORDER BY kart_adi""")
        kartlar = [dict(r) for r in (cur.fetchall() or [])]
        out, olculen, sapan, toplam_sapma = [], 0, 0, 0.0
        for k in kartlar:
            kid = k["id"]
            # Dönem başına TEK snapshot (aynı döneme birden çok içe aktarım
            # olabiliyor — en SON yazılan geçerlidir).
            cur.execute("""
                SELECT DISTINCT ON (donem)
                       donem::text AS donem, kesim_tarihi::text AS kesim,
                       donem_borcu::float AS borc,
                       donem_harcama::float AS harcama,
                       donem_odeme::float  AS odeme,
                       donem_faizi::float  AS faiz
                  FROM kart_ekstre_donem
                 WHERE kart_id=%s AND donem_borcu IS NOT NULL
                 ORDER BY donem, olusturma DESC
            """, (kid,))
            snaps = [dict(r) for r in (cur.fetchall() or [])]
            cur.execute("SELECT COALESCE(kesim_gunu,1) AS kg FROM kartlar WHERE id=%s", (kid,))
            _kg = int((cur.fetchone() or {}).get("kg") or 1)
            # 🧱 GEÇMİŞ BARİYERİ — defter ne zaman başlıyor? Bu tarihten ÖNCEKİ
            # dönemlerde "defter eksik" demek YANLIŞ ALARM olur: eksik olan
            # hareket değil, o günlere ait KAYIT HİÇ GİRİLMEMİŞ. İki durumu
            # ayırmadan sayı vermek, sahibi olmayan bir açığı kovalatır.
            cur.execute("""SELECT MIN(tarih)::text AS t FROM kart_hareketleri
                            WHERE kart_id=%s AND durum='aktif' AND islem_turu <> 'DEVIR'""",
                        (kid,))
            defter_basi = (cur.fetchone() or {}).get("t")
            if len(snaps) < 2:
                out.append({"kart_id": kid, "kart_adi": k["kart_adi"],
                            "durum": "TEK DÖNEM — ölçülemez",
                            "donem_sayisi": len(snaps), "donemler": []})
                continue
            donemler = []
            for onceki, simdi in zip(snaps, snaps[1:]):
                bas = onceki.get("kesim") or onceki.get("donem")
                bit = simdi.get("kesim") or simdi.get("donem")
                if not bas or not bit or bas >= bit:
                    donemler.append({"donem": simdi["donem"],
                                     "durum": "PENCERE KURULAMADI",
                                     "kesim_bas": bas, "kesim_bit": bit})
                    continue
                # 💳 TAKSİT YAYILIMI — ilk ölçümdeki en büyük yöntem hatası.
                # 12 taksitli 12.000 ₺'lik alışveriş ekstreye AY AY 1.000 ₺
                # olarak düşer; defterde ise TEK satır 12.000 ₺ durur. Ham
                # toplam alınırsa alışveriş ayında defter FAZLA, sonraki 11 ayda
                # EKSİK görünür — ikisi de sahte. Bu yüzden her hareket, taksit
                # sayısı kadar birer ay arayla PARÇALARA açılıp öyle toplanıyor.
                # Ham toplam da yanında raporlanıyor (yöntem şeffaf kalsın).
                # 🗓️ TAKSİT TAKVİMİ BANKANIN TAKVİMİDİR (2026-08-24, canlı ders)
                # İlk sürüm taksitleri İŞLEM TARİHİ YILDÖNÜMÜNE koyuyordu
                # (12 Şub → 12 Mar → 12 Nis…). Banka öyle yapmıyor: taksidi
                # EKSTRE KESİMİNDE fatura ediyor. WORLD ANNEM'de 12 Şubat'ta
                # yapılan 6 taksitli alım bankada Mart–Ağustos ekstrelerine
                # düşerken benim modelim Şubat–Temmuz'a koyuyordu → Temmuz
                # penceresine İKİ taksit sığdı, Ağustos'a HİÇ düşmedi. Kendi
                # düzeltmemin sonucunu bozan şey buydu (Temmuz +3.305,96 iken
                # −3.307,96'ya döndü — işaret değişimi model hatasının imzası).
                # Doğrusu: 1. taksit, alımdan SONRAKİ ilk kesimde başlar.
                # Tek çekimde davranış DEĞİŞMEZ (tarih neyse o) — patlama yüzeyi dar.
                cur.execute("""
                    SELECT COALESCE(SUM(pay),0)::float AS d, COUNT(*) AS n
                      FROM (
                        SELECT (CASE WHEN h.islem_turu='ODEME' THEN -h.tutar ELSE h.tutar END)
                               / GREATEST(COALESCE(h.taksit_sayisi,1),1) AS pay,
                               CASE
                                 WHEN COALESCE(h.taksit_sayisi,1) <= 1 THEN h.tarih
                                 ELSE (
                                   CASE WHEN EXTRACT(DAY FROM h.tarih) <= %s
                                        THEN date_trunc('month', h.tarih)::date
                                        ELSE (date_trunc('month', h.tarih)
                                              + INTERVAL '1 month')::date
                                   END
                                   + (%s - 1)
                                   + ((g.i-1) || ' month')::interval
                                 )::date
                               END AS vade
                          FROM kart_hareketleri h
                          CROSS JOIN LATERAL generate_series(
                                 1, GREATEST(COALESCE(h.taksit_sayisi,1),1)) AS g(i)
                         WHERE h.kart_id=%s AND h.durum='aktif' AND h.islem_turu <> 'DEVIR'
                      ) t
                     WHERE vade > %s::date AND vade <= %s::date
                """, (_kg, _kg, kid, bas, bit))
                dr = dict(cur.fetchone() or {})
                defter_degisim = float(dr.get("d") or 0)
                cur.execute("""
                    SELECT COALESCE(SUM(CASE WHEN islem_turu='ODEME' THEN -tutar
                                             ELSE tutar END),0)::float AS d,
                           COALESCE(SUM(CASE WHEN islem_turu='ODEME' THEN tutar
                                             ELSE 0 END),0)::float AS o,
                           COALESCE(SUM(CASE WHEN islem_turu='ODEME' THEN 0
                                             ELSE tutar END),0)::float AS h
                      FROM kart_hareketleri
                     WHERE kart_id=%s AND durum='aktif' AND islem_turu <> 'DEVIR'
                       AND tarih > %s::date AND tarih <= %s::date
                """, (kid, bas, bit))
                _hr = dict(cur.fetchone() or {})
                defter_ham = float(_hr.get("d") or 0)
                # 🔍 TARAF AYRIMI (2026-08-23) — "fark var" demek yetmiyor, İZ
                # sürebilmek için farkın HANGİ TARAFTAN geldiği lazım. Harcama
                # tarafında fark varsa deftere girmemiş bir HARCAMA; ödeme
                # tarafındaysa girmemiş/fazla girmiş bir ÖDEME vardır. Bu ayrım
                # olmadan 3.306 ₺'yi aramak samanlıkta iğne aramaktır.
                defter_harcama = float(_hr.get("h") or 0)
                defter_odeme = float(_hr.get("o") or 0)
                cur.execute("""SELECT COUNT(*) AS n FROM kart_hareketleri
                                WHERE kart_id=%s AND durum='aktif' AND islem_turu='DEVIR'
                                  AND tarih > %s::date AND tarih <= %s::date""",
                            (kid, bas, bit))
                devir_ic = int((cur.fetchone() or {}).get("n") or 0)
                banka_degisim = round(float(simdi["borc"]) - float(onceki["borc"]), 2)
                fark = round(banka_degisim - defter_degisim, 2)
                # 🎯 EKSTRENİN KENDİ İÇ ÇAPASI — banka kendi rakamıyla tutarlı mı?
                # İlk sürümde faiz AYRICA ekleniyordu ve çapa farkı tam olarak
                # −faiz çıkıyordu (WORLD ANNEM Ağu: −14.087,01 = DÖNEM FAİZİ'nin
                # kendisi). Bu, faizin `donem_harcama` İÇİNDE olduğunun kanıtı;
                # ayrıca eklemek MÜKERRER sayımdı. Düzeltince çapa 0,00'a oturdu
                # — yani ekstre okumaları BİRBİRİYLE TUTARLI. Faiz ayrı bilgi
                # olarak taşınıyor ama çapaya İKİNCİ KEZ girmiyor.
                # 🏦 İKİ BANKA GELENEĞİ (2026-08-24, Axess açılınca ölçüldü):
                # Yapı Kredi `dönem harcaması`na faizi KATIYOR, Axess KATMIYOR.
                # Tek yorumla çapa kurulursa bankalardan biri hep sapık görünür
                # (Axess'te tam dönem faizi kadar: 1.292,38 / 1.496,23).
                # İki yorum da denenir, tutan kabul edilir.
                capa = None
                if simdi.get("harcama") is not None or simdi.get("odeme") is not None:
                    _h = float(simdi.get("harcama") or 0)
                    _o = float(simdi.get("odeme") or 0)
                    _f = float(simdi.get("faiz") or 0)
                    _c1 = round(_h - _o, 2)        # faiz zaten harcamanın içinde
                    _c2 = round(_h - _o + _f, 2)   # faiz ayrı kalem
                    capa = _c1 if abs(banka_degisim - _c1) <= abs(banka_degisim - _c2) else _c2
                gecmis_yok = bool(defter_basi and bas < defter_basi)
                if gecmis_yok:
                    durum = "GEÇMİŞ YOK — ölçüm dışı"
                elif abs(fark) < 1:
                    durum = "TUTUYOR"
                else:
                    durum = "DEFTER EKSİK" if fark > 0 else "DEFTER FAZLA"
                donemler.append({
                    "donem": simdi["donem"], "kesim_bas": bas, "kesim_bit": bit,
                    "banka_degisim": banka_degisim,
                    "defter_degisim": round(defter_degisim, 2),
                    "defter_degisim_ham": round(defter_ham, 2),
                    "taksit_yayilim_etkisi": round(defter_degisim - defter_ham, 2),
                    "hareket_adet": int(dr.get("n") or 0),
                    "fark": fark,
                    "ekstre_ic_capa": capa,
                    "capa_farki": (None if capa is None else round(banka_degisim - capa, 2)),
                    "donem_faizi": simdi.get("faiz"),
                    # taraf ayrımı — farkın hangi taraftan geldiğini gösterir
                    "banka_harcama": simdi.get("harcama"),
                    "defter_harcama": round(defter_harcama, 2),
                    "harcama_farki": (None if simdi.get("harcama") is None
                                      else round(float(simdi["harcama"]) - defter_harcama, 2)),
                    "banka_odeme": simdi.get("odeme"),
                    "defter_odeme": round(defter_odeme, 2),
                    "odeme_farki": (None if simdi.get("odeme") is None
                                    else round(float(simdi["odeme"]) - defter_odeme, 2)),
                    "devir_pencerede": devir_ic,
                    "defter_basi": defter_basi,
                    "durum": durum,
                })
                if gecmis_yok:
                    continue
                olculen += 1
                if abs(fark) >= 1:
                    sapan += 1
                    toplam_sapma += abs(fark)
            olculebilir = [d for d in donemler if d.get("durum") in ("TUTUYOR", "DEFTER EKSİK", "DEFTER FAZLA")]
            out.append({"kart_id": kid, "kart_adi": k["kart_adi"],
                        "donem_sayisi": len(snaps),
                        "defter_basi": defter_basi,
                        "durum": ("ÖLÇÜLEBİLİR DÖNEM YOK" if not olculebilir
                                  else "TUTUYOR" if all(d["durum"] == "TUTUYOR" for d in olculebilir)
                                  else "SAPMA VAR"),
                        "donemler": donemler})
        return {
            "kartlar": out,
            "olculen_donem": olculen,
            "sapan_donem": sapan,
            "toplam_sapma": round(toplam_sapma, 2),
            "not": ("Bu ölçüm DEVİR yamasından bağımsızdır: iki ekstre arasındaki "
                    "DEĞİŞİM kıyaslanır, TOPLAM değil. 'fark' sıfırdan sapıyorsa o "
                    "dönemde deftere girmemiş (veya fazladan girmiş) hareket vardır. "
                    "'capa_farki' ise bankanın kendi rakamlarının tutarlılığını ölçer."),
        }


@app.get("/api/kartlar/devir-denetimi")
def kart_devir_denetimi():
    """🔬 DEVİR YAMASI DENETİMİ — SALT OKUR, hiçbir şey yazmaz.

    🔴 CODEX HÜKMÜ (2026-08-18, kodu okuyarak buldu):
    Bu sistemdeki DEVİR satırları HAM VERİ DEĞİLDİR. main.py'de iki yerde
    (`adj = donem_borcu - kart_borc(cur, kid)`) o günkü formüle göre hesaplanıp
    TEK SATIRLIK DEVİR olarak yazılıyorlar — yani bir DENKLEŞTİRME YAMASI
    (balancing plug). Formül ya da veri sonradan değişince yama BAYATLIYOR ve
    "mutabakat farkı" diye görünen şey aslında yamanın eskimişliği oluyor.

    Bugün bu tuzağa BEŞ KEZ düştüm: sentetik bir denkleştirme satırını kaynak
    veri sanıp etrafındaki ham geçmişi oynatmaya kalktım (geçmiş ekstre aktar,
    taksit dönüştür, devir öncesini sil, mutabakatı kesime sabitle...). Beşinde
    de ölçüm durdurdu; biri canlıda bir kartın 0,00'ını bozdu ve geri alındı.

    Bu rapor Codex'in koyduğu DURDURUCU KURALIN aracıdır:
      «DEVİR/override/fallback içeren hiçbir sayıya bakıp tarihsel veriyi
       DEĞİŞTİRME; önce o sayıyı ham HARCAMA/ODEME/FAIZ üzerinden bir İYİ ve
       bir KÖTÜ kartta birebir yeniden üret. Üretemiyorsan bozuk olan veri
       değil, TÜRETİLMİŞ KATMANDIR.»

    Ölçülenler:
      stored_devir   defterdeki DEVİR satır(lar)ının toplamı
      expected_devir ekstre_borcu − devir günündeki non-DEVİR borç
                     (yani "bugünkü kuralla yeniden hesaplansaydı ne olurdu")
      sapma          stored − expected → YAMANIN BAYATLIK ÖLÇÜSÜ
      same_day_net   devir günündeki diğer hareketlerin neti (sıra etkisi)
    """
    with db() as (conn, cur):
        cur.execute("""SELECT id::text AS id, kart_adi FROM kartlar
                        WHERE aktif=TRUE ORDER BY kart_adi""")
        kartlar = [dict(r) for r in (cur.fetchall() or [])]
        out = []
        for k in kartlar:
            kid = k["id"]
            cur.execute("""SELECT MAX(tarih)::text AS t, COALESCE(SUM(tutar),0) AS s
                             FROM kart_hareketleri
                            WHERE kart_id=%s AND durum='aktif' AND islem_turu='DEVIR'""", (kid,))
            dv = dict(cur.fetchone() or {})
            devir_ts, stored = dv.get("t"), float(dv.get("s") or 0)
            cur.execute("""SELECT kesim_tarihi::text AS kesim, donem_borcu
                             FROM kart_ekstre_donem
                            WHERE kart_id=%s AND donem_borcu IS NOT NULL
                            ORDER BY donem DESC, olusturma DESC LIMIT 1""", (kid,))
            sn = dict(cur.fetchone() or {})
            kesim, eb = sn.get("kesim"), (float(sn["donem_borcu"]) if sn.get("donem_borcu") is not None else None)
            satir = {"kart_id": kid, "kart_adi": k["kart_adi"],
                     "kesim": kesim, "devir_tarihi": devir_ts,
                     "stored_devir": round(stored, 2)}
            if not devir_ts or eb is None:
                satir.update({"expected_devir": None, "sapma": None, "same_day_net": None,
                              "durum": "devir yok" if not devir_ts else "ekstre snapshot yok"})
                out.append(satir); continue
            # Devir GÜNÜNE kadarki (dahil) non-DEVİR borç — ham hareketlerden
            # 🔴 FORMÜL DÜZELTMESİ (2026-08-18, kuru çalıştırma yakaladı):
            # Önce "devir GÜNÜNE KADARKİ ham borç" kullanıyordum. YANLIŞ —
            # defter TÜM hareketleri toplar, devirden sonrakileri de. O yüzden
            # yamanın doğru değeri `ekstre_borcu − TÜM ham hareketler`dir.
            # Yanlış formül, ÇALIŞAN iki kartı bayat gösteriyordu:
            #   Garanti Onur: ham(tümü) −113.261,15 + yama 597.238,22
            #                 = 483.977,07 = BANKA ✓ (yama GÜNCEL)
            #   ama eski formül "olması gereken 533.404,75" diyordu → tazeleseydim
            #   o kartın kuruşu kuruşuna tutan defterini 63.833 ₺ BOZACAKTIM.
            # 🔴 İKİNCİ FORMÜL DÜZELTMESİ (2026-08-18, canlı hasarla öğrenildi):
            # Ham SUM(tutar) kullanıyordum. YANLIŞ — `kart_borc()` TAKSİT
            # FARKINDALIDIR: taksitli alımı `tutar/adet × geçen taksit` sayar,
            # ham toplam ise `tutar`ın TAMAMINI sayar. İkisi taksitli kartlarda
            # ayrışır. Uyguladığımda Garanti Fethi ve OPET'in kuruşu kuruşuna
            # tutan defterlerini bozdu (72.307 ve 91.991 ₺); eski değerler geri
            # yüklendi, kalıcı hasar olmadı.
            # DOĞRUSU sistemin KENDİ formülü: main.py'deki yama yazıcıları da
            # `adj = donem_borcu - kart_borc(cur, kid)` kullanıyor. Denetim de
            # aynı fonksiyonu kullanmalı, yoksa "olması gereken" uydurma olur.
            # ⚠️ kart_borc DEVİR'i de içerir → ham = kart_borc − stored_devir
            ham = float(kart_borc(cur, kid)) - stored
            cur.execute("""SELECT COALESCE(SUM(CASE WHEN islem_turu='ODEME' THEN -tutar
                                                    ELSE tutar END),0) AS d
                             FROM kart_hareketleri
                            WHERE kart_id=%s AND durum='aktif' AND islem_turu<>'DEVIR'
                              AND tarih = %s::date""", (kid, devir_ts))
            aynigun = float((cur.fetchone() or {}).get("d") or 0)
            beklenen = round(eb - ham, 2)
            sapma = round(stored - beklenen, 2)
            satir.update({
                "ekstre_borcu": eb, "devir_gunune_kadar_ham": round(ham, 2),
                "expected_devir": beklenen, "sapma": sapma,
                "same_day_net": round(aynigun, 2),
                "devir_kesimden_sonra": bool(kesim and devir_ts > kesim),
                "durum": ("YAMA GÜNCEL" if abs(sapma) < 1 else "YAMA BAYAT"),
            })
            out.append(satir)
    bayat = [o for o in out if o.get("durum") == "YAMA BAYAT"]
    return {
        "kartlar": out,
        "bayat_adet": len(bayat),
        "bayat_toplam_sapma": round(sum(abs(o["sapma"]) for o in bayat), 2),
        "not": "SALT OKUR — hiçbir şey yazılmadı. `sapma`, DEVİR yamasının bugünkü "
               "kurala göre ne kadar eskidiğini ölçer. Bu rapor görülmeden hiçbir "
               "geçmiş hareket silinmemeli, ekstre taşınmamalı, taksit dönüştürülmemeli, "
               "DEVİR yeniden yazılmamalıdır (Codex durdurucu kuralı).",
    }


@app.get("/api/kartlar/taksit-plani")
def kart_taksit_plani_listele(kart_id: Optional[str] = None):
    """📅 Aktif taksit planları + AY AY gelecek yük takvimi.

    Bu, kart borcunun GÖRÜNMEYEN yarısıdır: ekstre "bu ay ne ödeyeceksin"
    der, bu görünüm "önümüzdeki aylarda ne çıkacak" der. Defterden AYRI
    yaşar (borç üretmez) — bkz. /kartlar/taksit-plani-kur gerekçesi.
    """
    with db() as (conn, cur):
        try:
            cur.execute("SAVEPOINT sp_tp")
            _k = " AND p.kart_id=%s" if kart_id else ""
            _p = [kart_id] if kart_id else []
            cur.execute(f"""
                SELECT p.id, p.kart_id, k.kart_adi, p.aciklama, p.toplam_tutar,
                       p.taksit_adedi, p.ilk_donem, p.durum
                  FROM kart_taksit_plani p
                  JOIN kartlar k ON k.id = p.kart_id
                 WHERE p.durum='aktif'{_k}
                 ORDER BY p.olusturma DESC
            """, _p)
            planlar = [dict(r) for r in (cur.fetchall() or [])]
            if not planlar:
                cur.execute("RELEASE SAVEPOINT sp_tp")
                return {"planlar": [], "gelecek_toplam": 0.0, "takvim": [],
                        "not": "Aktif taksit planı yok."}
            cur.execute(f"""
                SELECT d.plan_id, d.taksit_no, d.donem, d.tutar, d.durum
                  FROM kart_taksit_dilimi d
                  JOIN kart_taksit_plani p ON p.id = d.plan_id
                 WHERE p.durum='aktif'{_k}
                 ORDER BY d.donem, d.taksit_no
            """, _p)
            dilimler = [dict(r) for r in (cur.fetchall() or [])]
            cur.execute("RELEASE SAVEPOINT sp_tp")
        except Exception as e:  # noqa: BLE001 — tablo yoksa boş dön, sistem düşmesin
            try:
                cur.execute("ROLLBACK TO SAVEPOINT sp_tp"); cur.execute("RELEASE SAVEPOINT sp_tp")
            except Exception:  # noqa: BLE001
                pass
            return {"planlar": [], "gelecek_toplam": 0.0, "takvim": [],
                    "not": f"Taksit planı okunamadı: {str(e)[:100]}"}

    _pd = {}
    for d in dilimler:
        _pd.setdefault(d["plan_id"], []).append(d)
    out, takvim = [], {}
    for p in planlar:
        ds = _pd.get(p["id"], [])
        bek = [d for d in ds if d["durum"] == "beklenen"]
        for d in bek:
            _ay = str(d["donem"])[:7]
            takvim[_ay] = round(takvim.get(_ay, 0.0) + float(d["tutar"]), 2)
        out.append({
            "id": p["id"], "kart_adi": p["kart_adi"], "aciklama": p["aciklama"],
            "toplam_tutar": float(p["toplam_tutar"]), "taksit_adedi": int(p["taksit_adedi"]),
            "ilk_donem": str(p["ilk_donem"]) if p["ilk_donem"] else None,
            "odenen_taksit": len(ds) - len(bek), "kalan_taksit": len(bek),
            "kalan_tutar": round(sum(float(d["tutar"]) for d in bek), 2),
            "dilim_tutari": round(float(p["toplam_tutar"]) / int(p["taksit_adedi"]), 2),
        })
    out.sort(key=lambda x: -x["kalan_tutar"])
    return {
        "planlar": out,
        "gelecek_toplam": round(sum(o["kalan_tutar"] for o in out), 2),
        "takvim": [{"ay": a, "tutar": t} for a, t in sorted(takvim.items())],
        "not": "Bu tutarlar kart BORCUNA DAHİL DEĞİLDİR — henüz ekstreye girmediler. "
               "Ekstre 'bu ay ne ödeyeceksin' der; bu görünüm 'önümüzdeki aylarda ne "
               "çıkacak' der.",
    }


class TaksitPlanBody(BaseModel):
    kart_id: str
    kesim_tarihi: str
    dilimler: List["TaksitDonusumIslem"]
    uygula: bool = False


@app.post("/api/kartlar/taksit-plani-kur")
def kart_taksit_plani_kur(body: TaksitPlanBody):
    """📅 ADIM 8 (GÜVENLİ YOL) — gelecek taksitleri PLANA yazar, DEFTERE DEĞİL.

    🔴 NEDEN DÖNÜŞTÜRME DEĞİL: `/taksit-donustur` kuru çalıştırması ESER
    TİCARET'te +82.500 ₺ borç artışı gösterdi. Sebep: borç motoru taksiti
    KÜMÜLATİF sayar; geçmiş taksitler (Haziran 1/4, Temmuz 2/4) ise defterdeki
    DEVİR satırının içinde ZATEN var. Dönüştürmek onları ikinci kez
    borçlandırırdı ve şu an kuruşu kuruşuna tutan mutabakatı (0,00) bozardı.

    ✅ DOĞRU MODEL: geçmişe DOKUNMA, geleceği AYRI TUT.
    Kalan taksitler `kart_taksit_plani` + `kart_taksit_dilimi` tablolarına
    yazılır (Adım 2'de tam bunun için kurulmuşlardı). Kart defteri hiç
    değişmez → mutabakat bozulmaz; buna karşılık "önümüzdeki aylarda bu karttan
    ne çıkacak" sorusu ilk kez cevaplanır.

    Dilim durumu: bu dönem ve öncesi 'gerceklesti' (ekstrede zaten borç oldu),
    sonrası 'beklenen'. Böylece plan geçmişi de taşır ama borç üretmez.

    İdempotent: aynı kart + aynı açıklama + aynı toplam + aynı taksit adedi
    için ikinci kez plan açılmaz.
    """
    from datetime import date as _d
    try:
        _kesim = _d.fromisoformat(str(body.kesim_tarihi)[:10])
    except Exception:
        raise HTTPException(400, "kesim_tarihi YYYY-AA-GG olmalı")
    with db() as (conn, cur):
        cur.execute("SELECT kart_adi FROM kartlar WHERE id=%s AND aktif=TRUE", (body.kart_id,))
        _k = cur.fetchone()
        if not _k:
            raise HTTPException(404, "Kart bulunamadı")
        from database import ensure_kart_gercek_modeli
        ensure_kart_gercek_modeli(cur)

        kurulan, atlanan, ozet = [], [], 0.0
        for d in body.dilimler:
            no, adet = int(d.taksit_no or 0), int(d.taksit_sayisi or 0)
            if not (1 <= no <= adet <= 60):
                atlanan.append({"satici": d.satici, "neden": f"geçersiz taksit {no}/{adet}"})
                continue
            dilim = round(abs(float(d.dilim_tutari)), 2)
            toplam = round(float(d.alim_toplami or 0) or dilim * adet, 2)
            ad = (d.satici or "Taksitli alım")[:120]
            # İlk dönem = bu kesim − (no−1) ay
            _ay, _yil = _kesim.month - (no - 1), _kesim.year
            while _ay <= 0:
                _ay += 12; _yil -= 1
            ilk = _d(_yil, _ay, min(_kesim.day, 28))
            cur.execute("""SELECT id FROM kart_taksit_plani
                            WHERE kart_id=%s AND aciklama=%s AND durum='aktif'
                              AND ROUND(toplam_tutar::numeric,2)=ROUND(%s::numeric,2)
                              AND taksit_adedi=%s LIMIT 1""",
                        (body.kart_id, ad, toplam, adet))
            if cur.fetchone():
                atlanan.append({"satici": ad, "neden": "bu plan zaten kurulu"})
                continue
            kalan = max(0, adet - no)
            kurulan.append({"satici": ad, "toplam": toplam, "taksit": f"{no}/{adet}",
                            "ilk_donem": str(ilk), "kalan_taksit": kalan,
                            "kalan_tutar": round(kalan * dilim, 2)})
            ozet += kalan * dilim
            if not body.uygula:
                continue
            cur.execute("""INSERT INTO kart_taksit_plani
                (kart_id, aciklama, toplam_tutar, taksit_adedi, ilk_donem, durum)
                VALUES (%s,%s,%s,%s,%s,'aktif') RETURNING id""",
                (body.kart_id, ad, toplam, adet, ilk))
            pid = cur.fetchone()["id"]
            for s in range(1, adet + 1):
                _a, _y = ilk.month + (s - 1), ilk.year
                while _a > 12:
                    _a -= 12; _y += 1
                cur.execute("""INSERT INTO kart_taksit_dilimi
                    (plan_id, taksit_no, donem, tutar, durum)
                    VALUES (%s,%s,%s,%s,%s) ON CONFLICT (plan_id, taksit_no) DO NOTHING""",
                    (pid, s, _d(_y, _a, min(ilk.day, 28)), dilim,
                     "gerceklesti" if s <= no else "beklenen"))
            audit(cur, "kart_taksit_plani", pid, "INSERT")
    return {
        "kuru_calistirma": not body.uygula, "kart": _k["kart_adi"],
        "plan": kurulan, "atlanan": atlanan,
        "gelecek_taksit_yuku": round(ozet, 2),
        "not": "Kart DEFTERİNE DOKUNULMADI — borç değişmez, mutabakat bozulmaz. "
               "Bu plan yalnız «önümüzdeki aylarda ne çıkacak» sorusunu cevaplar."
               + ("" if body.uygula else " Hiçbir şey yazılmadı; uygula=true gönderin."),
    }


class TaksitDonusumIslem(BaseModel):
    tarih: str
    dilim_tutari: float          # bu dönemin taksit tutarı (ekstredeki satır)
    taksit_no: int
    taksit_sayisi: int
    alim_toplami: Optional[float] = None
    satici: Optional[str] = None


TaksitPlanBody.model_rebuild()


class TaksitDonusumBody(BaseModel):
    kart_id: str
    kesim_tarihi: str            # ekstrenin kesim tarihi (taksit takvimi buradan kurulur)
    dilimler: List[TaksitDonusumIslem]
    uygula: bool = False         # ⚠️ VARSAYILAN KURU ÇALIŞTIRMA


@app.post("/api/kartlar/taksit-donustur")
def kart_taksit_donustur(body: TaksitDonusumBody):
    """📅 ADIM 8 — İçe aktarılmış TEK ÇEKİM satırını TAKSİTLİ alıma çevirir.

    NEDEN GEREKLİ: ekstre "ESER TİCARET 3/4 · 41.250 ₺ (toplam 165.000)" diyor
    ama içe aktarım bunu TEK ÇEKİM 41.250 olarak yazıyor. Sonuç: bu dönemin
    borcu doğru, ama GELECEK TAKSİTLER sistemde yok. Canlı kanıt: bu kartta
    taksitli hareket sayısı SIFIR, oysa ekstre 5 taksitli alım gösteriyor ve
    87.890,59 ₺ gelecek yük taşıyor.

    ⚠️ NEDEN YENİ KAYIT DEĞİL, DÖNÜŞTÜRME: satır zaten deftere yazıldı. Üstüne
    bir de taksitli alım eklemek aynı alımı İKİ KEZ borçlandırırdı.

    ⚠️ NEDEN VARSAYILAN KURU ÇALIŞTIRMA (uygula=False): borç motoru taksiti
    KÜMÜLATİF sayar — `tutar/taksit_sayisi × geçen taksit adedi`. ESER'i
    165.000/4 taksit + başlangıç Haziran diye yazarsam Ağustos'ta ÜÇ taksit
    geçmiş sayılır (123.750 ₺) ve şu an KURUŞU KURUŞUNA tutan mutabakat
    (fark 0,00) bozulur. Geçmiş iki taksit önceki ekstrelerde zaten borç
    yazılmış olabilir. Bu yüzden uç önce NE OLACAĞINI hesaplar, borç etkisini
    gösterir; sahip görüp onaylamadan HİÇBİR ŞEY YAZILMAZ.

    Taksit takvimi: bu ekstre `taksit_no`'yu biller → başlangıç ayı =
    kesim ayı − (taksit_no − 1).
    """
    from datetime import date as _d
    try:
        _kesim = _d.fromisoformat(str(body.kesim_tarihi)[:10])
    except Exception:
        raise HTTPException(400, "kesim_tarihi YYYY-AA-GG olmalı")

    with db() as (conn, cur):
        cur.execute("SELECT id, kart_adi FROM kartlar WHERE id=%s AND aktif=TRUE", (body.kart_id,))
        _k = cur.fetchone()
        if not _k:
            raise HTTPException(404, "Kart bulunamadı")
        _borc_once = kart_borc(cur, body.kart_id)

        plan, atlanan = [], []
        for d in body.dilimler:
            no, adet = int(d.taksit_no or 0), int(d.taksit_sayisi or 0)
            if not (1 <= no <= adet <= 60):
                atlanan.append({"satici": d.satici, "neden": f"geçersiz taksit {no}/{adet}"})
                continue
            toplam = float(d.alim_toplami or 0) or round(float(d.dilim_tutari) * adet, 2)
            # Başlangıç ayı: bu ekstre no'ıncı taksidi billiyor
            _ay = _kesim.month - (no - 1)
            _yil = _kesim.year
            while _ay <= 0:
                _ay += 12; _yil -= 1
            _bas = _d(_yil, _ay, min(_kesim.day, 28))
            # Deftere yazılmış TEK ÇEKİM satırını bul (bu dönemin dilimi)
            cur.execute("""SELECT id, tutar, taksit_sayisi FROM kart_hareketleri
                            WHERE kart_id=%s AND durum='aktif' AND islem_turu='HARCAMA'
                              AND COALESCE(taksit_sayisi,1)=1
                              AND ROUND(tutar::numeric,2)=ROUND(%s::numeric,2)
                              AND tarih=%s::date LIMIT 1""",
                        (body.kart_id, float(d.dilim_tutari), str(d.tarih)[:10]))
            _r = cur.fetchone()
            if not _r:
                atlanan.append({"satici": d.satici,
                                "neden": "defterde eşleşen tek-çekim satırı yok "
                                         "(zaten dönüştürülmüş ya da hiç aktarılmamış)"})
                continue
            # Borç etkisi: motor kümülatif sayar → geçen taksit adedi = no
            _yeni_katki = round(toplam / adet * no, 2)
            plan.append({
                "hareket_id": _r["id"], "satici": d.satici,
                "eski_tutar": float(_r["tutar"]), "eski_katki": float(_r["tutar"]),
                "yeni_tutar": toplam, "taksit": f"{no}/{adet}",
                "baslangic": str(_bas), "yeni_katki": _yeni_katki,
                "borc_degisimi": round(_yeni_katki - float(_r["tutar"]), 2),
                "gelecek_yuk": round(toplam / adet * (adet - no), 2),
            })

        _fark = round(sum(p["borc_degisimi"] for p in plan), 2)
        if not body.uygula:
            return {
                "kuru_calistirma": True, "kart": _k["kart_adi"],
                "donusecek": len(plan), "atlanan": atlanan,
                "plan": plan,
                "borc_once": round(_borc_once, 2),
                "borc_sonra_tahmini": round(_borc_once + _fark, 2),
                "borc_degisimi": _fark,
                "gelecek_yuk_kazanimi": round(sum(p["gelecek_yuk"] for p in plan), 2),
                "uyari": ("⚠️ Bu dönüşüm kart borcunu DEĞİŞTİRİYOR. Motor taksiti kümülatif "
                          "sayar; geçmiş taksitler önceki ekstrelerde zaten borç yazılmış "
                          "olabilir — o zaman bu değişim ÇİFT SAYIM olur. Uygulamadan önce "
                          "farkı doğrulayın." if abs(_fark) > 1.0 else
                          "Borç değişmiyor — dönüşüm güvenli görünüyor."),
                "not": "Hiçbir şey yazılmadı. Uygulamak için uygula=true gönderin.",
            }

        for p in plan:
            cur.execute("""UPDATE kart_hareketleri
                              SET tutar=%s, taksit_sayisi=%s, baslangic_tarihi=%s::date,
                                  aciklama = COALESCE(aciklama,'') || %s
                            WHERE id=%s""",
                        (p["yeni_tutar"], int(p["taksit"].split("/")[1]), p["baslangic"],
                         f" [taksit {p['taksit']} — ekstreden dönüştürüldü]", p["hareket_id"]))
            audit(cur, "kart_hareketleri", p["hareket_id"], "TAKSIT_DONUSUM")
        kart_plan_guncelle_tx(cur)
        _borc_sonra = kart_borc(cur, body.kart_id)
    return {"kuru_calistirma": False, "donusturulen": len(plan), "atlanan": atlanan,
            "borc_once": round(_borc_once, 2), "borc_sonra": round(_borc_sonra, 2),
            "borc_degisimi": round(_borc_sonra - _borc_once, 2), "plan": plan}


@app.get("/api/kasa/finansman-dengesi")
def finansman_dengesi():
    """💠 KİM KİMİ FİNANSE EDİYOR — defterden TÜRETİLİR, borç olarak YAZILMAZ.

    🔴 SAHİP: "Alsancak kredisi bazen kasada para yoksa Gazze'den çekilir."
    Yani ödemeyi yapan şube, giderin sahibi olmayabilir. Aradaki fark, bir
    şubenin diğerini finanse ettiği anlamına gelir.

    ⚠️ NEDEN "BORÇ" DEĞİL (Codex): "varsayılan olarak TÜRETİLMİŞ net pozisyon
    olmalı, kayıtlı borç değil. Aksi halde mevcut sube_ici_borc ile kavga eder,
    SAHTE ALACAK/BORÇ ŞİŞMESİ üretir." Sahip de zaten şubeler arasında gerçek
    bir geri ödeme yapmıyor — "ortak kasa gibi düşün" diyor.
    Semantik sert ayrılır ve biri diğerini ASLA otomatik üretmez:
        /api/sube-ici-borc      → GERÇEK, kapanabilir borç (elle kaydedilir)
        bu görünüm              → ANALİTİK denge (defterden çıkar, kapanmaz)

    ⚠️ BİLİNMEYEN GİZLENMEZ: ödeyeni işaretlenmemiş hareketler ayrı sayılır.
    Geçmiş 28 kredi + 21 kart ödemesine geriye dönük ödeyen ATANMADI (Codex S5:
    "kanıt yoksa bilinmiyor bırak; yanlış atama dengeyi kirletir ve DOĞRUYMUŞ
    GİBİ konuşur"). Bu yüzden denge şimdilik yalnız işaretli hareketleri sayar
    ve kapsama oranını açıkça söyler.
    """
    with db() as (conn, cur):
        cur.execute("""
            SELECT COALESCE(kh.odeyen_sube_id, '') AS odeyen,
                   COALESCE(kh.maliyet_merkezi_id, '') AS maliyet,
                   COALESCE(kh.maliyet_merkezi_tipi, '') AS mtip,
                   COUNT(*) AS adet,
                   COALESCE(SUM(ABS(kh.tutar)), 0) AS tutar
              FROM kasa_hareketleri kh
             WHERE COALESCE(kh.durum,'aktif') = 'aktif'
               AND COALESCE(kh.kasa_etkisi, TRUE) = TRUE
               AND kh.tutar < 0
               AND kh.odeyen_sube_id IS NOT NULL
               AND kh.maliyet_merkezi_id IS NOT NULL
               AND kh.odeyen_sube_id <> kh.maliyet_merkezi_id
             GROUP BY 1,2,3
        """)
        capraz = [dict(r) for r in cur.fetchall()]
        cur.execute("""
            SELECT COUNT(*) FILTER (WHERE odeyen_sube_id IS NULL) AS odeyensiz,
                   COUNT(*) AS toplam,
                   COALESCE(SUM(ABS(tutar)) FILTER (WHERE odeyen_sube_id IS NULL), 0) AS odeyensiz_tutar
              FROM kasa_hareketleri
             WHERE COALESCE(durum,'aktif')='aktif' AND COALESCE(kasa_etkisi,TRUE)=TRUE
               AND tutar < 0
        """)
        k = dict(cur.fetchone() or {})
        cur.execute("SELECT id::text AS id, ad FROM subeler")
        ad = {r["id"]: r["ad"] for r in cur.fetchall()}

    net: Dict[str, float] = {}
    for c in capraz:
        o, m = c["odeyen"], c["maliyet"]
        net[o] = round(net.get(o, 0.0) + float(c["tutar"]), 2)     # finanse EDEN
        net[m] = round(net.get(m, 0.0) - float(c["tutar"]), 2)     # finanse EDİLEN
    satir = [{"sube_id": s, "ad": ad.get(s, s),
              "net": v, "rol": "finanse ediyor" if v > 0 else "finanse ediliyor"}
             for s, v in sorted(net.items(), key=lambda kv: -kv[1]) if abs(v) > 0.01]

    top = int(k.get("toplam") or 0)
    ody = int(k.get("odeyensiz") or 0)
    return {
        "capraz_odemeler": [
            {"odeyen_ad": ad.get(c["odeyen"], c["odeyen"]),
             "maliyet_ad": ad.get(c["maliyet"], c["maliyet"]),
             "maliyet_tipi": c["mtip"], "adet": int(c["adet"]),
             "tutar": round(float(c["tutar"]), 2)}
            for c in sorted(capraz, key=lambda x: -float(x["tutar"]))
        ],
        "net_denge": satir,
        "kapsama": {
            "cikis_hareketi": top,
            "odeyeni_bilinmeyen": ody,
            "odeyeni_bilinmeyen_tutar": round(float(k.get("odeyensiz_tutar") or 0), 2),
            "kapsama_yuzde": round(100.0 * (top - ody) / top, 1) if top else 0.0,
        },
        "not": "Bu bir BORÇ DEĞİL, analitik finansman dengesidir — defterden türer, "
               "kapanmaz, kimse kimseye geri ödeme yapmaz. Gerçek şube→şube borç için "
               "/api/sube-ici-borc ayrıdır ve biri diğerini otomatik ÜRETMEZ. "
               "Ödeyeni işaretlenmemiş hareketler bu dengeye GİRMEZ ve yukarıda sayılır.",
    }


@app.get("/api/kasa/sube-sessizlik")
def sube_sessizlik(esik_gun: int = 3):
    """🔇 SESSİZ ŞUBE DUYUSU — gider üretiyor ama ciro girmiyor.

    🔴 CANLI VAKA (2026-08-18, bu duyu tam bunu bulduktan sonra yazıldı):
    KÖYCEĞİZ ve ALSANCAK'ın son ciro kaydı **19 Haziran 2026**. 60 gündür
    ciro girilmemiş. AMA giderleri işlemeye devam ediyor:
        3 Ağu  KÖYCEĞİZ KİRA        43.500 ₺
        4 Ağu  POS DONANIM ÜCRETİ      949 ₺ (Köyceğiz) + 950 ₺ (Alsancak)
    POS donanım ücreti ödeniyorsa cihaz aktiftir — yani şube büyük olasılıkla
    satış yapıyor, sadece CİROSU SİSTEME GİRMİYOR.

    Neden kimse fark etmedi: hiçbir ekran "bir şeyin OLMAMASINI" göstermiyordu.
    Eksik ciro sessizdir — ekranda boşluk bırakmaz, alarm üretmez, sadece o
    şubeyi yapay olarak zarar ediyor gösterir. Kasa çekmecesindeki imkânsız
    −17.790 ₺ bunun tek görünür belirtisiydi ve onu da ancak defteri uçtan uca
    okuyunca gördük.

    Bu duyu o boşluğu KONUŞUR HALE getirir: gideri süren ama cirosu susan şube
    ya kapanmıştır (o zaman giderleri de durmalı) ya da veri girişi kopmuştur.
    İkisi de sahibin bilmesi gereken şeydir.

    🔴 SAHİP DÜZELTMESİ (2026-08-18, alarm kurulduktan HEMEN sonra):
    "Alsancak ve Köyceğiz ÖĞRENCİ BÖLGELERİNDE ve bundan dolayı ciro girmiyor,
    yani SEZONLUK."
    Yani bu iki şubede ciro yokluğu ARIZA DEĞİL, NORMAL. Alarm olduğu gibi
    bırakılsaydı her yaz aylarca eyleme dönüşmeyen kırmızı yakardı — sahibin
    uyarı bütçesi kuralının ("sürekli çıkan, eyleme dönüşmeyen uyarı YASAK")
    tam ihlali. Duyu bu yüzden şubenin FAALİYET DURUMUnu tanır:
        acik       → sessiz + giderli = KRİTİK (gerçek alarm)
        sezon_disi → alarm YOK; onun yerine SEZON MALİYETİ raporlanır
    Sezon dışı şubede asıl bilgi "ciro yok" değil, "kapalı dururken ne kadara
    mal oluyor" — kira işlemeye devam eder ve bu gerçek bir karardır.
    ⛔ ÖNERİ-ONLY: hüküm vermez, hiçbir şey yazmaz.
    """
    with db() as (conn, cur):
        # Lazy migration — faaliyet durumu (acik | sezon_disi | kapali)
        try:
            cur.execute("SAVEPOINT sp_faaliyet")
            cur.execute("""ALTER TABLE subeler ADD COLUMN IF NOT EXISTS
                           faaliyet_durumu TEXT DEFAULT 'acik'""")
            cur.execute("""ALTER TABLE subeler ADD COLUMN IF NOT EXISTS
                           faaliyet_notu TEXT""")
            cur.execute("RELEASE SAVEPOINT sp_faaliyet")
        except Exception:  # noqa: BLE001
            try:
                cur.execute("ROLLBACK TO SAVEPOINT sp_faaliyet")
                cur.execute("RELEASE SAVEPOINT sp_faaliyet")
            except Exception:  # noqa: BLE001
                pass
        cur.execute("""
            SELECT s.id::text AS sube_id, s.ad,
                   COALESCE(s.faaliyet_durumu, 'acik') AS faaliyet_durumu,
                   s.faaliyet_notu,
                   (SELECT MAX(c.tarih) FROM ciro c WHERE c.sube_id::text = s.id::text) AS son_ciro,
                   (SELECT COUNT(*) FROM ciro c WHERE c.sube_id::text = s.id::text) AS ciro_adet
              FROM subeler s
             WHERE COALESCE(s.aktif, TRUE) = TRUE
               AND UPPER(COALESCE(s.ad,'')) <> 'MERKEZ'
             ORDER BY s.ad
        """)
        subeler = [dict(r) for r in cur.fetchall()]
        out = []
        for s in subeler:
            son = s.get("son_ciro")
            # Ciro sustuktan SONRA gider üretilmiş mi? Asıl sinyal bu:
            # ciro yok + gider yok = kapanmış olabilir (sakin)
            # ciro yok + gider VAR = ya kayıt kopuk ya kapanış yarım (alarm)
            cur.execute("""
                SELECT COUNT(*) AS adet, COALESCE(SUM(ABS(tutar)),0) AS toplam
                  FROM kasa_hareketleri
                 WHERE sube_id::text = %s AND tutar < 0
                   AND COALESCE(durum,'aktif') = 'aktif'
                   AND COALESCE(kasa_etkisi, TRUE) = TRUE
                   AND (%s::date IS NULL OR tarih > %s::date)
            """, (s["sube_id"], son, son))
            g = dict(cur.fetchone() or {})
            gun = None
            if son:
                cur.execute("SELECT (CURRENT_DATE - %s::date) AS g", (son,))
                gun = int((cur.fetchone() or {}).get("g") or 0)
            sessiz = (gun is None) or (gun > esik_gun)
            giderli = int(g.get("adet") or 0) > 0
            faaliyet = (s.get("faaliyet_durumu") or "acik").strip().lower()
            gtut = round(float(g.get("toplam") or 0), 2)

            if faaliyet == "sezon_disi":
                # Sezon dışı şubede ciro yokluğu BEKLENEN durumdur → ALARM YOK.
                # Buradaki değerli bilgi "kapalı dururken ne kadara mal oluyor".
                durum = "sezon dışı"
                yorum = (f"sezon dışı · {gun} gündür kapalı · bu sürede {gtut:,.2f} ₺ "
                         f"sabit gider işledi (kapalı taşıma maliyeti)"
                         if gun is not None else "sezon dışı")
            elif faaliyet == "kapali":
                durum = "kapalı"
                yorum = (f"kapalı · bu sürede {gtut:,.2f} ₺ gider işledi — kapalı şubenin "
                         "gideri sürüyorsa sebebi olmalı (kira sözleşmesi vb.)"
                         if giderli else "kapalı, gideri de yok")
            elif sessiz and giderli:
                durum = "KRİTİK"
                yorum = (f"{gun} gündür ciro girilmemiş ama bu sürede {int(g.get('adet') or 0)} "
                         f"gider işlemiş ({gtut:,.2f} ₺) — şube açıksa ciro kaydı KOPMUŞ. "
                         "Sezonluksa şubeyi «sezon dışı» işaretleyin, bu uyarı susar.")
            elif sessiz:
                durum = "sessiz"
                yorum = f"{gun} gündür ciro yok, gideri de yok — kapanmış olabilir"
            else:
                durum = "akıyor"
                yorum = "ciro akıyor"

            out.append({
                "sube_id": s["sube_id"], "ad": s["ad"],
                "faaliyet_durumu": faaliyet,
                "faaliyet_notu": s.get("faaliyet_notu"),
                "son_ciro": str(son) if son else None,
                "ciro_adet": int(s.get("ciro_adet") or 0),
                "sessiz_gun": gun,
                "sonrasinda_gider_adet": int(g.get("adet") or 0),
                "sonrasinda_gider_tutar": gtut,
                "durum": durum,
                "yorum": yorum,
            })
    kritik = [o for o in out if o["durum"] == "KRİTİK"]
    sezon = [o for o in out if o["durum"] == "sezon dışı"]
    return {
        "esik_gun": esik_gun,
        "subeler": out,
        "kritik_adet": len(kritik),
        "kritik_toplam_gider": round(sum(o["sonrasinda_gider_tutar"] for o in kritik), 2),
        # Sezon dışı şubelerin KAPALI TAŞIMA MALİYETİ — alarm değil, karar verisi.
        # "Bu şubeyi sezon boyunca kapalı tutmak bana kaça mal oluyor" sorusunun
        # cevabı; kira sözleşmesi/devam kararı bunun üstünden verilir.
        "sezon_disi_adet": len(sezon),
        "sezon_disi_tasima_maliyeti": round(sum(o["sonrasinda_gider_tutar"] for o in sezon), 2),
        "not": "Öneri-only duyu: eksik ciro SESSİZDİR — ekranda boşluk bırakmaz, o şubeyi "
               "yapay zararda gösterir. SEZON DIŞI işaretli şube alarm ÜRETMEZ; onun yerine "
               "kapalı taşıma maliyeti raporlanır. Hüküm vermez, hiçbir kayıt yazmaz.",
    }


class SubeFaaliyetBody(BaseModel):
    faaliyet_durumu: str          # 'acik' | 'sezon_disi' | 'kapali'
    faaliyet_notu: Optional[str] = None


@app.post("/api/subeler/{sid}/faaliyet-durumu")
def sube_faaliyet_durumu(sid: str, b: SubeFaaliyetBody):
    """Şubenin faaliyet durumunu işaretle — sessizlik duyusunun yalancı alarm
    üretmemesi için (2026-08-18, sahip: "öğrenci bölgeleri, yani sezonluk")."""
    d = (b.faaliyet_durumu or "").strip().lower()
    if d not in ("acik", "sezon_disi", "kapali"):
        raise HTTPException(400, "faaliyet_durumu: acik | sezon_disi | kapali olmalı")
    with db() as (conn, cur):
        cur.execute("""ALTER TABLE subeler ADD COLUMN IF NOT EXISTS faaliyet_durumu TEXT DEFAULT 'acik'""")
        cur.execute("""ALTER TABLE subeler ADD COLUMN IF NOT EXISTS faaliyet_notu TEXT""")
        cur.execute("SELECT ad FROM subeler WHERE id::text=%s", (sid,))
        r = cur.fetchone()
        if not r:
            raise HTTPException(404, "Şube bulunamadı")
        cur.execute("""UPDATE subeler SET faaliyet_durumu=%s, faaliyet_notu=%s
                       WHERE id::text=%s""", (d, (b.faaliyet_notu or "").strip() or None, sid))
        audit(cur, "subeler", sid, "UPDATE")
    return {"success": True, "sube": dict(r).get("ad"), "faaliyet_durumu": d}


@app.get("/api/kasa-defteri")
def kasa_defteri(
    tarih_baslangic: Optional[str] = None,
    tarih_bitis: Optional[str] = None,
    islem_turu: Optional[str] = None,
    sube_id: Optional[str] = None,
    sadece_etkisiz: bool = False,
    sayfa: int = 1,
    boyut: int = 500,
):
    """📖 KASA DEFTERİ — TAM, SAYFALANABİLİR, SALT-OKUR.

    🔴 NEDEN AÇILDI (2026-08-17, sahip: "bu kasayı bugünün hareketleriyle
    beraber baştan sona detaylı incelendin mi?"):
    Kasayı UÇTAN UCA okumanın yolu YOKTU. Mevcut uçlar:
      /kasa              → sabit LIMIT 100, sayfalama YOK
      /kasa-detay        → yalnız islem_turu ÖZETİ (satır yok)
      /kasa/sube-bazli   → yalnız şube ÖZETİ (satır yok)
      /dis-kaynak        → tek türü, LIMIT 200
    Yani ~570 hareketlik defterin son 100'ü dışına kimse bakamıyordu. Bir
    bulguyu "özet rakamdan" doğrulamak, satırı görmeden hüküm vermektir; bu
    denetimde tam o hataya düştüm (67 hareketlik kovanın 18'ini görüp
    genelleme yaptım). Bu uç o gözü açar.

    DOKTRİN:
      · SALT-OKUR — hiçbir şey yazmaz, hiçbir şey düzeltmez
      · SESSİZ ELEME YASAK — `toplam` filtreye uyan TÜM satır sayısıdır;
        sayfa küçükse kaç satır dışarıda kaldığı `kalan` ile görünür
      · HATA ≠ BOŞ — geçersiz tarih 400 döner, sessizce "hepsi" göstermez
      · kasa_etkisi=FALSE satırlar da DÖNER (gizlenmez) ama toplamlarda
        AYRI sayılır: defterde durur, bakiyeyi oynatmaz — ikisi karışırsa
        "kasa tutmuyor" gibi sahte alarm doğar
    """
    import re as _re
    for _ad, _v in (("tarih_baslangic", tarih_baslangic), ("tarih_bitis", tarih_bitis)):
        if _v and not _re.match(r"^\d{4}-\d{2}-\d{2}$", _v.strip()):
            raise HTTPException(400, f"{_ad} YYYY-AA-GG biçiminde olmalı")
    sayfa = max(1, int(sayfa or 1))
    boyut = max(1, min(2000, int(boyut or 500)))

    kosul = ["durum='aktif'"]
    par: list = []
    if tarih_baslangic:
        kosul.append("tarih >= %s::date"); par.append(tarih_baslangic.strip())
    if tarih_bitis:
        kosul.append("tarih <= %s::date"); par.append(tarih_bitis.strip())
    if islem_turu:
        # Virgülle çoklu tür: 'ANLIK_GIDER,BORC_TAKSIT'
        turler = [t.strip().upper() for t in islem_turu.split(",") if t.strip()]
        if turler:
            kosul.append("UPPER(islem_turu) = ANY(%s)"); par.append(turler)
    if sube_id:
        s = sube_id.strip()
        if s.upper() in ("(MERKEZ)", "ATANMAMIS", "ATANMAMIŞ", "NULL", "-"):
            kosul.append("sube_id IS NULL")
        else:
            kosul.append("sube_id::text = %s"); par.append(s)
    if sadece_etkisiz:
        kosul.append("COALESCE(kasa_etkisi, TRUE) = FALSE")
    nere = " AND ".join(kosul)

    with db() as (conn, cur):
        cur.execute(f"""
            SELECT COUNT(*) AS n,
                   COALESCE(SUM(CASE WHEN COALESCE(kasa_etkisi,TRUE) AND tutar > 0
                                     THEN tutar ELSE 0 END), 0) AS giris,
                   COALESCE(SUM(CASE WHEN COALESCE(kasa_etkisi,TRUE) AND tutar < 0
                                     THEN ABS(tutar) ELSE 0 END), 0) AS cikis,
                   COUNT(*) FILTER (WHERE COALESCE(kasa_etkisi,TRUE) = FALSE) AS etkisiz_adet,
                   COALESCE(SUM(CASE WHEN COALESCE(kasa_etkisi,TRUE) = FALSE
                                     THEN tutar ELSE 0 END), 0) AS etkisiz_toplam,
                   COUNT(*) FILTER (WHERE sube_id IS NULL) AS subesiz_adet,
                   COALESCE(SUM(CASE WHEN sube_id IS NULL AND COALESCE(kasa_etkisi,TRUE)
                                     THEN tutar ELSE 0 END), 0) AS subesiz_net,
                   -- 💵 FİZİKSEL NAKİT (2026-08-18): `tutar` para POZİSYONUdur
                   -- (kart cirosu + banka ödemeleri dahil); `nakit_etki` ise
                   -- çekmecede gerçekten ne olduğudur. NULL = BİLİNMİYOR ve
                   -- ayrıca sayılır — 0 sayılsaydı "elde şu kadar nakit var"
                   -- diye SESSİZ YANLIŞ üretirdik.
                   COALESCE(SUM(CASE WHEN COALESCE(kasa_etkisi,TRUE)
                                     THEN nakit_etki ELSE 0 END), 0) AS nakit_net,
                   COUNT(*) FILTER (WHERE nakit_etki IS NULL) AS nakit_bilinmeyen_adet,
                   COALESCE(SUM(CASE WHEN nakit_etki IS NULL AND COALESCE(kasa_etkisi,TRUE)
                                     THEN ABS(tutar) ELSE 0 END), 0) AS nakit_bilinmeyen_tutar
            FROM kasa_hareketleri WHERE {nere}
        """, par)
        o = dict(cur.fetchone() or {})
        toplam = int(o.get("n") or 0)

        cur.execute(f"""
            SELECT id::text, tarih, islem_turu, tutar, aciklama, kaynak_tablo,
                   kaynak_id::text, ref_type, kasa_etkisi, sube_id::text,
                   odeme_yontemi, olusturma
            FROM kasa_hareketleri WHERE {nere}
            ORDER BY tarih DESC, olusturma DESC
            LIMIT %s OFFSET %s
        """, par + [boyut, (sayfa - 1) * boyut])
        satirlar = [dict(r) for r in cur.fetchall()]

        # Kırılımlar — filtrelenmiş küme için (özet ile satır AYNI kümeden gelsin)
        cur.execute(f"""
            SELECT islem_turu,
                   COUNT(*) AS adet,
                   COALESCE(SUM(CASE WHEN tutar > 0 THEN tutar ELSE 0 END), 0) AS giris,
                   COALESCE(SUM(CASE WHEN tutar < 0 THEN ABS(tutar) ELSE 0 END), 0) AS cikis,
                   COUNT(*) FILTER (WHERE sube_id IS NULL) AS subesiz
            FROM kasa_hareketleri WHERE {nere}
            GROUP BY islem_turu ORDER BY 4 DESC, 3 DESC
        """, par)
        tur_kirilim = [dict(r) for r in cur.fetchall()]

        cur.execute(f"""
            SELECT COALESCE(sube_id::text, '(atanmamış)') AS sube,
                   COUNT(*) AS adet,
                   COALESCE(SUM(CASE WHEN tutar > 0 THEN tutar ELSE 0 END), 0) AS giris,
                   COALESCE(SUM(CASE WHEN tutar < 0 THEN ABS(tutar) ELSE 0 END), 0) AS cikis,
                   -- Şube bazında ÇEKMECE (fiziksel nakit) — pozisyondan ayrı.
                   -- Kira şubenin bankasından çıkar: pozisyonu düşürür ama
                   -- çekmeceyi boşaltmaz. Bu sütun onu ayırır.
                   COALESCE(SUM(nakit_etki), 0) AS nakit_net,
                   COUNT(*) FILTER (WHERE nakit_etki IS NULL) AS nakit_bilinmeyen
            FROM kasa_hareketleri WHERE {nere}
            GROUP BY 1 ORDER BY 2 DESC
        """, par)
        sube_kirilim = [dict(r) for r in cur.fetchall()]

    giris, cikis = float(o.get("giris") or 0), float(o.get("cikis") or 0)
    return {
        "filtre": {
            "tarih_baslangic": tarih_baslangic, "tarih_bitis": tarih_bitis,
            "islem_turu": islem_turu, "sube_id": sube_id,
            "sadece_etkisiz": sadece_etkisiz, "sayfa": sayfa, "boyut": boyut,
        },
        "toplam": toplam,
        "donen": len(satirlar),
        "kalan": max(0, toplam - (sayfa - 1) * boyut - len(satirlar)),
        "ozet": {
            "giris": round(giris, 2),
            "cikis": round(cikis, 2),
            "net": round(giris - cikis, 2),
            # Defterde duran ama bakiyeyi OYNATMAYAN satırlar (DEVIR, ODEME_PLANI…)
            "kasa_etkisiz_adet": int(o.get("etkisiz_adet") or 0),
            "kasa_etkisiz_toplam": round(float(o.get("etkisiz_toplam") or 0), 2),
            # Şubesi çözülmemiş hareketler — "hangi kasadan çıktı" cevapsız
            "subesiz_adet": int(o.get("subesiz_adet") or 0),
            "subesiz_net": round(float(o.get("subesiz_net") or 0), 2),
            # 💵 ÇEKMECE (fiziksel nakit) — `net`ten AYRI kavram:
            #   net        = para POZİSYONU (kart cirosu + banka ödemeleri dahil)
            #   nakit_net  = çekmecede gerçekten olan
            # Aradaki fark modelin kendisidir, hata değildir.
            "nakit_net": round(float(o.get("nakit_net") or 0), 2),
            "nakit_bilinmeyen_adet": int(o.get("nakit_bilinmeyen_adet") or 0),
            "nakit_bilinmeyen_tutar": round(float(o.get("nakit_bilinmeyen_tutar") or 0), 2),
        },
        "tur_kirilim": tur_kirilim,
        "sube_kirilim": sube_kirilim,
        "hareketler": satirlar,
        "not": "Salt-okur kasa defteri. `net` yalnız kasa_etkisi=TRUE satırları "
               "sayar; etkisiz satırlar ayrı raporlanır (gizlenmez). Filtresiz "
               "çağrıda `net` = güncel kasa bakiyesi olmalıdır.",
    }


@app.post("/api/personel/sube-turet")
def personel_sube_turet(kuru: int = 1):
    """👤 ŞUBESİ TANIMSIZ PERSONELİN ŞUBESİNİ İZİNDEN TÜRET.

    4 personelin `sube_id`'si boş; maaşları kasada hep "atanmamış" kovasına
    düşüyor ve şube kârlılığından eksik kalıyor. Şube elle girilmemişse bile
    personelin İZİ var: yoklamada asıl şubesi, görev tamamlamada çalıştığı
    şube. Bu uç izleri sayar ve TEK şube baskınsa önerir/atar.

    Kanıt gücü sırası:
      1. gorev_yoklama.asil_sube_id — personelin kayıtlı asıl şubesi (en güçlü)
      2. gorev_yoklama.sube_id      — fiilen çalıştığı şube (misafir olabilir)
      3. gorev_tamamlama.sube_id    — görev yaptığı şube

    ⚠️ Birden çok şube çıkarsa ve baskın olan yoksa ATAMA YAPILMAZ — yanlış
    şubeye maaş yazmaktansa boş kalsın (öneri-only ilkesi).
    """
    _KAYNAKLAR = [
        ("gorev_yoklama", "asil_sube_id", 3, "yoklamada asıl şube"),
        ("gorev_yoklama", "sube_id", 2, "yoklamada çalıştığı şube"),
        ("gorev_tamamlama", "sube_id", 1, "görev tamamlama"),
    ]
    sonuc, atanan = [], 0
    with db() as (conn, cur):
        cur.execute("""SELECT id::text AS id, ad_soyad, aktif
                         FROM personel
                        WHERE COALESCE(TRIM(sube_id),'') = ''""")
        hedefler = [dict(r) for r in (cur.fetchall() or [])]
        cur.execute("SELECT id::text AS id, ad FROM subeler")
        ad_map = {r["id"]: r["ad"] for r in (cur.fetchall() or [])}
        for p in hedefler:
            skor: Dict[str, float] = {}
            kanit = []
            for tablo, kol, agirlik, etiket in _KAYNAKLAR:
                try:
                    cur.execute("SAVEPOINT sp_pt")
                    cur.execute(f"""
                        SELECT {kol}::text AS sid, COUNT(*) AS adet
                          FROM {tablo}
                         WHERE personel_id::text = %s
                           AND COALESCE(TRIM({kol}::text),'') <> ''
                         GROUP BY 1 ORDER BY 2 DESC
                    """, (p["id"],))
                    for r in (cur.fetchall() or []):
                        d = dict(r)
                        skor[d["sid"]] = skor.get(d["sid"], 0) + agirlik * int(d["adet"] or 0)
                        kanit.append({"kaynak": etiket, "sube": ad_map.get(d["sid"], d["sid"]),
                                      "adet": int(d["adet"] or 0)})
                    cur.execute("RELEASE SAVEPOINT sp_pt")
                except Exception:  # noqa: BLE001 — tablo/kolon yoksa atla
                    try:
                        cur.execute("ROLLBACK TO SAVEPOINT sp_pt"); cur.execute("RELEASE SAVEPOINT sp_pt")
                    except Exception:  # noqa: BLE001
                        pass
            sirali = sorted(skor.items(), key=lambda kv: -kv[1])
            oneri, gerekce = None, "iz bulunamadı"
            if sirali:
                _en, _puan = sirali[0]
                _ikinci = sirali[1][1] if len(sirali) > 1 else 0
                # Baskınlık şartı: en yüksek skor, ikincinin en az 2 katı olmalı
                if _puan >= 2 * max(1, _ikinci):
                    oneri, gerekce = _en, f"baskın iz (skor {_puan} / {_ikinci})"
                else:
                    gerekce = f"iki şube yakın (skor {_puan} / {_ikinci}) — sahip karar vermeli"
            if oneri and not kuru:
                cur.execute("UPDATE personel SET sube_id=%s WHERE id::text=%s AND COALESCE(TRIM(sube_id),'')=''",
                            (oneri, p["id"]))
                atanan += cur.rowcount or 0
            sonuc.append({
                "personel_id": p["id"], "ad_soyad": p["ad_soyad"], "aktif": p["aktif"],
                "onerilen_sube": ad_map.get(oneri) if oneri else None,
                "onerilen_sube_id": oneri, "gerekce": gerekce,
                "kanitlar": kanit[:6],
            })
        if not kuru:
            conn.commit()
    return {
        "kuru": bool(kuru), "hedef_personel": len(hedefler), "atanan": atanan,
        "sonuc": sonuc,
        "not": "Şube İZDEN türetilir: yoklamadaki asıl şube en güçlü kanıt. İki "
               "şube yakın çıkarsa atama YAPILMAZ — sahip karar verir.",
    }


@app.post("/api/kasa/sube-atama-denetimi")
def kasa_sube_atama_denetimi(kuru: int = 1):
    """🏪 ŞUBESİ ATANMAMIŞ KASA HAREKETLERİ — hangisi gerçekten merkezî?

    Kasa şube kırılımında 137 hareket / 1,58 M ₺ "Merkez / atanmamış" kovasında
    duruyor. İkiye ayrılır:
      · GERÇEKTEN MERKEZÎ  — kart borcu ödemesi, kredi taksiti, açılış devri,
        dış kaynak: bunların şubesi YOKTUR, kova doğrudur.
      · ATANABİLİR         — kaynak kaydında şube bilgisi var ama kasa satırına
        taşınmamış (personel maaşı → personelin şubesi, sabit gider → gider
        şubesi, anlık gider → gider şubesi …). Bunlar şube kârlılığını bozar.

    kuru=1 ölçer; kuru=0 atanabilir olanları kaynak kayıttan türetip yazar.
    """
    # kaynak_tablo → (şube kolonu, tablo) eşlemesi. Buradaki her satır
    # "kasa hareketinin şubesi ASLINDA şu kayıtta yazılı" demektir.
    _TUREME = [
        ("anlik_giderler", "anlik_giderler", "COALESCE(NULLIF(TRIM(t.sube_id),''), NULLIF(TRIM(t.sube),''))"),
        ("sabit_giderler", "sabit_giderler", "t.sube_id"),
        ("ciro", "ciro", "t.sube_id"),
        ("personel_aylik", "personel_aylik", "(SELECT p.sube_id FROM personel p WHERE p.id::text = t.personel_id::text)"),
        ("kasa_teslim", "kasa_teslim", "t.sube_id"),
    ]
    # Şubesi OLMAYAN, merkezde kalması DOĞRU olan işlem türleri.
    # ⚠️ Bu kontrol türetilebilirlikten ÖNCE gelir: kart borcu ödemesi
    # `odeme_plani` kaynaklı olsa bile şubesi yoktur.
    _MERKEZI = ("KART_ODEME", "KART_ODEME_IPTAL", "BORC_TAKSIT", "ACILIS_DEVRI",
                "DIS_KAYNAK", "DIS_KAYNAK_IPTAL", "VADELI_ODEME", "ODEME_IPTAL")
    with db() as (conn, cur):
        cur.execute("""
            SELECT COALESCE(islem_turu,'(yok)') AS islem_turu,
                   COALESCE(kaynak_tablo,'(yok)') AS kaynak_tablo,
                   COUNT(*) AS adet, COALESCE(SUM(tutar),0)::float AS tutar
              FROM kasa_hareketleri
             WHERE sube_id IS NULL
               AND COALESCE(durum,'aktif')='aktif'
               AND COALESCE(kasa_etkisi,TRUE)=TRUE
             GROUP BY 1,2 ORDER BY ABS(SUM(tutar)) DESC
        """)
        gruplar = []
        for r in (cur.fetchall() or []):
            d = dict(r)
            _tur = d["islem_turu"]
            _kt = d["kaynak_tablo"]
            _turetilebilir = any(k == _kt for k, _, _ in _TUREME)
            # MERKEZÎ önce: kart borcu `odeme_plani` kaynaklı olsa da şubesi yok.
            # PERSONEL_MAAS ayrı yoldan (açıklamadaki personel adı) atanır.
            d["sinif"] = ("merkezi" if _tur in _MERKEZI
                          else "atanabilir" if (_turetilebilir or _tur == "PERSONEL_MAAS")
                          else "belirsiz")
            d["tutar"] = round(float(d["tutar"] or 0), 2)
            gruplar.append(d)
        atanan = 0
        if not kuru:
            for kt, tablo, kol in _TUREME:
                try:
                    cur.execute("SAVEPOINT sp_sa")
                    cur.execute(f"""
                        UPDATE kasa_hareketleri kh
                           SET sube_id = s.id
                          FROM {tablo} t
                          JOIN subeler s ON s.id::text = ({kol})::text
                         WHERE kh.kaynak_tablo = %s
                           AND kh.kaynak_id::text = t.id::text
                           AND kh.sube_id IS NULL
                    """, (kt,))
                    atanan += cur.rowcount or 0
                    cur.execute("RELEASE SAVEPOINT sp_sa")
                except Exception as e:  # noqa: BLE001
                    try:
                        cur.execute("ROLLBACK TO SAVEPOINT sp_sa"); cur.execute("RELEASE SAVEPOINT sp_sa")
                    except Exception:  # noqa: BLE001
                        pass
                    logging.getLogger(__name__).warning(
                        "sube atama atlandi (%s): %s", kt, str(e)[:110])
            # 👤 MAAŞ — açıklamadaki personel adından türet.
            # Kasa kaydı `odeme_plani` kaynaklı ama planın kendi kaynağı BOŞ
            # (54/58 kayıtta NULL) — zincir kopuk. Tek sağlam iz açıklamadaki
            # ad: "Personel Maaş: MERVE KARABACAK — Haziran 2026 dönemi".
            # ⚠️ TEKİLLİK ŞART: ad birden çok personele uyuyorsa (MERVE /
            # MERVE AKTAŞ) atama YAPILMAZ — yanlış şubeye maaş yazmaktansa
            # atanmamış kalsın.
            try:
                cur.execute("SAVEPOINT sp_maas")
                # ⚠️ psycopg2: PARAMETRESİZ execute'ta '%%' escape EDİLMEZ —
                # literal '%%' olarak gider ve LIKE deseni bozulur (ilk denemede
                # atanan=0 dönmesinin sebebi buydu). Parametreli çağırıyoruz;
                # böylece %s dolu, %% doğru şekilde tek %'e iner.
                cur.execute("""
                    UPDATE kasa_hareketleri kh
                       SET sube_id = e.sube_id
                      FROM (
                        SELECT kh2.id AS hid, MIN(p.sube_id) AS sube_id
                          FROM kasa_hareketleri kh2
                          -- ⚠️ Kolon adı `ad` DEĞİL `ad_soyad` (personel tablosu).
                          -- Yanlış kolonla yazılan ilk sürüm hiç eşleşme
                          -- bulamıyordu ve sessizce 0 satır güncelliyordu.
                          JOIN personel p
                            ON UPPER(kh2.aciklama) LIKE '%%' || UPPER(TRIM(p.ad_soyad)) || '%%'
                         WHERE kh2.islem_turu = %s
                           AND kh2.sube_id IS NULL
                           AND COALESCE(p.sube_id,'') <> ''
                           -- Eşik 5 idi; canlıda "irem" (4 harf) gerçek bir
                           -- personel ve 18.998,25 ₺ maaşı eşleşmiyordu.
                           -- 4'e indirildi; yanlış eşleşmeye karşı asıl koruma
                           -- zaten aşağıdaki TEKİLLİK şartı (tek şube).
                           AND LENGTH(TRIM(p.ad_soyad)) >= 4
                         GROUP BY kh2.id
                        HAVING COUNT(DISTINCT p.sube_id) = 1
                      ) e
                     WHERE kh.id = e.hid AND kh.sube_id IS NULL
                """, ("PERSONEL_MAAS",))
                atanan += cur.rowcount or 0
                cur.execute("RELEASE SAVEPOINT sp_maas")
            except Exception as e:  # noqa: BLE001
                try:
                    cur.execute("ROLLBACK TO SAVEPOINT sp_maas")
                    cur.execute("RELEASE SAVEPOINT sp_maas")
                except Exception:  # noqa: BLE001
                    pass
                logging.getLogger(__name__).warning("maas sube atama: %s", str(e)[:110])
            conn.commit()
    _t = lambda s: round(sum(g["tutar"] for g in gruplar if g["sinif"] == s), 2)  # noqa: E731
    _a = lambda s: sum(g["adet"] for g in gruplar if g["sinif"] == s)  # noqa: E731
    return {
        "kuru": bool(kuru), "atanan_satir": atanan,
        "ozet": {
            "merkezi": {"adet": _a("merkezi"), "tutar": _t("merkezi"),
                        "ne_demek": "şubesi YOKTUR — kart borcu, kredi taksiti, "
                                    "açılış devri, dış kaynak. Kova doğru."},
            "atanabilir": {"adet": _a("atanabilir"), "tutar": _t("atanabilir"),
                           "ne_demek": "kaynak kayıtta şube var, kasa satırına "
                                       "taşınmamış — şube kârlılığını bozar."},
            "belirsiz": {"adet": _a("belirsiz"), "tutar": _t("belirsiz"),
                         "ne_demek": "kaynağı bilinmiyor ya da eşleme kuralı yok — "
                                     "sahip bakmalı."},
        },
        "gruplar": gruplar,
        "not": "kuru=1 ölçer, kuru=0 'atanabilir' olanları kaynak kayıttan türetir. "
               "'merkezi' sınıfa DOKUNULMAZ — onların şubesi gerçekten yoktur.",
    }


@app.post("/api/sube-kimlik-denetimi")
def sube_kimlik_denetimi(kuru: int = 1):
    """🏪 ŞUBE KİMLİĞİ TEK STANDARDA — 'sube' mi 'sube_id' mi, ad mı kimlik mi?

    Sistemde 50+ tablo `sube_id` kullanıyor ama `anlik_giderler` ayrıca `sube`
    taşıyor ve bir migration `sube_id = sube` diye KOPYALAMIŞ. Eğer `sube`
    şube ADI tutuyorsa `sube_id` de ad tutuyor demektir — yani "kimlik" alanı
    kimlik değil. Bugün KDV şube kırılımının ve kasa şube dolgusunun eksik
    kalmasının kökü bu.

    Bu uç ÖLÇER: her tabloda kaç satır kimlikle, kaç satır ADLA, kaçı hiç
    çözülemiyor. kuru=0 ile ad tutan satırları kanonik kimliğe çevirir.
    """
    _HEDEF = [("anlik_giderler", "sube"), ("anlik_giderler", "sube_id"),
              ("kasa_hareketleri", "sube_id"), ("sabit_giderler", "sube_id"),
              ("ciro", "sube_id"), ("siparis_talep", "sube_id")]
    rapor, duzeltilen = [], 0
    with db() as (conn, cur):
        cur.execute("SELECT id::text AS id, ad FROM subeler")
        _subeler = [dict(r) for r in (cur.fetchall() or [])]
        _ad2id = {(s["ad"] or "").strip().upper(): s["id"] for s in _subeler}
        for tablo, kol in _HEDEF:
            try:
                cur.execute("SAVEPOINT sp_sk")
                cur.execute(f"""
                    SELECT COUNT(*) FILTER (WHERE {kol} IS NULL OR TRIM({kol}::text)='') AS bos,
                           COUNT(*) FILTER (WHERE {kol}::text IN (SELECT id::text FROM subeler)) AS kimlikli,
                           COUNT(*) FILTER (WHERE UPPER(TRIM({kol}::text)) IN
                                            (SELECT UPPER(ad) FROM subeler)) AS adli,
                           COUNT(*) AS toplam
                      FROM {tablo}
                """)
                r = dict(cur.fetchone())
                cur.execute("RELEASE SAVEPOINT sp_sk")
                _coz = int(r["kimlikli"] or 0) + int(r["adli"] or 0) + int(r["bos"] or 0)
                rapor.append({
                    "tablo": tablo, "kolon": kol, "toplam": int(r["toplam"] or 0),
                    "kimlikli": int(r["kimlikli"] or 0), "adla_yazilmis": int(r["adli"] or 0),
                    "bos": int(r["bos"] or 0),
                    "cozulemeyen": max(0, int(r["toplam"] or 0) - _coz),
                })
            except Exception as e:  # noqa: BLE001 — kolon/tablo yoksa atla
                try:
                    cur.execute("ROLLBACK TO SAVEPOINT sp_sk"); cur.execute("RELEASE SAVEPOINT sp_sk")
                except Exception:  # noqa: BLE001
                    pass
                rapor.append({"tablo": tablo, "kolon": kol, "hata": str(e)[:90]})
        # ⚠️ YALNIZ KİMLİK KOLONLARI düzeltilir. `anlik_giderler.sube` bilinçli
        # olarak AD tutar (ekranlarda "MERKEZ" yazar); onu id'ye çevirmek UI'da
        # UUID gösterirdi. Ayrım net olsun: `sube` = görünen ad, `sube_id` = JOIN
        # anahtarı. Canlı ölçüm: 412 satırın 344'ü her iki kolonda da ADLA
        # yazılmış — `sube_id` üzerinden id ile JOIN yapan sorgular o satırları
        # sessizce kaçırıyordu.
        _DUZELTILEBILIR = {"sube_id"}
        if not kuru:
            for x in rapor:
                if x.get("hata") or not x.get("adla_yazilmis"):
                    continue
                if x.get("kolon") not in _DUZELTILEBILIR:
                    x["atlandi"] = "ad kolonu — bilinçli olarak ad tutar, dokunulmadı"
                    continue
                try:
                    cur.execute("SAVEPOINT sp_sd")
                    cur.execute(f"""
                        UPDATE {x['tablo']} t SET {x['kolon']} = s.id
                          FROM subeler s
                         WHERE UPPER(TRIM(t.{x['kolon']}::text)) = UPPER(s.ad)
                    """)
                    duzeltilen += cur.rowcount or 0
                    cur.execute("RELEASE SAVEPOINT sp_sd")
                except Exception as e:  # noqa: BLE001
                    try:
                        cur.execute("ROLLBACK TO SAVEPOINT sp_sd"); cur.execute("RELEASE SAVEPOINT sp_sd")
                    except Exception:  # noqa: BLE001
                        pass
                    x["duzeltme_hatasi"] = str(e)[:90]
            conn.commit()
    return {
        "kuru": bool(kuru), "subeler": [s["ad"] for s in _subeler],
        "rapor": rapor, "duzeltilen_satir": duzeltilen,
        "not": "'adla_yazilmis' > 0 ise o kolon KİMLİK değil AD tutuyor demektir; "
               "id ile yapılan JOIN'ler o satırları sessizce kaçırır. kuru=0 "
               "bunları kanonik id'ye çevirir (ad kolonu ayrıca durur).",
    }


@app.post("/api/kasa/iptal-cift-sayim-duzelt")
def kasa_iptal_cift_sayim_duzelt(kuru: int = 1):
    """🔧 İPTAL TERS KAYITLARININ ÇİFT SAYIMINI ÖLÇ/DÜZELT (2026-08-09).

    `iptal_kasa_hareketi()` orijinal hareketi `durum='iptal'` yapar — tutar
    kasa toplamından zaten çıkar. Ters kaydın DA `kasa_etkisi=TRUE` olması
    aynı düzeltmeyi iki kez uygular. CIRO_IPTAL için bu daha önce fark edilip
    kapatılmış, diğer türlerde açık kalmıştı.

    Kanıt (canlı ölçüm): 2.250 ₺'lik gider silindi → kasa +2.250 ₺ ARTTI.

    Bu uç, orijinali gerçekten iptal edilmiş ters kayıtları `kasa_etkisi=FALSE`
    yapar (SİLMEZ — audit izi kalır). kuru=1 yalnız ölçer.
    """
    _TURLER = ("ANLIK_GIDER_IPTAL", "CIRO_IPTAL", "CIRO_DUZELTME",
               "DIS_KAYNAK_IPTAL", "KART_ODEME_IPTAL", "VADELI_IPTAL", "ODEME_IPTAL")
    with db() as (conn, cur):
        # Ters kayıt "haksız" sayılır: kasa_etkisi hâlâ TRUE ve aynı kaynağın
        # orijinal hareketi zaten iptal edilmiş.
        cur.execute("""
            SELECT t.id, t.islem_turu, t.tutar::float AS tutar, t.tarih::text AS tarih,
                   COALESCE(t.aciklama,'') AS aciklama, t.kaynak_tablo, t.kaynak_id
              FROM kasa_hareketleri t
             WHERE t.islem_turu = ANY(%s)
               AND COALESCE(t.kasa_etkisi, TRUE) = TRUE
               AND COALESCE(t.durum,'aktif') = 'aktif'
               AND EXISTS (
                   SELECT 1 FROM kasa_hareketleri o
                    WHERE o.kaynak_id = t.kaynak_id
                      AND COALESCE(o.kaynak_tablo,'') = COALESCE(t.kaynak_tablo,'')
                      AND o.id <> t.id
                      AND COALESCE(o.durum,'') = 'iptal')
             ORDER BY t.tarih DESC
        """, (list(_TURLER),))
        satirlar = [dict(r) for r in (cur.fetchall() or [])]
        etki = round(sum(float(x["tutar"] or 0) for x in satirlar), 2)
        uygulandi = 0
        if not kuru and satirlar:
            cur.execute(
                "UPDATE kasa_hareketleri SET kasa_etkisi=FALSE WHERE id = ANY(%s)",
                ([x["id"] for x in satirlar],))
            uygulandi = cur.rowcount or 0
            conn.commit()
        kasa_yeni = guncel_kasa()
    return {
        "kuru": bool(kuru), "aday_satir": len(satirlar),
        "kasa_etkisi_tl": etki,
        "aciklama": ("Bu ters kayıtlar kasa toplamına HAKSIZ giriyor: orijinalleri "
                     "zaten iptal edilmiş. Pozitif toplam kasayı ŞİŞİRİR, negatif "
                     "toplam kasayı EKSİK gösterir."),
        "duzeltme_sonrasi_kasa": round(float(kasa_yeni or 0) - (etki if kuru else 0), 2),
        "uygulanan": uygulandi,
        "satirlar": satirlar[:50],
        "not": "kuru=1 ölçer, kuru=0 uygular. Kayıt SİLİNMEZ — yalnız kasa_etkisi "
               "FALSE olur, defterde iz kalır.",
    }


@app.get("/api/kasa/sube-bazli")
def kasa_sube_bazli():
    """🏪 ŞUBE KASALARI — merkez kasa, şube kasalarının TOPLAMIDIR.

    Sahip (2026-08-09): "her şubenin kasası var banka hesabı var, bu ayrım var;
    ödeme çıkışları hangi şubenin kasasından çıktığı belli olsun, sonunda da
    merkez kasada bu kasaların toplamı olsun."

    Eskiden kasa TEK havuzdu: 2,86 M ₺ görünüyordu ama hangi şubenin ne kadarı
    olduğu hiçbir yerde yoktu. Artık her hareket şubesini taşıyor (kaynak
    tablodan türetildi); şubesi çözülemeyen hareketler MERKEZ kovasında durur —
    gizlenmez, çünkü toplamın tutması için hepsi sayılmalı.

    ⚠️ Bu uç kasa bakiyesini DEĞİŞTİRMEZ, yalnız kırılımını gösterir:
       Σ şube kasaları + merkez (atanmamış) = guncel_bakiye
    """
    with db() as (conn, cur):
        toplam = guncel_kasa()
        cur.execute("""
            SELECT COALESCE(kh.sube_id,'(merkez)') AS sid,
                   COALESCE(s.ad,'Merkez / atanmamış') AS sube_adi,
                   COALESCE(SUM(CASE WHEN kh.tutar > 0 THEN kh.tutar END),0)::float AS giris,
                   COALESCE(SUM(CASE WHEN kh.tutar < 0 THEN ABS(kh.tutar) END),0)::float AS cikis,
                   COALESCE(SUM(kh.tutar),0)::float AS bakiye,
                   COUNT(*) AS hareket
              FROM kasa_hareketleri kh
              LEFT JOIN subeler s ON s.id::text = kh.sube_id::text
             WHERE COALESCE(kh.durum,'aktif')='aktif'
               AND COALESCE(kh.kasa_etkisi, TRUE) = TRUE
             GROUP BY 1,2
             ORDER BY bakiye DESC
        """)
        satirlar = [dict(r) for r in (cur.fetchall() or [])]
        # Ödeme yöntemi kırılımı — elden/havale ayrımı şube bazında da anlamlı
        try:
            cur.execute("""
                SELECT COALESCE(sube_id,'(merkez)') AS sid,
                       COALESCE(odeme_yontemi,'nakit') AS yontem,
                       COALESCE(SUM(ABS(tutar)),0)::float AS tutar
                  FROM kasa_hareketleri
                 WHERE tutar < 0 AND COALESCE(durum,'aktif')='aktif'
                   AND COALESCE(kasa_etkisi, TRUE) = TRUE
                 GROUP BY 1,2
            """)
            _y: Dict[str, Dict[str, float]] = {}
            for r in (cur.fetchall() or []):
                _y.setdefault(r["sid"], {})[r["yontem"]] = float(r["tutar"] or 0)
            for x in satirlar:
                x["cikis_yontem"] = _y.get(x["sid"], {})
        except Exception:  # noqa: BLE001 — kolon yoksa sessiz geç
            pass
    _sube_top = round(sum(x["bakiye"] for x in satirlar
                          if x["sid"] != "(merkez)"), 2)
    _merkez = round(sum(x["bakiye"] for x in satirlar
                        if x["sid"] == "(merkez)"), 2)
    _kirilim = round(_sube_top + _merkez, 2)
    return {
        "guncel_bakiye": round(float(toplam or 0), 2),
        "sube_toplami": _sube_top,
        "merkez_atanmamis": _merkez,
        "kirilim_toplami": _kirilim,
        # Tie-out: kırılım guncel_bakiye'yi TUTMALI. Tutmuyorsa kasa_etkisi=FALSE
        # ya da iptal edilmiş hareketler farkı yapıyordur — sessiz geçme, göster.
        "fark": round(float(toplam or 0) - _kirilim, 2),
        "satirlar": satirlar,
        "not": "Merkez kasa = şube kasalarının toplamı. Şubesi çözülemeyen "
               "hareketler 'Merkez / atanmamış' kovasında durur — gizlenmez, "
               "toplamın tutması için hepsi sayılır. Kırılım bakiyeyi değiştirmez.",
    }

# ── DIŞ KAYNAK GELİRİ (aile, kredi, ortak, vb.) ───────────────
class DisKaynakGelir(BaseModel):
    tarih: date
    kategori: str
    tutar: float
    aciklama: Optional[str] = None
    force: bool = False

@app.get("/api/dis-kaynak")
def dis_kaynak_listele(ay: str = None):
    import re as _re_ay
    ay_v = (ay or "").strip()
    ay_cond, ay_params = "", []
    if ay_v and ay_v.lower() != "hepsi" and _re_ay.match(r"^\d{4}-\d{2}$", ay_v):
        ay_cond = " AND to_char(tarih, 'YYYY-MM') = %s"
        ay_params = [ay_v]
    with db() as (conn, cur):
        cur.execute(f"""SELECT * FROM kasa_hareketleri
            WHERE islem_turu='DIS_KAYNAK' AND durum='aktif'{ay_cond}
            ORDER BY tarih DESC LIMIT 200""", ay_params)
        return [dict(r) for r in cur.fetchall()]

@app.post("/api/dis-kaynak")
def dis_kaynak_ekle(g: DisKaynakGelir):
    # 🔴 P2 (2026-08-13, EVV-PARA-N3): negatif/sıfır tutar reddedilmiyordu;
    # aşağıdaki abs() işareti SESSİZCE pozitife çeviriyordu → "-1000 girip geliri
    # geri alayım" diyen kullanıcı kasaya +1000 EKLİYORDU (2000 ₺ sapma).
    # Düzeltme DELETE ile yapılır; negatif gelire meşru ihtiyaç yok.
    try:
        g.tutar = round(float(g.tutar), 2)
    except (TypeError, ValueError):
        raise HTTPException(400, "Tutar sayı olmalı")
    if g.tutar <= 0:
        raise HTTPException(400, "Tutar pozitif olmalı")
    with db() as (conn, cur):
        if not g.force:
            cur.execute("""
                SELECT id FROM kasa_hareketleri WHERE islem_turu='DIS_KAYNAK' AND durum='aktif'
                AND tarih BETWEEN %s::date - INTERVAL '7 days' AND %s::date + INTERVAL '7 days'
                AND ABS(tutar - %s) < 1 AND aciklama LIKE %s
            """, (str(g.tarih), str(g.tarih), g.tutar, f"{g.kategori}%"))
            benzer = cur.fetchall()
            if benzer:
                return {"warning": True, "mesaj": f"Son 7 günde benzer kayıt var ({len(benzer)} adet). Yine de kaydetmek için force=true gönderin."}
        gid = str(uuid.uuid4())
        insert_kasa_hareketi(cur, g.tarih, 'DIS_KAYNAK', abs(g.tutar),
            f"{g.kategori}: {g.aciklama or ''}", 'dis_kaynak', gid)
        audit(cur, 'kasa_hareketleri', gid, 'DIS_KAYNAK')
    return {"id": gid, "success": True}

@app.delete("/api/dis-kaynak/{gid}")
def dis_kaynak_sil(gid: str):
    with db() as (conn, cur):
        # gid = kasa_hareketleri.id (frontend listeden alıyor)
        cur.execute("SELECT * FROM kasa_hareketleri WHERE id=%s AND islem_turu='DIS_KAYNAK'", (gid,))
        eski = cur.fetchone()
        if not eski: raise HTTPException(404, "Kayıt bulunamadı")
        # kaynak_id ile iptal et
        kaynak_id = eski['kaynak_id'] or gid
        iptal_kasa_hareketi(cur, kaynak_id, 'dis_kaynak', 'DIS_KAYNAK', 'DIS_KAYNAK_IPTAL', 'Dış kaynak iptali')
        audit(cur, 'kasa_hareketleri', gid, 'IPTAL', eski=eski)
    return {"success": True}

# ── ANLIQ GİDER (beklenmeyen giderler) ────────────────────────
class AnlikGider(BaseModel):
    tarih: date
    kategori: str
    tutar: float
    aciklama: Optional[str] = None
    sube: Optional[str] = "MERKEZ"
    # 💵 2026-08-09 (sahip): "elden ve havale diye ayrıştıralım; bazı ödemeler
    # nakit olsa bile elden ödenme ihtimali var, bu mutabakat doğru çalışmaz"
    #   elden  → kasadaki nakitten elden verildi → ELDE NAKİTİ AZALTIR
    #   havale → banka hesabından EFT            → elde nakiti etkilemez
    #   kart   → kredi kartı                     → kart borcunu büyütür
    #   nakit  → ESKİ/BELİRSİZ (elden mi havale mi seçilmemiş)
    odeme_yontemi: str = 'nakit'
    kart_id: Optional[str] = None
    kaynak_id: Optional[str] = None       # Değişken gider kaynağı (sabit_giderler.id)
    kaynak_tablo: Optional[str] = None    # 'sabit_giderler'
    tedarikci: Optional[str] = None       # V4: opsiyonel — dolarsa supplier_payment_event conf 1.0
    force: bool = False

def _sube_kanonik(cur, deger) -> Optional[str]:
    """🏪 Şube ADI ya da ID → kanonik sube_id (2026-08-09).

    Tablolar tutarsız: anlik_giderler'de kolon `sube`, sabit giderlerde
    `sube_id`; ikisi de bazen ad bazen id tutuyor (memory: 'sube ≠ sube_id'
    tuzağı). Kasa hareketine şube damgalarken tek kanonik kimliğe indiriyoruz.
    Çözülemezse None → 'Merkez / atanmamış' kovası; UYDURMA yapmaz.
    """
    v = str(deger or "").strip()
    # 'MERKEZ' bir şube DEĞİL, şube yokluğudur → kovası NULL. 'sube-merkez'
    # kimliği de aynı kapıya çıkar: subeler'de ad='MERKEZ' olan PASİF bir kayıt
    # var ve buraya ID olarak gelirse hayalet kasa yeniden doğar (2026-08-17,
    # canlıda 72 hareket / −1.840.501 ₺ bu yüzden birikmişti).
    if not v or v.upper() in ("MERKEZ", "SUBE-MERKEZ", "NONE", "NULL", "-"):
        return None
    try:
        cur.execute(
            "SELECT id::text AS id FROM subeler "
            "WHERE (id::text=%s OR UPPER(ad)=UPPER(%s)) "
            "  AND id::text <> 'sube-merkez' AND UPPER(COALESCE(ad,'')) <> 'MERKEZ' "
            "LIMIT 1", (v, v))
        r = cur.fetchone()
        return r["id"] if r else None
    except Exception:  # noqa: BLE001 — şube çözümü asla akışı kilitlemez
        return None


def _kanonik_kalan_limit(kart_id, yedek: float) -> float:
    """Ödeme guard'ları için KANONİK kalan limit (kartlar_listele — ekstre gerçeği
    + gelecek taksit + ortak limit havuzu). Kanonik okunamıyorsa eski defter
    hesabı (yedek) kullanılır — emniyet ağı, ödeme akışı asla kilitlenmez."""
    try:
        kl = kartlar_listele()
        kdata = kl if isinstance(kl, list) else (kl.get("kartlar") or [])
        for k in kdata:
            if str(k.get("id")) == str(kart_id):
                v = k.get("kalan_limit")
                return float(v) if v is not None else yedek
    except Exception:  # noqa: BLE001
        pass
    return yedek


def _kart_oneri_hesapla(tutar: float = 0) -> list:
    """Kart seçici TEK KAYNAK (2026-07-13, sahip şikayeti: 'ödeme yap alanında
    kart seçince gerçek limitler yok görünüyor'): eski seçici borcu defterden
    sıfırdan topluyordu (kart_borc) — ekstre içe aktarımı sonrası defter toplamı
    şişince kalan limit negatife düşüp HER kart 'Limit yetersiz' oluyordu.
    Artık kalan_limit/doluluk KANONİK kartlar_listele'den gelir (ekstre gerçeği +
    kesim-sonrası hareketler + gelecek taksit yükü + ortak limit havuzu dahil) —
    kart_borc_faiz_ozet ile aynı 'tek kaynak' ilkesi."""
    import calendar as _cal
    bugun = bugun_tr()
    kl = kartlar_listele()
    kdata = kl if isinstance(kl, list) else (kl.get("kartlar") or [])
    sonuc = []
    for k in kdata:
        limit = float(k.get('limit_tutar') or 0)
        kalan_limit = round(float(k.get('kalan_limit') or 0), 2)
        doluluk = float(k.get('limit_doluluk') or 0)
        faiz = float(k.get('faiz_orani') or 0)
        temel = {
            'kart_id': str(k['id']), 'kart_adi': k.get('kart_adi'), 'banka': k.get('banka'),
            'kalan_limit': kalan_limit, 'limit_tutar': limit, 'limit_doluluk': doluluk,
            'faiz_orani': faiz, 'kesim_gunu': k.get('kesim_gunu'),
            'son_odeme_gunu': k.get('son_odeme_gunu'), 'skor': 0, 'oneri': False,
        }
        if limit <= 0:
            sonuc.append({**temel, 'uygun': False,
                          'uygun_degil_neden': 'Kart limiti tanımlı değil — Kartlar ekranından limit gir'})
            continue
        if tutar > 0 and kalan_limit < tutar:
            sonuc.append({**temel, 'uygun': False, 'uygun_degil_neden': 'Limit yetersiz'})
            continue
        bugun_gun = bugun.day
        ay_sonu = _cal.monthrange(bugun.year, bugun.month)[1]
        kesim_gun = int(k.get('kesim_gunu') or 1)
        kesim_uzakligi = (kesim_gun - bugun_gun if kesim_gun >= bugun_gun
                          else (ay_sonu - bugun_gun) + kesim_gun)
        so_gun = int(k.get('son_odeme_gunu') or 10)
        son_odeme_uzakligi = (so_gun - bugun_gun if so_gun >= bugun_gun
                              else (ay_sonu - bugun_gun) + so_gun)
        temel.update({'kesim_uzakligi': kesim_uzakligi,
                      'son_odeme_uzakligi': son_odeme_uzakligi})
        if son_odeme_uzakligi <= 3:
            sonuc.append({**temel, 'uygun': False,
                          'uygun_degil_neden': f'Son ödeme {son_odeme_uzakligi} gün sonra — bu kart zaten ödenecek'})
            continue
        skor = ((kesim_uzakligi / 30.0) * 0.5 + (kalan_limit / limit) * 0.3
                - min(faiz / 5.0, 1.0) * 0.2)
        sonuc.append({**temel, 'uygun': True, 'uygun_degil_neden': None,
                      'skor': round(skor, 4)})
    uygunlar = [x for x in sonuc if x['uygun']]
    if uygunlar:
        en_iyi = max(uygunlar, key=lambda x: x['skor'])
        for x in sonuc:
            if x['kart_id'] == en_iyi['kart_id']:
                x['oneri'] = True
    sonuc.sort(key=lambda x: (-int(x['oneri']), -x['skor']))
    return sonuc


@app.get("/api/anlik-gider-kart-oneri")
def anlik_gider_kart_oneri(tutar: float = 0):
    """Anlık gider için kart önerisi — kanonik tek kaynaktan (_kart_oneri_hesapla)."""
    return _kart_oneri_hesapla(float(tutar or 0))

@app.get("/api/anlik-gider")
def anlik_gider_listele(durum: str = "aktif", include_pending: bool = False, include_summary: bool = False, ay: str = None):
    # Geriye uyum: eski include_pending=true => hepsi
    d = (durum or "aktif").strip().lower()
    if include_pending and d == "aktif":
        d = "hepsi"
    if d not in ("aktif", "onay_bekliyor", "hepsi"):
        raise HTTPException(400, "durum: aktif | onay_bekliyor | hepsi")

    # AY FİLTRESİ (YYYY-MM): verilirse o aya ait kayıtlar; 'hepsi'/boş => tüm aylar (geriye uyum)
    import re as _re_ay
    ay_v = (ay or "").strip()
    ay_cond = ""
    ay_params: list = []
    if ay_v and ay_v.lower() != "hepsi" and _re_ay.match(r"^\d{4}-\d{2}$", ay_v):
        ay_cond = " AND to_char(ag.tarih, 'YYYY-MM') = %s"
        ay_params = [ay_v]

    if d == "hepsi":
        durum_cond = "ag.durum IN ('aktif','onay_bekliyor')"
        limit = 300
    elif d == "onay_bekliyor":
        durum_cond = "ag.durum='onay_bekliyor'"
        limit = 300
    else:
        durum_cond = "ag.durum='aktif'"
        limit = 200

    with db() as (conn, cur):
        cur.execute(f"""
            SELECT ag.*, k.kart_adi, k.banka
            FROM anlik_giderler ag
            LEFT JOIN kartlar k ON k.id = ag.kart_id
            WHERE {durum_cond}{ay_cond}
            ORDER BY ag.tarih DESC, ag.olusturma DESC
            LIMIT {limit}
        """, ay_params)
        satirlar = [dict(r) for r in cur.fetchall()]

        if include_summary:
            cur.execute(
                """
                SELECT
                    COALESCE(COUNT(*), 0)::int AS adet,
                    COALESCE(SUM(tutar), 0) AS toplam
                FROM anlik_giderler
                -- 🏪 2026-08-09: eskiden `UPPER(sube) NOT IN ('','MERKEZ')` idi —
                -- yani ŞUBE ADIYLA karşılaştırıyordu. sube/sube_id kimliğe
                -- normalize edilince ('sube-merkez') bu karşılaştırma sessizce
                -- bozulur ve merkez giderleri "şube bekleyen" sayılırdı.
                -- Artık KİMLİKLE karşılaştırılıyor; ad değişse bile tutar.
                WHERE durum='onay_bekliyor'
                  AND COALESCE(NULLIF(TRIM(sube_id),''), NULLIF(TRIM(sube),'')) IS NOT NULL
                  AND COALESCE(NULLIF(TRIM(sube_id),''), NULLIF(TRIM(sube),''))
                      NOT IN ('sube-merkez', 'MERKEZ')
                """
            )
            rw = cur.fetchone() or {}
            # 🟡 P2 (2026-08-13, EVV-PARA-N9): frontend "Bu ay toplam" KPI'sı
            # ozet.toplam bekliyordu ama özet yalnız sube_bekleyen taşıyordu →
            # KPI sessizce LIMIT'li listeden toplanıyordu (ay 200 kaydı aşarsa
            # eksik). Gerçek dönem toplamı LIMIT'siz burada hesaplanır.
            cur.execute(f"""
                SELECT COALESCE(COUNT(*),0)::int AS adet, COALESCE(SUM(ag.tutar),0) AS toplam
                FROM anlik_giderler ag
                WHERE {durum_cond}{ay_cond}
            """, ay_params)
            _dt = cur.fetchone() or {}
            return {
                "satirlar": satirlar,
                "ozet": {
                    "toplam": float(_dt.get("toplam") or 0),
                    "adet": int(_dt.get("adet") or 0),
                    "sube_bekleyen": {
                        "adet": int(rw.get("adet") or 0),
                        "toplam": float(rw.get("toplam") or 0),
                    }
                },
            }
        return satirlar

@app.post("/api/anlik-gider")
def anlik_gider_ekle(g: AnlikGider):
    # 🔴 P0 (2026-08-12, Para modülü denetimi): negatif tutar reddedilmiyordu.
    # Kart modunda negatif tutar limit kontrolünü geçip kart_hareketleri'ne
    # NEGATİF HARCAMA yazıp kart borcunu DÜŞÜRÜYORDU = uydurma kredi/limit açma.
    # Düzeltmeler DELETE/ters-kayıtla yapılır; negatif tutara meşru ihtiyaç yok.
    # Kaynakta normalize + pozitiflik zorunlu (API doğrudan da çağrılabilir).
    try:
        g.tutar = round(float(g.tutar), 2)
    except (TypeError, ValueError):
        raise HTTPException(400, "Tutar sayı olmalı")
    if g.tutar <= 0:
        raise HTTPException(400, "Tutar pozitif olmalı")
    # 🔴 SAHİP TALİMATI (2026-08-14): "para çıkışlarında açıklama zorunlu hale getir!"
    # Canlı vaka: 26 Tem 42.000 ₺ anlık gider AÇIKLAMASIZ girilmiş — aylar sonra
    # neye gittiği kimse bilmiyor, belge de bağlanamıyor. Elle girilen serbest
    # metinli çıkışta ad zorunludur; sistem-üretimli akışlar (plan/vadeli/borç
    # ödemesi) kendi açıklamasını taşır, onlara dokunulmadı.
    if len((g.aciklama or "").strip()) < 3:
        raise HTTPException(400, "Açıklama zorunlu — para çıkışı adsız olamaz (neye ödendiğini yazın)")
    sv = (g.sube or "").strip()
    if sv and sv.upper() != "MERKEZ":
        raise HTTPException(
            400,
            "Şube anlık gideri CFO ekranından doğrudan yazılamaz. "
            "Şube personel panelinden girin; kayıt onay kuyruğuna düşer, onay sonrası kasaya işlenir.",
        )
    with db() as (conn, cur):
        if not g.force:
            cur.execute("""
                SELECT id FROM anlik_giderler WHERE durum='aktif'
                AND tarih BETWEEN %s::date - INTERVAL '7 days' AND %s::date + INTERVAL '7 days'
                AND ABS(tutar - %s) < 1 AND kategori = %s
            """, (str(g.tarih), str(g.tarih), g.tutar, g.kategori))
            benzer = cur.fetchall()
            if benzer:
                return {"warning": True, "mesaj": f"Son 7 günde benzer kayıt var ({len(benzer)} adet). Yine de kaydetmek için force=true gönderin."}

        # KART ile ödeme — kart validasyon
        if g.odeme_yontemi == 'kart':
            if not g.kart_id:
                raise HTTPException(400, "Kart seçimi zorunlu")
            cur.execute("SELECT * FROM kartlar WHERE id=%s AND aktif=TRUE FOR UPDATE", (g.kart_id,))
            kart = cur.fetchone()
            if not kart: raise HTTPException(404, "Kart bulunamadı")
            borc = kart_borc(cur, g.kart_id)
            kalan_limit = _kanonik_kalan_limit(g.kart_id, float(kart['limit_tutar']) - borc)
            if kalan_limit < g.tutar:
                raise HTTPException(400, f"Kart limiti yetersiz. Kalan: {kalan_limit:,.0f} ₺")

        gid = str(uuid.uuid4())
        # V4 (Ödeme Merkezi): opsiyonel tedarikçi kolonu — lazy migration
        cur.execute("ALTER TABLE anlik_giderler ADD COLUMN IF NOT EXISTS tedarikci TEXT")
        _ted = (g.tedarikci or "").strip() or None
        # sube_id GENERATED (2026-08); yazılmaz, sube'den türer.
        cur.execute("""INSERT INTO anlik_giderler
            (id,tarih,kategori,tutar,aciklama,sube,odeme_yontemi,kart_id,kaynak_id,kaynak_tablo,tedarikci)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (gid, g.tarih, g.kategori, g.tutar, g.aciklama, g.sube,
             g.odeme_yontemi, g.kart_id, g.kaynak_id, g.kaynak_tablo, _ted))

        if g.odeme_yontemi == 'kart':
            # Karta HARCAMA yaz — kasaya yazma
            # FIX O3 (2026-07-05): kaynak_id/kaynak_tablo ile anlık gidere KANONİK bağ
            # kur — silme artık açıklama-LIKE yerine bu bağı kullanır (yanlış kayıt iptali önlenir).
            hid = str(uuid.uuid4())
            cur.execute("""
                INSERT INTO kart_hareketleri
                    (id, kart_id, tarih, islem_turu, tutar, taksit_sayisi, aciklama, kaynak_id, kaynak_tablo)
                VALUES (%s, %s, %s, 'HARCAMA', %s, 1, %s, %s, 'anlik_giderler')
            """, (hid, g.kart_id, g.tarih, g.tutar,
                  f"Anlık gider: {g.aciklama or g.kategori}", gid))
        else:
            # NAKİT — kasaya yaz. Şube + ödeme yöntemi TAŞINIR (2026-08-09):
            # hangi şubenin kasasından çıktı ve elden mi havale mi.
            insert_kasa_hareketi(cur, g.tarih, 'ANLIK_GIDER', -abs(g.tutar),
                f"Anlık gider: {g.aciklama or g.kategori}", 'anlik_giderler', gid,
                sube_id=_sube_kanonik(cur, getattr(g, 'sube', None)),
                odeme_yontemi=getattr(g, 'odeme_yontemi', None))

        audit(cur, 'anlik_giderler', gid, 'INSERT')
        if g.odeme_yontemi == 'kart':
            kart_plan_guncelle_tx(cur)

    # V4 — supplier_payment.py başlığındaki vaat: ödeme ekranından tedarikçi
    # SEÇİLDİYSE olay KESİN güvenle (confidence=1.0) doğar. Hata-yutar, izole.
    if _ted:
        try:
            from supplier_payment import _ensure_tablo as _spe_ensure
            with db() as (_c2, cur2):
                _spe_ensure(cur2)
                cur2.execute("SELECT id FROM tedarikciler WHERE aktif=TRUE AND LOWER(TRIM(ad))=LOWER(%s) LIMIT 1", (_ted,))
                _tr = cur2.fetchone()
                cur2.execute("""
                    INSERT INTO supplier_payment_event
                        (tedarikci_id, tedarikci_ad, tutar, tarih, kaynak,
                         kaynak_tablo, kaynak_id, confidence, eslesme_yontemi, aciklama)
                    VALUES (%s,%s,%s,%s,%s,'anlik_giderler',%s,1.0,'manuel',%s)
                    ON CONFLICT (kaynak_tablo, kaynak_id) DO UPDATE
                    SET tedarikci_ad=EXCLUDED.tedarikci_ad, confidence=1.0,
                        eslesme_yontemi='manuel'
                """, ((dict(_tr)["id"] if _tr else None), _ted, abs(g.tutar), g.tarih,
                      ('kart' if g.odeme_yontemi == 'kart' else 'nakit'),
                      gid, (g.aciklama or g.kategori)))
        except Exception as _spe_e:  # noqa: BLE001
            logging.getLogger(__name__).warning(f"supplier_payment_event (v4) yazilamadi: {_spe_e}")

    return {"id": gid, "success": True}

@app.delete("/api/anlik-gider/{gid}")
def anlik_gider_sil(gid: str):
    with db() as (conn, cur):
        cur.execute("SELECT * FROM anlik_giderler WHERE id=%s AND durum='aktif'", (gid,))
        eski = cur.fetchone()
        if not eski: raise HTTPException(404, "Kayıt bulunamadı veya zaten iptal edilmiş")
        cur.execute("UPDATE anlik_giderler SET durum='iptal' WHERE id=%s", (gid,))
        if eski.get('odeme_yontemi') == 'kart' and eski.get('kart_id'):
            # FIX O3 (2026-07-05): kanonik bağ (kaynak_id) ile iptal — açıklama-LIKE
            # yanlış/çift kayıt iptal edebiliyordu. Eski kayıtlar (bağsız) için
            # tarih+tutar+LIMIT 1 ile TEK kayıt garantili fallback (LIKE tek başına değil).
            cur.execute("""
                UPDATE kart_hareketleri SET durum='iptal'
                WHERE kaynak_tablo='anlik_giderler' AND kaynak_id=%s
                  AND islem_turu='HARCAMA' AND durum='aktif'
            """, (gid,))
            if cur.rowcount == 0:
                cur.execute("""
                    UPDATE kart_hareketleri SET durum='iptal'
                    WHERE id = (
                        SELECT id FROM kart_hareketleri
                        WHERE kart_id=%s AND islem_turu='HARCAMA' AND durum='aktif'
                          AND tarih=%s AND ABS(tutar - %s) < 0.01
                          AND aciklama LIKE %s
                        ORDER BY id LIMIT 1
                    )
                """, (eski['kart_id'], eski['tarih'], float(eski.get('tutar') or 0),
                      f"%{eski.get('aciklama') or eski['kategori']}%"))
        else:
            # NAKİT — ters kasa kaydı
            iptal_kasa_hareketi(cur, gid, 'anlik_giderler', 'ANLIK_GIDER', 'ANLIK_GIDER_IPTAL', 'Anlık gider iptali')
        audit(cur, 'anlik_giderler', gid, 'IPTAL', eski=eski)
        # FIX KP3 (2026-07-05): nakit gider iptali nakit_giderler'i değiştirir → o günün kapanış
        # farkını tazele (yanlış zimmet önlenir). anlik_giderler.sube = kapanış farkı sorgusunun
        # (kasa_fark_recalc) kullandığı değerle aynı. Kart gideri ise / uyarı yoksa no-op. Hata-yutar.
        try:
            from kasa_fark_recalc import sube_gun_kapanis_recalc
            sube_gun_kapanis_recalc(cur, eski.get('sube'), eski['tarih'])
        except Exception:
            pass
    return {"success": True}

# ── KARTLAR ────────────────────────────────────────────────────
class KartModel(BaseModel):
    kart_adi: str
    banka: str
    limit_tutar: float
    kesim_gunu: int
    son_odeme_gunu: int
    faiz_orani: float = 0.0          # Akdi (yıllık) faiz oranı %
    asgari_oran: float = 40.0        # Bankanın asgari ödeme oranı (%)
    gecikme_faiz_orani: float = 0.0  # Asgari altı ödemede uygulanan yıllık % (0 → akdi×1.3 fallback)
    son_dort_hane: Optional[str] = None  # PDF ekstre eşleştirme için son 4 hane
    sahip: Optional[str] = None      # Kart sahibi (İşletme / Annem / ...) — sorumluluk ayrımı
    ortak_limit_grup: Optional[str] = None  # Aynı krediyi paylaşan kartlar (aynı etiket = ortak limit)

@app.get("/api/kartlar")
def kartlar_listele():
    with db() as (conn, cur):
        cur.execute("SELECT * FROM kartlar WHERE aktif=TRUE ORDER BY banka")
        kartlar = [dict(r) for r in cur.fetchall()]
        sonuc = []
        bugun = bugun_tr()
        # PERF N+1 FIX (2026-07-06): kart başına ayrı borç sorgusu yerine tüm kart borçları
        # TEK sorguda (finans_core.tum_kart_borclari — aynı kanonik formül, GROUP BY'lı hâli).
        _borc_map = tum_kart_borclari(cur)
        for k in kartlar:
            # ── CORE HESAPLAR ──────────────────────────────────
            borc     = _borc_map.get(str(k['id']), 0.0)
            ekstre_v = kart_ekstre(cur, k['id'], k['kesim_gunu'])
            aylik_taksit = ekstre_v["aylik_taksit"]

            # AKTİF DÖNEM hesabı — kart_aktif_donem önceki kapanmış dönemin
            # GERÇEK ödeme verisinden devreden anapara + faizi (KKDF/BSMV dahil)
            # hesaplar; aktif/önündeki ekstre buradan beslenir.
            try:
                aktif = kart_aktif_donem(cur, k['id'])
            except Exception:
                aktif = None
            if aktif:
                bu_ekstre        = float(aktif.get("ekstre_toplam") or 0)
                asgari_odeme     = float(aktif.get("asgari_tahmini") or 0)
                devreden_ana     = float(aktif.get("devreden_anapara") or 0)
                devreden_fz      = float(aktif.get("devreden_faiz") or 0)
                aktif_donem_ay   = aktif.get("ay")
                aktif_kesim      = aktif.get("kesim_tarihi")
                aktif_son_odeme  = aktif.get("son_odeme_tarihi")
                onceki_ekstre    = float(aktif.get("onceki_ekstre") or 0)
                onceki_asgari    = float(aktif.get("onceki_asgari") or 0)
                onceki_odenen    = float(aktif.get("onceki_odenen") or 0)
                onceki_durum     = aktif.get("onceki_durum") or "yok"
                bu_donem_odenen  = float(aktif.get("bu_donem_odenen") or 0)
            else:
                bu_ekstre        = ekstre_v["ekstre_toplam"]
                asgari_odeme     = bu_ekstre * kart_asgari_orani(k)
                devreden_ana     = 0.0
                devreden_fz      = ekstre_v.get("devreden_faiz", 0)
                aktif_donem_ay = aktif_kesim = aktif_son_odeme = None
                onceki_ekstre = onceki_asgari = onceki_odenen = 0.0
                onceki_durum  = "yok"
                bu_donem_odenen = 0.0

            # GERÇEK ekstre snapshot'ı (PDF yükle / manuel ekstre) varsa "bu ekstre"
            # ve "asgari ödeme" gösterimini onunla eşitle — kart_hareketleri
            # bazlı tahmin, büyük "devir" bakiyesini göremediği için gerçek
            # dönem borcundan çok düşük çıkabilir (panel ↔ ödeme planı tutarsızlığı).
            _kesim_for_ov = aktif_kesim or kesim_tarihi_hesapla(bugun.year, bugun.month, int(k['kesim_gunu']))
            _ov_borc, _ov_asgari = kart_ekstre_donem_override(cur, k['id'], _kesim_for_ov)
            _ekstre_gercek = _ov_borc is not None  # PDF/manuel ekstre snapshot'ı var mı?
            if _ov_borc is not None:
                bu_ekstre = _ov_borc
                asgari_odeme = _ov_asgari if _ov_asgari is not None else round(_ov_borc * kart_asgari_orani(k), 2)

            # Ekstre snapshot'ından: bankanın yazdığı kullanılabilir limit + kalan taksit
            # yükü (Worldcard "Kalan Toplam Taksit Tutarı"nı doğrudan basar).
            _kull_limit = None
            _kalan_taksit = None
            _snap_kesim = None   # gerçek ekstre kesim tarihi (anlık borç penceresi için)
            try:
                cur.execute(
                    """SELECT kullanilabilir_limit, kalan_taksit_tutari FROM kart_ekstre_donem
                       WHERE kart_id=%s AND donem = DATE_TRUNC('month', %s::date)
                       ORDER BY olusturma DESC LIMIT 1""",
                    (k['id'], _kesim_for_ov),
                )
                _klr = cur.fetchone()
                if not _klr or (_klr.get('kullanilabilir_limit') is None and _klr.get('kalan_taksit_tutari') is None):
                    # Fallback (donem mismatch): kartın EN SON snapshot'ındaki limit/taksit.
                    # _ov_borc fallback'i (kart_ekstre_donem_override) ile tutarlı.
                    cur.execute(
                        """SELECT kullanilabilir_limit, kalan_taksit_tutari FROM kart_ekstre_donem
                           WHERE kart_id=%s AND (kullanilabilir_limit IS NOT NULL OR kalan_taksit_tutari IS NOT NULL)
                           ORDER BY donem DESC, olusturma DESC LIMIT 1""",
                        (k['id'],),
                    )
                    _klr = cur.fetchone() or _klr
                if _klr:
                    if _klr.get('kullanilabilir_limit') is not None:
                        _kull_limit = float(_klr['kullanilabilir_limit'])
                    if _klr.get('kalan_taksit_tutari') is not None:
                        _kalan_taksit = float(_klr['kalan_taksit_tutari'])
                # ANLIK borç penceresi = GERÇEK ekstre kesim tarihi (teorik _kesim_for_ov
                # değil). Aksi halde kesim-sonrası manuel ödemeler yanlış pencerede kaçar
                # (ödeme borçtan düşmez / limit açılmaz). En son snapshot'ın kesim_tarihi.
                cur.execute(
                    """SELECT kesim_tarihi::text AS kt FROM kart_ekstre_donem
                       WHERE kart_id=%s AND donem_borcu IS NOT NULL AND kesim_tarihi IS NOT NULL
                       ORDER BY donem DESC, olusturma DESC LIMIT 1""",
                    (k['id'],),
                )
                _kr = cur.fetchone()
                if _kr and _kr.get('kt'):
                    _snap_kesim = _kr['kt']
            except Exception:
                pass

            # Gelecek ekstre = bir sonraki kesim için tek çekim + taksit payı.
            # Aktif dönem zaten "şu an açık" olan ekstre; "gelecek" demek
            # aktif dönemden BİR SONRAKİ kesim demek. Aralık:
            # (aktif_kesim, aktif_kesim + 1 ay] içine düşen tek çekim harcamalar.
            if aktif_kesim:
                cur.execute("""
                    SELECT COALESCE(SUM(tutar),0) AS gelecek
                    FROM kart_hareketleri
                    WHERE kart_id = %s AND durum = 'aktif' AND islem_turu = 'HARCAMA'
                      AND taksit_sayisi = 1
                      AND tarih >  %s::date
                      AND tarih <= %s::date + INTERVAL '1 month'
                """, (k['id'], aktif_kesim, aktif_kesim))
                gelecek_tek = float(cur.fetchone()['gelecek'])
            else:
                gelecek_tek = 0.0
            gelecek_ekstre = gelecek_tek + aylik_taksit

            limit = float(k['limit_tutar'])
            son_odeme_gun = k['son_odeme_gunu']
            # AKTİF dönemin son ödeme tarihi varsa onu kullan (kart_aktif_donem'den);
            # yoksa eski mantık (bu ayın son ödeme günü, geçmişse bir sonraki ay).
            if aktif_son_odeme:
                try:
                    son_odeme = datetime.strptime(aktif_son_odeme, "%Y-%m-%d").date()
                except Exception:
                    son_odeme = _safe_date(bugun.year, bugun.month, son_odeme_gun)
            else:
                # _safe_date: son_odeme_gun ay sonunu aşarsa (31 → Şubat) o ayın son gününe kırpar.
                # Ham date() Şubat'ta ValueError verip kart panelini çökertiyordu (K3).
                son_odeme = _safe_date(bugun.year, bugun.month, son_odeme_gun)
                if son_odeme < bugun:
                    if bugun.month == 12:
                        son_odeme = _safe_date(bugun.year+1, 1, son_odeme_gun)
                    else:
                        son_odeme = _safe_date(bugun.year, bugun.month+1, son_odeme_gun)
            gun_kaldi = (son_odeme - bugun).days

            cur.execute("""SELECT * FROM odeme_plani WHERE kart_id=%s AND durum='bekliyor'
                ORDER BY tarih ASC LIMIT 1""", (k['id'],))
            yaklasan = cur.fetchone()

            # Gelecek taksit yükü (DÖNGÜ-SABİT, snapshot'tan): Worldcard doğrudan
            # (kalan_taksit); diğerleri ekstre kesimindeki kullanılabilir limitten türetir
            # (limit − kullanılabilir − dönem borcu). Ödeme/harcama ile değişmez.
            if _kalan_taksit is not None:
                _gelecek_taksit = _kalan_taksit
            elif _kull_limit is not None:
                _gelecek_taksit = max(0.0, limit - _kull_limit - float(_ov_borc or 0))
            else:
                _gelecek_taksit = None
            # ANLIK güncel borç: ekstre dönem borcu + KESİM SONRASI defter hareketleri
            # (ödeme −, kullanım/harcama +). Böylece defteri değiştirmeden gerçek-zamanlı
            # borç çıkar; gelecek ay yeni ekstre yüklenince taban kendiliğinden döner.
            if _ekstre_gercek and _ov_borc is not None:
                try:
                    # Gerçek ekstre kesim tarihinden sonraki hareketler (ödeme −, harcama +).
                    # _snap_kesim yoksa teorik _kesim_for_ov'a düşer (geriye dönük uyumlu).
                    _pencere = _snap_kesim or str(_kesim_for_ov)
                    # 🔴 P0 DENKLEŞTİRME ÇİZGİSİ (2026-08-17, Ziraat vakası):
                    # DEVİR yaması (manuel-ekstre/toplu-devir) kartın borcunu O ANKİ
                    # ekstre rakamına eşitler — yani yama tarihine kadarki TÜM ödeme
                    # ve harcamaları ZATEN İÇİNDE taşır. Yama kesimden SONRA
                    # tarihliyse, aşağıdaki pencere o ödemeleri BİR KEZ DAHA düşüyordu.
                    # CANLI KANIT: Ziraat 3696 — kesim 18 Tem, 20 Tem ödeme 182.275,
                    # 25 Tem yama (borç 268.443'e eşitlendi = ödeme yamanın içinde),
                    # 28 Tem ödeme 86.168 → ekran 0,35 ₺ + "asgari ödendi ✓" YEŞİL
                    # gösteriyordu; gerçek borç 182.275,35 ₺ (sahte-yeşil ailesinin
                    # en tehlikeli vakası: kart ödenmiş sanılır).
                    # Kural: pencere = MAX(ekstre kesimi, son denkleştirme tarihi).
                    cur.execute("""
                        SELECT MAX(tarih)::text AS t FROM kart_hareketleri
                        WHERE kart_id=%s AND durum='aktif' AND islem_turu='DEVIR'
                    """, (k['id'],))
                    _devir_ts = str((cur.fetchone() or {}).get('t') or '')[:10]
                    if _devir_ts and _devir_ts > str(_pencere)[:10]:
                        _pencere = _devir_ts
                    cur.execute("""
                        SELECT COALESCE(SUM(CASE WHEN islem_turu='ODEME' THEN -tutar ELSE tutar END), 0) AS d
                        FROM kart_hareketleri
                        WHERE kart_id=%s AND durum='aktif' AND islem_turu <> 'DEVIR'
                          AND tarih > %s::date
                    """, (k['id'], _pencere))
                    _post = float((cur.fetchone() or {}).get('d') or 0)
                except Exception:
                    _post = 0.0
                _anlik = float(_ov_borc) + _post
            else:
                _anlik = borc
            # ASGARİ KARŞILANDI doğru ölçümü: GERÇEK snapshot kesim tarihinden SONRA
            # yapılan ödemeler bu ekstrenin asgarisine sayılır. kart_aktif_donem teorik
            # kesim kullandığı için kesim-sonrası ödemeleri kaçırıp bu_donem_odenen=0
            # bırakabiliyordu → asgari ödense bile kutucuk yeşile dönmüyordu.
            if _ekstre_gercek and _snap_kesim:
                try:
                    # Aynı denkleştirme çizgisi burada da geçerli (Ziraat vakası):
                    # yamadan ÖNCEKİ ödemeler yamanın içinde eridiği için bu
                    # ekstrenin asgarisine ikinci kez sayılamaz — yoksa kart
                    # ödenmemişken "asgari karşılandı ✓" yeşili yanıyordu.
                    _as_pencere = _snap_kesim
                    cur.execute("""
                        SELECT MAX(tarih)::text AS t FROM kart_hareketleri
                        WHERE kart_id=%s AND durum='aktif' AND islem_turu='DEVIR'
                    """, (k['id'],))
                    _dts = str((cur.fetchone() or {}).get('t') or '')[:10]
                    if _dts and _dts > str(_as_pencere)[:10]:
                        _as_pencere = _dts
                    cur.execute("""
                        SELECT COALESCE(SUM(tutar), 0) AS odenen
                        FROM kart_hareketleri
                        WHERE kart_id=%s AND durum='aktif' AND islem_turu='ODEME'
                          AND tarih > %s::date
                    """, (k['id'], _as_pencere))
                    bu_donem_odenen = float((cur.fetchone() or {}).get('odenen') or 0)
                except Exception:
                    pass
            # Kalan limit (GERÇEK ZAMANLI): limit − anlık borç − gelecek taksit yükü
            if _gelecek_taksit is not None:
                _kalan_limit = limit - _anlik - _gelecek_taksit
            else:
                _kalan_limit = limit - _anlik
            _doluluk = ((limit - _kalan_limit) / limit) if limit > 0 else 0
            # Toplam borç = anlık + gelecek taksit yükü. Taksit bilinmiyorsa (None) sadece
            # anlık (boş/None bırakma → frontend "Toplam Borç" hücresi eksik görünmesin).
            _toplam_taksitli = _anlik + (_gelecek_taksit or 0)
            # ── 🧭 ADIM 9: KANONİK ALANLAR LİSTEYE EKLENİYOR ──────────────
            # Ekranlar bugüne kadar `guncel_borc` / `anlik_borc` / `donem_borcu`
            # gibi BEŞ AYRI YOLDAN beslendi ve aynı ad farklı anlam taşıyordu
            # (17 Ağu: OPET aynı anda 508.023,92 ve 190.218,39 gösterdi).
            # Kanonik model (kart_bakiye_ozeti) 2026-08-17'de kuruldu ama
            # PARALEL duruyordu — hiçbir ekran ondan beslenmiyordu.
            # Burada aynı yanıta `kanonik_*` önekiyle ekleniyor: mevcut alanlar
            # AYNEN duruyor (hiçbir ekran bozulmaz), tüketiciler tek tek
            # taşındıkça eskiler emekliye ayrılabilir (Adım 11).
            # HATA-YUTAR: kanonik hesap düşse liste yine döner — kart ekranı
            # tek bir kartın hesabı yüzünden komple boş kalmasın.
            _kan = {}
            try:
                _kb = kart_bakiye_ozeti(cur, str(k["id"]))
                if _kb and not _kb.get("hata"):
                    _kan = {
                        "kanonik_ekstre_borcu": _kb.get("ekstre_borcu"),
                        "kanonik_anlik_borc": _kb.get("anlik_borc"),
                        "kanonik_defter_bakiye": _kb.get("defter_canli_bakiye"),
                        "kanonik_gelecek_taksit": _kb.get("gelecek_taksit_yuku"),
                        "kanonik_gelecek_taksit_kaynak": _kb.get("gelecek_taksit_kaynak"),
                        "kanonik_toplam_yukumluluk": _kb.get("toplam_yukumluluk"),
                        "kanonik_mutabakat_farki": _kb.get("mutabakat_farki"),
                        "kanonik_mutabakat_notu": _kb.get("mutabakat_notu"),
                        "kanonik_asgari": _kb.get("asgari_tutar"),
                        "kanonik_denklestirme_cizgisi": _kb.get("denklestirme_cizgisi"),
                    }
            except Exception as _ke:  # noqa: BLE001
                logger.warning("kanonik bakiye atlandi (%s): %s", k.get("kart_adi"), str(_ke)[:110])
                _kan = {"kanonik_hata": str(_ke)[:120]}

            sonuc.append({**k, **_kan,
                "guncel_borc": borc,
                # ANLIK borç = ekstre dönem borcu + kesim sonrası ödeme/kullanım (gerçek zamanlı).
                # Ekstresiz kartta defter borcu (borc). Ana gösterilen rakam budur.
                "anlik_borc": round(_anlik, 2),
                "kalan_limit": _kalan_limit,
                "limit_doluluk": _doluluk,
                # Gerçek (PDF/manuel) ekstre snapshot'ı varsa dönem borcu = ekstre borcu.
                "ekstre_gercek": _ekstre_gercek,
                "donem_borcu": (round(_ov_borc, 2) if _ov_borc is not None else None),
                # Bankanın yazdığı gerçek kullanılabilir limit (varsa); yoksa None.
                "kullanilabilir_limit": (round(_kull_limit, 2) if _kull_limit is not None else None),
                # Gelecek taksit yükü + taksitlerle toplam borç (taksitler artık göz ardı edilmez)
                "gelecek_taksit_anapara": (round(_gelecek_taksit, 2) if _gelecek_taksit is not None else None),
                "toplam_borc_taksitli": (round(_toplam_taksitli, 2) if _toplam_taksitli is not None else None),
                "asgari_odeme": asgari_odeme,
                "bu_donem_odenen": bu_donem_odenen,
                # Asgari estimasyon olduğundan kuruş/yuvarlama farkı için küçük tolerans
                # (×0.999 — kart_aktif_donem'deki önceki dönem kontrolüyle aynı desen).
                "asgari_karsilandi": asgari_odeme > 0 and bu_donem_odenen >= asgari_odeme * 0.999,
                "bu_ekstre": bu_ekstre,
                "devreden_anapara": devreden_ana,
                "devreden_faiz": devreden_fz,
                "tek_cekim": ekstre_v.get("tek_cekim", 0),
                "gelecek_ekstre": gelecek_ekstre,
                "aylik_taksit": aylik_taksit,
                "gun_kaldi": gun_kaldi,
                "son_odeme_tarihi": str(son_odeme),
                "aktif_donem":      aktif_donem_ay,
                "aktif_kesim":      aktif_kesim,
                "aktif_son_odeme":  aktif_son_odeme,
                "onceki_ekstre":    onceki_ekstre,
                "onceki_asgari":    onceki_asgari,
                "onceki_odenen":    onceki_odenen,
                "onceki_durum":     onceki_durum,
                "blink": gun_kaldi <= 0 and yaklasan is not None,
                "yaklasan_odeme": dict(yaklasan) if yaklasan else None
            })
        # ── ORTAK LİMİT HAVUZU: aynı grubu paylaşan kartlarda kalan limiti TEK
        #    havuzdan hesapla (çift sayma). grup_limit = gruptaki en büyük limit.
        gruplar = {}
        for s in sonuc:
            g = (s.get("ortak_limit_grup") or "").strip()
            if g:
                gruplar.setdefault(g, []).append(s)
        for g, uyeler in gruplar.items():
            if len(uyeler) < 2:
                continue
            grup_limit = max(float(u.get("limit_tutar") or 0) for u in uyeler)
            # Üye taahhüdü = dönem borcu (ekstre varsa) + gelecek taksit yükü → taksitler
            # paylaşılan limitten de düşülsün (göz ardı edilmesin).
            def _committed(u):
                # Anlık borç (kesim sonrası ödeme/kullanım dahil) + gelecek taksit yükü
                return float(u.get("anlik_borc") or u.get("guncel_borc") or 0) + float(u.get("gelecek_taksit_anapara") or 0)
            grup_borc = sum(_committed(u) for u in uyeler)
            grup_kalan = grup_limit - grup_borc
            for u in uyeler:
                u["ortak_grup_limit"] = round(grup_limit, 2)
                u["ortak_grup_borc"] = round(grup_borc, 2)
                u["ortak_grup_uye"] = len(uyeler)
                u["kalan_limit"] = round(grup_kalan, 2)  # paylaşılan → çift sayma yok
                u["limit_doluluk"] = (grup_borc / grup_limit) if grup_limit > 0 else 0
        return sonuc

def _kart_validate(k: KartModel):
    """Kart ekle/güncelle ortak doğrulaması — bozuk değerleri reddet."""
    if not (k.kart_adi or "").strip():
        raise HTTPException(400, "Kart adı zorunlu")
    if not (k.banka or "").strip():
        raise HTTPException(400, "Banka zorunlu")
    if float(k.limit_tutar or 0) < 0:
        raise HTTPException(400, "Limit negatif olamaz")
    if not (1 <= int(k.kesim_gunu or 0) <= 31):
        raise HTTPException(400, "Kesim günü 1–31 arası olmalı")
    if not (1 <= int(k.son_odeme_gunu or 0) <= 31):
        raise HTTPException(400, "Son ödeme günü 1–31 arası olmalı")
    if not (0 <= float(k.faiz_orani or 0) <= 500):
        raise HTTPException(400, "Faiz oranı 0–500 arası olmalı (yıllık %)")
    if not (0 <= float(k.gecikme_faiz_orani or 0) <= 500):
        raise HTTPException(400, "Gecikme faiz oranı 0–500 arası olmalı")
    if not (10 <= float(k.asgari_oran or 0) <= 100):
        raise HTTPException(400, "Asgari oran 10–100 arası olmalı (%)")


@app.post("/api/kartlar")
def kart_ekle(k: KartModel):
    _kart_validate(k)
    with db() as (conn, cur):
        # ÇİFT KART KORUMASI: aynı son 4 haneli AKTİF kart varsa yenisini AÇMA — mevcuti döndür.
        # (son_dort_hane eşleştirme anahtarı; çiftlenirse mutabakat bozulur. "Kartı Ekle"ye
        #  iki kez basmak / iki ekstreyi arka arkaya atmak güvenli → idempotent.)
        if k.son_dort_hane:
            _s4 = k.son_dort_hane.strip()[-4:]
            cur.execute("SELECT id::text, kart_adi FROM kartlar WHERE son_dort_hane=%s AND aktif=TRUE LIMIT 1", (_s4,))
            _ex = cur.fetchone()
            if _ex:
                _ex = dict(_ex)
                return {"id": _ex["id"], "success": True, "mevcut": True,
                        "mesaj": f"Bu son 4 hane (…{_s4}) zaten kayıtlı: {_ex['kart_adi']} — eşleştirildi, yeni kart açılmadı."}
        kid = str(uuid.uuid4())
        cur.execute("""INSERT INTO kartlar (id,kart_adi,banka,limit_tutar,kesim_gunu,son_odeme_gunu,faiz_orani,asgari_oran,gecikme_faiz_orani,son_dort_hane,sahip,ortak_limit_grup)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (kid, k.kart_adi, k.banka, k.limit_tutar, k.kesim_gunu, k.son_odeme_gunu,
             k.faiz_orani, k.asgari_oran, k.gecikme_faiz_orani,
             k.son_dort_hane.strip()[-4:] if k.son_dort_hane else None,
             (k.sahip or 'İşletme').strip() or 'İşletme',
             (k.ortak_limit_grup or '').strip() or None))
        audit(cur, 'kartlar', kid, 'INSERT')
    return {"id": kid, "success": True}

@app.put("/api/kartlar/{kid}")
def kart_guncelle(kid: str, k: KartModel):
    _kart_validate(k)
    with db() as (conn, cur):
        cur.execute("SELECT * FROM kartlar WHERE id=%s", (kid,))
        eski = cur.fetchone()
        if not eski: raise HTTPException(404)
        cur.execute("""UPDATE kartlar SET kart_adi=%s,banka=%s,limit_tutar=%s,
            kesim_gunu=%s,son_odeme_gunu=%s,faiz_orani=%s,asgari_oran=%s,gecikme_faiz_orani=%s,
            son_dort_hane=%s,sahip=%s,ortak_limit_grup=%s
            WHERE id=%s""",
            (k.kart_adi, k.banka, k.limit_tutar, k.kesim_gunu, k.son_odeme_gunu,
             k.faiz_orani, k.asgari_oran, k.gecikme_faiz_orani,
             k.son_dort_hane.strip()[-4:] if k.son_dort_hane else None,
             (k.sahip or 'İşletme').strip() or 'İşletme',
             (k.ortak_limit_grup or '').strip() or None, kid))
        audit(cur, 'kartlar', kid, 'UPDATE', eski=eski)
    return {"success": True}

@app.delete("/api/kartlar/{kid}")
def kart_sil(kid: str):
    with db() as (conn, cur):
        cur.execute("SELECT * FROM kartlar WHERE id=%s", (kid,))
        eski = cur.fetchone()
        if not eski: raise HTTPException(404)
        # K2 (2026-08-12, Kart denetimi): pasif kart aktif=TRUE özetlerden düşer →
        # borçlu kart pasife alınınca borcu SESSİZCE kaybolur. 'Pasife Al' bilinçli
        # akış olduğu için BLOKLAMIYORUZ ama borcu audit'e yaz + uyarı dön (görünsün).
        try:
            _borc = round(float(kart_borc(cur, kid) or 0), 2)
        except Exception:
            _borc = 0.0
        cur.execute("UPDATE kartlar SET aktif=FALSE WHERE id=%s", (kid,))
        audit(cur, 'kartlar', kid, 'PASIF', eski=eski, yeni={'pasifken_borc': _borc})
    if _borc > 0.01:
        return {"success": True,
                "uyari": f"Bu kartın {_borc:,.2f} ₺ aktif borcu vardı — pasif kart borcu özetlerde görünmez. Borç kapanmadıysa önce çözün."}
    return {"success": True}

class KartKaliciSilBody(BaseModel):
    onay_pin: Optional[str] = None


@app.post("/api/kartlar/{kid}/kalici-sil")
def kart_kalici_sil(kid: str, body: KartKaliciSilBody):
    """Kartı KALICI siler (kayıtlardan tamamen kaldırır). Yalnızca AKTİF işlemi
    olmayan kartlar (test/yanlış eklenen) silinebilir; işlemi olan kart silinemez →
    'Pasife Al' kullanılır (defter bütünlüğü). İşletme onayı (Merve Karabacak PIN) şart."""
    from operasyon_merkez_api import _isletme_onay_dogrula
    with db() as (conn, cur):
        cur.execute("SELECT id, kart_adi FROM kartlar WHERE id=%s", (kid,))
        k = cur.fetchone()
        if not k:
            raise HTTPException(404, "Kart bulunamadı")
        kart_adi = dict(k)["kart_adi"]
        onayci = _isletme_onay_dogrula(cur, body.onay_pin)  # PIN hatalı → 403
        # 🔴 P1 (2026-08-14, EVV-KART / Codex): guard yalnız AKTİF hareket sayıyordu —
        # 'ledger-sifirla' tüm hareketleri iptal'e çekince fiilen KULLANILMIŞ kart
        # "aktif hareket yok" diye kalıcı silmeden geçiyordu (sıfırla→sil zinciri
        # tüm izi yok ediyordu; docstring'deki "yalnız işlemsiz kart" kuralı deliniyordu).
        # Artık HERHANGİ bir hareket izi (iptal dahil) kalıcı silmeyi engeller.
        cur.execute(
            "SELECT COUNT(*) AS n FROM kart_hareketleri WHERE kart_id=%s", (kid,)
        )
        if int(dict(cur.fetchone())["n"]) > 0:
            raise HTTPException(
                409,
                "Bu kartın hareket İZİ var (iptal edilmişler dahil) → kalıcı silinemez; "
                "defter izi korunur. Kartı kullanımdan kaldırmak için 'Pasife Al' kullanın.",
            )
        # İşlemsiz kart → bağlı (iptal) kayıtları + snapshot + plan temizle, sonra kartı sil
        cur.execute("DELETE FROM kart_hareketleri WHERE kart_id=%s", (kid,))
        cur.execute("DELETE FROM kart_ekstre_donem WHERE kart_id=%s", (kid,))
        # onay_kuyrugu odeme_plani'na bağlı olabilir → önce o kartın planlarına ait
        # bekleyen onay kayıtlarını temizle, sonra TÜM odeme_plani (her durum) sil (FK).
        cur.execute(
            "UPDATE onay_kuyrugu SET durum='reddedildi' "
            "WHERE durum='bekliyor' AND kaynak_id IN (SELECT id FROM odeme_plani WHERE kart_id=%s)",
            (kid,),
        )
        cur.execute("DELETE FROM odeme_plani WHERE kart_id=%s", (kid,))
        cur.execute("DELETE FROM kartlar WHERE id=%s", (kid,))
        audit(cur, "kartlar", kid, "KALICI_SIL", yeni={"kart_adi": kart_adi, "onayci": onayci.get("ad_soyad")})
    return {"success": True, "kart_adi": kart_adi}


@app.post("/api/kartlar/{kid}/tam-sifirla")
def kart_tam_sifirla(kid: str, body: KartKaliciSilBody):
    """Kart TANIMINI (isim/banka/limit/son4hane) KORUR ama o karta ait TÜM
    geçmişi siler: kart_hareketleri (harcama/ödeme/faiz), kart_ekstre_donem
    (aylık ekstre snapshot'ları + faiz hesapları), bağlı ödeme planı ve
    ekstre-import kaynaklı anlık gider kayıtları. Faiz/gecikme faiz oranını
    da sıfırlar (bir sonraki ekstre yüklemesinde otomatik güncellenir).
    İşletme onayı (Merve Karabacak PIN) şart. kid='__hepsi__' → tüm aktif kartlar."""
    from operasyon_merkez_api import _isletme_onay_dogrula
    with db() as (conn, cur):
        onayci = _isletme_onay_dogrula(cur, body.onay_pin)  # PIN hatalı → 403
        if kid == "__hepsi__":
            cur.execute("SELECT id::text, kart_adi FROM kartlar WHERE aktif=TRUE")
            kartlar_l = [dict(r) for r in (cur.fetchall() or [])]
        else:
            cur.execute("SELECT id::text, kart_adi FROM kartlar WHERE id=%s", (kid,))
            r = cur.fetchone()
            if not r:
                raise HTTPException(404, "Kart bulunamadı")
            kartlar_l = [dict(r)]

        sonuc = []
        for k in kartlar_l:
            kk = k["id"]
            cur.execute(
                """UPDATE kasa_hareketleri SET durum='iptal'
                   WHERE kaynak_tablo='kart_hareketleri' AND durum='aktif'
                     AND kaynak_id IN (SELECT id FROM kart_hareketleri WHERE kart_id=%s)""",
                (kk,),
            )
            cur.execute(
                """DELETE FROM anlik_giderler
                   WHERE kaynak_tablo='ekstre_import'
                     AND kaynak_id IN (SELECT id FROM kart_hareketleri WHERE kart_id=%s)""",
                (kk,),
            )
            cur.execute("DELETE FROM kart_hareketleri WHERE kart_id=%s", (kk,))
            silinen_hareket = cur.rowcount
            cur.execute("DELETE FROM kart_ekstre_donem WHERE kart_id=%s", (kk,))
            silinen_donem = cur.rowcount
            cur.execute(
                "UPDATE onay_kuyrugu SET durum='reddedildi' "
                "WHERE durum='bekliyor' AND kaynak_id IN (SELECT id FROM odeme_plani WHERE kart_id=%s)",
                (kk,),
            )
            cur.execute("DELETE FROM odeme_plani WHERE kart_id=%s", (kk,))
            cur.execute("UPDATE kartlar SET faiz_orani=0, gecikme_faiz_orani=0 WHERE id=%s", (kk,))
            audit(cur, "kartlar", kk, "TAM_SIFIRLA",
                  yeni={"kart_adi": k["kart_adi"], "onayci": onayci.get("ad_soyad"),
                        "silinen_hareket": silinen_hareket, "silinen_donem": silinen_donem})
            sonuc.append({"kart_id": kk, "kart_adi": k["kart_adi"],
                           "silinen_hareket": silinen_hareket, "silinen_donem": silinen_donem})
    uyari_cache_clear()
    return {"success": True, "kartlar": sonuc}


@app.get("/api/kartlar/{kid}/taksitler")
def kart_taksitler(kid: str):
    """
    Kartın aktif taksitli harcamaları — kalan/geçen taksit dahil.
    """
    with db() as (conn, cur):
        cur.execute("SELECT * FROM kartlar WHERE id=%s AND aktif=TRUE", (kid,))
        if not cur.fetchone(): raise HTTPException(404, "Kart bulunamadı")
        return {
            "taksitler":      taksit_detay(cur, kid),
            "gelecek_yukler": gelecek_taksit_yuku(cur, kid, ay_sayisi=3),
        }

@app.get("/api/kartlar/taksit-yuku")
def tum_taksit_yuku():
    """Tüm aktif kartların önümüzdeki 3 aylık taksit yükü."""
    with db() as (conn, cur):
        return tum_kartlar_taksit_yuku(cur, ay_sayisi=3)

@app.get("/api/kartlar/{kid}/ekstre-forecast")
def kart_ekstre_forecast_endpoint(kid: str, aylar: int = 6, senaryo: str = "odenir"):
    """
    Kartın gelecek N ay ekstre tahmini — banka kesmeden önce.
    senaryo: 'tam' | 'odenir' | 'odenmez' (zincirleme faiz varsayımı)
    """
    aylar = max(1, min(12, int(aylar or 6)))
    if senaryo not in ("tam", "odenir", "odenmez"):
        senaryo = "odenir"
    with db() as (conn, cur):
        cur.execute("SELECT 1 FROM kartlar WHERE id=%s AND aktif=TRUE", (kid,))
        if not cur.fetchone():
            raise HTTPException(404, "Kart bulunamadı")
        return {
            "aylar": aylar,
            "senaryo": senaryo,
            "donemler": kart_ekstre_forecast(cur, kid, aylar, senaryo),
        }

@app.get("/api/kartlar/ekstre-forecast")
def tum_kartlar_ekstre_forecast_endpoint(aylar: int = 6, senaryo: str = "odenir"):
    """Tüm aktif kartların gelecek N ay ekstre forecast'i."""
    aylar = max(1, min(12, int(aylar or 6)))
    if senaryo not in ("tam", "odenir", "odenmez"):
        senaryo = "odenir"
    with db() as (conn, cur):
        return {
            "aylar": aylar,
            "senaryo": senaryo,
            "kartlar": tum_kartlar_ekstre_forecast(cur, aylar, senaryo),
        }

@app.put("/api/kartlar/{kid}/kesim-tarihi")
def kart_kesim_tarihi_guncelle(kid: str, body: dict):
    """
    Kartın son kesim tarihini ve toleransını güncelle.
    body: { son_kesim_tarihi: 'YYYY-MM-DD', kesim_tolerans: int }
    """
    with db() as (conn, cur):
        cur.execute("SELECT * FROM kartlar WHERE id=%s AND aktif=TRUE", (kid,))
        eski = cur.fetchone()
        if not eski: raise HTTPException(404, "Kart bulunamadı")
        son_kesim   = body.get('son_kesim_tarihi')
        tolerans    = body.get('kesim_tolerans', 0)
        cur.execute("""
            UPDATE kartlar
            SET son_kesim_tarihi = %s, kesim_tolerans = %s
            WHERE id = %s
        """, (son_kesim, tolerans, kid))
        audit(cur, 'kartlar', kid, 'KESIM_GUNCELLE', eski=eski)
    return {"success": True, "son_kesim_tarihi": son_kesim, "kesim_tolerans": tolerans}


class KartHareket(BaseModel):
    kart_id: str
    tarih: date
    islem_turu: str
    tutar: float
    taksit_sayisi: int = 1
    faiz_tutari: float = 0
    ana_para: float = 0
    aciklama: Optional[str] = None
    baslangic_tarihi: Optional[date] = None  # taksitli alımlar için
    harcama_tipi: Optional[str] = None       # 'isletme' | 'sahsi' | 'belirsiz'
    # 🏪 HANGİ KASADAN ÇIKTI (2026-08-17, sahip: "çıkışlar yapılırken bu kasaları
    # seçmedik çünkü öyle bir seçim gelmedi!"). Canlı defterde 21 kart ödemesinin
    # 21'i, 28 kredi taksidinin 28'i ŞUBESİZDİ — 4.005.571 ₺ para çıkışı "hangi
    # kasadan" sorusuna cevapsız kalıyordu. Boş bırakılırsa merkez kovası (NULL).
    sube_id: Optional[str] = None
    odeme_yontemi: Optional[str] = None      # 'elden' | 'havale' | 'nakit'

@app.get("/api/kart-hareketleri")
def kart_hareketleri(kart_id: Optional[str] = None, limit: int = 200):
    with db() as (conn, cur):
        if kart_id:
            cur.execute("""SELECT kh.*, k.banka, k.kart_adi FROM kart_hareketleri kh
                JOIN kartlar k ON k.id=kh.kart_id
                WHERE kh.kart_id=%s AND kh.durum='aktif' ORDER BY kh.tarih DESC LIMIT %s""", (kart_id, limit))
        else:
            cur.execute("""SELECT kh.*, k.banka, k.kart_adi FROM kart_hareketleri kh
                JOIN kartlar k ON k.id=kh.kart_id
                WHERE kh.durum='aktif' ORDER BY kh.tarih DESC LIMIT %s""", (limit,))
        return [dict(r) for r in cur.fetchall()]

@app.post("/api/kart-hareketleri")
def kart_hareket_ekle(h: KartHareket):
    # 🔴 SAHİP TALİMATI (2026-08-14): kart harcaması/ödemesi de PARA ÇIKIŞIDIR —
    # adsız olamaz. Bu uç yalnız ELLE girişin kapısıdır: ekstre importu, k1 tanısı,
    # vadeli-kart yolu ve kart analizi tabloya DOĞRUDAN yazar (grep ile doğrulandı),
    # bu guard'dan geçmez → otomatik akışlar etkilenmez.
    if h.islem_turu in ('HARCAMA', 'ODEME') and len((h.aciklama or '').strip()) < 3:
        raise HTTPException(400, "Açıklama zorunlu — para çıkışı adsız olamaz (neye ödendiğini yazın)")
    # 🔴 KART-013 (2026-09-02): kart defterinin sözleşmesi "POZİTİF BÜYÜKLÜK +
    # yönü islem_turu söyler"dir — kart borcu HARCAMA'ları toplayıp ODEME'leri
    # düşerek bulunur. Buraya negatif tutarlı bir HARCAMA girilirse borç ARTMAZ,
    # AZALIR. Kasa bacağı `-abs(h.tutar)` ile korunmuştu ama kart defteri
    # korunmamıştı: iki defter sessizce ayrışıyordu.
    # ⚠️ abs() ile sessizce düzeltMİYORUZ — işareti çevirmek kullanıcının ne
    # demek istediğini TAHMİN etmektir. Reddedip söylüyoruz.
    if h.islem_turu in ('HARCAMA', 'ODEME', 'FAIZ'):
        try:
            _t = float(h.tutar)
        except (TypeError, ValueError):
            raise HTTPException(400, "Geçerli bir tutar girin")
        if _t <= 0:
            raise HTTPException(
                400,
                f"Tutar pozitif olmalı. Kart defterinde yön işaretle değil işlem türüyle "
                f"belirtilir — iade/düzeltme için ODEME girin ya da kaydı iptal edin.")
    with db() as (conn, cur):
        hid = str(uuid.uuid4())
        faiz = abs(h.faiz_tutari) if h.faiz_tutari else 0
        ana  = abs(h.ana_para)   if h.ana_para   else 0
        # Taksitli alımda baslangic_tarihi = hareket tarihi (girilmemişse)
        bas_tarih = h.baslangic_tarihi or (h.tarih if h.taksit_sayisi > 1 else None)
        _htip = (h.harcama_tipi or '').strip().lower()
        if _htip not in ('isletme', 'sahsi', 'belirsiz'):
            _htip = 'belirsiz' if h.islem_turu == 'HARCAMA' else 'isletme'
        cur.execute("""INSERT INTO kart_hareketleri
            (id,kart_id,tarih,islem_turu,tutar,taksit_sayisi,faiz_tutari,ana_para,aciklama,baslangic_tarihi,harcama_tipi)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (hid, h.kart_id, h.tarih, h.islem_turu, h.tutar,
             h.taksit_sayisi, faiz, ana, h.aciklama, bas_tarih, _htip))
        if h.islem_turu == 'ODEME':
            # 🏪 ŞUBE + ÖDEME YÖNTEMİ TAŞINIR (2026-08-17): eskiden hiç
            # geçilmiyordu, bu yüzden 21 kart ödemesinin 21'i şubesizdi.
            # _sube_kanonik 'MERKEZ'i ve pasif şubeyi NULL'a indirir — uydurma
            # yapmaz, çözemezse merkez kovasında durur (gizlenmez).
            insert_kasa_hareketi(cur, str(h.tarih), 'KART_ODEME', -abs(h.tutar),
                h.aciklama or 'Kart ödemesi',
                'kart_hareketleri', hid, hid, 'KART_ODEME',
                sube_id=_sube_kanonik(cur, h.sube_id),
                odeme_yontemi=(h.odeme_yontemi or None))
        audit(cur, 'kart_hareketleri', hid, 'INSERT')
        if h.islem_turu in ('HARCAMA', 'ODEME', 'FAIZ'):
            kart_plan_guncelle_tx(cur)
    if h.islem_turu == 'ODEME':
        uyari_cache_clear()
    return {"id": hid, "success": True}

@app.delete("/api/kart-hareketleri/{hid}")
def kart_hareket_iptal(hid: str):
    with db() as (conn, cur):
        cur.execute("SELECT * FROM kart_hareketleri WHERE id=%s AND durum='aktif'", (hid,))
        eski = cur.fetchone()
        if not eski: raise HTTPException(404, "Kayıt bulunamadı veya zaten iptal edilmiş")
        cur.execute("UPDATE kart_hareketleri SET durum='iptal' WHERE id=%s", (hid,))
        kasa_iptal = None
        if eski['islem_turu'] == 'ODEME':
            # 🔴 KİLİTLİ KAYIT KUSURU (2026-08-18, canlı yakalandı):
            # Eskiden burada koşulsuz iptal_kasa_hareketi çağrılıyordu. O
            # fonksiyon kasa kaydı BULAMAZSA Exception atar → istek 500 verir →
            # kart hareketi de iptal olmaz. Sonuç: EKSTRE İMPORTUNDAN gelen bir
            # ödeme satırı HİÇ İPTAL EDİLEMİYORDU (import kasaya yazmaz, docstring
            # de bunu söyler). Canlı kanıt: Garanti 3018'e aktarılan 4 ödeme
            # satırı silinemedi — "İptal edilecek kayıt bulunamadı — KART_ODEME".
            # Yani yanlış aktarılmış bir ödeme defterde KALICI olarak sıkışıyordu.
            # Doğrusu: kasa karşılığı VARSA iptal et, YOKSA bu bir hata değildir —
            # kart tarafı yine de iptal olur ve durum yanıtta GÖRÜNÜR (sessiz geçme yok).
            cur.execute("""SELECT 1 FROM kasa_hareketleri
                            WHERE kaynak_id=%s AND islem_turu='KART_ODEME'
                              AND kasa_etkisi=TRUE AND durum='aktif' LIMIT 1""", (hid,))
            if cur.fetchone():
                iptal_kasa_hareketi(cur, hid, 'kart_hareketleri', 'KART_ODEME',
                                    'KART_ODEME_IPTAL', 'Kart ödemesi iptali')
                kasa_iptal = "iptal edildi"
            else:
                kasa_iptal = ("kasa karşılığı yok — bu satır kasaya hiç yazmamıştı "
                              "(ekstre importu kart borcuna yazar, kasaya dokunmaz)")
        # Ekstre import'tan otomatik açılmış eşlenik anlık gider varsa onu da iptal et
        if eski.get('kaynak_tablo') == 'ekstre_import' and eski['islem_turu'] == 'HARCAMA':
            cur.execute("UPDATE anlik_giderler SET durum='iptal' WHERE id=%s AND durum='aktif'", ("agk_" + hid,))
        audit(cur, 'kart_hareketleri', hid, 'IPTAL', eski=eski)
    return {"success": True, "kasa_iptal": kasa_iptal}


@app.post("/api/kart-hareketleri/{hid}/harcama-tipi")
def kart_hareket_tip_belirle(hid: str, tip: str):
    """Bir kart harcamasını şahsi / işletme / belirsiz olarak sınıflandırır (Faz K-A)."""
    t = (tip or '').strip().lower()
    if t not in ('isletme', 'sahsi', 'belirsiz'):
        raise HTTPException(400, "tip: isletme | sahsi | belirsiz")
    with db() as (conn, cur):
        cur.execute("UPDATE kart_hareketleri SET harcama_tipi=%s WHERE id=%s AND durum='aktif' RETURNING *", (t, hid))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Hareket bulunamadı")
        row = dict(row)

        # ⛔ ESKİ KOPYA MODELİ DİRİLTİLMEZ (2026-08-10, Codex denetimi #5).
        # Eskiden burada şahsi→iptal, işletme→(yeniden) OLUŞTUR yapılıyordu.
        # Kanonik modelde kart harcaması P&L'e kart defterinden girer; ikinci bir
        # gider kaydı çift sayımdır. Sınıflandırma değişince artık YALNIZ eski
        # kopyalar kapatılır — yenisi asla üretilmez.
        if row.get('kaynak_tablo') == 'ekstre_import' and row['islem_turu'] == 'HARCAMA':
            agid = "agk_" + hid
            cur.execute(
                "UPDATE anlik_giderler SET durum='arsiv' WHERE id=%s AND durum='aktif'",
                (agid,))
        if False:  # eski yol — bilinçli olarak ölü bırakıldı (geçmiş okunabilsin)
            agid = "agk_" + hid
            if t == 'sahsi':
                pass
            else:
                mevcut = None
                if mevcut:
                    pass
                else:
                    cur.execute(
                        """INSERT INTO anlik_giderler
                           (id, tarih, kategori, tutar, aciklama, sube, odeme_yontemi, kart_id, kaynak_id, kaynak_tablo)
                           VALUES (%s,%s,%s,%s,%s,'MERKEZ','kart',%s,%s,'ekstre_import')""",
                        (agid, row['tarih'], (row.get('kategori') or 'Diğer'), row['tutar'],
                         row.get('aciklama'), row['kart_id'], hid),
                    )
        # SATICI HAFIZASI: bu satıcıyı öğren → sonraki aynı satıcı otomatik önerilsin
        ogrenildi = None
        if t in ("isletme", "sahsi"):
            anahtar = _satici_anahtar(dict(row).get("aciklama"))
            if anahtar:
                cur.execute(
                    """INSERT INTO kart_satici_kural (anahtar, harcama_tipi, adet)
                       VALUES (%s, %s, 1)
                       ON CONFLICT (anahtar) DO UPDATE SET
                         harcama_tipi=EXCLUDED.harcama_tipi,
                         adet=kart_satici_kural.adet+1, guncelleme=NOW()""",
                    (anahtar, t),
                )
                ogrenildi = anahtar
        audit(cur, 'kart_hareketleri', hid, 'HARCAMA_TIPI', yeni={'tip': t})
    return {"success": True, "harcama_tipi": t, "ogrenilen_satici": ogrenildi}


@app.get("/api/kartlar/harcama-ozet")
def kart_harcama_ozet():
    """Şahsi/işletme/belirsiz kırılımı — kart başına ve toplam (sadece HARCAMA, aktif).
    'İşletmenin gerçek kart yükü' ile şahsi karışıklığı ayırır."""
    with db() as (conn, cur):
        cur.execute("""
            SELECT k.id::text AS kart_id, k.kart_adi, COALESCE(k.sahip,'İşletme') AS sahip,
                   kh.harcama_tipi,
                   COALESCE(SUM(kh.tutar),0)::float AS toplam,
                   COUNT(*)::int AS adet
            FROM kart_hareketleri kh
            JOIN kartlar k ON k.id = kh.kart_id
            WHERE kh.durum='aktif' AND kh.islem_turu='HARCAMA' AND k.aktif=TRUE
            GROUP BY k.id, k.kart_adi, k.sahip, kh.harcama_tipi
            ORDER BY k.kart_adi
        """)
        kartlar = {}
        gen = {'isletme': 0.0, 'sahsi': 0.0, 'belirsiz': 0.0}
        # 🟡 P2 (2026-08-14, EVV-KART): tip başına ADET yoktu — FE "sınıflandırılmayan
        # kaç hareket" sayısını 200-kesikli listeden sayıyordu (201+ harekette eksik).
        gen_adet = {'isletme': 0, 'sahsi': 0, 'belirsiz': 0}
        for r in (cur.fetchall() or []):
            r = dict(r)
            kid = r['kart_id']
            tip = r['harcama_tipi'] if r['harcama_tipi'] in gen else 'belirsiz'
            k = kartlar.setdefault(kid, {
                'kart_id': kid, 'kart_adi': r['kart_adi'], 'sahip': r['sahip'],
                'isletme': 0.0, 'sahsi': 0.0, 'belirsiz': 0.0, 'adet': 0,
            })
            k[tip] += float(r['toplam']); k['adet'] += int(r['adet'])
            gen[tip] += float(r['toplam'])
            gen_adet[tip] += int(r['adet'])
        return {
            "genel": {**gen, "toplam": round(sum(gen.values()), 2),
                      "isletme_adet": gen_adet['isletme'], "sahsi_adet": gen_adet['sahsi'],
                      "belirsiz_adet": gen_adet['belirsiz'],
                      "toplam_adet": sum(gen_adet.values())},
            "kartlar": sorted(kartlar.values(), key=lambda x: -(x['isletme'] + x['sahsi'] + x['belirsiz'])),
        }


@app.get("/api/kartlar/ekstre-ping")
def kart_ekstre_ping():
    """Teşhis: pdfplumber yüklü mü + yeni kod canlı mı (deploy doğrulama)."""
    try:
        import pdfplumber
        return {"ok": True, "pdfplumber": getattr(pdfplumber, "__version__", "?"), "marker": "e0-sync-v2"}
    except Exception as e:
        return {"ok": False, "hata": str(e), "marker": "e0-sync-v2"}


@app.get("/api/kartlar/borc-kocu")
def kart_borc_kocu(strateji: str = "cig", nakit: float = 0):
    """Borç ödeme koçu: hangi kartı önce kapat (çığ=en yüksek faiz / kartopu=en küçük
    borç), aylık faiz kaybı, eldeki nakitle bu ayki dağıtım önerisi.
    Standart borç-yönetimi çerçeveleri — kişiye özel mali tavsiye değildir."""
    nakit = max(0.0, float(nakit or 0))  # negatif nakit anlamsız → 0'a kırp
    from finans_core import son_odeme_tarihi_hesapla, kesim_tarihi_hesapla
    bugun = bugun_tr()
    with db() as (conn, cur):
        cur.execute("SELECT id::text, kart_adi, banka, COALESCE(sahip,'İşletme') AS sahip, "
                    "limit_tutar, faiz_orani, asgari_oran, kesim_gunu, son_odeme_gunu "
                    "FROM kartlar WHERE aktif=TRUE")
        kl = [dict(r) for r in (cur.fetchall() or [])]
        cur.execute("SELECT DISTINCT ON (kart_id) kart_id::text AS kid, asgari_tutar, "
                    "son_odeme_tarihi::text AS sot FROM kart_ekstre_donem ORDER BY kart_id, donem DESC")
        snap = {r["kid"]: dict(r) for r in (cur.fetchall() or [])}
        kartlar = []
        for k in kl:
            borc = float(kart_borc(cur, k["id"]) or 0)
            if borc <= 0.5:
                continue
            faiz_y = float(k.get("faiz_orani") or 0)
            aylik_faiz = round(borc * (faiz_y / 100 / 12), 2)
            s = snap.get(k["id"], {})
            asgari = float(s.get("asgari_tutar") or 0) or round(borc * float(k.get("asgari_oran") or 40) / 100, 2)
            so = s.get("sot")
            if not so:
                try:
                    kt = kesim_tarihi_hesapla(bugun.year, bugun.month, int(k.get("kesim_gunu") or 1))
                    so = str(son_odeme_tarihi_hesapla(kt, int(k.get("son_odeme_gunu") or 10)))
                except Exception:
                    so = None
            kartlar.append({
                "kart_id": k["id"], "kart_adi": k["kart_adi"], "sahip": k["sahip"],
                "borc": round(borc, 2), "faiz_yillik": faiz_y, "aylik_faiz": aylik_faiz,
                "asgari": round(asgari, 2), "son_odeme": so, "faiz_belirsiz": faiz_y <= 0,
                "onerilen_odeme": 0.0,
            })
        if strateji == "kartopu":
            kartlar.sort(key=lambda x: x["borc"])
        else:
            strateji = "cig"
            kartlar.sort(key=lambda x: (-x["faiz_yillik"], -x["borc"]))
        toplam_borc = round(sum(x["borc"] for x in kartlar), 2)
        toplam_aylik_faiz = round(sum(x["aylik_faiz"] for x in kartlar), 2)
        toplam_asgari = round(sum(x["asgari"] for x in kartlar), 2)
        kalan = float(nakit or 0)
        if kalan > 0:
            for x in kartlar:  # önce tüm asgariler (öncelik sırasıyla)
                pay = min(x["asgari"], x["borc"], kalan)
                x["onerilen_odeme"] = round(pay, 2); kalan = round(kalan - pay, 2)
                if kalan <= 0:
                    break
            for x in kartlar:  # kalanı önceliğe (çığ/kartopu sırası)
                if kalan <= 0:
                    break
                ek = min(kalan, round(x["borc"] - x["onerilen_odeme"], 2))
                if ek > 0:
                    x["onerilen_odeme"] = round(x["onerilen_odeme"] + ek, 2); kalan = round(kalan - ek, 2)
        return {
            "strateji": strateji, "nakit": float(nakit or 0),
            "toplam_borc": toplam_borc, "toplam_aylik_faiz": toplam_aylik_faiz,
            "toplam_asgari": toplam_asgari,
            "asgari_karsilaniyor": (float(nakit or 0) >= toplam_asgari) if nakit else None,
            "artan_nakit": round(kalan, 2) if nakit else 0,
            "oncelik": kartlar[0] if kartlar else None,
            "kartlar": kartlar,
        }


@app.get("/api/kartlar/borc-projeksiyon")
def kart_borc_projeksiyon(aylik: float, strateji: str = "cig"):
    """Borç kurtuluş projeksiyonu: aylık X ödersen kaç ayda biter + toplam faiz;
    sadece asgari ödersen ne kadar faiz/ay kaybedersin. Gerçek faiz simülasyonu
    (çığ/kartopu). Standart çerçeve — kişiye özel mali tavsiye değildir."""
    aylik = max(0.0, float(aylik or 0))  # negatif aylık anlamsız → 0'a kırp
    from datetime import date as _d
    bugun = bugun_tr()
    with db() as (conn, cur):
        cur.execute("SELECT id::text, faiz_orani, asgari_oran FROM kartlar WHERE aktif=TRUE")
        kl = [dict(r) for r in (cur.fetchall() or [])]
        kartlar = []
        for k in kl:
            b = float(kart_borc(cur, k["id"]) or 0)
            if b <= 0.5:
                continue
            cur.execute("SELECT asgari_tutar FROM kart_ekstre_donem WHERE kart_id=%s ORDER BY donem DESC LIMIT 1", (k["id"],))
            sr = cur.fetchone()
            asg = float((dict(sr).get("asgari_tutar") if sr else 0) or 0) or round(b * float(k.get("asgari_oran") or 40) / 100, 2)
            kartlar.append({"borc": b, "oran": float(k.get("faiz_orani") or 0) / 100 / 12,
                            "asgari": asg,
                            "asgari_oran": float(k.get("asgari_oran") or 40) / 100})

    toplam_borc = round(sum(c["borc"] for c in kartlar), 2)
    toplam_asgari = round(sum(c["asgari"] for c in kartlar), 2)

    def simule(butce, asgari_dinamik=False):
        # 🟡 P2 (2026-08-14, Codex): "sadece asgari" senaryosu İLK ayın asgarisini
        # sabit bütçe olarak kullanıyordu — gerçek minimum-only'de asgari borçla
        # birlikte HER AY AZALIR, süre uzar ve faiz büyür. asgari_dinamik=True
        # her ay asgariyi güncel borçtan yeniden hesaplar (kart asgari_oran'ı).
        cs = [dict(c) for c in kartlar]
        cs.sort(key=(lambda x: x["borc"]) if strateji == "kartopu" else (lambda x: -x["oran"]))
        ay = 0; tfaiz = 0.0; onceki = sum(c["borc"] for c in cs)
        while sum(c["borc"] for c in cs) > 1 and ay < 360:
            ay += 1
            for c in cs:
                f = c["borc"] * c["oran"]; c["borc"] += f; tfaiz += f
            if asgari_dinamik:
                for c in cs:
                    c["asgari"] = round(c["borc"] * c["asgari_oran"], 2)
                butce = sum(c["asgari"] for c in cs)
            kalan = butce
            for c in cs:
                if kalan <= 0:
                    break
                p = min(c["asgari"], c["borc"], kalan); c["borc"] = round(c["borc"] - p, 2); kalan = round(kalan - p, 2)
            for c in cs:
                if kalan <= 0:
                    break
                e = min(kalan, c["borc"]); c["borc"] = round(c["borc"] - e, 2); kalan = round(kalan - e, 2)
            simdi = sum(c["borc"] for c in cs)
            if simdi >= onceki - 0.01:  # bütçe faizi karşılamıyor → borç azalmıyor
                return {"ay": None, "toplam_faiz": round(tfaiz, 2), "bitmedi": True}
            onceki = simdi
        return {"ay": ay, "toplam_faiz": round(tfaiz, 2), "bitmedi": sum(c["borc"] for c in cs) > 1}

    def bitis(ay):
        if not ay:
            return None
        m = bugun.month - 1 + ay
        return str(_d(bugun.year + m // 12, m % 12 + 1, 1))

    verilen = simule(float(aylik or 0))
    asgari_only = simule(toplam_asgari, asgari_dinamik=True)
    tasarruf = None
    erken_ay = None
    if verilen.get("ay") and asgari_only.get("ay"):
        tasarruf = round(asgari_only["toplam_faiz"] - verilen["toplam_faiz"], 2)
        erken_ay = asgari_only["ay"] - verilen["ay"]
    return {
        "aylik": float(aylik or 0), "strateji": "kartopu" if strateji == "kartopu" else "cig",
        "toplam_borc": toplam_borc, "toplam_asgari": toplam_asgari,
        "verilen": {**verilen, "bitis_tarihi": bitis(verilen.get("ay"))},
        "asgari_only": {**asgari_only, "bitis_tarihi": bitis(asgari_only.get("ay"))},
        "tasarruf_faiz": tasarruf, "erken_ay": erken_ay,
    }


@app.get("/api/kartlar/analiz")
def kart_analiz_ozet():
    """Faz: saf ANALİZ görünümü — içe aktarılmış veriden (kart_hareketleri +
    kart_ekstre_donem). Yükleme yok; tek yükleme noktası Ekstre Yükle."""
    with db() as (conn, cur):
        cur.execute("""
            SELECT donem::text AS donem,
                   COALESCE(SUM(donem_borcu),0)::float AS borc,
                   COALESCE(SUM(donem_faizi),0)::float AS faiz
            FROM kart_ekstre_donem GROUP BY donem ORDER BY donem
        """)
        aylik = [dict(r) for r in (cur.fetchall() or [])]
        cur.execute("""
            SELECT COALESCE(NULLIF(kategori,''),'Diğer') AS kategori,
                   COALESCE(SUM(tutar),0)::float AS tutar, COUNT(*)::int AS adet
            FROM kart_hareketleri
            WHERE durum='aktif' AND islem_turu='HARCAMA' AND kategori IS NOT NULL
            GROUP BY COALESCE(NULLIF(kategori,''),'Diğer') ORDER BY tutar DESC
        """)
        kategori = [dict(r) for r in (cur.fetchall() or [])]
        cur.execute("""
            SELECT k.kart_adi, COALESCE(SUM(kh.tutar),0)::float AS harcama
            FROM kart_hareketleri kh JOIN kartlar k ON k.id=kh.kart_id
            WHERE kh.durum='aktif' AND kh.islem_turu='HARCAMA' AND k.aktif=TRUE
            GROUP BY k.kart_adi HAVING SUM(kh.tutar) > 0 ORDER BY harcama DESC
        """)
        kart_bazli = [dict(r) for r in (cur.fetchall() or [])]
        return {"aylik": aylik, "kategori": kategori, "kart_bazli": kart_bazli,
                "veri_var": bool(aylik or kategori or kart_bazli)}


@app.get("/api/kartlar/ekstre-arsiv")
def kart_ekstre_arsiv():
    """Faz: kart × ay ekstre ARŞİVİ — her kart için tüm aylık ekstre snapshot'ları
    (kart_ekstre_donem). Her (kart, donem) ayrı satır; aylar birikir. Salt görünürlük."""
    with db() as (conn, cur):
        _ekstre_belge_kolonlari(cur)
        cur.execute("""
            SELECT ked.kart_id::text AS kart_id,
                   k.kart_adi, k.banka, COALESCE(k.sahip,'İşletme') AS sahip,
                   k.son_dort_hane,
                   ked.donem::text            AS donem,
                   ked.kesim_tarihi::text     AS kesim_tarihi,
                   ked.son_odeme_tarihi::text AS son_odeme_tarihi,
                   COALESCE(ked.donem_borcu,0)::float   AS donem_borcu,
                   COALESCE(ked.asgari_tutar,0)::float  AS asgari_tutar,
                   COALESCE(ked.onceki_borc,0)::float   AS onceki_borc,
                   COALESCE(ked.donem_harcama,0)::float AS donem_harcama,
                   COALESCE(ked.donem_odeme,0)::float   AS donem_odeme,
                   COALESCE(ked.donem_faizi,0)::float   AS donem_faizi,
                   COALESCE(ked.kalan_taksit,0)::float  AS kalan_taksit,
                   ked.kaynak,
                   (ked.belge_pdf IS NOT NULL) AS belge_var
            FROM kart_ekstre_donem ked
            JOIN kartlar k ON k.id = ked.kart_id
            ORDER BY k.kart_adi, ked.donem DESC
        """)
        rows = [dict(r) for r in (cur.fetchall() or [])]
        kartlar = {}
        for r in rows:
            kid = r["kart_id"]
            grup = kartlar.get(kid)
            if grup is None:
                grup = {
                    "kart_id": kid, "kart_adi": r["kart_adi"], "banka": r["banka"],
                    "sahip": r["sahip"], "son_dort_hane": r["son_dort_hane"],
                    "donemler": [], "donem_adet": 0,
                    "toplam_faiz": 0.0, "son_donem": None,
                }
                kartlar[kid] = grup
            grup["donemler"].append({
                "donem": r["donem"], "kesim_tarihi": r["kesim_tarihi"],
                "son_odeme_tarihi": r["son_odeme_tarihi"],
                "donem_borcu": r["donem_borcu"], "asgari_tutar": r["asgari_tutar"],
                "onceki_borc": r["onceki_borc"], "donem_harcama": r["donem_harcama"],
                "donem_odeme": r["donem_odeme"], "donem_faizi": r["donem_faizi"],
                "kalan_taksit": r["kalan_taksit"], "kaynak": r["kaynak"],
                # PDF aslı arşivliyse indirme adresi — FE tıkla→aç bunu kullanır.
                "belge_var": bool(r.get("belge_var")),
                "belge_url": (f"/api/kartlar/ekstre-belge?kart_id={kid}&donem={str(r['donem'])[:10]}"
                              if r.get("belge_var") else None),
            })
            grup["donem_adet"] += 1
            grup["toplam_faiz"] = round(grup["toplam_faiz"] + r["donem_faizi"], 2)
            if grup["son_donem"] is None:  # donemler DESC sıralı → ilk gelen en yeni
                grup["son_donem"] = r["donem"]
        kart_list = sorted(kartlar.values(), key=lambda g: g["kart_adi"] or "")
        return {"kartlar": kart_list, "kart_adet": len(kart_list),
                "veri_var": bool(kart_list)}


@app.post("/api/kartlar/gercek-modeli-tasi")
def kart_gercek_modeli_tasi(kuru: bool = True):
    """📦 GEÇMİŞİ YENİ MODELE TAŞI (yol haritası ADIM 3/12, 2026-08-17).

    Mevcut veriyi ADIM 2'de kurulan tablolara KOPYALAR. Eski tablolara HİÇ
    dokunmaz — kart_hareketleri ve kart_ekstre_donem aynen kalır, sistem eskisi
    gibi çalışmaya devam eder. Amaç: yeni modelin gerçek veriyle doğrulanabilmesi.

    `kuru=true` (varsayılan): hiçbir şey yazmaz, ne taşınacağını SAYAR. Sahip
    rakamları görüp onaylayınca `kuru=false` ile gerçek taşıma yapılır.

    İDEMPOTANS: her kaydın kimliği ESKİ KİMLİĞİNDEN türetilir (kaynak_tur='legacy',
    kaynak_id=eski id) ve tekil kısıtla korunur → uç iki kez çağrılsa da mükerrer
    üretmez. Burada ON CONFLICT DO NOTHING MEŞRUDUR: kimlik içerik tahmini değil,
    eski satırın gerçek kimliğidir (bugünkü hash tuzağının tersi).

    TAŞINANLAR:
      1. kart_ekstre_donem.belge_pdf → kart_belge          (sha256 tekil)
      2. kart_ekstre_donem            → kart_donem          (kart+kesim tekil)
      3. kart_ekstre_donem            → kart_okuma_surumu   (legacy başlık)
      4. kart_hareketleri (DEVIR hariç) → kart_defter_kaydi
      5. kart_hareketleri (DEVIR)      → kart_acilis_devri  (kavram ayrımı!)
    """
    import hashlib
    sayac = {"belge": 0, "donem": 0, "surum": 0, "defter": 0, "acilis": 0,
             "atlanan_belge": 0, "hata": []}
    with db() as (conn, cur):
        try:
            from database import ensure_kart_gercek_modeli
            ensure_kart_gercek_modeli(cur)
        except Exception as e:
            raise HTTPException(500, f"Yeni tablolar kurulamadı: {str(e)[:120]}")

        # ── Kaynak: ekstre dönemleri (belge + dönem + okuma sürümü) ──────────
        cur.execute("""
            SELECT ked.id::text AS eski_id, ked.kart_id::text AS kart_id,
                   ked.donem::text AS donem, ked.kesim_tarihi::text AS kesim,
                   ked.son_odeme_tarihi::text AS son_odeme,
                   ked.donem_borcu::float AS borc, ked.asgari_tutar::float AS asgari,
                   ked.onceki_borc::float AS onceki, ked.donem_harcama::float AS harcama,
                   ked.donem_odeme::float AS odeme, ked.donem_faizi::float AS faiz,
                   ked.kalan_taksit_tutari::float AS kalan_taksit,
                   ked.kullanilabilir_limit::float AS kull_limit,
                   ked.kaynak, ked.belge_pdf, ked.belge_ad
              FROM kart_ekstre_donem ked
              JOIN kartlar k ON k.id = ked.kart_id
             ORDER BY ked.kart_id, ked.donem
        """)
        donemler = [dict(r) for r in (cur.fetchall() or [])]

        for d in donemler:
            kesim = d.get("kesim") or (d.get("donem") or "")[:10]
            if not kesim:
                sayac["hata"].append(f"kesim yok: {d['eski_id'][:8]}")
                continue

            # 1) BELGE — PDF aslı varsa sha256 ile
            belge_id = None
            if d.get("belge_pdf"):
                ham = bytes(d["belge_pdf"])
                sha = hashlib.sha256(ham).hexdigest()
                if not kuru:
                    from psycopg2 import Binary as _B
                    cur.execute(
                        """INSERT INTO kart_belge (sha256, dosya_adi, boyut, pdf, yukleyen)
                           VALUES (%s,%s,%s,%s,'legacy')
                           ON CONFLICT (sha256) DO NOTHING RETURNING id::text""",
                        (sha, d.get("belge_ad"), len(ham), _B(ham)))
                    r = cur.fetchone()
                    if r:
                        belge_id = r["id"]; sayac["belge"] += 1
                    else:
                        cur.execute("SELECT id::text FROM kart_belge WHERE sha256=%s", (sha,))
                        belge_id = (cur.fetchone() or {}).get("id")
                        sayac["atlanan_belge"] += 1
                else:
                    sayac["belge"] += 1

            # 2) DÖNEM — (kart, kesim) tekil
            donem_id = None
            if not kuru:
                cur.execute(
                    """INSERT INTO kart_donem (kart_id, kesim_tarihi, son_odeme_tarihi, durum)
                       VALUES (%s,%s::date,%s::date,'kabul')
                       ON CONFLICT (kart_id, kesim_tarihi) DO NOTHING RETURNING id::text""",
                    (d["kart_id"], kesim, d.get("son_odeme")))
                r = cur.fetchone()
                if r:
                    donem_id = r["id"]; sayac["donem"] += 1
                else:
                    cur.execute("SELECT id::text FROM kart_donem WHERE kart_id=%s AND kesim_tarihi=%s::date",
                                (d["kart_id"], kesim))
                    donem_id = (cur.fetchone() or {}).get("id")
            else:
                sayac["donem"] += 1

            # 3) OKUMA SÜRÜMÜ — eski snapshot'ın başlığı (legacy, çapa bilinmiyor)
            if belge_id and not kuru:
                baslik = {k: d.get(k) for k in
                          ("borc", "asgari", "onceki", "harcama", "odeme", "faiz",
                           "kalan_taksit", "kull_limit")}
                baslik["legacy_kaynak"] = d.get("kaynak")
                cur.execute(
                    """INSERT INTO kart_okuma_surumu
                           (belge_id, okuyucu_surumu, baslik, capa_durumu)
                       VALUES (%s, 'legacy/2026-08-17', %s::jsonb, 'bilinmiyor')
                       ON CONFLICT (belge_id, okuyucu_surumu) DO NOTHING
                       RETURNING id::text""",
                    (belge_id, json.dumps(baslik, default=str)))
                if cur.fetchone():
                    sayac["surum"] += 1
            elif belge_id:
                sayac["surum"] += 1

        # ── Kaynak: kart hareketleri ────────────────────────────────────────
        cur.execute("""
            SELECT h.id::text AS eski_id, h.kart_id::text AS kart_id,
                   h.tarih::text AS tarih, h.islem_turu, h.tutar::float AS tutar,
                   COALESCE(h.taksit_sayisi,1) AS taksit_sayisi,
                   h.aciklama, h.harcama_tipi, h.baslangic_tarihi::text AS bas,
                   h.durum
              FROM kart_hareketleri h JOIN kartlar k ON k.id=h.kart_id
             WHERE h.durum='aktif'
             ORDER BY h.kart_id, h.tarih
        """)
        hareketler = [dict(r) for r in (cur.fetchall() or [])]

        for h in hareketler:
            if str(h["islem_turu"]).upper() == "DEVIR":
                # 5) AÇILIŞ DEVRİ — kavram ayrımı: devir artık defter kaydı DEĞİL
                if not kuru:
                    cur.execute(
                        """INSERT INTO kart_acilis_devri
                               (kart_id, gecerlilik, tutar, dayanak, karar_veren)
                           VALUES (%s,%s::date,%s,%s,'legacy')
                           ON CONFLICT (kart_id, gecerlilik) DO NOTHING""",
                        (h["kart_id"], h["tarih"], h["tutar"],
                         f"legacy:{h['eski_id']} · {(h.get('aciklama') or '')[:80]}"))
                    if cur.rowcount:
                        sayac["acilis"] += 1
                else:
                    sayac["acilis"] += 1
                continue

            # 4) DEFTER KAYDI — kimlik ESKİ KİMLİKTEN (tekil kısıt korur)
            if not kuru:
                cur.execute(
                    """INSERT INTO kart_defter_kaydi
                           (kart_id, tarih, islem_turu, tutar, aciklama, harcama_tipi,
                            kaynak_tur, kaynak_id)
                       VALUES (%s,%s::date,%s,%s,%s,%s,'legacy',%s)
                       ON CONFLICT (kaynak_tur, kaynak_id) DO NOTHING""",
                    (h["kart_id"], h["tarih"], h["islem_turu"], h["tutar"],
                     h.get("aciklama"), h.get("harcama_tipi"), h["eski_id"]))
                if cur.rowcount:
                    sayac["defter"] += 1
            else:
                sayac["defter"] += 1

        # ── Doğrulama: taşınan toplamlar eski toplamlarla kuruş-eşit mi ──────
        dogrulama = {}
        if not kuru:
            cur.execute("""
                SELECT COALESCE(SUM(tutar),0)::float AS yeni FROM kart_defter_kaydi
                 WHERE kaynak_tur='legacy'""")
            yeni_top = float((cur.fetchone() or {}).get("yeni") or 0)
            cur.execute("""
                SELECT COALESCE(SUM(tutar),0)::float AS eski FROM kart_hareketleri
                 WHERE durum='aktif' AND islem_turu <> 'DEVIR'""")
            eski_top = float((cur.fetchone() or {}).get("eski") or 0)
            dogrulama = {
                "defter_eski_toplam": round(eski_top, 2),
                "defter_yeni_toplam": round(yeni_top, 2),
                "fark": round(yeni_top - eski_top, 2),
                "kurus_esit": abs(yeni_top - eski_top) < 0.01,
            }

    return {
        "kuru_kosu": kuru,
        "kaynak": {"donem_adet": len(donemler), "hareket_adet": len(hareketler)},
        "tasinan": sayac,
        "dogrulama": dogrulama,
        "not": ("KURU KOŞU — hiçbir şey yazılmadı, yalnız sayıldı. Gerçek taşıma için "
                "?kuru=false" if kuru else
                "Taşıma tamamlandı. Eski tablolara DOKUNULMADI; sistem eskisi gibi çalışıyor."),
    }


@app.get("/api/kartlar/bakiye-karsilastir")
def kart_bakiye_karsilastir():
    """🧭 ESKİ-YENİ BAKİYE KARŞILAŞTIRMASI (yol haritası ADIM 4/12, 2026-08-17).

    Kanonik modeli (finans_core.kart_bakiye_ozeti) bugünkü BEŞ ayrı yolla yan yana
    koyar. Amaç: tüketicileri taşımadan ÖNCE farkı ölçmek — 17 Ağustos denetiminde
    OPET WORLD aynı anda 508.023,92 ve 190.218,39 gösteriyordu.

    SALT OKUR: hiçbir şey yazmaz, hiçbir davranışı değiştirmez. ADIM 9'da ekranlar
    kanonik modele taşınırken bu uç "fark sıfır mı" kanıtı olarak kullanılacak.

    Dönen her satırda `uyusuyor` alanı: kanonik anlık borç ile eski panel yolu
    (borc-faiz-ozet → anlik_borc) 1 ₺ toleransında aynı mı?
    """
    satirlar = []
    with db() as (conn, cur):
        cur.execute("SELECT id::text FROM kartlar WHERE aktif=TRUE ORDER BY kart_adi")
        kart_idler = [r["id"] for r in (cur.fetchall() or [])]

        # Eski yol 1: kart_borc (defter, DEVİR dahil, takvim-payı taksit)
        eski_defter = tum_kart_borclari(cur)

        for kid in kart_idler:
            yeni = kart_bakiye_ozeti(cur, kid)
            if yeni.get("hata"):
                satirlar.append({"kart_id": kid, "hata": yeni["hata"]})
                continue

            # Eski yol 2: kart_aktif_donem (bu ekstre / asgari üreten yol)
            try:
                aktif = kart_aktif_donem(cur, kid) or {}
            except Exception as _e:
                aktif = {"hata": str(_e)[:80]}

            # Eski yol 3: snapshot override (son-snapshot'a düşen tuzaklı yol)
            try:
                ov_borc, ov_asgari = kart_ekstre_donem_override(
                    cur, kid, _safe_date(yeni.get("ekstre_kesim")) or date.today())
            except Exception:
                ov_borc, ov_asgari = None, None

            eski_defter_borc = round(float(eski_defter.get(kid) or 0), 2)
            k_anlik = yeni.get("anlik_borc")
            fark_defter = round(eski_defter_borc - float(yeni.get("defter_canli_bakiye") or 0), 2)

            satirlar.append({
                "kart_id": kid,
                "kart_adi": yeni.get("kart_adi"),
                # YENİ kanonik model
                "kanonik": {
                    "ekstre_borcu": yeni.get("ekstre_borcu"),
                    "anlik_borc": k_anlik,
                    "defter_canli_bakiye": yeni.get("defter_canli_bakiye"),
                    "gelecek_taksit_yuku": yeni.get("gelecek_taksit_yuku"),
                    # Rakamın NEREDEN geldiği (2026-08-18): plan=ÖLÇÜM ·
                    # ekstre_alani=bankanın yazdığı · limit_cikarimi=TAHMİN.
                    # Bu uç "fark sıfır mı" kanıtı olarak kullanılıyor; kaynağı
                    # taşımazsa tahmini rakam ölçüm sanılır. Canlı örnek:
                    # limit çıkarımı 65.810,05 derken ölçüm 87.890,59 çıktı.
                    "gelecek_taksit_kaynak": yeni.get("gelecek_taksit_kaynak"),
                    "toplam_yukumluluk": yeni.get("toplam_yukumluluk"),
                    "mutabakat_farki": yeni.get("mutabakat_farki"),
                    "asgari_tutar": yeni.get("asgari_tutar"),
                    "ekstre_donem": yeni.get("ekstre_donem"),
                    "denklestirme_cizgisi": yeni.get("denklestirme_cizgisi"),
                },
                # ESKİ yollar
                "eski": {
                    "kart_borc_defter": eski_defter_borc,
                    "aktif_donem_bu_ekstre": aktif.get("bu_ekstre"),
                    "aktif_donem_asgari": aktif.get("asgari_odeme"),
                    "snapshot_override_borc": round(ov_borc, 2) if ov_borc is not None else None,
                    "snapshot_override_asgari": round(ov_asgari, 2) if ov_asgari is not None else None,
                },
                "fark_defter_yolu": fark_defter,
                "defter_uyusuyor": abs(fark_defter) < 1.0,
            })

    uyusan = sum(1 for s in satirlar if s.get("defter_uyusuyor"))
    sapan = [s["kart_adi"] for s in satirlar if s.get("defter_uyusuyor") is False]
    return {
        "kart_adet": len(satirlar),
        "defter_uyusan": uyusan,
        "defter_sapan": sapan,
        "satirlar": satirlar,
        "not": "Salt okur karşılaştırma (ADIM 4/12). kanonik.anlik_borc EKRANDA "
               "gösterilmesi gereken sayıdır; kanonik.mutabakat_farki sıfırdan "
               "sapıyorsa defter ile banka arasında açıklanması gereken bir fark var "
               "— bu bilgidir, gizlenmez.",
    }


@app.get("/api/kartlar/borc-faiz-ozet")
def kart_borc_faiz_ozet():
    """Kart başına/toplam — borç (taksit dahil) + ödenen faiz + bu ay ekstre var mı.
    Borç/taksit/toplam TEK KAYNAK: kartlar_listele() (=/api/kartlar) → cep & masaüstü
    KartYönetimi AYNI rakamı gösterir (tutarsızlık olmaz). Faiz/son ekstre snapshot'tan."""
    from datetime import date as _date
    bugun = bugun_tr()
    # Borç + gelecek taksit + toplam: /api/kartlar ile birebir aynı kaynak
    kl = kartlar_listele()
    kdata = kl if isinstance(kl, list) else (kl.get("kartlar") or [])
    with db() as (conn, cur):
        cur.execute("""
            SELECT kart_id::text AS kart_id, COALESCE(SUM(donem_faizi),0)::float AS toplam_faiz,
                   MAX(donem)::text AS son_donem, COUNT(*)::int AS donem_adet
            FROM kart_ekstre_donem GROUP BY kart_id
        """)
        snap = {r["kart_id"]: dict(r) for r in (cur.fetchall() or [])}
    bu_ay = str(_date(bugun.year, bugun.month, 1))
    satirlar, toplam_borc, toplam_faiz, toplam_taksit, eksik = [], 0.0, 0.0, 0.0, []
    for k in kdata:
        kid = str(k.get("id"))
        anlik = float(k.get("anlik_borc") if k.get("anlik_borc") is not None else (k.get("guncel_borc") or 0))
        gt = float(k.get("gelecek_taksit_anapara") or 0)
        toplam_kart = float(k.get("toplam_borc_taksitli") if k.get("toplam_borc_taksitli") is not None else (anlik + gt))
        s = snap.get(kid, {})
        tf = float(s.get("toplam_faiz") or 0)
        son_donem = s.get("son_donem")
        bu_ay_var = bool(son_donem and son_donem[:7] == bu_ay[:7])
        toplam_borc += anlik; toplam_faiz += tf; toplam_taksit += gt
        if not bu_ay_var:
            eksik.append(k.get("kart_adi"))
        satirlar.append({
            "kart_id": kid, "kart_adi": k.get("kart_adi"), "banka": k.get("banka"),
            "sahip": k.get("sahip") or "İşletme", "limit": float(k.get("limit_tutar") or 0),
            "guncel_borc": round(anlik, 2), "toplam_odenen_faiz": round(tf, 2),
            "gelecek_taksit_anapara": round(gt, 2), "toplam_borc_taksitli": round(toplam_kart, 2),
            "son_ekstre_donem": son_donem, "ekstre_adet": int(s.get("donem_adet") or 0),
            "bu_ay_ekstre_var": bu_ay_var,
        })
    satirlar.sort(key=lambda x: -x["toplam_borc_taksitli"])
    return {
        "toplam_borc": round(toplam_borc, 2),                          # bu dönem borcu (anlık)
        "toplam_taksit": round(toplam_taksit, 2),                       # gelecek taksit anaparası
        "toplam_borc_taksitli": round(toplam_borc + toplam_taksit, 2),  # GERÇEK toplam (taksit dahil)
        "toplam_odenen_faiz": round(toplam_faiz, 2),
        "kart_adet": len(kdata),
        "bu_ay_eksik_ekstre": eksik,
        "kartlar": satirlar,
    }


@app.get("/api/kartlar/gelecek-ay-yuk")
def kart_gelecek_ay_yuk():
    """İZOLE/SALT-OKUR: gelecek ay tahmini ZORUNLU ödeme yükü + elde kalan strateji.
    Kullanıcı durumu: tüm nakit kredi/kartlara yatıyor → "gelecek ay en az ne ödemeliyim".
      1) Kart tahmini asgari (OTOMATİK): her kart için gelecek dönem borç tahmini ×
         asgari oran. Asgari oran TEK KAYNAKTAN gelir: kart_asgari_orani(k).
         Gelecek dönem = MAX( (anlık−bu ay asgari)×(1+AYLIK faiz) , gelecek ay ekstre planı )
         ⚠️ TOPLAM DEĞİL — odeme_plani'nın kartlı satırları ekstre yüklemesinde
         "dönem borcu ya da asgari" olarak açılır; yani anlık borcun İÇİNDEKİ
         aynı paradır. İkisini toplamak çift sayımdır (BORC-004, 2026-09-02).
      2) Kredi taksitleri (borc_envanteri, gelecek ay) — kesin.
      3) ZORUNLU YÜK = kart asgari + kredi taksiti (batmamak için minimum).
      4) Serbest nakit (kasa_bakiyesi) − zorunlu = ELDE KALAN → çığa (en pahalı borç).
      5) Ortalama aylık ödeme (son 90 gün kart ödemesi / 3) — tipik yük.
    Hiçbir veriyi DEĞİŞTİRMEZ; mevcut kartlar_listele + odeme_plani + borc_envanteri okur."""
    from datetime import date as _date
    from calendar import monthrange as _mr
    bugun = bugun_tr()
    ga_yil = bugun.year + (1 if bugun.month == 12 else 0)
    ga_ay = 1 if bugun.month == 12 else bugun.month + 1
    ga_bas = _date(ga_yil, ga_ay, 1)
    ga_bit = _date(ga_yil, ga_ay, _mr(ga_yil, ga_ay)[1])
    kl = kartlar_listele()
    kartlar = kl if isinstance(kl, list) else (kl.get("kartlar") or [])
    with db() as (conn, cur):
        serbest = float(kasa_bakiyesi(cur) or 0)
        cur.execute(
            """SELECT kart_id::text AS kid, COALESCE(SUM(odenecek_tutar),0)::float AS t
               FROM odeme_plani WHERE kart_id IS NOT NULL AND durum='bekliyor'
                 AND tarih BETWEEN %s AND %s GROUP BY kart_id""",
            (ga_bas, ga_bit),
        )
        kart_taksit = {r["kid"]: float(r["t"]) for r in (cur.fetchall() or [])}
        cur.execute(
            "SELECT COALESCE(SUM(aylik_taksit),0)::float AS t FROM borc_envanteri "
            "WHERE aktif=TRUE AND (kalan_vade IS NULL OR kalan_vade>0)"
        )
        kredi_taksit = float((cur.fetchone() or {}).get("t") or 0)
        try:
            cur.execute(
                "SELECT COALESCE(SUM(tutar),0)::float AS t FROM kart_hareketleri "
                "WHERE islem_turu='ODEME' AND durum='aktif' AND tarih >= CURRENT_DATE - 90"
            )
            son90 = float((cur.fetchone() or {}).get("t") or 0)
        except Exception:
            son90 = 0.0
    satir, t_asgari = [], 0.0
    for k in kartlar:
        anlik = float(k.get("anlik_borc") if k.get("anlik_borc") is not None else (k.get("guncel_borc") or 0))
        if anlik <= 0.5:
            continue
        bu_asg = float(k.get("asgari_odeme") or 0)
        # 🔴 BORC-001: kartlar.faiz_orani AKDİ (YILLIK) orandır — finans_core
        # ve borç koçu hep /100/12 ile aylığa çevirir. Burada /100 ile
        # doğrudan aylık sanılıyordu: 12 KAT şişkin faiz, üstelik çıktıya
        # "faiz_ay_pct" adıyla basılıyordu.
        faiz_ay = float(k.get("faiz_orani") or 0) / 100.0 / 12.0
        # 🔴 BORC-010: asgari oran limit eşiğiyle hardcode ediliyordu;
        # kartın kendi asgari_oran kolonu yok sayılıyordu. TEK KAYNAK:
        oran = kart_asgari_orani(k)
        kid = str(k.get("id"))
        kalan = max(0.0, anlik - bu_asg)
        faizli = kalan * (1 + faiz_ay)
        # 🔴 BORC-004: bu iki değer AYNI PARAYI ölçüyor. `taksit_dilim`,
        # ekstre yüklemesinde açılan kartlı odeme_plani satırıdır
        # (odenecek_tutar = dönem borcu ya da asgari) — yani zaten
        # anlik_borc'un içinde. Toplamak çift sayımdı. MAX alıyoruz:
        # plan GERÇEK bir ekstreden gelir, faizli ise TAHMİNDİR; hangisi
        # büyükse gelecek ayın yükü odur, ama ikisi ÜST ÜSTE binmez.
        taksit_dilim = kart_taksit.get(kid, 0.0)
        gelecek_donem = max(faizli, taksit_dilim)
        tahmini_asgari = gelecek_donem * oran
        t_asgari += tahmini_asgari
        satir.append({
            "kart_adi": k.get("kart_adi"), "anlik_borc": round(anlik, 2),
            "faiz_ay_pct": round(faiz_ay * 100, 4),
            "faiz_yil_pct": round(float(k.get("faiz_orani") or 0), 2),
            "asgari_oran_pct": round(oran * 100, 2),
            "gelecek_donem_tahmini": round(gelecek_donem, 2),
            # Kırılım görünür olsun: hangi bileşen kazandı, ekranda okunabilsin.
            "bilesen_faizli_tahmin": round(faizli, 2),
            "bilesen_ekstre_plani": round(taksit_dilim, 2),
            "bilesen_secilen": "ekstre_plani" if taksit_dilim > faizli else "faizli_tahmin",
            "tahmini_asgari": round(tahmini_asgari, 2),
        })
    satir.sort(key=lambda x: -x["tahmini_asgari"])
    zorunlu = t_asgari + kredi_taksit
    return {
        "gelecek_ay": f"{ga_yil}-{ga_ay:02d}",
        "kart_tahmini_asgari": round(t_asgari, 2),
        "kredi_taksiti": round(kredi_taksit, 2),
        "zorunlu_yuk": round(zorunlu, 2),
        "serbest_nakit": round(serbest, 2),
        "ekstra_kapasite": round(serbest - zorunlu, 2),   # +: çığa yatırılabilir, −: açık
        "ortalama_aylik_odeme": round(son90 / 3.0, 2),
        "kartlar": satir,
    }


# Şehir/jenerik kelimeler satıcı KİMLİĞİ değildir — anahtar olarak seçilirse
# hafıza alakasız harcamalara yayılır. Canlı vaka (2026-08-10): "KONYA SU"
# sınıflandırılınca anahtar "KONYA" öğrenildi ve "KONYA SELÇUKLU",
# "KONYA KENTPLAZA" gibi bambaşka harcamalara işletme önerisi çıkardı.
_SATICI_ATIL = {
    "KONYA", "ISTANBUL", "İSTANBUL", "IZMIR", "İZMİR", "ANKARA", "KARAMAN",
    "BURSA", "ANTALYA", "ADANA", "MERSIN", "MERSİN", "TR", "TUR", "TURKIYE",
    "FATURA", "ODEME", "ÖDEME", "OTOMATIK", "OTOMATİK", "TALIMAT", "TALİMAT",
    "SAN", "TIC", "TİC", "LTD", "STI", "ŞTI", "AS", "A.S", "ANONIM", "ANONİM",
    "SIRKETI", "ŞİRKETİ", "LIMITED", "LİMİTED", "MERKEZ", "SUBE", "ŞUBE",
}


def _satici_anahtar(aciklama: Optional[str]) -> Optional[str]:
    """Açıklamadan satıcı anahtarı (ilk ANLAMLI kelime) — hafıza eşleşmesi için.
    'METRO METRO GROSMARKET KOKONYA TR' → 'METRO'.
    Şehir/jenerik kelimeler atlanır; hepsi atılırsa ilk kelimeye düşülür
    (hiç anahtar üretmemektense zayıf anahtar üretmek yeğdir)."""
    import re as _re
    s = (aciklama or "").upper().strip()
    s = _re.sub(r"[^A-ZÇĞİÖŞÜ0-9 ]", " ", s)
    ilk = None
    for tok in s.split():
        if len(tok) < 3 or tok.isdigit():
            continue
        if ilk is None:
            ilk = tok
        if tok not in _SATICI_ATIL:
            return tok
    return ilk


def _ekstre_txn_map(t: dict) -> dict:
    """kart_analiz işlem dict → birleşik ekstre işlem formatı (tip/tutar/tarih/kategori)."""
    def _tr_kucuk(s: str) -> str:
        """Türkçe-duyarlı küçültme: İ→i, I→i (aksan birleşmesi olmadan).
        Python'un .lower()'ı 'İ' için 'i'+U+0307 üretir ve 'faiz' araması tutmaz."""
        return (str(s or "").replace("İ", "i").replace("I", "i")
                .replace("Ş", "ş").replace("Ğ", "ğ").replace("Ü", "ü")
                .replace("Ö", "ö").replace("Ç", "ç").lower())
    odeme = bool(t.get("odeme_mi"))
    kat = (t.get("kategori") or "")
    acik = (t.get("aciklama") or "")
    # 🔴 TÜRKÇE-İ TUZAĞI (2026-08-17 kart denetimi): "DÖNEM FAİZİ".lower() Python'da
    # 'faiz' ÜRETMEZ — büyük İ küçülünce 'i' + U+0307 (birleşen nokta) olur.
    # Eski kod bunu yalnız tam "DÖNEM FAİZİ" metniyle kurtarıyordu; BÜYÜK HARFLE
    # gelen "KREDİ FAİZİ · TAKSİT FAİZİ · GECİKME FAİZİ · LİMİT AŞIM FAİZİ ·
    # NAKİT AVANS FAİZİ" satırları HARCAMA sayılıyordu. Sonuç zinciri: faiz
    # işaretlenmeyince faiz_donemleri (aşağıda ~5350) boş kalır → motorun TAHMİNİ
    # faizi iptal edilmez → aynı dönemde ÇİFT FAİZ. Çözüm: aksan-duyarsız kat.
    _fk = _tr_kucuk(f"{kat} {acik}")
    faiz = "faiz" in _fk
    tip = "ODEME" if odeme else ("FAIZ" if faiz else "HARCAMA")
    tks = t.get("taksit")
    tsay = None
    if tks and "/" in str(tks):
        try:
            tsay = int(str(tks).split("/")[1])
        except (ValueError, IndexError):
            tsay = None
    return {
        "tarih": t.get("tarih"),
        "tutar": abs(float(t.get("tutar") or 0)),
        "tip": tip,
        "aciklama": acik,
        "kategori": kat or None,
        "taksit": tks,
        "taksit_anapara": t.get("taksit_anapara"),
        "taksit_sayisi": tsay,
    }


@app.post("/api/kartlar/ekstre-yukle")
def kart_ekstre_yukle(dosya: UploadFile = File(...)):
    """Faz E0: Banka kredi kartı ekstresi (PDF) yükle → ayrıştır → mutabakat ÖNİZLEME.

    ⚠️ "Hiçbir şey yazmaz" DEĞİL (docstring 2026-08-10'a kadar öyle diyordu, yanıltıcıydı):
    İŞLEM SATIRLARI yazılmaz — onlar sahibin şahsi/işletme sınıflandırmasından sonra
    /api/kartlar/ekstre-import ile gider. Ama kart son 4 haneden EŞLEŞİRSE ekstre
    ÖZETİ yazılır (_ekstre_eslesme_mutabakat): dönem borcu · asgari · faiz · taksit
    yükü snapshot'ı ve son ödeme planı güncellenir. Kart eşleşmezse hiçbir yazma olmaz.

    Worldcard + Enpara + Axess + Garanti + Ziraat desteklenir.
    Sync def: FastAPI threadpool'da çalışır, pdfplumber event-loop'u bloklamaz."""
    import io
    try:
        import pdfplumber
    except Exception:
        raise HTTPException(500, "pdfplumber yüklü değil (sunucu).")
    raw = dosya.file.read()
    if not raw:
        raise HTTPException(400, "Dosya boş veya yüklenemedi.")

    # ── AXESS/AKBANK: gömülü-font EBCDIC PDF → özel parser (OCR'a gerek yok).
    #    Diğer bankalardan ÖNCE dene; pdfplumber bunu çöp olarak okur.
    try:
        from ekstre_parser import is_axess, parse_axess
        if is_axess(raw):
            sonuc = parse_axess(raw)
            sonuc = _ekstre_eslesme_mutabakat(sonuc, raw)
            return _ekstre_pdf_arsivle(sonuc, raw, dosya.filename or "ekstre.pdf")
    except HTTPException:
        raise
    except Exception:
        pass  # fitz yok / parse hatası → normal akışa düş

    metin = ""
    try:
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            for pg in pdf.pages:
                metin += (pg.extract_text() or "") + "\n"
    except Exception as e:
        raise HTTPException(400, f"PDF okunamadı: {e}")
    if len(metin.strip()) < 40:
        raise HTTPException(400, "PDF'den metin çıkmadı — taranmış/görüntü ekstre olabilir (Axess gibi → OCR gerekir).")

    # BİRLEŞİK PARSER: başlık (borç/asgari/faiz/son4) = ekstre_parser;
    # işlemler = kart_analiz (4 banka + kategori) — tek transaction motoru.
    from ekstre_parser import parse_ekstre
    sonuc = parse_ekstre(metin)
    # ekstre_parser TAKSİT çıkarır (kart_analiz çıkarmaz) → taksit bilgisini sakla
    _taksit_lk = {}
    for _e in sonuc.get("islemler", []):
        if _e.get("taksit") and _e.get("tip") == "HARCAMA":
            _taksit_lk[(_e.get("tarih"), round(float(_e.get("tutar") or 0), 2))] = (
                _e.get("taksit"), _e.get("taksit_anapara"))
    try:
        import kart_analiz
        txns = kart_analiz.parse_pdf(raw)
    except Exception:
        txns = None
    if txns:
        sonuc["islemler"] = [_ekstre_txn_map(t) for t in txns]
        # kart_analiz (kategori+taksit) + ekstre_parser (yedek taksit) BİRLEŞTİR.
        # ⚠️ ÖNCELİK kart_analiz'dedir (2026-08-10): taksit satırını İŞLEMİN HEMEN
        # ALTINDAKİ satırdan okur. ekstre_parser (tarih,tutar) anahtarıyla eşleştirdiği
        # için satır kayması yapıyordu — canlıda KONYA KENTPLAZA'ya "1/2" (gerçeği 3/3,
        # son taksit) ve taksitsiz SOYTÜRKLER'e komşusunun "1/5"ini yazmıştı.
        # Yanlış taksit no'su = bitmiş taksidin yeniden borç yazılması demek.
        # YETKİ SINIRI: kart_analiz parser'ı bu ekstrede taksit ÜRETEBİLİYORSA
        # (en az bir satırda taksit var) o parser taksit-yetkilidir; yedeğe HİÇ
        # düşülmez. Aksi halde yedek, taksitsiz bir alıma komşusunun taksidini
        # yapıştırıyordu — canlıda tek çekim SOYTÜRKLER 50.000 ₺'ye KARABULUT'un
        # "1/5 · toplam 100.000" bilgisi bağlanmıştı (alım iki katına çıkardı).
        _ka_taksit_yetkili = any(i.get("taksit") for i in sonuc["islemler"])
        for _isl in sonuc["islemler"]:
            if _isl.get("tip") == "HARCAMA" and not _isl.get("taksit") and not _ka_taksit_yetkili:
                _info = _taksit_lk.get((_isl.get("tarih"), round(float(_isl.get("tutar") or 0), 2)))
                if _info:
                    _isl["taksit"] = _info[0]
                    _isl["taksit_anapara"] = _info[1]
                    try:
                        _isl["taksit_sayisi"] = int(str(_info[0]).split("/")[1])
                    except (ValueError, IndexError):
                        pass
        if not sonuc.get("banka_format") or sonuc.get("banka_format") == "bilinmiyor":
            sonuc["banka_format"] = (txns[0].get("banka") if txns else None) or sonuc.get("banka_format")
        # Kart sahibi header'dan gelmediyse kart_analiz'den al + temizle (Garanti vb.)
        if not (sonuc.get("kart_sahibi") or "").strip():
            import re as _re
            for _t in txns:
                _s = (_t.get("kart_sahibi") or "").strip()
                if not _s:
                    continue
                _s = _s.split("\n")[0].strip()
                _s = _re.sub(r"(?i)\s*(müşteri|musteri|kart)\s*limiti.*$", "", _s).strip()
                if _s:
                    sonuc["kart_sahibi"] = _s.title()
                    break
        sonuc.pop("hata", None)  # işlem bulunduysa parse başarısız sayma
    if sonuc.get("hata") and not txns:
        raise HTTPException(422, sonuc["hata"])

    # GARANTİ/BONUS: özet kutusundaki TARİHSİZ faiz/ücret satırları (DÖNEM FAİZİ vb.)
    # kart_analiz tarihli satır beklediği için kaçıyor → FAIZ işlemi olarak enjekte et.
    # (kart_analiz override'ından SONRA çağrılır ki işlemler silinmesin.)
    if sonuc.get("banka_format") == "garanti":
        try:
            from ekstre_parser import garanti_faiz_enjekte
            garanti_faiz_enjekte(sonuc, metin)
        except Exception:
            pass  # faiz enjeksiyonu başarısız olsa da ekstre okuması devam etsin
    # ZİRAAT/Bankkart: faiz satırları tarihli (kart_analiz FAIZ olarak yakalar) → burada
    # sadece dönem faizi özetini + yıllık oranları tamamla.
    elif sonuc.get("banka_format") == "ziraat":
        try:
            from ekstre_parser import ziraat_faiz_finalize
            ziraat_faiz_finalize(sonuc, metin)
        except Exception:
            pass

    sonuc = _ekstre_eslesme_mutabakat(sonuc, raw)
    return _ekstre_pdf_arsivle(sonuc, raw, dosya.filename or "ekstre.pdf")


def _ekstre_belge_kolonlari(cur):
    """kart_ekstre_donem'e PDF arşiv kolonları (idempotent DDL) — İz & Belge
    doktrini: rakamlar yazılıyordu ama BELGENİN ASLI saklanmıyordu; 'Ağustos
    ekstresini göster' dendiğinde sistem gösteremiyordu (2026-08-17, sahip)."""
    cur.execute("ALTER TABLE kart_ekstre_donem ADD COLUMN IF NOT EXISTS belge_pdf BYTEA")
    cur.execute("ALTER TABLE kart_ekstre_donem ADD COLUMN IF NOT EXISTS belge_ad TEXT")
    cur.execute("ALTER TABLE kart_ekstre_donem ADD COLUMN IF NOT EXISTS belge_ts TIMESTAMPTZ")


def _ekstre_pdf_arsivle(sonuc, raw: bytes, dosya_ad: str):
    """Yüklenen ekstre PDF'inin ASLINI dönem kaydına iliştirir (kart+dönem tekil).
    Kart eşleşmediyse ya da dönem yazılamadıysa sessizce False döner — arşiv
    başarısızlığı yükleme/mutabakat sonucunu ASLA düşürmez (ana iş rakamlar)."""
    kart = (sonuc or {}).get("eslesen_kart") or {}
    kesim = (sonuc or {}).get("kesim_tarihi")
    if not kart.get("id") or not kesim or not raw:
        sonuc["belge_arsivlendi"] = False
        return sonuc
    try:
        from psycopg2 import Binary as _PgBinary  # main.py'de modül-düzeyi psycopg2 yok
        with db() as (conn, cur):
            _ekstre_belge_kolonlari(cur)
            cur.execute(
                "UPDATE kart_ekstre_donem SET belge_pdf=%s, belge_ad=%s, belge_ts=NOW() "
                "WHERE kart_id=%s AND donem=DATE_TRUNC('month', %s::date)",
                (_PgBinary(raw), (dosya_ad or "ekstre.pdf")[:200], kart["id"], kesim),
            )
            sonuc["belge_arsivlendi"] = cur.rowcount > 0
    except Exception as _e:
        sonuc["belge_arsivlendi"] = False
        sonuc["belge_arsiv_hata"] = str(_e)[:120]
    return sonuc


@app.get("/api/kartlar/ekstre-belge")
def kart_ekstre_belge(kart_id: str, donem: str):
    """Arşivlenmiş ekstre PDF'inin aslını döner (kart + dönem YYYY-MM)."""
    d = (donem or "").strip()[:10]
    if len(d) == 7:
        d += "-01"
    with db() as (conn, cur):
        _ekstre_belge_kolonlari(cur)
        cur.execute(
            "SELECT belge_pdf, belge_ad FROM kart_ekstre_donem "
            "WHERE kart_id=%s AND donem=DATE_TRUNC('month', %s::date)",
            (kart_id, d),
        )
        r = cur.fetchone()
        if not r or not dict(r).get("belge_pdf"):
            raise HTTPException(404, "Bu dönem için arşivlenmiş ekstre PDF'i yok — PDF yüklendiğinde otomatik arşivlenir")
        r = dict(r)
        from fastapi.responses import Response as _Resp
        ad = (r.get("belge_ad") or "ekstre.pdf").replace('"', "")
        return _Resp(content=bytes(r["belge_pdf"]), media_type="application/pdf",
                     headers={"Content-Disposition": f'inline; filename="{ad}"'})


def _ekstre_eslesme_mutabakat(sonuc, raw: Optional[bytes] = None):
    """Ekstre sonucu → kart eşleştir (son 4 hane) + mutabakat + faiz/snapshot/CFO yaz.
    Hem normal (Worldcard/Enpara) hem Axess akışının ortak son adımı.

    raw: PDF ham baytları — GEOMETRİK İKİNCİ GÖZ için (2026-08-18). Verilmezse
    ikinci göz atlanır ve bu durum yanıtta GEREKÇESİYLE görünür (sessiz atlama
    yok). Eskiden bu fonksiyon `raw`u hiç almıyordu; geometrik blok onu
    kullanmaya çalışınca NameError veriyor ve hata-yutar sayesinde sessizce
    'hata' yazıyordu — canlıda 5 bankada da öyle görüldü."""
    sonuc["eslesen_kart"] = None
    sonuc["mutabakat"] = None

    # ── OKUMA DENETİMİ (tie-out) — 2026-08-10 ─────────────────────────────────
    # Sessiz yanlış ayrıştırma en tehlikeli hatadır: ekran dolu görünür, rakam
    # yanlıştır. Canlı vaka: Worldcard ayrıştırıcısı PUAN sütununu tutar sanıyordu
    # (TRENDYOL 719,99 → 14,00; KARABULUT 20.000 → 4,00) ve puan tablosundan iki
    # sahte harcama üretiyordu. Kimse fark etmeden kart borcuna yazılabilirdi.
    #
    # Ekstre kendi doğrulamasını taşır: başlıktaki "Dönem İçi Harcamalar" =
    # okunan harcama + faiz olmalı. Fark ölçülür ve YANITTA TAŞINIR — ekran
    # tutmuyorsa uyarır, tutuyorsa "okuma doğrulandı" der. Sayı uydurulmaz.
    try:
        _oku_h = sum(abs(float(i.get("tutar") or 0)) for i in (sonuc.get("islemler") or [])
                     if (i.get("tip") or "").upper() in ("HARCAMA", "FAIZ"))
        _oku_o = sum(abs(float(i.get("tutar") or 0)) for i in (sonuc.get("islemler") or [])
                     if (i.get("tip") or "").upper() == "ODEME")
        _bek_h = sonuc.get("donem_harcama")
        _bek_o = sonuc.get("donem_odeme")
        # ⚖️ BAĞIMSIZ KIYAS (2026-08-17, Axess vakası): donem_harcama bazı
        # parserlarda SATIRLARDAN türetilir (Axess) — satırı satırla kıyaslamak
        # totolojiydi ve faiz dahil/haric farkı yüzünden Axess HEP "tutmuyor"
        # diyordu. Başlıkta önceki borç + dönem borcu varsa muhasebe kimliği
        # kullanılır: önceki − ödeme + harcama + faiz = dönem borcu. İki ucu
        # da başlıktan (regex) gelir, satırlardan bağımsızdır — gerçek tie-out.
        _borc_b = sonuc.get("donem_borcu")
        _onceki_b = sonuc.get("onceki_borc")
        _kiyas_kaynak = "donem_harcama"
        if _borc_b is not None and _onceki_b is not None:
            _bek_h = float(_borc_b) - float(_onceki_b) + _oku_o
            _kiyas_kaynak = "borc_kimligi"
        _dn = {"okunan_harcama": round(_oku_h, 2), "okunan_odeme": round(_oku_o, 2),
               "islem_adet": len(sonuc.get("islemler") or []),
               "kiyas_kaynak": _kiyas_kaynak}
        if _bek_h is not None:
            _fark = round(_oku_h - float(_bek_h), 2)
            _dn.update({"beklenen_harcama": round(float(_bek_h), 2), "harcama_farki": _fark,
                        "harcama_tutuyor": abs(_fark) <= 1.0})
        if _bek_o is not None:
            _farko = round(_oku_o - float(_bek_o), 2)
            _dn.update({"beklenen_odeme": round(float(_bek_o), 2), "odeme_farki": _farko,
                        "odeme_tutuyor": abs(_farko) <= 1.0})
        _hepsi = [v for k, v in _dn.items() if k.endswith("_tutuyor")]
        _dn["saglam"] = bool(_hepsi) and all(_hepsi)
        _dn["mesaj"] = ("Okuma doğrulandı — ekstrenin kendi toplamıyla birebir."
                        if _dn["saglam"] else
                        ("Ekstrenin kendi toplamı ile okunan satırlar TUTMUYOR — "
                         "içe aktarmadan önce satırları gözden geçirin."
                         if _hepsi else
                         "Ekstre başlığında kıyas toplamı yok — okuma doğrulanamadı."))
        sonuc["okuma_denetimi"] = _dn

        # ── 📐 GEOMETRİK OKUYUCU — ÇAPA HAKEMLİĞİNDE İKİNCİ GÖZ (Adım 5) ──
        # Metin okuyucusu tutarı "satırın son sayısı" tahminiyle bulur ve bu
        # tahmin 2026-08-10..17 arasında ÜÇ canlı kusur üretti (Worldcard puan,
        # Garanti bonus sütunu, 'bosluk' hücresi = 42.193 ₺ kayıp). Kökü tek:
        # TUTARIN YERİ YAPISAL DEĞİL. Kanıt — aynı kavram iki bankada TERS
        # yönde durur: Garanti'de puan sütunu tutarın SOLUNDA (x≈445), Worldcard'da
        # SAĞINDA (x≈567). "Son sayıyı al" ikisinde aynı anda doğru olamaz.
        # ekstre_geometri tutarı SÜTUN KOORDİNATINDAN okur.
        #
        # ⚖️ AMA SESSİZCE DEVRALMAZ. Hangi okuyucunun kazandığına ÇAPA karar
        # verir: ekstrenin kendi başlık toplamına HANGİSİ DAHA YAKINSA o. Böylece
        # yeni okuyucu bir bankada kötüyse eskisini bozmaz — "sahte yeşil yasak"
        # doktrininin okuyucu tarafındaki karşılığı.
        # Şimdilik yalnız RAPORLAR (islemler'i DEĞİŞTİRMEZ): önce canlı ekstrelerde
        # kim kazanıyor görülecek, sonra hakem karar verici yapılacak.
        try:
            if raw is None:
                raise RuntimeError('PDF ham baytları bu akışta taşınmadı')
            import ekstre_geometri as _geo
            _g = _geo.geometrik_oku(raw)
            if _g.get("basarili"):
                _gs = _g["satirlar"]
                _gor = set(); _gded = []
                for _x in _gs:   # aynı tablo bazı PDF'lerde 2. sayfada TEKRAR basılır
                    _k = (_x["tarih"], _x["aciklama"][:30], round(_x["tutar"], 2), _x["odeme_mi"])
                    if _k in _gor:
                        continue
                    _gor.add(_k); _gded.append(_x)
                # 🪤 "MÜKERRER" SANILAN MEŞRU TEKRAR (2026-08-18, Enpara vakası):
                # Tekilleştirme anahtarı (tarih, açıklama, tutar) — ama AYNI GÜN
                # AYNI MARKETTEN İKİ KEZ alışveriş MEŞRUDUR ve bu anahtar onu
                # yutar. Canlı kanıt: ŞOK-KARAMAN YAVUZ SULTAN 550,00 ₺ aynı gün
                # iki kez; çapa tam −550,00 fark verdi. Ziraat'te ise tablo
                # gerçekten 2. sayfada TEKRAR basılıyor (78→39) ve orada
                # tekilleştirme ŞART.
                # Hangisi doğru? SAYIYA DEĞİL ÇAPAYA sor: her iki kümeyi de
                # ölç, ekstrenin kendi toplamına yakın olanı kullan.
                if _bek_h is not None and len(_gded) != len(_gs):
                    _h_ded = sum(x["tutar"] for x in _gded if not x.get("odeme_mi"))
                    _h_ham = sum(x["tutar"] for x in _gs if not x.get("odeme_mi"))
                    _o_ded = sum(x["tutar"] for x in _gded if x.get("odeme_mi"))
                    _o_ham = sum(x["tutar"] for x in _gs if x.get("odeme_mi"))
                    # beklenen harcama ödemeye bağlı (borç kimliği) → her küme
                    # kendi ödemesiyle kıyaslanır
                    _bek_ded = (float(_borc_b) - float(_onceki_b) + _o_ded
                                if (_borc_b is not None and _onceki_b is not None) else float(_bek_h))
                    _bek_ham = (float(_borc_b) - float(_onceki_b) + _o_ham
                                if (_borc_b is not None and _onceki_b is not None) else float(_bek_h))
                    _gtek = _gs if abs(_h_ham - _bek_ham) < abs(_h_ded - _bek_ded) - 0.01 else _gded
                else:
                    _gtek = _gded
                _gh = round(sum(x["tutar"] for x in _gtek if not x.get("odeme_mi")), 2)
                _go = round(sum(x["tutar"] for x in _gtek if x.get("odeme_mi")), 2)
                # ⚖️ ADİL KIYAS — ELMA İLE ELMA (2026-08-18 canlı ölçüm dersi):
                # Ekstrede faiz kalemleri (DÖNEM FAİZİ, KKDF+BSMV, GEÇ ÖDEME
                # FAİZİ, LİMİT AŞIM FAİZİ) TARİHSİZ satırlardır — tablonun içinde
                # değil, başlık bloğunda dururlar. Metin yolu bunları ayrıca
                # enjekte eder (garanti_faiz_enjekte / ziraat_faiz_finalize),
                # geometrik okuyucu ise tarihsiz satırı işlem saymaz.
                # Bu yüzden ham kıyas geometriği HAKSIZ YERE kaybettiriyordu:
                #   Garanti: metin 180.606,82 · geo 159.627,36 · fark tam 20.979,46
                #   = DÖNEM FAİZİ 15.722,62 + KKDF/BSMV 4.841,41 + GEÇ ÖDEME
                #     287,93 + LİMİT AŞIM 127,50  → kuruşu kuruşuna faiz kalemleri
                # Yani geometrik satırları DOĞRU okuyordu, sadece faiz onun işi
                # değildi. Kıyasa aynı faizi ekliyoruz; okuyucular yalnız
                # SATIR OKUMA becerisiyle yarışsın.
                # 🔬 DÜZELTME (2026-08-18, canlı ölçümle yakalandı): metin
                # yolundaki TÜM faizi eklemek AŞIRI DÜZELTMEYDİ. Bankalar faizi
                # farklı yerde tutuyor:
                #   Worldcard/Ziraat/Enpara → faiz TARİHLİ tablo satırı; geometrik
                #     onu ZATEN okuyor. Eklemek çift sayım oldu (World 0,00 → +14.087)
                #   Garanti → faiz TARİHSİZ başlık satırı; geometrik okumuyor.
                # Doğru ölçüt: yalnız GEOMETRİĞİN KENDİ tarihsiz satırlarındaki
                # faizi ekle. Böylece her banka kendi yapısına göre adil kıyaslanır
                # ve düzeltme kendi kendini sınırlar.
                # ⛔ DEVİR satırı faiz DEĞİLDİR — "önceki dönemden devir edilen
                #    tutar" eklenirse harcama borcun tamamı kadar şişerdi.
                import re as _re_f
                _faiz_dsn = _re_f.compile(
                    r"faiz|bsmv|kkdf|ücret|ucret|aidat|gecikme|limit\s*aşım|limit\s*asim", _re_f.I)
                _devir_dsn = _re_f.compile(r"devir|önceki\s*dönem|onceki\s*donem|geçen\s*dönem", _re_f.I)
                _enjekte_faiz = round(sum(
                    abs(float(t.get("tutar") or 0))
                    for t in (_g.get("tarihsiz_tutarli") or [])
                    if _faiz_dsn.search(t.get("aciklama") or "")
                    and not _devir_dsn.search(t.get("aciklama") or "")), 2)
                _gh_kiyas = round(_gh + _enjekte_faiz, 2)
                _gfark = round(_gh_kiyas - float(_bek_h), 2) if _bek_h is not None else None
                _metin_fark = _dn.get("harcama_farki")
                _kazanan = None
                if _gfark is not None and _metin_fark is not None:
                    _kazanan = "geometrik" if abs(_gfark) < abs(_metin_fark) - 0.01 else (
                        "metin" if abs(_metin_fark) < abs(_gfark) - 0.01 else "berabere")
                sonuc["geometrik_okuma"] = {
                    "satir": len(_gtek), "ham_satir": len(_gs),
                    "harcama": _gh, "odeme": _go,
                    "enjekte_faiz": _enjekte_faiz,     # başlıktan gelen, satırda yok
                    "kiyas_harcama": _gh_kiyas,        # satır + faiz = adil kıyas
                    "harcama_farki": _gfark,
                    "tutarsiz_satir": len(_g.get("tutarsiz") or []),
                    "kazanan": _kazanan,
                    "not": "Çapa hakemliğinde ikinci göz. Kıyasa yalnız geometriğin "
                           "KENDİ tarihsiz faizi eklenir — faizi tabloda tutan bankada "
                           "eklenecek bir şey bulunmaz, düzeltme kendini sınırlar.",
                }

                # ── ⚖️ HAKEM KARAR VERİYOR (sahip onayı 2026-08-18: "EVET") ──
                # Geometrik okuyucu ÇAPAYA DAHA YAKINSA satırlar ondan alınır.
                # Canlı ölçüm (5 banka) bu kararı destekliyor:
                #   Garanti 0,00 / 0,00 · Worldcard 0,00 / 0,00 · Enpara −550 / −550
                #   ZİRAAT −91.030,16 / 0,00  ← metin 91 K KAÇIRIYOR
                #   Axess: geometrik okuyamıyor (CID fontu) → devreye hiç girmez
                #
                # ⚠️ FAİZ KORUNUR: bu noktada sonuc["islemler"] içindeki FAIZ
                # satırları BAŞLIKTAN enjekte edilmiştir (garanti_faiz_enjekte /
                # ziraat_faiz_finalize) ve geometrik okuyucuda karşılığı YOKTUR.
                # Ham takas yapılsaydı Garanti'de 20.979,46 ₺ faiz SESSİZCE
                # düşerdi. Bu yüzden geometrik satırlar + mevcut FAİZ satırları
                # birleştirilir.
                # ⚠️ DENETİM YENİDEN HESAPLANIR: ekranda görünen çapa, GERÇEKTEN
                # içe aktarılacak satırların çapası olmalı — yoksa "yeşil" başka
                # bir kümeyi doğrular (sahte yeşil).
                if _kazanan == "geometrik":
                    # 🪤 ÇİFT FAİZ TUZAĞI (2026-08-18, takas yolunda TEKRAR düştüm
                    # ve canlı ölçümle yakaladım): metin yolunun FAİZ satırlarını
                    # koşulsuz eklemek, faizi TABLODA tutan bankada çift sayım
                    # üretiyor. Ziraat'te tam bu oldu: geo 107.749,59 (faiz dahil,
                    # tarihli satırlar) + enjekte 8.878,24 = 116.627,83 → çapa
                    # 0,00'dan +8.878,24'e kaydı.
                    # KURAL: metin faizini YALNIZCA geometriğin kendi TARİHSİZ
                    # faizi varsa ekle (_enjekte_faiz > 0). Sıfırsa geometrik
                    # faizi zaten satır olarak okumuştur — _ekstre_txn_map onu
                    # açıklamasından FAIZ tipine çevirir.
                    _faiz_satir = ([i for i in (sonuc.get("islemler") or [])
                                    if (i.get("tip") or "").upper() == "FAIZ"]
                                   if _enjekte_faiz > 0.009 else [])
                    _yeni = [_ekstre_txn_map({
                        "tarih": x["tarih"], "tutar": x["tutar"],
                        "aciklama": x["aciklama"], "kategori": None,
                        "odeme_mi": x.get("odeme_mi"),
                        "taksit": (f"{x['taksit_no']}/{x['taksit_sayisi']}"
                                   if x.get("taksit_no") and x.get("taksit_sayisi") else None),
                        "taksit_anapara": x.get("taksit_toplam"),
                    }) for x in _gtek]
                    sonuc["islemler"] = _yeni + _faiz_satir
                    # Çapayı YENİ kümeyle yeniden kur
                    _oku_h2 = sum(abs(float(i.get("tutar") or 0)) for i in sonuc["islemler"]
                                  if (i.get("tip") or "").upper() in ("HARCAMA", "FAIZ"))
                    _oku_o2 = sum(abs(float(i.get("tutar") or 0)) for i in sonuc["islemler"]
                                  if (i.get("tip") or "").upper() == "ODEME")
                    _bek_h2 = (float(_borc_b) - float(_onceki_b) + _oku_o2
                               if (_borc_b is not None and _onceki_b is not None) else _bek_h)
                    _f2 = round(_oku_h2 - float(_bek_h2), 2) if _bek_h2 is not None else None
                    _dn.update({
                        "okunan_harcama": round(_oku_h2, 2), "okunan_odeme": round(_oku_o2, 2),
                        "islem_adet": len(sonuc["islemler"]),
                        "okuyucu": "geometrik",
                        "okuyucu_gerekce": "Çapaya metin okuyucusundan DAHA YAKIN — "
                                           "satırlar sütun koordinatından okundu.",
                    })
                    if _bek_h2 is not None:
                        _dn.update({"beklenen_harcama": round(float(_bek_h2), 2),
                                    "harcama_farki": _f2, "harcama_tutuyor": abs(_f2) <= 1.0})
                    _h2 = [v for k, v in _dn.items() if k.endswith("_tutuyor")]
                    _dn["saglam"] = bool(_h2) and all(_h2)
                    _dn["mesaj"] = ("Okuma doğrulandı (geometrik okuyucu) — ekstrenin kendi "
                                    "toplamıyla birebir." if _dn["saglam"] else
                                    "Geometrik okuyucu seçildi ama çapa hâlâ tutmuyor — "
                                    "içe aktarmadan önce satırları gözden geçirin.")
                    sonuc["okuma_denetimi"] = _dn
                else:
                    _dn["okuyucu"] = "metin"

                # ── 📅 ADIM 8: TAKSİT DİLİMLERİ ─────────────────────────────
                # Taksit bilgisi PRATİKTE YALNIZ geometrik okuyucuda var:
                # ekstre_parser Garanti'de 0 taksit buluyor, geometrik 5 buluyor
                # (ESER 3/4 · TRENDYOL 1/3 · MARTI 1/2 · BİZİM MERMER 1/3 ·
                # U.S. POLO 5/6). Hangi okuyucu kazanırsa kazansın taksit bilgisi
                # geometrikten alınır — tutarı kim okursa okusun, taksit sütunu
                # geometrik okuyucunun gördüğü yerdir.
                _tks = {}
                for _x in _gtek:
                    if _x.get("taksit_sayisi"):
                        _tks[(_x["tarih"], round(_x["tutar"], 2))] = _x
                _kalan_yuk = 0.0
                _plan = []
                for _i in (sonuc.get("islemler") or []):
                    _m = _tks.get((_i.get("tarih"), round(abs(float(_i.get("tutar") or 0)), 2)))
                    if not _m:
                        continue
                    _no, _adet = _m.get("taksit_no"), _m.get("taksit_sayisi")
                    _i["taksit"] = f"{_no}/{_adet}" if _no else f"?/{_adet}"
                    _i["taksit_sayisi"] = _adet
                    _i["taksit_anapara"] = _m.get("taksit_toplam")
                    # ⚠️ BİTMİŞ TAKSİT YENİDEN BORÇ YAZILMAZ: 3/4 demek "4
                    # taksidin 3.'südür" — geriye YALNIZ 1 taksit kalmıştır.
                    # Toplam tutarı yeniden yazmak alımı baştan borçlandırır;
                    # bu, kart defterinde daha önce yaşanmış bir kusurdur.
                    _kalan_adet = max(0, int(_adet) - int(_no)) if _no else 0
                    _dilim = abs(float(_i.get("tutar") or 0))
                    _kalan_yuk += _kalan_adet * _dilim
                    _plan.append({
                        "tarih": _i.get("tarih"), "satici": _i.get("aciklama"),
                        "dilim_tutari": round(_dilim, 2),
                        "taksit_no": _no, "taksit_sayisi": _adet,
                        "kalan_taksit": _kalan_adet,
                        "kalan_tutar": round(_kalan_adet * _dilim, 2),
                        "alim_toplami": _m.get("taksit_toplam"),
                    })
                if _plan:
                    sonuc["taksit_plani"] = {
                        "satir": len(_plan),
                        "gelecek_taksit_yuku": round(_kalan_yuk, 2),
                        "dilimler": sorted(_plan, key=lambda p: -p["kalan_tutar"]),
                        "not": "Bu dönemin dilimi ekstrede zaten borçtur; buradaki "
                               "«kalan» GELECEK dönemlerde çıkacak yüktür. "
                               "Bitmiş taksitler (n/n) kalan üretmez — alımı baştan "
                               "borçlandırmak kart defterinde yaşanmış bir kusurdur.",
                    }
            else:
                sonuc["geometrik_okuma"] = {"basarili": False, "neden": _g.get("neden")}
        except Exception as _ge:  # noqa: BLE001 — ikinci göz ASLA yüklemeyi kilitlemez
            sonuc["geometrik_okuma"] = {"basarili": False, "neden": f"hata: {str(_ge)[:120]}"}
    except Exception as _e:
        sonuc["okuma_denetimi"] = {"saglam": None, "mesaj": f"Okuma denetimi çalışmadı: {_e}"}

    son4 = sonuc.get("son_dort")
    with db() as (conn, cur):
        # SATICI HAFIZASI: her işleme öneri tipi (hepsi 'belirsiz' başlar, hafıza öğrendikçe önerir)
        try:
            cur.execute("SELECT anahtar, harcama_tipi FROM kart_satici_kural")
            # ⚠️ ŞEHİR/JENERİK ANAHTAR KULLANILMAZ (2026-08-10): geçmişte
            # "KONYA SU" sınıflandırılırken anahtar "KONYA" öğrenilmişti ve
            # "KONYA SELÇUKLU", "KONYA KENTPLAZA" gibi bambaşka harcamalara
            # öneri yayıyordu. Yazma tarafı düzeltildi; okuma tarafında da
            # eski kirli kayıtlar SESSİZCE göz ardı edilir (veri temizliği
            # beklemeden koruma sağlar).
            _kurallar = {r["anahtar"]: r["harcama_tipi"]
                         for r in (cur.fetchall() or [])
                         if str(r["anahtar"]).upper() not in _SATICI_ATIL}
        except Exception:
            _kurallar = {}
        for _isl in sonuc.get("islemler", []):
            if _isl.get("tip") == "HARCAMA":
                _ak = _satici_anahtar(_isl.get("aciklama"))
                _isl["oneri_tipi"] = _kurallar.get(_ak) if _ak else None
        kart = None
        if son4:
            cur.execute(
                "SELECT id::text, kart_adi, banka, COALESCE(sahip,'İşletme') AS sahip, son_dort_hane "
                "FROM kartlar WHERE son_dort_hane=%s AND aktif=TRUE LIMIT 1",
                (son4,),
            )
            kart = cur.fetchone()
        if kart:
            kart = dict(kart)
            # ── ÇİFT YÜKLEME KORUMASI: bu kart için bu ayın ekstresi zaten var mı?
            #    (işlemler zaten idempotent yazılır; bu sadece kullanıcıyı net uyarır)
            kt_chk = sonuc.get("kesim_tarihi")
            if kt_chk:
                try:
                    cur.execute(
                        "SELECT donem_borcu, kaynak FROM kart_ekstre_donem "
                        "WHERE kart_id=%s AND donem=DATE_TRUNC('month', %s::date)",
                        (kart["id"], kt_chk),
                    )
                    _prev = cur.fetchone()
                    if _prev:
                        _prev = dict(_prev)
                        sonuc["donem_zaten_yuklendi"] = {
                            "donem": kt_chk[:7],
                            "onceki_borc": float(_prev.get("donem_borcu") or 0),
                            "kaynak": _prev.get("kaynak"),
                        }
                except Exception:
                    pass
            sistem_borc = kart_borc(cur, kart["id"])
            ekstre_borc = sonuc.get("donem_borcu") or 0
            sonuc["eslesen_kart"] = {
                "id": kart["id"], "kart_adi": kart["kart_adi"],
                "banka": kart["banka"], "sahip": kart["sahip"],
            }
            sonuc["mutabakat"] = {
                "sistem_borc": round(sistem_borc, 2),
                "ekstre_borc": round(ekstre_borc, 2),
                "fark": round(ekstre_borc - sistem_borc, 2),
                "tutar_uyumlu": abs(ekstre_borc - sistem_borc) < 1.0,
            }
            # İşlem-bazlı mutabakat: her ekstre satırını sistemdeki harekete eşle
            cur.execute(
                "SELECT tarih::text AS t, ROUND(tutar::numeric,2)::float AS tu, islem_turu, "
                "taksit_sayisi, baslangic_tarihi::text AS bas "
                "FROM kart_hareketleri WHERE kart_id=%s AND durum='aktif'",
                (kart["id"],),
            )
            mevcut = {}          # tek çekim / ödeme / faiz: (tarih, tutar, tip)
            mevcut_taksit = {}   # taksit: (baslangic, toplam, taksit_sayisi)
            for r in (cur.fetchall() or []):
                r = dict(r)
                if int(r.get("taksit_sayisi") or 1) > 1:
                    tk = (r.get("bas") or r["t"], round(float(r["tu"]), 2), int(r["taksit_sayisi"]))
                    mevcut_taksit[tk] = mevcut_taksit.get(tk, 0) + 1
                else:
                    key = (r["t"], round(float(r["tu"]), 2), r["islem_turu"])
                    mevcut[key] = mevcut.get(key, 0) + 1
            yeni_adet = 0
            for isl in sonuc.get("islemler", []):
                tsay = isl.get("taksit_sayisi")
                if tsay and tsay > 1 and isl.get("tip") == "HARCAMA":
                    # Taksitli: toplam (taksit_anapara) + başlangıç (tarih) + taksit sayısı ile eşle
                    total = round(float(isl.get("taksit_anapara") or 0), 2)
                    tk = (isl.get("tarih"), total, int(tsay))
                    if total <= 0:
                        isl["durum"] = "taksit"  # toplam bilinmiyor → elle
                    elif mevcut_taksit.get(tk, 0) > 0:
                        mevcut_taksit[tk] -= 1
                        isl["durum"] = "eslesti"
                    else:
                        isl["durum"] = "yeni"
                        yeni_adet += 1
                else:
                    tutar_r = round(float(isl.get("tutar") or 0), 2)
                    tip_i = isl.get("tip")
                    key = (isl.get("tarih"), tutar_r, tip_i)
                    if mevcut.get(key, 0) > 0:
                        mevcut[key] -= 1
                        isl["durum"] = "eslesti"
                    else:
                        # BULANIK EŞLEŞME (banka mutabakatı std): aynı tutar+tip, tarih ±3 gün.
                        # Banka işlem tarihi ile elle giriş tarihi farklı olabilir; açıklama umursanmaz.
                        _eslesti = False
                        try:
                            _it = date.fromisoformat(str(isl.get("tarih"))[:10])
                        except Exception:
                            _it = None
                        if _it:
                            for _k in list(mevcut.keys()):
                                if mevcut[_k] <= 0 or _k[1] != tutar_r or _k[2] != tip_i:
                                    continue
                                try:
                                    _kd = date.fromisoformat(str(_k[0])[:10])
                                except Exception:
                                    continue
                                if abs((_it - _kd).days) <= 3:
                                    mevcut[_k] -= 1
                                    isl["durum"] = "eslesti"
                                    isl["fuzzy_eslesme"] = True
                                    _eslesti = True
                                    break
                        if not _eslesti:
                            isl["durum"] = "yeni"
                            yeni_adet += 1
            # ── 🚧 DEVİR ÇİZGİSİ FRENİ (2026-08-18) ────────────────────────
            # 🔴 BUGÜN BİZZAT YAPTIĞIM HATA: Haziran kesimli GEÇMİŞ ekstreyi
            # aktardım ve kart borcunu 43.297,72 ₺ ŞİŞİRDİM. Sebep: Ağustos
            # ekstresinin "önceki borç" satırı (505.617,25 ₺) Haziran ve
            # Temmuz'un TÜM harcamalarını ZATEN İÇERİR. O ayların satırlarını
            # ayrıca eklemek aynı parayı ikinci kez borçlandırır. Mutabakat
            # 0,00'dan −43.297,72'ye düştü; 9 satırı izli iptal edip geri aldım.
            #
            # Bu, bugün ÜÇÜNCÜ kez görülen kalıp: TOPLU BİR RAKAMIN İÇİNDEKİ
            # DETAYI AYRICA EKLEMEK. Diğer ikisini sistem yakaladı (toplu ödeme
            # eşleşmesi · taksit çakışması); bunu YALNIZ İNSAN yakaladı — yani
            # tarama eksikti. Bu fren o boşluğu kapatır.
            #
            # Kural: yüklenen ekstrenin kesimi, kartın EN SON mutabık döneminden
            # ESKİYSE bu bir GEÇMİŞ ekstredir. Satırları zaten devir bakiyesinin
            # içindedir → aktarım varsayılan olarak KAPALI gelir.
            # ⛔ Ekstre yine de OKUNUR ve gösterilir: taksit zincirini görmek,
            #    okuma denetimi yapmak, belge arşivlemek için değerlidir.
            try:
                cur.execute("""SELECT MAX(kesim_tarihi) AS son FROM kart_ekstre_donem
                                WHERE kart_id=%s""", (kart["id"],))
                _sn = dict(cur.fetchone() or {})
                _son_kesim = _sn.get("son")
                _bu_kesim = sonuc.get("kesim_tarihi")
                if _son_kesim and _bu_kesim:
                    _bk = date.fromisoformat(str(_bu_kesim)[:10])
                    if _bk < _son_kesim:
                        _gecmis = [i for i in sonuc.get("islemler", [])
                                   if i.get("durum") == "yeni"]
                        for _i in _gecmis:
                            _i["durum"] = "devir_disi"
                            _i["devir_not"] = (
                                f"Bu ekstre {_bk} kesimli; sistemde {_son_kesim} kesimli "
                                "daha yeni bir ekstre mutabık. Bu satır o ekstrenin "
                                "«önceki borç» rakamının İÇİNDE zaten var — aktarılırsa "
                                "aynı para iki kez borç yazar.")
                        yeni_adet = max(0, yeni_adet - len(_gecmis))
                        sonuc["devir_cizgisi_uyarisi"] = {
                            "bu_kesim": str(_bk), "son_mutabik_kesim": str(_son_kesim),
                            "kilitlenen_satir": len(_gecmis),
                            "kilitlenen_tutar": round(sum(
                                abs(float(i.get("tutar") or 0)) for i in _gecmis), 2),
                            "mesaj": "🚧 GEÇMİŞ DÖNEM EKSTRESİ — satır aktarımı kapalı. "
                                     "Daha yeni bir ekstre zaten mutabık; bu ayın harcamaları "
                                     "onun «önceki borç» rakamının içinde. Ekstre yine de "
                                     "okundu: taksit zinciri, okuma denetimi ve belge arşivi "
                                     "için kullanılabilir.",
                        }
            except Exception as _de:  # noqa: BLE001
                sonuc["devir_cizgisi_uyarisi"] = {"hata": str(_de)[:120]}

            # ── 🔗 TAKSİT ÇAKIŞMA TARAMASI (2026-08-18) ────────────────────
            # 🔴 NEDEN: bir taksitli alım HER AY ekstrede yeniden görünür
            # (ESER TİCARET: Haziran 1/4 · Temmuz 2/4 · Ağustos 3/4 — hep
            # 41.250 ₺). Aynı alımın dilimi. Eğer plan bir ekstreden kurulduysa
            # (tutar=165.000, taksit_sayisi=4, başlangıç Haziran), motor o
            # dilimi ZATEN billiyor. Sonraki ekstrenin dilim satırını AYRICA
            # aktarmak aynı taksidi İKİ KEZ borçlandırır.
            # Satır-satır eşleme bunu göremez: defterdeki kayıt 165.000, ekstre
            # satırı 41.250 — tutarlar tutmaz, "yeni" sanılır.
            # Bu tarama, aktif taksit planlarının BU DÖNEME düşen dilimini
            # hesaplayıp ekstre satırıyla karşılaştırır.
            try:
                cur.execute("""SELECT tutar, taksit_sayisi,
                                      COALESCE(baslangic_tarihi, tarih) AS bas, aciklama
                                 FROM kart_hareketleri
                                WHERE kart_id=%s AND durum='aktif'
                                  AND islem_turu='HARCAMA' AND COALESCE(taksit_sayisi,1)>1""",
                            (kart["id"],))
                _planlar = [dict(r) for r in (cur.fetchall() or [])]
                _cak = []
                if _planlar:
                    _kes = sonuc.get("kesim_tarihi")
                    _kd = date.fromisoformat(str(_kes)[:10]) if _kes else None
                    for _i in sonuc.get("islemler", []):
                        if _i.get("durum") != "yeni" or _i.get("tip") != "HARCAMA":
                            continue
                        _tut = round(abs(float(_i.get("tutar") or 0)), 2)
                        for _p in _planlar:
                            _n = int(_p["taksit_sayisi"] or 1)
                            _dilim = round(float(_p["tutar"]) / _n, 2)
                            if abs(_dilim - _tut) > 1.0:
                                continue
                            # Bu ekstrenin ayı planın penceresinde mi?
                            _b = _p["bas"]
                            if _kd and _b:
                                _idx = (_kd.year - _b.year) * 12 + (_kd.month - _b.month)
                                if not (0 <= _idx < _n):
                                    continue
                            _i["durum"] = "taksit_kapsandi"
                            _i["taksit_not"] = (
                                f"Bu satır ({_tut:,.2f} ₺) defterdeki taksit planının "
                                f"({float(_p['tutar']):,.2f} ₺ / {_n} taksit) bu aya düşen "
                                "dilimidir — motor onu ZATEN billiyor. Aktarılırsa aynı "
                                "taksit iki kez borç yazar.")
                            _cak.append({"satir": _i.get("aciklama"), "dilim": _tut,
                                         "plan_toplam": float(_p["tutar"]), "taksit": _n})
                            yeni_adet = max(0, yeni_adet - 1)
                            break
                if _cak:
                    sonuc["taksit_cakisma_uyarisi"] = {
                        "adet": len(_cak), "satirlar": _cak,
                        "mesaj": "⚠️ Bu satırlar defterdeki taksit planlarının dilimleri. "
                                 "Aktarmayın — motor zaten billiyor.",
                    }
            except Exception as _ce:  # noqa: BLE001
                sonuc["taksit_cakisma_uyarisi"] = {"hata": str(_ce)[:120]}

            # ── 🧮 ADIM 7: TOPLU EŞLEŞME (bölünmüş ↔ toplu) ────────────────
            # 🔴 CANLI VAKA (2026-08-18, Garanti 3018): ekstre 22 Tem'de DÖRT
            # ayrı ödeme gösteriyordu (100.000 + 15.247 + 38.000 + 49.000 =
            # 202.247). Sistemde aynı para Ödeme Merkezi'nden TEK SATIR 202.247
            # olarak kayıtlıydı. Satır-satır eşleme "4 satır = 1 kayıt" ilişkisini
            # GÖREMEZ; dördü de "yeni" işaretlendi, aktarıldı ve ÖDEME ÇİFT
            # SAYILDI — kart borcu 202.247 ₺ eksik göründü. Elle fark ettim.
            # Bu tarama o boşluğu kapatır: aynı türden "yeni" satırların TOPLAMI
            # sistemdeki eşleşmemiş bir kayda denk geliyorsa hepsi işaretlenir.
            # ⛔ HÜKÜM VERMEZ, İÇE AKTARMAYI DURDURMAZ — uyarır. Kararı sahip verir
            #    (meşru olabilir: aynı gün gerçekten 4 ayrı ödeme + ayrı bir kayıt).
            try:
                _toplu = []
                for _tip in ("ODEME", "HARCAMA"):
                    _yeniler = [i for i in sonuc.get("islemler", [])
                                if i.get("durum") == "yeni" and i.get("tip") == _tip]
                    if len(_yeniler) < 2:
                        continue
                    _top = round(sum(abs(float(i.get("tutar") or 0)) for i in _yeniler), 2)
                    for _k, _adet in list(mevcut.items()):
                        if _adet <= 0 or _k[2] != _tip:
                            continue
                        if abs(float(_k[1]) - _top) > 1.0:
                            continue
                        for _i in _yeniler:
                            _i["durum"] = "toplu_eslesti"
                            _i["toplu_not"] = (
                                f"Bu satır tek başına sistemde yok, AMA {len(_yeniler)} "
                                f"{_tip.lower()} satırının toplamı ({_top:,.2f} ₺) sistemdeki "
                                f"{_k[0]} tarihli tek kayıtla birebir aynı. Aktarılırsa "
                                "AYNI PARA İKİ KEZ sayılır.")
                        _toplu.append({"tip": _tip, "satir": len(_yeniler), "toplam": _top,
                                       "sistem_tarih": _k[0], "sistem_tutar": float(_k[1])})
                        yeni_adet = max(0, yeni_adet - len(_yeniler))
                        break
                if _toplu:
                    sonuc["toplu_eslesme_uyarisi"] = {
                        "grup": _toplu,
                        "mesaj": "⚠️ Bölünmüş ekstre satırları, sistemdeki TOPLU bir kayıtla "
                                 "eşleşiyor. Bunları aktarmayın — aynı para iki kez sayılır. "
                                 "Satır-satır fren bu ilişkiyi göremez, bu tarama görür.",
                    }
            except Exception as _te:  # noqa: BLE001 — tarama yüklemeyi ASLA kilitlemez
                sonuc["toplu_eslesme_uyarisi"] = {"hata": str(_te)[:120]}
            sonuc["mutabakat"]["yeni_islem_adet"] = yeni_adet

            # ── BENZER GİDER UYARISI: "yeni" (eksik) HARCAMA kalemleri için, ±7 gün
            # içinde aynı tutarda kart ile girilmiş bir Anlık/Sabit Gider var mı kontrol et.
            # Varsa muhtemelen zaten sisteme girilmiş (bankaya farklı tarihte düşmüş) —
            # çift sayım riskine karşı kullanıcıyı uyar; "tüm eksikleri seç" bunu
            # otomatik işaretlemez, kullanıcı isterse elle seçer.
            for isl in sonuc.get("islemler", []):
                if isl.get("durum") != "yeni" or isl.get("tip") != "HARCAMA":
                    continue
                try:
                    _it = date.fromisoformat(str(isl.get("tarih"))[:10])
                except Exception:
                    continue
                _tutar = round(float(isl.get("tutar") or 0), 2)
                cur.execute("""
                    SELECT tarih::text AS tarih, aciklama, kategori
                    FROM anlik_giderler
                    WHERE durum='aktif' AND odeme_yontemi='kart'
                      AND ROUND(tutar::numeric,2) = %s
                      AND tarih BETWEEN %s::date - INTERVAL '7 days' AND %s::date + INTERVAL '7 days'
                    ORDER BY ABS(tarih - %s::date) ASC
                    LIMIT 1
                """, (_tutar, str(_it), str(_it), str(_it)))
                _benzer = cur.fetchone()
                if _benzer:
                    _benzer = dict(_benzer)
                    isl["benzer_gider_uyari"] = {
                        "tarih": _benzer["tarih"], "aciklama": _benzer["aciklama"],
                        "kategori": _benzer["kategori"],
                    }

            # Faiz oranlarını ekstreden GÜNCELLE (her ay otomatik — elle girmeye gerek yok)
            akdi = sonuc.get("akdi_faiz_yillik")
            gec = sonuc.get("gecikme_faiz_yillik")
            if akdi is not None and akdi > 0:
                cur.execute(
                    "UPDATE kartlar SET faiz_orani=%s, "
                    "gecikme_faiz_orani=COALESCE(%s, gecikme_faiz_orani) WHERE id=%s",
                    (akdi, gec, kart["id"]),
                )
                sonuc["faiz_guncellendi"] = {"faiz_orani": akdi, "gecikme_faiz_orani": gec}

            # Aylık ekstre SNAPSHOT'ı kaydet (kesim ayına göre; idempotent upsert)
            kt = sonuc.get("kesim_tarihi")
            if kt:
                try:
                    cur.execute(
                        """
                        INSERT INTO kart_ekstre_donem
                            (kart_id, donem, kesim_tarihi, son_odeme_tarihi, donem_borcu,
                             asgari_tutar, onceki_borc, donem_harcama, donem_odeme, donem_faizi,
                             kalan_taksit, kullanilabilir_limit, kalan_taksit_tutari, kaynak)
                        VALUES (%s, DATE_TRUNC('month', %s::date), %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'pdf')
                        ON CONFLICT (kart_id, donem) DO UPDATE SET
                            -- 🟡 P2 (2026-08-14, Codex): kaynak PDF yolunda hiç yazılmıyordu;
                            -- manuel dönem PDF'le düzeltilince rozet 'manuel' kalıyordu.
                            kaynak='pdf',
                            kesim_tarihi=EXCLUDED.kesim_tarihi, son_odeme_tarihi=EXCLUDED.son_odeme_tarihi,
                            donem_borcu=EXCLUDED.donem_borcu, asgari_tutar=EXCLUDED.asgari_tutar,
                            onceki_borc=EXCLUDED.onceki_borc, donem_harcama=EXCLUDED.donem_harcama,
                            donem_odeme=EXCLUDED.donem_odeme, donem_faizi=EXCLUDED.donem_faizi,
                            kalan_taksit=EXCLUDED.kalan_taksit,
                            kullanilabilir_limit=COALESCE(EXCLUDED.kullanilabilir_limit, kart_ekstre_donem.kullanilabilir_limit),
                            kalan_taksit_tutari=COALESCE(EXCLUDED.kalan_taksit_tutari, kart_ekstre_donem.kalan_taksit_tutari)
                        """,
                        (kart["id"], kt, kt, sonuc.get("son_odeme_tarihi"),
                         sonuc.get("donem_borcu"), sonuc.get("asgari_tutar"), sonuc.get("onceki_borc"),
                         sonuc.get("donem_harcama"), sonuc.get("donem_odeme"),
                         sonuc.get("donem_faizi") or 0, sonuc.get("kalan_taksit"),
                         sonuc.get("kullanilabilir_limit"), sonuc.get("kalan_taksit_tutari")),
                    )
                    sonuc["donem_kaydedildi"] = True
                    # Kart DURUM snapshot'ı (dönem borcu + kullanılabilir limit + kalan taksit +
                    # faiz) önizlemede DE kalıcı olsun → ekstre yüklenince toplam borç (taksit
                    # dahil) anında düzelir. "Borç uyumlu, yeni işlem yok" durumunda kullanıcı
                    # import etmese bile kart durumu güncellenir. İŞLEMLER (kart_hareketleri)
                    # yine yazılmaz; onlar ayrı ekstre-import adımında eklenir.
                    conn.commit()
                except Exception:
                    sonuc["donem_kaydedildi"] = False

            # A) ASGARİ/SON ÖDEME → CFO ödeme planı (gerçek banka değeriyle; çift olmaz, update-in-place)
            sot = sonuc.get("son_odeme_tarihi")
            asg = sonuc.get("asgari_tutar")
            brc = sonuc.get("donem_borcu")
            # Sistem başlangıcı 2026-06-01: Haziran ÖNCESİ son ödemeli ekstreden plan üretilmez
            # (eski ekstre tekrar yüklense bile hayalet 'vadesi geçmiş' oluşmasın).
            if sot and (asg or brc) and str(sot)[:10] >= '2026-06-01':
                acik = f"Kart ekstresi: {kart['kart_adi']} — asgari {asg}"
                cur.execute(
                    """UPDATE odeme_plani SET tarih=%s::date, odenecek_tutar=%s, asgari_tutar=%s,
                        aciklama=%s, referans_ay=DATE_TRUNC('month', %s::date)
                       WHERE kart_id=%s AND durum IN ('bekliyor','onay_bekliyor')
                         AND DATE_TRUNC('month', tarih) = DATE_TRUNC('month', %s::date)""",
                    (sot, (brc or asg), asg, acik, sot, kart["id"], sot),
                )
                if cur.rowcount == 0:
                    # R1b korumasi: ayni ay zaten ODENDIyse yeni bekleyen ekleme
                    #
                    # 🔴 EKSTRE YÜKLEMEYİ KOMPLE DURDURAN KUSUR (2026-08-24, canlı):
                    # Ziraat'in Ağustos ekstresi HTTP 500 ile reddedildi —
                    # "UniqueViolation: ux_odeme_plani_kart_donem (kart_id, referans_ay)".
                    # Sebep: yukarıdaki UPDATE yalnız 'bekliyor'/'onay_bekliyor'
                    # satırları tarıyor. O ay için plan satırı BAŞKA bir durumdaysa
                    # (ör. kesimden ÖNCE ödenmiş 'odendi') UPDATE 0 satır bulur,
                    # aşağıdaki NOT EXISTS de geçer ve INSERT tekil kısıtı çiğner.
                    # Sonuç: sahip ekstresini sisteme HİÇ yükleyemiyordu — üstelik
                    # hata mesajı bir SQL kısıt adıydı, ne yapacağını söylemiyordu.
                    # ON CONFLICT DO NOTHING: o ay için zaten bir plan varsa ikincisi
                    # açılmaz; ekstre yüklemesi bir plan satırı yüzünden ASLA durmaz.
                    # (Plan zaten tek gerçek değil — kasa izinden türer.)
                    cur.execute(
                        """INSERT INTO odeme_plani
                            (id, kart_id, tarih, referans_ay, odenecek_tutar, asgari_tutar, aciklama, durum)
                           SELECT %s, %s, %s::date, DATE_TRUNC('month', %s::date), %s, %s, %s, 'bekliyor'
                           WHERE NOT EXISTS (SELECT 1 FROM odeme_plani
                               WHERE kart_id=%s AND durum='odendi'
                                 AND DATE_TRUNC('month',tarih)=DATE_TRUNC('month',%s::date)
                                 AND COALESCE(odeme_tarihi, tarih) >= %s::date)
                           ON CONFLICT DO NOTHING""",
                        (str(uuid.uuid4()), kart["id"], sot, sot, (brc or asg), asg, acik,
                         kart["id"], sot, str(sonuc.get("kesim_tarihi") or sot)[:10]),
                    )
                sonuc["cfo_odeme_plani"] = {"son_odeme": sot, "asgari": asg, "borc": brc}
        elif son4:
            sonuc["eslestrme_notu"] = f"Son 4 hane '{son4}' ile eşleşen kart yok — kart tanımına son 4 haneyi girin."
    return sonuc


class ManuelEkstreBody(BaseModel):
    donem: str                              # kesim tarihi YYYY-MM-DD
    son_odeme: Optional[str] = None
    donem_borcu: float
    asgari_tutar: Optional[float] = None
    faiz_orani: Optional[float] = None
    gecikme_faiz_orani: Optional[float] = None


@app.post("/api/kartlar/{kid}/manuel-ekstre")
def kart_manuel_ekstre(kid: str, body: ManuelEkstreBody):
    """PDF okunamayan kartlar (Axess gibi) için ekstre özetini ELLE gir → aynı
    pipeline: snapshot + CFO ödeme planı + faiz + kart borcunu doğru değere çek.
    Borç düzeltmesi tek değiştirilebilir kayıtla yapılır (man_<id>); kasaya dokunmaz."""
    kesim = (body.donem or "")[:10]
    sot = (body.son_odeme or kesim)[:10]
    borc = float(body.donem_borcu or 0)
    asg = float(body.asgari_tutar or 0)
    with db() as (conn, cur):
        cur.execute("SELECT kart_adi FROM kartlar WHERE id=%s AND aktif=TRUE", (kid,))
        kr = cur.fetchone()
        if not kr:
            raise HTTPException(404, "Kart bulunamadı")
        kart_adi = dict(kr)["kart_adi"]
        # 1) Borcu hedef değere çek — tek değiştirilebilir DEVİR kaydı.
        #    DEVIR = açılış/devreden borç: gider sayılmaz, kasaya dokunmaz, borca eklenir.
        manid = "devir_" + kid
        cur.execute("DELETE FROM kart_hareketleri WHERE id IN (%s, %s)", (manid, "man_" + kid))
        diger = float(kart_borc(cur, kid) or 0)
        adj = round(borc - diger, 2)
        if abs(adj) > 0.01:
            cur.execute(
                """INSERT INTO kart_hareketleri
                   (id, kart_id, tarih, islem_turu, tutar, taksit_sayisi, aciklama, kaynak_tablo, kaynak_id)
                   VALUES (%s,%s,%s::date,'DEVIR',%s,1,%s,'devir',%s)""",
                (manid, kid, sot, adj,
                 "Açılış / devreden borç (ekstre bakiyesi)", manid),
            )
        # 2) snapshot
        cur.execute(
            """INSERT INTO kart_ekstre_donem
                (kart_id, donem, kesim_tarihi, son_odeme_tarihi, donem_borcu, asgari_tutar, kaynak)
               VALUES (%s, DATE_TRUNC('month',%s::date), %s::date, %s::date, %s, %s, 'manuel')
               ON CONFLICT (kart_id, donem) DO UPDATE SET
                 kesim_tarihi=EXCLUDED.kesim_tarihi, son_odeme_tarihi=EXCLUDED.son_odeme_tarihi,
                 donem_borcu=EXCLUDED.donem_borcu, asgari_tutar=EXCLUDED.asgari_tutar, kaynak='manuel'""",
            (kid, kesim, kesim, sot, borc, asg),
        )
        # 3) CFO ödeme planı — sistem başlangıcı 2026-06-01: Haziran öncesi son ödemeli
        #    ekstreden plan üretilmez (eski ekstre tekrar girilse de hayalet vadesi geçmiş olmasın)
        if (asg or borc) and str(sot)[:10] >= '2026-06-01':
            # TEK YAZICI (2026-08-08): kart planını üç ayrı yer yazıyordu ve
            # korumaları farklı anahtarlara bakıyordu → aynı ekstre 2-6 kez
            # plana düşüyordu. Artık dönem kimliği REFERANS AY ve tek kapı var.
            from kasa_service import kart_plani_upsert
            kart_plani_upsert(cur, kid, sot, (borc or asg), asg,
                              f"Kart ekstresi (manuel): {kart_adi} — asgari {asg}")
        # 4) faiz oranı
        if body.faiz_orani is not None and body.faiz_orani > 0:
            cur.execute("UPDATE kartlar SET faiz_orani=%s, gecikme_faiz_orani=COALESCE(%s,gecikme_faiz_orani) WHERE id=%s",
                        (body.faiz_orani, body.gecikme_faiz_orani, kid))
        yeni_borc = kart_borc(cur, kid)
        # K3 (2026-08-12, Kart denetimi): manuel-ekstre kart borcunu istemci girdisinden
        # (donem_borcu) yeniden yazıyor ama HİÇ audit yoktu. Koşulsuz (adj==0 dahil) iz bırak.
        audit(cur, 'kart_hareketleri', 'devir_' + kid, 'MANUEL_EKSTRE',
              yeni={'kart': kart_adi, 'donem': kesim, 'son_odeme': sot,
                    'donem_borcu': borc, 'asgari': asg, 'adj': adj,
                    'yeni_borc': round(yeni_borc, 2)})
    return {"success": True, "yeni_borc": round(yeni_borc, 2)}


@app.get("/api/kartlar/fatura-eslesme")
def kart_fatura_eslesme(ay_sayisi: int = 2, esik: float = 1000.0):
    """K2-D — KART HARCAMASI ↔ TEDARİKÇİ FATURASI köprüsü (öneri-only):
    1) EŞLEŞEN: işletme/belirsiz kart harcaması ile tutar ±%2 / tarih ±5 gün
       uyuşan fatura → 'faturası sistemde VAR, KDV indirimi ayrıştırılabilir'.
    2) FATURASIZ BÜYÜK: eşiği aşan işletme harcaması, eş faturası yok →
       belge/fatura isteme adayı (KDV indirimi kaçıyor olabilir).
    Hiçbir kaydı DEĞİŞTİRMEZ — farkındalık raporu."""
    n = max(1, min(6, int(ay_sayisi or 2)))
    with db() as (_, cur):
        cur.execute(
            """SELECT h.id, h.tarih::text AS tarih, k.kart_adi,
                      ROUND(h.tutar::numeric,2) AS tutar,
                      LEFT(COALESCE(h.aciklama,''),50) AS aciklama,
                      COALESCE(h.harcama_tipi,'belirsiz') AS tip
               FROM kart_hareketleri h JOIN kartlar k ON k.id = h.kart_id
               WHERE h.islem_turu='HARCAMA' AND h.durum='aktif'
                 AND COALESCE(h.harcama_tipi,'belirsiz') <> 'sahsi'
                 AND h.tarih >= DATE_TRUNC('month', CURRENT_DATE) - (%s || ' months')::interval
               ORDER BY h.tutar DESC""", (n - 1,))
        harcamalar = [dict(r) for r in cur.fetchall() or []]
        cur.execute(
            """SELECT id, tedarikci_ad, fatura_tarih::text AS tarih,
                      COALESCE(toplam_tutar,0)::float AS tutar
               FROM tedarikci_fatura
               WHERE fatura_tarih >= DATE_TRUNC('month', CURRENT_DATE) - (%s || ' months')::interval
                 AND COALESCE(durum,'') <> 'kopya'
                 AND COALESCE(toplam_tutar,0) > 0""", (n - 1,))
        faturalar = [dict(r) for r in cur.fetchall() or []]
    from datetime import date as _d
    import re as _re2

    # ⚠️ AD ÖRTÜŞMESİ ZORUNLU (2026-08-10 denetimi). Bu uç yalnız tutar+tarihe
    # bakıyordu ve canlıda ürettiği 3 "eşleşme"nin ÜÇÜ DE yanlıştı:
    #   GAZZE ELEKTRİK 20.977 ↔ APS GIDA (Red Bull tedarikçisi)
    #   ENERYA doğalgaz 1.410 ↔ Gelişim UYDU Sistemleri
    #   KÖYCEĞİZ İNTERNET 1.090 ↔ AGİT SEFA YÜCEAY (şahıs)
    # Ekran bunları "faturası sistemde VAR" diye gösteriyordu — sahip belge
    # istemeyi bırakır, KDV indirimi sessizce kaybolurdu. Tutar tek başına kanıt
    # değildir; hele değişken faturada hiç değildir.
    _ATIL = {"FATURA", "ODEME", "ODEMESI", "OTOMATIK", "TALIMAT", "SAN", "TIC",
             "LTD", "STI", "ANONIM", "SIRKETI", "LIMITED", "TICARET", "SANAYI",
             "KONYA", "ISTANBUL", "IZMIR", "ANKARA", "KARAMAN", "TR", "VE", "ILE",
             "GIDA", "ENERJI", "HIZMET", "HIZMETLERI", "URUN", "DIGER", "GENEL"}

    def _kelime_seti(x):
        t = str(x or "")
        for a, b in zip("çğıöşüÇĞİıÖŞÜâîû", "cgiosuCGIIOSUaiu"):
            t = t.replace(a, b)
        t = _re2.sub(r"[^A-Z0-9 ]", " ", t.upper())
        return {w for w in t.split() if len(w) >= 4 and w not in _ATIL and not w.isdigit()}

    kullanildi: set = set()
    eslesen, faturasiz_buyuk = [], []
    kdv_ayrisabilir = 0.0
    for h in harcamalar:
        tut = float(h["tutar"])
        aday = None
        h_kel = _kelime_seti(h["aciklama"])
        for f in faturalar:
            if f["id"] in kullanildi:
                continue
            if abs(f["tutar"] - tut) > max(5.0, tut * 0.02):
                continue
            # Ad örtüşmesi olmadan eşleştirme YOK — yanlış "faturası var" hükmü
            # gerçek bir belge takibini öldürür.
            if not (h_kel & _kelime_seti(f.get("tedarikci_ad"))):
                continue
            try:
                gunfark = abs((_d.fromisoformat(h["tarih"][:10])
                               - _d.fromisoformat(str(f["tarih"])[:10])).days)
            except Exception:  # noqa: BLE001
                continue
            if gunfark <= 5:
                aday = f
                break
        if aday:
            kullanildi.add(aday["id"])
            kdv_ayrisabilir += tut
            eslesen.append({"harcama_tarih": h["tarih"], "kart": h["kart_adi"],
                            "tutar": tut, "aciklama": h["aciklama"],
                            "fatura_tedarikci": aday["tedarikci_ad"],
                            "fatura_tarih": aday["tarih"]})
        elif tut >= float(esik) and h["tip"] == "isletme":
            faturasiz_buyuk.append({"tarih": h["tarih"], "kart": h["kart_adi"],
                                    "tutar": tut, "aciklama": h["aciklama"]})
    return {
        "eslesen": eslesen[:30],
        "eslesen_toplam": round(kdv_ayrisabilir, 2),
        "faturasiz_buyuk_isletme": faturasiz_buyuk[:30],
        "faturasiz_toplam": round(sum(x["tutar"] for x in faturasiz_buyuk), 2),
        "not": "ADAY eşleştirme — hüküm değil. 'Eşleşen' = faturası sistemde olan kart "
               "harcaması (KDV'si fatura kaleminden ayrıştırılabilir). 'Faturasız büyük' "
               "= belge isteme adayı; fatura alınmazsa indirilecek KDV kaçar. Kart "
               "giderleri şu an KDV-ayrıştırmasız gider yazılır (bilinçli model).",
    }


@app.get("/api/kartlar/taksit-takvimi")
def kart_taksit_takvimi(ay: int = 12):
    """K2-C — TAKSİT TAKVİMİ: gelecek N ay, kart × ay taksit yükü dağılımı
    (banka 'gelecek dönem taksitleriniz' ekranının karşılığı). Salt-okur."""
    n = max(3, min(24, int(ay or 12)))
    from finans_core import gelecek_taksit_yuku
    takvim, ay_toplam = [], {}
    with db() as (_, cur):
        cur.execute("SELECT id, kart_adi FROM kartlar WHERE aktif=TRUE ORDER BY kart_adi")
        for k in [dict(r) for r in cur.fetchall() or []]:
            try:
                aylar = gelecek_taksit_yuku(cur, k["id"], ay_sayisi=n) or []
            except Exception:
                aylar = []
            dolu = [a for a in aylar if (a.get("taksit_yuku") or 0) > 0]
            if not dolu:
                continue
            takvim.append({"kart": k["kart_adi"], "aylar": aylar,
                           "toplam_kalan": round(sum(a["taksit_yuku"] for a in aylar), 2)})
            for a in aylar:
                ay_toplam[a["ay"]] = round(ay_toplam.get(a["ay"], 0) + a["taksit_yuku"], 2)
    return {"kartlar": takvim,
            "ay_toplamlari": [{"ay": k2, "toplam_taksit": v}
                              for k2, v in sorted(ay_toplam.items())],
            "not": "Bilinen taksit sözleşmelerinin aylara dağılımı — yeni taksitli "
                   "alım yapıldıkça takvim büyür."}


@app.get("/api/kartlar/sahsi-cekim-raporu")
def kart_sahsi_cekim_raporu(ay_sayisi: int = 6):
    """K2-C — ŞAHSİ ÇEKİM RAPORU (sahip carisi): işletme kartlarından yapılan
    ŞAHSİ harcamaların ay × kart dökümü. P&L'e girmeyen ama kart borcunu
    büyüten kalemler — 'işletme, sahibin şahsi harcamasını finanse ediyor mu?'"""
    n = max(1, min(24, int(ay_sayisi or 6)))
    with db() as (_, cur):
        cur.execute(
            """SELECT TO_CHAR(h.tarih,'YYYY-MM') AS ay, k.kart_adi,
                      ROUND(SUM(h.tutar)::numeric,2) AS toplam, COUNT(*)::int AS adet
               FROM kart_hareketleri h JOIN kartlar k ON k.id = h.kart_id
               WHERE h.islem_turu='HARCAMA' AND h.durum='aktif'
                 AND h.harcama_tipi='sahsi'
                 AND h.tarih >= DATE_TRUNC('month', CURRENT_DATE) - (%s || ' months')::interval
               GROUP BY 1, 2 ORDER BY 1 DESC, toplam DESC""", (n - 1,))
        kirilim = [dict(r) for r in cur.fetchall() or []]
        for r in kirilim:
            r["toplam"] = float(r["toplam"])
        cur.execute(
            """SELECT h.tarih::text AS tarih, k.kart_adi,
                      ROUND(h.tutar::numeric,2) AS tutar,
                      LEFT(COALESCE(h.aciklama,''),60) AS aciklama
               FROM kart_hareketleri h JOIN kartlar k ON k.id = h.kart_id
               WHERE h.islem_turu='HARCAMA' AND h.durum='aktif'
                 AND h.harcama_tipi='sahsi'
               ORDER BY h.tarih DESC LIMIT 10""")
        son10 = [dict(r) for r in cur.fetchall() or []]
        for r in son10:
            r["tutar"] = float(r["tutar"])
    genel = round(sum(r["toplam"] for r in kirilim), 2)
    return {"ay_kart_kirilimi": kirilim, "son_10_sahsi_islem": son10,
            "genel_toplam": genel,
            "not": "Şahsi harcamalar P&L'e GİRMEZ ama kart borcunu ve faiz yükünü "
                   "büyütür — sahip carisi olarak izlenir; hüküm yok, görünürlük.",
    }


@app.delete("/api/kartlar/{kid}/ekstre-donem/{donem}")
def kart_ekstre_donem_sil(kid: str, donem: str):
    """Bir kart için tek bir ayın (dönem) ekstre verisini siler: o aya ait
    HARCAMA/FAIZ hareketlerini ve kart_ekstre_donem snapshot'ını kaldırır.
    Hatalı yüklenen ekstreyi silip 'Ekstre Yükle' ile yeniden yüklemek için kullanılır.
    ÖDEME/DEVIR hareketlerine (manuel girilen kasa hareketleri) dokunmaz."""
    with db() as (conn, cur):
        cur.execute("SELECT kart_adi, kesim_gunu FROM kartlar WHERE id=%s", (kid,))
        k = cur.fetchone()
        if not k:
            raise HTTPException(404, "Kart bulunamadı")
        # 🔴 P1 (2026-08-14, EVV-KART / Codex): silme penceresi TAKVİM AYI idi —
        # ekstre kesim DÖNGÜSÜ ise (önceki kesim, bu kesim]. 25 Temmuz kesimli
        # ekstrede 26-30 Haziran satırları doğaldır; takvim-ayı silme onları
        # BIRAKIYORDU (yanlış ekstrenin kalıntısı defterde kalıyordu). Pencere
        # artık kesim döngüsü: bu kesim = dönem ayında kesim günü; başlangıç =
        # bir önceki kesim. Kesim günü tanımsızsa eski takvim-ayı davranışı.
        # Pencere kaynağı ÖNCELİĞİ (Codex diff-review 2026-08-14): dönemin
        # SNAPSHOT'ındaki gerçek kesim_tarihi > kartın bugünkü kesim_gunu —
        # kesim günü sonradan değiştirilmişse yanlış pencere silerdi. Tarih
        # aritmetiği Python'da (SQL jimnastiği yerine).
        import calendar as _takvim
        from datetime import date as _tarih
        def _ay_gunu(y, m, g):
            return _tarih(y, m, min(int(g), _takvim.monthrange(y, m)[1]))
        _dm = None
        try:
            _dy, _dmo = int(str(donem)[:4]), int(str(donem)[5:7])
            _dm = (_dy, _dmo)
        except (ValueError, IndexError):
            pass
        _bit = None
        if _dm:
            cur.execute(
                "SELECT kesim_tarihi FROM kart_ekstre_donem "
                "WHERE kart_id=%s AND donem=DATE_TRUNC('month', %s::date)", (kid, donem))
            _snap = cur.fetchone()
            if _snap and dict(_snap).get("kesim_tarihi"):
                _bit = dict(_snap)["kesim_tarihi"]
            else:
                _kg = int(dict(k).get("kesim_gunu") or 0)
                if _kg >= 1:
                    _bit = _ay_gunu(_dm[0], _dm[1], _kg)
        if _bit is not None and _dm:
            _oy, _om = (_dm[0] - 1, 12) if _dm[1] == 1 else (_dm[0], _dm[1] - 1)
            cur.execute(
                "SELECT kesim_tarihi FROM kart_ekstre_donem "
                "WHERE kart_id=%s AND donem=make_date(%s,%s,1)", (kid, _oy, _om))
            _psnap = cur.fetchone()
            if _psnap and dict(_psnap).get("kesim_tarihi"):
                _bas = dict(_psnap)["kesim_tarihi"]
            else:
                _bas = _ay_gunu(_oy, _om, _bit.day)
            _pencere = " AND tarih > %s AND tarih <= %s"
            _pp = [_bas, _bit]
        else:
            _pencere = " AND DATE_TRUNC('month', tarih) = DATE_TRUNC('month', %s::date)"
            _pp = [donem]
        # K1 (2026-08-12): ekstre-import'tan açılmış EŞLENİK anlık giderleri de
        # kapat — yoksa dönem silinince o giderler ÖKSÜZ kalıp P&L'de asılı duruyordu.
        cur.execute(
            f"""UPDATE anlik_giderler SET durum='iptal'
               WHERE durum='aktif' AND id IN (
                 SELECT 'agk_' || id FROM kart_hareketleri
                 WHERE kart_id=%s AND islem_turu IN ('HARCAMA','FAIZ'){_pencere})""",
            [kid] + _pp,
        )
        iptal_anlik = cur.rowcount
        cur.execute(
            f"""DELETE FROM kart_hareketleri
               WHERE kart_id=%s AND islem_turu IN ('HARCAMA','FAIZ'){_pencere}""",
            [kid] + _pp,
        )
        silinen_hareket = cur.rowcount
        cur.execute(
            "DELETE FROM kart_ekstre_donem WHERE kart_id=%s AND donem = DATE_TRUNC('month', %s::date)",
            (kid, donem),
        )
        silinen_donem = cur.rowcount
        # K1: hard-delete artık izli — ne silindiği audit'e yazılır (forensik iz yoktu).
        audit(cur, 'kart_hareketleri', kid, 'EKSTRE_DONEM_SIL',
              yeni={'donem': donem, 'silinen_hareket': silinen_hareket,
                    'silinen_donem': silinen_donem, 'iptal_anlik_gider': iptal_anlik})
    return {
        "success": True,
        "kart_adi": dict(k)["kart_adi"],
        "silinen_hareket": silinen_hareket,
        "silinen_donem": silinen_donem,
        "iptal_anlik_gider": iptal_anlik,
    }


class KartLedgerSifirlaBody(BaseModel):
    onay_pin: Optional[str] = None     # İşletme onayı (Merve Karabacak PIN)
    hepsi: bool = False                # True → tüm kartlar; False → tek kart


@app.post("/api/kartlar/{kid}/ledger-sifirla")
def kart_ledger_sifirla(kid: str, body: KartLedgerSifirlaBody):
    """Bozuk/karışık kart hareketlerini temizler (durum='iptal') → kart borcu sıfırlanır.
    Açılış devri kurmadan ÖNCE çalıştırılır. İşletme onayı (Merve Karabacak PIN) şart.
    Bağlı kasa hareketleri (KART_ODEME/KART_FAIZ) da iptal edilir. Auditli.
    kid='__hepsi__' veya hepsi=True → tüm aktif kartlar."""
    from operasyon_merkez_api import _isletme_onay_dogrula
    with db() as (conn, cur):
        onayci = _isletme_onay_dogrula(cur, body.onay_pin)  # PIN hatalı → 403
        # K4 (2026-08-12, Kart denetimi): eskiden `body.hepsi or kid=='__hepsi__'` →
        # tek-kart rotasına {"hepsi":true} POST'lamak TÜM kartları sıfırlıyordu (PIN
        # yine şart ama footgun). Toplu sıfırlama YALNIZ açık __hepsi__ rotasından.
        if body.hepsi and kid != "__hepsi__":
            logger.warning(f"ledger-sifirla: tek-kart rotasına hepsi=true geldi (kid={kid}) — yok sayıldı")
        tum = kid == "__hepsi__"
        if tum:
            cur.execute("SELECT id::text FROM kartlar WHERE aktif=TRUE")
            kart_ids = [dict(r)["id"] for r in (cur.fetchall() or [])]
        else:
            cur.execute("SELECT id::text FROM kartlar WHERE id=%s AND aktif=TRUE", (kid,))
            r = cur.fetchone()
            if not r:
                raise HTTPException(404, "Kart bulunamadı")
            kart_ids = [dict(r)["id"]]
        toplam_iptal = 0
        for k in kart_ids:
            # bağlı kasa hareketlerini iptal et (varsa)
            cur.execute(
                """UPDATE kasa_hareketleri SET durum='iptal'
                   WHERE kaynak_tablo='kart_hareketleri' AND durum='aktif'
                     AND kaynak_id IN (SELECT id FROM kart_hareketleri WHERE kart_id=%s AND durum='aktif')""",
                (k,),
            )
            cur.execute(
                "UPDATE kart_hareketleri SET durum='iptal' WHERE kart_id=%s AND durum='aktif'", (k,),
            )
            toplam_iptal += cur.rowcount
            audit(cur, "kart_hareketleri", k, "LEDGER_SIFIRLA",
                  yeni={"onayci": onayci.get("ad_soyad"), "iptal_adet": cur.rowcount})
    return {"success": True, "kart_sayisi": len(kart_ids), "iptal_edilen_hareket": toplam_iptal}


class KasaAcilisBody(BaseModel):
    tutar: float
    onay_pin: Optional[str] = None
    tarih: Optional[str] = None        # default 2026-06-01


@app.post("/api/kasa/acilis-devri")
def kasa_acilis_devri(body: KasaAcilisBody):
    """Sistemin AÇILIŞ kasasını (devir) belirler — kasayı tek kayıtla gerçek tutara çeker.
    Önceki açılış kaydını siler, yenisini yazar. İşletme onayı (Merve Karabacak PIN) şart.
    NOT: Diğer kasa hareketlerine dokunmaz; sadece açılış/devir kalemini kurar. Auditli."""
    from operasyon_merkez_api import _isletme_onay_dogrula
    from finans_core import kasa_bakiyesi
    tarih = (body.tarih or "2026-06-01")[:10]
    hedef = float(body.tutar)
    with db() as (conn, cur):
        onayci = _isletme_onay_dogrula(cur, body.onay_pin)  # PIN hatalı → 403
        # Önceki açılış kaydını sil, kalan kasayı hesapla, FARK kadar düzeltme yaz →
        # kasa tam HEDEF tutara çekilir (mevcut hareketler ne olursa olsun, üstüne EKLEMEZ).
        cur.execute("DELETE FROM kasa_hareketleri WHERE islem_turu='ACILIS_DEVRI'")
        mevcut = kasa_bakiyesi(cur)
        duzeltme = round(hedef - mevcut, 2)
        insert_kasa_hareketi(
            cur, tarih, "ACILIS_DEVRI", duzeltme,
            f"Sistem açılış kasası (1 Haziran) — {hedef:,.0f}₺'ye çekildi",
            "sistem", "acilis_devri",
            idempotency_key=f"acilis_devri_{tarih}_{uuid.uuid4().hex[:10]}",
        )
        audit(cur, "kasa_hareketleri", "acilis_devri", "KASA_ACILIS",
              yeni={"hedef": hedef, "onceki": round(mevcut, 2), "duzeltme": duzeltme,
                    "tarih": tarih, "onayci": onayci.get("ad_soyad")})
        yeni_bakiye = kasa_bakiyesi(cur)
    return {"success": True, "kasa_bakiye": round(yeni_bakiye, 2), "duzeltme": duzeltme}


class TopluDevirBody(BaseModel):
    onay_pin: Optional[str] = None


@app.post("/api/kartlar/toplu-devir")
def kartlar_toplu_devir(body: TopluDevirBody):
    """Her kartın en son ekstre snapshot'ındaki (kart_ekstre_donem) dönem borcunu
    AÇILIŞ DEVRİ (islem_turu='DEVIR') olarak kurar → kart borçları Haziran'a temiz taşınır.
    Gider sayılmaz, kasaya dokunmaz. İşletme onayı şart. Auditli. Tekrar çalıştırılabilir."""
    from operasyon_merkez_api import _isletme_onay_dogrula
    with db() as (conn, cur):
        onayci = _isletme_onay_dogrula(cur, body.onay_pin)
        cur.execute("SELECT id::text, kart_adi FROM kartlar WHERE aktif=TRUE")
        kartlar = [dict(r) for r in (cur.fetchall() or [])]
        sonuc = []
        for k in kartlar:
            kid = k["id"]
            cur.execute(
                """SELECT donem_borcu, kesim_tarihi::text AS kt
                   FROM kart_ekstre_donem WHERE kart_id=%s ORDER BY donem DESC LIMIT 1""",
                (kid,),
            )
            r = cur.fetchone()
            if not r:
                continue
            r = dict(r)
            if r.get("donem_borcu") is None:
                continue
            borc = float(r["donem_borcu"])
            manid = "devir_" + kid
            cur.execute("DELETE FROM kart_hareketleri WHERE id IN (%s, %s)", (manid, "man_" + kid))
            diger = float(kart_borc(cur, kid) or 0)
            adj = round(borc - diger, 2)
            if abs(adj) > 0.01:
                cur.execute(
                    """INSERT INTO kart_hareketleri
                       (id, kart_id, tarih, islem_turu, tutar, taksit_sayisi, aciklama, kaynak_tablo, kaynak_id)
                       VALUES (%s,%s,%s::date,'DEVIR',%s,1,%s,'devir',%s)""",
                    (manid, kid, (r.get("kt") or "2026-05-31"), adj,
                     "Açılış / devreden borç (son ekstre bakiyesi)", manid),
                )
            sonuc.append({"kart": k["kart_adi"], "devir_borc": round(kart_borc(cur, kid), 2)})
        # Haziran ödeme planlarını (CFO hatırlatıcıları) üret — kart son ödeme günü + borç
        plan_sonuc = []
        try:
            from kasa_service import kart_plan_guncelle_tx
            plan_sonuc = kart_plan_guncelle_tx(cur)
        except Exception:
            pass
        audit(cur, "kart_hareketleri", "toplu", "TOPLU_DEVIR",
              yeni={"adet": len(sonuc), "plan": len(plan_sonuc), "onayci": onayci.get("ad_soyad")})
    return {"success": True, "kart_sayisi": len(sonuc), "kartlar": sonuc, "odeme_plani_uretildi": len(plan_sonuc)}


class EkstreImportIslem(BaseModel):
    tarih: Optional[str] = None
    tutar: float
    tip: str = "HARCAMA"      # HARCAMA | ODEME | FAIZ
    aciklama: Optional[str] = None
    harcama_tipi: Optional[str] = None  # isletme | sahsi | belirsiz
    kategori: Optional[str] = None      # ekstre kategorisi (Market, Akaryakıt...)
    taksit_sayisi: Optional[int] = None # taksitli alımda toplam taksit (Y)
    taksit_anapara: Optional[float] = None  # taksitli alımın TOPLAM tutarı


class EkstreImportBody(BaseModel):
    kart_id: str
    islemler: List[EkstreImportIslem]
    zorla: Optional[bool] = False  # çift-yazma frenini bilinçli geç


@app.post("/api/kartlar/ekstre-import")
def kart_ekstre_import(body: EkstreImportBody):
    """Faz E1: Ekstreden seçilen EKSİK işlemleri kart_hareketleri'ne yazar.
    İdempotent (deterministik id → çift import yok). Kasaya DOKUNMAZ (sadece kart
    borcu); taksit satırları kabul edilmez (v1). HARCAMA/ODEME/FAIZ.

    Ayrıca: şahsi OLMAYAN (işletme/belirsiz) HARCAMA satırları için, geçmişte
    girilmesi unutulmuş olabilecek harcamayı görünür kılmak amacıyla
    anlik_giderler'e de (odeme_yontemi='kart', kaynak_tablo='ekstre_import')
    bir kayıt eklenir — kart_hareketleri'ne TEKRAR yazmaz (çift borç oluşmaz),
    sadece CFO/Maliyet ekranlarında gider olarak görünür hale gelir. Şahsi
    (harcama_tipi='sahsi') satırlar kart borcuna yazılır ama anlık gidere
    yansıtılmaz."""
    import hashlib
    with db() as (conn, cur):
        cur.execute("SELECT id FROM kartlar WHERE id=%s AND aktif=TRUE", (body.kart_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Kart bulunamadı")
        yazilan, atlanan = 0, 0
        anlik_gider_yazilan = 0
        # ÇİFT-YAZMA FRENİ (2026-07-10, sahip: 'ekstreyi de yükleyince çift yazıyor'):
        # UI eşleştirmesi kuruş/tarih kaymasını kaçırabilir; sunucu ikinci hat olarak
        # elle girilmiş (ekstre_import OLMAYAN) eş kaydı arar — tutar ±1 TL, tarih ±5
        # gün, aynı tip. Eşi bulunan satır YAZILMAZ (atlanan_mevcut'a düşer);
        # bilinçli geçmek için body.zorla=true. Her manuel kayıt EN FAZLA BİR ekstre
        # satırını yutar (aynı güne iki eşit meşru harcama korunur).
        _kullanilan_manuel: set = set()
        atlanan_mevcut = []
        _zorla = bool(getattr(body, "zorla", False))
        faiz_donemleri: set = set()  # ekstreden faiz gelen YYYY-MM dönemleri (motor tahmini iptali için)
        for isl in body.islemler:
            tip = (isl.tip or "HARCAMA").upper()
            if tip not in ("HARCAMA", "ODEME", "FAIZ"):
                atlanan += 1; continue
            # TAKSİTLİ alım: tutar=TOPLAM (taksit_anapara), taksit_sayisi=Y, baslangic=tarih
            tsay = int(isl.taksit_sayisi or 1)
            is_taksit = tip == "HARCAMA" and tsay > 1 and float(isl.taksit_anapara or 0) > 0
            if is_taksit:
                tutar = round(abs(float(isl.taksit_anapara)), 2)
            else:
                tsay = 1
                tutar = abs(float(isl.tutar or 0))
            if tutar <= 0:
                atlanan += 1; continue
            # Faiz dönemini SADECE tutar>0 geçerli satır için işaretle (motor tahmini iptali);
            # sıfır tutarlı satır atlanırsa o dönem yanlışlıkla iptal tetiklenmesin.
            if tip == "FAIZ":
                _ft = (isl.tarih or str(bugun_tr()))[:7]
                if len(_ft) == 7:
                    faiz_donemleri.add(_ft)
            tarih = (isl.tarih or str(bugun_tr()))[:10]
            htip = (isl.harcama_tipi or "").strip().lower()
            if htip not in ("isletme", "sahsi", "belirsiz"):
                if tip == "HARCAMA":
                    # ⚠️ SATICI HAFIZASI ARTIK OTOMATİK YAZMAZ — YALNIZ ÖNERİR
                    # (sahip kararı 2026-08-11: "her zaman sorsun, aslında bazen
                    # işletme için de alınmış olabilir!").
                    # Eskiden kural bulununca harcama_tipi DOĞRUDAN yazılıyordu.
                    # Ama aynı satıcıdan bazen işletme bazen şahsi alım olur:
                    # A101'den ev alışverişi de yapılır, dükkâna peçete de alınır.
                    # Geçmişe bakıp bugünü hükme bağlamak sessiz yanlış üretir.
                    # Kural artık `oneri_tipi` olarak EKRANA taşınır; satır
                    # 'belirsiz' doğar ve sahip karar verir.
                    htip = "belirsiz"
                else:
                    htip = "isletme"
            if not _zorla and not is_taksit:
                # 🐞 PENCERE ±5 → ±10 (2026-08-03, FEZ vakası): banka valörü
                # 25.06, sistemdeki kısmi-vadeli kaydı 03.07 = 8 gün — fren
                # kaçırdı, aynı ödeme kart borcuna İKİ kez yazıldı (80K+20K).
                # Ekstre valör kaymaları 5 günü rahat aşıyor.
                cur.execute(
                    """SELECT id, tarih::text AS t, tutar::float AS tu
                       FROM kart_hareketleri
                       WHERE kart_id=%s AND durum='aktif' AND islem_turu=%s
                         AND COALESCE(kaynak_tablo,'') <> 'ekstre_import'
                         AND ABS(tutar - %s) <= 1.0
                         AND tarih BETWEEN %s::date - 10 AND %s::date + 10
                       ORDER BY ABS(tarih - %s::date), id""",
                    (body.kart_id, tip, tutar, tarih, tarih, tarih))
                _es = next((dict(r) for r in (cur.fetchall() or [])
                            if dict(r)["id"] not in _kullanilan_manuel), None)
                if _es:
                    _kullanilan_manuel.add(_es["id"])
                    atlanan += 1
                    atlanan_mevcut.append(
                        {"tarih": tarih, "tutar": tutar, "tip": tip,
                         "mevcut_kayit_tarihi": _es["t"], "mevcut_tutar": _es["tu"],
                         "aciklama": (isl.aciklama or "")[:60]})
                    continue
            anahtar = f"{body.kart_id}|{tarih}|{tutar:.2f}|{tip}|{tsay}|{(isl.aciklama or '')[:40]}"
            hid = "eks_" + hashlib.md5(anahtar.encode("utf-8")).hexdigest()[:24]
            cur.execute(
                """INSERT INTO kart_hareketleri
                   (id, kart_id, tarih, islem_turu, tutar, taksit_sayisi, baslangic_tarihi, aciklama,
                    harcama_tipi, kategori, kaynak_tablo, kaynak_id)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'ekstre_import',%s)
                   ON CONFLICT (id) DO NOTHING""",
                (hid, body.kart_id, tarih, tip, tutar, tsay,
                 (tarih if is_taksit else None),
                 (isl.aciklama or "Ekstre içe aktarım")[:200], htip,
                 (isl.kategori or None), hid),
            )
            if cur.rowcount > 0:
                yazilan += 1
                # ⛔ ESKİ PRENSİP KALDIRILDI (2026-08-10, sahip kararı + Codex):
                # Burada eskiden anlik_giderler'e de bir GİDER satırı yazılıyordu
                # ("şahsi olmayan harcamayı P&L'de görünür kıl"). Bu yanlıştı:
                # ödeme kanıtını maliyet kaydına çeviriyordu ve aynı para hem kart
                # defterinde hem gider defterinde durabiliyordu. Canlıda 166 kayıt /
                # 590.229,50 ₺ bu şekilde birikmişti.
                #
                # YENİ İLKE — GİDER = PARA ÇIKIŞI, TEK YERDEN:
                #   nakit çıkışı → anlik_giderler
                #   kart  çıkışı → kart_hareketleri  ← bu satır zaten yazıldı
                # P&L artık `gider_kanonik` görünümünden okuyor; kart harcaması
                # oradan sayılıyor. İkinci bir gider kaydı ÜRETİLMEZ.
                #
                # Geçmiş ekstre_import gider satırları SİLİNMEDİ; kanonik görünüm
                # onları zaten dışarıda bırakıyor (odeme_yontemi='kart'), arşiv
                # olarak denetlenebilir kalıyorlar.
            else:
                atlanan += 1  # zaten var (idempotent)

        # ── OTOMATİK: ekstreden GERÇEK banka faizi geldiyse, o dönem için motorun
        #    TAHMİNİ faizini iptal et (çift faiz olmasın). Motor faizi açıklamasında
        #    "kesim faizi" işaretini taşır; ekstre faizi (DÖNEM FAİZİ/Kredi faizi/KKDF…)
        #    taşımaz → yalnızca tahmin iptal olur, banka gerçeği kalır. İdempotent.
        motor_faizi_iptal = 0
        for _donem in faiz_donemleri:
            cur.execute(
                """
                UPDATE kart_hareketleri
                SET durum='iptal',
                    aciklama = aciklama || ' [ekstre gerçeği geldi — motor tahmini iptal]'
                WHERE kart_id=%s AND islem_turu='FAIZ' AND durum='aktif'
                  AND aciklama LIKE '%% kesim faizi %%'
                  AND to_char(tarih, 'YYYY-MM') = %s
                """,
                (body.kart_id, _donem),
            )
            motor_faizi_iptal += cur.rowcount or 0

        if yazilan or motor_faizi_iptal:
            try:
                # FIX K2 (2026-07-05): eskiden 'from motors import' idi ama fonksiyon
                # kasa_service'te → ImportError sessizce yutuluyordu (plan hiç güncellenmiyordu).
                from kasa_service import kart_plan_guncelle_tx
                kart_plan_guncelle_tx(cur)
            except Exception:
                pass
        yeni_borc = kart_borc(cur, body.kart_id)

    # ── 🔄 AP YENİDEN EŞLEŞTİRME (2026-08-08, sahip: "kart ekstresi yüklendiğinde
    # tekrar ödeme izlerini araştırıp sistemde var mı kontrol etmesi lazım").
    # Codex denetiminin işaret ettiği boşluk buydu: import yalnız kart borcunu
    # yazıyor, açık tedarikçi borçlarını yeniden değerlendirmiyordu.
    # Otomatik BAĞLAMA yapılır (KAPATMA değil): yeni ödeme kaydı açılmaz, sadece
    # "bu çekim şu tedarikçiye ait" damgası konur; borçtan düşme cari okumasıyla
    # olur. Yalnız ≥%95 güvenliler bağlanır, gerisi sahip onayına düşer.
    # HATA-YUTAR: eşleştirme patlarsa ekstre importu BOZULMAZ (para yazımı bitti).
    rematch = None
    if yazilan:
        try:
            from fatura_api import kart_izi_otomatik_tara
            rematch = kart_izi_otomatik_tara(gun=400, uygula=1)
        except Exception as e:  # noqa: BLE001
            logger.warning("ekstre sonrası AP rematch atlandı (yutuldu): %s", str(e)[:150])
            rematch = {"ok": False, "hata": str(e)[:120]}

    # 🔄 DENKLEŞTİRME YAMASINI OTOMATİK TAZELE (2026-08-18, sahip: "yoksa 7/7
    # birkaç gün sonra yine 3/7 olur").
    # 🔴 NEDEN ZORUNLU: DEVİR yaması `banka borcu − kart_borc()` formülüyle
    # yazılır. Her içe aktarım defteri büyütür ama yama SABİT kalır → aradaki
    # fark hemen yeniden açılır. Bugün elle tazeleyip 7 kartın 7'sini
    # bankayla kuruşu kuruşuna hizaladık; bu çağrı olmadan o kazanç BİR SONRAKİ
    # AKTARIMDA kaybolurdu.
    # Yalnız satır YAZILDIYSA çalışır (yazılmadıysa defter değişmemiştir).
    # HATA-YUTAR: tazeleme düşse bile içe aktarım BOZULMAZ — para yazımı bitti.
    devir_tazele = None
    if yazilan:
        try:
            devir_tazele = kart_devir_tazele(DevirTazeleBody(kart_id=body.kart_id, uygula=True))
            devir_tazele = {"tazelenen": devir_tazele.get("tazelenen"),
                            "duzeltme": devir_tazele.get("toplam_duzeltme")}
        except Exception as e:  # noqa: BLE001
            logger.warning("ekstre sonrası devir tazeleme atlandı (yutuldu): %s", str(e)[:150])
            devir_tazele = {"ok": False, "hata": str(e)[:120]}

    return {
        "yazilan": yazilan,
        "devir_tazele": devir_tazele,
        "atlanan_veya_mevcut": atlanan,
        "atlanan_mevcut_adet": len(atlanan_mevcut),
        "atlanan_mevcut": atlanan_mevcut[:20],
        "motor_tahmini_faiz_iptal": motor_faizi_iptal,
        # Geriye-uyum: ekran bu alanı okuyor. Artık HER ZAMAN 0 — içe aktarma
        # gider satırı üretmiyor (bkz. yukarıdaki "ESKİ PRENSİP KALDIRILDI").
        "anlik_gider_yazilan": anlik_gider_yazilan,
        "gider_modeli": "kanonik",
        "gider_notu": ("Gider satırı üretilmedi — kart harcaması P&L'e kart "
                       "defterinden (gider_kanonik) girer. Çift sayım yapısal "
                       "olarak imkânsız."),
        "yeni_sistem_borc": round(yeni_borc, 2),
        "ap_rematch": ({
            "otomatik_baglanan": rematch.get("otomatik_baglanan"),
            "otomatik_tutar": rematch.get("otomatik_tutar"),
            "sahip_onayi_bekleyen": rematch.get("sahip_onayi_bekleyen"),
            "aday_tutar": rematch.get("aday_tutar"),
        } if isinstance(rematch, dict) and rematch.get("otomatik_baglanan") is not None
            else rematch),
    }


@app.get("/api/kartlar/cift-kayit-tarama")
def kart_cift_kayit_tarama(kart_id: Optional[str] = None):
    """ÇİFT YAZMA TARAMASI (salt-okur): elle girilmiş kayıt ile ekstre_import
    kopyası eşleşen çiftleri listeler (aynı kart+tip, tutar ±1, tarih ±5 gün).
    Temizlik ayrı uçtadır — bu uç hiçbir şeyi değiştirmez."""
    with db() as (_, cur):
        cur.execute(
            """SELECT e.id AS ekstre_id, e.tarih::text AS ekstre_tarih,
                      e.tutar::float AS tutar, e.islem_turu AS tip,
                      LEFT(COALESCE(e.aciklama,''),60) AS ekstre_aciklama,
                      man.id AS manuel_id, man.tarih::text AS manuel_tarih,
                      man.tutar::float AS manuel_tutar,
                      LEFT(COALESCE(man.aciklama,''),60) AS manuel_aciklama,
                      k.kart_adi
               FROM kart_hareketleri e
               JOIN kartlar k ON k.id = e.kart_id
               JOIN LATERAL (
                   SELECT id, tarih, tutar, aciklama FROM kart_hareketleri m2
                   WHERE m2.kart_id = e.kart_id AND m2.durum='aktif'
                     AND m2.islem_turu = e.islem_turu
                     AND COALESCE(m2.kaynak_tablo,'') <> 'ekstre_import'
                     AND m2.islem_turu <> 'DEVIR'
                     AND ABS(m2.tutar - e.tutar) <= 1.0
                     AND m2.tarih BETWEEN e.tarih - 5 AND e.tarih + 5
                   ORDER BY ABS(m2.tarih - e.tarih) LIMIT 1
               ) man ON TRUE
               WHERE e.kaynak_tablo = 'ekstre_import' AND e.durum='aktif'
                 AND e.islem_turu IN ('HARCAMA','ODEME','FAIZ')
                 AND (%s::text IS NULL OR e.kart_id = %s)
               ORDER BY k.kart_adi, e.tarih DESC""",
            (kart_id, kart_id))
        adaylar = [dict(r) for r in cur.fetchall() or []]
    gider_adaylar, gider_hata = [], None
    try:
        with db() as (_, cur):
            cur.execute(
                """SELECT g.id AS gider_id, g.tarih::text AS gider_tarih,
                          g.tutar::float AS tutar,
                          LEFT(COALESCE(g.aciklama,''),60) AS gider_aciklama,
                          man.id AS elle_id, man.tarih::text AS elle_tarih,
                          man.tutar::float AS elle_tutar,
                          LEFT(COALESCE(man.aciklama,''),60) AS elle_aciklama,
                          COALESCE(k.kart_adi,'?') AS kart_adi
                   FROM anlik_giderler g
                   LEFT JOIN kartlar k ON k.id = g.kart_id
                   JOIN LATERAL (
                       SELECT id, tarih, tutar, aciklama FROM anlik_giderler m2
                       WHERE COALESCE(m2.kaynak_tablo,'') <> 'ekstre_import'
                         AND m2.odeme_yontemi = 'kart'
                         AND (m2.kart_id = g.kart_id OR m2.kart_id IS NULL)
                         AND ABS(m2.tutar - g.tutar) <= 1.0
                         AND m2.tarih::date BETWEEN g.tarih::date - 7 AND g.tarih::date + 7
                       ORDER BY ABS(m2.tarih::date - g.tarih::date) LIMIT 1
                   ) man ON TRUE
                   WHERE g.kaynak_tablo = 'ekstre_import'
                     AND (%s::text IS NULL OR g.kart_id = %s)
                   ORDER BY g.tarih DESC""",
                (kart_id, kart_id))
            gider_adaylar = [dict(r) for r in cur.fetchall() or []]
    except Exception as e:
        gider_hata = str(e)[:120]
    return {"aday_sayisi": len(adaylar), "adaylar": adaylar,
            "gider_aday_sayisi": len(gider_adaylar), "gider_adaylar": gider_adaylar,
            "gider_tarama_hatasi": gider_hata,
            "not": "ADAY çiftler — kesin hüküm değil (aynı tutarlı iki meşru işlem "
                   "olabilir). Temizlikte varsayılan: ekstre_import KOPYASI iptal "
                   "edilir (elle kayıt gider/kasa zincirine bağlı olabilir). "
                   "POST /api/kartlar/cift-kayit-temizle {ekstre_ids:[...], uygula:true}"}


@app.post("/api/kartlar/cift-kayit-temizle")
def kart_cift_kayit_temizle(body: dict):
    """Seçilen ekstre_import KOPYALARINI iptal eder (kart borcu düşer, bağlı
    anlik_giderler kaydı silinir). Yalnız ekstre_import kaynaklı id kabul edilir —
    elle kayıtlara DOKUNAMAZ. uygula=false → önizleme."""
    ids = [str(x) for x in (body or {}).get("ekstre_ids") or [] if str(x).startswith("eks_")]
    gider_ids = [str(x) for x in (body or {}).get("gider_ids") or [] if str(x).startswith("agk_")]
    uygula = bool((body or {}).get("uygula"))
    if not ids and not gider_ids:
        raise HTTPException(400, "ekstre_ids (eks_) veya gider_ids (agk_) zorunlu")
    if gider_ids and uygula:
        with db() as (conn, cur):
            cur.execute(
                """DELETE FROM anlik_giderler
                   WHERE id = ANY(%s) AND kaynak_tablo='ekstre_import'""", (gider_ids,))
            conn.commit()
    if not ids:
        return {"uygula": uygula, "islenen": 0, "gider_silinen": len(gider_ids) if uygula else 0,
                "gider_onizleme": gider_ids}
    with db() as (conn, cur):
        cur.execute(
            """SELECT id, kart_id, tarih::text AS tarih, islem_turu, tutar::float AS tutar
               FROM kart_hareketleri
               WHERE id = ANY(%s) AND kaynak_tablo='ekstre_import' AND durum='aktif'""",
            (ids,))
        kayitlar = [dict(r) for r in cur.fetchall() or []]
        if uygula and kayitlar:
            kids = [r["id"] for r in kayitlar]
            cur.execute(
                """UPDATE kart_hareketleri
                   SET durum='iptal',
                       aciklama = COALESCE(aciklama,'') || ' [çift-kayıt temizliği]'
                   WHERE id = ANY(%s)""", (kids,))
            cur.execute(
                """DELETE FROM anlik_giderler
                   WHERE kaynak_tablo='ekstre_import' AND kaynak_id = ANY(%s)""", (kids,))
            conn.commit()
    return {"uygula": uygula, "islenen": len(kayitlar), "kayitlar": kayitlar}


@app.post("/api/kartlar/ekstre-import-anlik-gider-backfill")
def kart_ekstre_import_anlik_gider_backfill(kart_id: Optional[str] = None):
    """⛔ EMEKLİ (2026-08-10, Codex denetimi). Bu uç ESKİ MODELİ DİRİLTİYOR.

    Kanonik gider modelinde kart harcaması P&L'e KART DEFTERİNDEN girer; ekstre
    satırı için ayrıca anlik_giderler kaydı üretmek çift sayım demektir. Bu uç
    tam olarak onu yapıyordu — biri çağırırsa arşivlediğimiz 166 kaydın benzerini
    yeniden üretir ve gideri sessizce şişirirdi.

    Silinmedi (geçmiş çağrılar 404 yerine açıklama görsün) ama ARTIK YAZMIYOR.
    """
    raise HTTPException(
        410,
        "Bu uç emekli edildi (2026-08-10). Kanonik gider modelinde kart harcaması "
        "P&L'e kart defterinden girer; ekstre satırı için ayrı gider kaydı "
        "üretmek çift sayımdır. Ölçüm için: GET /api/abonelik/maliyet-cift-sayim")
    with db() as (conn, cur):
        params: list = []
        kart_filter = ""
        if kart_id:
            kart_filter = "AND kart_id = %s"
            params.append(kart_id)
        cur.execute(
            f"""
            SELECT id, kart_id, tarih, tutar, aciklama, kategori
            FROM kart_hareketleri
            WHERE kaynak_tablo = 'ekstre_import'
              AND islem_turu = 'HARCAMA'
              AND durum = 'aktif'
              AND COALESCE(harcama_tipi, 'belirsiz') != 'sahsi'
              {kart_filter}
            """,
            params,
        )
        eklenen = 0
        for r in (cur.fetchall() or []):
            r = dict(r)
            agid = "agk_" + r["id"]
            cur.execute(
                """INSERT INTO anlik_giderler
                   (id, tarih, kategori, tutar, aciklama, sube, odeme_yontemi, kart_id, kaynak_id, kaynak_tablo)
                   VALUES (%s,%s,%s,%s,%s,'MERKEZ','kart',%s,%s,'ekstre_import')
                   ON CONFLICT (id) DO NOTHING""",
                (agid, r["tarih"], (r.get("kategori") or "Diğer"), r["tutar"],
                 r.get("aciklama"), r["kart_id"], r["id"]),
            )
            if cur.rowcount > 0:
                eklenen += 1
    return {"eklenen": eklenen}


# ── ÖDEME PLANI ────────────────────────────────────────────────
class OdemePlani(BaseModel):
    kart_id: str
    tarih: date
    odenecek_tutar: float
    asgari_tutar: Optional[float] = None
    aciklama: Optional[str] = None

class KismiOdeModel(BaseModel):
    odenen_tutar: float
    kalan_vade_tarihi: date
    odeme_yontemi: str = 'nakit'  # 'nakit' veya 'kart'
    kart_id: Optional[str] = None

@app.post("/api/odeme-plani")
def odeme_plani_ekle(o: OdemePlani):
    with db() as (conn, cur):
        oid = str(uuid.uuid4())
        asgari = o.asgari_tutar or o.odenecek_tutar * 0.4
        cur.execute("""INSERT INTO odeme_plani (id,kart_id,tarih,odenecek_tutar,asgari_tutar,aciklama)
            VALUES (%s,%s,%s,%s,%s,%s)""",
            (oid, o.kart_id, o.tarih, o.odenecek_tutar, asgari, o.aciklama))
        onay_ekle(cur, 'ODEME_PLANI', 'odeme_plani', oid,
            f"Ödeme planı", o.odenecek_tutar, o.tarih)
        audit(cur, 'odeme_plani', oid, 'INSERT')
    return {"id": oid, "success": True}

class VadeliOdeModel(BaseModel):
    odeme_yontemi: str = 'nakit'  # 'nakit' veya 'kart'
    kart_id: Optional[str] = None
    # 🏪 PARAYI HANGİ ŞUBE ÇIKARDI (2026-08-18, sahip: "hangisinin kasasında
    # para varsa ondan ödüyordur"). Giderin SAHİBİ değil, ÖDEYENİ.
    # Ödeyen ≠ giderin şubesi olduğunda otomatik şube borcu doğar
    # (netleştirmeli — bkz. sube_ici_borc_api.capraz_odeme_borcu_kur).
    # Boş bırakılırsa BİLİNMİYOR kalır; uydurma atama YAPILMAZ.
    odeyen_sube_id: Optional[str] = None
    # 'elden' | 'havale' — banka mutabakatını besler; boşsa belirsiz kalır.
    nakit_yontemi: Optional[str] = None


def _personel_maas_odeme_guard(cur, plan: dict) -> None:
    """FAZ 0 #3: Personel maaş planı ödenmeden önce DÖNEMİN kaydı 'onaylandi' olmalı.
    ARREARS kuralı: plan.referans_ay = ÖDEME ayı → kayıt bir ÖNCEKİ ayda (çalışma dönemi)
    aranır (maas_service.referans_to_donem). FIX 2026-07-04: guard referans ayının
    kendisine bakıyordu → Haziran kaydı onaylıyken 'kayıt girilmedi' diyordu.
    Eski konvansiyon (referans=çalışma ayı) planları için referans ayının kendisi de
    kabul edilir. Sadece kaynak_tablo='personel'; kira/borç/kart etkilenmez."""
    if (plan.get('kaynak_tablo') or '') != 'personel':
        return
    kid = plan.get('kaynak_id')
    ref = plan.get('referans_ay')
    if not kid or not ref:
        return  # eski/bağlanamayan kayıt — engelleme
    d_yil, d_ay = _maas_svc.referans_to_donem(ref)
    cur.execute(
        """SELECT yil, ay, durum FROM personel_aylik
           WHERE personel_id=%s AND ((yil=%s AND ay=%s) OR (yil=%s AND ay=%s))
           ORDER BY (CASE WHEN yil=%s AND ay=%s THEN 0 ELSE 1 END)""",
        (kid, d_yil, d_ay, ref.year, ref.month, d_yil, d_ay),
    )
    rows = cur.fetchall()
    if any((r or {}).get('durum') == 'onaylandi' for r in rows):
        return
    cur.execute("SELECT ad_soyad FROM personel WHERE id=%s", (kid,))
    ad = (cur.fetchone() or {}).get('ad_soyad') or 'Personel'
    donem_ad = f"{_maas_svc.TR_AYLAR[d_ay]} {d_yil}"
    if not rows:
        raise HTTPException(
            400,
            f"{ad}: {donem_ad} dönemi maaş kaydı henüz girilmedi. Personel > Aylık Maaş'ta "
            f"{donem_ad} dönemini seçip 'Vardiyadan Aktar' + 'Onayla' yapın "
            f"(fazla mesai dahil edilip tutar kilitlensin).",
        )
    raise HTTPException(
        400,
        f"{ad}: {donem_ad} dönemi maaş kaydı '{rows[0].get('durum')}' durumda. Ödemeden "
        f"önce 'Onayla' gerekli (tutarı kilitler, fazla mesai dahil olur).",
    )


@app.post("/api/odeme-plani/{oid}/ode")
def odeme_yap(oid: str, tutar: Optional[float] = None, body: VadeliOdeModel = VadeliOdeModel()):
    with db() as (conn, cur):
        # FOR UPDATE: eş zamanlı iki ödeme isteğinin aynı planı çift işlemesini önler
        cur.execute("SELECT * FROM odeme_plani WHERE id=%s FOR UPDATE", (oid,))
        plan = cur.fetchone()
        if not plan: raise HTTPException(404)
        if plan['durum'] == 'odendi': raise HTTPException(400, "Zaten ödendi")

        # FAZ 0 #3: personel maaşı onaysız ödenemez
        _personel_maas_odeme_guard(cur, dict(plan))

        # ÇİFT ÖDEME KAPISI (ters yön): sabit gider planı ödenmeden önce, o ay manuel
        # /fatura-ode ile (nakit veya kart) zaten ödendiyse engelle.
        if plan.get('kaynak_tablo') == 'sabit_giderler' and plan.get('kaynak_id'):
            cur.execute(
                """
                SELECT 1 FROM kasa_hareketleri
                WHERE kaynak_tablo='sabit_giderler' AND kaynak_id=%s
                  AND islem_turu IN ('SABIT_GIDER','FATURA_ODEMESI')
                  AND kasa_etkisi=true AND durum='aktif'
                  AND DATE_TRUNC('month', tarih) = DATE_TRUNC('month', %s::date)
                UNION ALL
                SELECT 1 FROM kart_hareketleri
                WHERE kaynak_tablo='fatura_giderleri' AND kaynak_id=%s
                  AND islem_turu='HARCAMA' AND durum='aktif'
                  AND DATE_TRUNC('month', tarih) = DATE_TRUNC('month', %s::date)
                LIMIT 1
                """,
                (plan['kaynak_id'], plan.get('referans_ay'), plan['kaynak_id'], plan.get('referans_ay')),
            )
            if cur.fetchone():
                raise HTTPException(400, "Bu gider bu ay zaten manuel ödenmiş — plan tekrar ödenemez")

        # 🔴 P0 (2026-08-13, EVV-ODE denetimi): ?tutar gelmezse varsayılan
        # ORİJİNAL odenecek_tutar idi — kısmi ödenmiş planda (odenen>0) "tamamını
        # öde" KALAN yerine TAM tutarı İKİNCİ kez ödetiyordu (300 ödenmiş 1000'lik
        # planda 1000 daha çıkar, toplam 1300). Varsayılan artık KALAN borçtur.
        _kalan_varsayilan = round(
            float(plan['odenecek_tutar'] or 0) - float(plan.get('odenen_tutar') or 0), 2)
        if _kalan_varsayilan <= 0:
            raise HTTPException(400, "Bu planın kalan borcu yok — zaten ödenmiş görünüyor")

        # KART seçildiyse kart akışına yönlendir.
        # ⚠️ ESKİDEN yalnız `kaynak_tablo='vadeli_alimlar'` planda açılıyordu
        # (canlı hata, 2026-08-31): tedarikçi cari ödemesi (`cari_odeme`) kartla
        # yapılmak istendiğinde bu şart tutmuyor, dal ATLANIYOR ve ödeme sessizce
        # NAKİT yoluna düşüyordu. Sonuç: sahip "kartla ödedim" diyor, sistem
        # "✓ ödendi" diyor, ama kart borcu ARTMIYOR ve kasadan 120.000 ₺ çıkmış
        # görünüyordu. Kanal sessizce değişti, kimse fark etmedi.
        # Kart seçimi KAYNAKTAN BAĞIMSIZ bir karardır: kart_id verildiyse kart.
        if body.odeme_yontemi == 'kart' and body.kart_id and plan.get('kaynak_tablo') in (
                'vadeli_alimlar', 'cari_odeme'):
            bugun = str(bugun_tr())
            odeme_tutari = tutar or _kalan_varsayilan
            # Kart validasyon — FOR UPDATE: eş zamanlı limit aşımını önler
            cur.execute("SELECT * FROM kartlar WHERE id=%s AND aktif=TRUE FOR UPDATE", (body.kart_id,))
            kart = cur.fetchone()
            if not kart: raise HTTPException(404, "Kart bulunamadı")
            borc = kart_borc(cur, body.kart_id)
            kalan_limit = _kanonik_kalan_limit(body.kart_id, float(kart['limit_tutar']) - borc)
            if kalan_limit < odeme_tutari:
                # OPET vakası (2026-08-14): limit "yetersiz" görünmesinin en sık
                # sebebi gerçekten dolu olması değil, ÖDENMİŞ ekstrenin deftere
                # girilmemiş olması — borç düşmeyince limit de açılmıyor. Sahibe
                # ne yapacağını söyle (kör "yetersiz" mesajı yol göstermiyordu).
                _ipucu = ""
                try:
                    # ⚠️ PENCERE TANIMI: [kesim, sonraki_kesim) — finansal_duyu_api
                    # FIN_KART_ODEME_GIRILMEMIS (D1) ve kasa_service kart planı
                    # yazıcısıyla BİREBİR AYNI. Önceki hâli `tarih >= kesim` idi:
                    # SONRAKİ dönemlere yapılmış bir ödeme, son ekstre penceresi
                    # bomboş olsa bile ipucunu susturuyordu (Codex H1).
                    cur.execute("""
                        SELECT kesim_tarihi FROM kart_ekstre_donem
                         WHERE kart_id=%s AND kesim_tarihi IS NOT NULL
                         ORDER BY donem DESC LIMIT 1
                    """, (body.kart_id,))
                    _ked = cur.fetchone()
                    _kt = _ked and _ked['kesim_tarihi']
                    if _kt:
                        _sy, _sa = ((_kt.year + 1, 1) if _kt.month == 12
                                    else (_kt.year, _kt.month + 1))
                        _sonraki = kesim_tarihi_hesapla(_sy, _sa, int(kart['kesim_gunu'] or 1))
                        cur.execute("""
                            SELECT 1 FROM kart_hareketleri
                             WHERE kart_id=%s AND islem_turu='ODEME'
                               AND COALESCE(durum,'aktif')='aktif'
                               AND tarih >= %s::date AND tarih < %s::date
                             LIMIT 1
                        """, (body.kart_id, _kt, _sonraki))
                        _odeme_var = cur.fetchone() is not None
                    else:
                        _odeme_var = True     # ekstre yoksa hüküm verme (ipucu susar)
                    if not _odeme_var:
                        _ipucu = (" — Not: bu kartın son ekstre ödemesi kayıtlı görünmüyor; "
                                  "gerçekte ödediyseniz önce karta ödemeyi girin, limit açılır.")
                except Exception:  # noqa: BLE001 — ipucu asla asıl hatayı gölgelemesin
                    _ipucu = ""
                raise HTTPException(400, f"Kart limiti yetersiz. Kalan: {kalan_limit:,.0f} ₺{_ipucu}")
            # Kart harcaması ekle — kasaya yazma
            hid = str(uuid.uuid4())
            # ⚠️ KAYNAK SABİT YAZILIYORDU (Codex denetimi main.py:8593,
            # 2026-09-01): kart dalı `cari_odeme` için de açıldı ama kart
            # hareketi HÂLÂ 'vadeli_alimlar' diye ve "Vadeli alım: ..."
            # açıklamasıyla kaydediliyordu. Kaynak-bazlı izleme/geri alma
            # yanlış tabloda arar; rapor yanlış kanala yazar. Kaydın kaynağı
            # TAHMİN EDİLMEZ, plandan okunur.
            _kkaynak = str(plan.get('kaynak_tablo') or 'vadeli_alimlar')
            _kbaslik = ("Cari ödeme" if _kkaynak == 'cari_odeme' else "Vadeli alım")
            cur.execute("""
                INSERT INTO kart_hareketleri
                    (id, kart_id, tarih, islem_turu, tutar, taksit_sayisi, aciklama, kaynak_id, kaynak_tablo)
                VALUES (%s, %s, %s, 'HARCAMA', %s, 1, %s, %s, %s)
            """, (hid, body.kart_id, bugun, odeme_tutari,
                  f"{_kbaslik}: {plan['aciklama']}",
                  plan.get('kaynak_id'), _kkaynak))
            audit(cur, 'kart_hareketleri', hid, 'VADELI_KART')
            # Plan kapat — KISMİ ÖDEME destekli (2026-08-08): kartla borcun bir
            # kısmı ödenebilir; kalan varsa satır 'bekliyor' kalır ve söz açık
            # kalır. Eskiden tutar ne olursa olsun 'odendi' yazılıp vadeli alım
            # kapatılıyordu → kalan borç kayboluyordu.
            _kart_onceki = float(plan.get('odenen_tutar') or 0)
            _kart_toplam = round(_kart_onceki + odeme_tutari, 2)
            _kart_kalan = round(float(plan['odenecek_tutar']) - _kart_toplam, 2)
            _kart_tam = _kart_kalan <= 0.01
            cur.execute(
                """UPDATE odeme_plani
                     SET durum=%s, odeme_tarihi=%s, odenen_tutar=%s
                   WHERE id=%s""",
                ('odendi' if _kart_tam else 'bekliyor',
                 (bugun if _kart_tam else plan.get('odeme_tarihi')), _kart_toplam, oid))
            cur.execute("""UPDATE onay_kuyrugu SET durum='onaylandi', onay_tarihi=NOW()
                WHERE kaynak_id=%s AND durum NOT IN ('onaylandi','reddedildi')""", (oid,))
            # ⚠️ YALNIZ VADELİ ALIMDA KAPAT (Codex denetimi main.py:8613):
            # `vadeli_alim_kapat` kaynak ayırmadan çağrılıyordu. `cari_odeme`
            # planında `kaynak_id` bir CARİ ÖDEME kimliğidir — o kimlikle
            # vadeli tabloya gitmek ya hiçbir şey kapatmaz (sessiz yarım iş)
            # ya da aynı kimliğe denk gelen ALAKASIZ bir vadeli kaydı kapatır.
            if _kart_tam and plan.get('kaynak_id') and _kkaynak == 'vadeli_alimlar':
                vadeli_alim_kapat(cur, plan['kaynak_id'], bugun)
            audit(cur, 'odeme_plani', oid, 'ODENDI_KART' if _kart_tam else 'KISMI_ODEME_KART')
            # Uyarı önbelleğini temizle — panelde uyarı hemen kalksın
            uyari_cache_clear()
            try:
                from supplier_payment import spe_tetikle as _spe
                _spe("odeme_yap_kart")
            except Exception:  # noqa: BLE001
                pass
            return {"success": True, "odeme_yontemi": "kart",
                    "tam_kapandi": _kart_tam,
                    "odenen_toplam": _kart_toplam,
                    "kalan_borc": max(0.0, _kart_kalan),
                    "mesaj": ("Borç tamamen kapandı" if _kart_tam
                              else f"Kısmi ödeme — kalan {max(0.0, _kart_kalan):,.2f} TL devrediyor")}

        bugun = str(bugun_tr())
        odenen = tutar or _kalan_varsayilan   # P0 fix: varsayılan = KALAN, orijinal değil
        # ── KISMİ ÖDEME (2026-08-08, sahip: "bazen içerdeki borç tutarının bir
        # kısmını bırakırız, üstüne yeni borçlar eklenir; ödeme illa fatura
        # tutarı kadar olmaz"). Eski davranış: tutar ne olursa olsun durum
        # 'odendi' yazılıyor, vadeli alım kapatılıyordu → KALAN BORÇ KAYBOLUYORDU
        # (canlı: 73 plan satırı 'ödendi' damgalı ama ödenen < ödenecek).
        # Yeni: ödenen BİRİKİR; kalan varsa satır 'bekliyor' kalır ve tekrar
        # ödenebilir. Borç ancak tamamı kapanınca 'odendi' olur.
        _onceki = float(plan.get('odenen_tutar') or 0)
        _toplam_odenen = round(_onceki + odenen, 2)
        _kalan = round(float(plan['odenecek_tutar']) - _toplam_odenen, 2)
        _tam_kapandi = _kalan <= 0.01
        cur.execute(
            """UPDATE odeme_plani
                 SET durum=%s, odeme_tarihi=%s, odenen_tutar=%s
               WHERE id=%s""",
            ('odendi' if _tam_kapandi else 'bekliyor',
             (bugun if _tam_kapandi else plan.get('odeme_tarihi')),
             _toplam_odenen, oid))

        # 🏪 Ödeyen şube + nakit yöntemi TAŞINIR (2026-08-18). Boş gelirse eski
        # davranış birebir sürer (BİLİNMİYOR) — uydurma atama yapılmaz.
        ana_para_kismi = kasa_ve_faiz_odeme_plani_tam_odeme(
            cur, dict(plan), oid, odenen, bugun,
            odeme_yontemi=getattr(body, 'nakit_yontemi', None),
            odeyen_sube_id=getattr(body, 'odeyen_sube_id', None),
        )

        # Onay kuyruğunu kapat — PROD-PANEL-002 FIX: SADECE kaynak_id ile eşleşme tablo-arası
        # id çakışmasında yanlış onayları kapatabiliyordu. kaynak_tablo ile skopla + durum='bekliyor'
        # (NOT IN(...) 'iptal'/'iptal_revize' gibi kapalı durumları da yanlış süpürüyordu).
        if plan.get('kaynak_tablo') and plan.get('kaynak_id'):
            cur.execute("""UPDATE onay_kuyrugu SET durum='onaylandi', onay_tarihi=NOW()
                WHERE durum='bekliyor'
                  AND ( (kaynak_tablo='odeme_plani' AND kaynak_id=%s)
                     OR (kaynak_tablo=%s AND kaynak_id=%s) )""",
                (oid, plan['kaynak_tablo'], plan['kaynak_id']))
        else:
            cur.execute("""UPDATE onay_kuyrugu SET durum='onaylandi', onay_tarihi=NOW()
                WHERE durum='bekliyor' AND kaynak_tablo='odeme_plani' AND kaynak_id=%s""",
                (oid,))
        audit(cur, 'odeme_plani', oid, 'ODEME', eski=plan)

        # Kaynak vadeli_alimlar ise tüm bağlı kayıtları atomik kapat — çift düşme engeli
        # ⚠️ YALNIZ TAM ÖDEMEDE: kısmi ödemede söz açık kalmalı, yoksa kalan borç
        # kuyruktan silinir (2026-08-08 kısmi ödeme düzeltmesi).
        if _tam_kapandi and plan.get('kaynak_tablo') == 'vadeli_alimlar' and plan.get('kaynak_id'):
            vadeli_alim_kapat(cur, plan['kaynak_id'], bugun)

        guncelle_borc_envanteri_odeme_plani_sonrasi(cur, plan, ana_para_kismi)

        # Faiz üretimi: /api/kartlar/faiz-uret endpoint'i veya ay sonu startup ile otomatik

        # Uyarı önbelleğini temizle — panelde uyarı hemen kalksın
        uyari_cache_clear()

    # ⚡ Kanonik ödeme katmanını ANINDA hizala (2026-08-08)
    try:
        from supplier_payment import spe_tetikle
        spe_tetikle("odeme_yap")
    except Exception:  # noqa: BLE001
        pass
    return {"success": True,
            "tam_kapandi": _tam_kapandi,
            "odenen_toplam": _toplam_odenen,
            "kalan_borc": max(0.0, _kalan),
            "mesaj": ("Borç tamamen kapandı" if _tam_kapandi
                      else f"Kısmi ödeme — kalan {max(0.0, _kalan):,.2f} TL devrediyor")}

@app.delete("/api/odeme-plani/{oid}")
def odeme_plani_sil(oid: str):
    with db() as (conn, cur):
        cur.execute("SELECT * FROM odeme_plani WHERE id=%s", (oid,))
        eski = cur.fetchone()
        if not eski: raise HTTPException(404)
        cur.execute("UPDATE odeme_plani SET durum='iptal' WHERE id=%s", (oid,))
        cur.execute("UPDATE onay_kuyrugu SET durum='reddedildi' WHERE kaynak_id=%s", (oid,))
        # Eğer ödeme zaten "odendi" durumundaysa kasa geri alınmalı
        if eski['durum'] == 'odendi':
            # FIX O1 (2026-07-05): Nakit ödeme kasaya KAYNAK türüyle yazılır
            # (SABIT_GIDER/PERSONEL_MAAS/BORC_TAKSIT/VADELI_ODEME) — 'ODEME' türüyle DEĞİL.
            # Eski kod 'ODEME' arıyordu → iptal_kasa_hareketi kaydı BULAMIYOR → ters kayıt
            # hiç oluşmuyor, para kasadan çıkmış görünmeye devam ediyordu (append-only #5
            # fiilen deliniyordu). Artık kaynak_tablo'dan doğru işlem türü türetilir.
            kaynak = eski.get('kaynak_tablo') or ''
            if eski.get('kart_id'):
                # ⚠️ BU DAL DA TAHMİN EDİYORDU (Codex denetimi main.py:1480,
                # 2026-09-01): 'KART_ODEME' sabiti aranıyordu, oysa yazma
                # tarafı türü `kasa_islem_turu(kaynak_tablo)`dan üretiyor.
                # kaynak_tablo boş bir kart-ekstre planında yazılan 'ODEME',
                # aranan 'KART_ODEME' — iptal kaydı BULUNAMAZ ve para kasadan
                # çıkmış görünmeye DEVAM EDER. Bugün `cari_odeme` için
                # düzeltilen hatanın birebir ikizi. Kural her dalda aynı:
                # TÜRÜ TAHMİN ETME, DEFTERDEN OKU.
                for islem in _kasa_iptal_turleri(cur, oid, 'KART_ODEME'):
                    iptal_kasa_hareketi(
                        cur, oid, 'odeme_plani', islem, islem + '_IPTAL',
                        f"Ödeme iptali: {eski['aciklama']}")
            elif kaynak == 'vadeli_alimlar' and eski.get('kaynak_id'):
                # vadeli kasa kaydı kaynak_id=vadeli_id ile bağlı (plan_id değil)
                for islem in _kasa_iptal_turleri(
                        cur, eski['kaynak_id'], 'VADELI_ODEME'):
                    iptal_kasa_hareketi(
                        cur, eski['kaynak_id'], 'vadeli_alimlar',
                        islem, islem + '_IPTAL',
                        f"Ödeme iptali: {eski['aciklama']}")
            else:
                # ══════════════════════════════════════════════════════════
                # 🔎 TÜRÜ TAHMİN ETME — OKU (canlı hata, 2026-08-31)
                # ══════════════════════════════════════════════════════════
                # Tür haritadan TÜRETİLİYORDU; kaynak haritada yoksa yazma
                # tarafı bir varsayılan, iptal tarafı BAŞKA bir varsayılan
                # kullanıyor ve iptal kaydı bulamıyordu ("cari_odeme" vakası:
                # yazıldı KART_ODEME, arandı ODEME → 120.000 ₺ kasadan çıkmış
                # görünmeye devam etti). Haritayı düzeltmek yetmez: GEÇMİŞTE
                # farklı türle yazılmış satırlar da geri alınabilmeli.
                # Kaydın türü zaten defterde YAZILI — tahmin etmek yerine
                # okuyoruz. Birden çok tür varsa hepsi ayrı ayrı iptal edilir.
                # Tek okuyucu: _kasa_iptal_turleri (üç dal da AYNI kuralı
                # kullanır — aynı soru iki yerde ayrı hesaplanmaz).
                for islem in _kasa_iptal_turleri(
                        cur, oid, kasa_islem_turu(kaynak)):
                    iptal_kasa_hareketi(
                        cur, oid, 'odeme_plani', islem, islem + '_IPTAL',
                        f"Ödeme iptali: {eski['aciklama']}")
        audit(cur, 'odeme_plani', oid, 'IPTAL', eski=eski)
    return {"success": True}

# ── ONAY KUYRUGU ───────────────────────────────────────────────
@app.get("/api/onay-kuyrugu")
def onay_listele(durum: str = "bekliyor", limit: int = 300):
    d = (durum or "bekliyor").strip().lower()
    lim = max(1, min(int(limit or 300), 1000))
    with db() as (conn, cur):
        if d == "bekliyor":
            cur.execute(
                """
                SELECT *
                FROM onay_kuyrugu
                WHERE durum='bekliyor'
                ORDER BY tarih ASC, olusturma ASC
                LIMIT %s
                """,
                (lim,),
            )
        elif d == "gecmis":
            cur.execute(
                """
                SELECT *
                FROM onay_kuyrugu
                WHERE durum IN ('onaylandi','reddedildi')
                ORDER BY COALESCE(onay_tarihi, olusturma) DESC
                LIMIT %s
                """,
                (lim,),
            )
        elif d == "hepsi":
            cur.execute(
                """
                SELECT *
                FROM onay_kuyrugu
                ORDER BY
                    CASE WHEN durum='bekliyor' THEN 0 ELSE 1 END,
                    COALESCE(onay_tarihi, olusturma) DESC
                LIMIT %s
                """,
                (lim,),
            )
        else:
            raise HTTPException(400, "durum: bekliyor | gecmis | hepsi")
        return [dict(r) for r in cur.fetchall()]


def _onayla_tx(cur, oid: str):
    cur.execute("SELECT * FROM onay_kuyrugu WHERE id=%s FOR UPDATE", (oid,))
    onay = cur.fetchone()
    if not onay:
        raise HTTPException(404)
    # Zaten onaylanmış — çift onay engeli
    if onay['durum'] != 'bekliyor':
        raise HTTPException(400, f"Bu işlem zaten '{onay['durum']}' durumunda, tekrar onaylanamaz.")
    tutar = float(onay['tutar'])
    tarih = str(onay['tarih'])
    GIDER_TURLERI = {'KART_ODEME', 'ANLIK_GIDER', 'VADELI_ODEME', 'PERSONEL_MAAS', 'PERSONEL_AVANS', 'SABIT_GIDER', 'BORC_TAKSIT', 'FATURA_ODEMESI', 'ODEME_PLANI'}
    GELIR_TURLERI = {'CIRO', 'CIRO_DUZELTME', 'DIS_KAYNAK', 'KASA_GIRIS', 'KASA_DUZELTME'}
    islem_turu = onay['islem_turu']
    KASA_FARK_TURLERI = {'KAPANIS_KASA_FARK', 'ACILIS_KASA_FARK'}
    # 🔗 PARASIZ ONAY TÜRLERİ (2026-08-15, F5): bunlar KASAYA HİÇ DOKUNMAZ —
    # onay bir BAĞ kurar (fatura↔teslimat), para hareketi doğurmaz. Ayrı küme
    # olmasının sebebi aşağıdaki "bilinmeyen işlem türü" uyarısıdır: meşru bir
    # tür sahte uyarı üretmemeli (sahte alarm, alarm körlüğünün ilk adımıdır).
    BAG_TURLERI = {'TESLIMAT_FATURA_ESLESME'}
    if islem_turu in BAG_TURLERI:
        signed_tutar = 0.0          # kasa yazımı YOK; aşağıdaki dal para geçirmez
    elif islem_turu in GIDER_TURLERI:
        signed_tutar = -abs(tutar)
    elif islem_turu in GELIR_TURLERI:
        signed_tutar = abs(tutar)
    elif islem_turu in KASA_FARK_TURLERI:
        signed_tutar = tutar  # işaret korunur; kasaya yazılmayacak
    else:
        signed_tutar = tutar
        logger.warning(f"Bilinmeyen işlem türü onaylandı: {islem_turu}, tutar={tutar}")

    # 🔒 SAVUNMA (2026-08-13, EVV-YUK): kaynağı PASİFE alınmış SABIT_GIDER onayı
    # kasadan para çıkarmasın — sil-yolu artık bekleyen onayı iptal ediyor ama
    # eski/yarış kalıntısı kayıtlar için ikinci kemer.
    if islem_turu == 'SABIT_GIDER' and (onay.get('kaynak_tablo') or '') == 'sabit_giderler' and onay.get('kaynak_id'):
        cur.execute("SELECT aktif FROM sabit_giderler WHERE id=%s", (onay['kaynak_id'],))
        _sg = cur.fetchone()
        if _sg is not None and _sg.get('aktif') is False:
            raise HTTPException(400, "Bu gider kapatılmış — onaylanamaz (kayıt iptal edildi sayın, listeyi yenileyin)")

    # ODEME_PLANI onaylandığında kasa_etkisi True olmalı
    # Plan oluşumu = niyet (False), onay = gerçekleşme (True)
    # islem_turu değişmez — anlam korunur, sadece davranış eklenir
    if islem_turu == 'ODEME_PLANI':
        # Önce planı kapat; kasa yalnız plan gerçekten kapatıldıysa (/ode ile aynı — çift kasa önlemi)
        cur.execute("SELECT * FROM odeme_plani WHERE id=%s", (onay['kaynak_id'],))
        plan_row = cur.fetchone()
        if not plan_row:
            raise HTTPException(404, "Ödeme planı bulunamadı")
        plan_dict = dict(plan_row)
        kaynak_tablo = plan_dict.get('kaynak_tablo')
        odenen_onay = float(onay['tutar'])
        # 🔴 P1 (2026-08-12, Ödeme modülü denetimi): onay yolu planı KOŞULSUZ
        # 'odendi' yazıp odenen_tutar'ı EZİYORDU (önceki kısmi ödemeyi biriktirmeden,
        # _kalan kontrolü olmadan) → kalan borç KAYBOLUYORDU. Bu, /ode'de 2026-08-08'de
        # düzeltilen "kısmi ödeme kalanı korur" invariant'ının onay-yolundaki ikiziydi.
        # Artık /ode ile AYNI partial-aware mantık: biriktir + _kalan; borç ancak
        # tamamı kapanınca 'odendi' olur, kalan varsa 'bekliyor' kalır.
        _onceki_od = float(plan_dict.get('odenen_tutar') or 0)
        _toplam_od = round(_onceki_od + odenen_onay, 2)
        _kalan_od = round(float(plan_dict['odenecek_tutar']) - _toplam_od, 2)
        _tam_kapandi_onay = _kalan_od <= 0.01
        cur.execute("""
            UPDATE odeme_plani SET durum=%s, odeme_tarihi=%s, odenen_tutar=%s
            WHERE id=%s AND durum IN ('bekliyor','onay_bekliyor')
        """, ('odendi' if _tam_kapandi_onay else 'bekliyor',
              tarih if _tam_kapandi_onay else plan_dict.get('odeme_tarihi'),
              _toplam_od, onay['kaynak_id']))
        plan_guncellendi = cur.rowcount > 0
        if plan_guncellendi:
            # Kasa BU ödeme kadar yazılır (para gerçekten çıktı — tam/kısmi fark etmez,
            # /ode ile aynı). vadeli kapatma + borç envanteri YALNIZ tam kapanışta.
            ana_onay = kasa_ve_faiz_odeme_plani_tam_odeme(
                cur, plan_dict, onay['kaynak_id'], odenen_onay, tarih,
                anapara_aciklama=f"Onaylandı: {onay['aciklama']}",
            )
            if _tam_kapandi_onay and kaynak_tablo == 'vadeli_alimlar' and plan_dict.get('kaynak_id'):
                vadeli_alim_kapat(cur, plan_dict['kaynak_id'], tarih)
            guncelle_borc_envanteri_odeme_plani_sonrasi(cur, plan_dict, ana_onay)
    elif islem_turu == 'VADELI_ODEME':
        # Eşzamanlı iki onayın aynı vadeli kaydı çift düşmesini engelle.
        cur.execute("SELECT id FROM vadeli_alimlar WHERE id=%s FOR UPDATE", (onay['kaynak_id'],))
        if not cur.fetchone():
            raise HTTPException(404, "Vadeli alım kaydı bulunamadı")
        # ÇİFT ÖDEME GUARD: Kısmi ödeme + tam ödeme farklı kaynak_id ile tutulabildi — tek yerden topla
        onceki_odenen = vadeli_kasadan_odenen_toplam(cur, onay['kaynak_id'])
        # FIX O2/B4 (2026-07-05): "kalan kadar yaz" yorumu vardı ama kod TAM tutarı
        # ikinci kez yazıyordu → kısmi ödenmiş (0 < önceki < tutar) vadelide ÇİFT DÜŞME.
        # Artık gerçekten KALAN kadar yazılır.
        kalan_yazilacak = abs(signed_tutar) - onceki_odenen
        if kalan_yazilacak <= 0.01:
            logger.warning(f"VADELI_ODEME çift ödeme engellendi — kaynak_id={onay['kaynak_id']}")
            # Kasa zaten yazılmış, sadece onay kuyruğunu kapat ve tabloları güncelle
        else:
            insert_kasa_hareketi(cur, tarih, islem_turu, -abs(kalan_yazilacak),
                f"Onaylandı{' (kalan)' if onceki_odenen > 0.01 else ''}: {onay['aciklama']}",
                onay['kaynak_tablo'], onay['kaynak_id'], ref_id=oid, ref_type='ONAY')
        # Tüm bağlı kayıtları atomik kapat — çift düşme engeli
        vadeli_alim_kapat(cur, onay['kaynak_id'], tarih)
    elif (
        islem_turu == "ANLIK_GIDER"
        and (onay.get("kaynak_tablo") or "") == "anlik_giderler"
        and onay.get("kaynak_id")
    ):
        kid = str(onay["kaynak_id"])
        cur.execute(
            "SELECT id, durum FROM anlik_giderler WHERE id=%s FOR UPDATE",
            (kid,),
        )
        ag = cur.fetchone()
        if not ag:
            raise HTTPException(404, "Anlık gider kaydı bulunamadı")
        st = ag["durum"]
        if st == "onay_bekliyor":
            cur.execute(
                "UPDATE anlik_giderler SET durum='aktif' WHERE id=%s",
                (kid,),
            )
        elif st != "aktif":
            raise HTTPException(
                400,
                f"Anlık gider bu durumda onaylanamaz: {st}",
            )
        cur.execute(
            """
            SELECT COALESCE(COUNT(*), 0)::int AS n
            FROM kasa_hareketleri
            WHERE kaynak_id=%s AND islem_turu='ANLIK_GIDER'
              AND durum='aktif' AND kasa_etkisi=true
            """,
            (kid,),
        )
        n = int((cur.fetchone() or {}).get("n") or 0)
        if n == 0:
            insert_kasa_hareketi(
                cur,
                tarih,
                islem_turu,
                signed_tutar,
                f"Onaylandı: {onay['aciklama']}",
                "anlik_giderler",
                kid,
                ref_id=oid,
                ref_type="ONAY",
            )

        # ── RAPOR CACHE HOOK ── anlık gider onaylanınca özet güncellensin
        try:
            cur.execute("SELECT sube, tarih FROM anlik_giderler WHERE id=%s", (kid,))
            _ag = cur.fetchone()
            if _ag and _ag.get("sube"):
                from rapor_cache import gunluk_ozet_yenile
                gunluk_ozet_yenile(cur, str(_ag["sube"]), _ag.get("tarih"), kaynak='event_gider')
        except Exception:
            pass
    elif islem_turu in KASA_FARK_TURLERI:
        # Kasa farkı onayı = "Merkez gördü ve kabul etti" — kasa bakiyesini ETKİLEMEZ.
        # Gerçek fiziksel açık varsa Operasyon Merkezi → Kasa Uyumsuzluğu → "Gerçek Açık" akışı kullanılır.
        # Bağlı sube_operasyon_uyari satırını da çözüldü olarak işaretle — aksi halde
        # onay kuyruğundaki onay CFO/Merkez panelindeki uyarıya yansımıyordu.
        if (onay.get('kaynak_tablo') or '') == 'sube_operasyon_uyari' and onay.get('kaynak_id'):
            cur.execute(
                """
                UPDATE sube_operasyon_uyari
                SET okundu=TRUE,
                    cozum_notu=COALESCE(cozum_notu, %s),
                    cozum_ts=COALESCE(cozum_ts, NOW()),
                    cozum_personel_ad=COALESCE(cozum_personel_ad, %s)
                WHERE id=%s
                """,
                ("Onay kuyruğundan onaylandı (Merkez)", "Merkez (Onay Kuyruğu)", onay['kaynak_id']),
            )
    elif islem_turu == 'TESLIMAT_FATURA_ESLESME':
        # ── F5 · ÖNERİ ONAYI → FATURA ↔ TESLİMAT BAĞI (2026-08-15) ───────────
        # Öneriyi yazan: belge_talep_api.teslimat_fatura_oneri_tara (yükleme anı
        # + gece koşusu). Burada YALNIZ sahibin "evet"i uygulanır.
        #
        # ⚠️ PARA YOK: bu onay kasaya HİÇBİR ŞEY yazmaz — bağ kurmak ödeme değildir
        # (fatura borcu zaten kendi hattından yürüyor). Bu yüzden dal `insert_kasa_hareketi`
        # ÇAĞIRMAZ ve aşağıdaki genel `else` bloğuna DÜŞMEZ.
        #
        # ⚠️ HTTP ÇAĞRISI YOK: bağ, aynı transaction içinde iç fonksiyonla kurulur.
        # Uç üzerinden çağrılsaydı bağ ayrı bir tx'te commit olur, onay burada
        # düşerse "bağ kuruldu ama onay bekliyor" hayaleti kalırdı.
        #
        # ⚠️ Guard'lar AYNEN geçerli (tarih yönü / çoklu aday / zaten bağlı):
        # fatura_bagla_uygula içinde dururlar. Öneri yazılırken de aynı eşikler
        # uygulandığı için normalde tetiklenmezler; arada durum değiştiyse
        # (fatura başkasına bağlandı) onay 4xx ile REDDEDİLİR — sessiz geçmez.
        from belge_talep_api import fatura_bagla_uygula, _oneri_faturasini_coz
        tid_bt = str(onay.get('kaynak_id') or '').strip()
        fid_bt = _oneri_faturasini_coz(onay.get('aciklama'))
        if not tid_bt or not fid_bt:
            raise HTTPException(
                400, "Eşleşme önerisi okunamadı — açıklamadaki [fatura:<id>] damgası yok. "
                     "Bu öneri elle kapatılmalı; bağı Belge Merkezi'nden kurun.")
        fatura_bagla_uygula(cur, tid_bt, fid_bt, onay_kaynagi='oneri-onayi')
    elif islem_turu in ("CIRO", "CIRO_DUZELTME"):
        # Ciro kaynak kaydı varsa satırı kilitleyerek eşzamanlı onay/yazım çakışmasını azalt.
        if (onay.get("kaynak_tablo") or "") == "ciro" and onay.get("kaynak_id"):
            cur.execute("SELECT id FROM ciro WHERE id=%s FOR UPDATE", (onay["kaynak_id"],))
        insert_kasa_hareketi(cur, tarih, islem_turu, signed_tutar,
            f"Onaylandı: {onay['aciklama']}", onay['kaynak_tablo'], onay['kaynak_id'],
            ref_id=oid, ref_type='ONAY')
    else:
        # Maaş/sabit/borç taksit onayında çift ödeme riskini kapat:
        # aynı kaynak için aynı ayda aktif kasa kaydı varsa tekrar yazma.
        if islem_turu in ("SABIT_GIDER", "BORC_TAKSIT", "PERSONEL_MAAS") and onay.get("kaynak_id"):
            kaynak_tablo = (onay.get("kaynak_tablo") or "").strip().lower()
            kid = str(onay["kaynak_id"])
            if kaynak_tablo == "personel":
                cur.execute("SELECT id FROM personel WHERE id=%s FOR UPDATE", (kid,))
            elif kaynak_tablo == "sabit_giderler":
                cur.execute("SELECT id FROM sabit_giderler WHERE id=%s FOR UPDATE", (kid,))
            elif kaynak_tablo == "borc_envanteri":
                cur.execute("SELECT id FROM borc_envanteri WHERE id=%s FOR UPDATE", (kid,))

            cur.execute(
                """
                SELECT COALESCE(COUNT(*), 0)::int AS n
                FROM kasa_hareketleri
                WHERE kaynak_id=%s
                  AND islem_turu=%s
                  AND durum='aktif'
                  AND kasa_etkisi=true
                  AND DATE_TRUNC('month', tarih) = DATE_TRUNC('month', %s::date)
                """,
                (kid, islem_turu, tarih),
            )
            onceki = int((cur.fetchone() or {}).get("n") or 0)
            if onceki > 0:
                raise HTTPException(409, f"{islem_turu} için bu ay ödeme zaten işlenmiş.")

        insert_kasa_hareketi(cur, tarih, islem_turu, signed_tutar,
            f"Onaylandı: {onay['aciklama']}", onay['kaynak_tablo'], onay['kaynak_id'],
            ref_id=oid, ref_type='ONAY')
        # SABIT_GIDER / BORC_TAKSIT: bağlı odeme_plani'nı odendi yap — yuk_7'den çıksın.
        # KÖK DÜZELTME: önceden "aynı ay" (DATE_TRUNC) eşleşmesi kullanılıyordu; ödeme
        # planın vade ayından farklı ayda yapılınca (geç/erken ödeme) hiçbir plan
        # kapanmıyor, panelde "vadesi geçmiş" kalıyordu. Artık ay'dan bağımsız olarak
        # bu kaynağın EN ESKİ açık planı kapatılır (bir ödeme = en eski yükümlülük).
        if islem_turu in ('SABIT_GIDER', 'BORC_TAKSIT', 'PERSONEL_MAAS'):
            cur.execute("""
                UPDATE odeme_plani SET durum='odendi', odeme_tarihi=%s
                WHERE id = (
                    SELECT id FROM odeme_plani
                    WHERE kaynak_tablo=%s AND kaynak_id=%s
                    AND durum IN ('bekliyor','onay_bekliyor')
                    ORDER BY tarih ASC
                    LIMIT 1
                )
            """, (tarih, onay['kaynak_tablo'], onay['kaynak_id']))
    # Onay durumunu güncelle — vadeli_alim_kapat bazı kayıtları önceden onaylanmış yapabilir
    cur.execute("UPDATE onay_kuyrugu SET durum='onaylandi', onay_tarihi=NOW() WHERE id=%s AND durum='bekliyor'", (oid,))
    if cur.rowcount == 0:
        cur.execute("SELECT durum FROM onay_kuyrugu WHERE id=%s", (oid,))
        st = cur.fetchone()
        if not st or st['durum'] != 'onaylandi':
            raise HTTPException(409, "Eş zamanlı onay çakışması — işlem zaten onaylandı.")
    audit(cur, 'onay_kuyrugu', oid, 'ONAYLANDI', eski=onay)
    return {"success": True}


@app.post("/api/onay-kuyrugu/toplu-onayla",
          dependencies=[Depends(merkez_mutasyon_korumasi)])
def toplu_onayla(body: dict, x_evvel_oturum: Optional[str] = Header(default=None)):
    _ak_ad, _ak_kaynak = aktor_bilgisi(x_evvel_oturum)
    """
    Seçili onayları tek seferde onayla.
    body: { ids: [id1, id2, ...] }
    Her onay kendi transaction'ında işlenir — biri başarısız olursa diğerleri etkilenmez.
    """
    ids = body.get('ids', [])
    if not ids:
        raise HTTPException(400, "Onay listesi boş")

    sonuclar = []
    with db() as (conn, cur):
        # R5 (2026-08-12): id'leri KANONİK sırada işle — iki eşzamanlı toplu onay aynı
        # id'leri farklı sırada kilitleyip deadlock'a girmesin. (Sonuç yine per-item.)
        for i, oid in enumerate(sorted(ids)):
            sp = f"sp_toplu_onay_{i}"
            cur.execute(f"SAVEPOINT {sp}")
            try:
                _onayla_tx(cur, oid)
                cur.execute(f"RELEASE SAVEPOINT {sp}")
                sonuclar.append({"id": oid, "durum": "onaylandi"})
            except HTTPException as e:
                cur.execute(f"ROLLBACK TO SAVEPOINT {sp}")
                cur.execute(f"RELEASE SAVEPOINT {sp}")
                sonuclar.append({"id": oid, "durum": "hata", "mesaj": str(e.detail)})
            except Exception as e:
                cur.execute(f"ROLLBACK TO SAVEPOINT {sp}")
                cur.execute(f"RELEASE SAVEPOINT {sp}")
                sonuclar.append({"id": oid, "durum": "hata", "mesaj": str(e)})

    onaylanan = sum(1 for s in sonuclar if s["durum"] == "onaylandi")
    return {
        "toplam": len(ids),
        "onaylanan": onaylanan,
        "hata": len(ids) - onaylanan,
        "sonuclar": sonuclar,
    }

# 🔴 ONAY-002 (2026-09-02): bu üç uç PARA TAŞIR (onay → kasaya/karta yazım)
# ama hiçbirinde kapı yoktu. `merkez_mutasyon_korumasi` kodda vardı ve
# sube_panel.py'de kullanılıyordu — main.py onu import bile etmiyordu.
# ⚠️ DÜRÜST SINIR: bu guard yalnız EVVEL_MERKEZ_MUTASYON_ANAHTARI ortamda
# TANIMLIYSA ısırır; tanımsızsa davranış değişmez. Yani bu, kapıyı takmaktır —
# kapıyı KİLİTLEMEK sahibin anahtarı tanımlamasına bağlı (bkz. güvenlik backlog).
# Kapıdan bağımsız olarak aktör artık HER onayda deftere yazılıyor: jeton
# geçerliyse 'oturum', değilse 'anonim' — cevapsız soru cevapsız görünsün.
@app.post("/api/onay-kuyrugu/{oid}/onayla",
          dependencies=[Depends(merkez_mutasyon_korumasi)])
def onayla(oid: str, x_evvel_oturum: Optional[str] = Header(default=None)):
    _ak_ad, _ak_kaynak = aktor_bilgisi(x_evvel_oturum)
    with db() as (conn, cur):
        _s = _onayla_tx(cur, oid)
        audit(cur, 'onay_kuyrugu', oid, 'ONAYLANDI',
              aktor=_ak_ad, aktor_kaynak=_ak_kaynak)
        return _s

class ReddetModel(BaseModel):
    neden: str = 'hata'  # 'hata' veya 'surec_bitti'

@app.post("/api/onay-kuyrugu/{oid}/reddet",
          dependencies=[Depends(merkez_mutasyon_korumasi)])
def reddet(oid: str, body: ReddetModel = ReddetModel(),
           x_evvel_oturum: Optional[str] = Header(default=None)):
    _ak_ad, _ak_kaynak = aktor_bilgisi(x_evvel_oturum)
    with db() as (conn, cur):
        # FIX O5 (2026-07-06): zaten 'onaylandi' (kasaya/karta yazılmış) bir onay reddet ile
        # sessizce iptal edilirse plan iptal olur ama kasa izi kalır → ters kayıt zinciri delinir
        # (dokunulmaz #5). Reddet yalnız bekleyen onaylar içindir; onaylanmışı geri almak için
        # ilgili kaydın iptal/ters-kayıt akışı kullanılmalı (o akış kasadan da düşer).
        # R2 (2026-08-12, Onay denetimi): satırı FOR UPDATE KİLİTLE — eşzamanlı _onayla_tx
        # parayı taşırken reddet'in planı iptal edip sessizce success dönmesini engelle;
        # rowcount doğrula (yarış/çift-ret false-success'i kapat).
        cur.execute("SELECT * FROM onay_kuyrugu WHERE id=%s FOR UPDATE", (oid,))
        _kilit = cur.fetchone()
        if not _kilit:
            raise HTTPException(404, "Onay bulunamadı")
        _durum = _kilit.get("durum") if isinstance(_kilit, dict) else _kilit["durum"]
        if _durum == "onaylandi":
            raise HTTPException(400, "Bu onay zaten onaylanmış (kasaya işlenmiş olabilir). Reddetmek yerine ilgili kaydı iptal/ters-kayıt akışından geri alın — o akış kasa izini de düzeltir.")
        if _durum != "bekliyor":
            raise HTTPException(409, f"Onay '{_durum}' durumunda — reddedilemez (eşzamanlı değişim olabilir).")
        cur.execute("UPDATE onay_kuyrugu SET durum='reddedildi', onay_tarihi=NOW() WHERE id=%s AND durum='bekliyor'", (oid,))
        if cur.rowcount == 0:
            raise HTTPException(409, "Eş zamanlı onay/ret çakışması — durum değişti.")
        # R3 (2026-08-12): reddet artık audit'lenir (onay'da vardı, reddet'te yoktu).
        audit(cur, 'onay_kuyrugu', oid, 'REDDEDILDI', eski=_kilit)

        onay = _kilit
        if onay:
            kt  = onay.get("kaynak_tablo") or ""
            kid = onay.get("kaynak_id") or ""

            if kt == "anlik_giderler" and kid and onay.get("islem_turu") == "ANLIK_GIDER":
                cur.execute(
                    "UPDATE anlik_giderler SET durum='reddedildi' WHERE id=%s AND durum='onay_bekliyor'",
                    (kid,),
                )
            # R1 (2026-08-12, Onay denetimi): eskiden `WHERE id=%s OR kaynak_id=%s`
            # SKOPSUZDU → kaynak_id paylaşan TÜM kardeş planları (ör. bir borcun tüm
            # taksitleri) iptal ediyordu. _onayla_tx TEK (en eski) planı işler → reddet
            # de simetrik olmalı:
            #   • neden='surec_bitti' → kaynak tamamen kapanıyor: TÜM açık plan iptal + kaynak pasif
            #   • neden='hata'        → yalnız BU onaya karşılık gelen TEK plan
            #                           (kid bir plan id ise o, değilse (kt,kid) için en eski)
            if body.neden == 'surec_bitti' and kt and kid:
                cur.execute("""UPDATE odeme_plani SET durum='iptal'
                    WHERE kaynak_tablo=%s AND kaynak_id=%s AND durum IN ('bekliyor','onay_bekliyor')""",
                    (kt, kid))
                if kt == 'sabit_giderler':
                    cur.execute("UPDATE sabit_giderler SET aktif=FALSE WHERE id=%s", (kid,))
                elif kt == 'personel':
                    cur.execute("UPDATE personel SET aktif=FALSE WHERE id=%s", (kid,))
                elif kt == 'borc_envanteri':
                    cur.execute("UPDATE borc_envanteri SET aktif=FALSE WHERE id=%s", (kid,))
            elif kid:
                cur.execute("""UPDATE odeme_plani SET durum='iptal'
                    WHERE id = COALESCE(
                        (SELECT id FROM odeme_plani WHERE id=%s AND durum IN ('bekliyor','onay_bekliyor')),
                        (SELECT id FROM odeme_plani WHERE kaynak_tablo=%s AND kaynak_id=%s
                           AND durum IN ('bekliyor','onay_bekliyor') ORDER BY tarih ASC LIMIT 1))""",
                    (kid, kt, kid))

            # ── Şube bildirimi ─────────────────────────────────
            # Kaynak tablodan sube_id'yi bul
            _sube_id = None
            _sube_lookup = {
                "anlik_giderler":  "SELECT sube     AS sid FROM anlik_giderler WHERE id=%s",
                "siparis_talep":   "SELECT sube_id  AS sid FROM siparis_talep  WHERE id=%s",
                "ciro":            "SELECT sube_id  AS sid FROM ciro            WHERE id=%s",
                "anlik_gider":     "SELECT sube     AS sid FROM anlik_giderler WHERE id=%s",
                "odeme_plani":     "SELECT sube_id  AS sid FROM odeme_plani    WHERE id=%s",
            }
            _q = _sube_lookup.get(kt)
            if _q and kid:
                try:
                    cur.execute(_q, (kid,))
                    _r = cur.fetchone()
                    _sube_id = str(_r["sid"]) if _r and _r.get("sid") else None
                except Exception:
                    pass

            if _sube_id:
                _aciklama = (onay.get("aciklama") or "").strip()
                _tutar    = onay.get("tutar")
                _neden    = (body.neden or "").strip() if body.neden else ""
                _mesaj_parcalari = [f"❌ Onay reddedildi: {_aciklama}" if _aciklama else "❌ Onay talebiniz reddedildi."]
                if _tutar:
                    _mesaj_parcalari.append(f"Tutar: {float(_tutar):,.2f} ₺")
                if _neden:
                    _mesaj_parcalari.append(f"Neden: {_neden}")
                _mesaj = " — ".join(_mesaj_parcalari)
                try:
                    cur.execute(
                        """
                        INSERT INTO sube_merkez_mesaj
                            (id, sube_id, mesaj, oncelik, ttl_saat)
                        VALUES (%s, %s, %s, 'yuksek', 48)
                        """,
                        (str(uuid.uuid4()), _sube_id, _mesaj),
                    )
                except Exception:
                    pass

    return {"success": True}

# ── CİRO ───────────────────────────────────────────────────────
class CiroModel(BaseModel):
    tarih: date
    sube_id: str
    nakit: float = 0
    pos: float = 0
    online: float = 0
    aciklama: Optional[str] = None
    force: bool = False

@app.get("/api/ciro")
def ciro_listele(limit: int = 200, ay: str = None):
    import re as _re_ay
    ay_v = (ay or "").strip()
    ay_cond, ay_params = "", []
    if ay_v and ay_v.lower() != "hepsi" and _re_ay.match(r"^\d{4}-\d{2}$", ay_v):
        ay_cond = " AND to_char(c.tarih, 'YYYY-MM') = %s"
        ay_params = [ay_v]
    with db() as (conn, cur):
        cur.execute(f"""
            SELECT
                c.*,
                s.ad as sube_adi,
                COALESCE(s.pos_oran, 0) as pos_oran,
                COALESCE(s.online_oran, 0) as online_oran,
                ROUND(c.pos    * COALESCE(s.pos_oran,    0) / 100.0, 2) as pos_kesinti,
                ROUND(c.online * COALESCE(s.online_oran, 0) / 100.0, 2) as online_kesinti,
                ROUND(c.pos    * COALESCE(s.pos_oran,    0) / 100.0 +
                      c.online * COALESCE(s.online_oran, 0) / 100.0, 2) as toplam_yanan
            FROM ciro c
            LEFT JOIN subeler s ON s.id = c.sube_id
            WHERE c.durum = 'aktif'{ay_cond}
            ORDER BY c.tarih DESC
            LIMIT %s
        """, ay_params + [limit])
        return [dict(r) for r in cur.fetchall()]

@app.post("/api/ciro")
def ciro_ekle(c: CiroModel):
    nakit = float(c.nakit or 0)
    pos   = float(c.pos or 0)
    online = float(c.online or 0)
    toplam = nakit + pos + online
    # 🔴 P1 (2026-08-12, Para modülü denetimi): negatif/sıfır ciro engeli yoktu →
    # negatif net_tutar kasaya girebilir ya da anlamsız 0 ciro yazılabilirdi.
    # (0 satış = KAYIT YOK; "girilmemiş ciro" alarmı yokluğu yakalar.)
    if nakit < 0 or pos < 0 or online < 0:
        raise HTTPException(400, "Ciro alanları negatif olamaz")
    if toplam <= 0:
        raise HTTPException(400, "Ciro toplamı pozitif olmalı")
    # 🟡 P3 (2026-08-13, EVV-PARA-N15): gelecek tarihli ciro anlamsız — henüz
    # yaşanmamış günün satışı kasaya girip tüm gün bazlı mutabakatları bozar.
    # Geçmiş gün düzeltmesi meşru ve serbest kalır.
    if c.tarih > bugun_tr():
        raise HTTPException(400, "Gelecek tarihli ciro girilemez")
    with db() as (conn, cur):
        # Aynı şube+tarih için ciro yazımlarını transaction bazında seri hale getir.
        lock_key = f"ciro:{c.sube_id}:{c.tarih}"
        cur.execute("SELECT pg_advisory_xact_lock(hashtext(%s))", (lock_key,))
        # Sert koruma: aynı şube+tarih için aktif ciro birden fazla olamaz.
        cur.execute(
            """
            SELECT id, (nakit+pos+online) AS toplam
            FROM ciro
            WHERE durum='aktif' AND tarih=%s AND sube_id=%s
            FOR UPDATE
            """,
            (str(c.tarih), c.sube_id),
        )
        mevcut = cur.fetchone()
        if mevcut:
            mevcut_tutar = float(mevcut.get("toplam") or 0)
            raise HTTPException(
                409,
                f"Bu şube için {c.tarih} tarihinde aktif ciro zaten var ({mevcut_tutar:,.0f} ₺).",
            )

        # Şube oranlarını çek
        cur.execute("SELECT COALESCE(pos_oran,0) as pos_oran, COALESCE(online_oran,0) as online_oran FROM subeler WHERE id=%s", (c.sube_id,))
        oran = cur.fetchone()
        pos_oran    = float(oran['pos_oran'])    if oran else 0.0
        online_oran = float(oran['online_oran']) if oran else 0.0

        pos_kesinti    = pos    * pos_oran    / 100.0
        online_kesinti = online * online_oran / 100.0
        net_tutar      = nakit + (pos - pos_kesinti) + (online - online_kesinti)

        # force sadece UX uyarılarını bypass eder; sert duplicate engeli yukarıda uygulanır.
        cid = str(uuid.uuid4())
        # Teknik duplicate koruması: son 5 saniye içinde birebir aynı istek geldi mi?
        if not c.force:
            cur.execute("""
                SELECT id FROM ciro WHERE durum='aktif'
                AND tarih=%s AND sube_id=%s
                AND nakit=%s AND pos=%s AND online=%s
                AND olusturma >= NOW() - INTERVAL '5 seconds'
            """, (c.tarih, c.sube_id, c.nakit, c.pos, c.online))
            if cur.fetchone():
                return {"id": None, "success": False, "duplicate": True,
                        "mesaj": "Aynı istek son 5 saniye içinde zaten gönderildi."}

        # Ciro tablosuna yaz
        cur.execute("""INSERT INTO ciro (id,tarih,sube_id,nakit,pos,online,aciklama)
            VALUES (%s,%s,%s,%s,%s,%s,%s)""",
            (cid, c.tarih, c.sube_id, c.nakit, c.pos, c.online, c.aciklama))

        # Kasaya NET tutar yaz (komisyon zaten düşülmüş)
        # POS/Online kesinti ayrıca yazılmıyor — net tutar içinde zaten yok
        # Panel komisyon tutarını ciro tablosundan hesaplıyor (bilgi amaçlı)
        insert_kasa_hareketi(cur, c.tarih, 'CIRO', net_tutar,
            f'Ciro girişi (net) — pos:%{pos_oran} online:%{online_oran}',
            'ciro', cid, ref_id=cid, ref_type='CIRO')

        audit(cur, 'ciro', cid, 'INSERT')
    return {"id": cid, "success": True, "net_tutar": net_tutar,
            "pos_kesinti": pos_kesinti, "online_kesinti": online_kesinti}


@app.put("/api/ciro/{cid}")
def ciro_guncelle(cid: str, c: CiroModel):
    """
    Ciro güncelleme — ledger immutable mantığı korunur:
    1. Eski kasa hareketi ters kayıtla iptal edilir
    2. Yeni tutarla yeni kasa hareketi yazılır
    3. Ciro tablosu güncellenir
    Audit trail eksiksiz kalır.
    """
    nakit  = float(c.nakit  or 0)
    pos    = float(c.pos    or 0)
    online = float(c.online or 0)
    # 🔴 P1 (2026-08-12): POST ile aynı negatif/sıfır koruması PUT'ta da olmalı.
    if nakit < 0 or pos < 0 or online < 0:
        raise HTTPException(400, "Ciro alanları negatif olamaz")
    if (nakit + pos + online) <= 0:
        raise HTTPException(400, "Ciro toplamı pozitif olmalı")

    with db() as (conn, cur):
        cur.execute("SELECT * FROM ciro WHERE id=%s AND durum='aktif'", (cid,))
        eski = cur.fetchone()
        if not eski:
            raise HTTPException(404, "Ciro kaydı bulunamadı veya iptal edilmiş")

        # Şube oranlarını çek — güncel oranla hesapla
        sube_id = c.sube_id or eski['sube_id']
        # 🔴 P1 (2026-08-12, Codex): PUT şube değiştirince hedef şube+tarihte zaten
        # aktif ciro varsa İKİ aktif ciro + çift ledger oluşuyordu (POST'taki sert
        # 409 koruması PUT'ta yoktu). Aynı korumayı buraya taşı: tarih SABİT kalır
        # (kasa eski['tarih']'e yazılıyor), çakışma anahtarı = (hedef sube_id, tarih).
        _tarih = eski['tarih']
        cur.execute("SELECT pg_advisory_xact_lock(hashtext(%s))", (f"ciro:{sube_id}:{_tarih}",))
        cur.execute(
            "SELECT id FROM ciro WHERE durum='aktif' AND tarih=%s AND sube_id=%s AND id<>%s FOR UPDATE",
            (str(_tarih), sube_id, cid),
        )
        if cur.fetchone():
            raise HTTPException(409, f"Hedef şube için {_tarih} tarihinde aktif ciro zaten var.")
        cur.execute("SELECT COALESCE(pos_oran,0) as pos_oran, COALESCE(online_oran,0) as online_oran FROM subeler WHERE id=%s", (sube_id,))
        oran = cur.fetchone()
        pos_oran    = float(oran['pos_oran'])    if oran else 0.0
        online_oran = float(oran['online_oran']) if oran else 0.0

        pos_kesinti    = pos    * pos_oran    / 100.0
        online_kesinti = online * online_oran / 100.0
        net_tutar      = nakit + (pos - pos_kesinti) + (online - online_kesinti)

        # 1. Eski kasa hareketini iptal et (ters kayıt)
        iptal_kasa_hareketi(cur, cid, 'ciro', 'CIRO', 'CIRO_DUZELTME',
                            f'Ciro düzeltme — eski tutar iptal')

        # 2. Ciro tablosunu güncelle
        cur.execute("""
            UPDATE ciro SET nakit=%s, pos=%s, online=%s, aciklama=%s, sube_id=%s
            WHERE id=%s
        """, (nakit, pos, online, c.aciklama, sube_id, cid))

        # 3. Yeni net tutarla kasa hareketi yaz
        insert_kasa_hareketi(cur, eski['tarih'], 'CIRO', net_tutar,
            f'Ciro düzeltme (net) — pos:%{pos_oran} online:%{online_oran}',
            'ciro', cid, ref_id=cid, ref_type='CIRO_GUNCELLEME')

        audit(cur, 'ciro', cid, 'GUNCELLEME', eski=eski)

    return {"success": True, "net_tutar": net_tutar,
            "pos_kesinti": pos_kesinti, "online_kesinti": online_kesinti}

@app.delete("/api/ciro/{cid}")
def ciro_sil(cid: str):
    with db() as (conn, cur):
        cur.execute("SELECT * FROM ciro WHERE id=%s AND durum='aktif'", (cid,))
        eski = cur.fetchone()
        if not eski: raise HTTPException(404, "Kayıt bulunamadı veya zaten iptal edilmiş")

        # Ciroyu iptal et
        cur.execute("UPDATE ciro SET durum='iptal' WHERE id=%s", (cid,))

        # Ledger: tüm silmelerle aynı model — tek merkez
        iptal_kasa_hareketi(cur, cid, 'ciro', 'CIRO', 'CIRO_IPTAL', 'Ciro iptali')

        audit(cur, 'ciro', cid, 'IPTAL', eski=eski)
        # FIX KP3 (2026-07-05): ciro iptali z_nakit'i (ciro.nakit) değiştirir → o günün kapanış
        # farkını tazele (masum personele eski/yanlış zimmet kalmasın). Tek tetikleme noktası;
        # uyarı yoksa no-op. Hata-yutar: recalc sorunu ciro iptalini bozmasın.
        try:
            from kasa_fark_recalc import sube_gun_kapanis_recalc
            sube_gun_kapanis_recalc(cur, eski['sube_id'], eski['tarih'])
        except Exception:
            pass
    return {"success": True}

# ── PERSONEL ───────────────────────────────────────────────────
class PersonelModel(BaseModel):
    ad_soyad: str
    gorev: Optional[str] = None
    calisma_turu: str = 'surekli'
    maas: float = 0
    saatlik_ucret: float = 0
    yemek_ucreti: float = 0
    yol_ucreti: float = 0
    odeme_gunu: int = 1
    baslangic_tarihi: Optional[str] = None  # string olarak alıp None/boş kontrolü yapılır
    sube_id: Optional[str] = None
    notlar: Optional[str] = None
    telefon: Optional[str] = None

    def baslangic_date(self):
        if not self.baslangic_tarihi or self.baslangic_tarihi.strip() == '':
            return None
        try:
            from datetime import date as _date
            return _date.fromisoformat(self.baslangic_tarihi)
        except ValueError:
            return None

def _personel_api_row(r: dict) -> dict:
    d = dict(r)
    d["panel_pin_tanimli"] = bool((d.get("panel_pin_hash") or "").strip())
    d.pop("panel_pin_salt", None)
    d.pop("panel_pin_hash", None)
    if "panel_yonetici" in d and d["panel_yonetici"] is not None:
        d["panel_yonetici"] = bool(d["panel_yonetici"])
    # 🌅 Erken açılış izni (2026-08-29): ekran bu alanı okumazsa herkesi
    # "07:00" gösterir ve verilen izin görünmez olurdu.
    d["erken_acilis_izni"] = bool(d.get("erken_acilis_izni"))
    return d


@app.get("/api/personel")
def personel_listele(aktif: Optional[bool] = None):
    with db() as (conn, cur):
        if aktif is not None:
            cur.execute(
                """
                SELECT p.*, s.ad as sube_adi,
                       opv.odeme_durumu, opv.odeme_tarihi, opv.odenecek_tutar, opv.odenen_tutar
                FROM personel p
                LEFT JOIN subeler s ON s.id = p.sube_id
                LEFT JOIN LATERAL (
                    SELECT
                        op.durum AS odeme_durumu,
                        op.tarih AS odeme_tarihi,
                        op.odenecek_tutar,
                        op.odenen_tutar
                    FROM odeme_plani op
                    WHERE op.kaynak_tablo='personel'
                      AND op.kaynak_id = p.id
                      AND op.durum != 'iptal'
                      AND DATE_TRUNC('month', op.tarih) = DATE_TRUNC('month', CURRENT_DATE)
                    ORDER BY
                        CASE WHEN op.durum='odendi' THEN 0 WHEN op.durum='onay_bekliyor' THEN 1 ELSE 2 END,
                        op.olusturma DESC
                    LIMIT 1
                ) opv ON TRUE
                WHERE p.aktif=%s
                ORDER BY p.ad_soyad
                """,
                (aktif,),
            )
        else:
            cur.execute(
                """
                SELECT p.*, s.ad as sube_adi,
                       opv.odeme_durumu, opv.odeme_tarihi, opv.odenecek_tutar, opv.odenen_tutar
                FROM personel p
                LEFT JOIN subeler s ON s.id = p.sube_id
                LEFT JOIN LATERAL (
                    SELECT
                        op.durum AS odeme_durumu,
                        op.tarih AS odeme_tarihi,
                        op.odenecek_tutar,
                        op.odenen_tutar
                    FROM odeme_plani op
                    WHERE op.kaynak_tablo='personel'
                      AND op.kaynak_id = p.id
                      AND op.durum != 'iptal'
                      AND DATE_TRUNC('month', op.tarih) = DATE_TRUNC('month', CURRENT_DATE)
                    ORDER BY
                        CASE WHEN op.durum='odendi' THEN 0 WHEN op.durum='onay_bekliyor' THEN 1 ELSE 2 END,
                        op.olusturma DESC
                    LIMIT 1
                ) opv ON TRUE
                ORDER BY p.ad_soyad
                """
            )
        return [_personel_api_row(dict(r)) for r in cur.fetchall()]

@app.post("/api/personel")
def personel_ekle(p: PersonelModel):
    with db() as (conn, cur):
        pid = str(uuid.uuid4())
        cur.execute("""INSERT INTO personel
            (id,ad_soyad,gorev,calisma_turu,maas,saatlik_ucret,yemek_ucreti,yol_ucreti,odeme_gunu,baslangic_tarihi,sube_id,notlar,telefon)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (pid, p.ad_soyad, p.gorev, p.calisma_turu, p.maas, p.saatlik_ucret,
             p.yemek_ucreti, p.yol_ucreti, p.odeme_gunu, p.baslangic_date(), p.sube_id, p.notlar,
             (p.telefon or '').strip() or None))
        audit(cur, 'personel', pid, 'INSERT')
    return {"id": pid, "success": True}

@app.put("/api/personel/{pid}")
def personel_guncelle(pid: str, p: PersonelModel):
    with db() as (conn, cur):
        cur.execute("SELECT * FROM personel WHERE id=%s", (pid,))
        eski = cur.fetchone()
        if not eski: raise HTTPException(404)
        cur.execute("""UPDATE personel SET ad_soyad=%s,gorev=%s,calisma_turu=%s,maas=%s,
            saatlik_ucret=%s,yemek_ucreti=%s,yol_ucreti=%s,odeme_gunu=%s,
            baslangic_tarihi=%s,sube_id=%s,notlar=%s,telefon=%s WHERE id=%s""",
            (p.ad_soyad, p.gorev, p.calisma_turu, p.maas, p.saatlik_ucret,
             p.yemek_ucreti, p.yol_ucreti, p.odeme_gunu, p.baslangic_date(),
             p.sube_id, p.notlar, (p.telefon or '').strip() or None, pid))
        audit(cur, 'personel', pid, 'UPDATE', eski=eski)
    return {"success": True}

@app.post("/api/personel/{pid}/cikis")
def personel_cikis(pid: str, neden: str = ""):
    with db() as (conn, cur):
        cur.execute("UPDATE personel SET aktif=FALSE, cikis_tarihi=%s WHERE id=%s",
            (str(bugun_tr()), pid))
        # KURAL (2026-07-03): "Dönem hakedişi aktiflikten bağımsızdır."
        # Sadece ÇALIŞILMAMIŞ (gelecek dönem) planlar iptal edilir — simülasyondan çıksın.
        # Çalışılan dönemin (çıkış ayı dahil, kısmi hakediş) planı KALIR: ay sonunda
        # kasadan doğru oranda ödenir, ödendikten sonra listeden doğal düşer.
        # DİKKAT — referans_ay konvansiyonu = ÖDEME AYI (çalışma ayı + 1, a168871 sync ekseni):
        # bugün çıkan kişinin çalıştığı son dönemin planı referans_ay = gelecek ayın 1'ini taşır,
        # o yüzden koruma sınırı "+1 ay"dır. (Önceki filtre bunu bilmediği için 10 günlük
        # hakedişi de siliyordu — kullanıcı şikayetiyle düzeltildi 2026-07-04.)
        cur.execute("""
            UPDATE odeme_plani SET durum='iptal'
            WHERE kaynak_tablo='personel' AND kaynak_id=%s
            AND durum IN ('bekliyor','onay_bekliyor')
            AND referans_ay > DATE_TRUNC('month', %s::date + INTERVAL '1 month')
        """, (pid, str(bugun_tr())))
        cur.execute("""
            UPDATE onay_kuyrugu SET durum='reddedildi'
            WHERE kaynak_tablo='personel' AND kaynak_id=%s
            AND durum='bekliyor'
        """, (pid,))
        audit(cur, 'personel', pid, 'CIKIS')
    return {"success": True}

@app.delete("/api/personel/{pid}")
def personel_sil(pid: str):
    with db() as (conn, cur):
        cur.execute("SELECT * FROM personel WHERE id=%s", (pid,))
        eski = cur.fetchone()
        if not eski: raise HTTPException(404)
        cur.execute("DELETE FROM personel WHERE id=%s", (pid,))
        audit(cur, 'personel', pid, 'DELETE', eski=eski)
    return {"success": True}



class PersonelAylikModel(BaseModel):
    calisma_saati: float = 0
    fazla_mesai_saat: float = 0
    bayram_mesai_saat: float = 0
    eksik_gun: float = 0
    raporlu_gun: float = 0
    rapor_kesinti: bool = False
    manuel_duzeltme: float = 0
    not_aciklama: Optional[str] = None

def _maas_kayit_kilit_guard(cur, pid: str, yil: int, ay: int) -> None:
    """FAZ 0 #5: Onaylı (kilitli) veya ödenmiş maaş kaydı sessizce taslağa
    döndürülemez. Düzeltme için kullanıcı önce '🔓 Kilidi Aç' demeli."""
    cur.execute(
        "SELECT durum FROM personel_aylik WHERE personel_id=%s AND yil=%s AND ay=%s",
        (pid, yil, ay),
    )
    r = cur.fetchone()
    if r and (r.get("durum") == "onaylandi"):
        raise HTTPException(
            400,
            "Maaş kaydı onaylı (kilitli). Değiştirmek için önce '🔓 Kilidi Aç' yapın.",
        )
    # ARREARS: (yil, ay) = ÇALIŞMA dönemi → ödenmiş plan referans_ay = ödeme ayı (dönem+1).
    # FIX 2026-07-04: MAKE_DATE(yil,ay,1) çalışma ayına bakıyordu → kilit hiç devreye
    # girmiyordu. Eski konvansiyon planları için çalışma ayı da kontrol edilir.
    cur.execute(
        """
        SELECT 1 FROM odeme_plani
        WHERE kaynak_tablo='personel' AND kaynak_id=%s AND durum='odendi'
          AND referans_ay IN (%s::date, MAKE_DATE(%s, %s, 1))
        LIMIT 1
        """,
        (pid, str(_maas_svc.maas_odeme_tarihi(yil, ay)), yil, ay),
    )
    if cur.fetchone():
        raise HTTPException(
            400,
            "Bu dönemin maaşı ödenmiş — kayıt değiştirilemez. Düzeltme için ek ödeme / "
            "gelecek aydan mahsup gerekir (geçmiş kayıt değişmez).",
        )


# ── MAAŞ ÇEKİRDEĞİ: maas_service.py (TEK MERKEZ) ───────────────
# Hesap + plan yazımı + dönem (arrears) kuralı + sabitler artık maas_service'te.
# Eski isimler geriye-uyum için aynen bağlanır (bu dosyadaki çağıranlar değişmez).
import maas_service as _maas_svc
from maas_service import (
    TR_AYLAR as _TR_AYLAR,
    maas_hesapla,
    personel_donem_orani as _personel_donem_orani,
    vardiya_kayit_dict as _personel_vardiya_kayit_dict,
    maas_odeme_tarihi as _personel_maas_odeme_tarihi,
    vardiya_takip_hesap as _vardiya_takip_hesap,
    kanonik_net as _kanonik_net,
    aylik_vardiya_senkronize as _personel_aylik_vardiya_senkronize,
    bordro_anomali_tara as _bordro_anomali_tara,
)


def _personel_odeme_plani_senkronize(cur, p: dict, yil: int, ay: int, net: float):
    """Geriye-uyum sarmalayıcı: tek plan yazıcı = maas_service.odeme_plani_esitle.
    (Eski dönüş sözleşmesi korunur: plan id veya None.)"""
    r = _maas_svc.odeme_plani_esitle(cur, p, yil, ay, net, guncelle=True)
    return (r or {}).get("id")


@app.get("/api/personel-aylik")
def personel_aylik_listele(yil: int = None, ay: int = None):
    """Bu ay için tüm personelin aylik kayıtlarını döner. Kayıt yoksa tahmini tutar ile döner."""
    bugun = bugun_tr()
    yil = yil or bugun.year
    ay  = ay  or bugun.month
    maas_odeme_tarihi = _personel_maas_odeme_tarihi(yil, ay)
    with db() as (conn, cur):
        # KURAL: "Dönem hakedişi aktiflikten bağımsızdır" — ayrılan personel, çıkış tarihi bu
        # dönemle kesişiyorsa maaş ekranında KALIR (son dönem hakedişi görünür/ödenir).
        cur.execute("""SELECT * FROM personel
                       WHERE aktif=TRUE
                          OR (cikis_tarihi IS NOT NULL AND cikis_tarihi >= MAKE_DATE(%s,%s,1))
                       ORDER BY ad_soyad""", (yil, ay))
        personeller = cur.fetchall()
        import calendar as _cal
        _son_gun = _cal.monthrange(yil, ay)[1]

        # PERF N+1 FIX (2026-07-06): eskiden KİŞİ BAŞI 2 SQL + tam vardiya-takip hesabı
        # (her biri KENDİ DB bağlantısını açıp tüm sorgu setini çalıştırıyordu) → ~20 personelde
        # ~13sn. Üç veri de TOPLU çekilir; döngü içinde yalnız map lookup kalır. Davranış birebir:
        # 1) Vardiya Takip TÜM personel TEK çağrıda (aynı kanonik fonksiyon, personel filtresi yok).
        _vt_map = {}
        try:
            from gorev_api import vardiya_takip as _vt_all_fn
            _vt_res = _vt_all_fn(yil, ay)
            _vt_map = {str(r.get("personel_id")): r for r in ((_vt_res or {}).get("personeller") or [])}
        except Exception as _e_vt:
            logging.getLogger(__name__).warning("personel-aylik toplu vardiya takip hatasi: %s", _e_vt)
        # 2) personel_aylik kayıtları tek sorgu → map
        cur.execute("SELECT * FROM personel_aylik WHERE yil=%s AND ay=%s", (yil, ay))
        _kayit_map = {str(r["personel_id"]): dict(r) for r in (cur.fetchall() or [])}
        # 3) ödeme planları tek sorgu — kişi başı LIMIT 1 ile AYNI öncelik sırası (DISTINCT ON)
        cur.execute(
            """
            SELECT DISTINCT ON (op.kaynak_id)
                op.kaynak_id, op.id::text AS odeme_id, op.durum AS odeme_durumu,
                op.tarih AS odeme_tarihi, op.odenecek_tutar, op.odenen_tutar
            FROM odeme_plani op
            WHERE op.kaynak_tablo='personel' AND op.durum != 'iptal'
              AND op.referans_ay = DATE_TRUNC('month', %s::date)
            ORDER BY op.kaynak_id,
                CASE WHEN op.durum='odendi' THEN 0 WHEN op.durum='onay_bekliyor' THEN 1 ELSE 2 END,
                op.olusturma DESC
            """,
            (str(maas_odeme_tarihi),),
        )
        _plan_map = {str(r["kaynak_id"]): dict(r) for r in (cur.fetchall() or [])}

        sonuc = []
        for p in personeller:
            # dönem içi çalışma aralığı (başlangıç/çıkış kırpması)
            _d1, _d2 = date(yil, ay, 1), date(yil, ay, _son_gun)
            _eff1 = max(_d1, p['baslangic_tarihi']) if p.get('baslangic_tarihi') else _d1
            _eff2 = min(_d2, p['cikis_tarihi']) if p.get('cikis_tarihi') else _d2
            if _eff1 > _eff2:
                continue  # bu dönemde hiç çalışmamış
            # PERF N+1 FIX: toplu map lookup (eski kişi-başı sorgularla birebir aynı sonuç)
            kayit = _kayit_map.get(str(p['id']))
            plan = _plan_map.get(str(p['id'])) or {}
            if kayit:
                net = float(kayit['hesaplanan_net'] or 0)
                durum = kayit['durum']
            else:
                # Tahmini hesap — ay içinde başlayan/ayrılan için gün oranlı
                if p['calisma_turu'] == 'surekli':
                    net = float(p['maas'] or 0) + float(p['yemek_ucreti'] or 0) + float(p['yol_ucreti'] or 0)
                    _gun = (_eff2 - _eff1).days + 1
                    if _gun < _son_gun:
                        net = round(net * _gun / _son_gun, 2)
                else:
                    net = 0  # Part-time saat girilmeden tahmin yapılamaz
                durum = 'tahmini'
                kayit = {}

            # KANONİK: vardiya alanları ve kayıtsız tahmin = Vardiya Takip kurgusu (tek merkez)
            # PERF N+1 FIX: kişi-başı ayrı hesap yerine toplu map (aynı kanonik fonksiyon)
            vt = _vt_map.get(str(p['id']))
            _saat_kaynagi = 'vardiya_atama'
            if vt is not None:
                vk = {'toplam_ay_saat': float(vt.get('toplam_planlanan_saat') or 0),
                      'ek_mesai_haftalik_toplam': float(vt.get('toplam_fazla_mesai_saat') or 0),
                      'haftalik_limit': 0}
                # SABİT MESAİ FALLBACK (sahip doktrini 2026-08-07): vardiya ataması
                # TEYİT katmanıdır — yoksa sabit tanımlı mesaiden aktarılır. Part-time
                # hakedişi saat×ücret olduğu için atama yokken 0 ₺ çıkıyordu.
                # aylik_vardiya_senkronize ile AYNI kanonik fonksiyon kullanılır.
                if p['calisma_turu'] != 'surekli' and vk['toplam_ay_saat'] <= 0:
                    try:
                        _s, _k = _maas_svc.sabit_mesai_saati(cur, dict(p), yil, ay)
                        if _s > 0:
                            # Saatlik ücret de tanımsızsa sahip kararı: 99 ₺/saat
                            # (tek kaynak: maas_service.VARSAYILAN_SAATLIK_UCRET).
                            _su = float(p['saatlik_ucret'] or 0)
                            if _su <= 0:
                                _su = _maas_svc.VARSAYILAN_SAATLIK_UCRET
                                _k = f"{_k}+varsayilan_ucret"
                            vk['toplam_ay_saat'] = _s
                            _saat_kaynagi = _k
                            vt = dict(vt)
                            vt['toplam_planlanan_saat'] = _s
                            vt['net_hakediş'] = round(_s * _su, 2) + float(p['yol_ucreti'] or 0)
                    except Exception as _e_sm:
                        logging.getLogger(__name__).warning("sabit mesai fallback: %s", _e_sm)
            else:
                vk = _vv2.personel_ay_vardiya_maas_kaynagi(cur, p['id'], yil, ay)
            if not kayit:
                if vt is not None:
                    net = float(vt.get('net_hakediş') or 0)
                    kayit = {'calisma_saati': vk['toplam_ay_saat'],
                             'fazla_mesai_saat': vk['ek_mesai_haftalik_toplam']}
                else:
                    kayit = _personel_vardiya_kayit_dict(cur, p, yil, ay)
                    kayit.pop("_vardiya", None)
                    net = maas_hesapla(dict(p), kayit, yil, ay)
                durum = 'vardiya_tahmini'

            sonuc.append({
                'personel_id': p['id'],
                'ad_soyad': p['ad_soyad'],
                'gorev': p['gorev'],
                'calisma_turu': p['calisma_turu'],
                'maas': float(p['maas'] or 0),
                'saatlik_ucret': float(p['saatlik_ucret'] or 0),
                'yemek_ucreti': float(p['yemek_ucreti'] or 0),
                'yol_ucreti': float(p['yol_ucreti'] or 0),
                'sube_id': p['sube_id'],
                'kayit_id': kayit.get('id'),
                # Saatin NEREDEN geldiği: 'vardiya_atama' (gerçek plan) ·
                # 'sabit_tanim_haftalik' (personelin haftalık tanımı) ·
                # 'varsayilan_gunluk' (tanım da yok → 9,5 sa/gün varsayıldı, UYARI).
                'saat_kaynagi': _saat_kaynagi,
                'calisma_saati': float(kayit.get('calisma_saati') or 0),
                'fazla_mesai_saat': float(kayit.get('fazla_mesai_saat') or 0),
                'bayram_mesai_saat': float(kayit.get('bayram_mesai_saat') or 0),
                'eksik_gun': float(kayit.get('eksik_gun') or 0),
                'raporlu_gun': float(kayit.get('raporlu_gun') or 0),
                'rapor_kesinti': kayit.get('rapor_kesinti', False),
                'manuel_duzeltme': float(kayit.get('manuel_duzeltme') or 0),
                'not_aciklama': kayit.get('not_aciklama'),
                'hesaplanan_net': net,
                'avans_mahsup': float(kayit.get('avans_mahsup') or 0),
                'mahsup_devir': float(kayit.get('mahsup_devir') or 0),
                'durum': durum,
                'odeme_id': plan.get('odeme_id'),
                'odeme_durumu': plan.get('odeme_durumu'),
                'odeme_tarihi': plan.get('odeme_tarihi'),
                'odenecek_tutar': float(plan.get('odenecek_tutar') or 0),
                'odenen_tutar': float(plan.get('odenen_tutar') or 0),
                'vardiya_ay_toplam_saat': vk.get('toplam_ay_saat', 0),
                'vardiya_ek_mesai_saat': vk.get('ek_mesai_haftalik_toplam', 0),
                'vardiya_haftalik_limit': vk.get('haftalik_limit', 0),
                'aktif': bool(p.get('aktif')),
                'cikis_tarihi': str(p['cikis_tarihi']) if p.get('cikis_tarihi') else None,
            })
        return {'yil': yil, 'ay': ay, 'personeller': sonuc,
                'toplam_tahmini': sum(r['hesaplanan_net'] for r in sonuc)}


@app.post("/api/personel-aylik/vardiya-sync")
def personel_aylik_vardiya_sync(yil: int = None, ay: int = None):
    """Secilen ayin vardiya verisini aylik maas kayitlarina ve odeme planina senkronlar."""
    bugun = bugun_tr()
    yil = yil or bugun.year
    ay = ay or bugun.month
    with db() as (conn, cur):
        # KURAL: dönem hakedişi aktiflikten bağımsız — ayrılanın son dönemi de vardiyadan
        # güncel veri almaya devam eder (çıkış tarihi dönemle kesişiyorsa)
        cur.execute("""SELECT * FROM personel
                       WHERE aktif=TRUE
                          OR (cikis_tarihi IS NOT NULL AND cikis_tarihi >= MAKE_DATE(%s,%s,1))
                       ORDER BY ad_soyad""", (yil, ay))
        personeller = cur.fetchall()
        sonuc = []
        for p in personeller:
            r = _personel_aylik_vardiya_senkronize(cur, dict(p), yil, ay)
            sonuc.append({"personel_id": p["id"], "ad_soyad": p["ad_soyad"], **r})
        audit(cur, 'personel_aylik', f"{yil}-{ay:02d}", 'VARDIYA_SYNC', yeni={'adet': len(sonuc)})
    return {
        "success": True,
        "yil": yil,
        "ay": ay,
        "adet": len(sonuc),
        "toplam_net": sum(float(r.get("hesaplanan_net") or 0) for r in sonuc),
        "personeller": sonuc,
    }


@app.get("/api/personel-aylik/onay-kuyrugu")
def personel_aylik_onay_kuyrugu(yil: int = None, ay: int = None):
    """BORDRO ONAY KUYRUĞU — "istisnaya göre onay" (approval by exception).

    Dönemin taslak kayıtlarını tarayıp ikiye ayırır:
      • TEMİZ   → hiçbir anomali yok, toplu onaya uygun
      • İNCELE  → gözle teyit isteyen kayıt (sıfır net, sapma, elle düzeltme, kısmi dönem...)

    Salt-okur. Hiçbir kaydı onaylamaz — sahip kararını hazırlar, yerine geçmez.
    Kural motoru: maas_service.bordro_anomali_tara (tek çekirdek).
    """
    bugun = bugun_tr()
    yil = yil or bugun.year
    ay  = ay  or bugun.month
    with db() as (conn, cur):
        sonuc = _bordro_anomali_tara(cur, yil, ay)
    # Gecikme ölçüsü: ödeme günü geçtiyse bekleyen bordro artık sessiz değil.
    _od = date.fromisoformat(sonuc["odeme_tarihi"])
    _gecikme = (bugun - _od).days
    sonuc["gecikme_gun"] = _gecikme if _gecikme > 0 else 0
    sonuc["odeme_gunu_gecti"] = _gecikme > 0 and sonuc["ozet"]["bekleyen_adet"] > 0
    return sonuc


class TopluOnayModel(BaseModel):
    yil: int = None
    ay: int = None
    kapsam: str = "temiz"          # temiz | secili
    personel_idler: list = None    # kapsam='secili' ise zorunlu
    kuru: bool = True             # varsayılan PROVA — yazmadan önce ne olacağını gösterir


@app.post("/api/personel-aylik/toplu-onayla")
def personel_aylik_toplu_onayla(body: TopluOnayModel):
    """Dönemin bordro kayıtlarını TOPLU onaylar. Ödeme YAPMAZ, kasa hareketi üretmez.

    kapsam='temiz'  → yalnız anomalisiz kayıtlar (varsayılan; şüpheli olan elde kalır)
    kapsam='secili' → personel_idler listesindekiler (anomalili olsa da sahip bilerek onaylar)

    kuru=True (varsayılan) hiçbir şey yazmaz; ne olacağını döner. Yazmak için kuru=False.
    Onay = hesabın KİLİDİ. Ödeme ayrı adımdır ve kendi guard'ından geçer.
    """
    bugun = bugun_tr()
    yil = body.yil or bugun.year
    ay  = body.ay  or bugun.month
    kapsam = (body.kapsam or "temiz").lower()
    if kapsam not in ("temiz", "secili"):
        raise HTTPException(400, "kapsam 'temiz' veya 'secili' olmalı")
    with db() as (conn, cur):
        tarama = _bordro_anomali_tara(cur, yil, ay)
        if kapsam == "temiz":
            hedefler = tarama["temiz"]
            neden = "anomali taramasından temiz geçti"
        else:
            istenen = set(str(x) for x in (body.personel_idler or []))
            if not istenen:
                raise HTTPException(400, "kapsam='secili' için personel_idler gerekli")
            hedefler = [k for k in (tarama["temiz"] + tarama["incele"])
                        if str(k["personel_id"]) in istenen]
            neden = "sahip elle seçti"
        # Kimlik üzerinden karşılaştır (dict eşitliği kırılgan olur)
        _hedef_id = {str(k["personel_id"]) for k in hedefler}
        _kalan = [k for k in tarama["incele"] if str(k["personel_id"]) not in _hedef_id]
        if body.kuru:
            return {
                "kuru": True, "yil": yil, "ay": ay, "donem": tarama["donem"],
                "onaylanacak_adet": len(hedefler),
                "onaylanacak_tutar": round(sum(k["hesaplanan_net"] for k in hedefler), 2),
                "neden": neden,
                "kayitlar": [{"personel_id": k["personel_id"], "ad_soyad": k["ad_soyad"],
                              "net": k["hesaplanan_net"],
                              "anomali_adet": len(k["anomaliler"])} for k in hedefler],
                "dokunulmayan": {
                    "incele_adet": len(_kalan),
                    "incele_tutar": round(sum(k["hesaplanan_net"] for k in _kalan), 2),
                },
                "mesaj": "PROVA — hiçbir kayıt değişmedi. Uygulamak için kuru=false gönderin.",
            }
        onaylanan, atlanan = [], []
        for k in hedefler:
            cur.execute(
                """UPDATE personel_aylik SET durum='onaylandi'
                    WHERE personel_id=%s AND yil=%s AND ay=%s AND durum='taslak'""",
                (k["personel_id"], yil, ay),
            )
            (onaylanan if cur.rowcount else atlanan).append(k)
        audit(cur, 'personel_aylik', f"{yil}-{ay:02d}", 'TOPLU_ONAY',
              yeni={'kapsam': kapsam, 'adet': len(onaylanan),
                    'tutar': round(sum(k["hesaplanan_net"] for k in onaylanan), 2)})
    return {
        "success": True, "kuru": False, "yil": yil, "ay": ay, "donem": tarama["donem"],
        "onaylanan_adet": len(onaylanan),
        "onaylanan_tutar": round(sum(k["hesaplanan_net"] for k in onaylanan), 2),
        "atlanan_adet": len(atlanan),
        "kasa_etkisi": False,
        "mesaj": (f"{len(onaylanan)} bordro kaydı onaylandı ({neden}). "
                  "Ödeme, Ödeme Merkezi'nden ayrıca yapılır — bu adım para hareketi üretmedi."),
    }


@app.post("/api/personel-aylik/{pid}")
def personel_aylik_kaydet(pid: str, body: PersonelAylikModel, yil: int = None, ay: int = None):
    """Personel aylık kaydını girer/günceller ve maaşı hesaplar."""
    bugun = bugun_tr()
    yil = yil or bugun.year
    ay  = ay  or bugun.month
    maas_odeme_tarihi = _personel_maas_odeme_tarihi(yil, ay)
    with db() as (conn, cur):
        # aktif şartı YOK: ayrılan personelin son dönem hakedişi girilebilmeli (dönem hakedişi kuralı)
        cur.execute("SELECT * FROM personel WHERE id=%s", (pid,))
        p = cur.fetchone()
        if not p: raise HTTPException(404, "Personel bulunamadı")
        _maas_kayit_kilit_guard(cur, pid, yil, ay)

        kayit_dict = body.dict()
        # KANONİK: saat/fazla mesai VARDİYA TAKİP'ten gelir (elle girilse de kaynak takiptir —
        # kullanıcı kararı 2026-07-04); bayram/eksik/rapor/manuel/not katmanı body'den uygulanır.
        vt = _vardiya_takip_hesap(p["id"], yil, ay)
        if vt is not None:
            kayit_dict["calisma_saati"] = float(vt.get("toplam_planlanan_saat") or 0)
            kayit_dict["fazla_mesai_saat"] = float(vt.get("toplam_fazla_mesai_saat") or 0)
            net = _kanonik_net(dict(p), vt, kayit_dict)
        else:
            net = maas_hesapla(dict(p), kayit_dict, yil, ay)
        net, _avm, _avd = _maas_svc.avans_mahsup_uygula(cur, dict(p), yil, ay, net)

        kid = str(uuid.uuid4())
        cur.execute("""
            INSERT INTO personel_aylik
                (id, personel_id, yil, ay, calisma_saati, fazla_mesai_saat, bayram_mesai_saat,
                 eksik_gun, raporlu_gun, rapor_kesinti, manuel_duzeltme,
                 not_aciklama, hesaplanan_net, avans_mahsup, mahsup_devir, durum)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'taslak')
            ON CONFLICT (personel_id, yil, ay) DO UPDATE SET
                calisma_saati=%s, fazla_mesai_saat=%s, bayram_mesai_saat=%s,
                eksik_gun=%s, raporlu_gun=%s, rapor_kesinti=%s, manuel_duzeltme=%s,
                not_aciklama=%s, hesaplanan_net=%s, avans_mahsup=%s, mahsup_devir=%s, durum='taslak'
        """, (kid, pid, yil, ay,
                kayit_dict["calisma_saati"], kayit_dict["fazla_mesai_saat"], body.bayram_mesai_saat,
                body.eksik_gun, body.raporlu_gun, body.rapor_kesinti,
                body.manuel_duzeltme, body.not_aciklama, net, _avm, _avd,
                kayit_dict["calisma_saati"], kayit_dict["fazla_mesai_saat"], body.bayram_mesai_saat,
                body.eksik_gun, body.raporlu_gun, body.rapor_kesinti,
                body.manuel_duzeltme, body.not_aciklama, net, _avm, _avd))

        # Bağlı ödeme planını gerçek tutarla güncelle
        _personel_odeme_plani_senkronize(cur, dict(p), yil, ay, net)

        audit(cur, 'personel_aylik', kid, 'KAYDET', yeni={'net': net, 'yil': yil, 'ay': ay})
    return {"success": True, "hesaplanan_net": net}


@app.post("/api/personel-aylik/{pid}/vardiya-aktar")
def personel_aylik_vardiya_aktar(pid: str, yil: int = None, ay: int = None):
    """
    Vardiya atamalarından (planlı/onaylı) aylık maaş satırına aktarır ve neti yeniden hesaplar.

    - **Sürekli (TAM):** ``fazla_mesai_saat`` ← haftalık limit üstü toplam ek mesai (57h vb.).
    - **Part-time:** ``calisma_saati`` ← ay içi toplam atanmış saat.
    Bayram, eksik gün, manuel düzeltme vb. mevcut kayıttan korunur.
    """
    bugun = bugun_tr()
    yil = yil or bugun.year
    ay = ay or bugun.month
    with db() as (conn, cur):
        # aktif şartı YOK: ayrılanın son ayı vardiyadan pro-rata aktarılabilmeli
        # (personel_calisma_araligi çıkış tarihinden sonrasını zaten kırpar)
        cur.execute("SELECT * FROM personel WHERE id=%s", (pid,))
        p = cur.fetchone()
        if not p:
            raise HTTPException(404, "Personel bulunamadı")
        _maas_kayit_kilit_guard(cur, pid, yil, ay)

        # KANONİK (kullanıcı kararı 2026-07-04): aktarım = tek kişilik vardiya-sync ile AYNI yol.
        # Saat/fazla mesai Vardiya Takip kurgusundan (günlük 9.5 üstü), manuel alanlar korunur.
        cur.execute(
            "SELECT * FROM personel_aylik WHERE personel_id=%s AND yil=%s AND ay=%s",
            (pid, yil, ay),
        )
        row = cur.fetchone()
        bayram = float((row or {}).get("bayram_mesai_saat") or 0)
        eksik = float((row or {}).get("eksik_gun") or 0)
        raporlu = float((row or {}).get("raporlu_gun") or 0)
        rapor_k = bool((row or {}).get("rapor_kesinti") or False)
        manuel = float((row or {}).get("manuel_duzeltme") or 0)
        not_a = (row or {}).get("not_aciklama")

        vt = _vardiya_takip_hesap(pid, yil, ay)
        if vt is not None:
            calisma = float(vt.get("toplam_planlanan_saat") or 0)
            fazla = float(vt.get("toplam_fazla_mesai_saat") or 0)
            # FIX 2026-07-04: kanonik yolda vk tanımsız kalıyordu → uç 500 veriyordu
            vk = {"kaynak": "vardiya_takip",
                  "toplam_planlanan_saat": calisma,
                  "fazla_mesai_saat": fazla}
        else:
            # savunma: takip hesabı alınamazsa eski kaynak
            vk = _vv2.personel_ay_vardiya_maas_kaynagi(cur, pid, yil, ay)
            ct = (p.get("calisma_turu") or "surekli")
            calisma = float((row or {}).get("calisma_saati") or 0)
            fazla = float(vk["ek_mesai_haftalik_toplam"]) if ct == "surekli" else 0.0
            if ct != "surekli":
                calisma = float(vk["toplam_ay_saat"])
                bayram = 0.0

        kayit_dict = {
            "calisma_saati": calisma,
            "fazla_mesai_saat": fazla,
            "bayram_mesai_saat": bayram,
            "eksik_gun": eksik,
            "raporlu_gun": raporlu,
            "rapor_kesinti": rapor_k,
            "manuel_duzeltme": manuel,
            "not_aciklama": not_a,
        }
        net = _kanonik_net(dict(p), vt, kayit_dict) if vt is not None else maas_hesapla(dict(p), kayit_dict, yil, ay)
        net, _avm, _avd = _maas_svc.avans_mahsup_uygula(cur, dict(p), yil, ay, net)

        kid = str(uuid.uuid4())
        cur.execute("""
            INSERT INTO personel_aylik
                (id, personel_id, yil, ay, calisma_saati, fazla_mesai_saat, bayram_mesai_saat,
                 eksik_gun, raporlu_gun, rapor_kesinti, manuel_duzeltme,
                 not_aciklama, hesaplanan_net, avans_mahsup, mahsup_devir, durum)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'taslak')
            ON CONFLICT (personel_id, yil, ay) DO UPDATE SET
                calisma_saati=%s, fazla_mesai_saat=%s, bayram_mesai_saat=%s,
                eksik_gun=%s, raporlu_gun=%s, rapor_kesinti=%s, manuel_duzeltme=%s,
                not_aciklama=%s, hesaplanan_net=%s, avans_mahsup=%s, mahsup_devir=%s, durum='taslak'
        """, (kid, pid, yil, ay,
                calisma, fazla, bayram,
                eksik, raporlu, rapor_k,
                manuel, not_a, net, _avm, _avd,
                calisma, fazla, bayram,
                eksik, raporlu, rapor_k,
                manuel, not_a, net, _avm, _avd))

        _personel_odeme_plani_senkronize(cur, dict(p), yil, ay, net)

        audit(cur, 'personel_aylik', kid, 'VARDIYA_AKTAR', yeni={'net': net, 'yil': yil, 'ay': ay})
    return {"success": True, "hesaplanan_net": net, "vardiya": vk}


@app.post("/api/personel-aylik/{pid}/onayla")
def personel_aylik_onayla(pid: str, yil: int = None, ay: int = None):
    """Maaş hesabını kilitler; ödeme yapmaz, kasa hareketi oluşturmaz."""
    bugun = bugun_tr()
    yil = yil or bugun.year
    ay  = ay  or bugun.month
    maas_odeme_tarihi = _personel_maas_odeme_tarihi(yil, ay)
    with db() as (conn, cur):
        cur.execute("""
            UPDATE personel_aylik SET durum='onaylandi'
            WHERE personel_id=%s AND yil=%s AND ay=%s AND durum='taslak'
        """, (pid, yil, ay))
        if cur.rowcount == 0:
            raise HTTPException(400, "Kayıt bulunamadı veya zaten onaylandı")
        cur.execute(
            """
            SELECT durum
            FROM odeme_plani
            WHERE kaynak_tablo='personel'
              AND kaynak_id=%s
              AND durum != 'iptal'
              AND referans_ay = DATE_TRUNC('month', %s::date)
            ORDER BY
              CASE WHEN durum='odendi' THEN 0 WHEN durum='onay_bekliyor' THEN 1 ELSE 2 END,
              olusturma DESC
            LIMIT 1
            """,
            (pid, str(maas_odeme_tarihi)),
        )
        plan = cur.fetchone()
    return {
        "success": True,
        "mesaj": "Maaş hesabı onaylandı. Ödeme paneldeki ödeme planından yapılır.",
        "kasa_etkisi": False,
        "odeme_durumu": (plan or {}).get("odeme_durumu"),
    }

@app.post("/api/personel-aylik/{pid}/kilit-ac")
def personel_aylik_kilit_ac(pid: str, yil: int = None, ay: int = None):
    """Onaylanmış (kilitli) maaş kaydını düzeltme için taslağa döndürür — son dakika
    fazla mesai/raporlu gibi durumlar için. ÖDENMİŞSE açılmaz (geçmiş değişmez;
    o durumda ek ödeme/mahsup gerekir)."""
    bugun = bugun_tr()
    yil = yil or bugun.year
    ay  = ay  or bugun.month
    maas_odeme_tarihi = _personel_maas_odeme_tarihi(yil, ay)
    with db() as (conn, cur):
        # Ödenmiş mi? Ödendiyse kilit açılmaz.
        cur.execute(
            """
            SELECT 1 FROM odeme_plani
            WHERE kaynak_tablo='personel' AND kaynak_id=%s AND durum='odendi'
              AND referans_ay = DATE_TRUNC('month', %s::date)
            LIMIT 1
            """,
            (pid, str(maas_odeme_tarihi)),
        )
        if cur.fetchone():
            raise HTTPException(
                400,
                "Bu ayın maaşı ödenmiş — kilit açılamaz. Düzeltme için ek ödeme / "
                "gelecek aydan mahsup gerekir (geçmiş kayıt değişmez).",
            )
        cur.execute(
            "UPDATE personel_aylik SET durum='taslak' WHERE personel_id=%s AND yil=%s AND ay=%s AND durum='onaylandi'",
            (pid, yil, ay),
        )
        if cur.rowcount == 0:
            raise HTTPException(400, "Onaylı kayıt bulunamadı (zaten taslak olabilir).")
        audit(cur, 'personel_aylik', pid, 'KILIT_AC', yeni={'yil': yil, 'ay': ay})
    return {"success": True, "mesaj": "Kilit açıldı — düzeltip tekrar onaylayın."}


@app.delete("/api/personel-aylik/{pid}")
def personel_aylik_sil(pid: str, yil: int = None, ay: int = None):
    """Personelin aylık maaş kaydını siler. Sadece taslak durumdakiler silinebilir."""
    bugun = bugun_tr()
    yil = yil or bugun.year
    ay  = ay  or bugun.month
    with db() as (conn, cur):
        cur.execute("SELECT * FROM personel_aylik WHERE personel_id=%s AND yil=%s AND ay=%s",
            (pid, yil, ay))
        kayit = cur.fetchone()
        if not kayit:
            raise HTTPException(404, "Kayıt bulunamadı")
        if kayit['durum'] == 'onaylandi':
            raise HTTPException(400, "Onaylanmış kayıt silinemez")
        cur.execute("DELETE FROM personel_aylik WHERE personel_id=%s AND yil=%s AND ay=%s",
            (pid, yil, ay))
        # Ödeme planını tahmini tutara geri döndür
        cur.execute("SELECT * FROM personel WHERE id=%s", (pid,))
        p = cur.fetchone()
        if p and p['calisma_turu'] == 'surekli':
            tahmini = float(p['maas'] or 0) + float(p['yemek_ucreti'] or 0) + float(p['yol_ucreti'] or 0)
            _personel_odeme_plani_senkronize(cur, dict(p), yil, ay, tahmini)
        audit(cur, 'personel_aylik', str(kayit['id']), 'DELETE')
    return {"success": True}

@app.get("/api/personel-aylik/{pid}/gecmis")
def personel_aylik_gecmis(pid: str):
    """Personelin son 12 aylık maaş geçmişini döner."""
    with db() as (conn, cur):
        cur.execute("""
            SELECT yil, ay, hesaplanan_net, durum, calisma_saati,
                   fazla_mesai_saat, bayram_mesai_saat, eksik_gun, manuel_duzeltme
            FROM personel_aylik WHERE personel_id=%s
            ORDER BY yil DESC, ay DESC LIMIT 12
        """, (pid,))
        return [dict(r) for r in cur.fetchall()]

# ── SABİT GİDERLER ─────────────────────────────────────────────
class SabitGider(BaseModel):
    gider_adi: str
    kategori: str
    tutar: float = 0        # degisken tipte 0 olabilir
    tip: str = 'sabit'      # 'sabit' = tutar belli | 'degisken' = tutar sonradan belli
    periyot: str = 'aylik'
    odeme_gunu: int = 1
    baslangic_tarihi: Optional[date] = None
    sube_id: Optional[str] = None
    gecerlilik_tarihi: Optional[date] = None
    sozlesme_sure_ay: Optional[int] = None
    kira_artis_periyot: Optional[str] = None
    kira_artis_tarihi: Optional[date] = None
    sozlesme_bitis_tarihi: Optional[date] = None
    # 💵 2026-08-09 (sahip): "elden ve havale diye ayrıştıralım; bazı ödemeler
    # nakit olsa bile elden ödenme ihtimali var, bu mutabakat doğru çalışmaz"
    #   elden  → kasadaki nakitten elden verildi → ELDE NAKİTİ AZALTIR
    #   havale → banka hesabından EFT            → elde nakiti etkilemez
    #   kart   → kredi kartı                     → kart borcunu büyütür
    #   nakit  → ESKİ/BELİRSİZ (elden mi havale mi seçilmemiş)
    odeme_yontemi: str = 'nakit'
    kart_id: Optional[str] = None  # Kart talimatı için
    stopaj_oran: float = 0         # kira stopajı: şahıstan işyeri kirasında 0.20; 0=stopajsız

KIRA_ARTIS_PERIYOT_MAP = {"6ay": 6, "1yil": 12, "2yil": 24, "5yil": 60}

@app.get("/api/sabit-giderler")
def sabit_giderler_listele():
    with db() as (conn, cur):
        cur.execute("""
            SELECT sg.*, s.ad as sube_adi,
              -- Bu ay ödendi mi? Nakit kasa / kart talimatı / ödenmiş plan — herhangi biri.
              (
                EXISTS (
                  SELECT 1 FROM kasa_hareketleri kh
                  WHERE kh.kaynak_tablo='sabit_giderler' AND kh.kaynak_id=sg.id
                    AND kh.islem_turu='SABIT_GIDER' AND kh.kasa_etkisi=true AND kh.durum='aktif'
                    AND DATE_TRUNC('month', kh.tarih) = DATE_TRUNC('month', CURRENT_DATE)
                )
                OR EXISTS (
                  SELECT 1 FROM kart_hareketleri kt
                  WHERE kt.kaynak_tablo='fatura_giderleri' AND kt.kaynak_id=sg.id
                    AND kt.islem_turu='HARCAMA' AND kt.durum='aktif'
                    AND DATE_TRUNC('month', kt.tarih) = DATE_TRUNC('month', CURRENT_DATE)
                )
                OR EXISTS (
                  SELECT 1 FROM odeme_plani op
                  WHERE op.kaynak_tablo='sabit_giderler' AND op.kaynak_id=sg.id
                    AND op.durum='odendi'
                    AND op.referans_ay = DATE_TRUNC('month', CURRENT_DATE)
                )
              ) AS bu_ay_odendi
            FROM sabit_giderler sg
            LEFT JOIN subeler s ON s.id=sg.sube_id
            ORDER BY sg.kategori, sg.gider_adi""")
        return [dict(r) for r in cur.fetchall()]

@app.post("/api/sabit-giderler")
def sabit_gider_ekle(g: SabitGider):
    with db() as (conn, cur):
        gid = str(uuid.uuid4())
        # Kira artış tarihi: periyot seçildiyse başlangıçtan hesapla
        kira_artis_tarihi = g.kira_artis_tarihi  # manuel girilmişse koru
        if g.baslangic_tarihi and g.kira_artis_periyot and g.kira_artis_periyot in KIRA_ARTIS_PERIYOT_MAP:
            kira_artis_tarihi = ay_ekle(g.baslangic_tarihi, KIRA_ARTIS_PERIYOT_MAP[g.kira_artis_periyot])
        # Sözleşme bitiş tarihi hesapla
        sozlesme_bitis = None
        if g.baslangic_tarihi and g.sozlesme_sure_ay:
            sozlesme_bitis = ay_ekle(g.baslangic_tarihi, g.sozlesme_sure_ay)
        cur.execute("""INSERT INTO sabit_giderler
            (id,gider_adi,kategori,tutar,tip,periyot,odeme_gunu,baslangic_tarihi,sube_id,
             sozlesme_sure_ay,kira_artis_periyot,kira_artis_tarihi,sozlesme_bitis_tarihi,
             odeme_yontemi,kart_id,stopaj_oran)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (gid, g.gider_adi, g.kategori, g.tutar, g.tip, g.periyot, g.odeme_gunu,
             g.baslangic_tarihi, g.sube_id or None,
             g.sozlesme_sure_ay, g.kira_artis_periyot, kira_artis_tarihi, sozlesme_bitis,
             g.odeme_yontemi, g.kart_id or None, max(0.0, min(1.0, float(g.stopaj_oran or 0)))))
        # Degisken gider: onay kuyruğuna girme — motor da plan üretmez, sadece hatırlatır
        # Kart talimatı: motor otomatik işler, onay kuyruğuna girme
        if g.tip == 'sabit' and g.odeme_yontemi != 'kart':
            onay_ekle(cur, 'SABIT_GIDER', 'sabit_giderler', gid,
                f"Sabit gider: {g.gider_adi}", g.tutar, bugun_tr())
        audit(cur, 'sabit_giderler', gid, 'INSERT')
    return {"id": gid, "success": True}

@app.put("/api/sabit-giderler/{gid}")
def sabit_gider_guncelle(gid: str, g: SabitGider):
    with db() as (conn, cur):
        cur.execute("SELECT * FROM sabit_giderler WHERE id=%s", (gid,))
        eski = cur.fetchone()
        if not eski: raise HTTPException(404)

        # Eksik alanları eski kayıttan tamamla — None kontrolü: 0 ve False korunmalı
        def _pick(yeni, eski_val, default=None):
            """Yeni değer None ise eskiyi al. 0 ve False geçerli değerlerdir."""
            return yeni if yeni is not None else (eski_val if eski_val is not None else default)

        gider_adi     = g.gider_adi   or eski['gider_adi']
        kategori      = g.kategori    or eski['kategori']
        periyot       = g.periyot     or eski['periyot'] or 'aylik'
        odeme_gunu    = _pick(g.odeme_gunu, eski['odeme_gunu'], 1)
        sube_id       = g.sube_id     or eski['sube_id']
        odeme_yontemi = g.odeme_yontemi or eski.get('odeme_yontemi') or 'nakit'
        kart_id       = g.kart_id     or eski.get('kart_id')

        # Eğer gecerlilik_tarihi belirtilmişse: eski kaydı kapat, yeni kayıt aç
        if g.gecerlilik_tarihi:
            # Eski kaydı kapat
            cur.execute("UPDATE sabit_giderler SET aktif=FALSE WHERE id=%s", (gid,))
            audit(cur, 'sabit_giderler', gid, 'KAPATILDI', eski=eski)
            # Eski sabit gidere ait bu ayki bekleyen ödeme planlarını iptal et
            cur.execute("""
                UPDATE odeme_plani SET durum='iptal'
                WHERE kaynak_tablo='sabit_giderler'
                AND kaynak_id=%s
                AND durum IN ('bekliyor','onay_bekliyor')
                AND EXTRACT(YEAR FROM tarih) = EXTRACT(YEAR FROM %s::date)
                AND EXTRACT(MONTH FROM tarih) = EXTRACT(MONTH FROM %s::date)
            """, (gid, str(g.gecerlilik_tarihi), str(g.gecerlilik_tarihi)))
            # Onay kuyruğundaki eski kaydı da iptal et
            cur.execute("""
                UPDATE onay_kuyrugu SET durum='reddedildi'
                WHERE kaynak_id=%s AND durum='bekliyor'
            """, (gid,))
            # Yeni kayıt aç — gecerlilik_tarihi'nden itibaren
            yeni_id = str(uuid.uuid4())
            kira_artis_tarihi_g = g.kira_artis_tarihi
            if g.gecerlilik_tarihi and g.kira_artis_periyot and g.kira_artis_periyot in KIRA_ARTIS_PERIYOT_MAP:
                kira_artis_tarihi_g = ay_ekle(g.gecerlilik_tarihi, KIRA_ARTIS_PERIYOT_MAP[g.kira_artis_periyot])
            sozlesme_bitis = None
            if g.gecerlilik_tarihi and g.sozlesme_sure_ay:
                sozlesme_bitis = ay_ekle(g.gecerlilik_tarihi, g.sozlesme_sure_ay)
            tip = g.tip or eski.get('tip') or 'sabit'
            cur.execute("""INSERT INTO sabit_giderler
                (id,gider_adi,kategori,tutar,tip,periyot,odeme_gunu,baslangic_tarihi,sube_id,
                 sozlesme_sure_ay,kira_artis_periyot,kira_artis_tarihi,sozlesme_bitis_tarihi,
                 odeme_yontemi,kart_id,stopaj_oran)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (yeni_id, gider_adi, kategori, g.tutar, tip, periyot,
                 odeme_gunu, g.gecerlilik_tarihi, sube_id,
                 g.sozlesme_sure_ay, g.kira_artis_periyot, kira_artis_tarihi_g, sozlesme_bitis,
                 odeme_yontemi, kart_id or None, max(0.0, min(1.0, float(g.stopaj_oran or 0)))))
            # KRİTİK 3: degisken gider onay kuyruğuna girmesin
            if tip == 'sabit' and odeme_yontemi != 'kart':
                onay_ekle(cur, 'SABIT_GIDER', 'sabit_giderler', yeni_id,
                    f"Sabit gider güncellendi: {gider_adi}", g.tutar, g.gecerlilik_tarihi)
            audit(cur, 'sabit_giderler', yeni_id, 'INSERT_GUNCELLEME')
            return {"success": True, "yeni_id": yeni_id}
        else:
            # Tarih belirtilmemişse — sadece bu kaydı güncelle
            tip_guncelle = g.tip or eski.get('tip') or 'sabit'
            cur.execute("""UPDATE sabit_giderler SET gider_adi=%s,kategori=%s,tutar=%s,
                tip=%s,periyot=%s,odeme_gunu=%s,baslangic_tarihi=%s,sube_id=%s,
                odeme_yontemi=%s,kart_id=%s,stopaj_oran=%s WHERE id=%s""",
                (gider_adi, kategori, g.tutar, tip_guncelle, periyot, odeme_gunu,
                 g.baslangic_tarihi, sube_id, odeme_yontemi, kart_id or None,
                 max(0.0, min(1.0, float(g.stopaj_oran or 0))), gid))
            # 🔴 P1 (2026-08-13, EVV-YUK / Codex): tutar yerinde güncellenince
            # BEKLEYEN onay satırının tutarı eski kalıyordu — 1.000₺ açılıp 1.500₺'ye
            # düzeltilen gider onaylanınca kasa 1.000 düşüyor, kayıt 1.500 görünüyordu.
            # Diff-review eki: tür DEĞİŞTİYSE (değişken ya da kart talimatı — artık
            # nakit onay gerektirmeyen hâller) bekleyen onay senkron değil İPTAL edilir;
            # yoksa stale kayıt pasif tür için kasadan nakit düşürürdü.
            if tip_guncelle == 'sabit' and odeme_yontemi != 'kart':
                cur.execute(
                    """UPDATE onay_kuyrugu SET tutar=%s, aciklama=%s
                       WHERE kaynak_tablo='sabit_giderler' AND kaynak_id=%s
                         AND islem_turu='SABIT_GIDER' AND durum='bekliyor'""",
                    (g.tutar, f"Sabit gider güncellendi: {gider_adi}", gid))
            else:
                cur.execute(
                    """UPDATE onay_kuyrugu SET durum='iptal'
                       WHERE kaynak_tablo='sabit_giderler' AND kaynak_id=%s
                         AND islem_turu='SABIT_GIDER' AND durum='bekliyor'""",
                    (gid,))
            audit(cur, 'sabit_giderler', gid, 'UPDATE', eski=eski)
        return {"success": True}

@app.delete("/api/sabit-giderler/{gid}")
def sabit_gider_sil(gid: str):
    with db() as (conn, cur):
        cur.execute("SELECT * FROM sabit_giderler WHERE id=%s AND aktif=TRUE", (gid,))
        eski = cur.fetchone()
        if not eski: raise HTTPException(404, "Kayıt bulunamadı veya zaten pasif")
        cur.execute("UPDATE sabit_giderler SET aktif=FALSE WHERE id=%s", (gid,))
        # Bekleyen ödeme planlarını iptal et — yetim plan kasayı/borç projeksiyonunu şişirmesin
        cur.execute(
            """
            UPDATE odeme_plani SET durum='iptal'
            WHERE kaynak_tablo='sabit_giderler' AND kaynak_id=%s
              AND durum IN ('bekliyor','onay_bekliyor')
            """,
            (gid,),
        )
        iptal_plan = cur.rowcount or 0
        # İlişkili bekleyen onay kuyruğu kayıtlarını da iptal et
        try:
            cur.execute(
                """
                UPDATE onay_kuyrugu SET durum='iptal'
                WHERE kaynak_tablo='odeme_plani' AND durum='bekliyor'
                  AND kaynak_id IN (
                    SELECT id FROM odeme_plani
                    WHERE kaynak_tablo='sabit_giderler' AND kaynak_id=%s
                  )
                """,
                (gid,),
            )
            # 🔴 P1 (2026-08-13, EVV-YUK / Codex): gider EKLENİRKEN açılan DOĞRUDAN
            # SABIT_GIDER onayı (kaynak_tablo='sabit_giderler') iptal EDİLMİYORDU —
            # kapatılan gidere bekleyen onay verilirse kasa çıkışı yazılırdı.
            cur.execute(
                """
                UPDATE onay_kuyrugu SET durum='iptal'
                WHERE kaynak_tablo='sabit_giderler' AND kaynak_id=%s AND durum='bekliyor'
                """,
                (gid,),
            )
        except Exception:
            pass
        audit(cur, 'sabit_giderler', gid, 'PASIF', eski=eski)
    return {"success": True, "iptal_edilen_plan": iptal_plan}

@app.get("/api/sabit-giderler/uyarilar")
def sabit_gider_uyarilar():
    """
    Kira/Abonelik uyarıları — iki bağımsız uyarı tipi:
    KIRA_ARTIS   : artış tarihi yaklaşıyor veya geçti → ödeme planı DURDU, tutar güncellenmeli
    SOZLESME_BITIS: sözleşme bitiyor veya bitti → uzatma/yenileme gerekiyor
    """
    bugun = bugun_tr()
    with db() as (conn, cur):
        cur.execute("""
            SELECT id, gider_adi, kategori, tutar, kira_artis_tarihi,
                   sozlesme_bitis_tarihi, kira_artis_periyot
            FROM sabit_giderler
            WHERE aktif = TRUE
            AND kategori IN ('Kira', 'Abonelik')
            AND (kira_artis_tarihi IS NOT NULL OR sozlesme_bitis_tarihi IS NOT NULL)
        """)
        kayitlar = cur.fetchall()

    uyarilar = []
    for r in kayitlar:

        # ── KIRA ARTIS ─────────────────────────────────────────────
        if r['kira_artis_tarihi']:
            gun_kalan = (r['kira_artis_tarihi'] - bugun).days

            if gun_kalan < 0:
                # Artış tarihi geçti — ödeme planı durdurulmuş, KRİTİK
                uyarilar.append({
                    'id': r['id'],
                    'tip': 'KIRA_ARTIS',
                    'seviye': 'KRITIK',
                    'durduruldu': True,        # plan üretimi durdu — sayaç için
                    'renk': 'red',
                    'gider_adi': r['gider_adi'],
                    'mesaj': (
                        f"⛔ {r['gider_adi']} — kira artış tarihi {abs(gun_kalan)} gün önce geçti! "
                        f"Yeni tutar girilene kadar ödeme planı üretilmiyor."
                    ),
                    'alt_mesaj': 'Mevcut tutar: ' + '{:,.0f} ₺'.format(float(r['tutar'])) + ' · Yeni tutarı ve artış tarihini güncelleyin',
                    'aksiyon': 'TUTAR_GUNCELLE',
                    'gun_kalan': gun_kalan,
                    'tarih': str(r['kira_artis_tarihi']),
                    'tutar': float(r['tutar'])
                })
            elif gun_kalan <= 15:
                # Artış yaklaşıyor — UYARI, plan henüz durmuş değil
                uyarilar.append({
                    'id': r['id'],
                    'tip': 'KIRA_ARTIS',
                    'seviye': 'UYARI',
                    'durduruldu': False,
                    'renk': 'yellow',
                    'gider_adi': r['gider_adi'],
                    'mesaj': f"⚠️ {r['gider_adi']} — kira artış tarihi {gun_kalan} gün sonra.",
                    'alt_mesaj': 'Mevcut tutar: ' + '{:,.0f} ₺'.format(float(r['tutar'])) + ' · Şimdiden yeni tutarı hazırlayın',
                    'aksiyon': 'TUTAR_GUNCELLE',
                    'gun_kalan': gun_kalan,
                    'tarih': str(r['kira_artis_tarihi']),
                    'tutar': float(r['tutar'])
                })

        # ── SÖZLEŞME BİTİŞ ─────────────────────────────────────────
        if r['sozlesme_bitis_tarihi']:
            gun_kalan = (r['sozlesme_bitis_tarihi'] - bugun).days

            if gun_kalan < 0:
                # Sözleşme süresi doldu — KRİTİK, ödeme planı durdurulmuş
                uyarilar.append({
                    'id': r['id'],
                    'tip': 'SOZLESME_BITIS',
                    'seviye': 'KRITIK',
                    'durduruldu': True,        # plan üretimi durdu — sayaç için
                    'renk': 'red',
                    'gider_adi': r['gider_adi'],
                    'mesaj': (
                        f"⛔ {r['gider_adi']} — sözleşme süresi {abs(gun_kalan)} gün önce doldu! "
                        f"Yenilenene kadar ödeme planı üretilmiyor."
                    ),
                    'alt_mesaj': 'Sözleşmeyi yenileyin: yeni süre ve başlangıç tarihini girin',
                    'aksiyon': 'SOZLESME_UZAT',
                    'gun_kalan': gun_kalan,
                    'tarih': str(r['sozlesme_bitis_tarihi']),
                    'tutar': float(r['tutar'])
                })
            elif gun_kalan <= 30:
                # Sözleşme yaklaşıyor — UYARI, plan henüz durmuş değil
                uyarilar.append({
                    'id': r['id'],
                    'tip': 'SOZLESME_BITIS',
                    'seviye': 'UYARI',
                    'durduruldu': False,
                    'renk': 'yellow',
                    'gider_adi': r['gider_adi'],
                    'mesaj': f"📋 {r['gider_adi']} — sözleşme {gun_kalan} gün sonra bitiyor.",
                    'alt_mesaj': 'Yenileme için hazırlık yapın',
                    'aksiyon': 'SOZLESME_UZAT',
                    'gun_kalan': gun_kalan,
                    'tarih': str(r['sozlesme_bitis_tarihi']),
                    'tutar': float(r['tutar'])
                })

    # Kritikler önce, sonra uyarılar; kendi içinde gün_kalan'a göre sırala
    uyarilar.sort(key=lambda x: (0 if x['seviye'] == 'KRITIK' else 1, x['gun_kalan']))
    return {"uyarilar": uyarilar, "adet": len(uyarilar)}

@app.get("/api/sabit-giderler/odenenler")
def sabit_gider_odenenler():
    """Gerçekleşmiş sabit gider ödemeleri — CFO görünürlük katmanı.
    İki yol birleştirilir: (a) ödeme planından ödenenler (odeme_plani),
    (b) manuel /fatura-ode ile ödenenler (kasa SABIT_GIDER, kaynak_tablo=sabit_giderler).
    Plan-nakit kasa kaydı kaynak_tablo='odeme_plani' taşıdığından çakışmaz; ayrıca aynı
    gider+ay için plan ödemesi varsa manuel satır NOT EXISTS ile elenir (çift sayım yok)."""
    with db() as (conn, cur):
        cur.execute("""
            SELECT * FROM (
                SELECT
                    op.id,
                    op.aciklama,
                    op.odenen_tutar,
                    op.odenecek_tutar,
                    op.odeme_tarihi,
                    op.tarih as plan_tarihi,
                    op.kaynak_id,
                    COALESCE(sg.gider_adi, op.aciklama) as gider_adi,
                    COALESCE(sg.kategori, '') as kategori
                FROM odeme_plani op
                LEFT JOIN sabit_giderler sg ON sg.id = op.kaynak_id
                WHERE op.durum = 'odendi'
                AND op.kaynak_tablo = 'sabit_giderler'

                UNION ALL

                SELECT
                    kh.id,
                    kh.aciklama,
                    ABS(kh.tutar) as odenen_tutar,
                    ABS(kh.tutar) as odenecek_tutar,
                    kh.tarih as odeme_tarihi,
                    kh.tarih as plan_tarihi,
                    kh.kaynak_id,
                    COALESCE(sg.gider_adi, kh.aciklama) as gider_adi,
                    COALESCE(sg.kategori, '') as kategori
                FROM kasa_hareketleri kh
                LEFT JOIN sabit_giderler sg ON sg.id = kh.kaynak_id
                WHERE kh.islem_turu = 'SABIT_GIDER'
                AND kh.kasa_etkisi = true AND kh.durum = 'aktif'
                AND kh.kaynak_tablo = 'sabit_giderler'
                AND NOT EXISTS (
                    SELECT 1 FROM odeme_plani op2
                    WHERE op2.kaynak_tablo='sabit_giderler' AND op2.kaynak_id=kh.kaynak_id
                      AND op2.durum='odendi'
                      AND op2.referans_ay = DATE_TRUNC('month', kh.tarih)
                )
            ) t
            ORDER BY t.odeme_tarihi DESC
            LIMIT 50
        """)
        return [dict(r) for r in cur.fetchall()]

@app.get("/api/sabit-giderler/odemeler")
def sabit_gider_odemeler(ay: str = None):
    """Ödenmiş + bekleyen + gecikmiş sabit giderler — CFO dashboard.
    Nakit: kasa_hareketleri SABIT_GIDER
    Kart: kart_hareketleri kaynak_tablo=sabit_giderler (kart talimatı)
    """
    with db() as (conn, cur):
        # Nakit ödenenler — kasa_hareketleri
        cur.execute("""
            SELECT
                kh.tarih,
                ABS(kh.tutar) as tutar,
                kh.aciklama,
                COALESCE(sg.gider_adi, kh.aciklama) as gider_adi,
                COALESCE(sg.kategori, '') as kategori,
                'odendi' as durum,
                'nakit' as odeme_yontemi,
                NULL as banka,
                NULL as kart_adi,
                kh.olusturma
            FROM kasa_hareketleri kh
            LEFT JOIN sabit_giderler sg ON sg.id = kh.kaynak_id
            WHERE kh.islem_turu = 'SABIT_GIDER'
            AND kh.kasa_etkisi = true AND kh.durum = 'aktif'
            ORDER BY kh.tarih DESC
            LIMIT 200
        """)
        nakit_odenenler = [dict(r) for r in cur.fetchall()]

        # Kart ödenenler — kart_hareketleri (kart talimatı ile)
        cur.execute("""
            SELECT
                kh.tarih,
                kh.tutar,
                kh.aciklama,
                COALESCE(sg.gider_adi, kh.aciklama) as gider_adi,
                COALESCE(sg.kategori, '') as kategori,
                'odendi' as durum,
                'kart' as odeme_yontemi,
                k.banka,
                k.kart_adi,
                kh.olusturma
            FROM kart_hareketleri kh
            JOIN kartlar k ON k.id = kh.kart_id
            LEFT JOIN sabit_giderler sg ON sg.id = kh.kaynak_id
            WHERE kh.islem_turu = 'HARCAMA' AND kh.durum = 'aktif'
            -- 🟡 P2 (2026-08-13, Codex): fiili kart talimatı yazımı 'fatura_giderleri'
            -- kaynak_tablo'suyla da oluyor (bu_ay_odendi kontrolü onu arıyor);
            -- burada yalnız 'sabit_giderler' sayılınca kartla ödenen gider bu
            -- listede kayboluyordu. İki kaynak da sayılır.
            AND kh.kaynak_tablo IN ('sabit_giderler', 'fatura_giderleri')
            ORDER BY kh.tarih DESC
            LIMIT 200
        """)
        kart_odenenler = [dict(r) for r in cur.fetchall()]

        odenenler = nakit_odenenler + kart_odenenler
        odenenler.sort(key=lambda x: str(x['tarih']), reverse=True)

        # Bekleyen + gecikmiş — odeme_plani üzerinden (nakit: bekliyor, kart: motor zaten işledi)
        cur.execute("""
            SELECT
                op.tarih,
                op.odenecek_tutar as tutar,
                op.aciklama,
                COALESCE(sg.gider_adi, op.aciklama) as gider_adi,
                COALESCE(sg.kategori, '') as kategori,
                CASE
                    WHEN op.tarih < CURRENT_DATE THEN 'gecikti'
                    ELSE 'bekliyor'
                END as durum,
                op.olusturma
            FROM odeme_plani op
            LEFT JOIN sabit_giderler sg ON sg.id = op.kaynak_id
            WHERE op.kaynak_tablo = 'sabit_giderler'
            AND op.durum IN ('bekliyor', 'onay_bekliyor')
            AND op.tarih >= DATE '2026-06-01'   -- sistem başlangıcı: Haziran öncesi gösterilmez
            ORDER BY op.tarih ASC
        """)
        bekleyenler = [dict(r) for r in cur.fetchall()]

        # Özet
        nakit_odenen = sum(float(r['tutar']) for r in nakit_odenenler)
        kart_odenen  = sum(float(r['tutar']) for r in kart_odenenler)
        toplam_odenen = nakit_odenen + kart_odenen
        toplam_bekleyen = sum(float(r['tutar']) for r in bekleyenler)
        geciken = [r for r in bekleyenler if r['durum'] == 'gecikti']

        return {
            "odenenler": odenenler,
            "bekleyenler": bekleyenler,
            "ozet": {
                "toplam_odenen": toplam_odenen,
                "nakit_odenen": nakit_odenen,
                "kart_odenen": kart_odenen,
                "toplam_bekleyen": toplam_bekleyen,
                "geciken_adet": len(geciken),
                "geciken_tutar": sum(float(r['tutar']) for r in geciken),
                "odenenler": odenenler
            }
        }

@app.post("/api/odeme-plani/gecmis-temizle")
def odeme_plani_gecmis_temizle(baslangic: str = "2026-06-01", uygula: bool = False):
    """Sistem başlangıç tarihinden ÖNCEKİ bekleyen/onay_bekleyen ödeme planlarını
    borç listesinden çıkarır (durum='iptal' — SİLİNMEZ, geri alınabilir).

    - uygula=False (varsayılan): yalnızca önizleme — neyin iptal edileceğini listeler.
    - uygula=True: iptal işlemini uygular.
    Haziran (ve sonrası) ödemelere dokunmaz. 'odendi' kayıtlar korunur.
    """
    from datetime import date as _date
    try:
        kesim = _date.fromisoformat(baslangic)
    except Exception:
        raise HTTPException(400, "baslangic tarihi geçersiz (YYYY-MM-DD bekleniyor)")

    with db() as (conn, cur):
        cur.execute(
            """
            SELECT id::text, tarih, odenecek_tutar, durum, kaynak_tablo, aciklama
            FROM odeme_plani
            WHERE durum IN ('bekliyor','onay_bekliyor') AND tarih < %s
            ORDER BY tarih ASC
            """,
            (kesim,),
        )
        adaylar = [dict(r) for r in (cur.fetchall() or [])]
        toplam = sum(float(r["odenecek_tutar"] or 0) for r in adaylar)
        liste = [
            {
                "tarih": str(r["tarih"])[:10],
                "tutar": float(r["odenecek_tutar"] or 0),
                "durum": r["durum"],
                "kaynak": r["kaynak_tablo"],
                "aciklama": str(r["aciklama"] or "")[:60],
            }
            for r in adaylar
        ]

        if not uygula:
            return {
                "onizleme": True,
                "baslangic": str(kesim),
                "iptal_edilecek_adet": len(adaylar),
                "iptal_edilecek_tutar": round(toplam, 2),
                "kayitlar": liste,
                "not": "uygula=true ile iptal edilir. Hiçbiri silinmez, durum='iptal' olur.",
            }

        cur.execute(
            """
            UPDATE odeme_plani
            SET durum='iptal',
                aciklama = COALESCE(aciklama,'') || ' · iptal: sistem başlangıcı ' || %s
            WHERE durum IN ('bekliyor','onay_bekliyor') AND tarih < %s
            """,
            (str(kesim), kesim),
        )
        iptal_adet = cur.rowcount
        # FIX MN10 (2026-07-06): iptal edilen planlara bağlı BEKLEYEN onay kuyruğu kayıtları
        # güncellenmiyordu → yetim onaylar panelde sonsuza dek "bekliyor" görünüyordu.
        onay_kapatilan = 0
        if adaylar:
            _plan_ids = [r["id"] for r in adaylar]
            cur.execute(
                """
                UPDATE onay_kuyrugu SET durum='reddedildi', onay_tarihi=NOW()
                WHERE durum NOT IN ('onaylandi','reddedildi') AND kaynak_id = ANY(%s)
                """,
                (_plan_ids,),
            )
            onay_kapatilan = cur.rowcount
        return {
            "onizleme": False,
            "baslangic": str(kesim),
            "iptal_edilen_adet": iptal_adet,
            "kapatilan_onay_adet": onay_kapatilan,
            "iptal_edilen_tutar": round(toplam, 2),
            "kayitlar": liste,
        }


@app.post("/api/sabit-giderler/odenmis-plan-esitle")
def sabit_gider_odenmis_plan_esitle(uygula: bool = False):
    """Geri-doldurma: kasada ÖDENMİŞ (SABIT_GIDER) kaydı olduğu hâlde bağlı ödeme planı
    hâlâ 'bekliyor'/'onay_bekliyor' kalmış sabit giderleri 'odendi' yapar.
    Bunlar 'ödenmiş giderler'de görünüp panelde aynı anda 'vadesi geçmiş borç' görünen
    çift-gösterim kayıtlarıdır. Önizleme (uygula=False) → liste; uygula=True → düzeltir.
    İlişkili bekleyen onay kuyruğu da kapatılır."""
    with db() as (conn, cur):
        cur.execute("""
            SELECT op.id::text, op.tarih, op.odenecek_tutar, op.durum, op.kaynak_id::text,
                   COALESCE(sg.gider_adi, op.aciklama) AS gider_adi
            FROM odeme_plani op
            LEFT JOIN sabit_giderler sg ON sg.id = op.kaynak_id
            WHERE op.kaynak_tablo='sabit_giderler'
              AND op.durum IN ('bekliyor','onay_bekliyor')
              AND EXISTS (
                  SELECT 1 FROM kasa_hareketleri kh
                  WHERE kh.kaynak_tablo='sabit_giderler' AND kh.kaynak_id=op.kaynak_id
                    AND kh.islem_turu='SABIT_GIDER' AND kh.kasa_etkisi=true AND kh.durum='aktif'
                    AND DATE_TRUNC('month', kh.tarih) = DATE_TRUNC('month', op.tarih)
              )
            ORDER BY op.tarih ASC
        """)
        adaylar = [dict(r) for r in (cur.fetchall() or [])]
        liste = [{"tarih": str(r["tarih"])[:10], "tutar": float(r["odenecek_tutar"] or 0),
                  "durum": r["durum"], "gider_adi": r["gider_adi"]} for r in adaylar]
        if not uygula:
            return {"onizleme": True, "esitlenecek_adet": len(adaylar), "kayitlar": liste,
                    "not": "uygula=true ile bu planlar 'odendi' yapılır."}
        esit = 0
        if adaylar:
            cur.execute("""
                UPDATE odeme_plani op SET durum='odendi',
                    odeme_tarihi=COALESCE(op.odeme_tarihi, CURRENT_DATE)
                WHERE op.kaynak_tablo='sabit_giderler'
                  AND op.durum IN ('bekliyor','onay_bekliyor')
                  AND EXISTS (
                      SELECT 1 FROM kasa_hareketleri kh
                      WHERE kh.kaynak_tablo='sabit_giderler' AND kh.kaynak_id=op.kaynak_id
                        AND kh.islem_turu='SABIT_GIDER' AND kh.kasa_etkisi=true AND kh.durum='aktif'
                        AND DATE_TRUNC('month', kh.tarih) = DATE_TRUNC('month', op.tarih)
                  )
            """)
            esit = cur.rowcount
            # İlişkili bekleyen onayları da kapat
            cur.execute("""
                UPDATE onay_kuyrugu ok SET durum='onaylandi', onay_tarihi=NOW()
                WHERE ok.kaynak_tablo='sabit_giderler' AND ok.durum='bekliyor'
                AND EXISTS (
                    SELECT 1 FROM kasa_hareketleri kh
                    WHERE kh.kaynak_tablo='sabit_giderler' AND kh.kaynak_id=ok.kaynak_id
                      AND kh.islem_turu='SABIT_GIDER' AND kh.kasa_etkisi=true AND kh.durum='aktif'
                      AND DATE_TRUNC('month', kh.tarih) = DATE_TRUNC('month', ok.tarih)
                )
            """)
        return {"onizleme": False, "esitlenen_adet": esit, "kayitlar": liste}


@app.get("/api/kartlar/odeme-eslestir")
def kart_odeme_eslestir(gun: int = 120, oran_tol: float = 0.005, tutar_tol: float = 1.0):
    """💳 EKSTRE = ÖDEME KANITI — kartla ödenmiş faturayı kuyrukta açık bırakma.

    Sahip (2026-08-10): "faturalar şu anda karttan ödendiği için otomatik talimat
    kart ekstresi gelene kadar açık kalabilir; ama ekstre yüklendiğinde ekstreler
    bir öncenin fatura ödemesini yapar — burada TARİH önemli, eşleşmeyi yapmalı ve
    'karttan ödendi' olarak düzenlemeli."

    İŞLEYİŞ: elektrik/su/internet gibi otomatik talimatlı faturalar karttan çeker.
    Fatura sisteme girildiğinde ödeme planına düşer ve ödenmiş görünmez; para
    aslında ÇIKMIŞTIR ama izi kart ekstresindedir. Ekstre gelince o iz belirir.
    Bu uç iki tarafı buluşturur.

    MUHASEBE (çift sayım yok): fatura kartla ödendiyse fatura borcu kapanır,
    kart borcu zaten kart_hareketleri'nde durur, KASA'ya DOKUNULMAZ — kasa çıkışı
    kart ekstresi ödendiğinde olur. Aynı para iki kez çıkmış görünmez.

    ⚠️ BAĞLAMA ≠ KAPATMA: bu uç yalnız ÖNERİ üretir (salt-okur). Kapatma
    /api/odeme-plani/{oid}/karttan-odendi ile ve sahip onayıyla yapılır.
    """
    with db() as (conn, cur):
        cur.execute(
            """SELECT op.id, op.tarih::text AS vade, op.aciklama, op.kaynak_tablo,
                      COALESCE(op.odenecek_tutar,0)::float AS tutar,
                      COALESCE(op.odenen_tutar,0)::float   AS odenen,
                      va.tedarikci AS ted
                 FROM odeme_plani op
                 LEFT JOIN vadeli_alimlar va
                        ON op.kaynak_tablo='vadeli_alimlar' AND va.id = op.kaynak_id
                WHERE op.durum='bekliyor'
                  AND op.kart_id IS NULL
                  AND COALESCE(op.kaynak_tablo,'') NOT IN ('personel', 'borc_envanteri')
                  AND op.tarih >= CURRENT_DATE - %s""",
            (int(gun),),
        )
        planlar = [dict(r) for r in (cur.fetchall() or [])]
        cur.execute(
            """SELECT h.id, h.tarih::text AS tarih, h.aciklama, ABS(h.tutar)::float AS tutar,
                      k.kart_adi
                 FROM kart_hareketleri h
                 JOIN kartlar k ON k.id = h.kart_id
                WHERE h.islem_turu='HARCAMA' AND COALESCE(h.durum,'aktif')='aktif'
                  AND COALESCE(h.harcama_tipi,'belirsiz') <> 'sahsi'
                  AND h.tarih >= CURRENT_DATE - %s""",
            (int(gun) + 60,),
        )
        hareketler = [dict(r) for r in (cur.fetchall() or [])]

    import re as _re
    from datetime import date as _d
    # Türkçe sadeleştirme tek merkezden (fatura_api._sadele) — ayrı bir kopya
    # yazmak "ÖDEMESİ"/"ODEMESI" gibi Türkçe-I sapmalarını geri getirirdi.
    try:
        from fatura_api import _sadele as _sad
    except Exception:  # noqa: BLE001
        def _sad(x):  # emniyet ağı: sadeleştirme olmasa da eşleştirme çalışsın
            return str(x or "")

    def _kelimeler(s):
        s = _sad(str(s or "")).upper()
        s = _re.sub(r"[^A-Z0-9 ]", " ", s)
        # Jenerik kelimeler kanıt sayılmaz (fatura/ödeme/ay adları vb.) — aksi
        # halde "FATURA" kelimesi her şeyi her şeye eşler.
        atil = {"FATURA", "ODEME", "ODEMESI", "TAKSIT", "TAKSITI", "TR", "LTD", "STI",
                "SAN", "TIC", "ANONIM", "SIRKETI", "LIMITED", "SIRKET", "TICARET",
                "SANAYI", "VE", "ILE", "KONYA", "ISTANBUL", "IZMIR", "ANKARA",
                "KARAMAN", "MERKEZ", "SUBE", "CARI", "BORC", "KREDI", "GIDER",
                "GIDERI", "SABIT", "VADELI", "ALIM", "ALIMI", "KISMI", "AIDAT",
                "HIZMETLERI", "HIZMET", "URUN", "GENEL", "DIGER", "TUTAR"}
        return {w for w in s.split() if len(w) >= 4 and w not in atil}

    def _gun(a, b):
        try:
            return abs((_d.fromisoformat(a[:10]) - _d.fromisoformat(b[:10])).days)
        except Exception:
            return 999

    kullanilan = set()
    oneriler = []
    for p in planlar:
        kalan = round(p["tutar"] - p["odenen"], 2)
        if kalan <= 0:
            continue
        p_kel = _kelimeler(p["aciklama"]) | _kelimeler(p.get("ted"))
        adaylar = []
        for h in hareketler:
            if h["id"] in kullanilan:
                continue
            fark = abs(h["tutar"] - kalan)
            if fark > max(tutar_tol, kalan * oran_tol):
                continue
            # TARİH KURALI (sahibin vurgusu): ekstre satırı, vadesi GELMİŞ bir
            # faturanın ödemesidir. Kart çekimi vadeden biraz önce de olabilir
            # (otomatik talimat erken çeker), çok sonra da (gecikmiş talimat).
            g = _gun(h["tarih"], p["vade"])
            if g > 45:
                continue
            ortak = p_kel & _kelimeler(h["aciklama"])
            # ⚠️ AD ÖRTÜŞMESİ ZORUNLU (2026-08-10 öz-denetim): ilk sürüm yalnız
            # tutar+tarihe bakıyordu ve canlıda ürettiği 9 önerinin HEPSİ yanlıştı
            # ("AKALIN 684 ₺ ↔ OVOLT ŞARJ 659 ₺", "banka kredi taksiti ↔ EVA
            # MUTFAK"). Sıfır öneri, dokuz yanlış öneriden iyidir — yanlış öneri
            # sahip onaylarsa gerçek bir borcu sahte kapatır.
            if not ortak:
                continue
            # GÜVEN: tutar tek başına kanıt değil (aynı tutarlı iki fatura olur).
            skor = 0
            skor += 40 if fark <= 0.01 else (25 if fark <= max(1.0, kalan * 0.005) else 10)
            skor += 30 if g <= 7 else (20 if g <= 20 else 10)
            skor += 15 * min(2, len(ortak))
            adaylar.append({"hareket": h, "skor": skor, "gun_fark": g,
                            "tutar_fark": round(fark, 2), "ortak_kelime": sorted(ortak)})
        if not adaylar:
            continue
        adaylar.sort(key=lambda x: -x["skor"])
        en = adaylar[0]
        kullanilan.add(en["hareket"]["id"])
        oneriler.append({
            "plan_id": p["id"], "plan_aciklama": p["aciklama"], "tedarikci": p.get("ted"),
            "vade": p["vade"], "kalan_tutar": kalan, "kaynak_tablo": p["kaynak_tablo"],
            "kart_hareket_id": en["hareket"]["id"], "kart_adi": en["hareket"]["kart_adi"],
            "kart_aciklama": en["hareket"]["aciklama"], "kart_tarih": en["hareket"]["tarih"],
            "kart_tutar": en["hareket"]["tutar"],
            "skor": en["skor"], "gun_fark": en["gun_fark"], "tutar_fark": en["tutar_fark"],
            "ortak_kelime": en["ortak_kelime"],
            "guven": ("yuksek" if en["skor"] >= 85 else "orta" if en["skor"] >= 60 else "dusuk"),
            "diger_aday": len(adaylar) - 1,
        })
    oneriler.sort(key=lambda x: (-x["skor"], -x["kalan_tutar"]))
    _y = [o for o in oneriler if o["guven"] == "yuksek"]
    return {
        "gun": int(gun), "oneri_adet": len(oneriler),
        "oneri_tutar": round(sum(o["kalan_tutar"] for o in oneriler), 2),
        "yuksek_guven_adet": len(_y),
        "yuksek_guven_tutar": round(sum(o["kalan_tutar"] for o in _y), 2),
        "taranan_plan": len(planlar), "taranan_kart_hareketi": len(hareketler),
        "oneriler": oneriler,
        "not": ("Salt-okur öneri. Kapatma sahibin onayıyla /api/odeme-plani/{id}/"
                "karttan-odendi ucundan yapılır; kasaya dokunulmaz çünkü para "
                "kart ekstresi ödenirken çıkar."),
    }


@app.post("/api/odeme-plani/{oid}/karttan-odendi")
def odeme_plani_karttan_odendi(oid: str, kart_hareket_id: str = "", gerekce: str = "", kuru: int = 1):
    """Kuyruk kalemini 'karttan ödendi' olarak kapatır — KASA HAREKETİ YARATMAZ.

    Para kart ekstresi ödendiğinde çıkacak; burada ikinci bir kasa kaydı açmak
    çift sayım olurdu. Hangi kart hareketinin kanıt olduğu açıklamaya damgalanır,
    denetim izi kalır ve karar geri alınabilir.
    """
    with db() as (conn, cur):
        cur.execute("""SELECT id, tarih::text AS vade, aciklama, durum,
                              COALESCE(odenecek_tutar,0)::float AS tutar,
                              COALESCE(odenen_tutar,0)::float AS odenen
                         FROM odeme_plani WHERE id=%s FOR UPDATE""", (oid,))
        p = cur.fetchone()
        if not p:
            raise HTTPException(404, "Plan kalemi bulunamadı")
        p = dict(p)
        if p["durum"] != "bekliyor":
            raise HTTPException(409, f"Kalem '{p['durum']}' durumunda — yalnız 'bekliyor' kapatılabilir")
        kalan = round(float(p["tutar"]) - float(p["odenen"]), 2)
        kanit = ""
        if kart_hareket_id:
            cur.execute("""SELECT h.tarih::text AS tarih, ABS(h.tutar)::float AS tutar,
                                  h.aciklama, k.kart_adi
                             FROM kart_hareketleri h JOIN kartlar k ON k.id=h.kart_id
                            WHERE h.id=%s""", (kart_hareket_id,))
            kh = cur.fetchone()
            if not kh:
                raise HTTPException(404, "Kart hareketi bulunamadı")
            kh = dict(kh)
            kanit = f"{kh['kart_adi']} · {kh['tarih']} · {kh['tutar']:,.2f} ₺".replace(",", ".")
        _damga = (f"{p['aciklama'] or ''} · [karttan ödendi {bugun_tr().isoformat()}"
                  f"{' — ' + kanit if kanit else ''}"
                  f"{': ' + gerekce.strip() if gerekce.strip() else ''}]")[:500]
        if kuru:
            return {"kuru": True, "plan_id": oid, "vade": p["vade"], "kapanacak_tutar": kalan,
                    "kanit": kanit or None, "yeni_aciklama": _damga, "kasa_etkisi": False,
                    "mesaj": "PROVA — hiçbir kayıt değişmedi. Uygulamak için kuru=0 gönderin."}
        cur.execute(
            """UPDATE odeme_plani
                  SET durum='odendi', odenen_tutar=%s, odeme_tarihi=CURRENT_DATE, aciklama=%s
                WHERE id=%s AND durum='bekliyor'""",
            (p["tutar"], _damga, oid),
        )
        if cur.rowcount == 0:
            raise HTTPException(409, "Kalem bu sırada değişti — tekrar deneyin")
        audit(cur, 'odeme_plani', oid, 'KARTTAN_ODENDI',
              yeni={'tutar': kalan, 'kanit': kanit, 'kasa_etkisi': False})
    return {"success": True, "plan_id": oid, "kapanan_tutar": kalan, "kanit": kanit or None,
            "kasa_etkisi": False,
            "mesaj": ("Kalem 'karttan ödendi' olarak kapatıldı. Kasa hareketi oluşmadı — "
                      "para kart ekstresi ödendiğinde çıkacak.")}


@app.post("/api/odeme-plani/{oid}/cari-odemesiyle-kapat")
def odeme_plani_cari_odemesiyle_kapat(oid: str, gerekce: str = "", kuru: int = 1):
    """💳 CARİ ÖDEMESİYLE KAPAT — para zaten çıkmış, yalnız kuyruk kapanır.

    Sahip (2026-08-09): "FEZ'e yapılmış ödeme var, fazla tutar olsa da;
    kapanmış, kalan borcun vadesi diğer aya gitmesi lazımdı."

    Canlı vaka: FEZ'e 27.07'de 70.000 ₺ ödendi ama ödeme CARİ HESABA yapıldı
    ("Cari borç ödemesi — FEZ", kaynak anlik_giderler). Plan kalemi ise
    vadeli_alimlar kaynaklı; ikisi arasında bağ yok, bu yüzden
    /odenmis-plan-esitle bunu yakalayamıyor. Sonuç: 35.148 ₺'lik fatura
    ödendiği hâlde kuyrukta duruyor ve gecikmiş görünüyor.

    ⚠️ KASA HAREKETİ YARATMAZ. Para zaten bir kez çıktı; ikinci kayıt açmak
    çift sayım olurdu. Bu uç yalnızca kuyruk satırını kapatır ve NEDEN
    kapatıldığını açıklamaya damgalar — denetim izi kalır, geri alınabilir.
    """
    with db() as (conn, cur):
        cur.execute("""SELECT id, tarih::text AS vade, aciklama, durum,
                              COALESCE(odenecek_tutar,0)::float AS tutar,
                              COALESCE(odenen_tutar,0)::float AS odenen,
                              kaynak_tablo, kaynak_id
                         FROM odeme_plani WHERE id=%s FOR UPDATE""", (oid,))
        p = cur.fetchone()
        if not p:
            raise HTTPException(404, "Plan kalemi bulunamadı")
        p = dict(p)
        if p["durum"] != "bekliyor":
            raise HTTPException(409, f"Kalem '{p['durum']}' durumunda — yalnız 'bekliyor' kapatılabilir")
        kalan = round(float(p["tutar"]) - float(p["odenen"]), 2)
        _damga = (f"{p['aciklama'] or ''} · [cari ödemesiyle kapatıldı "
                  f"{bugun_tr().isoformat()}{': ' + gerekce.strip() if gerekce.strip() else ''}]")[:500]
        if kuru:
            return {"kuru": True, "plan_id": oid, "vade": p["vade"],
                    "kapanacak_tutar": kalan, "aciklama": p["aciklama"],
                    "yeni_aciklama": _damga,
                    # Sürüm işareti: v2 = vadeli_alimlar'a DOKUNMAYAN sürüm.
                    # Deploy yayılmadan uygulamamak için çağıran bunu kontrol eder
                    # (v1 canlıda cari hesabı 35.148 ₺ fazla düşürmüştü).
                    "surum": "v2-vadeli-dokunmaz",
                    "not": "kuru=0 ile uygulanır. Kasa hareketi YARATILMAZ, "
                           "vadeli_alimlar kaydına DOKUNULMAZ (cari hesap korunur)."}
        cur.execute("""UPDATE odeme_plani
                          SET durum='odendi', odeme_tarihi=%s,
                              odenen_tutar=%s, aciklama=%s
                        WHERE id=%s AND durum='bekliyor'""",
                    (bugun_tr(), float(p["tutar"]), _damga, oid))
        _n = cur.rowcount or 0
        # ⛔ VADELİ ALIMA DOKUNULMAZ (2026-08-10, canlıda yaşandı ve geri alındı)
        # İlk sürüm kaynak `vadeli_alimlar` kaydını da 'odendi' yapıyordu.
        # Ama cari ekstre `vadeli_alimlar durum='odendi'` satırlarını ÖDEME İZİ
        # sayar — yani aynı 70.000 ₺'lik ödeme İKİNCİ KEZ düşüyor:
        #     FEZ cari açık 51.428,59 → 16.280,59  (35.148 ₺ fazladan)
        # Bu uç yalnız KUYRUĞU düzeltir; cari hesap zaten doğrudur ve ona
        # dokunmak çift sayım yaratır. Cari ESAS, kuyruk YORUMDUR.
        audit(cur, 'odeme_plani', oid, 'CARI_ODEMESIYLE_KAPAT',
              eski={"durum": "bekliyor", "kalan": kalan},
              yeni={"durum": "odendi", "gerekce": gerekce[:200]})
    return {"success": bool(_n), "plan_id": oid, "kapatilan_tutar": kalan,
            "kasa_hareketi": "YARATILMADI — para daha önce çıkmıştı",
            "geri_alma": "Ödeme Merkezi'nden kalemi yeniden açabilirsiniz"}


@app.post("/api/odeme-plani/odenmis-plan-esitle")
def odeme_plani_odenmis_esitle(uygula: bool = False):
    """Geri-doldurma (TÜM kaynaklar): kasada ÖDENMİŞ kasa hareketi olduğu hâlde bağlı
    ödeme planı hâlâ 'bekliyor'/'onay_bekliyor' kalmış kayıtları 'odendi' yapar.
    /api/sabit-giderler/odenmis-plan-esitle'nin genel hâli — borc_envanteri, vadeli_alimlar,
    personel ve sabit_giderler için de çalışır.

    Sebep: ödeme yolları planı 'odendi' yaparken DATE_TRUNC('month', plan.tarih)=ödeme ayı
    filtresi kullanır; ödeme planın vade ayından farklı bir ayda yapıldıysa UPDATE 0 satır
    eşler → para kasadan çıkar ama plan 'bekliyor' kalır, panelde gecikmiş borç görünür.

    Kasa kaydı iki şekilde bağlanmış olabilir:
      (a) kaynak_tablo='odeme_plani' AND kaynak_id=plan.id (borç/kart anapara)
      (b) kaynak_tablo=plan.kaynak_tablo AND kaynak_id=plan.kaynak_id (sabit/vadeli/personel)
    Sadece kart_id'siz (nakit/kaynak bağlı) planlar — kart planlarının kendi guard'ı vardır.
    Önizleme (uygula=False) → liste; uygula=True → düzeltir + ilişkili onay kuyruğunu kapatır."""
    # İz tanımı = odeme_plani_kasa_mutabakat ile AYNI: (a) plan-bağlı iz ay filtresiz,
    # (b) kaynak-bağlı iz sadece plan ayında (tekrarlayan kalem güvenliği).
    _esit_kosul = """
        op.kart_id IS NULL
        AND op.durum IN ('bekliyor','onay_bekliyor')
        AND EXISTS (
            SELECT 1 FROM kasa_hareketleri kh
            WHERE kh.kasa_etkisi = TRUE AND kh.durum = 'aktif'
              AND (
                    (kh.kaynak_tablo = 'odeme_plani' AND kh.kaynak_id = op.id)
                 OR (op.kaynak_id IS NOT NULL
                     AND kh.kaynak_tablo = op.kaynak_tablo AND kh.kaynak_id = op.kaynak_id
                     AND DATE_TRUNC('month', kh.tarih) = DATE_TRUNC('month', op.tarih))
              )
        )
    """
    with db() as (conn, cur):
        cur.execute(f"""
            SELECT op.id::text, op.tarih, op.odenecek_tutar, op.durum,
                   op.kaynak_tablo, op.kaynak_id::text, op.aciklama
            FROM odeme_plani op
            WHERE {_esit_kosul}
            ORDER BY op.tarih ASC
        """)
        adaylar = [dict(r) for r in (cur.fetchall() or [])]
        liste = [{"tarih": str(r["tarih"])[:10], "tutar": float(r["odenecek_tutar"] or 0),
                  "durum": r["durum"], "kaynak_tablo": r["kaynak_tablo"],
                  "aciklama": r["aciklama"]} for r in adaylar]
        if not uygula:
            return {"onizleme": True, "esitlenecek_adet": len(adaylar), "kayitlar": liste,
                    "not": "uygula=true ile bu planlar 'odendi' yapılır."}
        esit = 0
        if adaylar:
            cur.execute(f"""
                UPDATE odeme_plani op SET durum='odendi',
                    odeme_tarihi=COALESCE(op.odeme_tarihi, CURRENT_DATE)
                WHERE {_esit_kosul}
            """)
            esit = cur.rowcount
            # İlişkili bekleyen onayları da kapat (plan id veya kaynak id ile bağlı)
            cur.execute("""
                UPDATE onay_kuyrugu ok SET durum='onaylandi', onay_tarihi=NOW()
                WHERE ok.durum='bekliyor'
                AND EXISTS (
                    SELECT 1 FROM odeme_plani op
                    WHERE op.durum='odendi'
                      AND (ok.kaynak_id = op.id::text OR
                           (ok.kaynak_tablo = op.kaynak_tablo AND ok.kaynak_id = op.kaynak_id::text))
                )
            """)
        return {"onizleme": False, "esitlenen_adet": esit, "kayitlar": liste}


@app.get("/api/odeme-plani/mukerrer-tara")
def odeme_plani_mukerrer_tara():
    """Tanı (read-only): aktif mükerrer ödeme planı grupları + mükerrer-engel index'i kurulu mu.
    Aynı (kaynak_tablo, kaynak_id, referans_ay) için >1 aktif (iptal değil) plan = çift ödeme riski."""
    with db() as (conn, cur):
        cur.execute(
            """
            SELECT kaynak_tablo, kaynak_id, referans_ay::text AS referans_ay,
                   COUNT(*) AS adet,
                   ARRAY_AGG(id::text) AS plan_idler,
                   ARRAY_AGG(durum) AS durumlar,
                   ARRAY_AGG(aciklama) AS aciklamalar
            FROM odeme_plani
            WHERE durum <> 'iptal' AND kaynak_id IS NOT NULL AND referans_ay IS NOT NULL
            GROUP BY kaynak_tablo, kaynak_id, referans_ay
            HAVING COUNT(*) > 1
            ORDER BY adet DESC
            """
        )
        gruplar = [dict(r) for r in (cur.fetchall() or [])]
        cur.execute(
            "SELECT 1 FROM pg_indexes WHERE indexname = 'uq_odeme_plani_kaynak_ay_aktif'"
        )
        index_var = cur.fetchone() is not None
    return {
        "mukerrer_grup_adet": len(gruplar),
        "mukerrer_engel_index_kurulu": index_var,
        "gruplar": gruplar,
    }


@app.post("/api/odeme-plani/mukerrer-temizle")
def odeme_plani_mukerrer_temizle(uygula: bool = False):
    """Aktif mükerrer ödeme planlarını teke indirir (en eski kalır, fazlalar 'iptal').
    Ödenmiş kopyaların PARA ayağı da geri alınır (geri alınabilir, hiçbir şey silinmez):
      - kasa_hareketleri: kaynak_id=plan_id (idempotency) ile iptal (nakit ödenmişse).
      - kart_hareketleri: kart talimatı (kaynak_tablo=sabit_giderler) için eşleşen fazla HARCAMA iptal.
    uygula=False → önizleme. Mükerrer kalmazsa UNIQUE index kurulur."""
    rapor = {"onizleme": (not uygula), "gruplar": [], "iptal_plan": [],
             "iptal_kart_hareketi": [], "iptal_kasa": [], "index_kuruldu": False}
    with db() as (conn, cur):
        cur.execute("""
            SELECT kaynak_tablo, kaynak_id, referans_ay
            FROM odeme_plani
            WHERE durum <> 'iptal' AND kaynak_id IS NOT NULL AND referans_ay IS NOT NULL
            GROUP BY kaynak_tablo, kaynak_id, referans_ay
            HAVING COUNT(*) > 1
        """)
        gruplar = [dict(r) for r in (cur.fetchall() or [])]
        for g in gruplar:
            cur.execute("""
                SELECT id::text, durum, odenecek_tutar, kart_id, aciklama
                FROM odeme_plani
                WHERE durum <> 'iptal'
                  AND kaynak_tablo IS NOT DISTINCT FROM %s
                  AND kaynak_id = %s AND referans_ay = %s
                ORDER BY olusturma ASC, id ASC
            """, (g["kaynak_tablo"], g["kaynak_id"], g["referans_ay"]))
            rows = [dict(r) for r in (cur.fetchall() or [])]
            tutulan, fazlalar = rows[0], rows[1:]
            rapor["gruplar"].append({
                "kaynak_tablo": g["kaynak_tablo"], "referans_ay": str(g["referans_ay"]),
                "toplam": len(rows), "tutulan_plan": tutulan["id"],
                "iptal_edilecek": [f["id"] for f in fazlalar],
                "aciklama": str(tutulan.get("aciklama") or "")[:50],
            })
            if not uygula:
                continue
            for f in fazlalar:
                cur.execute("""UPDATE odeme_plani SET durum='iptal',
                    aciklama = COALESCE(aciklama,'') || ' · iptal: mukerrer temizlik'
                    WHERE id=%s""", (f["id"],))
                rapor["iptal_plan"].append({"id": f["id"],
                    "tutar": float(f["odenecek_tutar"] or 0), "onceki_durum": f["durum"]})
                if f["durum"] == "odendi":
                    cur.execute("""UPDATE kasa_hareketleri SET durum='iptal'
                        WHERE kaynak_id=%s AND durum='aktif' RETURNING id::text, tutar""", (f["id"],))
                    for kr in (cur.fetchall() or []):
                        rapor["iptal_kasa"].append({"id": kr["id"], "tutar": float(kr["tutar"])})
                        # FIX MN2 (2026-07-05): mükerrer temizlikte UPDATE durum='iptal' DOĞRU
                        # (fazla kopyayı geri al; ters kayıt eklemek kasayı yanlış sıfırlardı) —
                        # ancak append-only iz eksikti. Audit kaydıyla denetim izi tamamlanır.
                        audit(cur, 'kasa_hareketleri', kr["id"], 'MUKERRER_IPTAL',
                              eski={'durum': 'aktif', 'tutar': float(kr["tutar"])},
                              yeni={'durum': 'iptal', 'neden': 'mukerrer plan temizligi'})
                    if g["kaynak_tablo"] == "sabit_giderler":
                        cur.execute("""
                            UPDATE kart_hareketleri SET durum='iptal'
                            WHERE id = (
                                SELECT id FROM kart_hareketleri
                                WHERE kaynak_tablo='sabit_giderler' AND kaynak_id=%s
                                  AND islem_turu='HARCAMA' AND durum='aktif'
                                  AND ABS(tutar - %s) < 0.01
                                ORDER BY tarih DESC, id DESC LIMIT 1
                            )
                            RETURNING id::text, tutar
                        """, (g["kaynak_id"], float(f["odenecek_tutar"] or 0)))
                        for kr in (cur.fetchall() or []):
                            rapor["iptal_kart_hareketi"].append({"id": kr["id"], "tutar": float(kr["tutar"])})
        if uygula:
            cur.execute("""
                SELECT COUNT(*) AS n FROM (
                  SELECT 1 FROM odeme_plani
                  WHERE durum <> 'iptal' AND kaynak_id IS NOT NULL AND referans_ay IS NOT NULL
                  GROUP BY kaynak_tablo, kaynak_id, referans_ay HAVING COUNT(*)>1
                ) d
            """)
            if int(cur.fetchone()["n"]) == 0:
                cur.execute("""
                    CREATE UNIQUE INDEX IF NOT EXISTS uq_odeme_plani_kaynak_ay_aktif
                    ON odeme_plani (kaynak_tablo, kaynak_id, referans_ay)
                    WHERE durum <> 'iptal' AND kaynak_id IS NOT NULL AND referans_ay IS NOT NULL
                """)
                rapor["index_kuruldu"] = True
    return rapor


@app.post("/api/odeme-plani/personel-arrears-tarih-duzelt")
def personel_arrears_tarih_duzelt(uygula: bool = False):
    """Mevcut bekleyen PERSONEL maaş planlarının ödeme tarihini arrears (geçmiş ay)
    modeline çeker: tarih = çalışma ayı (referans_ay) + 1 ay'ın 1'i. Ödenmişlere
    dokunmaz. İdempotent (zaten ileri tarihliyse atlar). uygula=False önizleme."""
    with db() as (conn, cur):
        where = """
            WHERE kaynak_tablo='personel' AND durum IN ('bekliyor','onay_bekliyor')
              AND referans_ay IS NOT NULL
              AND tarih < (referans_ay + INTERVAL '1 month')::date
        """
        cur.execute(
            "SELECT id::text, referans_ay::text AS calisma_ay, tarih::text AS eski_tarih, "
            "((referans_ay + INTERVAL '1 month')::date)::text AS yeni_tarih, aciklama "
            "FROM odeme_plani " + where + " ORDER BY referans_ay"
        )
        adaylar = [dict(r) for r in (cur.fetchall() or [])]
        if uygula:
            cur.execute(
                "UPDATE odeme_plani SET tarih = (referans_ay + INTERVAL '1 month')::date " + where
            )
        return {"onizleme": (not uygula), "adet": len(adaylar), "kayitlar": adaylar}


# ── FATURA ÖDEMESİ ────────────────────────────────────────────

class FaturaOdemeModel(BaseModel):
    sabit_gider_id: str       # Hangi değişken gider ödeniyor
    tutar: float              # Fatura tutarı
    tarih: date               # Ödeme tarihi
    odeme_yontemi: str = 'nakit'
    kart_id: Optional[str] = None
    aciklama: Optional[str] = None

@app.post("/api/fatura-ode")
def fatura_ode(body: FaturaOdemeModel):
    """
    Değişken sabit gider (elektrik, su vb.) fatura ödemesi.
    Kasaya FATURA_ODEMESI olarak yazılır, kaynak sabit_giderler tablosuna bağlanır.
    """
    try:
        _tutar_chk = float(body.tutar)
    except (TypeError, ValueError):
        raise HTTPException(400, "Geçerli bir tutar girin")
    if _tutar_chk <= 0:
        raise HTTPException(400, "Tutar 0'dan büyük olmalı")
    with db() as (conn, cur):
        # FIX SEC2 (2026-07-06): gider-bazlı yarış kilidi — "bu ay ödendi mi" kontrolü ile kasa
        # yazımı arasında kilit yoktu; eşzamanlı iki istek ikisi de 'ödenmemiş' görüp aynı faturayı
        # iki kez kasadan düşürebiliyordu. Kilit istekleri sıraya sokar; ikincisi dedup'a takılır.
        cur.execute("SELECT pg_advisory_xact_lock(hashtext(%s))", (f"fatura-ode:{body.sabit_gider_id}",))
        # Sabit gideri kontrol et
        cur.execute("SELECT * FROM sabit_giderler WHERE id=%s AND aktif=TRUE FOR UPDATE", (body.sabit_gider_id,))
        gider = cur.fetchone()
        if not gider:
            raise HTTPException(404, "Gider bulunamadı")
        _tip = (gider.get('tip') or 'sabit')
        if _tip not in ('degisken', 'sabit'):
            raise HTTPException(400, "Geçersiz gider tipi")
        # Değişken (elektrik/su) → FATURA_ODEMESI · Sabit (kira vb.) → SABIT_GIDER
        _kasa_turu = 'FATURA_ODEMESI' if _tip == 'degisken' else 'SABIT_GIDER'

        # Bu ay zaten ödendi mi? (a) manuel kasa ödemesi (b) ödeme planından ödenmiş.
        # ÇİFT ÖDEME KAPISI: plan yolu kasaya kaynak_id=plan_id yazar; bu yüzden manuel
        # dedup'ı plan ödemesini de görmeli — yoksa aynı kira iki yoldan ödenip kasadan
        # iki kez düşer.
        cur.execute("""
            SELECT 1 FROM kasa_hareketleri
            WHERE kaynak_id=%s AND kaynak_tablo='sabit_giderler'
            AND islem_turu=%s AND kasa_etkisi=true AND durum='aktif'
            AND EXTRACT(YEAR FROM tarih) = EXTRACT(YEAR FROM %s::date)
            AND EXTRACT(MONTH FROM tarih) = EXTRACT(MONTH FROM %s::date)
        """, (body.sabit_gider_id, _kasa_turu, str(body.tarih), str(body.tarih)))
        if cur.fetchone():
            raise HTTPException(400, "Bu ay için zaten ödeme yapılmış")
        # Bu giderin bu ayki ödeme planı zaten ödenmiş mi? (kart veya nakit — her yöntem)
        cur.execute("""
            SELECT 1 FROM odeme_plani
            WHERE kaynak_tablo='sabit_giderler' AND kaynak_id=%s
              AND durum='odendi'
              AND referans_ay = DATE_TRUNC('month', %s::date)
        """, (body.sabit_gider_id, str(body.tarih)))
        if cur.fetchone():
            raise HTTPException(400, "Bu ay için ödeme planından zaten ödendi — tekrar ödeme yapılamaz")

        aciklama = body.aciklama or f"Fatura: {gider['gider_adi']}"

        if body.odeme_yontemi == 'kart':
            if not body.kart_id:
                raise HTTPException(400, "Kart seçimi zorunlu")
            cur.execute("SELECT * FROM kartlar WHERE id=%s AND aktif=TRUE FOR UPDATE", (body.kart_id,))
            kart = cur.fetchone()
            if not kart:
                raise HTTPException(404, "Kart bulunamadı")
            # Mevcut kart borcunu hesapla — limit kontrolü, FOR UPDATE: eş zamanlı limit aşımını önler
            borc = kart_borc(cur, body.kart_id)
            kalan_limit = _kanonik_kalan_limit(body.kart_id, float(kart['limit_tutar']) - borc)
            if kalan_limit < body.tutar:
                raise HTTPException(400, f"Kart limiti yetersiz. Kalan: {kalan_limit:,.0f} ₺")
            # Karta HARCAMA yaz — kaynak_tablo fatura_giderleri
            fid = str(uuid.uuid4())   # kart yolunda fid = kart_hareketleri kaydı
            cur.execute("""
                INSERT INTO kart_hareketleri
                    (id, kart_id, tarih, islem_turu, tutar, taksit_sayisi, aciklama, kaynak_id, kaynak_tablo)
                VALUES (%s, %s, %s, 'HARCAMA', %s, 1, %s, %s, 'fatura_giderleri')
            """, (fid, body.kart_id, str(body.tarih), body.tutar, aciklama, body.sabit_gider_id))
            audit(cur, 'kart_hareketleri', fid, 'FATURA_KART')
            kart_plan_guncelle_tx(cur)
        else:
            # Kasaya yaz
            fid = str(uuid.uuid4())   # nakit yolunda fid = kasa_hareketleri kaydı
            insert_kasa_hareketi(cur, str(body.tarih), _kasa_turu, -abs(body.tutar),
                aciklama, 'sabit_giderler', body.sabit_gider_id,
                ref_id=fid, ref_type=_kasa_turu)

        # Bağlı ödeme planını ÖDENDİ yap — yoksa kasa kaydı oluşsa da plan 'bekliyor'
        # kalıp CFO panelde "vadesi geçmiş borç" + ödenmedi olarak görünür (çift gösterim).
        # KÖK DÜZELTME: "aynı ay" filtresi yerine ay'dan bağımsız EN ESKİ açık plan
        # kapatılır (geç/erken ödemede de doğru çalışır). Üstteki dedup guard'ları aynı
        # ay için ikinci ödemeyi zaten engellediğinden tek plan kapanır.
        cur.execute("""
            UPDATE odeme_plani SET durum='odendi', odeme_tarihi=%s
            WHERE id = (
                SELECT id FROM odeme_plani
                WHERE kaynak_tablo='sabit_giderler' AND kaynak_id=%s
                AND durum IN ('bekliyor','onay_bekliyor')
                ORDER BY tarih ASC
                LIMIT 1
            )
        """, (str(body.tarih), body.sabit_gider_id))
        # Bekleyen onay kuyruğu kaydını kapat (en eski) — onay merkezden düşsün.
        cur.execute("""
            UPDATE onay_kuyrugu SET durum='onaylandi', onay_tarihi=NOW()
            WHERE id = (
                SELECT id FROM onay_kuyrugu
                WHERE kaynak_tablo='sabit_giderler' AND kaynak_id=%s AND durum='bekliyor'
                ORDER BY tarih ASC
                LIMIT 1
            )
        """, (body.sabit_gider_id,))

        audit(cur, 'sabit_giderler', body.sabit_gider_id, 'FATURA_ODENDI',
              yeni={'tutar': body.tutar, 'tarih': str(body.tarih)})
    return {"success": True, "id": fid}


class FaturaVadeModel(BaseModel):
    sabit_gider_id: str
    tutar: float
    tarih: Optional[date] = None  # vade; verilmezse bu ayın odeme_gunu


@app.post("/api/fatura-vadeye-yaz")
def fatura_vadeye_yaz(body: FaturaVadeModel):
    """DEĞİŞKEN GİDER KURGUSU (kullanıcı, 2026-07-04): hatırlatma tutarı SORAR; kullanıcı
    bu ayın fatura tutarını girip 'henüz ödenmedi' derse fatura o ayın ÖDEME PLANINA
    yazılır (bekliyor) → CFO vade listesinde GERÇEK tutarla izlenir, ödenince kasa izi
    ile kapanır ("kasa izi = tek gerçek"). Kasa ETKİLENMEZ (para henüz çıkmadı).
    İdempotent: aynı gider + ay için bekleyen plan varsa tutarı/vadesi güncellenir."""
    try:
        _t = float(body.tutar)
    except (TypeError, ValueError):
        raise HTTPException(400, "Geçerli bir tutar girin")
    if _t <= 0:
        raise HTTPException(400, "Tutar 0'dan büyük olmalı")
    bugun = bugun_tr()
    with db() as (conn, cur):
        cur.execute("SELECT * FROM sabit_giderler WHERE id=%s AND aktif=TRUE", (body.sabit_gider_id,))
        gider = cur.fetchone()
        if not gider:
            raise HTTPException(404, "Gider bulunamadı")
        _tip = (gider.get('tip') or 'sabit')
        _kasa_turu = 'FATURA_ODEMESI' if _tip == 'degisken' else 'SABIT_GIDER'

        # Bu ay zaten ödendiyse vadeye yazmak anlamsız (çift gösterim/çift ödeme kapısı)
        cur.execute("""
            SELECT 1 FROM kasa_hareketleri
            WHERE kaynak_id=%s AND kaynak_tablo='sabit_giderler'
            AND islem_turu=%s AND kasa_etkisi=true AND durum='aktif'
            AND EXTRACT(YEAR FROM tarih)=%s AND EXTRACT(MONTH FROM tarih)=%s
        """, (body.sabit_gider_id, _kasa_turu, bugun.year, bugun.month))
        if cur.fetchone():
            raise HTTPException(400, "Bu ay için zaten ödeme yapılmış — vadeye yazılamaz")
        cur.execute("""
            SELECT 1 FROM odeme_plani
            WHERE kaynak_tablo='sabit_giderler' AND kaynak_id=%s
              AND durum='odendi' AND referans_ay = DATE_TRUNC('month', %s::date)
        """, (body.sabit_gider_id, str(bugun)))
        if cur.fetchone():
            raise HTTPException(400, "Bu ayın faturası zaten ödendi — vadeye yazılamaz")

        # Vade: verilen tarih ya da bu ayın odeme_gunu (ay sonuna kırpılır)
        if body.tarih:
            vade = body.tarih
        else:
            import calendar as _cal
            g_gun = int(gider.get('odeme_gunu') or bugun.day)
            vade = date(bugun.year, bugun.month,
                        min(g_gun, _cal.monthrange(bugun.year, bugun.month)[1]))
        aciklama = f"Fatura: {gider['gider_adi']} — {_TR_AYLAR[bugun.month]} {bugun.year}"

        # İdempotent yazım: bekleyen plan varsa güncelle, yoksa oluştur
        cur.execute("""
            UPDATE odeme_plani
            SET tarih=%s, odenecek_tutar=%s, asgari_tutar=%s, aciklama=%s
            WHERE kaynak_tablo='sabit_giderler' AND kaynak_id=%s
              AND durum IN ('bekliyor','onay_bekliyor')
              AND referans_ay = DATE_TRUNC('month', %s::date)
            RETURNING id
        """, (str(vade), _t, _t, aciklama, body.sabit_gider_id, str(bugun)))
        row = cur.fetchone()
        if not row:
            pid = str(uuid.uuid4())
            cur.execute("""
                INSERT INTO odeme_plani
                    (id, kart_id, tarih, referans_ay, odenecek_tutar, asgari_tutar,
                     aciklama, durum, kaynak_tablo, kaynak_id)
                SELECT %s, NULL, %s, DATE_TRUNC('month', %s::date), %s, %s,
                       %s, 'bekliyor', 'sabit_giderler', %s
                WHERE NOT EXISTS (
                    SELECT 1 FROM odeme_plani
                    WHERE kaynak_tablo='sabit_giderler' AND kaynak_id=%s
                      AND durum != 'iptal'
                      AND referans_ay = DATE_TRUNC('month', %s::date)
                )
                RETURNING id
            """, (pid, str(vade), str(bugun), _t, _t, aciklama,
                  body.sabit_gider_id, body.sabit_gider_id, str(bugun)))
            row = cur.fetchone()
        if not row:
            raise HTTPException(400, "Bu ay için plan yazılamadı (mevcut kayıt engelledi)")
        audit(cur, 'odeme_plani', str(row['id']), 'FATURA_VADEYE_YAZ',
              yeni={'tutar': _t, 'vade': str(vade), 'gider': gider['gider_adi']})
        uyari_cache_clear()
    return {"success": True, "plan_id": str(row['id']), "vade": str(vade), "tutar": _t}


@app.get("/api/fatura-gecmis/{gider_id}")
def fatura_gecmis(gider_id: str):
    """Bir değişken giderin geçmiş fatura ödemelerini döner."""
    with db() as (conn, cur):
        cur.execute("""
            SELECT tarih, ABS(tutar) as tutar, aciklama, 'nakit' as yontem
            FROM kasa_hareketleri
            WHERE kaynak_id=%s AND kaynak_tablo='sabit_giderler'
            AND islem_turu IN ('FATURA_ODEMESI','SABIT_GIDER') AND kasa_etkisi=true AND durum='aktif'
            ORDER BY tarih DESC LIMIT 12
        """, (gider_id,))
        nakit = [dict(r) for r in cur.fetchall()]

        cur.execute("""
            SELECT kh.tarih, kh.tutar, kh.aciklama, 'kart' as yontem, k.banka, k.kart_adi
            FROM kart_hareketleri kh
            JOIN kartlar k ON k.id = kh.kart_id
            WHERE kh.kaynak_id=%s AND kh.kaynak_tablo='fatura_giderleri'
            AND kh.islem_turu='HARCAMA' AND kh.durum='aktif'
            ORDER BY kh.tarih DESC LIMIT 12
        """, (gider_id,))
        kart = [dict(r) for r in cur.fetchall()]

        gecmis = nakit + kart
        gecmis.sort(key=lambda x: str(x['tarih']), reverse=True)
        return gecmis

@app.get("/api/sabit-giderler/{gid}/gecmis")
def sabit_gider_gecmis(gid: str):
    """Sabit giderin ödeme geçmişi — kasa_hareketleri + odeme_plani."""
    with db() as (conn, cur):
        cur.execute("SELECT * FROM sabit_giderler WHERE id=%s", (gid,))
        gider = cur.fetchone()
        if not gider: raise HTTPException(404, "Gider bulunamadı")

        # Ödenen — kasa_hareketleri
        cur.execute("""
            SELECT tarih, ABS(tutar) as tutar, aciklama, islem_turu
            FROM kasa_hareketleri
            WHERE kaynak_id = %s AND kaynak_tablo = 'sabit_giderler'
            AND kasa_etkisi = true AND durum = 'aktif' AND tutar < 0
            ORDER BY tarih DESC
        """, (gid,))
        odenenler = [{"tarih": str(r['tarih']), "tutar": float(r['tutar']),
                      "aciklama": r['aciklama'] or '', "durum": "odendi"} for r in cur.fetchall()]

        # Bekleyen — odeme_plani
        cur.execute("""
            SELECT tarih, odenecek_tutar, durum, aciklama
            FROM odeme_plani
            WHERE kaynak_id = %s AND kaynak_tablo = 'sabit_giderler'
            AND durum IN ('bekliyor','onay_bekliyor')
            ORDER BY tarih ASC
        """, (gid,))
        bekleyenler = [{"tarih": str(r['tarih']), "tutar": float(r['odenecek_tutar']),
                        "aciklama": r['aciklama'] or '', "durum": r['durum']} for r in cur.fetchall()]

        toplam_odenen = sum(r['tutar'] for r in odenenler)
        son_tutar = odenenler[0]['tutar'] if odenenler else None
        return {
            "gider": {"id": str(gider['id']), "gider_adi": gider['gider_adi'],
                      "kategori": gider['kategori'], "tutar": float(gider['tutar'])},
            "ozet": {"toplam_odenen": round(toplam_odenen, 2),
                     "odeme_adedi": len(odenenler)},
            # son_tutar: son ödeme tutarı — fatura modalı otomatik öneri için
            "son_tutar": round(son_tutar, 2) if son_tutar is not None else None,
            "son_tarih": odenenler[0]['tarih'] if odenenler else None,
            "odenenler": odenenler,
            "bekleyenler": bekleyenler,
        }

@app.get("/api/anlik-gider/gecmis")
def anlik_gider_gecmis(kategori: str = None, limit: int = 100):
    """Anlık gider geçmişi — isteğe bağlı kategori filtresi."""
    with db() as (conn, cur):
        if kategori:
            cur.execute("""
                SELECT tarih, ABS(tutar) as tutar, aciklama, kategori, odeme_yontemi
                FROM kasa_hareketleri
                WHERE islem_turu = 'ANLIK_GIDER' AND durum = 'aktif' AND tutar < 0
                AND aciklama ILIKE %s
                ORDER BY tarih DESC LIMIT %s
            """, (f"%{kategori}%", limit))
        else:
            cur.execute("""
                SELECT tarih, ABS(tutar) as tutar, aciklama, islem_turu as kategori, odeme_yontemi
                FROM kasa_hareketleri
                WHERE islem_turu = 'ANLIK_GIDER' AND durum = 'aktif' AND tutar < 0
                ORDER BY tarih DESC LIMIT %s
            """, (limit,))
        satirlar = [{"tarih": str(r['tarih']), "tutar": float(r['tutar']),
                     "aciklama": r['aciklama'] or '', 
                     "odeme_yontemi": r.get('odeme_yontemi', 'nakit')} for r in cur.fetchall()]

        # Kategori özeti
        cur.execute("""
            SELECT
                SPLIT_PART(aciklama, ' - ', 1) as kat,
                COUNT(*) as adet,
                SUM(ABS(tutar)) as toplam
            FROM kasa_hareketleri
            WHERE islem_turu = 'ANLIK_GIDER' AND durum = 'aktif' AND tutar < 0
            GROUP BY kat ORDER BY toplam DESC LIMIT 10
        """)
        kategoriler = [{"kategori": r['kat'] or 'Diğer',
                        "adet": int(r['adet']), "toplam": float(r['toplam'])} for r in cur.fetchall()]

        return {"satirlar": satirlar, "kategoriler": kategoriler,
                "toplam": sum(r['tutar'] for r in satirlar)}


@app.get("/api/bilgi-teslim-kayitlari")
def bilgi_teslim_kayitlari(sube_id: Optional[str] = None, gun: int = 30, limit: int = 300):
    """
    Şubelerden merkeze iletilen bilgi/not kayıtları.
    """
    gun_sayi = max(1, min(365, int(gun)))
    lim = max(1, min(1000, int(limit)))
    with db() as (conn, cur):
        qp: List[Any] = [gun_sayi]
        q = """
            SELECT n.id, n.sube_id, s.ad AS sube_adi, n.metin,
                   n.personel_id, n.personel_ad, n.olusturma
            FROM sube_merkez_not n
            LEFT JOIN subeler s ON s.id = n.sube_id
            WHERE n.olusturma >= (NOW() - (%s * INTERVAL '1 day'))
        """
        if sube_id:
            q += " AND n.sube_id=%s"
            qp.append(sube_id)
        q += " ORDER BY n.olusturma DESC LIMIT %s"
        qp.append(lim)
        cur.execute(q, qp)
        satirlar = []
        for r in cur.fetchall():
            d = dict(r)
            if d.get("olusturma"):
                d["olusturma"] = str(d["olusturma"])
            satirlar.append(d)
        return {"gun_sayi": gun_sayi, "sube_id": sube_id, "limit": lim, "satirlar": satirlar}


# ── VADELİ ALIMLAR ─────────────────────────────────────────────
class VadeliAlim(BaseModel):
    aciklama: str
    tutar: float
    vade_tarihi: date
    tedarikci: str          # Zorunlu — kart takibi ve raporlar için
    force: bool = False
    # TEDARIKCI_ACIK_BAKIYE (çoklu açık borç): ayri=yeni satır, ilave=seçilen/hedefe ekle.
    # Tek açık borçta birleştirme artık varsayılan — ayri ile ayrı satır zorlanır.
    tedarikci_karari: Optional[str] = None
    # API/çoklu açık borçta ilave hedefi; tek satırda tedarikci_karari=ilave yeter
    birlestir_vadeli_id: Optional[str] = None


def _vadeli_tedarikci_norm(s: str) -> str:
    return (s or "").strip().lower()


# ─────────────────────────────────────────────────────────────────────────
# 🛡️ İSTEK İZİ — para yazan uçlarda ağ retry'ına karşı idempotency defteri
#
# NEDEN (PARA-011, 2026-09-02): vadeli alım ucunda aynı tedarikçide TEK açık
# borç varsa, ikinci istek "ilave" sayılıp tutarı ÜSTÜNE topluyordu. Yani
# ağ retry'ı — kullanıcı hiçbir şey yapmadan — borcu SESSİZCE İKİYE KATLIYOR.
# 7 günlük "benzer kayıt" uyarısı bu dala hiç girmiyordu.
#
# Çözüm: ödeme ANLAMINI taşıyan alanlardan parmak izi üret, pencere içinde
# aynı parmak gelirse İŞLEME, önceki sonucu geri ver. Kullanıcı gerçekten
# ikinci bir kayıt istiyorsa force=true ile geçebilir — yani kapı kilit değil,
# BİLİNÇLİ GEÇİŞ gerektiren bir eşik.
#
# Pencere neden 180 sn: retry saniyeler içinde olur; 3 dakika arayla girilen
# birebir aynı borç gerçek bir ikinci kayıttır. Yanlış tarafa düşmenin bedeli
# asimetrik — mükerrer borç sessizdir, engellenen kayıt ekranda görünür.
ISTEK_IZI_PENCERE_SN = 180


def _ensure_istek_izi(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS istek_izi (
            parmak     TEXT PRIMARY KEY,
            kapsam     TEXT NOT NULL,
            sonuc      JSONB,
            olusturma  TIMESTAMPTZ NOT NULL DEFAULT NOW()
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS ix_istek_izi_ts ON istek_izi (olusturma DESC)")


def _istek_parmak(kapsam: str, *parcalar) -> str:
    ham = kapsam + "|" + "|".join(
        str(p if p is not None else "").strip().lower() for p in parcalar)
    return hashlib.md5(ham.encode("utf-8")).hexdigest()


def _istek_izi_tazeyse(cur, parmak: str):
    """Pencere içinde aynı parmak varsa ÖNCEKİ sonucu döner; yoksa None."""
    _ensure_istek_izi(cur)
    cur.execute(
        """SELECT sonuc FROM istek_izi
           WHERE parmak=%s AND olusturma >= NOW() - (%s || ' seconds')::interval""",
        (parmak, ISTEK_IZI_PENCERE_SN),
    )
    r = cur.fetchone()
    if not r:
        return None
    s = r.get("sonuc")
    if isinstance(s, dict):
        return s
    try:
        return json.loads(s) if s else {}
    except Exception:
        return {}


def _istek_izi_yaz(cur, parmak: str, kapsam: str, sonuc: dict):
    _ensure_istek_izi(cur)
    cur.execute(
        """INSERT INTO istek_izi (parmak, kapsam, sonuc, olusturma)
           VALUES (%s, %s, %s::jsonb, NOW())
           ON CONFLICT (parmak) DO UPDATE SET sonuc=EXCLUDED.sonuc, olusturma=NOW()""",
        (parmak, kapsam, json.dumps(sonuc, default=str)),
    )


def _vadeli_bekleyen_ayni_tedarikci(cur, tedarikci: str):
    t = _vadeli_tedarikci_norm(tedarikci)
    if not t:
        return []
    cur.execute(
        """
        SELECT id, aciklama, tutar, vade_tarihi, tedarikci
        FROM vadeli_alimlar
        WHERE durum = 'bekliyor'
          AND LOWER(TRIM(COALESCE(tedarikci, ''))) = %s
        ORDER BY vade_tarihi
        """,
        (t,),
    )
    return cur.fetchall()


def _vadeli_aktif_plan_bul(cur, vadeli_id: str):
    cur.execute(
        """
        SELECT id, odenecek_tutar
        FROM odeme_plani
        WHERE kaynak_tablo='vadeli_alimlar' AND kaynak_id=%s
        AND durum IN ('bekliyor','onay_bekliyor')
        ORDER BY olusturma DESC, tarih DESC, id DESC
        LIMIT 1
        """,
        (vadeli_id,),
    )
    return cur.fetchone()


def _vadeli_diger_aktif_planlari_iptal(cur, vadeli_id: str, aktif_plan_id: str):
    cur.execute(
        """
        UPDATE odeme_plani SET durum='iptal'
        WHERE kaynak_tablo='vadeli_alimlar' AND kaynak_id=%s
        AND id<>%s
        AND durum IN ('bekliyor','onay_bekliyor')
        RETURNING id
        """,
        (vadeli_id, aktif_plan_id),
    )
    iptal_edilenler = cur.fetchall()
    iptal_ids = [row["id"] for row in iptal_edilenler]
    if iptal_ids:
        cur.execute(
            """
            UPDATE onay_kuyrugu SET durum='reddedildi', onay_tarihi=NOW()
            WHERE durum='bekliyor'
            AND kaynak_tablo='odeme_plani'
            AND kaynak_id = ANY(%s)
            """,
            (iptal_ids,),
        )


def _vadeli_borcla_birlestir(cur, hedef_id: str, v: VadeliAlim) -> dict:
    cur.execute(
        "SELECT * FROM vadeli_alimlar WHERE id=%s FOR UPDATE",
        (hedef_id,),
    )
    eski = cur.fetchone()
    if not eski:
        raise HTTPException(404, "Birleştirilecek vadeli kaydı bulunamadı")
    if eski["durum"] != "bekliyor":
        raise HTTPException(
            400,
            "Sadece bekleyen vadeli borcuna eklenebilir — ödenmiş veya iptal satıra eklenemez.",
        )
    if _vadeli_tedarikci_norm(eski.get("tedarikci")) != _vadeli_tedarikci_norm(v.tedarikci):
        raise HTTPException(400, "Tedarikçi eşleşmiyor — birleştirme yapılamaz.")
    ek = float(v.tutar)
    if ek <= 0:
        raise HTTPException(400, "Eklenecek tutar sıfırdan büyük olmalı")
    yeni_toplam = float(eski["tutar"]) + ek
    a_eski = (eski.get("aciklama") or "").strip()
    a_yeni = (v.aciklama or "").strip()
    yeni_aciklama = (
        f"{a_eski} + {a_yeni}" if a_eski and a_yeni else (a_yeni or a_eski or "Vadeli alım")
    )
    ted = (v.tedarikci or "").strip() or (eski.get("tedarikci") or "")
    cur.execute(
        """UPDATE vadeli_alimlar SET tutar=%s, vade_tarihi=%s, aciklama=%s, tedarikci=%s WHERE id=%s""",
        (yeni_toplam, v.vade_tarihi, yeni_aciklama, ted, hedef_id),
    )
    prow = _vadeli_aktif_plan_bul(cur, hedef_id)
    if prow:
        pid = prow["id"]
        cur.execute(
            """
            UPDATE odeme_plani SET
                tarih=%s,
                referans_ay=DATE_TRUNC('month', %s::date),
                odenecek_tutar=%s,
                asgari_tutar=%s,
                aciklama=%s
            WHERE id=%s
            """,
            (
                v.vade_tarihi,
                str(v.vade_tarihi),
                yeni_toplam,
                yeni_toplam,
                f"Vadeli Alım: {yeni_aciklama}",
                pid,
            ),
        )
    else:
        pid = str(uuid.uuid4())
        cur.execute(
            """
            INSERT INTO odeme_plani
                (id, kart_id, tarih, referans_ay, odenecek_tutar, asgari_tutar, aciklama, durum, kaynak_tablo, kaynak_id)
            VALUES (%s, NULL, %s, DATE_TRUNC('month', %s::date), %s, %s, %s, 'bekliyor', 'vadeli_alimlar', %s)
            """,
            (
                pid,
                v.vade_tarihi,
                str(v.vade_tarihi),
                yeni_toplam,
                yeni_toplam,
                f"Vadeli Alım: {yeni_aciklama}",
                hedef_id,
            ),
        )
    _vadeli_diger_aktif_planlari_iptal(cur, hedef_id, pid)
    cur.execute(
        """
        UPDATE onay_kuyrugu SET tutar=%s, tarih=%s
        WHERE durum='bekliyor' AND islem_turu='VADELI_ODEME'
        AND kaynak_tablo='vadeli_alimlar' AND kaynak_id=%s
        """,
        (yeni_toplam, v.vade_tarihi, hedef_id),
    )
    cur.execute(
        """
        UPDATE onay_kuyrugu SET tutar=%s, tarih=%s
        WHERE durum='bekliyor' AND islem_turu='ODEME_PLANI'
        AND kaynak_tablo='odeme_plani'
        AND kaynak_id IN (
            SELECT id FROM odeme_plani
            WHERE kaynak_tablo='vadeli_alimlar' AND kaynak_id=%s
            AND durum IN ('bekliyor','onay_bekliyor')
        )
        """,
        (yeni_toplam, v.vade_tarihi, hedef_id),
    )
    audit(
        cur,
        "vadeli_alimlar",
        hedef_id,
        "BORC_EKLE",
        eski=dict(eski),
        yeni={
            "tutar": yeni_toplam,
            "vade_tarihi": str(v.vade_tarihi),
            "aciklama": yeni_aciklama,
        },
    )
    return {
        "id": hedef_id,
        "success": True,
        "birlestirildi": True,
        "onceki_tutar": float(eski["tutar"]),
        "eklenen": ek,
        "yeni_toplam": yeni_toplam,
    }


@app.get("/api/vadeli-alimlar")
def vadeli_listele(durum: str = "bekliyor", gun: int = 30):
    d = (durum or "bekliyor").strip().lower()
    g = max(1, min(int(gun or 30), 365))
    with db() as (conn, cur):
        if d == "bekliyor":
            cur.execute(
                """
                SELECT *, (vade_tarihi - CURRENT_DATE) as gun_kaldi
                FROM vadeli_alimlar
                WHERE durum='bekliyor'
                ORDER BY vade_tarihi
                """
            )
        elif d == "odendi":
            cur.execute(
                """
                SELECT *, (vade_tarihi - CURRENT_DATE) as gun_kaldi
                FROM vadeli_alimlar
                WHERE durum='odendi'
                  AND odeme_tarihi >= CURRENT_DATE - (%s || ' days')::interval
                ORDER BY odeme_tarihi DESC NULLS LAST, vade_tarihi DESC
                """,
                (g,),
            )
        elif d == "hepsi":
            cur.execute(
                """
                SELECT *, (vade_tarihi - CURRENT_DATE) as gun_kaldi
                FROM vadeli_alimlar
                ORDER BY
                    CASE WHEN durum='bekliyor' THEN 0 WHEN durum='odendi' THEN 1 ELSE 2 END,
                    COALESCE(odeme_tarihi, vade_tarihi) DESC
                """
            )
        else:
            raise HTTPException(400, "durum: bekliyor | odendi | hepsi")
        return [dict(r) for r in cur.fetchall()]


@app.get("/api/vadeli-alimlar/gecmis")
def vadeli_gecmis(limit: int = 120, ay: str = None):
    lim = max(1, min(int(limit or 120), 500))
    # 🟡 P1 (2026-08-13, EVV-ODE): ay filtresi istemcideydi ama pencere "son N
    # kayıt" olduğu için eski bir ay pencere dışında kalınca ekran sessizce
    # başka ayın kayıtlarına düşüyordu. Ay artık sunucuda süzülür.
    import re as _re_ay
    ay_v = (ay or "").strip()
    ay_cond = ""
    ay_params: list = []
    if ay_v and _re_ay.match(r"^\d{4}-\d{2}$", ay_v):
        ay_cond = " WHERE to_char(q.tarih, 'YYYY-MM') = %s"
        ay_params = [ay_v]
    with db() as (conn, cur):
        cur.execute(
            """
            SELECT *
            FROM (
                SELECT
                    kh.tarih,
                    ABS(kh.tutar) AS tutar,
                    'nakit'::text AS odeme_yontemi,
                    kh.aciklama,
                    va.id AS vadeli_id,
                    va.aciklama AS vadeli_aciklama,
                    va.tedarikci
                FROM kasa_hareketleri kh
                LEFT JOIN vadeli_alimlar va ON va.id = kh.kaynak_id
                WHERE kh.kaynak_tablo='vadeli_alimlar'
                  AND kh.islem_turu='VADELI_ODEME'
                  AND kh.kasa_etkisi=TRUE
                  AND kh.durum='aktif'

                UNION ALL

                SELECT
                    kht.tarih,
                    kht.tutar,
                    'kart'::text AS odeme_yontemi,
                    kht.aciklama,
                    va.id AS vadeli_id,
                    va.aciklama AS vadeli_aciklama,
                    va.tedarikci
                FROM kart_hareketleri kht
                LEFT JOIN vadeli_alimlar va ON va.id = kht.kaynak_id
                WHERE kht.kaynak_tablo='vadeli_alimlar'
                  AND kht.islem_turu='HARCAMA'
                  AND kht.durum='aktif'
            ) q
            """ + ay_cond + """
            ORDER BY q.tarih DESC
            LIMIT %s
            """,
            ay_params + [lim],
        )
        satirlar = [dict(r) for r in cur.fetchall()]
        toplam = sum(float(r.get("tutar") or 0) for r in satirlar)
        return {"satirlar": satirlar, "ozet": {"adet": len(satirlar), "toplam": toplam}}

@app.post("/api/vadeli-alimlar")
def vadeli_ekle(v: VadeliAlim):
    try:
        if float(v.tutar) <= 0:
            raise HTTPException(400, "Tutar 0'dan büyük olmalı")
    except (TypeError, ValueError):
        raise HTTPException(400, "Geçerli bir tutar girin")
    if not (v.tedarikci or "").strip():
        raise HTTPException(400, "Tedarikçi zorunlu")
    with db() as (conn, cur):
        birlestir = (v.birlestir_vadeli_id or "").strip()
        karar = (v.tedarikci_karari or "").strip().lower()

        # 🛡️ PARA-011: retry koruması. Parmak izi ödeme ANLAMINI taşıyan
        # alanlardan üretilir. Aşağıdaki "tek açık borç varsa üstüne topla"
        # dalı olmasaydı bu guard'a gerek yoktu — ama o dal retry'ı sessizce
        # borç katlamaya çeviriyor, o yüzden kapı ucun EN BAŞINDA duruyor.
        _parmak = _istek_parmak(
            "vadeli_ekle", v.tedarikci, "%.2f" % float(v.tutar),
            v.vade_tarihi, v.aciklama, birlestir, karar)
        if not v.force:
            _onceki = _istek_izi_tazeyse(cur, _parmak)
            if _onceki:
                return {**_onceki, "tekrar": True, "mesaj": (
                    "Aynı istek az önce işlendi — mükerrer borç YAZILMADI. "
                    "Gerçekten ikinci bir kayıt istiyorsanız force=true gönderin.")}

        def _bitir(sonuc):
            """Yazma yapan her dönüş buradan geçer — iz olmadan para yazılmaz."""
            _istek_izi_yaz(cur, _parmak, "vadeli_ekle", sonuc)
            return sonuc

        if birlestir:
            return _bitir(_vadeli_borcla_birlestir(cur, birlestir, v))

        acik = _vadeli_bekleyen_ayni_tedarikci(cur, v.tedarikci)
        if acik and not v.force:
            if karar == "ayri":
                pass
            elif karar == "ilave":
                if len(acik) == 1:
                    return _bitir(_vadeli_borcla_birlestir(cur, acik[0]["id"], v))
                raise HTTPException(
                    400,
                    "Bu tedarikçide birden fazla açık borç var — birlestir_vadeli_id ile hedef satırı gönderin.",
                )
            elif len(acik) == 1:
                # Aynı toptancı/tedarikçide tek bekleyen borç: tutarı üstüne topla.
                # ⚠️ Bu dal ONAY SORMADAN yazar — retry koruması yukarıdaki
                # parmak izi kapısıyla sağlanıyor (PARA-011).
                return _bitir(_vadeli_borcla_birlestir(cur, acik[0]["id"], v))
            else:
                return {
                    "warning": True,
                    "kod": "TEDARIKCI_ACIK_BAKIYE",
                    "mesaj": (
                        "Bu tedarikçi için birden fazla bekleyen vadeli borç var. "
                        "Hangi satıra ekleneceğini seçin veya ayrı satır olarak kaydedin."
                    ),
                    "mevcut_borc": [dict(r) for r in acik],
                }

        if not v.force:
            cur.execute("""
                SELECT id FROM vadeli_alimlar WHERE durum='bekliyor'
                AND vade_tarihi BETWEEN %s::date - INTERVAL '7 days' AND %s::date + INTERVAL '7 days'
                AND ABS(tutar - %s) < 1
            """, (str(v.vade_tarihi), str(v.vade_tarihi), v.tutar))
            benzer = cur.fetchall()
            if benzer:
                return {"warning": True, "mesaj": f"Son 7 günde benzer kayıt var ({len(benzer)} adet). Yine de kaydetmek için force=true gönderin."}
        vid = str(uuid.uuid4())
        cur.execute("""INSERT INTO vadeli_alimlar (id,aciklama,tutar,vade_tarihi,tedarikci)
            VALUES (%s,%s,%s,%s,%s)""",
            (vid, v.aciklama, v.tutar, v.vade_tarihi, v.tedarikci))
        # odeme_plani'na kaynak bağlı plan ekle — simülasyon ve karar motoru görsün
        pid = str(uuid.uuid4())
        cur.execute("""
            INSERT INTO odeme_plani
                (id, kart_id, tarih, referans_ay, odenecek_tutar, asgari_tutar, aciklama, durum, kaynak_tablo, kaynak_id)
            SELECT %s, NULL, %s, DATE_TRUNC('month', %s::date), %s, %s, %s, 'bekliyor', 'vadeli_alimlar', %s
            WHERE NOT EXISTS (
                SELECT 1 FROM odeme_plani
                WHERE kaynak_tablo = 'vadeli_alimlar'
                AND kaynak_id = %s
                AND durum != 'iptal'
            )
        """, (pid, v.vade_tarihi, str(v.vade_tarihi), float(v.tutar), float(v.tutar),
              f"Vadeli Alım: {v.aciklama}", vid, vid))
        audit(cur, 'vadeli_alimlar', vid, 'INSERT')
        _sonuc = {"id": vid, "success": True}
        _bitir(_sonuc)
    return _sonuc

@app.put("/api/vadeli-alimlar/{vid}")
def vadeli_guncelle(vid: str, v: VadeliAlim):
    try:
        if float(v.tutar) <= 0:
            raise HTTPException(400, "Tutar 0'dan büyük olmalı")
    except (TypeError, ValueError):
        raise HTTPException(400, "Geçerli bir tutar girin")
    with db() as (conn, cur):
        cur.execute("SELECT * FROM vadeli_alimlar WHERE id=%s", (vid,))
        eski = cur.fetchone()
        if not eski: raise HTTPException(404)
        cur.execute("""UPDATE vadeli_alimlar SET aciklama=%s,tutar=%s,vade_tarihi=%s,tedarikci=%s WHERE id=%s""",
            (v.aciklama, v.tutar, v.vade_tarihi, v.tedarikci, vid))
        # Bağlı odeme_plani'nı da güncelle — İKİ ADIM (UniqueViolation fix, 2026-07-04):
        # 1) tarih/tutar/açıklama HER ZAMAN güncellenir (fiili vade = tarih alanıdır).
        cur.execute("""
            UPDATE odeme_plani SET
                tarih=%s,
                odenecek_tutar=%s,
                asgari_tutar=%s,
                aciklama=%s
            WHERE kaynak_tablo='vadeli_alimlar' AND kaynak_id=%s
            AND durum IN ('bekliyor','onay_bekliyor')
        """, (v.vade_tarihi, float(v.tutar), float(v.tutar),
              f"Vadeli Alım: {v.aciklama}", vid))
        # 2) referans_ay (ay-gruplama etiketi) SADECE hedef ayda çakışma yoksa taşınır.
        #    Aynı vadeli alımın o ayda başka (örn. kısmi ödemeden ÖDENMİŞ) planı varsa
        #    uq_odeme_plani_kaynak_ay_aktif indeksi patlıyordu — kullanıcı bildirimi.
        cur.execute("""
            UPDATE odeme_plani op SET referans_ay=DATE_TRUNC('month', %s::date)
            WHERE op.kaynak_tablo='vadeli_alimlar' AND op.kaynak_id=%s
            AND op.durum IN ('bekliyor','onay_bekliyor')
            AND NOT EXISTS (
                SELECT 1 FROM odeme_plani q
                WHERE q.kaynak_tablo='vadeli_alimlar' AND q.kaynak_id=%s
                AND q.id <> op.id AND q.durum <> 'iptal'
                AND q.referans_ay = DATE_TRUNC('month', %s::date)
            )
        """, (str(v.vade_tarihi), vid, vid, str(v.vade_tarihi)))
        # Bekleyen onay tutarı planla aynı kalsın (düzenleme sonrası eski tutarla çift/eksik kasa olmasın)
        cur.execute("""
            UPDATE onay_kuyrugu SET tutar=%s, tarih=%s
            WHERE durum='bekliyor' AND islem_turu='VADELI_ODEME'
            AND kaynak_tablo='vadeli_alimlar' AND kaynak_id=%s
        """, (float(v.tutar), v.vade_tarihi, vid))
        cur.execute("""
            UPDATE onay_kuyrugu SET tutar=%s, tarih=%s
            WHERE durum='bekliyor' AND islem_turu='ODEME_PLANI'
            AND kaynak_tablo='odeme_plani'
            AND kaynak_id IN (
                SELECT id FROM odeme_plani
                WHERE kaynak_tablo='vadeli_alimlar' AND kaynak_id=%s
                AND durum IN ('bekliyor','onay_bekliyor')
            )
        """, (float(v.tutar), v.vade_tarihi, vid))
        audit(cur, 'vadeli_alimlar', vid, 'UPDATE', eski=eski)
    return {"success": True}

@app.delete("/api/vadeli-alimlar/{vid}")
def vadeli_sil(vid: str):
    with db() as (conn, cur):
        cur.execute("SELECT * FROM vadeli_alimlar WHERE id=%s AND durum='bekliyor'", (vid,))
        eski = cur.fetchone()
        if not eski: raise HTTPException(404, "Kayıt bulunamadı veya zaten ödenmiş/iptal edilmiş")
        cur.execute("UPDATE vadeli_alimlar SET durum='iptal' WHERE id=%s", (vid,))
        # Bağlı odeme_plani'nı iptal et — simülasyondan çıkar
        cur.execute("""
            UPDATE odeme_plani SET durum='iptal'
            WHERE kaynak_tablo='vadeli_alimlar' AND kaynak_id=%s
            AND durum IN ('bekliyor','onay_bekliyor')
        """, (vid,))
        # Guard: kasa hareketi varsa ters kayıt yaz, yoksa sadece durum değiştir
        cur.execute("""
            SELECT id FROM kasa_hareketleri
            WHERE kaynak_id=%s AND islem_turu='VADELI_ODEME' AND durum='aktif'
        """, (vid,))
        if cur.fetchone():
            iptal_kasa_hareketi(cur, vid, 'vadeli_alimlar', 'VADELI_ODEME', 'VADELI_IPTAL', 'Vadeli alım iptali')
        audit(cur, 'vadeli_alimlar', vid, 'IPTAL', eski=eski)
    return {"success": True}


class VadeliGeriAlModel(BaseModel):
    # 'bekliyor' → ödeme YANLIŞ işaretlendi, borç geri açılır (plan da geri açılır)
    # 'iptal'    → kaydın kendisi hatalı/mükerrerdi, tamamen kapanır (SÜTAŞ çift-kayıt vakası)
    mod: str = 'bekliyor'
    aciklama: Optional[str] = None


@app.get("/api/vadeli-alimlar/{vid}/iz-onizle")
def vadeli_iz_onizle(vid: str):
    """🔎 KURU ÇALIŞTIRMA — bir vadeli kaydının ARKASINDA hangi izler var?

    Sahip kuralı: para değiştiren işlem ÖNCE kuru çalışır ve listesi okunur.
    `odeme-geri-al` üç ize dokunur (kart · kasa · plan) ama hangisinin GERÇEK
    olduğunu ancak bakarak anlarız:

      • kasa izi VARSA  → geri alma parayı kasaya GERİ EKLER. Bu ancak o para
        gerçekten çıkmadıysa doğrudur.
      • kasa izi YOKSA  → kayıt yalnız KÂĞIT ÜSTÜNDE kapatılmıştır (tipik
        örnek: gece self-heal'i bir nakit ödemeyi görüp sözü 'odendi'
        işaretler). Geri almak hiçbir parayı hareket ettirmez.

    SALT OKUR. Hiçbir satır değişmez.
    """
    with db() as (_, cur):
        cur.execute("SELECT id, tedarikci, tutar::float AS tutar, durum, "
                    "vade_tarihi::text AS vade, odeme_tarihi::text AS odeme_tarihi, "
                    "COALESCE(aciklama,'') AS aciklama "
                    "FROM vadeli_alimlar WHERE id=%s", (vid,))
        v = cur.fetchone()
        if not v:
            raise HTTPException(404, "Vadeli alım bulunamadı")
        v = dict(v)
        cur.execute("""SELECT id, tarih::text AS tarih, tutar::float AS tutar,
                              islem_turu, COALESCE(aciklama,'') AS aciklama
                         FROM kart_hareketleri
                        WHERE kaynak_tablo='vadeli_alimlar' AND kaynak_id=%s
                          AND COALESCE(durum,'aktif')='aktif'""", (vid,))
        kart = [dict(r) for r in cur.fetchall() or []]
        cur.execute("""SELECT id, tarih::text AS tarih, tutar::float AS tutar,
                              islem_turu, COALESCE(kasa_etkisi,TRUE) AS kasa_etkisi,
                              COALESCE(aciklama,'') AS aciklama
                         FROM kasa_hareketleri
                        WHERE kaynak_id=%s AND COALESCE(durum,'aktif')='aktif'""",
                    (vid,))
        kasa = [dict(r) for r in cur.fetchall() or []]
        cur.execute("""SELECT id, durum, odenecek_tutar::float AS odenecek,
                              COALESCE(odenen_tutar,0)::float AS odenen
                         FROM odeme_plani
                        WHERE kaynak_tablo='vadeli_alimlar' AND kaynak_id=%s""",
                    (vid,))
        plan = [dict(r) for r in cur.fetchall() or []]
    _kasa_etkili = [k for k in kasa if k.get("kasa_etkisi")]
    return {
        "vadeli": v,
        "kart_izleri": kart,
        "kasa_izleri": kasa,
        "planlar": plan,
        "kasa_etkili_iz_adet": len(_kasa_etkili),
        "geri_alinirsa_kasaya_donecek_tl": round(
            sum(abs(float(k.get("tutar") or 0)) for k in _kasa_etkili), 2),
        "hukum": (
            "KAGIT UZERINDE KAPANMIS — geri alma hicbir parayi hareket "
            "ettirmez (kasa izi yok, kart izi yok)."
            if not kart and not _kasa_etkili else
            "GERCEK PARA IZI VAR — geri alma kasayi/kart borcunu DEGISTIRIR. "
            "Once bu paranin gercekten cikip cikmadigini dogrulayin."),
        "not": ("SALT OKUR. Geri almak icin: "
                "POST /api/vadeli-alimlar/{vid}/odeme-geri-al {mod: 'iptal'|'bekliyor'}"),
    }


@app.post("/api/vadeli-alimlar/{vid}/odeme-geri-al")
def vadeli_odeme_geri_al(vid: str, body: VadeliGeriAlModel):
    """ÖDENMİŞ vadeli alımı geri alır — çift kayıt / yanlış ödeme işareti düzeltmesi.

    vadeli_sil yalnız 'bekliyor' kaydı iptal eder; ödenmiş kayıtta üç iz birden
    açık kalır (kart HARCAMA satırı · kasa VADELI_ODEME satırı · odeme_plani
    'odendi'). Bu uç üçünü de kaynağından (kaynak_id=vid) bulup geri alır:
      - kart izi  → durum='iptal' (kart borcu düşer, plan yeniden üretilir)
      - kasa izi  → ters kayıt (iptal_kasa_hareketi — defter silinmez)
      - plan      → mod'a göre 'iptal' ya da 'bekliyor' (odeme alanları sıfırlanır)
      - vadeli    → mod
    Cari ekstre ödeme izini vadeli_alimlar.durum='odendi'den okuduğu için
    yalnız kart iznini iptal etmek EKSTREYİ DÜZELTMEZ — bu yüzden tek kapı.
    Her adım audit'li; hiçbir satır silinmez.
    """
    mod = (body.mod or '').strip().lower()
    if mod not in ('bekliyor', 'iptal'):
        raise HTTPException(400, "mod: bekliyor | iptal")
    with db() as (conn, cur):
        cur.execute("SELECT * FROM vadeli_alimlar WHERE id=%s AND durum='odendi' FOR UPDATE", (vid,))
        v = cur.fetchone()
        if not v:
            raise HTTPException(404, "Ödenmiş vadeli alım bulunamadı (yalnız durum='odendi' geri alınabilir)")

        # 1) KART izleri — vadeli_ode kart yolunda kaynak_id=vid ile yazılır
        cur.execute("""
            SELECT id FROM kart_hareketleri
            WHERE kaynak_tablo='vadeli_alimlar' AND kaynak_id=%s
              AND islem_turu='HARCAMA' AND durum='aktif'
        """, (vid,))
        kart_izleri = [r['id'] for r in cur.fetchall() or []]
        for hid in kart_izleri:
            cur.execute("UPDATE kart_hareketleri SET durum='iptal' WHERE id=%s", (hid,))
            audit(cur, 'kart_hareketleri', hid, 'VADELI_ODEME_GERI_AL')

        # 2) KASA izi — ters kayıt (defter append-only kalır)
        cur.execute("""
            SELECT id FROM kasa_hareketleri
            WHERE kaynak_id=%s AND islem_turu='VADELI_ODEME' AND durum='aktif'
        """, (vid,))
        kasa_izi_var = bool(cur.fetchone())
        if kasa_izi_var:
            iptal_kasa_hareketi(cur, vid, 'vadeli_alimlar', 'VADELI_ODEME',
                                'VADELI_ODEME_GERI_AL', 'Vadeli ödeme geri alındı')

        # 3) PLAN — ödeme alanları sıfırlanır; mod'a göre kapanır ya da geri açılır
        plan_durum = 'iptal' if mod == 'iptal' else 'bekliyor'
        cur.execute("""
            UPDATE odeme_plani SET durum=%s, odeme_tarihi=NULL, odenen_tutar=NULL
            WHERE kaynak_tablo='vadeli_alimlar' AND kaynak_id=%s AND durum='odendi'
        """, (plan_durum, vid))

        # 4) VADELİ kaydın kendisi
        cur.execute("UPDATE vadeli_alimlar SET durum=%s WHERE id=%s", (mod, vid))
        audit(cur, 'vadeli_alimlar', vid, 'ODEME_GERI_AL',
              eski=v, yeni={'mod': mod, 'aciklama': (body.aciklama or '').strip() or None})

        if kart_izleri:
            kart_plan_guncelle_tx(cur)
        uyari_cache_clear()
    return {
        "success": True, "mod": mod,
        "kart_izi_iptal": len(kart_izleri), "kasa_izi_ters_kayit": kasa_izi_var,
        "plan_durum": plan_durum,
    }

@app.get("/api/vadeli-alimlar/{vid}/kart-oneri")
def vadeli_kart_oneri(vid: str):
    """Vadeli alım ödemesi için kart önerisi — kanonik tek kaynaktan
    (_kart_oneri_hesapla; kalan limit ekstre gerçeği + taksit + ortak havuz dahil)."""
    with db() as (conn, cur):
        cur.execute("SELECT * FROM vadeli_alimlar WHERE id=%s", (vid,))
        v = cur.fetchone()
        if not v: raise HTTPException(404)
        odeme_tutari = float(v['tutar'])
    sonuc = _kart_oneri_hesapla(odeme_tutari)
    return {
        'vadeli_alim': {'id': str(v['id']), 'aciklama': v['aciklama'], 'tutar': odeme_tutari},
        'kartlar': sonuc,
        'oneri_var': any(k['oneri'] for k in sonuc)
    }


@app.post("/api/vadeli-alimlar/{vid}/ode")
def vadeli_ode(vid: str, body: VadeliOdeModel = VadeliOdeModel()):
    with db() as (conn, cur):
        # FIX SEC2 (2026-07-06): FOR UPDATE — eşzamanlı iki tam-ödeme isteği aynı vadeliyi iki
        # kez kasadan düşürebiliyordu. İkinci istek ilki commit edene kadar bekler; vadeli
        # kapanınca durum='bekliyor' bulamaz (404) veya ödendi-guard'ına takılır.
        cur.execute("SELECT * FROM vadeli_alimlar WHERE id=%s AND durum='bekliyor' FOR UPDATE", (vid,))
        v = cur.fetchone()
        if not v: raise HTTPException(404)

        # ÇİFT ÖDEME GUARD — bağlı aktif odeme_plani varsa zaten ödenmemiş demektir.
        # Önce planı bul ki ödenecek gerçek tutarı (plan) belirleyelim; kart limiti de
        # bu tutara göre kontrol edilsin.
        aktif_plan = _vadeli_aktif_plan_bul(cur, vid)
        if not aktif_plan:
            odenen = vadeli_kasadan_odenen_toplam(cur, vid)
            if odenen >= float(v['tutar']):
                raise HTTPException(400, "Bu vadeli alım zaten tam olarak kasaya işlenmiş, tekrar ödeme yapılamaz.")
        plan = aktif_plan
        if not plan:
            # SELF-HEAL: aktif ödeme planı yoksa (eski kayıt veya planı 'sistem başlangıcı'
            # temizliğinde iptal olmuş vadeli) ödeme anında plan üret → "plan bulunamadı"
            # kırmızı hatası yerine ödeme sorunsuz tamamlanır. vadeli_ekle ile aynı INSERT.
            pid_yeni = str(uuid.uuid4())
            cur.execute(
                """
                INSERT INTO odeme_plani
                    (id, kart_id, tarih, referans_ay, odenecek_tutar, asgari_tutar,
                     aciklama, durum, kaynak_tablo, kaynak_id)
                VALUES (%s, NULL, %s, DATE_TRUNC('month', %s::date), %s, %s, %s,
                        'bekliyor', 'vadeli_alimlar', %s)
                RETURNING id, odenecek_tutar
                """,
                (pid_yeni, v['vade_tarihi'], str(v['vade_tarihi']), float(v['tutar']),
                 float(v['tutar']), f"Vadeli Alım: {v.get('aciklama') or ''}", vid),
            )
            plan = cur.fetchone()
        if not plan:
            raise HTTPException(400, "Bu vadeli alım için ödeme planı bulunamadı")
        _vadeli_diger_aktif_planlari_iptal(cur, vid, plan["id"])

        bugun = str(bugun_tr())
        tutar = float(plan['odenecek_tutar'])  # vadeli_alimlar.tutar değil, planın tutarı

        # KART seçildiyse validasyon + limit kontrolü (PLAN tutarı üzerinden)
        if body.odeme_yontemi == 'kart':
            if not body.kart_id:
                raise HTTPException(400, "Kart seçimi zorunlu")
            cur.execute("SELECT * FROM kartlar WHERE id=%s AND aktif=TRUE FOR UPDATE", (body.kart_id,))
            kart = cur.fetchone()
            if not kart: raise HTTPException(404, "Kart bulunamadı")
            borc = kart_borc(cur, body.kart_id)
            kalan_limit = _kanonik_kalan_limit(body.kart_id, float(kart['limit_tutar']) - borc)
            if kalan_limit < tutar:
                raise HTTPException(400, f"Kart limiti yetersiz. Kalan: {kalan_limit:,.0f} ₺")

        # Onay kuyruğunda bekleyen VADELI_ODEME varsa kapat
        cur.execute("""
            UPDATE onay_kuyrugu SET durum='reddedildi'
            WHERE kaynak_id=%s AND islem_turu='VADELI_ODEME' AND durum='bekliyor'
        """, (vid,))

        if body.odeme_yontemi == 'kart':
            # KART: kart borcuna HARCAMA ekle — kasaya yazma yok
            hid = str(uuid.uuid4())
            cur.execute("""
                INSERT INTO kart_hareketleri
                    (id, kart_id, tarih, islem_turu, tutar, taksit_sayisi, aciklama, kaynak_id, kaynak_tablo)
                VALUES (%s, %s, %s, 'HARCAMA', %s, 1, %s, %s, 'vadeli_alimlar')
            """, (hid, body.kart_id, bugun, tutar, f"Vadeli alım: {v['aciklama']}", vid))
            audit(cur, 'kart_hareketleri', hid, 'VADELI_KART')
            kart_plan_guncelle_tx(cur)
            cur.execute("UPDATE odeme_plani SET durum='odendi', odeme_tarihi=%s, odenen_tutar=%s WHERE id=%s",
                (bugun, tutar, plan['id']))
            vadeli_alim_kapat(cur, vid, bugun)
            audit(cur, 'vadeli_alimlar', vid, 'ODENDI_KART')
            uyari_cache_clear()
            return {"success": True, "odeme_yontemi": "kart", "kart_id": body.kart_id}

        # NAKİT: doğrudan kasaya yaz — onay kuyruğuna girmez
        insert_kasa_hareketi(
            cur, bugun, 'VADELI_ODEME', -abs(tutar),
            f"Vadeli alım ödemesi: {v['aciklama']}",
            'vadeli_alimlar', vid,
            str(uuid.uuid4()), 'VADELI_ALIMLAR'
        )
        cur.execute("UPDATE odeme_plani SET durum='odendi', odeme_tarihi=%s, odenen_tutar=%s WHERE id=%s",
            (bugun, tutar, plan['id']))
        cur.execute("""
            UPDATE onay_kuyrugu SET durum='onaylandi', onay_tarihi=NOW()
            WHERE durum NOT IN ('onaylandi','reddedildi')
            AND (kaynak_id=%s OR kaynak_id=%s)
        """, (plan['id'], vid))
        vadeli_alim_kapat(cur, vid, bugun)
        audit(cur, 'vadeli_alimlar', vid, 'ODENDI_NAKIT')
        uyari_cache_clear()
    return {"success": True, "odeme_yontemi": "nakit"}

@app.post("/api/vadeli-alimlar/{vid}/kismi-ode")
def vadeli_kismi_ode(vid: str, body: KismiOdeModel):
    """
    Vadeli alım kısmi ödeme.
    Nakit: ödenen kısım kasadan düşer, kalan yeni plan olarak bekler.
    Kart: ödenen kısım kart harcamasına eklenir (kasaya yazılmaz), kalan yeni plan bekler.
    """
    with db() as (conn, cur):
        cur.execute("SELECT * FROM vadeli_alimlar WHERE id=%s AND durum='bekliyor'", (vid,))
        v = cur.fetchone()
        if not v: raise HTTPException(404, "Vadeli alım bulunamadı veya zaten ödendi")

        # KART seçildiyse validasyon
        if body.odeme_yontemi == 'kart':
            if not body.kart_id:
                raise HTTPException(400, "Kart seçimi zorunlu")
            cur.execute("SELECT * FROM kartlar WHERE id=%s AND aktif=TRUE FOR UPDATE", (body.kart_id,))
            kart = cur.fetchone()
            if not kart: raise HTTPException(404, "Kart bulunamadı")
            borc = kart_borc(cur, body.kart_id)
            kalan_limit = _kanonik_kalan_limit(body.kart_id, float(kart['limit_tutar']) - borc)
            if kalan_limit < body.odenen_tutar:
                raise HTTPException(400, f"Kart limiti yetersiz. Kalan: {kalan_limit:,.0f} ₺")

        # Bağlı aktif odeme_plani'nı bul — yoksa zaten ödenmiş demektir
        plan = _vadeli_aktif_plan_bul(cur, vid)
        if not plan:
            raise HTTPException(400, "Bu vadeli alım için aktif ödeme planı bulunamadı — zaten ödenmiş olabilir.")

    # kismi_odeme_yap — nakit/kart bilgisini body üzerinden taşır
    return kismi_odeme_yap(plan['id'], body)

# ── BORÇLAR ────────────────────────────────────────────────────
class BorcModel(BaseModel):
    kurum: str
    borc_turu: str = 'Kredi'
    toplam_borc: Optional[float] = None
    aylik_taksit: float
    kalan_vade: Optional[int] = None
    toplam_vade: Optional[int] = None
    baslangic_tarihi: Optional[date] = None
    odeme_gunu: int = 1
    # Borç Navigasyon (takvim/amortisman) için ek alanlar — hepsi opsiyonel; gönderilmezse
    # mevcut değer korunur (PUT'ta COALESCE). toplam_borc artık KALAN ANAPARA olarak kullanılır.
    faiz_orani: Optional[float] = None        # AYLIK faiz % (ör. 3.46)
    odemesiz_ay: Optional[int] = None         # ödemesiz dönem (ay); ilk taksit gecikir
    ilk_taksit_tarihi: Optional[date] = None  # ilk taksit tarihi (ödemesiz dönem varsa kritik)


def _ensure_borc_kolonlar(cur) -> None:
    """borc_envanteri'ne takvim alanlarını idempotent ekle (hata-yutar)."""
    try:
        cur.execute("ALTER TABLE borc_envanteri ADD COLUMN IF NOT EXISTS faiz_orani NUMERIC")
        cur.execute("ALTER TABLE borc_envanteri ADD COLUMN IF NOT EXISTS odemesiz_ay INTEGER DEFAULT 0")
        cur.execute("ALTER TABLE borc_envanteri ADD COLUMN IF NOT EXISTS ilk_taksit_tarihi DATE")
    except Exception:
        pass

class SubeGuncelleModel(BaseModel):
    pos_oran: float = 0
    online_oran: float = 0
    acilis_saati: Optional[str] = None
    kapanis_saati: Optional[str] = None
    yogun_saat_baslangic: Optional[str] = None
    yogun_saat_bitis: Optional[str] = None
    ortusme_gerekli: bool = False
    vardiya_yazilsin: bool = True
    acilis_sadece_part: bool = False
    kapanis_sadece_part: bool = False
    min_personel: int = 1
    yogun_saat_ek_personel: int = 0
    # Aynı gün açılış slotuna yazılabilecek üst kişi sayısı (örn. Alsancak = 1); boş = sınır yok
    acilis_max_kisi: Optional[int] = None
    sube_tipi: Optional[str] = None


class KasaDuzeltModel(BaseModel):
    baslangic: date
    bitis: Optional[date] = None

def _borc_siradaki_taksit(borc: dict):
    """
    Sayaç bazlı sıradaki taksit: no = toplam_vade - kalan_vade + 1.
    Vade tarihi = ilk_taksit_tarihi + (no-1) ay; ilk taksit tarihi yoksa
    başlangıç ayı + odeme_gunu üzerinden tahmin edilir.

    ⚠ NEDEN TAKVİM AYI DEĞİL (KOÇ FİNANS dersi, 2026-07-28): gecikmeli ödemede
    iki FARKLI taksit aynı takvim ayına düşebilir — Haziran taksiti (vade 29.06)
    04.07'de ödendi, Temmuz taksiti (vade 29.07) aynı ay içinde kaydedilmek
    istendi. Ay-bazlı kontrol bunu "çift ödeme" sanıp meşru taksidi bloke etti.
    Taksidin kimliği takvim ayı değil TAKSİT NUMARASIDIR; kalan_vade her ödemede
    düştüğü için sayaç kasa iziyle birlikte ilerler (kasa izi = tek gerçek).
    Dönüş: (taksit_no | None, vade_tarihi | None). None = vade takibi yapılamıyor
    (toplam_vade tanımsız) ya da tüm taksitler bitti.
    """
    toplam_vade = borc.get('toplam_vade')
    kalan_vade = borc.get('kalan_vade')
    if not toplam_vade or kalan_vade is None:
        return None, None
    no = int(toplam_vade) - int(kalan_vade) + 1
    if no < 1 or no > int(toplam_vade):
        return None, None
    ilk = borc.get('ilk_taksit_tarihi')
    if isinstance(ilk, str):
        try:
            ilk = date.fromisoformat(ilk[:10])
        except ValueError:
            ilk = None
    if not ilk:
        bas = borc.get('baslangic_tarihi')
        if isinstance(bas, str):
            try:
                bas = date.fromisoformat(bas[:10])
            except ValueError:
                bas = None
        if bas:
            gun = int(borc.get('odeme_gunu') or 1)
            gun = min(gun, calendar.monthrange(bas.year, bas.month)[1])
            # kredi genelde başlangıçtan bir ay sonra ilk taksite döner
            ilk = ay_ekle(date(bas.year, bas.month, gun), 1)
    if not ilk:
        return no, None
    return no, ay_ekle(ilk, no - 1)


@app.get("/api/borclar")
def borclar_listele():
    """
    Borç listesi + taksit dönemi durumu. Frontend "Öde" butonunu ve
    güncel/gecikmiş rozetini bunlara göre yönetir:
    - siradaki_taksit_no / siradaki_taksit_vade: sayaç bazlı sıradaki taksit
    - vadesi_gecmis_odenmemis: sıradaki taksitin vadesi geçti (gecikmiş)
    - bu_ay_odendi: GERİYE UYUM alanı — artık "vadesi gelmiş tüm taksitler
      ödendi" anlamında (dönem bazlı; takvim ayı DEĞİL — KOÇ FİNANS dersi)
    """
    with db() as (conn, cur):
        cur.execute("SELECT * FROM borc_envanteri ORDER BY kurum")
        borclar = [dict(r) for r in cur.fetchall()]
        if not borclar:
            return []
        ids = [b['id'] for b in borclar]
        cur.execute(
            """
            SELECT kaynak_id, MAX(tarih) AS son_odeme
            FROM kasa_hareketleri
            WHERE kaynak_tablo='borc_envanteri'
              AND kaynak_id = ANY(%s)
              AND kasa_etkisi = TRUE
              AND tutar < 0
              AND durum='aktif'
            GROUP BY kaynak_id
            """,
            (ids,),
        )
        odeme_map = {str(r['kaynak_id']): r for r in cur.fetchall()}
        bugun = date.today()
        for b in borclar:
            rec = odeme_map.get(str(b['id'])) or {}
            b['son_odeme'] = str(rec['son_odeme']) if rec.get('son_odeme') else None
            no, vade = _borc_siradaki_taksit(b)
            b['siradaki_taksit_no'] = no
            b['siradaki_taksit_vade'] = str(vade) if vade else None
            b['vadesi_gecmis_odenmemis'] = bool(vade and vade <= bugun)
            # Dönem bazlı "güncel" bayrağı: vadesi gelmiş ödenmemiş taksit yok.
            # Vade takibi yapılamayan borçta (toplam_vade tanımsız) eski takvim-ayı
            # davranışı korunur — regresyon olmasın.
            if no is None and b.get('toplam_vade'):
                b['bu_ay_odendi'] = True   # tüm taksitler bitti
            elif no is None:
                son = rec.get('son_odeme')
                b['bu_ay_odendi'] = bool(
                    son and son.year == bugun.year and son.month == bugun.month
                )
            else:
                b['bu_ay_odendi'] = not b['vadesi_gecmis_odenmemis'] if vade else False
        return borclar


class BorcOdemeBody(BaseModel):
    tutar: Optional[float] = None
    tarih: Optional[str] = None
    aciklama: Optional[str] = None
    # CAS: frontend'in ekranda gördüğü sıradaki taksit no'su. Sunucudakiyle
    # uyuşmazsa (başka pencere/çift tık az önce ödedi) 409 döner — yanlış
    # taksite ödeme yazılamaz. Vade takipli kredilerde zorunlu.
    beklenen_taksit_no: Optional[int] = None


@app.post("/api/borclar/{bid}/ode")
def borc_ode(bid: str, body: BorcOdemeBody):
    """
    Manuel taksit ödemesi (nakit kasa düşümü):
    - Kasadan düşülür (islem_turu=BORC_TAKSIT, tutar negatif).
    - borc_envanteri: kalan_vade -= 1, toplam_borc -= ödenen (0'a kırpılır).
    - Aynı ay tekrar ödeme engellenir (idempotency + bu_ay_odendi kontrolü).
    - Frontend butonu "✓ Bu Ay Ödendi" olur.
    """
    with db() as (conn, cur):
        cur.execute("SELECT * FROM borc_envanteri WHERE id=%s FOR UPDATE", (bid,))
        borc = cur.fetchone()
        if not borc:
            raise HTTPException(404, "Borç bulunamadı")
        if not borc['aktif']:
            raise HTTPException(400, "Pasif borç için ödeme yapılamaz")
        kalan_vade = int(borc['kalan_vade'] or 0)
        if borc['kalan_vade'] is not None and kalan_vade <= 0:
            raise HTTPException(400, "Bu borç için kalan taksit yok")

        tutar = float(body.tutar) if body.tutar else float(borc['aylik_taksit'] or 0)
        if tutar <= 0:
            raise HTTPException(400, "Geçerli tutar girin")

        tarih = (body.tarih or date.today().isoformat())[:10]

        # ÇİFT ÖDEME KAPISI — TAKSİT DÖNEMİ BAZLI (KOÇ FİNANS dersi, 2026-07-28).
        # Eski kapı TAKVİM AYINA bakıyordu: gecikmeli ödemede iki farklı taksit
        # aynı takvim ayına düşünce (Haziran taksiti 04.07'de ödendi, Temmuz
        # taksiti 28.07'de kaydedilmek istendi) meşru ödemeyi 409'la bloke etti.
        # Taksidin kimliği takvim ayı değil TAKSİT NO'sudur: kalan_vade her
        # ödemede düştüğü için sayaç kasa iziyle birlikte ilerler.
        taksit_no, taksit_vade = _borc_siradaki_taksit(dict(borc))
        if borc['toplam_vade']:
            # CAS: ekranda görülen taksit ile sunucudaki sıradaki taksit aynı mı?
            # Çift tık / ikinci pencere aynı no'yu gönderir → ilki öder, ikincisi
            # burada 409 alır (kalan_vade düştüğü için no ilerlemiştir).
            if body.beklenen_taksit_no is None:
                raise HTTPException(
                    409,
                    "Ödeme ekranı eski sürümde — sayfayı yenileyip (Ctrl+F5) tekrar deneyin",
                )
            if taksit_no is None:
                raise HTTPException(400, "Bu borcun tüm taksitleri zaten ödenmiş görünüyor")
            if int(body.beklenen_taksit_no) != int(taksit_no):
                raise HTTPException(
                    409,
                    f"Sıradaki taksit değişti (şimdi {taksit_no}/{borc['toplam_vade']}) — "
                    "bu taksit az önce başka bir yerden ödenmiş olabilir; listeyi yenileyin",
                )
            idem = f"borc-ode:{bid}:taksit-{taksit_no}"
        else:
            # Vade takibi yapılamayan borç (toplam_vade tanımsız): eski takvim-ayı
            # koruması aynen korunur — regresyon yok.
            cur.execute(
                """
                SELECT 1 FROM kasa_hareketleri kh
                WHERE kh.islem_turu='BORC_TAKSIT' AND kh.kasa_etkisi=TRUE AND kh.durum='aktif'
                  AND EXTRACT(YEAR FROM kh.tarih)  = EXTRACT(YEAR FROM %s::date)
                  AND EXTRACT(MONTH FROM kh.tarih) = EXTRACT(MONTH FROM %s::date)
                  AND (
                        (kh.kaynak_tablo='borc_envanteri' AND kh.kaynak_id=%s)
                     OR (kh.kaynak_tablo='odeme_plani' AND kh.kaynak_id IN (
                            SELECT id FROM odeme_plani
                            WHERE kaynak_tablo='borc_envanteri' AND kaynak_id=%s))
                  )
                LIMIT 1
                """,
                (tarih, tarih, bid, bid),
            )
            if cur.fetchone():
                raise HTTPException(409, "Bu ay için zaten ödeme kaydı var (manuel veya plandan)")
            idem = f"borc-ode:{bid}:{tarih[:7]}"  # ay bazlı idempotent (vadesiz borç)

        aciklama = (body.aciklama or f"{borc['kurum']} — {borc['borc_turu']} taksiti").strip()
        if borc['toplam_vade'] and taksit_no:
            aciklama = f"{aciklama} ({taksit_no}/{borc['toplam_vade']}. taksit"
            aciklama += f", vade {taksit_vade})" if taksit_vade else ")"
        ref_id = str(uuid.uuid4())
        insert_kasa_hareketi(
            cur, tarih, 'BORC_TAKSIT', -abs(tutar), aciklama,
            'borc_envanteri', bid, ref_id, 'BORC_TAKSIT', idempotency_key=idem,
        )

        # Borç kaydını güncelle
        yeni_kalan = (kalan_vade - 1) if borc['kalan_vade'] is not None else None
        yeni_toplam = max(0.0, float(borc['toplam_borc'] or 0) - tutar)
        # Vade ile kapan; vade tanımsızsa (NULL) toplam borç sıfırlanınca kapan.
        # FIX MN9 (2026-07-06): toplam_borc da NULL ise (0 sayılıp) İLK ödemede borç yanlışlıkla
        # kapanıyordu — iki alan da tanımsızken kapanma kararı VERİLEMEZ, borç açık kalır.
        kapansin = (yeni_kalan is not None and yeni_kalan <= 0) or (
            yeni_kalan is None and borc['toplam_borc'] is not None and yeni_toplam <= 0
        )
        cur.execute(
            """
            UPDATE borc_envanteri
            SET kalan_vade=%s,
                toplam_borc=%s,
                aktif = CASE WHEN %s THEN FALSE ELSE aktif END
            WHERE id=%s
            """,
            (yeni_kalan, yeni_toplam, kapansin, bid),
        )
        audit(cur, 'borc_envanteri', bid, 'ODEME', eski=dict(borc))

        # Bu ay için bekleyen bir odeme_plani varsa kapat (çift ödemeyi önler)
        cur.execute(
            """
            UPDATE odeme_plani
            SET durum='odendi', odenen_tutar=%s, odeme_tarihi=%s
            WHERE kaynak_tablo='borc_envanteri' AND kaynak_id=%s
              AND durum IN ('bekliyor','onay_bekliyor')
              AND EXTRACT(YEAR FROM tarih)  = EXTRACT(YEAR FROM %s::date)
              AND EXTRACT(MONTH FROM tarih) = EXTRACT(MONTH FROM %s::date)
            """,
            (tutar, tarih, bid, tarih, tarih),
        )

        return {
            "success": True,
            "odenen":     tutar,
            "tarih":      tarih,
            "kalan_vade": yeni_kalan,
            "toplam_borc": yeni_toplam,
            "kapandi":    kapansin,
        }

def _borc_validate(b: BorcModel):
    """Borç ekle/güncelle ortak doğrulaması — negatif/tutarsız değerleri reddet."""
    if not (b.kurum or "").strip():
        raise HTTPException(400, "Kurum/alacaklı adı zorunlu")
    if b.aylik_taksit is None or float(b.aylik_taksit) <= 0:
        raise HTTPException(400, "Aylık taksit 0'dan büyük olmalı")
    if b.toplam_borc is not None and float(b.toplam_borc) < 0:
        raise HTTPException(400, "Toplam borç negatif olamaz")
    if not (1 <= int(b.odeme_gunu or 0) <= 31):
        raise HTTPException(400, "Ödeme günü 1–31 arası olmalı")
    if b.kalan_vade is not None and int(b.kalan_vade) < 0:
        raise HTTPException(400, "Kalan vade negatif olamaz")
    if b.toplam_vade is not None and int(b.toplam_vade) < 0:
        raise HTTPException(400, "Toplam vade negatif olamaz")
    if (b.kalan_vade is not None and b.toplam_vade is not None
            and int(b.kalan_vade) > int(b.toplam_vade)):
        raise HTTPException(400, "Kalan vade, toplam vadeden büyük olamaz")
    # FIX MN8 (2026-07-06): faiz/ödemesiz ay validasyonu yoktu — negatif/aşırı faiz girilip
    # Borç Koçu motoruna kirli veri akabiliyordu (kartlardaki 0-500 deseniyle aynı).
    # EVV-SAG-N8 (2026-08-14): mesaj "yıllık %" diyordu ama sözleşme AYLIK —
    # v2 kredi formu "Aylık faiz %" ister, canlı değerler 2,99-4,375 bandında ve
    # borç navigasyon motoru bu alanı AYLIK oran olarak kullanır.
    if b.faiz_orani is not None and not (0 <= float(b.faiz_orani) <= 500):
        raise HTTPException(400, "Faiz oranı 0–500 arası olmalı (aylık %)")
    if b.odemesiz_ay is not None and int(b.odemesiz_ay) < 0:
        raise HTTPException(400, "Ödemesiz ay negatif olamaz")


@app.post("/api/borclar")
def borc_ekle(b: BorcModel):
    _borc_validate(b)
    with db() as (conn, cur):
        _ensure_borc_kolonlar(cur)
        bid = str(uuid.uuid4())
        cur.execute("""INSERT INTO borc_envanteri
            (id,kurum,borc_turu,toplam_borc,aylik_taksit,kalan_vade,toplam_vade,baslangic_tarihi,odeme_gunu,faiz_orani,odemesiz_ay,ilk_taksit_tarihi)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (bid, b.kurum, b.borc_turu, b.toplam_borc, b.aylik_taksit, b.kalan_vade, b.toplam_vade,
             b.baslangic_tarihi, b.odeme_gunu, b.faiz_orani, (b.odemesiz_ay or 0), b.ilk_taksit_tarihi))
        audit(cur, 'borc_envanteri', bid, 'INSERT')
    return {"id": bid, "success": True}

@app.put("/api/borclar/{bid}")
def borc_guncelle(bid: str, b: BorcModel):
    _borc_validate(b)
    with db() as (conn, cur):
        _ensure_borc_kolonlar(cur)
        cur.execute("SELECT * FROM borc_envanteri WHERE id=%s", (bid,))
        eski = cur.fetchone()
        if not eski: raise HTTPException(404)
        # Ek alanlar (faiz/ödemesiz/ilk taksit) gönderilmezse MEVCUT korunur (COALESCE) —
        # eski borçlar UI'ı bu alanları yollamıyor, sıfırlanmasın.
        cur.execute("""UPDATE borc_envanteri SET kurum=%s,borc_turu=%s,toplam_borc=%s,aylik_taksit=%s,
            kalan_vade=%s,toplam_vade=%s,baslangic_tarihi=%s,odeme_gunu=%s,
            faiz_orani=COALESCE(%s,faiz_orani),
            odemesiz_ay=COALESCE(%s,odemesiz_ay),
            ilk_taksit_tarihi=COALESCE(%s,ilk_taksit_tarihi)
            WHERE id=%s""",
            (b.kurum, b.borc_turu, b.toplam_borc, b.aylik_taksit, b.kalan_vade, b.toplam_vade,
             b.baslangic_tarihi, b.odeme_gunu, b.faiz_orani, b.odemesiz_ay, b.ilk_taksit_tarihi, bid))

        # Bekleyen ödeme planı senkronu — eski tutar/yetim plan kalmasın
        if b.kalan_vade is not None and int(b.kalan_vade) <= 0:
            # Borç bitti → bu giderin bekleyen planlarını + onaylarını iptal et
            cur.execute(
                """UPDATE odeme_plani SET durum='iptal'
                   WHERE kaynak_tablo='borc_envanteri' AND kaynak_id=%s
                     AND durum IN ('bekliyor','onay_bekliyor')""",
                (bid,),
            )
            cur.execute(
                """UPDATE onay_kuyrugu SET durum='reddedildi', onay_tarihi=NOW()
                   WHERE kaynak_tablo='odeme_plani' AND durum='bekliyor'
                     AND kaynak_id IN (SELECT id FROM odeme_plani
                                       WHERE kaynak_tablo='borc_envanteri' AND kaynak_id=%s)""",
                (bid,),
            )
        else:
            # Aktif borç → bekleyen planların tutarını yeni aylık taksite çek (eski tutarla ödenmesin)
            cur.execute(
                """UPDATE odeme_plani
                   SET odenecek_tutar=%s, asgari_tutar=%s
                   WHERE kaynak_tablo='borc_envanteri' AND kaynak_id=%s
                     AND durum IN ('bekliyor','onay_bekliyor')""",
                (b.aylik_taksit, b.aylik_taksit, bid),
            )
        audit(cur, 'borc_envanteri', bid, 'UPDATE', eski=eski)
    return {"success": True}


@app.get("/api/borclar/{bid}/gecmis")
def borc_gecmis(bid: str):
    """
    Bir borcun tüm ödeme geçmişi:
    - Ödenen taksitler (kasa_hareketleri)
    - Bekleyen / gelecek ödemeler (odeme_plani)
    - Özet: toplam ödenen, kalan, ilerleme
    """
    with db() as (conn, cur):
        cur.execute("SELECT * FROM borc_envanteri WHERE id=%s", (bid,))
        borc = cur.fetchone()
        if not borc: raise HTTPException(404, "Borç bulunamadı")

        # Ödenen taksitler — kasa_hareketleri
        cur.execute("""
            SELECT tarih, tutar, aciklama, islem_turu, durum
            FROM kasa_hareketleri
            WHERE kaynak_tablo = 'borc_envanteri'
            AND kaynak_id = %s AND kasa_etkisi = true
            AND tutar < 0
            ORDER BY tarih DESC
        """, (bid,))
        odenenler = [{
            "tarih":    str(r['tarih']),
            "tutar":    abs(float(r['tutar'])),
            "aciklama": r['aciklama'] or '',
            "durum":    "odendi",
        } for r in cur.fetchall()]

        # Bekleyen ödemeler — odeme_plani
        cur.execute("""
            SELECT id, tarih, odenecek_tutar, asgari_tutar, durum, aciklama
            FROM odeme_plani
            WHERE kaynak_tablo = 'borc_envanteri'
            AND kaynak_id = %s
            AND durum IN ('bekliyor', 'onay_bekliyor')
            ORDER BY tarih ASC
        """, (bid,))
        bekleyenler = [{
            "tarih":   str(r['tarih']),
            "tutar":   float(r['odenecek_tutar']),
            "aciklama": r['aciklama'] or '',
            "durum":   r['durum'],
            "plan_id": str(r['id']),
        } for r in cur.fetchall()]

        # Özet hesapla
        toplam_odenen   = sum(r['tutar'] for r in odenenler)
        toplam_beklenen = sum(r['tutar'] for r in bekleyenler)
        toplam_borc     = float(borc['toplam_borc'] or 0)
        aylik_taksit    = float(borc['aylik_taksit'] or 0)
        kalan_vade      = int(borc['kalan_vade'] or 0)
        toplam_vade     = int(borc['toplam_vade'] or 0)
        gecen_taksit    = max(0, toplam_vade - kalan_vade) if toplam_vade else len(odenenler)
        ilerleme_pct    = min(100, max(0, round(gecen_taksit / toplam_vade * 100))) if toplam_vade else 0

        return {
            "borc": {
                "id":              str(borc['id']),
                "kurum":           borc['kurum'],
                "borc_turu":       borc['borc_turu'],
                "toplam_borc":     toplam_borc,
                "aylik_taksit":    aylik_taksit,
                "kalan_vade":      kalan_vade,
                "toplam_vade":     toplam_vade,
                "baslangic":       str(borc['baslangic_tarihi']) if borc['baslangic_tarihi'] else None,
                "aktif":           borc['aktif'],
            },
            "ozet": {
                "toplam_odenen":   round(toplam_odenen, 2),
                "toplam_beklenen": round(toplam_beklenen, 2),
                "kalan_borc":      round(max(0, toplam_borc - toplam_odenen), 2),
                "gecen_taksit":    gecen_taksit,
                "kalan_taksit":    kalan_vade,
                "ilerleme_pct":    ilerleme_pct,
            },
            "odenenler":   odenenler,
            "bekleyenler": bekleyenler,
        }

@app.delete("/api/borclar/{bid}")
def borc_sil(bid: str):
    with db() as (conn, cur):
        cur.execute("SELECT * FROM borc_envanteri WHERE id=%s", (bid,))
        eski = cur.fetchone()
        if not eski: raise HTTPException(404)
        cur.execute("UPDATE borc_envanteri SET aktif=FALSE WHERE id=%s", (bid,))
        # Bağlı bekleyen planları iptal et — panelde görünmesin
        cur.execute("""
            UPDATE odeme_plani SET durum='iptal'
            WHERE kaynak_tablo='borc_envanteri' AND kaynak_id=%s
            AND durum IN ('bekliyor','onay_bekliyor')
        """, (bid,))
        cur.execute("""
            UPDATE onay_kuyrugu SET durum='reddedildi'
            WHERE kaynak_tablo='borc_envanteri' AND kaynak_id=%s
            AND durum='bekliyor'
        """, (bid,))
        audit(cur, 'borc_envanteri', bid, 'PASIF', eski=eski)
    return {"success": True}

# ── ŞUBELER ────────────────────────────────────────────────────
@app.get("/api/subeler")
def subeler():
    with db() as (conn, cur):
        cur.execute("SELECT * FROM subeler ORDER BY ad")
        return [dict(r) for r in cur.fetchall()]


@app.post("/api/subeler/{sid}/sezon")
def sube_sezon_ayarla(sid: str, body: dict = None):
    """Şubeyi sezonluk KAPAT/AÇ (Köyceğiz/Alsancak sezon dışı). Kapalıyken canlı
    operasyon görünümünde ve atama dropdown'larında gizlenir. Şube kaydı silinmez."""
    kapali = bool((body or {}).get("sezon_kapali", True))
    with db() as (conn, cur):
        cur.execute("UPDATE subeler SET sezon_kapali=%s WHERE id=%s", (kapali, sid))
        if cur.rowcount == 0:
            raise HTTPException(404, "Şube bulunamadı")
    return {"success": True, "sube_id": sid, "sezon_kapali": kapali}


def _sube_katalog_stok_garantile(cur, sube_id: str) -> int:
    """Verilen şube için tüm aktif siparis_urun kayıtlarına karşılık sube_depo_stok satırı ekler.
    Fiziksel havuz ürünleri (su_adet, bardak vb.) atlanır — onlar text-kodlu satırlarla izlenir.
    Mevcut satırları değiştirmez (ON CONFLICT DO NOTHING). Eklenen satır sayısını döner."""
    _HAVUZ = (
        'bardak_kucuk', 'bardak_buyuk', 'bardak_plastik', 'su_adet',
        'redbull_adet', 'soda_adet', 'cookie_adet', 'pasta_adet',
        'sut_litre', 'surup_adet', 'kahve_paket', 'karton_bardak',
        'kapak_adet', 'pecete_paket', 'diger_sarf',
    )
    cur.execute("""
        INSERT INTO sube_depo_stok (id, sube_id, kalem_kodu, kalem_adi, mevcut_adet)
        SELECT gen_random_uuid()::text, %s, su.id::text, su.ad, 0
        FROM siparis_urun su
        WHERE su.aktif = TRUE
          AND (su.depo_stok_kalem_kodu IS NULL
               OR su.depo_stok_kalem_kodu NOT IN %s)
        ON CONFLICT (sube_id, kalem_kodu) DO NOTHING
    """, (sube_id, _HAVUZ))
    return cur.rowcount


@app.post("/api/subeler")
def sube_olustur(body: dict):
    """Yeni şube oluştur. Zorunlu: id (text), ad (text). İsteğe bağlı: adres."""
    sid = (body.get("id") or "").strip()
    ad  = (body.get("ad") or "").strip()
    adres = (body.get("adres") or "").strip() or None
    if not sid or not ad:
        raise HTTPException(400, "id ve ad zorunludur")
    import re
    if not re.match(r'^[a-z0-9_-]+$', sid):
        raise HTTPException(400, "id yalnızca küçük harf, rakam, _ veya - içerebilir")
    with db() as (conn, cur):
        cur.execute("SELECT id FROM subeler WHERE id=%s", (sid,))
        if cur.fetchone():
            raise HTTPException(409, f"'{sid}' id'li şube zaten mevcut")
        cur.execute(
            "INSERT INTO subeler (id, ad, adres) VALUES (%s, %s, %s)",
            (sid, ad, adres),
        )
        # 1-to-1: yeni şube için tüm aktif katalog ürünlerinin stok satırını garantile
        eklenen = _sube_katalog_stok_garantile(cur, sid)
    return {"success": True, "sube_id": sid, "stok_satirlari_eklendi": eklenen}


@app.put("/api/subeler/{sid}")
def sube_guncelle(sid: str, body: SubeGuncelleModel):
    pos_oran = float(body.pos_oran)
    online_oran = float(body.online_oran)
    if not (0 <= pos_oran <= 100) or not (0 <= online_oran <= 100):
        raise HTTPException(400, "Oran 0-100 arasında olmalı")
    if body.min_personel < 1:
        raise HTTPException(400, "Minimum personel en az 1 olmalı")
    if body.yogun_saat_ek_personel < 0:
        raise HTTPException(400, "Yoğun saat ek personel 0 veya daha büyük olmalı")
    sube_tipi = (body.sube_tipi or "").strip().lower() or None
    if sube_tipi in ("sevkiyat",):
        sube_tipi = "depo"
    elif sube_tipi in ("merkez",):
        sube_tipi = "karma"
    if sube_tipi is not None and sube_tipi not in ("normal", "depo", "karma"):
        raise HTTPException(400, "sube_tipi: normal | depo | karma")
    with db() as (conn, cur):
        cur.execute("SELECT id FROM subeler WHERE id=%s", (sid,))
        if not cur.fetchone():
            raise HTTPException(404, "Şube bulunamadı")
        cur.execute(
            """
            UPDATE subeler
            SET pos_oran=%s,
                online_oran=%s,
                acilis_saati=%s,
                kapanis_saati=%s,
                yogun_saat_baslangic=%s,
                yogun_saat_bitis=%s,
                ortusme_gerekli=%s,
                vardiya_yazilsin=%s,
                acilis_sadece_part=%s,
                kapanis_sadece_part=%s,
                min_personel=%s,
                yogun_saat_ek_personel=%s,
                acilis_max_kisi=%s,
                sube_tipi=COALESCE(%s, sube_tipi)
            WHERE id=%s
            """,
            (
                pos_oran,
                online_oran,
                body.acilis_saati,
                body.kapanis_saati,
                body.yogun_saat_baslangic,
                body.yogun_saat_bitis,
                bool(body.ortusme_gerekli),
                bool(body.vardiya_yazilsin),
                bool(body.acilis_sadece_part),
                bool(body.kapanis_sadece_part),
                int(body.min_personel),
                int(body.yogun_saat_ek_personel),
                body.acilis_max_kisi,
                sube_tipi,
                sid,
            ),
        )
    # Vardiya planı açık şubelerde: mesai saatleri kaydedilince AUTO slotları güncelle
    # (Pzt–Paz tüm günler — gün matrisi `aktif_gunler` ile aynı hizada kalır)
    if bool(body.vardiya_yazilsin):
        try:
            with db() as (conn, cur):
                sonuc = _vv2.slotlari_sube_saatlerinden_uret(
                    cur,
                    sid,
                    mod="yenile",
                    acilis_dakika=60,
                    kapanis_dakika=60,
                    normal_slot_dakika=120,
                    hafta_ici=False,
                    aktif_gunler=None,
                )
            if not sonuc.get("basarili"):
                logging.warning(
                    "Şube %s güncellendi; AUTO slot yenilenemedi: %s",
                    sid,
                    sonuc.get("mesaj"),
                )
        except Exception:
            logging.exception("Şube %s güncellendi; AUTO slot yenileme hatası", sid)
    return {"success": True}

@app.get("/api/subeler/{sid}/kasa-onizle")
def kasa_onizle(sid: str, baslangic: date, bitis: date = None):
    """
    Seçilen tarih aralığındaki ciro kayıtları için kasa düzeltme önizlemesi.
    Düzeltme yapmaz — sadece etki hesaplar.
    """
    bitis = bitis or bugun_tr()
    with db() as (conn, cur):
        cur.execute("SELECT * FROM subeler WHERE id=%s", (sid,))
        sube = cur.fetchone()
        if not sube:
            raise HTTPException(404, "Şube bulunamadı")
        pos_oran = float(sube['pos_oran'] or 0)
        online_oran = float(sube['online_oran'] or 0)

        cur.execute("""
            SELECT c.id, c.tarih, c.nakit, c.pos, c.online, c.toplam,
                   kh.tutar as kasa_tutar, kh.id as kasa_id
            FROM ciro c
            JOIN kasa_hareketleri kh ON kh.ref_id = c.id
                AND kh.ref_type IN ('CIRO', 'CIRO_GUNCELLEME')
                AND kh.islem_turu = 'CIRO'
                AND kh.durum = 'aktif'
            WHERE c.sube_id = %s AND c.durum = 'aktif'
            AND c.tarih BETWEEN %s AND %s
            ORDER BY c.tarih
        """, (sid, baslangic, bitis))
        kayitlar = cur.fetchall()

        satirlar = []
        toplam_fark = 0
        for k in kayitlar:
            dogru_tutar = float(k['nakit']) + float(k['pos']) * (1 - pos_oran/100) + float(k['online']) * (1 - online_oran/100)
            mevcut_tutar = float(k['kasa_tutar'])
            fark = dogru_tutar - mevcut_tutar
            if abs(fark) > 0.01:
                satirlar.append({
                    "ciro_id": k['id'],
                    "tarih": str(k['tarih']),
                    "nakit": float(k['nakit']),
                    "pos": float(k['pos']),
                    "online": float(k['online']),
                    "mevcut_kasa": mevcut_tutar,
                    "dogru_kasa": dogru_tutar,
                    "fark": fark
                })
                toplam_fark += fark

        return {
            "sube_adi": sube['ad'],
            "pos_oran": pos_oran,
            "online_oran": online_oran,
            "baslangic": str(baslangic),
            "bitis": str(bitis),
            "etkilenen_kayit": len(satirlar),
            "toplam_fark": toplam_fark,
            "satirlar": satirlar
        }

@app.post("/api/subeler/{sid}/kasa-duzelt")
def kasa_duzelt(sid: str, body: KasaDuzeltModel):
    """
    Onaylanan tarih aralığındaki kasa kayıtlarını düzeltir.
    Her kayıt için: eski kasa kaydı iptal edilir + doğru tutarla yeni kayıt yazılır.
    """
    baslangic = body.baslangic
    bitis = body.bitis or bugun_tr()

    with db() as (conn, cur):
        cur.execute("SELECT * FROM subeler WHERE id=%s", (sid,))
        sube = cur.fetchone()
        if not sube:
            raise HTTPException(404, "Şube bulunamadı")
        pos_oran = float(sube['pos_oran'] or 0)
        online_oran = float(sube['online_oran'] or 0)

        cur.execute("""
            SELECT c.id as ciro_id, c.tarih, c.nakit, c.pos, c.online,
                   kh.id as kasa_id, kh.tutar as kasa_tutar
            FROM ciro c
            JOIN kasa_hareketleri kh ON kh.ref_id = c.id
                AND kh.ref_type IN ('CIRO', 'CIRO_GUNCELLEME')
                AND kh.islem_turu = 'CIRO'
                AND kh.durum = 'aktif'
            WHERE c.sube_id = %s AND c.durum = 'aktif'
            AND c.tarih BETWEEN %s AND %s
        """, (sid, baslangic, bitis))
        kayitlar = cur.fetchall()

        duzeltilen = 0
        toplam_fark = 0

        for k in kayitlar:
            pos_tutari = float(k['pos'])
            online_tutari = float(k['online'])
            dogru_tutar = float(k['nakit']) + pos_tutari * (1 - pos_oran/100) + online_tutari * (1 - online_oran/100)
            mevcut_tutar = float(k['kasa_tutar'])
            fark = dogru_tutar - mevcut_tutar

            if abs(fark) < 0.01:
                continue

            # 1) Eski POS_KESINTI / ONLINE_KESINTI kayıtlarını iptal et
            cur.execute("""
                UPDATE kasa_hareketleri SET durum='iptal'
                WHERE ref_id = %s AND islem_turu = 'POS_KESINTI' AND durum='aktif'
            """, (k['ciro_id'],))

            # 3) FIX MN1 (2026-07-05): Eski CIRO kaydını UPDATE tutar ile EZMEK append-only
            # #5 ihlaliydi (ikinci düzeltmede eski değer kalıcı kaybolur). Artık eski kayıt
            # İPTAL edilir (tutar korunur, iz kalır) + yeni CIRO kaydı kanonik yoldan yazılır.
            # Net kasa etkisi AYNI (eski iptal → sadece yeni sayılır); davranış bit-bit korunur.
            cur.execute("UPDATE kasa_hareketleri SET durum='iptal' WHERE id=%s AND durum='aktif'",
                        (k['kasa_id'],))
            insert_kasa_hareketi(
                cur, k['tarih'], 'CIRO', dogru_tutar,
                f'POS/Online kesinti düzeltmesi (pos:%{pos_oran}, online:%{online_oran})',
                'ciro', k['ciro_id'], ref_id=str(uuid.uuid4()), ref_type='CIRO_DUZELTME',
            )

            # 4) Yeni POS_KESINTI kaydı yaz — paneldeki finansman maliyeti buradan hesaplanır
            # FIX MN3 (2026-07-06): ham INSERT idempotency'sizdi → düzeltme iki kez tetiklenirse
            # (çift tık/retry) aynı ciroya ÇİFT kesinti yazılıp kasa fazladan düşüyordu. Merkezi
            # yazıcıya (insert_kasa_hareketi) çevrildi: deterministik ref_id'den türeyen
            # idempotency anahtarı birebir tekrarı sessizce yutar (ON CONFLICT).
            pos_kesinti = pos_tutari * pos_oran / 100
            online_kesinti = online_tutari * online_oran / 100
            if pos_kesinti > 0.01:
                insert_kasa_hareketi(
                    cur, k['tarih'], 'POS_KESINTI', -pos_kesinti,
                    f'POS komisyon kesintisi (%{pos_oran})',
                    'ciro', k['ciro_id'] + '_pos',
                    ref_id=k['ciro_id'] + '_pos', ref_type='POS_KESINTI',
                )
            if online_kesinti > 0.01:
                insert_kasa_hareketi(
                    cur, k['tarih'], 'ONLINE_KESINTI', -online_kesinti,
                    f'Online komisyon kesintisi (%{online_oran})',
                    'ciro', k['ciro_id'] + '_online',
                    ref_id=k['ciro_id'] + '_online', ref_type='ONLINE_KESINTI',
                )

            audit(cur, 'kasa_hareketleri', k['kasa_id'], 'DUZELTME',
                  eski={'tutar': mevcut_tutar}, yeni={'tutar': dogru_tutar})

            duzeltilen += 1
            toplam_fark += fark

    return {"success": True, "duzeltilen": duzeltilen, "toplam_fark": toplam_fark}

# ── İŞLEM DEFTERİ (LEDGER) ─────────────────────────────────────
@app.get("/api/ledger")
def ledger(limit: int = 200, islem_turu: Optional[str] = None, ay: str = None):
    import re as _re_ay
    ay_v = (ay or "").strip()
    ay_aktif = bool(ay_v and ay_v.lower() != "hepsi" and _re_ay.match(r"^\d{4}-\d{2}$", ay_v))
    with db() as (conn, cur):
        sql = "SELECT * FROM kasa_hareketleri WHERE durum='aktif'"
        params: list = []
        if islem_turu:
            sql += " AND islem_turu=%s"
            params.append(islem_turu)
        if ay_aktif:
            sql += " AND to_char(tarih, 'YYYY-MM') = %s"
            params.append(ay_v)
        sql += " ORDER BY tarih DESC, olusturma DESC LIMIT %s"
        params.append(limit)
        cur.execute(sql, params)
        rows = [dict(r) for r in cur.fetchall()]

        # Ay verildiyse {rows, ozet} döner (kartlar aylık olsun); yoksa geriye uyum: düz dizi.
        if not ay_aktif:
            return rows
        cur.execute(
            """
            SELECT
                COALESCE(SUM(CASE WHEN tutar > 0 THEN tutar ELSE 0 END), 0)        AS toplam_gelir,
                COALESCE(SUM(CASE WHEN tutar < 0 THEN -tutar ELSE 0 END), 0)       AS toplam_gider,
                COALESCE(SUM(CASE WHEN islem_turu LIKE '%%IPTAL%%' THEN ABS(tutar) ELSE 0 END), 0) AS toplam_iptal
            FROM kasa_hareketleri
            WHERE durum='aktif' AND to_char(tarih, 'YYYY-MM') = %s
            """,
            (ay_v,),
        )
        oz = dict(cur.fetchone() or {})
        return {
            "rows": rows,
            "ozet": {
                "toplam_gelir": float(oz.get("toplam_gelir") or 0),
                "toplam_gider": float(oz.get("toplam_gider") or 0),
                "toplam_iptal": float(oz.get("toplam_iptal") or 0),
            },
        }

# ── EXCEL IMPORT ───────────────────────────────────────────────
from fastapi import UploadFile, File
import io

@app.post("/api/excel-import")
async def excel_import(dosya: UploadFile = File(...)):
    try:
        import openpyxl
        content = await dosya.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        
        detay = {}
        toplam = 0

        with db() as (conn, cur):
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2: continue
                
                headers = [str(h).strip().lower() if h else '' for h in rows[0]]
                eklenen = 0
                hata = 0
                atlanan = []
                satir_no = 1  # header=0, veri=1'den başlar

                for row in rows[1:]:
                    satir_no += 1
                    if all(v is None for v in row): continue
                    d = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}

                    # FIX MN4 (2026-07-06): her satır kendi SAVEPOINT'inde işlenir. Önceden tek
                    # transaction'dı → bir satır constraint ihlali yapınca DB "aborted" olup kalan
                    # satırlar toptan başarısız, commit'te 1-29 sessizce rollback ama "29 eklendi"
                    # YANLIŞ raporlanıyordu. Savepoint hatalı satırı izole eder; eklenen=gerçek commit.
                    cur.execute("SAVEPOINT sp_xls_row")
                    try:
                        # Tarih düzelt
                        def fix_date(v):
                            if v is None: return None
                            if hasattr(v, 'strftime'): return v.strftime('%Y-%m-%d')
                            return str(v)[:10]

                        sn = sheet_name.lower().strip()

                        if sn == 'ciro':
                            sube_id = 'sube-merkez'
                            cur.execute("SELECT id, COALESCE(pos_oran,0) as pos_oran, COALESCE(online_oran,0) as online_oran FROM subeler WHERE LOWER(ad)=LOWER(%s)", (str(d.get('sube','MERKEZ')),))
                            sube_row = cur.fetchone()
                            if sube_row:
                                sube_id     = sube_row['id']
                                pos_oran_x  = float(sube_row['pos_oran'])
                                online_oran_x = float(sube_row['online_oran'])
                            else:
                                cur.execute("SELECT COALESCE(pos_oran,0) as pos_oran, COALESCE(online_oran,0) as online_oran FROM subeler WHERE id='sube-merkez'")
                                merkez = cur.fetchone()
                                pos_oran_x    = float(merkez['pos_oran'])    if merkez else 0.0
                                online_oran_x = float(merkez['online_oran']) if merkez else 0.0
                            cid   = str(uuid.uuid4())
                            nakit = float(d.get('nakit')  or 0)
                            pos   = float(d.get('pos')    or 0)
                            online= float(d.get('online') or 0)
                            # Normal ciro girişiyle aynı prensip: komisyon düşülüp net kasaya
                            pos_kesinti_x    = pos    * pos_oran_x    / 100.0
                            online_kesinti_x = online * online_oran_x / 100.0
                            net_tutar_x = nakit + (pos - pos_kesinti_x) + (online - online_kesinti_x)
                            cur.execute("""INSERT INTO ciro (id,tarih,sube_id,nakit,pos,online,aciklama)
                                VALUES (%s,%s,%s,%s,%s,%s,%s) ON CONFLICT DO NOTHING""",
                                (cid, fix_date(d.get('tarih')), sube_id, nakit, pos, online, str(d.get('aciklama') or '')))
                            if cur.rowcount > 0:
                                insert_kasa_hareketi(cur, fix_date(d.get('tarih')), 'CIRO',
                                    net_tutar_x,
                                    f'Excel import (net) — pos:%{pos_oran_x} online:%{online_oran_x}',
                                    'ciro', cid, ref_id=cid, ref_type='CIRO')
                                eklenen += 1
                            else:
                                atlanan.append({"satir": satir_no, "sebep": "duplicate", "veri": f"{d.get('tarih')} / {d.get('sube','')}"})

                        elif sn == 'kartlar':
                            cur.execute("""INSERT INTO kartlar (id,kart_adi,banka,limit_tutar,kesim_gunu,son_odeme_gunu,faiz_orani)
                                VALUES (%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (kart_adi) DO NOTHING""",
                                (str(uuid.uuid4()), str(d.get('kart_adi','')).upper(),
                                 str(d.get('banka','')), float(d.get('limit_tutar') or 0),
                                 int(d.get('kesim_gunu') or 15), int(d.get('son_odeme_gunu') or 25),
                                 float(d.get('faiz_orani') or 0)))
                            if cur.rowcount > 0:
                                eklenen += 1
                            else:
                                atlanan.append({"satir": satir_no, "sebep": "duplicate", "veri": str(d.get('kart_adi',''))})

                        elif sn == 'kart_hareketleri':
                            kart_adi = str(d.get('kart_adi','')).upper()
                            cur.execute("SELECT id FROM kartlar WHERE UPPER(kart_adi)=%s", (kart_adi,))
                            k = cur.fetchone()
                            if not k:
                                # EVV-SIS (Codex): eskiden SESSİZ continue — satır ne
                                # hata ne atlanan sayacına giriyordu (görünmez kayıp).
                                # hata sayacı da artar (toast/KPI/iz defteri buna bakar).
                                atlanan.append({"satir": satir_no, "sebep": f"kart bulunamadı: {kart_adi[:40]}", "veri": kart_adi[:40]})
                                hata += 1
                                cur.execute("RELEASE SAVEPOINT sp_xls_row")
                                continue
                            islem = str(d.get('islem_turu','HARCAMA')).upper()
                            hid = str(uuid.uuid4())
                            cur.execute("""INSERT INTO kart_hareketleri (id,kart_id,tarih,islem_turu,tutar,taksit_sayisi,aciklama)
                                VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                                (hid, k['id'], fix_date(d.get('tarih')),
                                 islem, float(d.get('tutar') or 0),
                                 int(d.get('taksit_sayisi') or 1), str(d.get('aciklama') or '')))
                            # HARCAMA kasayı etkilemez
                            # ODEME -> onay kuyruğuna girer (kasadan düşmesi onay gerektirir)
                            if islem == 'ODEME':
                                cur.execute("""INSERT INTO onay_kuyrugu (id,islem_turu,kaynak_tablo,kaynak_id,aciklama,tutar,tarih)
                                    VALUES (%s,'KART_ODEME','kart_hareketleri',%s,'Excel import kart ödemesi',%s,%s)""",
                                    (str(uuid.uuid4()), hid, float(d.get('tutar') or 0), fix_date(d.get('tarih'))))
                            eklenen += 1

                        elif sn == 'borclar':
                            cur.execute("""INSERT INTO borc_envanteri (id,kurum,borc_turu,toplam_borc,aylik_taksit,kalan_vade,odeme_gunu)
                                VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                                (str(uuid.uuid4()), str(d.get('kurum','')),
                                 str(d.get('borc_turu','Kredi')),
                                 float(d.get('toplam_borc') or 0),
                                 float(d.get('aylik_taksit') or 0),
                                 int(d.get('kalan_vade') or 0),
                                 int(d.get('odeme_gunu') or 1)))
                            eklenen += 1

                        elif sn == 'personel':
                            sube_id = None
                            cur.execute("SELECT id FROM subeler WHERE LOWER(ad)=LOWER(%s)", (str(d.get('sube','MERKEZ')),))
                            r = cur.fetchone()
                            if r: sube_id = r['id']
                            cur.execute("""INSERT INTO personel (id,ad_soyad,gorev,calisma_turu,maas,yemek_ucreti,yol_ucreti,odeme_gunu,sube_id)
                                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                                (str(uuid.uuid4()), str(d.get('ad_soyad','')),
                                 str(d.get('gorev') or ''),
                                 str(d.get('calisma_turu','surekli')),
                                 float(d.get('maas') or 0),
                                 float(d.get('yemek_ucreti') or 0),
                                 float(d.get('yol_ucreti') or 0),
                                 int(d.get('odeme_gunu') or 1), sube_id))
                            eklenen += 1

                        elif sn == 'sabit_giderler':
                            sube_id = None
                            cur.execute("SELECT id FROM subeler WHERE LOWER(ad)=LOWER(%s)", (str(d.get('sube','MERKEZ')),))
                            r = cur.fetchone()
                            if r: sube_id = r['id']
                            cur.execute("""INSERT INTO sabit_giderler (id,gider_adi,kategori,tutar,tip,periyot,odeme_gunu,sube_id,odeme_yontemi)
                                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'nakit')""",
                                (str(uuid.uuid4()), str(d.get('gider_adi','')),
                                 str(d.get('kategori','Diğer')),
                                 float(d.get('tutar') or 0),
                                 str(d.get('tip','sabit')),
                                 str(d.get('periyot','aylik')),
                                 int(d.get('odeme_gunu') or 1), sube_id))
                            eklenen += 1

                        elif sn == 'vadeli_alimlar':
                            cur.execute("""INSERT INTO vadeli_alimlar (id,aciklama,tutar,vade_tarihi,tedarikci)
                                VALUES (%s,%s,%s,%s,%s)""",
                                (str(uuid.uuid4()), str(d.get('aciklama','')),
                                 float(d.get('tutar') or 0),
                                 fix_date(d.get('vade_tarihi')),
                                 str(d.get('tedarikci') or '')))
                            eklenen += 1

                        cur.execute("RELEASE SAVEPOINT sp_xls_row")
                    except Exception as ex:
                        cur.execute("ROLLBACK TO SAVEPOINT sp_xls_row")
                        cur.execute("RELEASE SAVEPOINT sp_xls_row")
                        hata += 1
                        atlanan.append({"satir": satir_no, "sebep": str(ex)[:100], "veri": str(list(d.values())[:3])})

                if eklenen > 0 or hata > 0 or atlanan:
                    detay[sheet_name] = {'eklenen': eklenen, 'hata': hata, 'atlanan': atlanan}
                    toplam += eklenen

        # DUYU 6/6 (2026-07-29): IMPORT İZ DEFTERİ — hata-yutar append-only yazıcı.
        # Kim/ne zaman/hangi dosya/kaç satır — "bu sayılar nereden geldi"nin izi.
        # Yazıcı hatası import'u ASLA engellemez (duyu kuralı).
        try:
            with db() as (conn, cur):
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS import_izi (
                        id         TEXT PRIMARY KEY DEFAULT gen_random_uuid()::text,
                        dosya_adi  TEXT,
                        kaynak     TEXT NOT NULL DEFAULT 'excel-import',
                        toplam_eklenen INT,
                        hata_sayisi    INT,
                        detay_json     TEXT,
                        olusturma  TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                )
                _hata_toplam = sum(int(v.get('hata') or 0) for v in detay.values())
                cur.execute(
                    "INSERT INTO import_izi (dosya_adi, kaynak, toplam_eklenen, hata_sayisi, detay_json) "
                    "VALUES (%s,%s,%s,%s,%s)",
                    (str(getattr(dosya, 'filename', '') or '')[:200], 'excel-import',
                     int(toplam), _hata_toplam,
                     json.dumps({k: {'eklenen': v.get('eklenen'), 'hata': v.get('hata')}
                                 for k, v in detay.items()}, ensure_ascii=False)[:2000]),
                )
        except Exception:
            pass  # iz yazıcı sessiz düşer — import sonucu etkilenmez

        return {"success": True, "toplam": toplam, "detay": detay}
    except ImportError:
        raise HTTPException(500, "openpyxl kurulu değil")
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/import-izi")
def import_izi_liste(limit: int = 50):
    """SALT-OKUR: import iz defteri — son yüklemeler (dosya, eklenen, hata, zaman)."""
    lim = max(1, min(200, int(limit or 50)))
    with db() as (conn, cur):
        cur.execute("SELECT to_regclass('public.import_izi') AS t")
        if not (cur.fetchone() or {}).get("t"):
            return {"kayitlar": [], "toplam": 0, "not": "Henüz iz yok — ilk yüklemeyle başlar."}
        cur.execute(
            "SELECT dosya_adi, kaynak, toplam_eklenen, hata_sayisi, detay_json, olusturma "
            "FROM import_izi ORDER BY olusturma DESC LIMIT %s",
            (lim,),
        )
        out = []
        for r in cur.fetchall() or []:
            d = dict(r)
            d["olusturma"] = str(d.get("olusturma") or "")[:16]
            try:
                d["detay"] = json.loads(d.pop("detay_json") or "{}")
            except Exception:
                d["detay"] = {}
            out.append(d)
    return {"kayitlar": out, "toplam": len(out)}


# ── ÇIFT KAYIT KONTROL ENDPOINTLERİ ───────────────────────────

@app.get("/api/ciro/kontrol")
def ciro_kontrol(tarih: str, tutar: float, sube_id: str = None):
    with db() as (conn, cur):
        cur.execute("""
            SELECT id, tarih, nakit+pos+online as toplam, sube_id FROM ciro
            WHERE durum='aktif'
            AND tarih = %s
            AND ABS((nakit+pos+online) - %s) < 1
            AND (%s IS NULL OR sube_id = %s)
        """, (tarih, tutar, sube_id, sube_id))
        benzer = [dict(r) for r in cur.fetchall()]
        return {"benzer": benzer, "var": len(benzer) > 0}

@app.get("/api/ciro/eksik-gunler")
def ciro_eksik_gunler(gun: int = 14):
    """📅 HANGİ ŞUBE HANGİ GÜN CİRO GİRMEMİŞ?

    Sahip (2026-08-09): "Panel bir şubenin cirosunu 8 Ağustos için gösteriyor,
    neden?" Cevap: en son girilen ciro oydu — 8 Ağustos'ta yalnız ZAFER
    girmiş, TEMA girmemişti. Panel doğruydu, EKSİK OLAN VERİYDİ.

    Sistem bu eksikliği hiçbir yerde SÖYLEMİYORDU: ciro girilmeyince kayıt
    yok, kayıt yoksa alarm da yok — "yokluğun alarmı" kurulmamıştı. Bu uç
    tam da onu ölçer: şube açıksa ve sezon kapalı değilse, o gün ciro
    BEKLENİR; yoksa eksiktir.

    ⚠️ Sezon kapalı ve pasif şubeler sayılmaz (hayalet eksik üretmesin).
    ⚠️ BUGÜN ayrı raporlanır — gün bitmediği için "eksik" demek erken.
    """
    bugun = bugun_tr()
    bas = bugun - timedelta(days=max(1, min(gun, 90)))
    with db() as (conn, cur):
        cur.execute("""SELECT id::text AS id, ad FROM subeler
                        WHERE COALESCE(aktif,TRUE)=TRUE
                          AND COALESCE(sezon_kapali,FALSE)=FALSE
                          AND id <> 'sube-merkez'
                        ORDER BY ad""")
        subeler = [dict(r) for r in (cur.fetchall() or [])]
        cur.execute("""SELECT sube_id::text AS sid, tarih::text AS tarih
                         FROM ciro
                        WHERE COALESCE(durum,'aktif')='aktif' AND tarih >= %s""",
                    (bas,))
        var = {(r["sid"], r["tarih"]) for r in (cur.fetchall() or [])}
    eksikler, bugun_bekleyen = [], []
    for i in range((bugun - bas).days + 1):
        t = bas + timedelta(days=i)
        ts = t.isoformat()
        for s in subeler:
            if (s["id"], ts) in var:
                continue
            kayit = {"tarih": ts, "sube_id": s["id"], "sube_adi": s["ad"],
                     "gun_once": (bugun - t).days}
            (bugun_bekleyen if t == bugun else eksikler).append(kayit)
    eksikler.sort(key=lambda x: (x["tarih"], x["sube_adi"]), reverse=True)
    _sube_ozet: Dict[str, int] = {}
    for e in eksikler:
        _sube_ozet[e["sube_adi"]] = _sube_ozet.get(e["sube_adi"], 0) + 1
    return {
        "pencere_gun": gun, "bugun": bugun.isoformat(),
        "aktif_sube": [s["ad"] for s in subeler],
        # ⚠️ BİRİM AYRIMI (2026-08-26) — "4 gün mü, 2 gün mü?"
        # `eksik_adet` ŞUBE×GÜN sayar (ZAFER 24 Ağu + TEMA 24 Ağu = 2 kayıt).
        # Panel bunu "N gün" diye yazıyordu; takvimde ise 2 gün eksikti ve
        # kardeş ekran (BAKIŞ, motors.ciro_eksik_gunler) GÜN sayıyordu.
        # Aynı sabah aynı sahibe iki farklı sayı: 4 ve 2. Sahip iki ekrana
        # birden güvenemezse hiçbirine güvenmez.
        # Çözüm sayıyı değiştirmek DEĞİL, BİRİMİ ADLANDIRMAK: ikisi de doğru,
        # ikisi de farklı şeyi ölçüyor. Artık ikisi de sunucudan gelir ve
        # ekran hangisini yazdığını söyleyebilir.
        "eksik_adet": len(eksikler),                                  # şube×gün
        "eksik_gun_adet": len({e["tarih"] for e in eksikler}),        # takvim günü
        "eksik_sube_ozet": _sube_ozet,
        "eksikler": eksikler[:60],
        # Bugün henüz bitmedi — eksik DEĞİL, "bekleniyor"
        "bugun_bekleyen": bugun_bekleyen,
        "not": "Sezon kapalı ve pasif şubeler sayılmaz. Bugün 'eksik' değil "
               "'bekleniyor' sayılır — gün kapanmadan ciro beklenmez.",
    }


@app.get("/api/anlik-gider/kontrol")
def anlik_gider_kontrol(tarih: str, tutar: float, kategori: str = None):
    with db() as (conn, cur):
        cur.execute("""
            SELECT id, tarih, tutar, kategori FROM anlik_giderler
            WHERE durum='aktif'
            AND tarih BETWEEN %s::date - INTERVAL '7 days' AND %s::date + INTERVAL '7 days'
            AND ABS(tutar - %s) < 1
            AND (%s IS NULL OR kategori = %s)
        """, (tarih, tarih, tutar, kategori, kategori))
        benzer = [dict(r) for r in cur.fetchall()]
        return {"benzer": benzer, "var": len(benzer) > 0}

@app.get("/api/dis-kaynak/kontrol")
def dis_kaynak_kontrol(tarih: str, tutar: float, kategori: str = None):
    with db() as (conn, cur):
        cur.execute("""
            SELECT id, tarih, tutar, aciklama FROM kasa_hareketleri
            WHERE islem_turu='DIS_KAYNAK' AND durum='aktif'
            AND tarih BETWEEN %s::date - INTERVAL '7 days' AND %s::date + INTERVAL '7 days'
            AND ABS(tutar - %s) < 1
            AND (%s IS NULL OR aciklama LIKE %s)
        """, (tarih, tarih, tutar, kategori, f"{kategori}%"))
        benzer = [dict(r) for r in cur.fetchall()]
        return {"benzer": benzer, "var": len(benzer) > 0}

@app.get("/api/vadeli-panel-detay")
def vadeli_panel_detay():
    """
    Panel Vadeli Borç kartına tıklanınca açılan detay.
    SADECE kaynak_id ile çalışır — aciklama eşleşmesi yok, risk yok.
    """
    with db() as (conn, cur):
        yil = bugun_tr().year
        ay = bugun_tr().month

        # Bu ay ödeme yapılan vadeli alımlar — SADECE kaynak_id ile
        cur.execute("""
            SELECT DISTINCT
                va.id, va.aciklama, va.tutar, va.vade_tarihi,
                va.tedarikci, va.durum,
                (va.vade_tarihi - CURRENT_DATE) as gun_kaldi
            FROM vadeli_alimlar va
            WHERE (
                EXISTS (
                    SELECT 1 FROM kasa_hareketleri kh
                    WHERE kh.kaynak_id = va.id::text
                    AND kh.islem_turu = 'VADELI_ODEME'
                    AND kh.kasa_etkisi = true AND kh.durum = 'aktif'
                    AND EXTRACT(YEAR FROM kh.tarih) = %s
                    AND EXTRACT(MONTH FROM kh.tarih) = %s
                )
                OR EXISTS (
                    SELECT 1 FROM kart_hareketleri kh
                    WHERE kh.kaynak_id = va.id::text
                    AND kh.kaynak_tablo = 'vadeli_alimlar'
                    AND kh.islem_turu = 'HARCAMA' AND kh.durum = 'aktif'
                    AND EXTRACT(YEAR FROM kh.tarih) = %s
                    AND EXTRACT(MONTH FROM kh.tarih) = %s
                )
            )
            ORDER BY va.vade_tarihi DESC
        """, (yil, ay, yil, ay))
        vadeli_liste = cur.fetchall()

        sonuc = []
        for v in vadeli_liste:
            vid = str(v['id'])

            # Nakit ödemeler — kaynak_id ile
            cur.execute("""
                SELECT ABS(kh.tutar) as tutar, kh.tarih,
                    'nakit' as yontem, kh.aciklama,
                    NULL as banka, NULL as kart_adi
                FROM kasa_hareketleri kh
                WHERE kh.kaynak_id = %s
                AND kh.islem_turu = 'VADELI_ODEME'
                AND kh.kasa_etkisi = true AND kh.durum = 'aktif'
                AND EXTRACT(YEAR FROM kh.tarih) = %s
                AND EXTRACT(MONTH FROM kh.tarih) = %s
                ORDER BY kh.tarih DESC
            """, (vid, yil, ay))
            odemeler = [dict(r) for r in cur.fetchall()]

            # Kart ödemeleri — SADECE kaynak_id ile
            cur.execute("""
                SELECT kh.tutar, kh.tarih, 'kart' as yontem,
                    kh.aciklama, k.banka, k.kart_adi
                FROM kart_hareketleri kh
                JOIN kartlar k ON k.id = kh.kart_id
                WHERE kh.kaynak_id = %s
                AND kh.kaynak_tablo = 'vadeli_alimlar'
                AND kh.islem_turu = 'HARCAMA' AND kh.durum = 'aktif'
                AND EXTRACT(YEAR FROM kh.tarih) = %s
                AND EXTRACT(MONTH FROM kh.tarih) = %s
                ORDER BY kh.tarih DESC
            """, (vid, yil, ay))
            odemeler += [dict(r) for r in cur.fetchall()]

            nakit_toplam = sum(float(o['tutar']) for o in odemeler if o['yontem'] == 'nakit')
            kart_toplam  = sum(float(o['tutar']) for o in odemeler if o['yontem'] == 'kart')

            sonuc.append({
                'id': vid,
                'aciklama': v['aciklama'],
                'tutar': float(v['tutar']),
                'vade_tarihi': str(v['vade_tarihi']),
                'tedarikci': v['tedarikci'],
                'durum': v['durum'],
                'gun_kaldi': int(v['gun_kaldi']) if v['gun_kaldi'] is not None else None,
                'nakit_odenen': nakit_toplam,
                'kart_odenen': kart_toplam,
                'toplam_odenen': nakit_toplam + kart_toplam,
                'odemeler': odemeler,
            })

        return sonuc

@app.get("/api/vadeli-odeme-detay")
def vadeli_odeme_detay(kaynak: str = 'kart'):
    """
    Panel kart kırılımı detay — 💳 Kart tıklanınca açılır.
    kaynak='kart' → bu ay kartla yapılan vadeli ödemeleri listeler.
    kaynak='nakit' → bu ay nakitle yapılan vadeli ödemeleri listeler.
    """
    with db() as (conn, cur):
        bugun = bugun_tr()
        yil, ay = bugun.year, bugun.month

        if kaynak == 'kart':
            cur.execute("""
                SELECT
                    kh.tarih,
                    kh.tutar,
                    kh.aciklama,
                    k.banka,
                    k.kart_adi
                FROM kart_hareketleri kh
                JOIN kartlar k ON k.id = kh.kart_id
                WHERE kh.islem_turu = 'HARCAMA'
                AND kh.durum = 'aktif'
                AND kh.kaynak_tablo = 'vadeli_alimlar'
                AND EXTRACT(YEAR FROM kh.tarih) = %s
                AND EXTRACT(MONTH FROM kh.tarih) = %s
                ORDER BY kh.tarih DESC
            """, (yil, ay))
        else:
            cur.execute("""
                SELECT
                    kh.tarih,
                    ABS(kh.tutar) as tutar,
                    kh.aciklama,
                    NULL as banka,
                    NULL as kart_adi
                FROM kasa_hareketleri kh
                WHERE kh.islem_turu = 'VADELI_ODEME'
                AND kh.kasa_etkisi = true
                AND kh.durum = 'aktif'
                AND EXTRACT(YEAR FROM kh.tarih) = %s
                AND EXTRACT(MONTH FROM kh.tarih) = %s
                ORDER BY kh.tarih DESC
            """, (yil, ay))

        return [dict(r) for r in cur.fetchall()]

@app.get("/api/vadeli-alimlar/ozet")
def vadeli_ozet():
    """Vadeli alımlar özet — Panel kartı için. Nakit + kart dahil."""
    with db() as (conn, cur):
        # Bu ay nakit ödenen (kasa_hareketleri)
        cur.execute("""
            SELECT COALESCE(SUM(ABS(kh.tutar)), 0) as nakit
            FROM kasa_hareketleri kh
            WHERE kh.islem_turu = 'VADELI_ODEME'
            AND kh.kasa_etkisi = true AND kh.durum = 'aktif'
            AND EXTRACT(YEAR FROM kh.tarih) = EXTRACT(YEAR FROM CURRENT_DATE)
            AND EXTRACT(MONTH FROM kh.tarih) = EXTRACT(MONTH FROM CURRENT_DATE)
        """)
        nakit_odenen = float(cur.fetchone()['nakit'])

        # Bu ay kartla ödenen (kart_hareketleri — tam ve kısmi)
        cur.execute("""
            SELECT COALESCE(SUM(kh.tutar), 0) as kart
            FROM kart_hareketleri kh
            WHERE kh.islem_turu = 'HARCAMA' AND kh.durum = 'aktif'
            AND kh.kaynak_tablo = 'vadeli_alimlar'
            AND EXTRACT(YEAR FROM kh.tarih) = EXTRACT(YEAR FROM CURRENT_DATE)
            AND EXTRACT(MONTH FROM kh.tarih) = EXTRACT(MONTH FROM CURRENT_DATE)
        """)
        kart_odenen = float(cur.fetchone()['kart'])

        toplam_odenen = nakit_odenen + kart_odenen

        # Bekleyen vadeli alımlar
        cur.execute("""
            SELECT COALESCE(SUM(tutar), 0) as toplam_bekleyen, COUNT(*) as adet
            FROM vadeli_alimlar WHERE durum = 'bekliyor'
        """)
        row = cur.fetchone()
        toplam_bekleyen = float(row['toplam_bekleyen'])
        bekleyen_adet = int(row['adet'])

        # Geciken vadeli alımlar (sistem başlangıcı: Haziran öncesi sayılmaz)
        cur.execute("""
            SELECT COUNT(*) as adet FROM vadeli_alimlar
            WHERE durum = 'bekliyor' AND vade_tarihi < CURRENT_DATE
              AND vade_tarihi >= DATE '2026-06-01'
        """)
        geciken_adet = int(cur.fetchone()['adet'])

        return {
            "toplam_odenen": toplam_odenen,
            "toplam_bekleyen": toplam_bekleyen,
            "bekleyen_adet": bekleyen_adet,
            "geciken_adet": geciken_adet
        }

@app.get("/api/vadeli-alimlar/kontrol")
def vadeli_kontrol(vade_tarihi: str, tutar: float):
    with db() as (conn, cur):
        cur.execute("""
            SELECT id, aciklama, tutar, vade_tarihi FROM vadeli_alimlar
            WHERE durum='bekliyor'
            AND vade_tarihi BETWEEN %s::date - INTERVAL '7 days' AND %s::date + INTERVAL '7 days'
            AND ABS(tutar - %s) < 1
        """, (vade_tarihi, vade_tarihi, tutar))
        benzer = [dict(r) for r in cur.fetchall()]
        return {"benzer": benzer, "var": len(benzer) > 0}


@app.post("/api/odeme-plani/{oid}/odendi")
def odeme_odendi(oid: str, manuel_tutar: Optional[float] = None):
    """Geriye dönük uyumluluk — /ode endpoint'ine yönlendirir."""
    return odeme_yap(oid, tutar=manuel_tutar)


@app.post("/api/odeme-plani/{oid}/ertele")
def odeme_ertele(oid: str, yeni_tarih: Optional[date] = None):
    """Ödemeyi ertele — sadece tarih güncellenir, yeni kayıt açılmaz."""
    with db() as (conn, cur):
        cur.execute(
            "SELECT * FROM odeme_plani WHERE id=%s AND durum IN ('bekliyor','onay_bekliyor')",
            (oid,),
        )
        o = cur.fetchone()
        if not o: raise HTTPException(404)
        mevcut = o["tarih"]
        yeni = yeni_tarih or (mevcut + timedelta(days=4))

        # KURAL (kullanıcı, 2026-07-03): vade HER YÖNE değiştirilebilir (öne çekme dahil).
        # Tek sınır: bugünden geriye atılamaz (defter/simülasyon bozulmasın).
        # Eski "≤7 güne zorla +4" dayatması KALDIRILDI (FAZ0 bulgu #7 çözüldü) —
        # kullanıcının seçtiği tarih aynen uygulanır.
        if yeni < bugun_tr():
            raise HTTPException(400, "Vade bugünden geriye alınamaz")
        if yeni == mevcut:
            raise HTTPException(400, "Tarih değişmedi — mevcut vade zaten bu")
        # Ödeme planı tarihini güncelle
        cur.execute("UPDATE odeme_plani SET tarih=%s WHERE id=%s", (yeni, oid))
        # Onay kuyruğundaki tarihi de güncelle — PROD-PANEL-002 ailesi: kaynak_tablo ile skopla
        # (SADECE kaynak_id eşleşmesi tablo-arası id çakışmasında yanlış onayın tarihini kaydırıyordu)
        if o.get('kaynak_tablo') and o.get('kaynak_id'):
            cur.execute("""
                UPDATE onay_kuyrugu SET tarih=%s
                WHERE durum='bekliyor'
                  AND ( (kaynak_tablo='odeme_plani' AND kaynak_id=%s)
                     OR (kaynak_tablo=%s AND kaynak_id=%s) )
            """, (yeni, oid, o['kaynak_tablo'], o['kaynak_id']))
        else:
            cur.execute("""
                UPDATE onay_kuyrugu SET tarih=%s
                WHERE durum='bekliyor' AND kaynak_tablo='odeme_plani' AND kaynak_id=%s
            """, (yeni, oid))
        audit(cur, 'odeme_plani', oid, 'ERTELE')
        # Uyarı önbelleğini temizle — erteleme sonrası uyarı gizlensin
        uyari_cache_clear()
    return {"success": True, "yeni_tarih": str(yeni)}


@app.post("/api/odeme-plani/{oid}/kismi-ode")
def kismi_odeme_yap(oid: str, body: KismiOdeModel):
    """
    Kısmi ödeme — plan bölünür:
    1. Ödenen kısım: nakitte kasadan (VADELI_ODEME), kartta kart harcamasına yazılır.
    2. vadeli_alimlar.tutar = kalan borç; yeni vade ile tek satırda devam eder.
    3. Kalan için yeni odeme_plani + gerekirse VADELI_ODEME onayı.
    Sonradan yeni mal için ayrı satır yerine POST /vadeli-alimlar + birlestir_vadeli_id ile toplam borç artırılabilir.
    """
    with db() as (conn, cur):
        # FOR UPDATE: eş zamanlı çift kısmi ödeme isteğini engeller
        cur.execute("SELECT * FROM odeme_plani WHERE id=%s AND durum IN ('bekliyor','onay_bekliyor') FOR UPDATE", (oid,))
        plan = cur.fetchone()
        if not plan: raise HTTPException(404, "Plan bulunamadı veya zaten ödendi")

        # FAZ 0 #3: personel maaşı onaysız (kısmi de olsa) ödenemez
        _personel_maas_odeme_guard(cur, dict(plan))

        # 🔴 P0 (2026-08-13, EVV-ODE denetimi): "toplam" ORİJİNAL odenecek_tutar
        # idi — birikmiş odenen_tutar düşülmüyordu. Kısmi ödenmiş planda (300/1000)
        # ikinci kısmi 600 ödenince kalan 1000−600=400 hesaplanıyor, önceki 300
        # defterden düşüyordu (gerçek kalan 100). Referans artık KALAN borçtur.
        _mevcut_odenen = float(plan.get('odenen_tutar') or 0)
        toplam = round(float(plan['odenecek_tutar'] or 0) - _mevcut_odenen, 2)
        odenen = body.odenen_tutar
        kaynak = plan.get('kaynak_tablo') or ''
        if toplam <= 0:
            raise HTTPException(400, "Bu planın kalan borcu yok — zaten ödenmiş görünüyor")

        if odenen <= 0:
            raise HTTPException(400, "Ödenen tutar sıfırdan büyük olmalı")
        if odenen > toplam:
            if kaynak == 'vadeli_alimlar':
                raise HTTPException(400, "Odenen tutar kalan borctan buyuk olamaz")
            raise HTTPException(400, "Tam ödeme için normal ödeme ekranını kullanın")
        bugun = str(bugun_tr())

        # 🔴 P1 (2026-08-13, EVV-ODE denetimi): /ode kart yolunda FOR UPDATE +
        # limit kontrolü var, kismi-ode kart yolunda YOKTU — limiti aşan kısmi
        # ödeme sessizce kart borcuna yazılıyordu. Simetri kuruldu.
        if getattr(body, 'odeme_yontemi', 'nakit') == 'kart' and getattr(body, 'kart_id', None):
            cur.execute("SELECT * FROM kartlar WHERE id=%s AND aktif=TRUE FOR UPDATE", (body.kart_id,))
            _kk = cur.fetchone()
            if not _kk:
                raise HTTPException(404, "Kart bulunamadı")
            _kb = kart_borc(cur, body.kart_id)
            _klim = _kanonik_kalan_limit(body.kart_id, float(_kk['limit_tutar']) - _kb)
            if _klim < odenen:
                raise HTTPException(400, f"Kart limiti yetersiz. Kalan: {_klim:,.0f} ₺")
            # Kart yazımı yalnız vadeli_alimlar kaynağında var; diğer kaynaklarda
            # eski kod "kart" seçilse bile parayı SESSİZCE kasadan çıkarıyordu.
            if kaynak != 'vadeli_alimlar':
                raise HTTPException(400,
                    "Kısmi KART ödemesi yalnız vadeli alım kalemlerinde desteklenir — "
                    "bu kalemde nakit/havale kullanın ya da tam ödeme yapın")

        if odenen >= toplam:
            if kaynak == 'vadeli_alimlar' and plan.get('kaynak_id'):
                vid = plan['kaynak_id']
                _vadeli_diger_aktif_planlari_iptal(cur, vid, oid)

                if getattr(body, 'odeme_yontemi', 'nakit') == 'kart' and getattr(body, 'kart_id', None):
                    hid = str(uuid.uuid4())
                    cur.execute("""
                        INSERT INTO kart_hareketleri
                            (id, kart_id, tarih, islem_turu, tutar, taksit_sayisi, aciklama, kaynak_id, kaynak_tablo)
                        VALUES (%s, %s, %s, 'HARCAMA', %s, 1, %s, %s, 'vadeli_alimlar')
                    """, (
                        hid,
                        body.kart_id,
                        bugun,
                        toplam,
                        f"Vadeli alÄ±m kapanÄ±ÅŸ: {plan['aciklama']} ({int(toplam):,} â‚º)",
                        vid,
                    ))
                    audit(cur, 'kart_hareketleri', hid, 'VADELI_KART_KAPANIS')
                    kart_plan_guncelle_tx(cur)
                else:
                    insert_kasa_hareketi(
                        cur,
                        bugun,
                        'VADELI_ODEME',
                        -abs(toplam),
                        f"Vadeli alÄ±m kapanÄ±ÅŸ: {plan['aciklama']}",
                        'vadeli_alimlar',
                        vid,
                        oid,
                        'KISMI_ODE_KAPANIS',
                    )

                cur.execute(
                    "UPDATE odeme_plani SET durum='odendi', odeme_tarihi=%s, odenen_tutar=%s WHERE id=%s",
                    (bugun, round(_mevcut_odenen + toplam, 2), oid),  # P0 fix: birikimli
                )
                cur.execute("""
                    UPDATE onay_kuyrugu SET durum='onaylandi', onay_tarihi=NOW()
                    WHERE durum NOT IN ('onaylandi','reddedildi')
                    AND (kaynak_id=%s OR kaynak_id=%s)
                """, (oid, vid))
                vadeli_alim_kapat(cur, vid, bugun)
                audit(cur, 'odeme_plani', oid, 'KISMI_ODE_KAPANIS',
                      eski={'tutar': toplam}, yeni={'odenen': toplam, 'kalan': 0})
                uyari_cache_clear()
                return {"success": True, "odenen": toplam, "kalan": 0, "yeni_plan_id": None, "kapandi": True}

            if kaynak != 'vadeli_alimlar':
                raise HTTPException(400, "Tam ödeme için normal ödeme ekranını kullanın")
            raise HTTPException(400, "Odenen tutar kalan borctan buyuk olamaz")

        kalan = toplam - odenen
        if kalan <= 0:
            raise HTTPException(400, "Kalan tutar hesaplanamadı")

        bugun = str(bugun_tr())

        # Kart kesim ödemesi (kaynak_tablo boş + kart_id dolu) ise plan kapanmaz —
        # kullanıcı asgari tutara ulaşana kadar panel hatırlatıcısı devam eder.
        is_kart_plan = (not plan.get('kaynak_tablo')) and plan.get('kart_id')

        if is_kart_plan:
            # Birikimli ödenen: eski odenen_tutar + bu ödeme
            mevcut_odenen = float(plan.get('odenen_tutar') or 0)
            yeni_total_odenen = mevcut_odenen + odenen
            asgari = float(plan.get('asgari_tutar') or 0)
            # Asgari karşılandıysa plan kapanır; değilse bekliyor kalır (uyarı devam eder).
            asgari_tamam = (asgari > 0 and yeni_total_odenen >= asgari * 0.999)
            if asgari_tamam:
                cur.execute(
                    "UPDATE odeme_plani SET durum='odendi', odeme_tarihi=%s, odenen_tutar=%s WHERE id=%s",
                    (bugun, yeni_total_odenen, oid)
                )
            else:
                cur.execute(
                    "UPDATE odeme_plani SET odenen_tutar=%s WHERE id=%s",
                    (yeni_total_odenen, oid)
                )
        else:
            # Diğer kaynaklar (sabit_gider, personel, vadeli, borc_envanteri) — eski davranış:
            # Eski planı odendi yap, kalan için ayrı plan oluşturulur.
            # P0 fix: odenen_tutar BİRİKİMLİ yazılır (önceki kısmi ödemeler kaybolmasın).
            cur.execute("UPDATE odeme_plani SET durum='odendi', odeme_tarihi=%s, odenen_tutar=%s WHERE id=%s",
                (bugun, round(_mevcut_odenen + odenen, 2), oid))

        # FIX K1 (2026-07-06): ödeme-olayı sıra kimliği — plan + kümülatif kuruş, deterministik.
        # Kart kesim planı birden çok kısmi ödemeye açık kalır; eskiden kart ODEME satırı sabit
        # id=f"kodm_{oid}" + ON CONFLICT ile yazıldığından 2. ödeme kart borcunu DÜŞÜRMÜYORDU
        # (nakit kasadan yine çıkıyordu → borç kalıcı şişik). Kasa idempotency anahtarı da
        # ref_id=oid+tarih+tutar'dan türediği için aynı gün aynı tutarlı 2. ödeme kasada sessizce
        # yutulup plan sayacı yine artıyordu. Artık İKİ defter de bu sıra kimliğinden türer:
        # meşru yeni ödeme → kümülatif değişir → ikisi de yazar; commit-öncesi kopan birebir
        # retry → kümülatif aynı → ikisi de idempotent atlar. Defterler hep senkron.
        _odeme_seq = int(round((yeni_total_odenen if is_kart_plan else odenen) * 100))

        # Eski plana ait TUM acik onaylari kapat
        # Hem plan_id hem de kaynağın id'si (sabit_gider, personel vb.) ile ara
        _kaynak_id = plan.get('kaynak_id') or oid
        cur.execute("""
            UPDATE onay_kuyrugu SET durum='onaylandi', onay_tarihi=NOW()
            WHERE durum NOT IN ('onaylandi','reddedildi')
            AND (kaynak_id=%s OR kaynak_id=%s)
        """, (oid, _kaynak_id))

        # 2. Kasaya sadece ödenen kadar yaz (nakit) VEYA kart harcamasına ekle (kart)
        kaynak = plan.get('kaynak_tablo') or ''
        islem_t = kasa_islem_turu(kaynak)   # iptal tarafiyla AYNI kaynak

        if kaynak == 'vadeli_alimlar' and getattr(body, 'odeme_yontemi', 'nakit') == 'kart' and getattr(body, 'kart_id', None):
            # KART: kasaya yazma — kart harcamasına ekle
            hid = str(uuid.uuid4())
            cur.execute("""
                INSERT INTO kart_hareketleri
                    (id, kart_id, tarih, islem_turu, tutar, taksit_sayisi, aciklama, kaynak_id, kaynak_tablo)
                VALUES (%s, %s, %s, 'HARCAMA', %s, 1, %s, %s, 'vadeli_alimlar')
            """, (hid, body.kart_id, bugun, odenen,
                  f"Kısmi vadeli alım: {plan['aciklama']} ({int(odenen):,} / {int(toplam):,} ₺)",
                  plan.get('kaynak_id')))
            audit(cur, 'kart_hareketleri', hid, 'VADELI_KART_KISMI')
        else:
            # NAKİT: kasaya yaz — vadeli için kaynak vadeli_alimlar (tam ödeme / onay guard ile aynı anahtar)
            kasa_kt = (
                "vadeli_alimlar"
                if kaynak == "vadeli_alimlar" and plan.get("kaynak_id")
                else "odeme_plani"
            )
            kasa_kid = (
                plan["kaynak_id"]
                if kasa_kt == "vadeli_alimlar"
                else oid
            )
            insert_kasa_hareketi(
                cur,
                bugun,
                islem_t,
                -abs(odenen),
                f"Kısmi ödeme: {plan['aciklama']} ({int(odenen):,} / {int(toplam):,} ₺)",
                kasa_kt,
                kasa_kid,
                oid,
                "KISMI_ODE",
                # FIX K1: dedup anahtarı ödeme-olayı bazlı — aynı plana aynı gün aynı tutarlı
                # 2. MEŞRU ödeme artık sessizce yutulmaz (kümülatif seq farklı), birebir retry
                # ise yine idempotent (aynı seq → aynı anahtar).
                idempotency_key=f"kismi|{oid}|{_odeme_seq}",
            )

        # Kaynak vadeli_alimlar ise tutarı ve vadeyi güncelle (kapatma — kalan borç devam ediyor)
        if kaynak == 'vadeli_alimlar' and plan.get('kaynak_id'):
            cur.execute("""
                UPDATE vadeli_alimlar SET tutar=%s, vade_tarihi=%s WHERE id=%s
            """, (kalan, body.kalan_vade_tarihi, plan['kaynak_id']))
            # Eski onay kuyruğundaki bekleyen VADELI_ODEME kayıtlarını kapat — yenisi açılacak
            cur.execute("""
                UPDATE onay_kuyrugu SET durum='reddedildi', onay_tarihi=NOW()
                WHERE kaynak_tablo='vadeli_alimlar' AND kaynak_id=%s
                AND islem_turu='VADELI_ODEME' AND durum='bekliyor'
            """, (plan['kaynak_id'],))

        # Kaynak borc_envanteri ise: toplam_borc'u ÖDENEN kadar düş. Kısmi ödeme TAM taksit
        # sayılmaz → kalan_vade düşürülmez (kalan için aşağıda yeni plan açılıyor). Önceden bu
        # güncelleme hiç yapılmıyordu → kasadan para çıkıp borç hep yüksek görünüyordu.
        if kaynak == 'borc_envanteri' and plan.get('kaynak_id') and odenen > 0:
            cur.execute(
                """
                UPDATE borc_envanteri
                SET toplam_borc = GREATEST(0, COALESCE(toplam_borc, 0) - %s)
                WHERE id=%s
                """,
                (odenen, plan['kaynak_id']),
            )

        # 3. Kalan için yeni plan oluştur
        # referans_ay: yeni vade tarihinin ayı — eski planın ay'ını kopyalama, motor o ayı tekrar üretmesin diye
        # NOT: Kart kesim ödemesinde yeni plan oluşturulmaz — mevcut plan aynı ayın asgarisi için
        # panelde 'bekliyor' olarak kalır (asgari-ödenen farkı uyarı olarak görünür).
        yeni_id = None
        if not is_kart_plan:
            yeni_referans_ay = str(body.kalan_vade_tarihi)  # DATE_TRUNC('month') DB'de yapılır
            yeni_id = str(uuid.uuid4())
            if kaynak == 'vadeli_alimlar' and plan.get('kaynak_id'):
                _vadeli_diger_aktif_planlari_iptal(cur, plan['kaynak_id'], oid)
            cur.execute("""
                INSERT INTO odeme_plani
                    (id, kart_id, tarih, referans_ay, odenecek_tutar, asgari_tutar, aciklama, durum, kaynak_tablo, kaynak_id)
                VALUES (%s, %s, %s, DATE_TRUNC('month', %s::date), %s, %s, %s, 'bekliyor', %s, %s)
            """, (
                yeni_id,
                plan.get('kart_id'),
                str(body.kalan_vade_tarihi),
                str(body.kalan_vade_tarihi),
                kalan, kalan,
                f"{plan['aciklama']} (kalan)",
                plan.get('kaynak_tablo'),
                plan.get('kaynak_id')
            ))
            if kaynak == 'vadeli_alimlar' and plan.get('kaynak_id'):
                _vadeli_diger_aktif_planlari_iptal(cur, plan['kaynak_id'], yeni_id)

        # Nakit kısmi ödeme kart_id'li planda ise kart borcunu düşür
        if plan.get('kart_id') and odenen > 0 and not (
            kaynak == 'vadeli_alimlar'
            and getattr(body, 'odeme_yontemi', 'nakit') == 'kart'
        ):
            cur.execute("""
                INSERT INTO kart_hareketleri
                    (id, kart_id, tarih, islem_turu, tutar, aciklama, kaynak_id, kaynak_tablo)
                VALUES (%s, %s, %s, 'ODEME', %s, %s, %s, 'odeme_plani')
                ON CONFLICT DO NOTHING
            """, (
                # FIX K1: id ödeme-olayı bazlı (eskiden plan-bazlı sabit kodm_{oid} idi →
                # 2. kısmi ödeme ON CONFLICT'e takılıp kart borcunu düşürmüyordu).
                # Kasa dedup anahtarıyla AYNI seq → iki defter senkron. Birebir retry'da
                # aynı seq → ON CONFLICT yine korur.
                f"kodm_{oid}_{_odeme_seq}",
                plan['kart_id'],
                bugun,
                abs(odenen),
                f"Kısmi ödeme: {plan.get('aciklama', '')}",
                oid,
            ))

        # Vadeli alımlar kalanı odeme_plani'nda 'bekliyor' olarak bırakılır — onay kuyruğuna girmez.
        # Kart kesim ödemesinde zaten yeni plan oluşmadığı için onay da yok.
        # Diğer kaynak türleri (sabit_giderler, personel vb.) onay akışını kullanmaya devam eder.
        if yeni_id and kaynak != 'vadeli_alimlar':
            onay_ekle(cur, 'ODEME_PLANI', 'odeme_plani', yeni_id,
                f"Kısmi ödeme kalanı: {plan['aciklama']} ({int(kalan):,} ₺)",
                kalan, body.kalan_vade_tarihi)

        audit(cur, 'odeme_plani', oid, 'KISMI_ODE',
              eski={'tutar': toplam}, yeni={'odenen': odenen, 'kalan': kalan, 'yeni_plan': yeni_id})

        # Uyarı önbelleğini temizle — kısmi ödeme sonrası uyarı güncellensin
        uyari_cache_clear()

    return {"success": True, "odenen": odenen, "kalan": kalan, "yeni_plan_id": yeni_id}

@app.get("/api/kasa-detay")
def kasa_detay_endpoint():
    """Kasa'yı işlem türü bazında gösterir — her türün ne kadar etki yaptığını döker."""
    try:
        return kasa_detay()
    except Exception as e:
        raise HTTPException(500, str(e))

# ── FAİZ SİSTEMİ (TEK ENTRY POINT) ────────────────────────────
# Doküman: faiz otomatik çalışır, manuel giriş yoktur.
# Tüm faiz hesabı finans_core.faiz_hesapla_ve_yaz üzerinden geçer.
# Eski 5 endpoint → 1 endpoint.

@app.get("/api/kart-faiz")
def kart_faiz_listele(kart_id: str = None):
    """Kart bazlı faiz geçmişi."""
    with db() as (conn, cur):
        if kart_id:
            cur.execute("""
                SELECT kh.*, k.kart_adi, k.banka
                FROM kart_hareketleri kh
                JOIN kartlar k ON k.id = kh.kart_id
                WHERE kh.kart_id = %s AND kh.islem_turu = 'FAIZ'
                AND kh.durum = 'aktif'
                ORDER BY kh.tarih DESC
            """, (kart_id,))
        else:
            cur.execute("""
                SELECT kh.*, k.kart_adi, k.banka
                FROM kart_hareketleri kh
                JOIN kartlar k ON k.id = kh.kart_id
                WHERE kh.islem_turu = 'FAIZ' AND kh.durum = 'aktif'
                ORDER BY kh.tarih DESC
            """)
        return [dict(r) for r in cur.fetchall()]

@app.post("/api/kartlar/faiz-uret")
def faiz_uret(body: dict = {}):
    """
    Faiz hesapla ve yaz — tek entry point.
    body: { kart_id: str (opsiyonel), donem: 'YYYY-MM' (opsiyonel) }
    kart_id verilmezse tüm aktif kartlar işlenir.
    donem verilmezse bu ay işlenir.
    """
    kart_id = body.get('kart_id')
    donem   = body.get('donem')
    try:
        with db() as (conn, cur):
            if kart_id:
                sonuc = faiz_hesapla_ve_yaz(cur, kart_id, donem)
                audit(cur, 'kart_hareketleri',
                      sonuc.get('id', kart_id), 'FAIZ_OTOMATIK')
                return sonuc
            else:
                sonuclar = tum_kartlar_faiz_hesapla(cur, donem)
                for s in sonuclar:
                    if s.get('id'):
                        audit(cur, 'kart_hareketleri', s['id'], 'FAIZ_OTOMATIK')
                return {
                    "donem":   donem or bugun_tr().strftime('%Y-%m'),
                    "kartlar": sonuclar,
                    "yazilan": sum(1 for s in sonuclar if s.get('durum') == 'yazildi'),
                }
    except Exception as e:
        raise HTTPException(500, str(e))

# ── HAFTALIK FIRE RAPORU ──────────────────────────────────────
@app.get("/api/fire-haftalik")
def fire_haftalik_rapor(
    sube_id: Optional[str] = None,
    hafta_sayisi: int = 4,
    yeniden_hesapla: bool = False,
):
    """
    Kalem bazlı haftalık fire raporu.
    yeniden_hesapla=true → geçen haftayı tekrar hesaplar (test/düzeltme).
    """
    from operasyon_stok_motor import haftalik_fire_hesapla as _hfh
    from datetime import date as _date, timedelta as _td

    hafta_sayisi = max(1, min(int(hafta_sayisi), 12))
    with db() as (conn, cur):
        if yeniden_hesapla and sube_id:
            gecen_hf = _date.today() - _td(days=_date.today().weekday() + 7)
            r = _hfh(cur, sube_id, gecen_hf)
            conn.commit()

        q = """
            SELECT f.*, s.ad AS sube_adi
            FROM sube_fire_haftalik f
            JOIN subeler s ON s.id = f.sube_id
            WHERE f.hafta_baslangic >= CURRENT_DATE - (%s * INTERVAL '7 days')
        """
        params: list = [hafta_sayisi]
        if sube_id:
            q += " AND f.sube_id = %s"
            params.append(sube_id)
        q += " ORDER BY f.sube_id, f.hafta_baslangic DESC"
        cur.execute(q, tuple(params))
        rows = []
        for r in cur.fetchall():
            d = dict(r)
            if isinstance(d.get("kalemler"), str):
                import json as _j
                try:
                    d["kalemler"] = _j.loads(d["kalemler"])
                except Exception:
                    pass
            rows.append(d)
    return {"raporlar": rows, "toplam": len(rows)}


# ── KASA TUTARLILIK KONTROLÜ ──────────────────────────────────
@app.get("/api/kasa-kontrol")
def kasa_kontrol():
    """Ciro kayıtlarında kasa anomalisi var mı? Varsa listeler."""
    with db() as (conn, cur):
        cur.execute("""
            SELECT * FROM v_kasa_anomali
            WHERE durum != 'OK'
            LIMIT 50
        """)
        anomaliler = [dict(r) for r in cur.fetchall()]
        cur.execute("SELECT COUNT(*) as toplam FROM v_kasa_anomali")
        toplam = cur.fetchone()['toplam']
        cur.execute("SELECT COUNT(*) as sorunlu FROM v_kasa_anomali WHERE durum != 'OK'")
        sorunlu = cur.fetchone()['sorunlu']
        return {
            "toplam_ciro": toplam,
            "sorunlu": sorunlu,
            "saglikli": toplam - sorunlu,
            "anomaliler": anomaliler
        }

# ── TOPLU ÖDEME (tek transaction) ──────────────────────────────
@app.post("/api/toplu-odeme")
def toplu_odeme(payload: dict):
    """
    Birden fazla ödemeyi tek transaction'da uygular.
    Biri başarısız olursa hepsi rollback.
    payload: { odemeler: [{odeme_id, tutar}] }
    """
    odemeler = payload.get('odemeler', [])
    if not odemeler:
        raise HTTPException(400, "Ödeme listesi boş")
    
    with db() as (conn, cur):
        # FIX MN5 (2026-07-05): kasa yeterlilik kontrolü ile ödeme uygulaması arasında
        # kilit yoktu → iki eşzamanlı toplu-ödeme aynı bakiyeyi görüp kasayı negatife
        # düşürebiliyordu. Global advisory lock (transaction sonunda otomatik düşer)
        # kasadan-çıkaran toplu işlemleri serileştirir; plan FOR UPDATE tekli /ode ile yarışı keser.
        cur.execute("SELECT pg_advisory_xact_lock(hashtext('kasa-toplu-cikis'))")
        # PROD-PANEL-001 FIX: Kasa yeterlilik kontrolü, client'ın gönderdiği `tutar`a değil,
        # SERVER'ın hesapladığı KALAN borca (odenecek − odenen) göre yapılır. Aksi halde tutar
        # boş gelen satırlar toplamdan düşüyor ve yetersiz kasada işlem başlayabiliyordu.
        _oids = [i.get('odeme_id') for i in odemeler if i.get('odeme_id')]
        _kalan_map: dict = {}
        if _oids:
            cur.execute(
                """SELECT id::text AS id,
                          (COALESCE(odenecek_tutar,0) - COALESCE(odenen_tutar,0)) AS kalan
                   FROM odeme_plani WHERE id = ANY(%s) AND durum <> 'odendi'""",
                (_oids,),
            )
            for r in cur.fetchall():
                _kalan_map[str(r['id'])] = float(r['kalan'] or 0)
        # Backend kasa kontrolü — core'dan
        kasa = kasa_bakiyesi(cur)
        toplam = sum(v for v in _kalan_map.values() if v > 0)
        if toplam > 0 and kasa - toplam < -1:
            raise HTTPException(400, f"Kasa yetersiz. Kasa: {kasa:,.0f}₺ · Toplam ödeme: {toplam:,.0f}₺")

        basarili = []
        for item in odemeler:
            oid = item.get('odeme_id')
            tutar = item.get('tutar')
            if not oid:
                continue
            cur.execute("SELECT * FROM odeme_plani WHERE id=%s FOR UPDATE", (oid,))
            plan = cur.fetchone()
            if not plan:
                raise HTTPException(404, f"Ödeme bulunamadı: {oid}")
            if plan['durum'] == 'odendi':
                continue  # Zaten ödendi, atla
            # PROD-PANEL-001 FIX: toplu-öde = TAM KAPATMA. Booking her zaman server'ın hesapladığı
            # KALAN borç kadar (odenecek − odenen); client `tutar` yalnızca DOĞRULAMA için. Kısmi/
            # negatif/fazla tutar plan'ı 'odendi' yapıp eksik-hatalı kasa hareketi yazamaz → defter
            # bozulmasın. Kısmi ödeme için /api/odeme-plani/{oid}/kismi-ode kullanılır (kuruş-int kıyas).
            kalan_krs = int(round((float(plan['odenecek_tutar'] or 0) - float(plan.get('odenen_tutar') or 0)) * 100))
            if kalan_krs <= 0:
                continue  # ödenecek kalan yok
            if tutar is not None and int(round(float(tutar) * 100)) != kalan_krs:
                raise HTTPException(400, f"Toplu ödeme tam kapatma yapar (kalan {kalan_krs/100:.2f}₺); kısmi/farklı tutar için kısmi-öde kullanın: {oid}")
            # 🔴 PERS-011 (2026-09-02): maaş onay kapısı tekli /ode ve /kismi-ode
            # yollarında vardı, TOPLU ödemede YOKTU. Yani onaylanmamış bir maaş
            # tek tek ödenemezken toplu listede işaretlenerek kapatılabiliyordu —
            # kapı, etrafından dolaşılabildiği sürece kapı değildir.
            # Guard hata atarsa TÜM parti geri alınır (tek transaction): onaysız
            # bir kalem yüzünden partinin geri kalanını sessizce ödemek, sorunu
            # görünmez kılardı.
            _personel_maas_odeme_guard(cur, dict(plan))
            odenen = kalan_krs / 100.0
            bugun = str(bugun_tr())
            # P1 (Codex diff-review 2026-08-13): tam kapatmada odenen_tutar
            # BİRİKİMLİ olmalı (önceki kısmi ödemeler + bu kalan) — kalan ile
            # overwrite edilirse /ode-/kismi-ode'de kurulan tutarlılık bozulur.
            cur.execute("UPDATE odeme_plani SET durum='odendi', odeme_tarihi=%s, odenen_tutar=%s WHERE id=%s",
                        (bugun, float(plan['odenecek_tutar'] or 0), oid))
            plan_d = dict(plan)
            ana_t = kasa_ve_faiz_odeme_plani_tam_odeme(
                cur, plan_d, oid, odenen, bugun,
                anapara_aciklama=f"Toplu ödeme: {plan['aciklama']}",
            )
            if plan.get('kaynak_tablo') == 'vadeli_alimlar' and plan.get('kaynak_id'):
                vadeli_alim_kapat(cur, plan['kaynak_id'], bugun)
            guncelle_borc_envanteri_odeme_plani_sonrasi(cur, plan_d, ana_t)
            # Onay kuyruğunu kapat — PROD-PANEL-002 FIX (toplu): kaynak_tablo ile skopla (tablo-arası
            # id çakışması yanlış onay kapatmasın) + durum='bekliyor' (iptal/revize süpürülmesin).
            if plan.get('kaynak_tablo') and plan.get('kaynak_id'):
                cur.execute("""UPDATE onay_kuyrugu SET durum='onaylandi', onay_tarihi=NOW()
                    WHERE durum='bekliyor'
                      AND ( (kaynak_tablo='odeme_plani' AND kaynak_id=%s)
                         OR (kaynak_tablo=%s AND kaynak_id=%s) )""",
                    (oid, plan['kaynak_tablo'], plan['kaynak_id']))
            else:
                cur.execute("""UPDATE onay_kuyrugu SET durum='onaylandi', onay_tarihi=NOW()
                    WHERE durum='bekliyor' AND kaynak_tablo='odeme_plani' AND kaynak_id=%s""",
                    (oid,))
            audit(cur, 'odeme_plani', oid, 'TOPLU_ODEME', eski=plan)
            basarili.append(oid)
        # Hepsi başarılıysa commit (with db() otomatik commit eder)
    return {"success": True, "uygulanan": len(basarili), "odemeler": basarili}

# ── AY SONU RAPOR (Excel) ──────────────────────────────────────
@app.get("/api/rapor/aylik")
def aylik_rapor(yil: int = None, ay: int = None):
    import calendar as cal
    bugun = bugun_tr()
    yil = yil or bugun.year
    ay  = ay  or bugun.month
    ay_basi = date(yil, ay, 1)
    ay_son  = date(yil, ay, cal.monthrange(yil, ay)[1])

    donem_key = f"{yil}-{ay:02d}"

    with db() as (conn, cur):
        # Mühürlü dönem mi? → değişmez snapshot'ı döndür (NRF dönem kapanışı)
        try:
            cur.execute("SELECT ozet_json, muhurleyen_ad, muhur_ts FROM rapor_kapanis WHERE donem=%s", (donem_key,))
            _seal = cur.fetchone()
        except Exception:
            _seal = None
        if _seal:
            sd = dict(_seal)
            snap = sd.get("ozet_json")
            if isinstance(snap, str):
                snap = json.loads(snap)
            if isinstance(snap, dict):
                snap["muhur"] = {
                    "muhurlu": True,
                    "muhurleyen_ad": sd.get("muhurleyen_ad"),
                    "muhur_ts": str(sd.get("muhur_ts") or ""),
                }
                return snap

        # 0. Ay başı kasa
        cur.execute("SELECT COALESCE(SUM(tutar),0) as v FROM kasa_hareketleri WHERE kasa_etkisi=true AND durum='aktif' AND tarih < %s", (ay_basi,))
        baslangic_kasa = float(cur.fetchone()['v'])

        # 1. Özet
        cur.execute("""
            SELECT
                COALESCE(SUM(CASE WHEN islem_turu='CIRO'          THEN tutar  ELSE 0 END),0) as ciro_toplam,
                COALESCE(SUM(CASE WHEN islem_turu='DIS_KAYNAK'    THEN tutar  ELSE 0 END),0) as dis_kaynak_toplam,
                COALESCE(SUM(CASE WHEN islem_turu='DEVIR'         THEN tutar  ELSE 0 END),0) as devir_toplam,
                -- 🔴 RAPOR-006 (2026-09-02): kart_toplam KART_FAIZ'i de içeriyordu,
                -- ama gider dağılımında "Kart Faizi" AYRI satır olarak da basılıyordu
                -- (Rapor.jsx ve Excel çıktısı) → faiz İKİ KEZ sayılıyordu.
                -- Ayrıca drill detayı (kart_detay) odeme_plani ANAPARAsından gelir,
                -- faiz içermez — yani başlık ile detay da tutmuyordu.
                -- Artık iki alan AYRIK: kart_toplam = ödeme, kart_faiz_toplam = faiz.
                COALESCE(SUM(CASE WHEN islem_turu='KART_ODEME'    THEN ABS(tutar) ELSE 0 END),0) as kart_toplam,
                COALESCE(SUM(CASE WHEN islem_turu='KART_FAIZ'     THEN ABS(tutar) ELSE 0 END),0) as kart_faiz_toplam,
                COALESCE(SUM(CASE WHEN islem_turu='ANLIK_GIDER'   THEN ABS(tutar) ELSE 0 END),0) as anlik_toplam,
                COALESCE(SUM(CASE WHEN islem_turu='VADELI_ODEME'  THEN ABS(tutar) ELSE 0 END),0) as vadeli_toplam,
                COALESCE(SUM(CASE WHEN islem_turu='PERSONEL_MAAS' THEN ABS(tutar) ELSE 0 END),0) as maas_toplam,
                COALESCE(SUM(CASE WHEN islem_turu='SABIT_GIDER'   THEN ABS(tutar) ELSE 0 END),0) as sabit_toplam,
                COALESCE(SUM(CASE WHEN islem_turu='BORC_TAKSIT'   THEN ABS(tutar) ELSE 0 END),0) as borc_taksit_toplam,
                COALESCE(SUM(CASE WHEN islem_turu='FATURA_ODEMESI' THEN ABS(tutar) ELSE 0 END),0) as fatura_toplam,
                COALESCE(SUM(CASE WHEN islem_turu='POS_KESINTI'   THEN ABS(tutar) ELSE 0 END),0) as pos_kesinti_toplam,
                COALESCE(SUM(CASE WHEN tutar > 0 AND islem_turu != 'DEVIR' THEN tutar ELSE 0 END),0) as toplam_gelir,
                COALESCE(SUM(CASE WHEN tutar < 0 THEN ABS(tutar) ELSE 0 END),0) as toplam_gider,
                COALESCE(SUM(tutar),0) as net_kasa_degisim
            FROM kasa_hareketleri
            WHERE durum='aktif' AND kasa_etkisi=true AND tarih BETWEEN %s AND %s
        """, (ay_basi, ay_son))
        ozet = dict(cur.fetchone())
        ozet['baslangic_kasa'] = baslangic_kasa
        ozet['bitis_kasa']     = baslangic_kasa + float(ozet['net_kasa_degisim'])
        ozet['net_kar_zarar']  = float(ozet['toplam_gelir']) - float(ozet['toplam_gider'])

        # 1b. Ciro breakdown
        cur.execute("""
            SELECT COALESCE(SUM(nakit),0) as nakit, COALESCE(SUM(pos),0) as pos,
                   COALESCE(SUM(online),0) as online, COUNT(*) as islem_sayisi
            FROM ciro WHERE durum='aktif' AND tarih BETWEEN %s AND %s
        """, (ay_basi, ay_son))
        cbd = dict(cur.fetchone())
        ozet['ciro_nakit']  = float(cbd['nakit'])
        ozet['ciro_pos']    = float(cbd['pos'])
        ozet['ciro_online'] = float(cbd['online'])
        ozet['ciro_islem']  = int(cbd['islem_sayisi'])

        # 2. Şube bazlı ciro
        cur.execute("""
            SELECT COALESCE(s.ad,'Tanımsız') as sube,
                   COALESCE(SUM(c.toplam),0) as ciro, COALESCE(SUM(c.nakit),0) as nakit,
                   COALESCE(SUM(c.pos),0) as pos, COALESCE(SUM(c.online),0) as online,
                   COUNT(*) as islem_sayisi
            FROM ciro c LEFT JOIN subeler s ON s.id=c.sube_id
            WHERE c.durum='aktif' AND c.tarih BETWEEN %s AND %s
            GROUP BY s.ad ORDER BY ciro DESC
        """, (ay_basi, ay_son))
        sube_ciro = [dict(r) for r in cur.fetchall()]

        # 3. Sabit gider detay
        cur.execute("""
            SELECT COALESCE(sg.gider_adi, kh.aciklama) as gider_adi,
                   COALESCE(sg.kategori,'') as kategori,
                   ABS(kh.tutar) as odenen, kh.tarih::text as odeme_tarihi
            FROM kasa_hareketleri kh
            LEFT JOIN sabit_giderler sg ON sg.id=kh.kaynak_id
            WHERE kh.islem_turu='SABIT_GIDER' AND kh.durum='aktif' AND kh.kasa_etkisi=true
            AND kh.tarih BETWEEN %s AND %s ORDER BY kh.tarih
        """, (ay_basi, ay_son))
        sabit_detay = [dict(r) for r in cur.fetchall()]

        # 4. Personel detay
        cur.execute("""
            SELECT COALESCE(p.ad_soyad, kh.aciklama) as ad_soyad,
                   COALESCE(p.gorev,'') as gorev,
                   ABS(kh.tutar) as odenen, kh.tarih::text as odeme_tarihi
            FROM kasa_hareketleri kh
            LEFT JOIN personel p ON p.id=kh.kaynak_id
            WHERE kh.islem_turu='PERSONEL_MAAS' AND kh.durum='aktif' AND kh.kasa_etkisi=true
            AND kh.tarih BETWEEN %s AND %s ORDER BY kh.tarih
        """, (ay_basi, ay_son))
        personel_detay = [dict(r) for r in cur.fetchall()]

        # 5. Anlık gider kategori
        cur.execute("""
            SELECT kategori, COUNT(*) as adet, COALESCE(SUM(tutar),0) as toplam
            FROM anlik_giderler WHERE durum='aktif' AND tarih BETWEEN %s AND %s
            GROUP BY kategori ORDER BY toplam DESC
        """, (ay_basi, ay_son))
        anlik_kategoriler = [dict(r) for r in cur.fetchall()]

        # 6. Kart detay — odeme_plani üzerinden
        cur.execute("""
            SELECT k.kart_adi, k.banka,
                   COALESCE(SUM(op.odenen_tutar),0) as anapara,
                   0 as faiz, COUNT(*) as adet
            FROM odeme_plani op JOIN kartlar k ON k.id=op.kart_id
            WHERE op.durum='odendi' AND op.kart_id IS NOT NULL
            AND op.odeme_tarihi BETWEEN %s AND %s
            GROUP BY k.kart_adi, k.banka ORDER BY anapara DESC
        """, (ay_basi, ay_son))
        kart_detay = [dict(r) for r in cur.fetchall()]

        # 7. Günlük kümülatif
        cur.execute("""
            SELECT tarih::text,
                   COALESCE(SUM(CASE WHEN tutar>0 THEN tutar ELSE 0 END),0) as giris,
                   COALESCE(SUM(CASE WHEN tutar<0 THEN ABS(tutar) ELSE 0 END),0) as cikis,
                   SUM(tutar) as net
            FROM kasa_hareketleri
            WHERE durum='aktif' AND kasa_etkisi=true AND tarih BETWEEN %s AND %s
            GROUP BY tarih ORDER BY tarih
        """, (ay_basi, ay_son))
        gunluk = [dict(r) for r in cur.fetchall()]
        kumulatif = baslangic_kasa
        for g in gunluk:
            kumulatif += float(g['net'])
            g['kasa'] = round(kumulatif, 2)

        # 8. Önceki ay karşılaştırma
        if ay == 1: onceki_yil, onceki_ay = yil-1, 12
        else: onceki_yil, onceki_ay = yil, ay-1
        ob = date(onceki_yil, onceki_ay, 1)
        os_ = date(onceki_yil, onceki_ay, cal.monthrange(onceki_yil, onceki_ay)[1])
        cur.execute("""
            SELECT COALESCE(SUM(CASE WHEN islem_turu='CIRO' THEN tutar ELSE 0 END),0) as ciro,
                   COALESCE(SUM(CASE WHEN tutar>0 AND islem_turu!='DEVIR' THEN tutar ELSE 0 END),0) as gelir,
                   COALESCE(SUM(CASE WHEN tutar<0 THEN ABS(tutar) ELSE 0 END),0) as gider
            FROM kasa_hareketleri WHERE durum='aktif' AND kasa_etkisi=true AND tarih BETWEEN %s AND %s
        """, (ob, os_))
        onceki = dict(cur.fetchone())

        en_karli = max(sube_ciro, key=lambda x: x['ciro']) if sube_ciro else None

        # ───────────────────────────────────────────────────────────────
        # FAZ 1 KARNE — KPI'lar + 12 aylık trend + şube karnesi + yön. özeti
        # ───────────────────────────────────────────────────────────────
        aylar_kisa = ['', 'Oca', 'Şub', 'Mar', 'Nis', 'May', 'Haz', 'Tem', 'Ağu', 'Eyl', 'Eki', 'Kas', 'Ara']

        def _tl(v):
            try:
                return f"{float(v):,.0f} ₺".replace(',', '.')
            except Exception:
                return "0 ₺"

        ciro_t = float(ozet.get('ciro_toplam') or 0)
        gelir_t = float(ozet.get('toplam_gelir') or 0)
        gider_t = float(ozet.get('toplam_gider') or 0)
        net_t = float(ozet.get('net_kar_zarar') or 0)
        pos_ciro = float(ozet.get('ciro_pos') or 0)
        pos_kesinti = float(ozet.get('pos_kesinti_toplam') or 0)
        # Devam eden ayda geçen güne kadar; geçmiş ayda tam ay gün sayısı (adil run-rate)
        ay_suruyor = (bugun.year == yil and bugun.month == ay)
        gecen_gun = bugun.day if ay_suruyor else ay_son.day
        onc_gun = os_.day  # önceki ayın gün sayısı
        gun_say = gecen_gun
        gunluk_gider = (gider_t / gun_say) if gun_say else 0.0
        kpi = {
            "net_kar_marji": round(net_t / gelir_t * 100, 1) if gelir_t else None,
            "gider_ciro_orani": round(gider_t / ciro_t * 100, 1) if ciro_t else None,
            "pos_yanan_orani": round(pos_kesinti / pos_ciro * 100, 2) if pos_ciro else None,
            "pos_kesinti_toplam": round(pos_kesinti, 2),
            "gunluk_ortalama_gider": round(gunluk_gider, 2),
            "runway_gun": int(float(ozet.get('bitis_kasa') or 0) / gunluk_gider) if gunluk_gider > 0 else None,
            "bitis_kasa": float(ozet.get('bitis_kasa') or 0),
        }

        # 12 aylık trend
        ty12, tm12 = yil, ay - 11
        while tm12 <= 0:
            tm12 += 12
            ty12 -= 1
        trend_basi = date(ty12, tm12, 1)
        cur.execute("""
            SELECT to_char(date_trunc('month', tarih),'YYYY-MM') AS ay,
                   COALESCE(SUM(CASE WHEN islem_turu='CIRO' THEN tutar ELSE 0 END),0) AS ciro,
                   COALESCE(SUM(CASE WHEN tutar>0 AND islem_turu!='DEVIR' THEN tutar ELSE 0 END),0) AS gelir,
                   COALESCE(SUM(CASE WHEN tutar<0 THEN ABS(tutar) ELSE 0 END),0) AS gider
            FROM kasa_hareketleri
            WHERE durum='aktif' AND kasa_etkisi=true AND tarih BETWEEN %s AND %s
            GROUP BY 1 ORDER BY 1
        """, (trend_basi, ay_son))
        _tr_map = {r['ay']: dict(r) for r in cur.fetchall()}
        trend12 = []
        _cy, _cm = ty12, tm12
        for _ in range(12):
            key = f"{_cy}-{_cm:02d}"
            row = _tr_map.get(key)
            ge = float(row['gelir']) if row else 0.0
            gi = float(row['gider']) if row else 0.0
            trend12.append({
                "ay": key,
                "ay_kisa": aylar_kisa[_cm],
                "ciro": float(row['ciro']) if row else 0.0,
                "gelir": ge, "gider": gi, "net": ge - gi,
            })
            _cm += 1
            if _cm > 12:
                _cm = 1
                _cy += 1

        # Önceki ay şube ciro (büyüme için)
        cur.execute("""
            SELECT COALESCE(s.ad,'Tanımsız') AS sube, COALESCE(SUM(c.toplam),0) AS ciro
            FROM ciro c LEFT JOIN subeler s ON s.id=c.sube_id
            WHERE c.durum='aktif' AND c.tarih BETWEEN %s AND %s GROUP BY s.ad
        """, (ob, os_))
        _onc_sube = {r['sube']: float(r['ciro']) for r in cur.fetchall()}

        # Şube risk sinyalleri (denetim uyarıları aylık roll-up)
        _risk_sube = {}
        try:
            cur.execute("""
                SELECT COALESCE(s.ad,'Tanımsız') AS sube, COUNT(*) AS c
                FROM sube_operasyon_uyari u LEFT JOIN subeler s ON s.id=u.sube_id
                WHERE u.tarih BETWEEN %s AND %s
                  AND u.tip IN ('ACILIS_KASA_FARK','FIRE_TESPITI')
                GROUP BY s.ad
            """, (ay_basi, ay_son))
            _risk_sube = {r['sube']: int(r['c']) for r in cur.fetchall()}
        except Exception:
            _risk_sube = {}

        def _harf(skor):
            if skor >= 85:
                return "A"
            if skor >= 70:
                return "B"
            if skor >= 55:
                return "C"
            return "D"

        sube_karne = []
        for s in sube_ciro:
            ad = s['sube']
            ci = float(s.get('ciro') or 0)
            onc_ci = _onc_sube.get(ad, 0.0)
            # Adil kıyas: günlük ortalama (run-rate) — yarım ay tam ayı haksız ezmesin
            bu_gunluk = (ci / gecen_gun) if gecen_gun else 0.0
            onc_gunluk = (onc_ci / onc_gun) if onc_gun else 0.0
            buyume = round((bu_gunluk - onc_gunluk) / onc_gunluk * 100, 1) if onc_gunluk > 0 else None
            risk = int(_risk_sube.get(ad, 0))
            pay = round(ci / ciro_t * 100, 1) if ciro_t else 0.0
            skor = 80.0
            if buyume is not None:
                skor += max(-15.0, min(15.0, buyume))
            # Risk: ham sayı değil, gün başına sıklık (her gün kasa farkı = ağır ceza)
            risk_siklik = (risk / gecen_gun) if gecen_gun else 0.0
            skor -= min(30.0, risk_siklik * 40.0)
            sube_karne.append({
                "sube": ad, "ciro": ci, "pay_yuzde": pay, "buyume_yuzde": buyume,
                "risk_sinyali": risk, "islem_sayisi": s.get('islem_sayisi') or 0,
                "harf": _harf(skor), "skor": round(skor),
            })
        sube_karne.sort(key=lambda x: x['ciro'], reverse=True)

        # ── FAZ 2-A: DENETİM & RİSK ÖZETİ (aylık roll-up) ──
        denetim_ozeti = {
            "kasa": {"acik_tl": 0.0, "fazla_tl": 0.0, "acik_gun": 0, "olay": 0},
            "kasa_sube": [],
            "fire": {"toplam_bildirim": 0, "toplam_adet": 0, "sebepler": []},
            "uyumsuzluk": {"acik_adet": 0, "bekleyen_fark": 0},
        }
        # 1) Kasa disiplini (fark_tl: negatif=açık, pozitif=fazla)
        try:
            cur.execute("""
                SELECT COALESCE(s.ad,'Tanımsız') AS sube,
                       COUNT(*) AS olay,
                       COALESCE(SUM(CASE WHEN u.fark_tl < 0 THEN ABS(u.fark_tl) ELSE 0 END),0) AS acik_tl,
                       COALESCE(SUM(CASE WHEN u.fark_tl > 0 THEN u.fark_tl ELSE 0 END),0) AS fazla_tl,
                       COUNT(DISTINCT u.tarih) FILTER (WHERE u.fark_tl < 0) AS acik_gun
                FROM sube_operasyon_uyari u LEFT JOIN subeler s ON s.id=u.sube_id
                WHERE u.tarih BETWEEN %s AND %s
                  AND u.fark_tl IS NOT NULL AND u.fark_tl <> 0
                GROUP BY s.ad ORDER BY acik_tl DESC
            """, (ay_basi, ay_son))
            for r in cur.fetchall():
                rd = dict(r)
                denetim_ozeti["kasa_sube"].append({
                    "sube": rd["sube"],
                    "acik_tl": float(rd["acik_tl"]), "fazla_tl": float(rd["fazla_tl"]),
                    "acik_gun": int(rd["acik_gun"]), "olay": int(rd["olay"]),
                })
            denetim_ozeti["kasa"]["acik_tl"] = round(sum(x["acik_tl"] for x in denetim_ozeti["kasa_sube"]), 2)
            denetim_ozeti["kasa"]["fazla_tl"] = round(sum(x["fazla_tl"] for x in denetim_ozeti["kasa_sube"]), 2)
            denetim_ozeti["kasa"]["acik_gun"] = sum(x["acik_gun"] for x in denetim_ozeti["kasa_sube"])
            denetim_ozeti["kasa"]["olay"] = sum(x["olay"] for x in denetim_ozeti["kasa_sube"])
        except Exception:
            pass
        # 2) Fire / zayi (sebep dağılımı)
        try:
            cur.execute("""
                SELECT sebep_label, COUNT(*) AS adet, COALESCE(SUM(toplam_adet),0) AS urun_adet
                FROM sube_fire_bildirim
                WHERE tarih BETWEEN %s AND %s
                GROUP BY sebep_label ORDER BY adet DESC
            """, (ay_basi, ay_son))
            sebepler = [dict(r) for r in cur.fetchall()]
            denetim_ozeti["fire"]["sebepler"] = [
                {"sebep": s["sebep_label"], "adet": int(s["adet"]), "urun_adet": int(s["urun_adet"])}
                for s in sebepler
            ]
            denetim_ozeti["fire"]["toplam_bildirim"] = sum(s["adet"] for s in denetim_ozeti["fire"]["sebepler"])
            denetim_ozeti["fire"]["toplam_adet"] = sum(s["urun_adet"] for s in denetim_ozeti["fire"]["sebepler"])
        except Exception:
            pass
        # 3) Çözülmemiş sevkiyat uyumsuzluğu (şu an açık — dönemden bağımsız risk)
        try:
            cur.execute("""
                SELECT COUNT(*) AS adet,
                       COALESCE(SUM(GREATEST(0, COALESCE(sevk_adet,0)-COALESCE(kabul_adet,0))),0) AS bekleyen_fark
                FROM stok_yolda WHERE durum='kabul_uyusmazlik'
            """)
            ur = dict(cur.fetchone() or {})
            denetim_ozeti["uyumsuzluk"]["acik_adet"] = int(ur.get("adet") or 0)
            denetim_ozeti["uyumsuzluk"]["bekleyen_fark"] = int(ur.get("bekleyen_fark") or 0)
        except Exception:
            pass

        # ── FAZ 2-B: NAKİT AKIŞ & PROJEKSİYON ──
        projeksiyon = {
            "mevcut_kasa": float(ozet.get('bitis_kasa') or 0),
            "gunluk_gelir": 0.0, "gunluk_gider": 0.0, "net_gunluk": 0.0,
            "aylik_sabit_gider": 0.0, "aylik_maas": 0.0,
            "bekleyen_taksit_90": 0.0,
            "ufuklar": [], "runway_gun": None,
        }
        try:
            # 90 günlük run-rate (gelir/gider)
            cur.execute("""
                SELECT COALESCE(SUM(CASE WHEN tutar>0 AND islem_turu!='DEVIR' THEN tutar ELSE 0 END),0) AS gelir,
                       COALESCE(SUM(CASE WHEN tutar<0 THEN ABS(tutar) ELSE 0 END),0) AS gider
                FROM kasa_hareketleri
                WHERE durum='aktif' AND kasa_etkisi=true AND tarih >= CURRENT_DATE - INTERVAL '90 days'
            """)
            rr = dict(cur.fetchone() or {})
            g_gelir = float(rr.get("gelir") or 0) / 90.0
            g_gider_runrate = float(rr.get("gider") or 0) / 90.0

            # Bilinen sabit yükler
            try:
                cur.execute("SELECT COALESCE(SUM(tutar),0) AS v FROM sabit_giderler WHERE aktif=TRUE")
                projeksiyon["aylik_sabit_gider"] = float(dict(cur.fetchone() or {}).get("v") or 0)
            except Exception:
                pass
            try:
                # Sürekli (tam zamanlı) personel: sabit aylık maaş.
                cur.execute("""
                    SELECT COALESCE(SUM(maas),0) AS v FROM personel
                    WHERE aktif=TRUE AND COALESCE(calisma_turu,'surekli')='surekli'
                """)
                aylik_maas_surekli = float(dict(cur.fetchone() or {}).get("v") or 0)

                # Part-time personel: sabit maaş yok — son 30 günde planlanan/onaylı
                # vardiya saatleri × saatlik_ucret ile tahmini aylık maliyet.
                cur.execute("""
                    SELECT COALESCE(SUM(
                        EXTRACT(EPOCH FROM (
                            CASE WHEN va.bitis_saat <= va.baslangic_saat
                                 THEN (va.bitis_saat::time + INTERVAL '24h') - va.baslangic_saat::time
                                 ELSE va.bitis_saat::time - va.baslangic_saat::time END
                        ))/3600.0 * COALESCE(p.saatlik_ucret,0)
                    ),0) AS v
                    FROM vardiya_atama va
                    JOIN personel p ON p.id = va.personel_id
                    WHERE va.tarih >= CURRENT_DATE - INTERVAL '30 days'
                      AND va.tarih < CURRENT_DATE
                      AND va.durum IN ('planli','onayli')
                      AND p.aktif = TRUE
                      AND COALESCE(p.calisma_turu,'surekli') != 'surekli'
                """)
                aylik_maas_parttime = float(dict(cur.fetchone() or {}).get("v") or 0)

                projeksiyon["aylik_maas_surekli"] = round(aylik_maas_surekli, 2)
                projeksiyon["aylik_maas_parttime"] = round(aylik_maas_parttime, 2)
                projeksiyon["aylik_maas"] = round(aylik_maas_surekli + aylik_maas_parttime, 2)
            except Exception:
                pass

            # Günlük gider TABANI: bilinen sabit yükler kasaya tam işlenmemiş olabilir.
            # Run-rate ile (sabit+maaş)/30'un büyüğünü al — ne eksik say, ne çift say.
            sabit_gunluk = (projeksiyon["aylik_sabit_gider"] + projeksiyon["aylik_maas"]) / 30.0
            g_gider = max(g_gider_runrate, sabit_gunluk)
            projeksiyon["gunluk_gelir"] = round(g_gelir, 2)
            projeksiyon["gunluk_gider"] = round(g_gider, 2)
            projeksiyon["gunluk_gider_runrate"] = round(g_gider_runrate, 2)
            projeksiyon["gunluk_gider_sabit_taban"] = round(sabit_gunluk, 2)
            projeksiyon["net_gunluk"] = round(g_gelir - g_gider, 2)

            # Bekleyen taksitler — ufuk bazlı (kesin tarihli yük)
            tk = {30: 0.0, 60: 0.0, 90: 0.0}
            try:
                cur.execute("""
                    SELECT
                      COALESCE(SUM(CASE WHEN odeme_tarihi <= CURRENT_DATE + INTERVAL '30 days' THEN odenecek_tutar ELSE 0 END),0) AS t30,
                      COALESCE(SUM(CASE WHEN odeme_tarihi <= CURRENT_DATE + INTERVAL '60 days' THEN odenecek_tutar ELSE 0 END),0) AS t60,
                      COALESCE(SUM(CASE WHEN odeme_tarihi <= CURRENT_DATE + INTERVAL '90 days' THEN odenecek_tutar ELSE 0 END),0) AS t90
                    FROM odeme_plani
                    WHERE durum='bekliyor' AND odeme_tarihi >= CURRENT_DATE
                      AND odeme_tarihi <= CURRENT_DATE + INTERVAL '90 days'
                """)
                trow = dict(cur.fetchone() or {})
                tk = {30: float(trow.get("t30") or 0), 60: float(trow.get("t60") or 0), 90: float(trow.get("t90") or 0)}
            except Exception:
                pass
            projeksiyon["bekleyen_taksit_90"] = tk[90]

            mevcut = projeksiyon["mevcut_kasa"]
            net_g = projeksiyon["net_gunluk"]
            for n in (30, 60, 90):
                projeksiyon["ufuklar"].append({
                    "gun": n,
                    "beklenen_gelir": round(g_gelir * n, 2),
                    "beklenen_gider": round(g_gider * n, 2),
                    "bekleyen_taksit": round(tk[n], 2),
                    "projekte_kasa": round(mevcut + net_g * n, 2),
                })
            if net_g < 0:
                projeksiyon["runway_gun"] = int(mevcut / abs(net_g)) if abs(net_g) > 0 else None
        except Exception:
            pass

        # Yönetici özeti — otomatik cümleler
        yonetici_ozeti = []
        onc_ciro = float(onceki.get('ciro') or 0)
        if onc_ciro > 0 and gecen_gun and onc_gun:
            cd = ((ciro_t / gecen_gun) - (onc_ciro / onc_gun)) / (onc_ciro / onc_gun) * 100
            _kiyas_not = " (günlük ortalama, ay sürüyor)" if ay_suruyor else ""
            yonetici_ozeti.append({
                "tip": "iyi" if cd >= 0 else "uyari",
                "metin": f"Ciro geçen aya göre %{abs(cd):.1f} {'yüksek' if cd >= 0 else 'düşük'} hızda ({_tl(ciro_t)}){_kiyas_not}.",
            })
        else:
            yonetici_ozeti.append({"tip": "notr", "metin": f"Bu ay ciro: {_tl(ciro_t)}."})
        yonetici_ozeti.append({
            "tip": "iyi" if net_t >= 0 else "uyari",
            # ⚠️ DİL DÜZELTMESİ (2026-08-07 denetimi): bu rakam KÂR DEĞİL, kasa
            # değişimidir (net_kar_zarar = toplam_gelir − toplam_gider ve
            # net_kasa_degisim ile birebir aynı). İçinde dış kaynak geliri
            # (sahibin koyduğu para = sermaye) GELİR sayılır, borç anaparası
            # GİDER sayılır, tahakkuk etmiş ama ödenmemiş maaş HİÇ sayılmaz —
            # canlı Ağustos: maas_toplam 0 iken bordro 68.344 ₺ tahakkuk etmişti.
            # Gerçek P&L kârı Maliyet modülünde ayrı hesaplanır (vergi öncesi kâr).
            # Klasik Rapor.jsx zaten "Net Nakit Akışı (kasa) — kâr değil" diyordu;
            # yönetici özeti cümlesi ise "Net kâr" diyerek iki farklı gerçek üretiyordu.
            "metin": (f"Kasa {'arttı' if net_t >= 0 else 'azaldı'}: {_tl(abs(net_t))} "
                      f"(giren {_tl(gelir_t)} − çıkan {_tl(gider_t)}). "
                      f"Bu KÂR DEĞİL, nakit hareketi — gerçek kâr için Maliyet ekranı."),
        })
        if en_karli:
            yonetici_ozeti.append({
                "tip": "iyi",
                "metin": f"En güçlü şube: {en_karli['sube']} ({_tl(float(en_karli['ciro']))}).",
            })
        if _risk_sube:
            _en_risk = max(_risk_sube.items(), key=lambda kv: kv[1])
            if _en_risk[1] > 0:
                yonetici_ozeti.append({
                    "tip": "uyari",
                    "metin": f"Dikkat: {_en_risk[0]} şubesinde {_en_risk[1]} denetim sinyali (kasa farkı / anomali / fire).",
                })
        if kpi["runway_gun"] is not None:
            yonetici_ozeti.append({
                "tip": "iyi" if kpi["runway_gun"] >= 30 else "uyari",
                "metin": f"Ay sonu kasa {_tl(kpi['bitis_kasa'])} — mevcut gider hızıyla ~{kpi['runway_gun']} gün dayanır.",
            })
        if kpi["pos_yanan_orani"]:
            yonetici_ozeti.append({
                "tip": "notr",
                "metin": f"POS kesintisi cironun %{kpi['pos_yanan_orani']:.2f}'i ({_tl(pos_kesinti)}).",
            })
        _kasa_acik = denetim_ozeti["kasa"]["acik_tl"]
        if _kasa_acik > 0:
            yonetici_ozeti.append({
                "tip": "uyari",
                "metin": f"Kasa açığı: {denetim_ozeti['kasa']['acik_gun']} şube-gün, toplam {_tl(_kasa_acik)} eksik kaydedildi.",
            })
        if denetim_ozeti["uyumsuzluk"]["acik_adet"] > 0:
            yonetici_ozeti.append({
                "tip": "uyari",
                "metin": f"{denetim_ozeti['uyumsuzluk']['acik_adet']} çözülmemiş sevkiyat uyumsuzluğu var (toplam {denetim_ozeti['uyumsuzluk']['bekleyen_fark']} adet fark).",
            })
        if projeksiyon["net_gunluk"] < 0 and projeksiyon["runway_gun"] is not None:
            yonetici_ozeti.append({
                "tip": "uyari",
                "metin": f"Nakit erime uyarısı: günlük net −{_tl(abs(projeksiyon['net_gunluk']))}, mevcut kasa ~{projeksiyon['runway_gun']} gün sonra tükenir.",
            })
        elif projeksiyon["net_gunluk"] > 0:
            _p90 = next((u for u in projeksiyon["ufuklar"] if u["gun"] == 90), None)
            if _p90:
                yonetici_ozeti.append({
                    "tip": "iyi",
                    "metin": f"Nakit trendi pozitif: günlük net +{_tl(projeksiyon['net_gunluk'])}, 90 gün projeksiyon ~{_tl(_p90['projekte_kasa'])}.",
                })

    aylar = ['','Ocak','Şubat','Mart','Nisan','Mayıs','Haziran','Temmuz','Ağustos','Eylül','Ekim','Kasım','Aralık']
    return {
        "donem": f"{yil}-{ay:02d}",
        "donem_label": f"{aylar[ay]} {yil}",
        "ozet": ozet,
        "sube_ciro": sube_ciro,
        "sabit_detay": sabit_detay,
        "personel_detay": personel_detay,
        "anlik_kategoriler": anlik_kategoriler,
        "kart_detay": kart_detay,
        "gunluk": gunluk,
        "onceki_ay": onceki,
        "en_karli_sube": en_karli,
        "kpi": kpi,
        "trend12": trend12,
        "sube_karne": sube_karne,
        "yonetici_ozeti": yonetici_ozeti,
        "denetim_ozeti": denetim_ozeti,
        "projeksiyon": projeksiyon,
        "muhur": {"muhurlu": False},
    }


class RaporMuhurleBody(BaseModel):
    yil: int
    ay: int
    muhurleyen_ad: Optional[str] = None


@app.post("/api/rapor/aylik/muhurle")
def aylik_rapor_muhurle(body: RaporMuhurleBody):
    """Aylık raporu mühürle — değişmez snapshot al (NRF dönem kapanışı).
    Mühürlü dönem bir daha değişmez; GET artık snapshot'ı döndürür."""
    yil, ay = int(body.yil), int(body.ay)
    donem_key = f"{yil}-{ay:02d}"
    with db() as (conn, cur):
        ensure_rapor_kapanis(cur)
        cur.execute("SELECT 1 FROM rapor_kapanis WHERE donem=%s", (donem_key,))
        if cur.fetchone():
            raise HTTPException(409, "Bu dönem zaten mühürlenmiş.")

    # Canlı raporu hesapla (mühür yokken live döner) ve dondur
    snap = aylik_rapor(yil, ay)
    snap.pop("muhur", None)
    payload = json.dumps(snap, ensure_ascii=False, sort_keys=True, default=str)
    h = hashlib.sha256(payload.encode("utf-8")).hexdigest()[:16]
    ad = (body.muhurleyen_ad or "CFO").strip() or "CFO"

    with db() as (conn, cur):
        ensure_rapor_kapanis(cur)
        cur.execute("""
            INSERT INTO rapor_kapanis (donem, ozet_json, muhurleyen_ad, hash)
            VALUES (%s, %s::jsonb, %s, %s)
            ON CONFLICT (donem) DO NOTHING
        """, (donem_key, payload, ad, h))
    return {"success": True, "donem": donem_key, "hash": h, "muhurleyen_ad": ad}

@app.get("/api/banka-mutabakat")
def banka_mutabakat(yil: int = None, ay: int = None):
    """Üçlü nakit mutabakatı (kasa→teslim→banka): dönem boyunca şubelerden TESLİM
    alınan nakit (ara teslim + kapanış teslimi) vs BANKAYA yatan; ayrıca kümülatif
    'elde/yolda nakit' (henüz bankaya gitmemiş havuz). Gösterge amaçlı."""
    import calendar as cal
    bugun = bugun_tr()
    yil = yil or bugun.year
    ay = ay or bugun.month
    ay_basi = date(yil, ay, 1)
    ay_son = date(yil, ay, cal.monthrange(yil, ay)[1])
    # 🔴 P1 (2026-08-13, EVV-PARA-N1 — ÇİFT SAYIM): eski kod kasa_teslim'i TÜR
    # FİLTRESİZ topluyor (ara + gun_sonu) ve üstüne KAPANIS event'inin teslim'ini
    # AYRICA ekliyordu. Oysa kapanış akışı (sube_operasyon.py) gün-sonu teslimi
    # HEM event.teslim'e HEM kasa_teslim'e (teslim_turu='gun_sonu') yazar → aynı
    # para iki kez sayılıyordu. CANLI KANIT: 2026-08 gerçek teslim 92.000 ₺ iken
    # donem_teslim 184.000 ₺, kümülatif elde_nakit 1,53 M ₺ şişkindi.
    # DOĞRU MODEL: ara = kasa_teslim(teslim_turu='ara'); gün-sonu = şube+gün
    # bazında kasa_teslim.gun_sonu ile event.teslim'den YALNIZ BİRİ (kasa_teslim
    # izi öncelikli — düzeltme akışları onu günceller; event-only eski günler
    # kaybolmasın diye FULL OUTER JOIN + COALESCE).
    _GUNSONU_DEDUPE = """
        SELECT COALESCE(SUM(COALESCE(kt.t, ev.t)), 0) AS v
        FROM (
            SELECT sube_id::text AS sube_id, tarih, SUM(tutar) AS t
            FROM kasa_teslim
            WHERE teslim_turu='gun_sonu' AND tarih {op} %s {op2}
            GROUP BY 1, 2
        ) kt
        FULL OUTER JOIN (
            SELECT sube_id::text AS sube_id, tarih, SUM(teslim) AS t
            FROM sube_operasyon_event
            WHERE tip='KAPANIS' AND durum='tamamlandi' AND teslim IS NOT NULL
              AND tarih {op} %s {op2}
            GROUP BY 1, 2
        ) ev USING (sube_id, tarih)
    """
    with db() as (conn, cur):
        cur.execute("SELECT COALESCE(SUM(tutar),0) AS v FROM kasa_teslim WHERE teslim_turu='ara' AND tarih BETWEEN %s AND %s", (ay_basi, ay_son))
        teslim_ara = float(cur.fetchone()["v"])
        cur.execute(_GUNSONU_DEDUPE.format(op="BETWEEN", op2="AND %s"),
                    (ay_basi, ay_son, ay_basi, ay_son))
        teslim_kap = float(cur.fetchone()["v"])
        donem_teslim = round(teslim_ara + teslim_kap, 2)
        cur.execute("SELECT COALESCE(SUM(tutar),0) AS v, COUNT(*) AS c FROM banka_yatirimlari WHERE tarih BETWEEN %s AND %s", (ay_basi, ay_son))
        _y = dict(cur.fetchone())
        donem_yatan = float(_y["v"]); yatan_adet = int(_y["c"])
        # Kümülatif elde nakit (tüm zaman teslim − tüm zaman yatan) — aynı dedupe
        cur.execute("SELECT COALESCE(SUM(tutar),0) AS v FROM kasa_teslim WHERE teslim_turu='ara' AND tarih <= %s", (ay_son,))
        kum_ara = float(cur.fetchone()["v"])
        cur.execute(_GUNSONU_DEDUPE.format(op="<=", op2=""), (ay_son, ay_son))
        kum_kap = float(cur.fetchone()["v"])
        cur.execute("SELECT COALESCE(SUM(tutar),0) AS v FROM banka_yatirimlari WHERE tarih <= %s", (ay_son,))
        kum_yatan = float(cur.fetchone()["v"])
        # ── 💵 ELDEN vs HAVALE (2026-08-09, sahip: "bazı ödemeler nakit olsa bile
        # elden ödenme ihtimali var ve bu mutabakat doğru çalışmaz")
        # Eski formül: elde = teslim − bankaya yatan. ELDEN ödenen para hiç
        # düşülmüyordu; kasadan çıkıp tedarikçiye gitmiş nakit hâlâ "elde duruyor"
        # sayılıyordu. Havale ise banka hesabından çıkar — elde nakiti etkilemez.
        # Kova kova ayırıyoruz; sınıflanmamış 'nakit' kayıtları BELİRSİZ'dir ve
        # elde nakiti bir ARALIK olarak verir (alt sınır ↔ üst sınır).
        def _cikis(yontemler) -> tuple:
            try:
                cur.execute(
                    """SELECT COALESCE(SUM(ABS(tutar)),0) AS v, COUNT(*) AS c
                       FROM kasa_hareketleri
                       WHERE tutar < 0 AND COALESCE(durum,'aktif')='aktif'
                         AND COALESCE(kasa_etkisi, TRUE) = TRUE
                         AND tarih <= %s
                         AND COALESCE(odeme_yontemi,'nakit') = ANY(%s)""",
                    (ay_son, list(yontemler)))
                r = dict(cur.fetchone())
                return round(float(r["v"] or 0), 2), int(r["c"] or 0)
            except Exception:  # noqa: BLE001 — kolon yoksa sessiz geç
                return 0.0, 0
        elden_odenen, elden_adet = _cikis(("elden",))
        havale_odenen, havale_adet = _cikis(("havale", "eft"))
        belirsiz_nakit, belirsiz_adet = _cikis(("nakit",))
        _havuz = round((kum_ara + kum_kap) - kum_yatan, 2)
        # ÜST SINIR: elden ödendiği KESİN olanlar düşülür
        elde_nakit = round(_havuz - elden_odenen, 2)
        # ALT SINIR: belirsizlerin tamamı elden ödenmiş varsayımı. Bu bir ÜST
        # KORKU senaryosudur, gerçekçi değildir — belirsiz kovada kart borcu
        # ödemesi/havale de var (canlıda 210 kayıt · 6,26 M ₺ hepsi 'nakit'
        # varsayılanıyla duruyor). Eldeki nakit NEGATİF olamaz; 0'da kırpılır,
        # yoksa ekranda −4,8 M ₺ gibi anlamsız bir korku rakamı çıkıyordu.
        elde_nakit_alt = round(max(0.0, _havuz - elden_odenen - belirsiz_nakit), 2)
        # Sınıflama olgunluğu: belirsizlerin payı düştükçe aralık daralır
        _siniflanan = elden_odenen + havale_odenen
        siniflama_pct = round(
            100.0 * _siniflanan / (_siniflanan + belirsiz_nakit), 1
        ) if (_siniflanan + belirsiz_nakit) > 0 else 100.0
        # Şube bazlı dönem teslim — aynı gün-sonu dedupe (EVV-PARA-N1):
        # ara teslim + şube+gün bazında COALESCE(kasa_teslim.gun_sonu, event.teslim)
        cur.execute("""
            SELECT COALESCE(s.ad,'?') AS sube,
              COALESCE((SELECT SUM(kt.tutar) FROM kasa_teslim kt
                        WHERE kt.sube_id=s.id AND kt.teslim_turu='ara'
                          AND kt.tarih BETWEEN %s AND %s),0)
            + COALESCE((SELECT SUM(COALESCE(kt2.t, ev2.t)) FROM (
                  SELECT tarih, SUM(tutar) AS t FROM kasa_teslim
                  WHERE sube_id=s.id AND teslim_turu='gun_sonu'
                    AND tarih BETWEEN %s AND %s GROUP BY tarih
               ) kt2
               FULL OUTER JOIN (
                  SELECT tarih, SUM(teslim) AS t FROM sube_operasyon_event
                  WHERE sube_id=s.id AND tip='KAPANIS' AND durum='tamamlandi'
                    AND teslim IS NOT NULL AND tarih BETWEEN %s AND %s GROUP BY tarih
               ) ev2 USING (tarih)),0) AS teslim
            FROM subeler s ORDER BY teslim DESC
        """, (ay_basi, ay_son, ay_basi, ay_son, ay_basi, ay_son))
        sube_teslim = [{"sube": r["sube"], "teslim": float(r["teslim"] or 0)} for r in cur.fetchall() if float(r["teslim"] or 0) > 0]
    return {
        "donem": f"{yil}-{ay:02d}",
        "donem_teslim": donem_teslim,
        "teslim_ara": round(teslim_ara, 2),
        "teslim_kapanis": round(teslim_kap, 2),
        "donem_yatan": round(donem_yatan, 2),
        "yatan_adet": yatan_adet,
        "donem_fark": round(donem_teslim - donem_yatan, 2),
        "elde_nakit": elde_nakit,
        # 💵 ELDEN / HAVALE AYRIMI — sahip 2026-08-09
        "elde_nakit_alt": elde_nakit_alt,
        "elden_odenen": elden_odenen, "elden_adet": elden_adet,
        "havale_odenen": havale_odenen, "havale_adet": havale_adet,
        "belirsiz_nakit": belirsiz_nakit, "belirsiz_adet": belirsiz_adet,
        "siniflama_pct": siniflama_pct,
        "elde_nakit_not": (
            "Elde nakit = teslim alınan − bankaya yatan − ELDEN ödenen. Havale "
            "banka hesabından çıkar, elde nakiti etkilemez. Ödeme yöntemi "
            "seçilmemiş kayıtlar 'belirsiz'dir; alt sınır 'hepsi elden ödendi' "
            "korku senaryosudur ve gerçekçi değildir (belirsiz kovada kart borcu "
            "ödemesi de var). Yeni giderlerde elden/havale seçildikçe aralık daralır."),
        "sube_teslim": sube_teslim,
    }


@app.get("/api/rapor/aylik/excel")
def aylik_rapor_excel(yil: int = None, ay: int = None):
    """
    Aylık raporu Excel olarak indir.
    aylik_rapor() verisini openpyxl ile XLSX'e çevirir.
    """
    import io
    import calendar as cal
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from fastapi.responses import StreamingResponse
    except ImportError:
        raise HTTPException(500, "openpyxl kurulu değil")

    # Aynı rapor verisini çek
    rapor = aylik_rapor(yil, ay)
    o     = rapor["ozet"]
    aylar = ['','Ocak','Şubat','Mart','Nisan','Mayıs','Haziran',
             'Temmuz','Ağustos','Eylül','Ekim','Kasım','Aralık']

    wb = openpyxl.Workbook()

    # ── RENK & STİL TANIMLAR ──────────────────────────────────
    def stil(ws, cell, deger, bold=False, renk=None, bg=None, sayi=False, hizala='left'):
        c = ws[cell] if isinstance(cell, str) else cell
        c.value = deger
        if bold:       c.font = Font(bold=True, size=11)
        if renk:       c.font = Font(bold=bold, color=renk, size=11)
        if bg:         c.fill = PatternFill("solid", fgColor=bg)
        if sayi:       c.number_format = '#,##0'
        c.alignment = Alignment(horizontal=hizala, vertical='center')
        return c

    BASLIK_BG   = "1E2A3A"
    BASLIK_FG   = "FFFFFF"
    ALT_BASLIK  = "2D4A6A"
    SARI_BG     = "FFF3CD"
    YESIL_BG    = "D4EDDA"
    KIRMIZI_BG  = "F8D7DA"
    GRI_BG      = "F8F9FA"
    KENAR       = Side(style='thin', color='CCCCCC')

    def border(c):
        c.border = Border(
            left=KENAR, right=KENAR, top=KENAR, bottom=KENAR
        )
        return c

    def baslik_satiri(ws, row, text, col_start=1, col_end=6, bg=BASLIK_BG):
        c = ws.cell(row=row, column=col_start, value=text)
        c.font = Font(bold=True, color=BASLIK_FG, size=12)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row, end_column=col_end)
        ws.row_dimensions[row].height = 24
        return c

    # ════════════════════════════════════════════════════════════
    # SAYFA 1: ÖZET
    # ════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Özet"
    ws1.column_dimensions['A'].width = 32
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 20

    # Başlık
    baslik_satiri(ws1, 1, f"EVVEL ERP — AYLIK FİNANSAL RAPOR", 1, 3)
    baslik_satiri(ws1, 2, f"{rapor['donem_label']} · {rapor['donem']}", 1, 3, ALT_BASLIK)
    ws1.row_dimensions[2].height = 20

    r = 4
    baslik_satiri(ws1, r, "KASA ÖZETİ", 1, 3, "2C3E50"); r += 1

    for label, key, bg in [
        ("Ay Başı Kasa",    "baslangic_kasa", GRI_BG),
        ("Toplam Gelir",    "toplam_gelir",   YESIL_BG),
        ("Toplam Gider",    "toplam_gider",   KIRMIZI_BG),
        # FIX MN7 (2026-07-06): kasa-bazlı sayı = nakit akışı, kâr değil (kredi girişi kârı şişirir)
        ("Net Nakit Akışı (kasa)", "net_kar_zarar",  YESIL_BG if float(o.get("net_kar_zarar",0)) >= 0 else KIRMIZI_BG),
        ("Ay Sonu Kasa",    "bitis_kasa",     SARI_BG),
    ]:
        val = float(o.get(key, 0) or 0)
        c1 = ws1.cell(row=r, column=1, value=label)
        c1.fill = PatternFill("solid", fgColor=bg)
        c1.font = Font(bold=True, size=11)
        c1.alignment = Alignment(horizontal='left', vertical='center')
        c2 = ws1.cell(row=r, column=2, value=val)
        c2.number_format = '#,##0'
        c2.fill = PatternFill("solid", fgColor=bg)
        c2.font = Font(bold=True, size=11)
        c2.alignment = Alignment(horizontal='right', vertical='center')
        border(c1); border(c2)
        r += 1

    r += 1
    baslik_satiri(ws1, r, "GELİR DAĞILIMI", 1, 3, "27AE60"); r += 1
    for label, key in [
        ("Nakit Ciro",       "ciro_nakit"),
        ("POS Ciro",         "ciro_pos"),
        ("Online Ciro",      "ciro_online"),
        ("Dış Kaynak",       "dis_kaynak_toplam"),
    ]:
        val = float(o.get(key, 0) or 0)
        if val == 0: continue
        c1 = ws1.cell(row=r, column=1, value=label)
        c1.alignment = Alignment(horizontal='left')
        c2 = ws1.cell(row=r, column=2, value=val)
        c2.number_format = '#,##0'
        c2.alignment = Alignment(horizontal='right')
        toplam = float(o.get("toplam_gelir", 1) or 1)
        c3 = ws1.cell(row=r, column=3, value=f"%{round(val/toplam*100)}")
        c3.alignment = Alignment(horizontal='right')
        border(c1); border(c2); border(c3)
        r += 1

    r += 1
    baslik_satiri(ws1, r, "GİDER DAĞILIMI", 1, 3, "E74C3C"); r += 1
    for label, key in [
        ("Kart Ödemeleri",    "kart_toplam"),
        ("Anlık Giderler",    "anlik_toplam"),
        ("Personel Maaşları", "maas_toplam"),
        ("Sabit Giderler",    "sabit_toplam"),
        ("Vadeli Ödemeler",   "vadeli_toplam"),
        ("Borç Taksitleri",   "borc_taksit_toplam"),
        ("Fatura Giderleri",  "fatura_toplam"),
        ("Kart Faizi",        "kart_faiz_toplam"),
        ("POS Komisyon",      "pos_kesinti_toplam"),
    ]:
        val = float(o.get(key, 0) or 0)
        if val == 0: continue
        c1 = ws1.cell(row=r, column=1, value=label)
        c1.alignment = Alignment(horizontal='left')
        c2 = ws1.cell(row=r, column=2, value=val)
        c2.number_format = '#,##0'
        c2.alignment = Alignment(horizontal='right')
        toplam_g = float(o.get("toplam_gider", 1) or 1)
        c3 = ws1.cell(row=r, column=3, value=f"%{round(val/toplam_g*100)}")
        c3.alignment = Alignment(horizontal='right')
        border(c1); border(c2); border(c3)
        r += 1

    # ════════════════════════════════════════════════════════════
    # SAYFA 2: GÜNLÜK KASA
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Günlük Kasa")
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 16
    ws2.column_dimensions['E'].width = 16

    baslik_satiri(ws2, 1, f"GÜNLÜK KASA SEYRİ — {rapor['donem_label']}", 1, 5)
    r = 2
    headers = ["Tarih", "Giriş", "Çıkış", "Net", "Kümülatif Kasa"]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=r, column=col, value=h)
        c.font = Font(bold=True, color=BASLIK_FG, size=10)
        c.fill = PatternFill("solid", fgColor=ALT_BASLIK)
        c.alignment = Alignment(horizontal='center', vertical='center')
        border(c)
    r += 1

    for g in (rapor.get("gunluk") or []):
        giris = float(g.get("giris", 0) or 0)
        cikis = float(g.get("cikis", 0) or 0)
        net   = float(g.get("net", 0) or 0)
        kasa  = float(g.get("kasa", 0) or 0)
        tarih = str(g.get("tarih",""))
        bg = KIRMIZI_BG if kasa < 0 else ("FFFFFF" if r % 2 == 0 else GRI_BG)
        for col, val in enumerate([tarih, giris, cikis, net, kasa], 1):
            c = ws2.cell(row=r, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=bg)
            if col > 1: c.number_format = '#,##0'
            c.alignment = Alignment(horizontal='right' if col>1 else 'left', vertical='center')
            border(c)
        r += 1

    # ════════════════════════════════════════════════════════════
    # SAYFA 3: ŞUBE CİRO
    # ════════════════════════════════════════════════════════════
    if rapor.get("sube_ciro"):
        ws3 = wb.create_sheet("Şube Ciro")
        for col, w in zip('ABCDEF', [20,16,14,14,14,10]):
            ws3.column_dimensions[col].width = w
        baslik_satiri(ws3, 1, f"ŞUBE BAZLI CİRO — {rapor['donem_label']}", 1, 6)
        r = 2
        for col, h in enumerate(["Şube","Toplam Ciro","Nakit","POS","Online","İşlem"], 1):
            c = ws3.cell(row=r, column=col, value=h)
            c.font = Font(bold=True, color=BASLIK_FG, size=10)
            c.fill = PatternFill("solid", fgColor=ALT_BASLIK)
            c.alignment = Alignment(horizontal='center', vertical='center')
            border(c)
        r += 1
        for s in rapor["sube_ciro"]:
            for col, val in enumerate([
                s.get("sube",""), float(s.get("ciro",0) or 0),
                float(s.get("nakit",0) or 0), float(s.get("pos",0) or 0),
                float(s.get("online",0) or 0), int(s.get("islem_sayisi",0) or 0)
            ], 1):
                c = ws3.cell(row=r, column=col, value=val)
                if col > 1: c.number_format = '#,##0'
                c.alignment = Alignment(horizontal='right' if col>1 else 'left')
                border(c)
            r += 1

    # ════════════════════════════════════════════════════════════
    # SAYFA 4: GİDER DETAYLARI
    # ════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Gider Detayları")
    ws4.column_dimensions['A'].width = 10
    ws4.column_dimensions['B'].width = 30
    ws4.column_dimensions['C'].width = 16
    ws4.column_dimensions['D'].width = 14
    ws4.column_dimensions['E'].width = 16

    r = 1
    # Sabit giderler
    if rapor.get("sabit_detay"):
        baslik_satiri(ws4, r, "SABİT GİDERLER", 1, 5, "2980B9"); r += 1
        for col, h in enumerate(["Tarih","Gider Adı","Kategori","Ödenen"], 1):
            c = ws4.cell(row=r, column=col, value=h)
            c.font = Font(bold=True, color=BASLIK_FG, size=10)
            c.fill = PatternFill("solid", fgColor=ALT_BASLIK)
            border(c)
        r += 1
        toplam_sabit = 0
        for g in rapor["sabit_detay"]:
            odenen = float(g.get("odenen", 0) or 0)
            toplam_sabit += odenen
            for col, val in enumerate([
                str(g.get("odeme_tarihi",""))[:10],
                g.get("gider_adi",""), g.get("kategori",""), odenen
            ], 1):
                c = ws4.cell(row=r, column=col, value=val)
                if col == 4: c.number_format = '#,##0'
                border(c)
            r += 1
        c = ws4.cell(row=r, column=3, value="TOPLAM")
        c.font = Font(bold=True)
        c2 = ws4.cell(row=r, column=4, value=toplam_sabit)
        c2.number_format = '#,##0'; c2.font = Font(bold=True)
        r += 2

    # Personel giderleri
    if rapor.get("personel_detay"):
        baslik_satiri(ws4, r, "PERSONEL MAAŞLARI", 1, 5, "8E44AD"); r += 1
        for col, h in enumerate(["Tarih","Ad Soyad","Görev","Ödenen"], 1):
            c = ws4.cell(row=r, column=col, value=h)
            c.font = Font(bold=True, color=BASLIK_FG, size=10)
            c.fill = PatternFill("solid", fgColor=ALT_BASLIK)
            border(c)
        r += 1
        toplam_maas = 0
        for p in rapor["personel_detay"]:
            odenen = float(p.get("odenen", 0) or 0)
            toplam_maas += odenen
            for col, val in enumerate([
                str(p.get("odeme_tarihi",""))[:10],
                p.get("ad_soyad",""), p.get("gorev",""), odenen
            ], 1):
                c = ws4.cell(row=r, column=col, value=val)
                if col == 4: c.number_format = '#,##0'
                border(c)
            r += 1
        c = ws4.cell(row=r, column=3, value="TOPLAM")
        c.font = Font(bold=True)
        c2 = ws4.cell(row=r, column=4, value=toplam_maas)
        c2.number_format = '#,##0'; c2.font = Font(bold=True)
        r += 2

    # Anlık gider kategorileri
    if rapor.get("anlik_kategoriler"):
        baslik_satiri(ws4, r, "ANLIK GİDER KATEGORİLERİ", 1, 5, "D35400"); r += 1
        for col, h in enumerate(["Kategori","İşlem Adedi","Toplam"], 1):
            c = ws4.cell(row=r, column=col, value=h)
            c.font = Font(bold=True, color=BASLIK_FG, size=10)
            c.fill = PatternFill("solid", fgColor=ALT_BASLIK)
            border(c)
        r += 1
        for g in rapor["anlik_kategoriler"]:
            for col, val in enumerate([
                g.get("kategori",""), int(g.get("adet",0) or 0),
                float(g.get("toplam",0) or 0)
            ], 1):
                c = ws4.cell(row=r, column=col, value=val)
                if col == 3: c.number_format = '#,##0'
                border(c)
            r += 1
        r += 1

    # Kart ödemeleri
    if rapor.get("kart_detay"):
        baslik_satiri(ws4, r, "KART ÖDEMELERİ", 1, 5, "C0392B"); r += 1
        for col, h in enumerate(["Kart","Banka","Ödeme Adedi","Anapara"], 1):
            c = ws4.cell(row=r, column=col, value=h)
            c.font = Font(bold=True, color=BASLIK_FG, size=10)
            c.fill = PatternFill("solid", fgColor=ALT_BASLIK)
            border(c)
        r += 1
        for k in rapor["kart_detay"]:
            for col, val in enumerate([
                k.get("kart_adi",""), k.get("banka",""),
                int(k.get("adet",0) or 0), float(k.get("anapara",0) or 0)
            ], 1):
                c = ws4.cell(row=r, column=col, value=val)
                if col == 4: c.number_format = '#,##0'
                border(c)
            r += 1

    # Excel'i belleğe yaz
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    dosya_adi = f"evvel-rapor-{rapor['donem']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={dosya_adi}"}
    )


def health():
    return {
        "status": "ok",
        "version": "EVVEL-ERP-2.0",
        "build": "v2.4",
        "build_date": "2026-03-27",
        "features": [
            "odeme_tipi_duzeltme",
            "kira_artis_periyot",
            "sozlesme_stop",
            "inline_edit_modal",
            "durdurulmus_sabit_zorunlu",
            "plan_iptal_on_stop",
        ]
    }


@app.post("/api/sistem-sifirla")
def sistem_sifirla(body: dict = {}):
    """Seçili tabloları siler. body: {onay: 'EVET_SIL', tablolar: [...]}"""
    if body.get('onay') != 'EVET_SIL':
        raise HTTPException(400, "Onay gerekli")

    # İzin verilen tablolar — şubeler, kartlar, master katalog (merkez_stok_kart,
    # siparis_kategori, siparis_urun, tedarikciler) ve vardiya kural tabloları asla silinmez.
    IZINLI = {
        # ── Finans / Muhasebe ─────────────────────────────────────────
        'ciro':                 'ciro',
        'ciro_taslak':          'ciro_taslak',
        'kasa':                 'kasa_hareketleri',
        'kart_hareketleri':     'kart_hareketleri',
        'banka_yatirimlari':    'banka_yatirimlari',
        'anlik_gider':          'anlik_giderler',
        'vadeli_alim':          'vadeli_alimlar',
        'sabit_gider':          'sabit_giderler',
        'borc':                 'borc_envanteri',
        'odeme_plani':          'odeme_plani',
        'onay_kuyrugu':         'onay_kuyrugu',
        'x_rapor_kayit':        'x_rapor_kayit',
        # ── Personel (aylık + ataması; personel tanımı ayrı) ──────────
        'personel':             'personel',
        'personel_aylik':       'personel_aylik',
        'personel_risk_sinyal': 'personel_risk_sinyal',
        'personel_takip':       'personel_takip',
        # ── Şube operasyon (test verileri için) ───────────────────────
        'sube_acilis':          'sube_acilis',
        'sube_operasyon_event': 'sube_operasyon_event',
        'sube_operasyon_uyari': 'sube_operasyon_uyari',
        'sube_operasyon_ozet':  'sube_operasyon_ozet',
        'sube_fire_haftalik':   'sube_fire_haftalik',
        'operasyon_defter':     'operasyon_defter',
        'sube_kasa_gun_acma':   'sube_kasa_gun_acma',
        'kapanis_kayit':        'kapanis_kayit',
        'sube_depo_stok':       'sube_depo_stok',
        'stok_yolda':           'stok_yolda',
        'sube_skor':            'sube_skor',
        'sube_merkez_mesaj':    'sube_merkez_mesaj',
        'sube_merkez_not':      'sube_merkez_not',
        'panel_pin_guvenlik':   'panel_pin_guvenlik',
        'operasyon_guvenlik_olay':       'operasyon_guvenlik_olay',
        'operasyon_guvenlik_alarm_durum':'operasyon_guvenlik_alarm_durum',
        'motor_analitik_olay':  'motor_analitik_olay',
        'kasa_teslim':          'kasa_teslim',
        # ── Sipariş akışı ─────────────────────────────────────────────
        'siparis_talep':        'siparis_talep',
        'siparis_ozel_talep':   'siparis_ozel_talep',
        'siparis_sevk_eksik':   'siparis_sevk_eksik',
        'merkez_stok_sevk':     'merkez_stok_sevk',
        # ── Denetim ──────────────────────────────────────────────────
        # 🔴 VERI-011 (2026-09-02): 'audit_log' BU LİSTEDEN ÇIKARILDI.
        # Veriyi silen düğmenin, o silmenin izini tutan defteri de silebilmesi
        # denetimin kendisini anlamsız kılar — silinen şeyin silindiğini
        # söyleyecek tek kayıt oydu. Denetim defteri veriyle birlikte gitmez.
    }

    # ⚠️ KRİTİK ANAHTARLAR — "boş liste = hepsi" kısayoluna DAHİL DEĞİL.
    # kasa_teslim para emanetinin (custody) zincirini tutar: kimin kasasında
    # ne kadar para olduğunu söyleyen tek kayıt. Silinebilir kalsın (demo
    # sıfırlaması meşru), ama ASLA "hepsini temizle" ile yanlışlıkla değil —
    # adı adıyla istenmeli.
    KRITIK = {'kasa_teslim'}

    # EVV-SIS (2026-08-15, Codex + diff-review 2 katman): v2 `tablolar: []`
    # gönderince 400 dönüyordu ("Sıfırla" fiilen ölüydü). FAIL-CLOSED sıkı hali:
    # yalnız AÇIK boş liste [] = hepsi; liste-dışı her tip (None/""/0) → 400;
    # bilinmeyen tablo anahtarı sessizce filtrelenmez, İSTEK REDDEDİLİR.
    _istenen_ham = body.get('tablolar', [])
    if not isinstance(_istenen_ham, list):
        raise HTTPException(400, "tablolar bir liste olmalı (boş liste = tüm izinli tablolar)")
    _bilinmeyen = [k for k in _istenen_ham if k not in IZINLI]
    if _bilinmeyen:
        raise HTTPException(400, f"Bilinmeyen tablo anahtarı: {', '.join(str(b)[:30] for b in _bilinmeyen[:5])}")
    istenen = _istenen_ham if _istenen_ham else [k for k in IZINLI.keys() if k not in KRITIK]
    silincekler = [IZINLI[k] for k in istenen if k in IZINLI]

    if not silincekler:
        raise HTTPException(400, "Silinecek tablo seçilmedi")

    # 🔴 VERI-002/012 (2026-09-02): bu ucun TEK koruması istemcinin gönderdiği
    # `onay:'EVET_SIL'` metniydi. O metin frontend kaynağında yazılı — yani
    # kapı değil, tabela. Aynı ekrandaki kasa açılışı PIN isterken TRUNCATE
    # istemiyordu; `ledger-sifirla` ise PIN'liydi. Tutarsızlık kapandı.
    from operasyon_merkez_api import _isletme_onay_dogrula

    # 🧪 KURU ÇALIŞTIRMA: ne silineceğini ÖNCE saydırıp okumadan yıkıcı
    # işlem çalıştırılmaz. kuru=true PIN istemez (hiçbir şey silmez).
    _kuru = bool(body.get('kuru'))

    with db() as (conn, cur):
        if not _kuru:
            onayci = _isletme_onay_dogrula(cur, body.get('onay_pin'))  # PIN hatalı → 403

        # Sayım: silinmeden ÖNCE, hem kuru modda hem gerçek silmede.
        sayimlar = {}
        for _t in silincekler:
            try:
                with savepoint(cur, "sp_sayim"):
                    cur.execute(f"SELECT COUNT(*) AS c FROM {_t}")
                    sayimlar[_t] = int((cur.fetchone() or {}).get("c") or 0)
            except Exception:
                sayimlar[_t] = None   # tablo yoksa sıfır yazmıyoruz — bilinmiyor
        _toplam = sum(v for v in sayimlar.values() if isinstance(v, int))

        if _kuru:
            return {"basarili": True, "kuru": True, "silinecek": silincekler,
                    "satir_sayilari": sayimlar, "toplam_satir": _toplam,
                    "mesaj": (f"KURU ÇALIŞTIRMA — hiçbir şey silinmedi. "
                              f"{len(silincekler)} tabloda {_toplam} satır silinecekti.")}

        # İz ÖNCE yazılır: TRUNCATE'ten sonra sayıları okuyacak kimse kalmaz.
        audit(cur, 'sistem', 'sistem-sifirla', 'SISTEM_SIFIRLA',
              yeni={'tablolar': silincekler, 'satir_sayilari': sayimlar,
                    'toplam_satir': _toplam},
              aktor=onayci.get('ad_soyad'), aktor_id=str(onayci.get('id')),
              aktor_kaynak='isletme_pin')
        cur.execute(f"TRUNCATE TABLE {', '.join(silincekler)} CASCADE")

    return {"basarili": True, "silinen": silincekler,
            "satir_sayilari": sayimlar, "toplam_satir": _toplam,
            "onaylayan": onayci.get('ad_soyad'),
            "mesaj": f"{len(silincekler)} tablo temizlendi ({_toplam} satır)."}

# Şube personel paneli HTML (SPA catch-all'dan önce).
# Üretim imajında dosya yalnızca static/'tedir (Dockerfile cp). Geliştiricide düzenlenen
# kaynak genelde kökteki sube_panel.html olduğundan — varsa kök önceliklidir; böylece
# /sube-panel yanlışlıkla React SPA catch-all'a düşmez ve eski static kopyası ezmez.
_kok_panel = pathlib.Path("sube_panel.html")
_static_panel = pathlib.Path("static/sube_panel.html")
_sube_panel_path = _kok_panel if _kok_panel.is_file() else _static_panel
if _sube_panel_path.exists():
    from fastapi.responses import FileResponse as _FileResponseSube

    _sube_headers = {
        "Cache-Control": "no-cache, no-store, must-revalidate",
        "Pragma": "no-cache",
        "Expires": "0",
    }

    @app.get("/sube-panel")
    @app.get("/sube-panel/{sube_id:path}")
    async def serve_sube_panel_html(sube_id: str = ""):
        _ = sube_id
        return _FileResponseSube(
            str(_sube_panel_path),
            media_type="text/html",
            headers=_sube_headers,
        )

    @app.get("/sube")
    @app.get("/sube/{sube_id:path}")
    async def serve_sube_personel_html(sube_id: str = ""):
        _ = sube_id
        return _FileResponseSube(
            str(_sube_panel_path),
            media_type="text/html",
            headers=_sube_headers,
        )

# ══════════════════════════════════════════════════════════════
# VARDİYA v2 — SLOT BAZLI YENİ SİSTEM
# ══════════════════════════════════════════════════════════════
class _V2SlotIn(BaseModel):
    sube_id: str
    ad: str
    tip: str = "normal"
    baslangic_saat: str   # "HH:MM"
    bitis_saat: str
    gece_vardiyasi: bool = False
    min_personel: int = 1
    ideal_personel: int = 1
    aktif_gunler: List[int] = Field(default_factory=lambda: [1,2,3,4,5,6,7])
    aktif: bool = True
    sira: int = 0


class _V2SlotUretIn(BaseModel):
    """Şube açılış/kapanış/yoğun saatlerinden AUTO: slot üretimi."""
    sube_id: str
    mod: str = "yenile"  # yenile | ekle
    acilis_dakika: int = 60
    kapanis_dakika: int = 60
    normal_slot_dakika: int = 120
    hafta_ici: bool = False
    aktif_gunler: Optional[List[int]] = None


class _V2KisitIn(BaseModel):
    max_gunluk_saat: float = 9.5
    max_haftalik_saat: float = 57.0
    izinli_subeler: List[str] = Field(default_factory=list)
    yasak_subeler: List[str] = Field(default_factory=list)
    calisilabilir_saat_min: Optional[str] = None
    calisilabilir_saat_max: Optional[str] = None
    min_gecis_dk: int = 30
    # YENİ: hibrit preset + ders saatleri + yemek molası
    vardiya_preset_json: Dict[str, Any] = Field(default_factory=dict)
    gun_saat_kisitlari_json: Dict[str, Any] = Field(default_factory=dict)
    yemek_sube_id: Optional[str] = None

class _V2AtamaIn(BaseModel):
    personel_id: str
    slot_id: str
    tarih: str            # "YYYY-MM-DD"
    baslangic_saat: Optional[str] = None
    bitis_saat: Optional[str] = None
    override: bool = False
    aciklama: Optional[str] = None
    # True: saatler boşsa preset/mesai/kısa dilim (otomatik doldur / motor / sürükle önizleme)
    otomatik_saat_cozumu: bool = False
    # True: kayıt `onayli` — taşınamaz kilidi için UI (otomatik planlar varsayılan `planli`)
    kesinlestir: bool = False


class _MotorHaftaIn(BaseModel):
    """Haftalık otomatik plan motoru (`vardiya_plan_motor`)."""
    pazartesi: str              # ISO gün — haftanın herhangi bir günü olabilir (Pzt’ye normalize edilir)
    max_rounds: int = 120
    tasima_izni: bool = True    # Başka slottan taşımayı dene (min kontenjan korunur)
    dry_run: bool = False       # True ise işlem sonunda rollback (önizleme)


class _V2SubeGunHedefIn(BaseModel):
    """Şube × gün hedef kişi; `hedef_personel` null/omit → hedef kaldırılır."""
    sube_id: str
    tarih: str
    hedef_personel: Optional[int] = None


class _V2PersonelGunNiyetIn(BaseModel):
    personel_id: str
    tarih: str
    kasitli_bos: bool


class _V2GunKilitIn(BaseModel):
    tarih: str
    kilitli: bool = True
    aciklama: Optional[str] = None


class _V2IzinIn(BaseModel):
    personel_id: str
    baslangic_tarih: str
    bitis_tarih: str
    tip: str = "mazeret"
    aciklama: Optional[str] = None
    # Aynı ISO haftasında ikinci kayıt: 409 sonrası kullanıcı onayı ile True
    force: bool = False


def _t(s: Optional[str]):
    if not s:
        return None
    h, m = str(s).split(":")[:2]
    from datetime import time as _time
    return _time(int(h), int(m))


# ── SLOT CRUD ──
@app.get("/api/vardiya/v2/slot")
def v2_slot_liste(sube_id: Optional[str] = None, aktif_mi: Optional[bool] = None):
    with db() as (conn, cur):
        sql = "SELECT * FROM vardiya_slot WHERE 1=1"
        params: List = []
        if sube_id:
            sql += " AND sube_id = %s"; params.append(sube_id)
        if aktif_mi is not None:
            sql += " AND aktif = %s"; params.append(aktif_mi)
        sql += " ORDER BY sube_id, sira, baslangic_saat"
        cur.execute(sql, tuple(params))
        return {"slotlar": [dict(r) for r in cur.fetchall()]}

@app.post("/api/vardiya/v2/slot")
def v2_slot_ekle(s: _V2SlotIn):
    sid = str(uuid.uuid4())
    with db() as (conn, cur):
        cur.execute("""
            INSERT INTO vardiya_slot
                (id, sube_id, ad, tip, baslangic_saat, bitis_saat, gece_vardiyasi,
                 min_personel, ideal_personel, aktif_gunler, aktif, sira)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (sid, s.sube_id, s.ad, s.tip, _t(s.baslangic_saat), _t(s.bitis_saat),
              s.gece_vardiyasi, s.min_personel, s.ideal_personel,
              s.aktif_gunler, s.aktif, s.sira))
    return {"id": sid}

@app.put("/api/vardiya/v2/slot/{sid}")
def v2_slot_guncelle(sid: str, s: _V2SlotIn):
    with db() as (conn, cur):
        cur.execute("""
            UPDATE vardiya_slot SET
                sube_id=%s, ad=%s, tip=%s, baslangic_saat=%s, bitis_saat=%s,
                gece_vardiyasi=%s, min_personel=%s, ideal_personel=%s,
                aktif_gunler=%s, aktif=%s, sira=%s
            WHERE id=%s
        """, (s.sube_id, s.ad, s.tip, _t(s.baslangic_saat), _t(s.bitis_saat),
              s.gece_vardiyasi, s.min_personel, s.ideal_personel,
              s.aktif_gunler, s.aktif, s.sira, sid))
        if cur.rowcount == 0:
            raise HTTPException(404, "Slot bulunamadı")
    return {"basarili": True}

@app.delete("/api/vardiya/v2/slot/{sid}")
def v2_slot_sil(sid: str):
    with db() as (conn, cur):
        cur.execute("DELETE FROM vardiya_slot WHERE id = %s", (sid,))
        if cur.rowcount == 0:
            raise HTTPException(404, "Slot bulunamadı")
    return {"basarili": True}


@app.post("/api/vardiya/v2/slot/uret")
def v2_slot_sube_saatlerinden_uret(body: _V2SlotUretIn):
    """Aşama 2 slot motoru: `subeler.acilis_saati` / `kapanis_saati` / yoğun aralığı → `AUTO:` slotlar."""
    with db() as (conn, cur):
        sonuc = _vv2.slotlari_sube_saatlerinden_uret(
            cur,
            body.sube_id,
            mod=body.mod,
            acilis_dakika=body.acilis_dakika,
            kapanis_dakika=body.kapanis_dakika,
            normal_slot_dakika=body.normal_slot_dakika,
            aktif_gunler=body.aktif_gunler,
            hafta_ici=body.hafta_ici,
        )
    if not sonuc.get("basarili"):
        raise HTTPException(409, sonuc)
    return sonuc


# ── PERSONEL KISIT ──
@app.get("/api/vardiya/v2/kisit/{pid}")
def v2_kisit_getir(pid: str):
    with db() as (conn, cur):
        return _vv2.personel_kisit_getir(cur, pid)

@app.put("/api/vardiya/v2/kisit/{pid}")
def v2_kisit_kaydet(pid: str, k: _V2KisitIn):
    """
    Eski PostgreSQL kurulumlarında `personel_kisit` tablosunda ayrı NOT NULL `id`
    kolonu olabiliyor; INSERT `id` vermezse NotNullViolation oluşur. Kolon varsa
    UUID üretilir.
    """
    with db() as (conn, cur):
        cur.execute(
            """
            SELECT 1 FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = 'personel_kisit'
              AND column_name = 'id'
            """
        )
        has_id_col = cur.fetchone() is not None
        tmin = _t(k.calisilabilir_saat_min)
        tmax = _t(k.calisilabilir_saat_max)
        if has_id_col:
            kid = str(uuid.uuid4())
            cur.execute(
                """
                INSERT INTO personel_kisit
                    (id, personel_id, max_gunluk_saat, max_haftalik_saat,
                     izinli_subeler, yasak_subeler,
                     calisilabilir_saat_min, calisilabilir_saat_max, min_gecis_dk)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (personel_id) DO UPDATE SET
                    max_gunluk_saat = EXCLUDED.max_gunluk_saat,
                    max_haftalik_saat = EXCLUDED.max_haftalik_saat,
                    izinli_subeler = EXCLUDED.izinli_subeler,
                    yasak_subeler = EXCLUDED.yasak_subeler,
                    calisilabilir_saat_min = EXCLUDED.calisilabilir_saat_min,
                    calisilabilir_saat_max = EXCLUDED.calisilabilir_saat_max,
                    min_gecis_dk = EXCLUDED.min_gecis_dk,
                    guncelleme = NOW()
                """,
                (
                    kid,
                    pid,
                    k.max_gunluk_saat,
                    k.max_haftalik_saat,
                    k.izinli_subeler,
                    k.yasak_subeler,
                    tmin,
                    tmax,
                    k.min_gecis_dk,
                ),
            )
        else:
            cur.execute(
                """
                INSERT INTO personel_kisit
                    (personel_id, max_gunluk_saat, max_haftalik_saat,
                     izinli_subeler, yasak_subeler,
                     calisilabilir_saat_min, calisilabilir_saat_max, min_gecis_dk)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (personel_id) DO UPDATE SET
                    max_gunluk_saat = EXCLUDED.max_gunluk_saat,
                    max_haftalik_saat = EXCLUDED.max_haftalik_saat,
                    izinli_subeler = EXCLUDED.izinli_subeler,
                    yasak_subeler = EXCLUDED.yasak_subeler,
                    calisilabilir_saat_min = EXCLUDED.calisilabilir_saat_min,
                    calisilabilir_saat_max = EXCLUDED.calisilabilir_saat_max,
                    min_gecis_dk = EXCLUDED.min_gecis_dk,
                    guncelleme = NOW()
                """,
                (
                    pid,
                    k.max_gunluk_saat,
                    k.max_haftalik_saat,
                    k.izinli_subeler,
                    k.yasak_subeler,
                    tmin,
                    tmax,
                    k.min_gecis_dk,
                ),
            )
        # YENİ alanlar (preset, ders saatleri, yemek molası) — UPDATE ayrı.
        # Eski/yeni schema fark etmez; kolonlar yoksa migration zaten ekleyecek.
        import json as _json
        try:
            cur.execute(
                """
                UPDATE personel_kisit
                SET vardiya_preset_json = %s::jsonb,
                    gun_saat_kisitlari_json = %s::jsonb,
                    yemek_sube_id = %s,
                    guncelleme = NOW()
                WHERE personel_id = %s
                """,
                (
                    _json.dumps(k.vardiya_preset_json or {}),
                    _json.dumps(k.gun_saat_kisitlari_json or {}),
                    k.yemek_sube_id,
                    pid,
                ),
            )
        except Exception:
            # Eski schema'da kolon yoksa sessiz geç (migration sonrasında çalışacak)
            pass
    return {"basarili": True}


# ── ATAMA ──
@app.get("/api/vardiya/v2/gun")
def v2_gun_planini_getir(tarih: str, sube_id: Optional[str] = None):
    from datetime import datetime as _dt
    t = _dt.strptime(tarih, "%Y-%m-%d").date()
    with db() as (conn, cur):
        return _vv2.gun_planini_getir(cur, t, sube_id)


class _V2GecmisKilitIn(BaseModel):
    pin: str


@app.post("/api/vardiya/v2/gecmis-kilit-ac")
def v2_gecmis_kilit_ac(body: _V2GecmisKilitIn):
    """Geçmiş haftada vardiya düzenleme KİLİDİ — kasın hatayla geçmişi bozmaması için.
    İşletme (Merve Karabacak / sahip) 4 haneli PIN'i ile açılır. PIN hatalıysa 403."""
    from operasyon_merkez_api import _isletme_onay_dogrula
    with db() as (conn, cur):
        onayci = _isletme_onay_dogrula(cur, body.pin)  # PIN hatalı → 403
    return {"ok": True, "onayci": onayci.get("ad_soyad")}


@app.put("/api/vardiya/v2/sube-gun-hedef")
def v2_sube_gun_hedef_kaydet(body: _V2SubeGunHedefIn):
    """Şube için seçilen güne hedef kişi sayısı (vardiya_sube_gun_hedef)."""
    from datetime import datetime as _dt
    t = _dt.strptime(body.tarih[:10], "%Y-%m-%d").date()
    with db() as (conn, cur):
        _vv2.sube_gun_hedef_kaydet(cur, body.sube_id, t, body.hedef_personel)
    return {"basarili": True}


@app.get("/api/vardiya/v2/hafta-personel-tablo")
def v2_hafta_personel_tablo(pazartesi: str):
    """Seçilen haftanın Pazartesi tarihi (veya hafta içi herhangi bir gün — sunucu Pazartesi’ye normalize eder)."""
    from datetime import datetime as _dt
    t = _dt.strptime(pazartesi[:10], "%Y-%m-%d").date()
    with db() as (conn, cur):
        return _vv2.hafta_personel_tablosu(cur, t)


@app.get("/api/vardiya/v2/hafta-sube-tablo")
def v2_hafta_sube_tablo(pazartesi: str):
    """Şube × hafta görünümü — her şube için 7 gün, her günde çalışan personel listesi."""
    from datetime import datetime as _dt
    t = _dt.strptime(pazartesi[:10], "%Y-%m-%d").date()
    with db() as (conn, cur):
        return _vv2.sube_haftalik_gorunum(cur, t)


@app.post("/api/vardiya/v2/motor/hafta-doldur")
def v2_motor_hafta_doldur(body: _MotorHaftaIn):
    """Haftalık plan motoru — eksik slotları önceliklendirir; doğrudan veya taşıma ile doldurur."""
    import vardiya_plan_motor as _vpm
    from datetime import datetime as _dt
    pzt = _dt.strptime(body.pazartesi[:10], "%Y-%m-%d").date()
    mr = max(15, min(int(body.max_rounds or 120), 400))
    with db() as (conn, cur):
        out = _vpm.hafta_otomatik_planla(
            cur,
            pazartesi_gun=pzt,
            max_rounds=mr,
            tasima_izni=bool(body.tasima_izni),
            kullanici_id=None,
            aciklama_etiketi="[Otomatik plan motoru]",
        )
        if getattr(body, "dry_run", False):
            conn.rollback()
            out["dry_run"] = True
            base = out.get("mesaj") or ""
            out["mesaj"] = base + " (dry-run — veritabanı geri alındı)"
    return out


@app.post("/api/vardiya/v2/atama/check")
def v2_atama_check(a: _V2AtamaIn):
    """Atama yapılmadan önce uyarıları döner — UI bunu çağırıp gösterir.

    Şube çerçevesi taşması (`slot_band_disinda`) uyarıdır; kritik blok listesinde yer almaz.
    Esas zamanlar istek gövdesindeki / önerilen atama başlangıç–bitişidir (`vardiya_v2` docstring).
    """
    from datetime import datetime as _dt
    t = _dt.strptime(a.tarih, "%Y-%m-%d").date()
    with db() as (conn, cur):
        uyarilar = _vv2.atama_uyarilari(
            cur, a.personel_id, a.slot_id, t,
            _t(a.baslangic_saat), _t(a.bitis_saat),
            otomatik_saat_cozumu=bool(a.otomatik_saat_cozumu),
        )
        gd = _vv2.personel_gun_durumu(cur, a.personel_id, t)
    has_cakisma = any(u.get("tip") == "cakisma" for u in uyarilar)
    # Override yalnızca çakışma dışı kritikler için; çakışma fiziksel blok (override edilemez).
    override_gerekir = any(
        u.get("seviye") == "kritik" and u.get("tip") != "cakisma" for u in uyarilar
    )
    return {
        "uyarilar": uyarilar,
        "kritik_var": any(u["seviye"] == "kritik" for u in uyarilar),
        "cakisma_var": has_cakisma,
        "override_gerekir": override_gerekir,
        "personel_gun": {
            "durum": gd.get("durum"),
            "kalan_saat": gd.get("kalan_saat"),
            "toplam_saat": gd.get("toplam_saat"),
            "max_gunluk_saat": gd.get("max_gunluk_saat", 9.0),
            "atama_sayisi": gd.get("atama_sayisi", 0),
        },
    }

@app.post("/api/vardiya/v2/atama")
def v2_atama_olustur(a: _V2AtamaIn):
    from datetime import datetime as _dt
    t = _dt.strptime(a.tarih, "%Y-%m-%d").date()
    with db() as (conn, cur):
        sonuc = _vv2.atama_olustur(
            cur, a.personel_id, a.slot_id, t,
            _t(a.baslangic_saat), _t(a.bitis_saat),
            override=a.override,
            aciklama=a.aciklama,
            otomatik_saat_cozumu=bool(a.otomatik_saat_cozumu),
            durum="onayli" if bool(a.kesinlestir) else "planli",
        )
        if not sonuc.get('basarili'):
            raise HTTPException(409, sonuc)
        return sonuc


@app.post("/api/vardiya/v2/assign")
def v2_assign(a: _V2AtamaIn):
    """Önce atama/check, sonra atama — spec’te sık geçen /assign ismi için POST alias (gövde /atama ile aynı)."""
    return v2_atama_olustur(a)


class _V2AtamaSerbestIn(BaseModel):
    """Serbest-saat atama: slot seçmeden, şube + özel saatle. Gün penceresi atamaya göre şekillenir."""
    personel_id: str
    sube_id: str
    tarih: str            # "YYYY-MM-DD"
    baslangic_saat: str   # "08:30"
    bitis_saat: str       # "01:30" → ertesi gün (gece otomatik)
    override: bool = False
    aciklama: Optional[str] = None
    kesinlestir: bool = False


@app.post("/api/vardiya/v2/serbest-slot-hazirla")
def v2_serbest_slot_hazirla():
    """Her aktif şubeye 'Serbest' satırı (slot) oluşturur/garantiler — kullanıcı slot
    kurmadan grid'de her şubede hazır bir satıra sürükleyip atayabilsin. İdempotent."""
    with db() as (conn, cur):
        cur.execute("SELECT id FROM subeler WHERE aktif=TRUE")
        sids = [dict(r)["id"] for r in (cur.fetchall() or [])]
        for sid in sids:
            _vv2.serbest_slot_getir_olustur(cur, sid)
    return {"success": True, "hazirlanan_sube": len(sids)}


@app.get("/api/vardiya/v2/iscilik-ozet")
def v2_iscilik_ozet(tarih: str):
    """Gün bazında işçilik maliyeti + ciro tahminine göre işçilik % (şube + toplam)."""
    from datetime import datetime as _dt
    t = _dt.strptime(tarih, "%Y-%m-%d").date()
    with db() as (conn, cur):
        return _vv2.iscilik_ozet(cur, t)


@app.post("/api/vardiya/v2/atama-serbest")
def v2_atama_serbest(a: _V2AtamaSerbestIn):
    """Slot kurma derdi olmadan serbest saatle atama. Şubenin 'Serbest' slot'unu
    otomatik kullanır/oluşturur; bitiş < başlangıç ise gece vardiyası otomatik."""
    from datetime import datetime as _dt
    t = _dt.strptime(a.tarih, "%Y-%m-%d").date()
    with db() as (conn, cur):
        slot_id = _vv2.serbest_slot_getir_olustur(cur, a.sube_id)
        sonuc = _vv2.atama_olustur(
            cur, a.personel_id, slot_id, t,
            _t(a.baslangic_saat), _t(a.bitis_saat),
            override=a.override,
            aciklama=a.aciklama,
            durum="onayli" if bool(a.kesinlestir) else "planli",
        )
        if not sonuc.get('basarili'):
            raise HTTPException(409, sonuc)
        return sonuc


@app.delete("/api/vardiya/v2/atama/{aid}")
def v2_atama_iptal(aid: str):
    with db() as (conn, cur):
        sonuc = _vv2.atama_iptal(cur, aid)
        if not sonuc.get('basarili'):
            raise HTTPException(404, sonuc.get('mesaj', 'Bulunamadı'))
        return sonuc


@app.post("/api/vardiya/v2/gun-temizle")
def v2_gun_temizle(tarih: str, sube_id: Optional[str] = None):
    """Bir günün tüm atamalarını iptal eder (opsiyonel: tek şube)."""
    from datetime import datetime as _dt
    t = _dt.strptime(tarih, "%Y-%m-%d").date()
    with db() as (conn, cur):
        return _vv2.gun_temizle(cur, t, sube_id)


@app.get("/api/vardiya/v2/hafta-personel-tablo")
def v2_hafta_personel_tablo(pazartesi: str):
    """Tulipi PDF formatında personel × 7 gün haftalık görünüm."""
    from datetime import datetime as _dt
    p = _dt.strptime(pazartesi, "%Y-%m-%d").date()
    # Pazartesi'ye normalize et
    p = p - timedelta(days=p.weekday())
    with db() as (conn, cur):
        return _vv2.personel_haftalik_gorunum(cur, p)


@app.post("/api/vardiya/v2/gun-kopyala")
def v2_gun_kopyala(kaynak: str, hedef: str,
                    sube_id: Optional[str] = None,
                    temizle: bool = True):
    """Kaynak günün atamalarını hedef güne kopyalar."""
    from datetime import datetime as _dt
    k = _dt.strptime(kaynak, "%Y-%m-%d").date()
    h = _dt.strptime(hedef, "%Y-%m-%d").date()
    with db() as (conn, cur):
        return _vv2.gun_kopyala(cur, k, h, sube_id, temizle)


@app.put("/api/vardiya/v2/personel-gun")
def v2_personel_gun_niyet(body: _V2PersonelGunNiyetIn):
    from datetime import datetime as _dt
    t = _dt.strptime(body.tarih, "%Y-%m-%d").date()
    with db() as (conn, cur):
        _vv2.personel_gun_niyet_kaydet(cur, body.personel_id, t, body.kasitli_bos)
        _vv2.personel_gun_durumu(cur, body.personel_id, t)
    return {"basarili": True}


@app.get("/api/vardiya/v2/gun-kilit")
def v2_gun_kilit_getir(tarih: str):
    from datetime import datetime as _dt
    t = _dt.strptime(tarih, "%Y-%m-%d").date()
    with db() as (conn, cur):
        return {"tarih": tarih, "kilitli": _vv2.gun_kilit_mi(cur, t)}


@app.put("/api/vardiya/v2/gun-kilit")
def v2_gun_kilit_kaydet(k: _V2GunKilitIn):
    from datetime import datetime as _dt
    t = _dt.strptime(k.tarih, "%Y-%m-%d").date()
    with db() as (conn, cur):
        _vv2.gun_kilit_kaydet(cur, t, k.kilitli, k.aciklama or "")
    return {"basarili": True}


# ── İZİN ──
@app.get("/api/vardiya/v2/izin")
def v2_izin_liste(personel_id: Optional[str] = None,
                  baslangic: Optional[str] = None,
                  bitis: Optional[str] = None):
    with db() as (conn, cur):
        sql = (
            "SELECT i.*, TRIM(COALESCE(p.ad_soyad, '')) AS _personel_full "
            "FROM personel_izin i JOIN personel p ON p.id=i.personel_id WHERE 1=1"
        )
        params: List = []
        if personel_id:
            sql += " AND i.personel_id = %s"; params.append(personel_id)
        if baslangic:
            sql += " AND i.bitis_tarih >= %s"; params.append(baslangic)
        if bitis:
            sql += " AND i.baslangic_tarih <= %s"; params.append(bitis)
        sql += " ORDER BY i.baslangic_tarih DESC"
        cur.execute(sql, tuple(params))
        izinler = []
        for r in cur.fetchall():
            d = dict(r)
            a, s = _vardiya_personel_ad_split(d.pop("_personel_full", None))
            d["personel_ad"] = a or "(isimsiz)"
            d["personel_soyad"] = s
            izinler.append(d)
        return {"izinler": izinler}


@app.get("/api/vardiya/v2/izin-hafta-ozet")
def v2_izin_hafta_ozet(pazartesi: str):
    """
    Seçilen tarihin ait olduğu ISO haftasında (Pzt–Paz) izin kaydı olmayan aktif personel listesi.
    `pazartesi` herhangi bir gün olabilir; haftanın Pazartesi’sine normalize edilir.
    """
    from datetime import datetime as _dt
    raw = _dt.strptime(pazartesi[:10], "%Y-%m-%d").date()
    mon = raw - timedelta(days=raw.weekday())
    with db() as (conn, cur):
        return _vv2.izin_hafta_ozet(cur, mon)


@app.post("/api/vardiya/v2/izin")
def v2_izin_ekle(i: _V2IzinIn):
    from datetime import datetime as _dt
    iid = str(uuid.uuid4())
    d1 = _dt.strptime(i.baslangic_tarih[:10], "%Y-%m-%d").date()
    d2 = _dt.strptime(i.bitis_tarih[:10], "%Y-%m-%d").date()
    if d2 < d1:
        raise HTTPException(400, "Bitiş tarihi başlangıçtan önce olamaz.")
    with db() as (conn, cur):
        if not bool(i.force):
            cak = _vv2.personel_izin_baska_ayni_iso_haftada(cur, i.personel_id, d1, d2)
            if cak:
                parcalar = []
                for x in cak[:4]:
                    parcalar.append(
                        f"{x.get('baslangic_tarih')} → {x.get('bitis_tarih')} ({x.get('tip') or '?'})"
                    )
                ozet = " | ".join(parcalar)
                if len(cak) > 4:
                    ozet += f" (+{len(cak) - 4} kayıt daha)"
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "Bu personel için aynı takvim haftasında (Pzt–Paz) zaten izin kaydı var. "
                        "Bu hafta içinde tekrar izin tanımlıyorsunuz; yasal planlamada çift kayıt oluşmaması için "
                        "önce mevcut kaydı düzenleyin veya silin. Mevcut: "
                        + ozet
                        + " Yine de eklemek için gövdede «force»: true gönderin."
                    ),
                )
        cur.execute("""
            INSERT INTO personel_izin
                (id, personel_id, baslangic_tarih, bitis_tarih, tip, aciklama)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (iid, i.personel_id, i.baslangic_tarih, i.bitis_tarih,
              i.tip, i.aciklama))
        _vv2.personel_gun_state_yenile_tarih_araligi(cur, i.personel_id, d1, d2)
    return {"id": iid}

@app.delete("/api/vardiya/v2/izin/{iid}")
def v2_izin_sil(iid: str):
    with db() as (conn, cur):
        cur.execute(
            "SELECT personel_id, baslangic_tarih, bitis_tarih FROM personel_izin WHERE id = %s",
            (iid,),
        )
        r = cur.fetchone()
        if not r:
            raise HTTPException(404, "İzin bulunamadı")
        pid = r["personel_id"]
        d1 = r["baslangic_tarih"]
        d2 = r["bitis_tarih"]
        cur.execute("DELETE FROM personel_izin WHERE id = %s", (iid,))
        if cur.rowcount == 0:
            raise HTTPException(404, "İzin bulunamadı")
        _vv2.personel_gun_state_yenile_tarih_araligi(cur, pid, d1, d2)
    return {"basarili": True}


# ── RAPORLAR (Aşama 7) ──
@app.get("/api/vardiya/v2/rapor/fazla-mesai")
def v2_rapor_fazla_mesai(
    baslangic: str,
    bitis: str,
    limit: int = 500,
):
    from datetime import datetime as _dt
    d1 = _dt.strptime(baslangic[:10], "%Y-%m-%d").date()
    d2 = _dt.strptime(bitis[:10], "%Y-%m-%d").date()
    with db() as (conn, cur):
        return _vv2.rapor_fazla_mesai(cur, d1, d2, limit=limit)


@app.get("/api/vardiya/v2/rapor/izinli-calisti")
def v2_rapor_izinli_calisti(
    baslangic: str,
    bitis: str,
    limit: int = 500,
):
    from datetime import datetime as _dt
    d1 = _dt.strptime(baslangic[:10], "%Y-%m-%d").date()
    d2 = _dt.strptime(bitis[:10], "%Y-%m-%d").date()
    with db() as (conn, cur):
        return _vv2.rapor_izinli_calisti(cur, d1, d2, limit=limit)


@app.get("/api/vardiya/v2/rapor/devamsizlik")
def v2_rapor_devamsizlik(
    baslangic: str,
    bitis: str,
    limit: int = 500,
):
    """Vardiyada planlı ama hiç yoklaması olmayan + kapanışı unutulmuş personel."""
    from datetime import datetime as _dt
    d1 = _dt.strptime(baslangic[:10], "%Y-%m-%d").date()
    d2 = _dt.strptime(bitis[:10], "%Y-%m-%d").date()
    with db() as (conn, cur):
        return _vv2.rapor_devamsizlik(cur, d1, d2, limit=limit)


# ── PRESET (TAM/PART/ARACI/AÇILIŞ/KAPANIŞ) ──
class _V2PresetIn(BaseModel):
    kod: str
    ad: str
    bas_saat: str
    bit_saat: str
    gece_vardiyasi: bool = False
    renk: Optional[str] = None
    sira: int = 0
    aktif: bool = True


@app.get("/api/vardiya/v2/preset")
def v2_preset_listele():
    with db() as (conn, cur):
        return {"presetler": _vv2.vardiya_preset_listele(cur)}


@app.get("/api/vardiya/v2/preset-admin")
def v2_preset_admin_liste():
    """Tüm preset satırları (pasif dahil) — sistem yönetimi ekranı."""
    with db() as (conn, cur):
        return {"presetler": _vv2.vardiya_preset_listele_hepsi(cur)}


@app.post("/api/vardiya/v2/preset")
def v2_preset_kaydet(p: _V2PresetIn):
    with db() as (conn, cur):
        cur.execute("""
            INSERT INTO vardiya_preset (kod, ad, bas_saat, bit_saat, gece_vardiyasi, renk, sira, aktif)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (kod) DO UPDATE SET
                ad = EXCLUDED.ad,
                bas_saat = EXCLUDED.bas_saat,
                bit_saat = EXCLUDED.bit_saat,
                gece_vardiyasi = EXCLUDED.gece_vardiyasi,
                renk = EXCLUDED.renk,
                sira = EXCLUDED.sira,
                aktif = EXCLUDED.aktif
        """, (p.kod, p.ad, _t(p.bas_saat), _t(p.bit_saat), p.gece_vardiyasi,
              p.renk, p.sira, p.aktif))
    return {"basarili": True}


@app.delete("/api/vardiya/v2/preset/{kod}")
def v2_preset_sil(kod: str):
    with db() as (conn, cur):
        cur.execute("UPDATE vardiya_preset SET aktif = FALSE WHERE kod = %s", (kod,))
        if cur.rowcount == 0:
            raise HTTPException(404, "Preset bulunamadı")
    return {"basarili": True}


@app.get("/api/vardiya/v2/personel-onerilen-saat")
def v2_personel_onerilen_saat(
    personel_id: str,
    tarih: str,
    slot_id: Optional[str] = None,
):
    """
    Personelin verilen tarih için önerilen vardiya saatini döner.
    Öncelik: personel gün preset (vardiya_preset_json) → yoksa slot + günlük limit ile mesai bandı
    (tam/part ayrımı yok; uzun slotlarda tam gün varsayılmaz).
    `slot_id` verilmezse yalnızca JSON preset dönebilir.
    """
    from datetime import datetime as _dt
    t = _dt.strptime(tarih, "%Y-%m-%d").date()
    with db() as (conn, cur):
        preset = _vv2.personel_gun_preset(cur, personel_id, t)
        kaynak = "personel_json" if preset else None
        if not preset and slot_id:
            preset = _vv2.slot_mesai_onerilen_saat(cur, personel_id, t, slot_id)
            if preset:
                kaynak = "mesai_slot"
        return {"preset": preset, "tarih": tarih, "kaynak": kaynak}


@app.get("/api/vardiya/v2/personel-kisit-serbest-saat")
def v2_personel_kisit_serbest_saat(
    personel_id: str,
    tarih: str,
    slot_id: Optional[str] = None,
):
    """
    Personelin o gün ders/kısıt saatlerinin dışındaki ilk uygun zaman dilimini döner.
    slot_id verilirse yalnızca o slot'un saat çerçevesi içinde arar.
    Dönüş: {bas_saat, bit_saat, kisit_var, yasaklar[], mesaj}
    """
    from datetime import datetime as _dt
    t = _dt.strptime(tarih, "%Y-%m-%d").date()
    with db() as (conn, cur):
        cur.execute(
            "SELECT gun_saat_kisitlari_json FROM personel_kisit WHERE personel_id = %s",
            (personel_id,),
        )
        row = cur.fetchone()
        gsk = (row["gun_saat_kisitlari_json"] if row else {}) or {}
        if isinstance(gsk, str):
            import json as _json
            try:
                gsk = _json.loads(gsk)
            except Exception:
                gsk = {}

        gun_kisa = _vv2.GUN_KISALTMA[t.weekday()]
        yasak_listesi = gsk.get(gun_kisa) or []

        arama_bas_t = _vv2._parse_saat_metni("08:00")
        arama_bit_t = _vv2._parse_saat_metni("22:00")
        if slot_id:
            cur.execute(
                "SELECT baslangic_saat, bitis_saat FROM vardiya_slot WHERE id = %s",
                (slot_id,),
            )
            sr = cur.fetchone()
            if sr and sr["baslangic_saat"] and sr["bitis_saat"]:
                arama_bas_t = sr["baslangic_saat"]
                arama_bit_t = sr["bitis_saat"]

        def _to_min(tt):
            return tt.hour * 60 + tt.minute

        def _to_str(m):
            return f"{m // 60:02d}:{m % 60:02d}"

        yasaklar = []
        for ys in yasak_listesi:
            yb = _vv2._parse_saat_metni(ys.get("yasak_bas"))
            yt = _vv2._parse_saat_metni(ys.get("yasak_bit"))
            if yb and yt:
                yasaklar.append({"bas": yb, "bit": yt, "neden": ys.get("neden", "Kısıt")})
        yasaklar.sort(key=lambda x: x["bas"])
        yasak_out = [{"neden": y["neden"], "bas": str(y["bas"]), "bit": str(y["bit"])} for y in yasaklar]

        if not yasaklar:
            return {
                "bas_saat": str(arama_bas_t),
                "bit_saat": str(arama_bit_t),
                "kisit_var": False,
                "yasaklar": [],
                "mesaj": "Bu gün için kısıt tanımlı değil.",
            }

        r_bas = _to_min(arama_bas_t)
        r_bit = _to_min(arama_bit_t)
        current = r_bas

        for ys in yasaklar:
            yb = _to_min(ys["bas"])
            yt = _to_min(ys["bit"])
            if yt <= current:
                continue
            if yb > current:
                free_end = min(yb, r_bit)
                if free_end > current:
                    return {
                        "bas_saat": _to_str(current),
                        "bit_saat": _to_str(free_end),
                        "kisit_var": True,
                        "yasaklar": yasak_out,
                        "mesaj": f"Kısıt başlamadan önceki serbest dilim önerildi ({_to_str(current)}–{_to_str(free_end)}).",
                    }
            current = max(current, yt)

        if current < r_bit:
            return {
                "bas_saat": _to_str(current),
                "bit_saat": _to_str(r_bit),
                "kisit_var": True,
                "yasaklar": yasak_out,
                "mesaj": f"Kısıtlar bittikten sonraki serbest dilim önerildi ({_to_str(current)}–{_to_str(r_bit)}).",
            }

        return {
            "bas_saat": None,
            "bit_saat": None,
            "kisit_var": True,
            "yasaklar": yasak_out,
            "mesaj": "Bu gün tüm saatler kısıtlı — uygun boş dilim bulunamadı.",
        }


# ── OVERRIDE LOG ──
@app.get("/api/vardiya/v2/override-log")
def v2_override_log_liste(limit: int = 100, personel_id: Optional[str] = None):
    with db() as (conn, cur):
        sql = (
            "SELECT o.*, TRIM(COALESCE(p.ad_soyad, '')) AS _personel_full "
            "FROM vardiya_override_log o LEFT JOIN personel p ON p.id=o.personel_id "
            "WHERE 1=1"
        )
        params: List = []
        if personel_id:
            sql += " AND o.personel_id = %s"; params.append(personel_id)
        sql += " ORDER BY o.ts DESC LIMIT %s"
        params.append(min(max(int(limit), 1), 1000))
        cur.execute(sql, tuple(params))
        kayitlar = []
        for r in cur.fetchall():
            d = dict(r)
            a, s = _vardiya_personel_ad_split(d.pop("_personel_full", None))
            d["personel_ad"] = a or "(isimsiz)"
            d["personel_soyad"] = s
            kayitlar.append(d)
        return {"kayitlar": kayitlar}


# ── WhatsApp Endpoints ──────────────────────────────────────────────────────

@app.get("/api/whatsapp/onizle")
def whatsapp_onizle(tarih: Optional[str] = None):
    """Mesaj içeriğini gönderme olmadan önizler (test amaçlı)."""
    from whatsapp_bildirim import gunluk_ozet_mesaj_olustur
    hedef = date.fromisoformat(tarih) if tarih else None
    try:
        mesaj = gunluk_ozet_mesaj_olustur(hedef)
        return {"mesaj": mesaj}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/whatsapp/durum")
def whatsapp_durum():
    """Green API bağlantı/yetki durumunu salt-okur (mesaj göndermez). Tedarikçi
    siparişleri WhatsApp'tan düşmüyorsa buradan teşhis: state='authorized' değilse
    WhatsApp telefonu Green API'den kopmuş demektir → yeniden QR ile bağlanmalı."""
    from whatsapp_bildirim import wa_instance_durum
    return wa_instance_durum()


@app.get("/api/whatsapp/gonder")
@app.post("/api/whatsapp/gonder")
def whatsapp_manuel_gonder(tarih: Optional[str] = None):
    """Mesajı hemen gönderir (manuel tetikleme, test amaçlı)."""
    from whatsapp_bildirim import gunluk_ozet_gonder
    hedef = date.fromisoformat(tarih) if tarih else None
    try:
        sonuc = gunluk_ozet_gonder(hedef)
        if not sonuc.get("basarili"):
            raise HTTPException(status_code=500, detail=sonuc.get("hata", "Gönderim başarısız"))
        return sonuc
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# Frontend
if pathlib.Path("static/index.html").exists():
    from fastapi.responses import FileResponse
    from fastapi import Request as _Req

    # assets önce mount edilmeli — wildcard route kapmadan
    app.mount("/assets", StaticFiles(directory="static/assets"), name="assets")

    _idx_path = pathlib.Path("static/index.html")
    _spa_headers = {
        "Cache-Control": "no-cache, no-store, must-revalidate",
        "Pragma": "no-cache",
        "Expires": "0",
    }

    @app.get("/admin")
    @app.get("/admin/{admin_path:path}")
    async def serve_admin_spa(admin_path: str = ""):
        _ = admin_path
        if _idx_path.exists():
            return FileResponse(str(_idx_path), headers=_spa_headers)
        from fastapi.responses import JSONResponse
        return JSONResponse({"detail": "Frontend not built"}, status_code=404)

    @app.get("/")
    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str = "", request: _Req = None):
        """SPA routing — index.html'i her zaman no-cache ile sun"""
        import pathlib as _pl
        if full_path.startswith("api/") or full_path.startswith("assets/"):
            from fastapi.responses import JSONResponse
            return JSONResponse({"detail": "Not found"}, status_code=404)
        if _idx_path.exists():
            return FileResponse(str(_idx_path), headers=_spa_headers)
        return JSONResponse({"detail": "Frontend not built"}, status_code=404)
